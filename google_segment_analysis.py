from __future__ import annotations

"""Verified Alphabet segment adapter for GOOG / GOOGL.

Alphabet's 2025 Form 10-K contains clean 2023-2025 segment and revenue-disaggregation
figures, but the generic SEC HTML table parser can occasionally fail on the filing's inline
XBRL/HTML structure. This adapter is deliberately narrow: it supplies only company-reported
values verified against Alphabet's 2025 Form 10-K and never estimates missing segment economics.

Values below are USD billions. The underlying filing reports USD millions.
"""

from openpyxl.styles import Alignment, Font

from segment_analysis_v2 import _write_sheet


SOURCE_URL = "https://www.sec.gov/Archives/edgar/data/1652044/000165204426000018/goog-20251231.htm"
YEARS = [2023, 2024, 2025]

SEGMENTS = [
    ("Google Services", [272.543, 304.930, 342.721], [95.858, 121.263, 139.404]),
    ("Google Cloud", [33.088, 43.229, 58.705], [1.716, 6.112, 13.910]),
    ("Other Bets", [1.527, 1.648, 1.537], [-4.095, -4.444, -7.515]),
]

BUSINESS_LINES = [
    ("Google Search & other", [175.033, 198.084, 224.532]),
    ("YouTube ads", [31.510, 36.147, 40.367]),
    ("Google Network", [31.312, 30.359, 29.792]),
    ("Google subscriptions, platforms, and devices", [34.688, 40.340, 48.030]),
    ("Google Cloud", [33.088, 43.229, 58.705]),
    ("Other Bets", [1.527, 1.648, 1.537]),
]

STATUS = (
    "VERIFIED OFFICIAL — Alphabet 2025 Form 10-K. The generic SEC HTML table parser can fail "
    "on this filing, so 2023-2025 reportable-segment revenue, segment operating income (loss), "
    "and disclosed revenue groups are populated from a verified filing adapter. Values are USD "
    "billions converted from the filing's USD millions; no segment values are estimated."
)


def ensure_google_segment_analysis(wb, ticker):
    """Build the canonical Segment Analysis sheet for GOOG/GOOGL from verified filing values."""
    t = str(ticker or "").upper().strip()
    if t not in {"GOOG", "GOOGL"}:
        return None

    ws = _write_sheet(wb, t, YEARS, SEGMENTS, BUSINESS_LINES, SOURCE_URL, STATUS)

    for r in range(7, 10):
        ws.cell(r, 15, "Revenue + profitability")
        ws.cell(r, 16, "Alphabet 2025 Form 10-K — Note 15")
        ws.cell(r, 16).hyperlink = SOURCE_URL
        ws.cell(r, 16).font = Font(color="008000")

    # _write_sheet reserves ten rows; the six disclosed revenue groups occupy rows 21:26.
    for r in range(21, 27):
        ws.cell(r, 8, "Alphabet 2025 Form 10-K — revenue disaggregation")
        ws.cell(r, 8).hyperlink = SOURCE_URL
        ws.cell(r, 8).font = Font(color="008000")

    source_row = 33
    ws.cell(source_row + 3, 2, (
        "Reportable-segment rows follow Alphabet's segment disclosure. Segment revenue excludes "
        "Alphabet-level hedging gains/losses, and segment operating income excludes Alphabet-level "
        "activities. Therefore segment totals need not equal consolidated revenue/operating income. "
        "No undisclosed product revenue or profit is estimated."
    ))
    ws.cell(source_row + 3, 2).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[3].height = 62
    return ws
