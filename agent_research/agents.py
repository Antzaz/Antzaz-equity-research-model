"""Phase-1 specialist agents for the equity research pipeline."""

from __future__ import annotations

import hashlib
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .core import AgentContext, AgentResult, OpenAIResearchClient, normalize_json_value, utc_now_iso


class BaseAgent:
    name = "Base Agent"

    def __init__(self, ai_client: OpenAIResearchClient | None = None):
        self.ai_client = ai_client

    def run(self, ctx: AgentContext) -> AgentResult:
        raise NotImplementedError


class FilingsAgent(BaseAgent):
    name = "Filings & Financials Agent"

    def run(self, ctx: AgentContext) -> AgentResult:
        wb = load_workbook(ctx.workbook_path, read_only=True, data_only=False)
        rows: list[dict[str, Any]] = []
        if "Filings" in wb.sheetnames:
            ws = wb["Filings"]
            for r in range(1, min(ws.max_row, 40) + 1):
                values = [normalize_json_value(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 8) + 1)]
                form = str(values[0] or "").strip().upper() if values else ""
                if form in {"10-K", "10-Q", "8-K", "DEF 14A", "20-F", "6-K"}:
                    rows.append({
                        "form": values[0],
                        "report_date": values[1] if len(values) > 1 else None,
                        "filed_date": values[2] if len(values) > 2 else None,
                        "source": values[3] if len(values) > 3 else None,
                        "source_type": values[4] if len(values) > 4 else None,
                    })
        wb.close()

        payload = {
            "ticker": ctx.ticker,
            "generated_at": utc_now_iso(),
            "workbook": str(ctx.workbook_path),
            "filings": rows,
        }
        artifact = ctx.write_json("filings_snapshot.json", payload)
        ctx.shared["filings"] = payload

        latest = rows[0] if rows else None
        summary = f"Captured {len(rows)} filing reference(s)."
        if latest:
            summary += f" Latest listed filing: {latest.get('form')} filed {latest.get('filed_date') or 'date unavailable'}."
        else:
            summary += " No recognized filing rows were found in the workbook."

        narrative = None
        if self.ai_client:
            narrative = self.ai_client.analyze(
                self.name,
                "Identify filing coverage gaps and the next primary-source documents a human analyst should review. Do not claim to have read a filing when only metadata is supplied.",
                payload,
            )
        return AgentResult(self.name, "PASS" if rows else "WARN", summary, payload, [str(artifact)], narrative)


class KPIEarningsAgent(BaseAgent):
    name = "KPI / Earnings Agent"

    _HEADER_ALIASES = {
        "kpi / evidence": "kpi",
        "current": "current",
        "unit / comparison": "unit_comparison",
        "signal": "signal",
        "investment read-through": "investment_read_through",
        "as of": "as_of",
        "source url": "source_url",
        "data type": "data_type",
    }

    def _extract(self, workbook: Path) -> list[dict[str, Any]]:
        wb = load_workbook(workbook, read_only=True, data_only=False)
        if "AI Impact Analysis" not in wb.sheetnames:
            wb.close()
            return []
        ws = wb["AI Impact Analysis"]
        header_row = None
        mapping: dict[int, str] = {}
        for r in range(1, min(ws.max_row, 80) + 1):
            row_map: dict[int, str] = {}
            for c in range(1, min(ws.max_column, 14) + 1):
                raw = ws.cell(r, c).value
                key = str(raw or "").strip().lower()
                if key in self._HEADER_ALIASES:
                    row_map[c] = self._HEADER_ALIASES[key]
            if "kpi" in row_map.values() and "signal" in row_map.values():
                header_row = r
                mapping = row_map
                break
        if header_row is None:
            wb.close()
            return []

        records: list[dict[str, Any]] = []
        blank_streak = 0
        for r in range(header_row + 1, min(ws.max_row, header_row + 35) + 1):
            item = {name: normalize_json_value(ws.cell(r, c).value) for c, name in mapping.items()}
            if not item.get("kpi"):
                blank_streak += 1
                if blank_streak >= 2:
                    break
                continue
            blank_streak = 0
            records.append(item)
        wb.close()
        return records

    @staticmethod
    def _fingerprint(kpis: list[dict[str, Any]]) -> str:
        raw = json.dumps(kpis, sort_keys=True, ensure_ascii=False, default=str).encode("utf-8")
        return hashlib.sha256(raw).hexdigest()[:16]

    def _update_history(self, ctx: AgentContext, kpis: list[dict[str, Any]]) -> tuple[Path, dict[str, Any] | None, bool]:
        path = ctx.research_data_dir / "kpi_history.json"
        if path.exists():
            try:
                history = json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                history = {"ticker": ctx.ticker, "snapshots": []}
        else:
            history = {"ticker": ctx.ticker, "snapshots": []}
        history.setdefault("ticker", ctx.ticker)
        snapshots = history.setdefault("snapshots", [])
        previous = snapshots[-1] if snapshots else None
        fp = self._fingerprint(kpis)
        changed = not previous or previous.get("fingerprint") != fp
        if changed:
            snapshots.append({
                "captured_at": utc_now_iso(),
                "fingerprint": fp,
                "workbook": ctx.workbook_path.name,
                "kpis": kpis,
            })
            path.write_text(json.dumps(history, indent=2, ensure_ascii=False, default=str) + "\n", encoding="utf-8")
        return path, previous, changed

    def run(self, ctx: AgentContext) -> AgentResult:
        kpis = self._extract(ctx.workbook_path)
        history_path, previous, changed = self._update_history(ctx, kpis)
        payload = {
            "ticker": ctx.ticker,
            "generated_at": utc_now_iso(),
            "workbook": str(ctx.workbook_path),
            "changed_from_previous_snapshot": changed,
            "previous_snapshot": previous,
            "kpis": kpis,
            "history_path": str(history_path),
        }
        artifact = ctx.write_json("kpi_snapshot.json", payload)
        ctx.shared["kpis"] = payload
        summary = f"Captured {len(kpis)} KPI/evidence row(s)."
        summary += " New evidence snapshot stored." if changed else " Evidence matches the previous snapshot."
        if not kpis:
            summary += " AI Impact Analysis did not contain a recognizable evidence table."

        narrative = None
        if self.ai_client and kpis:
            narrative = self.ai_client.analyze(
                self.name,
                "Classify the evidence into adoption, monetization, capital intensity, moat, and cannibalization. Highlight what changed versus the previous snapshot only when the JSON supports that comparison.",
                payload,
            )
        return AgentResult(self.name, "PASS" if kpis else "WARN", summary, payload, [str(artifact), str(history_path)], narrative)


class ThesisMonitorAgent(BaseAgent):
    name = "Thesis Monitor Agent"

    @staticmethod
    def _signal_score(signal: Any) -> int:
        text = str(signal or "").strip().lower()
        if not text:
            return 0
        if "very strong" in text:
            return 2
        if "key risk" in text:
            return -2
        if "strong" in text or "positive" in text:
            return 1
        if "risk" in text or "weak" in text or "negative" in text:
            return -1
        return 0

    def run(self, ctx: AgentContext) -> AgentResult:
        kpi_payload = ctx.shared.get("kpis") or {}
        kpis = kpi_payload.get("kpis") or []
        scored = []
        total = 0
        for item in kpis:
            score = self._signal_score(item.get("signal"))
            total += score
            scored.append({"kpi": item.get("kpi"), "signal": item.get("signal"), "score": score})
        if not kpis:
            classification = "Insufficient evidence"
        elif total >= 4:
            classification = "Evidence strengthened"
        elif total <= -4:
            classification = "Evidence weakened"
        else:
            classification = "Mixed / broadly unchanged"

        payload = {
            "ticker": ctx.ticker,
            "generated_at": utc_now_iso(),
            "classification": classification,
            "evidence_score": total,
            "scored_signals": scored,
            "snapshot_changed": kpi_payload.get("changed_from_previous_snapshot"),
            "important_note": "The deterministic evidence score is a triage aid, not an investment rating or valuation input.",
        }
        artifact = ctx.write_json("thesis_monitor.json", payload)
        ctx.shared["thesis_monitor"] = payload
        narrative = None
        if self.ai_client and kpis:
            ai_evidence = {
                "kpis": kpis,
                "previous_snapshot": kpi_payload.get("previous_snapshot"),
                "deterministic_triage": payload,
            }
            narrative = self.ai_client.analyze(
                self.name,
                "Assess whether the evidence strengthens, weakens, or leaves unchanged the investment thesis across growth, margins, capital intensity, moat, and disruption risk. Name explicit falsification questions. Do not alter DCF assumptions; propose changes for analyst review only.",
                ai_evidence,
            )
        return AgentResult(self.name, "PASS" if kpis else "WARN", f"{classification}; evidence triage score {total:+d}.", payload, [str(artifact)], narrative)


class ResearchQAAgent(BaseAgent):
    name = "Research QA / Skeptic Agent"

    @staticmethod
    def _parse_as_of(value: Any) -> date | None:
        text = str(value or "").strip()
        if not text:
            return None
        try:
            return date.fromisoformat(text[:10])
        except Exception:
            pass
        match = re.fullmatch(r"(\d{4})-Q([1-4])", text.upper())
        if match:
            year, quarter = int(match.group(1)), int(match.group(2))
            month = quarter * 3
            day = 30 if month in {6, 9} else 31
            return date(year, month, day)
        return None

    def run(self, ctx: AgentContext) -> AgentResult:
        issues: list[dict[str, str]] = []
        kpi_payload = ctx.shared.get("kpis") or {}
        kpis = kpi_payload.get("kpis") or []
        seen: set[str] = set()
        today = datetime.now().date()

        if not ctx.workbook_path.exists():
            issues.append({"severity": "ERROR", "issue": "Research workbook is missing."})
        if not kpis:
            issues.append({"severity": "WARN", "issue": "No KPI evidence rows were extracted."})
        for item in kpis:
            name = str(item.get("kpi") or "").strip()
            key = name.lower()
            if key in seen:
                issues.append({"severity": "WARN", "issue": f"Duplicate KPI evidence row: {name}"})
            seen.add(key)
            if not item.get("source_url"):
                issues.append({"severity": "WARN", "issue": f"Missing source URL for KPI: {name}"})
            if not item.get("data_type"):
                issues.append({"severity": "WARN", "issue": f"Missing disclosure/data type for KPI: {name}"})
            as_of = self._parse_as_of(item.get("as_of"))
            if not as_of:
                issues.append({"severity": "WARN", "issue": f"Unparseable or missing as-of date for KPI: {name}"})
            elif (today - as_of).days > 550:
                issues.append({"severity": "WARN", "issue": f"Potentially stale KPI (>550 days): {name} ({item.get('as_of')})"})

        errors = sum(1 for x in issues if x["severity"] == "ERROR")
        warnings = sum(1 for x in issues if x["severity"] == "WARN")
        status = "FAIL" if errors else ("WARN" if warnings else "PASS")
        payload = {
            "ticker": ctx.ticker,
            "generated_at": utc_now_iso(),
            "status": status,
            "errors": errors,
            "warnings": warnings,
            "issues": issues,
            "controls": [
                "AI narrative is never treated as a primary data source.",
                "Source URLs and disclosure types are checked for KPI evidence.",
                "Stale and duplicate KPI evidence is flagged.",
                "The deterministic valuation engine remains separate from agent reasoning.",
            ],
        }
        artifact = ctx.write_json("qa_report.json", payload)
        ctx.shared["qa"] = payload
        narrative = None
        if self.ai_client:
            narrative = self.ai_client.analyze(
                self.name,
                "Act as an internal research auditor. Find unsupported leaps, missing evidence, and reasons the analyst should not update valuation assumptions yet. Keep the audit grounded in the supplied QA and evidence payloads.",
                {"qa": payload, "kpis": kpis, "filings": ctx.shared.get("filings")},
            )
        summary = f"QA {status}: {errors} error(s), {warnings} warning(s)."
        return AgentResult(self.name, status, summary, payload, [str(artifact)], narrative)
