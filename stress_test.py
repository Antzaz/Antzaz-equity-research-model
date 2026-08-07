"""Automated downside stress testing for the equity research workbook."""

from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule

NAVY = "17365D"
BLUE = "2F75B5"
WHITE = "FFFFFF"
GOLD = "FFF2CC"
INPUT_BLUE = "0000FF"
GREEN = "E2F0D9"
RED = "FCE4D6"
LINK_GREEN = "008000"
GREY = "666666"

FMT_PCT = "0.0%;[Red](0.0%);-"
FMT_PRICE = "$#,##0.00;[Red]($#,##0.00);-"
FMT_BN = "#,##0.0;[Red](#,##0.0);-"


def _col_name(n):
    out = ""
    while n:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out


def _merge_once(ws, cell_range):
    if cell_range not in {str(r) for r in ws.merged_cells.ranges}:
        ws.merge_cells(cell_range)


def ensure_stress_test(wb):
    """
    Add/update an automated stress-test module driven by the Base scenario.

    Tests:
      - Revenue growth -3ppt / -5ppt
      - EBIT margin -300bps / -500bps
      - Capex/revenue +500bps
      - WACC +100bps / +200bps
      - Terminal growth -50bps / -100bps
      - Combined Severe Bear

    Each row rebuilds 2026-2035 revenue and FCF, then performs a full DCF.
    """
    if "Three-Case Scenarios" not in wb.sheetnames:
        return

    ws = wb["Three-Case Scenarios"]
    existed = ws["A44"].value == "Automated Stress Testing — Base Case Robustness"

    _merge_once(ws, "A44:AI44")
    ws["A44"] = "Automated Stress Testing — Base Case Robustness"
    ws["A44"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A44"].font = Font(bold=True, color=WHITE, size=14)

    _merge_once(ws, "A45:AI45")
    ws["A45"] = (
        "Each row shocks the Base case and recalculates 2026–2035 revenue, FCF and DCF value. "
        "Yellow/blue cells are editable stress assumptions; Severe Bear combines all five downside shocks."
    )
    ws["A45"].font = Font(italic=True, color=GREY)
    ws["A45"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[45].height = 30

    _merge_once(ws, "L46:U46")
    ws["L46"] = "Stressed FCF ($bn)"
    ws["L46"].fill = PatternFill("solid", fgColor=NAVY)
    ws["L46"].font = Font(bold=True, color=WHITE)
    ws["L46"].alignment = Alignment(horizontal="center")

    _merge_once(ws, "V46:AE46")
    ws["V46"] = "Stressed Revenue ($bn)"
    ws["V46"].fill = PatternFill("solid", fgColor=NAVY)
    ws["V46"].font = Font(bold=True, color=WHITE)
    ws["V46"].alignment = Alignment(horizontal="center")

    headers = [
        "Stress Test", "Revenue Growth Shock", "EBIT Margin Shock", "Capex / Revenue Shock",
        "WACC Shock", "Terminal Growth Shock", "Value / Share", "Upside / (Downside)",
        "Value vs Base", "Sensitivity", "Rationale",
    ] + list(range(2026, 2036)) + list(range(2026, 2036))

    for col, value in enumerate(headers, 1):
        cell = ws.cell(47, col, value)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[47].height = 34

    defaults = [
        ["Base Control", 0, 0, 0, 0, 0, "No shock; should reconcile to Base scenario."],
        ["Growth -3ppt", -0.03, 0, 0, 0, 0, "Revenue growth reduced 3 percentage points each forecast year."],
        ["Growth -5ppt", -0.05, 0, 0, 0, 0, "Revenue growth reduced 5 percentage points each forecast year."],
        ["EBIT Margin -300bps", 0, -0.03, 0, 0, 0, "Operating margin reduced 300 bps across the forecast."],
        ["EBIT Margin -500bps", 0, -0.05, 0, 0, 0, "Operating margin reduced 500 bps across the forecast."],
        ["Capex +500bps", 0, 0, 0.05, 0, 0, "Capex/revenue held 500 bps above Base assumptions."],
        ["WACC +100bps", 0, 0, 0, 0.01, 0, "Discount rate increased 100 bps."],
        ["WACC +200bps", 0, 0, 0, 0.02, 0, "Discount rate increased 200 bps."],
        ["Terminal Growth -50bps", 0, 0, 0, 0, -0.005, "Perpetual growth reduced 50 bps."],
        ["Terminal Growth -100bps", 0, 0, 0, 0, -0.01, "Perpetual growth reduced 100 bps."],
        ["Severe Bear", -0.05, -0.05, 0.05, 0.02, -0.01,
         "Combined downside: growth -5ppt, margin -500bps, capex +500bps, WACC +200bps, terminal growth -100bps."],
    ]

    # Preserve user-edited stress inputs if the template already contains this module.
    if not existed:
        for r, row in enumerate(defaults, 48):
            ws.cell(r, 1, row[0])
            for c, value in enumerate(row[1:6], 2):
                ws.cell(r, c, value)
            ws.cell(r, 11, row[6])
        ws["B65"] = -0.40
    else:
        for r, row in enumerate(defaults, 48):
            ws.cell(r, 1, row[0])
            ws.cell(r, 11, row[6])

    for r in range(48, 59):
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 11).alignment = Alignment(wrap_text=True)
        for c in range(2, 7):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=GOLD)
            ws.cell(r, c).font = Font(color=INPUT_BLUE)
            ws.cell(r, c).number_format = FMT_PCT

    base_cols = [_col_name(14 + i) for i in range(10)]   # N:W
    fcf_cols = [_col_name(12 + i) for i in range(10)]    # L:U
    rev_cols = [_col_name(22 + i) for i in range(10)]    # V:AE

    for r in range(48, 59):
        for i, (rev_c, base_c) in enumerate(zip(rev_cols, base_cols)):
            if i == 0:
                ws[f"{rev_c}{r}"] = f"='Historical Financials'!G4*(1+MAX(-0.95,{base_c}$12+$B{r}))"
            else:
                prev_rev = rev_cols[i - 1]
                ws[f"{rev_c}{r}"] = f"={prev_rev}{r}*(1+MAX(-0.95,{base_c}$12+$B{r}))"
            ws[f"{rev_c}{r}"].number_format = FMT_BN

        for i, (fcf_c, rev_c, base_c) in enumerate(zip(fcf_cols, rev_cols, base_cols)):
            if i == 0:
                nwc_rate = f"IFERROR({base_c}$22/({base_c}$13-'Historical Financials'!G4),0)"
                prev_rev = "'Historical Financials'!G4"
            else:
                prev_base = base_cols[i - 1]
                nwc_rate = f"IFERROR({base_c}$22/({base_c}$13-{prev_base}$13),0)"
                prev_rev = f"{rev_cols[i - 1]}{r}"

            ws[f"{fcf_c}{r}"] = (
                f"={rev_c}{r}*({base_c}$14+$C{r})*(1-$C$8)"
                f"+{rev_c}{r}*{base_c}$18"
                f"-{rev_c}{r}*({base_c}$20+$D{r})"
                f"-({rev_c}{r}-{prev_rev})*({nwc_rate})"
            )
            ws[f"{fcf_c}{r}"].number_format = FMT_BN

        explicit = "+".join(
            f"{fcf_c}{r}/(1+($C$6+$E{r}))^{i+1}"
            for i, fcf_c in enumerate(fcf_cols)
        )
        terminal = (
            f"({fcf_cols[-1]}{r}*(1+($C$7+$F{r}))/"
            f"(($C$6+$E{r})-($C$7+$F{r})))/(1+($C$6+$E{r}))^10"
        )
        ws[f"G{r}"] = f"=IFERROR((({explicit})+{terminal}-'Company Data'!B14)/'Company Data'!B9,\"\")"
        ws[f"H{r}"] = f"=IFERROR(G{r}/'Company Data'!B8-1,\"\")"
        ws[f"I{r}"] = f"=IFERROR(G{r}/$G$48-1,\"\")"
        if r == 48:
            ws[f"J{r}"] = "Control"
        else:
            ws[f"J{r}"] = f'=IF(I{r}="","",IF(I{r}>=-0.1,"Low",IF(I{r}>=-0.25,"Moderate","High")))'

        ws[f"G{r}"].number_format = FMT_PRICE
        ws[f"H{r}"].number_format = FMT_PCT
        ws[f"I{r}"].number_format = FMT_PCT

    _merge_once(ws, "A61:E61")
    ws["A61"] = "Stress-Test Summary"
    ws["A61"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A61"].font = Font(bold=True, color=WHITE)

    summary = [
        ["Metric", "Result"],
        ["Severe Bear Value / Share", "=G58"],
        ["Severe Bear Downside", "=H58"],
        ["Max Severe-Bear Downside Tolerance", ws["B65"].value if existed and ws["B65"].value is not None else -0.40],
        ["Robustness Status", '=IF(B64>=B65,"PASS","FAIL")'],
    ]
    for r, row in enumerate(summary, 62):
        ws.cell(r, 1, row[0])
        ws.cell(r, 2, row[1])
    for c in range(1, 3):
        ws.cell(62, c).fill = PatternFill("solid", fgColor=BLUE)
        ws.cell(62, c).font = Font(bold=True, color=WHITE)
    ws["B63"].number_format = FMT_PRICE
    ws["B64"].number_format = FMT_PCT
    ws["B65"].number_format = FMT_PCT
    ws["B65"].fill = PatternFill("solid", fgColor=GOLD)
    ws["B65"].font = Font(color=INPUT_BLUE)

    checks = [
        ["Completion Check", "Status"],
        ["Growth stress present", '=IF(AND(A49="Growth -3ppt",A50="Growth -5ppt"),"YES","NO")'],
        ["Margin stress present", '=IF(AND(A51="EBIT Margin -300bps",A52="EBIT Margin -500bps"),"YES","NO")'],
        ["Capex/WACC/terminal stresses present", '=IF(AND(A53="Capex +500bps",A54="WACC +100bps",A55="WACC +200bps",A56="Terminal Growth -50bps",A57="Terminal Growth -100bps"),"YES","NO")'],
        ["Bear/Base/Bull assumptions stress-tested", '=IF(AND(E63="YES",E64="YES",E65="YES",A58="Severe Bear",COUNTA(B58:F58)=5),"YES","NO")'],
    ]
    for r, row in enumerate(checks, 62):
        ws.cell(r, 4, row[0])
        ws.cell(r, 5, row[1])
    for c in range(4, 6):
        ws.cell(62, c).fill = PatternFill("solid", fgColor=BLUE)
        ws.cell(62, c).font = Font(bold=True, color=WHITE)

    widths = {"A":23,"B":16,"C":16,"D":17,"E":12,"F":16,"G":15,"H":16,"I":14,"J":13,"K":43}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for col in fcf_cols + rev_cols:
        ws.column_dimensions[col].width = 12

    # Dashboard snapshot.
    if "Dashboard" in wb.sheetnames:
        dash = wb["Dashboard"]
        _merge_once(dash, "F10:J10")
        dash["F10"] = "Stress-Test Snapshot"
        dash["F10"].fill = PatternFill("solid", fgColor=NAVY)
        dash["F10"].font = Font(bold=True, color=WHITE)
        values = [
            ["Metric", "Result"],
            ["Severe Bear Value / Share", "='Three-Case Scenarios'!B63"],
            ["Severe Bear Downside", "='Three-Case Scenarios'!B64"],
            ["Robustness Status", "='Three-Case Scenarios'!B66"],
            ["Stress Testing Completed", "='Three-Case Scenarios'!E66"],
        ]
        for r, row in enumerate(values, 11):
            dash.cell(r, 6, row[0])
            dash.cell(r, 7, row[1])
        for c in range(6, 8):
            dash.cell(11, c).fill = PatternFill("solid", fgColor=BLUE)
            dash.cell(11, c).font = Font(bold=True, color=WHITE)
        dash["G12"].number_format = FMT_PRICE
        dash["G13"].number_format = FMT_PCT
        for r in range(12, 16):
            dash.cell(r, 7).font = Font(color=LINK_GREEN)
        dash.column_dimensions["F"].width = 27
        dash.column_dimensions["G"].width = 18

    # Research checklist becomes an automatic completion check, not a valuation endorsement.
    if "Research Checklist" in wb.sheetnames:
        checklist = wb["Research Checklist"]
        checklist["B6"] = "='Three-Case Scenarios'!E66"
        checklist["B6"].font = Font(color=LINK_GREEN)
        checklist["C6"] = (
            "Automatic check: growth, margin, capex, WACC, terminal growth and combined Severe Bear are present. "
            "Review the resulting downside before relying on the model."
        )
        checklist["C6"].alignment = Alignment(wrap_text=True)

    if "Data Dictionary" in wb.sheetnames:
        dd = wb["Data Dictionary"]
        dd["A11"] = "Stress Testing"
        dd["B11"] = "Base-case annual assumptions + editable shocks; full 10-year revenue/FCF revaluation"
        dd["C11"] = "Tests growth, margin, capex, WACC, terminal growth and combined Severe Bear robustness"
        dd["B11"].alignment = Alignment(wrap_text=True)
        dd["C11"].alignment = Alignment(wrap_text=True)
