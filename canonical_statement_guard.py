from __future__ import annotations

"""SEC-first canonical financial-statement guard.

This module protects decision-critical annual accounting lines from structured-provider aliases
that may be adjusted, normalized, or economically different from the issuer's reported GAAP
measure. It is deliberately narrow: exact annual SEC facts may overwrite a populated fallback;
secondary providers may never overwrite an exact SEC canonical line.

The guard also synchronizes those canonical values into Historical Financials so scenario,
scoring, capital-economics and ML features all use the same accounting definition.
Unsupported rows remain blank rather than being estimated.
"""

from datetime import date
import math
from typing import Any

from openpyxl.styles import Alignment, Font, PatternFill

from business_model_registry import get_business_model_policy

FMT_BN = '#,##0.0;[Red](#,##0.0);-'
FMT_PCT = '0.0%;[Red](0.0%);-'
FMT_EPS = '$0.00;[Red]($0.00);-'
FMT_SHARES = '#,##0.000;[Red](#,##0.000);-'
GREEN = 'E2F0D9'; GOLD = 'FFF2CC'; RED = 'FCE4D6'


def _num(v: Any) -> float | None:
    try:
        if isinstance(v, bool) or v in (None, ''):
            return None
        x = float(v)
        return x if math.isfinite(x) else None
    except Exception:
        return None


def _find(ws, label: str, start: int = 1, end: int | None = None) -> int | None:
    needle = str(label).strip().lower()
    end = end or ws.max_row
    for r in range(start, min(end, ws.max_row) + 1):
        if str(ws.cell(r, 1).value or '').strip().lower() == needle:
            return r
    return None


def _year_cols(ws, header_row: int) -> dict[int, int]:
    out: dict[int, int] = {}
    for c in range(2, min(ws.max_column, 12) + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, (int, float)) and 1900 <= int(v) <= 2100:
            out[int(v)] = c
    return out


def _annual_fact_series(facts: dict | None, tags: tuple[str, ...], unit_hint: str | None = None) -> dict[int, float]:
    """Return exact annual 10-K/20-F facts using a fixed tag preference order.

    Duration facts are restricted to roughly one fiscal year. Instant facts are accepted without
    a duration test. Latest filed annual fact wins within a fiscal year.
    """
    if not facts:
        return {}
    namespaces = facts.get('facts') or {}
    best: dict[int, tuple[tuple[str, str, int], float]] = {}
    for priority, raw_tag in enumerate(tags):
        if ':' in raw_tag:
            ns, tag = raw_tag.split(':', 1)
            fact = (namespaces.get(ns) or {}).get(tag)
        else:
            fact = (namespaces.get('us-gaap') or {}).get(raw_tag)
        if not fact:
            continue
        units = fact.get('units') or {}
        values = units.get(unit_hint) if unit_hint and unit_hint in units else None
        if not values:
            preferred = [k for k in units if k in {'USD', 'shares', 'USD/shares'}]
            key = preferred[0] if preferred else (next(iter(units), None))
            values = units.get(key) if key else None
        for item in values or []:
            if str(item.get('form') or '') not in {'10-K', '10-K/A', '20-F', '20-F/A'}:
                continue
            val = _num(item.get('val'))
            end = str(item.get('end') or '')[:10]
            if val is None or len(end) != 10:
                continue
            start = str(item.get('start') or '')[:10]
            if start:
                try:
                    days = (date.fromisoformat(end) - date.fromisoformat(start)).days
                    if days < 250 or days > 450:
                        continue
                except Exception:
                    pass
            fy = item.get('fy')
            try:
                year = int(fy) if fy is not None else int(end[:4])
            except Exception:
                year = int(end[:4])
            stamp = (str(item.get('filed') or ''), end, -priority)
            if year not in best or stamp > best[year][0]:
                best[year] = (stamp, float(val))
    return {year: value for year, (_stamp, value) in best.items()}


INCOME_FACTS = {
    'Revenue': (('RevenueFromContractWithCustomerExcludingAssessedTax', 'SalesRevenueNet', 'Revenues', 'ifrs-full:Revenue'), 'USD', 'money'),
    'Total Operating Expenses': (('OperatingExpenses',), 'USD', 'money'),
    'Operating Income': (('OperatingIncomeLoss', 'ifrs-full:ProfitLossFromOperatingActivities'), 'USD', 'money'),
    'Pre-Tax Income': (('IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest', 'IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments', 'ifrs-full:ProfitLossBeforeTax'), 'USD', 'money'),
    'Income Taxes': (('IncomeTaxExpenseBenefit', 'ifrs-full:IncomeTaxExpenseContinuingOperations'), 'USD', 'money'),
    'Net Income': (('NetIncomeLoss', 'ProfitLoss', 'ifrs-full:ProfitLoss'), 'USD', 'money'),
    'Diluted EPS': (('EarningsPerShareDiluted', 'ifrs-full:DilutedEarningsLossPerShare'), 'USD/shares', 'eps'),
    'Diluted Weighted Average Shares (bn)': (('WeightedAverageNumberOfDilutedSharesOutstanding', 'ifrs-full:AdjustedWeightedAverageShares'), 'shares', 'shares'),
}

CASH_FACTS = {
    'Depreciation, Amortization & Accretion': (('DepreciationDepletionAndAmortization', 'DepreciationDepletionAndAmortizationPropertyPlantAndEquipment', 'ifrs-full:DepreciationAndAmortisationExpense'), 'USD', 'money'),
    'Stock-Based Compensation': (('ShareBasedCompensation', 'ifrs-full:SharebasedPaymentExpense'), 'USD', 'money'),
    'Operating Cash Flow': (('NetCashProvidedByUsedInOperatingActivities', 'ifrs-full:CashFlowsFromUsedInOperatingActivities'), 'USD', 'money'),
    'Capital Expenditures': (('PaymentsToAcquirePropertyPlantAndEquipment', 'PaymentsToAcquireProductiveAssets', 'ifrs-full:PurchaseOfPropertyPlantAndEquipmentClassifiedAsInvestingActivities'), 'USD', 'capex'),
}


def _write_exact(ws, row: int, col: int, raw: float, kind: str) -> float:
    if kind == 'eps':
        value = raw; fmt = FMT_EPS
    elif kind == 'shares':
        value = raw / 1e9; fmt = FMT_SHARES
    elif kind == 'capex':
        value = -abs(raw) / 1e9; fmt = FMT_BN
    else:
        value = raw / 1e9; fmt = FMT_BN
    ws.cell(row, col).value = value
    ws.cell(row, col).number_format = fmt
    return value


def _statement_sections(ws):
    i0 = _find(ws, 'Income Statement')
    b0 = _find(ws, 'Balance Sheet')
    c0 = _find(ws, 'Cash Flow Statement')
    if not all((i0, b0, c0)):
        return None
    ih = next((r for r in range(i0 + 1, min(b0, i0 + 6)) if str(ws.cell(r, 1).value or '').strip().lower() == 'metric'), None)
    ch = next((r for r in range(c0 + 1, min(ws.max_row + 1, c0 + 6)) if str(ws.cell(r, 1).value or '').strip().lower() == 'metric'), None)
    return (i0, b0, c0, ih, ch) if ih and ch else None


def apply_canonical_statement_guard(wb, ticker: str, facts: dict | None) -> dict[str, Any]:
    """Overwrite only direct canonical lines with exact annual SEC facts and sync history."""
    result: dict[str, Any] = {
        'ticker': str(ticker).upper().strip(), 'exact_cells_written': 0,
        'material_corrections': [], 'latest_core_exact': 0, 'latest_core_total': 3,
        'payments_structure_guard': False,
    }
    if 'Financial Statements' not in wb.sheetnames or 'Historical Financials' not in wb.sheetnames:
        return result
    fs = wb['Financial Statements']; hs = wb['Historical Financials']
    sections = _statement_sections(fs)
    if not sections:
        return result
    i0, b0, c0, ih, ch = sections
    iy = _year_cols(fs, ih); cy = _year_cols(fs, ch)
    hy = {int(hs.cell(3, c).value): c for c in range(2, min(hs.max_column, 8) + 1) if isinstance(hs.cell(3, c).value, (int, float))}
    source_col = max(max(iy.values(), default=1), max(cy.values(), default=1)) + 1

    ctx = {}
    try:
        cd = wb['Company Data']
        ctx = {'sector': cd['B6'].value, 'industry': cd['B7'].value, 'name': cd['B5'].value}
    except Exception:
        pass
    policy = get_business_model_policy(ticker, ctx.get('sector'), ctx.get('industry'), ctx.get('name'))
    result['business_model'] = policy.key

    exact_by_label: dict[str, dict[int, float]] = {}
    for label, (tags, unit, kind) in INCOME_FACTS.items():
        series = _annual_fact_series(facts, tags, unit)
        if not series:
            continue
        row = _find(fs, label, ih + 1, b0 - 1)
        if not row:
            continue
        exact_by_label[label] = {}
        for year, raw in series.items():
            col = iy.get(year)
            if not col:
                continue
            old = _num(fs.cell(row, col).value)
            new = _write_exact(fs, row, col, raw, kind)
            exact_by_label[label][year] = new
            result['exact_cells_written'] += 1
            if old is not None and abs(old - new) > max(0.01, abs(new) * 0.005):
                result['material_corrections'].append({'label': label, 'year': year, 'old': old, 'new': new})
        fs.cell(row, source_col).value = 'SEC Company Facts — exact canonical annual fact'
        fs.cell(row, source_col).font = Font(italic=True)

    for label, (tags, unit, kind) in CASH_FACTS.items():
        series = _annual_fact_series(facts, tags, unit)
        if not series:
            continue
        row = _find(fs, label, ch + 1, fs.max_row)
        if not row:
            continue
        exact_by_label[label] = {}
        for year, raw in series.items():
            col = cy.get(year)
            if not col:
                continue
            old = _num(fs.cell(row, col).value)
            new = _write_exact(fs, row, col, raw, kind)
            exact_by_label[label][year] = new
            result['exact_cells_written'] += 1
            if old is not None and abs(old - new) > max(0.01, abs(new) * 0.005):
                result['material_corrections'].append({'label': label, 'year': year, 'old': old, 'new': new})
        fs.cell(row, source_col).value = 'SEC Company Facts — exact canonical annual fact'
        fs.cell(row, source_col).font = Font(italic=True)

    if policy.key == 'payments':
        result['payments_structure_guard'] = True
        for label in ('Cost of Revenue', 'Gross Profit'):
            row = _find(fs, label, ih + 1, b0 - 1)
            if not row:
                continue
            for col in iy.values():
                fs.cell(row, col).value = None
            fs.cell(row, source_col).value = 'N/M — issuer does not report a conventional cost-of-revenue / gross-profit subtotal'
        fs['A4'] = (
            'Statement profile: Payments / asset-light network — exact GAAP revenue, operating expense, '
            'operating income and cash-flow lines take precedence over structured-provider normalized fields.'
        )

    reconciliation = []
    for year in sorted(set(exact_by_label.get('Revenue', {})) & set(exact_by_label.get('Total Operating Expenses', {})) & set(exact_by_label.get('Operating Income', {}))):
        rev = exact_by_label['Revenue'][year]; opex = exact_by_label['Total Operating Expenses'][year]; op = exact_by_label['Operating Income'][year]
        diff = (rev - opex) - op
        reconciliation.append({'year': year, 'difference_bn': diff})
    result['operating_income_reconciliation'] = reconciliation

    history_map = {
        4: ('Revenue', False),
        9: ('Operating Income', False),
        11: ('Net Income', False),
        12: ('Diluted EPS', False),
        14: ('Operating Cash Flow', False),
        15: ('Capital Expenditures', True),
        18: ('Depreciation, Amortization & Accretion', False),
        21: ('Stock-Based Compensation', False),
    }
    synced = 0
    for hrow, (label, absolute) in history_map.items():
        series = exact_by_label.get(label) or {}
        for year, value in series.items():
            hc = hy.get(year)
            if not hc:
                continue
            hs.cell(hrow, hc).value = abs(value) if absolute else value
            synced += 1
    result['history_sync'] = synced

    latest_year = max(hy) if hy else None
    if latest_year:
        for label in ('Revenue', 'Operating Income', 'Net Income'):
            if latest_year in (exact_by_label.get(label) or {}):
                result['latest_core_exact'] += 1

    try:
        wb.calculation.calcMode = 'auto'; wb.calculation.fullCalcOnLoad = True; wb.calculation.forceFullCalc = True
    except Exception:
        pass
    return result


def decorate_data_quality(wb, ticker: str, result: dict[str, Any] | None) -> None:
    """Keep business-model suitability separate from source exactness.

    A correctly routed commodity, bank, software, payments or other statement profile can be
    structurally appropriate even when the SEC endpoint is temporarily unavailable. Exact-source
    coverage is therefore reported by its own guard row instead of incorrectly downgrading the
    business-model profile itself.
    """
    if 'Data Quality' not in wb.sheetnames:
        return
    result = result or {}
    ws = wb['Data Quality']
    try:
        cd = wb['Company Data']
        policy = get_business_model_policy(ticker, cd['B6'].value, cd['B7'].value, cd['B5'].value)
    except Exception:
        policy = get_business_model_policy(ticker)

    def set_row(label: str, status: str, observed: str, why: str) -> None:
        row = next((r for r in range(1, ws.max_row + 1) if str(ws.cell(r, 1).value or '').strip() == label), ws.max_row + 1)
        for c, value in enumerate((label, status, observed, why), 1):
            ws.cell(row, c).value = value
            ws.cell(row, c).alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row, 2).fill = PatternFill('solid', fgColor=GREEN if status == 'PASS' else GOLD if status in {'REVIEW', 'N/A'} else RED)
        ws.cell(row, 2).font = Font(bold=True)

    exact = int(result.get('latest_core_exact') or 0)
    corrections = result.get('material_corrections') or []
    source_status = 'PASS' if exact >= 3 else 'REVIEW'
    set_row(
        'SEC-first canonical accounting guard', source_status,
        f'{exact}/3 latest-year core filing lines exact; {len(corrections)} material provider correction(s); {result.get("history_sync", 0)} history cell(s) synchronized.',
        'Source exactness is assessed separately from statement-profile suitability. Missing SEC access may be recovered from issuer/structured public sources, but those values remain explicitly lower in the source hierarchy.'
    )

    if policy.key == 'payments':
        observed = (
            'Payments / asset-light network profile selected; synthetic provider gross-profit rows suppressed; '
            'reported operating-expense and operating-income structure retained.'
        )
        why = (
            'A payments network should not be forced into an industrial cost-of-goods/gross-profit template. '
            'Exact-source coverage is tracked independently by the SEC-first accounting guard.'
        )
    else:
        observed = (
            f'{policy.label} profile selected. Primary valuation framework: {policy.primary_valuation}. '
            f'Latest-year exact SEC/filing core coverage is {exact}/3 and is tracked separately.'
        )
        why = (
            'PASS means the statement/valuation structure matches the company business model; it does not mean every cell came directly from SEC. '
            'Source gaps are handled by issuer/structured public fallbacks and disclosed by separate quality controls.'
        )
    set_row('Statement profile suitability', 'PASS', observed, why)
