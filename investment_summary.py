"""Decision-focused investment summary for generated equity-research workbooks.

The summary is intentionally auditable. It converts existing workbook outputs into:
1) a compact investment-decision dashboard,
2) a numeric operating / valuation snapshot, and
3) a transparent quantitative score bridge.

The score bridge always shows every intended dimension. Missing dimensions are not silently
dropped: their base weight remains visible, their effective weight becomes zero, and available
dimensions are reweighted transparently. The bridge starts from a neutral 50 and shows the
positive / negative point impact of every dimension.
"""

from __future__ import annotations

import math
import statistics
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

NAVY = "17365D"
BLUE = "2F75B5"
WHITE = "FFFFFF"
LIGHT = "F5F9FC"
GOLD = "FFF2CC"
PALE_GREEN = "E2F0D9"
PALE_RED = "FCE4D6"
PALE_YELLOW = "FFF2CC"
GREY = "666666"
BLACK = "000000"
LINK_GREEN = "008000"

FMT_PCT = '0.0%;[Red](0.0%);-'
FMT_PPT = '0.0 "ppt";[Red](0.0 "ppt");-'
FMT_PRICE = '$#,##0.00;[Red]($#,##0.00);-'
FMT_BN = '#,##0.0;[Red](#,##0.0);-'
FMT_MULT = '0.0x;[Red](0.0x);-'
FMT_SCORE = '0.0'
FMT_INT = '0;[Red](0);-'

THIN = Side(style="thin", color="D9E1F2")
MEDIUM = Side(style="medium", color=NAVY)

DIMENSIONS = [
    ("Absolute valuation", 30),
    ("Relative valuation", 15),
    ("Growth", 10),
    ("Profitability", 10),
    ("FCF quality", 10),
    ("Balance sheet", 10),
    ("Stress resilience", 10),
    ("Bayesian skew", 5),
]


def _fill(color):
    return PatternFill("solid", fgColor=color)


def _num(v, default=None):
    try:
        if isinstance(v, bool) or v in (None, ""):
            return default
        x = float(v)
        return x if math.isfinite(x) else default
    except Exception:
        return default


def _clamp(x, lo=0, hi=100):
    return max(lo, min(hi, x))


def _find_row(ws, label, col=1, contains=False):
    if ws is None:
        return None
    needle = str(label).strip().lower()
    for r in range(1, ws.max_row + 1):
        text = str(ws.cell(r, col).value or "").strip().lower()
        if (needle in text) if contains else (text == needle):
            return r
    return None


def _label_value(wb, sheet, label, value_col=2, contains=False):
    if sheet not in wb.sheetnames:
        return None
    ws = wb[sheet]
    r = _find_row(ws, label, 1, contains)
    return ws.cell(r, value_col).value if r else None


def _any_label_value(wb, sheet, label, search_cols=(1,), value_offset=1):
    """Find a label in one of several columns and return the cell immediately to its right."""
    if sheet not in wb.sheetnames:
        return None
    ws = wb[sheet]
    needle = str(label).strip().lower()
    for c in search_cols:
        for r in range(1, ws.max_row + 1):
            text = str(ws.cell(r, c).value or "").strip().lower()
            if text == needle:
                return ws.cell(r, c + value_offset).value
    return None


def _style_title(ws, text):
    for c in range(1, 11):
        ws.cell(1, c).fill = _fill(NAVY)
        ws.cell(2, c).fill = _fill(NAVY)
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, color=WHITE, size=19)
    ws.sheet_view.showGridLines = False


def _section(ws, row, title, end=10):
    for c in range(1, end + 1):
        cell = ws.cell(row, c)
        cell.fill = _fill(NAVY)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.border = Border(bottom=MEDIUM)
    ws.cell(row, 1, title)


def _header(ws, row, headers):
    for c, v in enumerate(headers, 1):
        cell = ws.cell(row, c, v)
        cell.fill = _fill(BLUE)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)


def _score_fill(score):
    if score is None:
        return _fill(LIGHT)
    if score >= 65:
        return _fill(PALE_GREEN)
    if score < 40:
        return _fill(PALE_RED)
    return _fill(PALE_YELLOW)


def _latest_num(ws, row, start=2, end=7):
    if ws is None:
        return None
    for c in range(end, start - 1, -1):
        v = _num(ws.cell(row, c).value)
        if v is not None:
            return v
    return None


def _history_cagr(ws, row):
    if ws is None:
        return None
    points = []
    for c in range(2, 8):
        y = ws.cell(3, c).value
        v = _num(ws.cell(row, c).value)
        if isinstance(y, (int, float)) and v is not None and v > 0:
            points.append((int(y), v))
    if len(points) < 2:
        return None
    y0, v0 = points[0]
    y1, v1 = points[-1]
    years = max(1, y1 - y0)
    return (v1 / v0) ** (1 / years) - 1 if v0 > 0 and v1 > 0 else None


def _scenario_base_cagr(wb):
    if "Three-Case Scenarios" not in wb.sheetnames:
        return None
    ws = wb["Three-Case Scenarios"]
    growth = [_num(ws.cell(12, c).value) for c in range(14, 24)]
    growth = [g for g in growth if g is not None]
    if not growth:
        return None
    product = 1.0
    for g in growth:
        product *= 1 + g
    return product ** (1 / len(growth)) - 1


def _data_quality(wb):
    if "Data Quality" not in wb.sheetnames:
        return 0, 0, 0
    ws = wb["Data Quality"]
    passed = review = failed = 0
    for r in range(1, ws.max_row + 1):
        s = str(ws.cell(r, 2).value or "").upper().strip()
        if s == "PASS":
            passed += 1
        elif s == "REVIEW":
            review += 1
        elif s == "FAIL":
            failed += 1
    return passed, review, failed


def _advanced_score(wb, label):
    if "Advanced Analytics" not in wb.sheetnames:
        return None
    ws = wb["Advanced Analytics"]
    r = _find_row(ws, label, 1)
    return _num(ws.cell(r, 2).value) if r else None


def _bayes_prob(wb, scenario):
    if "Base Rates & Probabilities" not in wb.sheetnames:
        return None
    ws = wb["Base Rates & Probabilities"]
    r = _find_row(ws, scenario, 1)
    return _num(ws.cell(r, 4).value) if r else None


def _peer_snapshot(wb):
    """Read target and peer medians directly from the authoritative Peer Comps sheet."""
    out = {"target": {}, "medians": {}, "direct_medians": {}, "coverage": None}
    if "Peer Comps" not in wb.sheetnames:
        return out
    ws = wb["Peer Comps"]
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(3, c).value or "").strip()
        if h:
            headers[h] = c

    wanted = {
        "Forward P/E",
        "EV/Revenue",
        "EV/EBITDA",
        "Revenue Growth",
        "Operating Margin",
        "ROE",
        "Data Coverage %",
    }
    if not wanted.intersection(headers):
        return out

    target_row = None
    peer_rows = []
    direct_rows = []
    for r in range(4, min(ws.max_row, 40) + 1):
        ticker = str(ws.cell(r, headers.get("Ticker", 2)).value or "").strip()
        if not ticker:
            continue
        peer_type = str(ws.cell(r, headers.get("Peer Type", 16)).value or "").strip()
        if peer_type == "Target classification":
            target_row = r
        else:
            peer_rows.append(r)
            if peer_type in {"Direct business-model peer", "Exact industry"}:
                direct_rows.append(r)

    if target_row is None:
        target_row = 4 if str(ws.cell(4, headers.get("Ticker", 2)).value or "").strip() else None

    metric_headers = [
        "Forward P/E",
        "EV/Revenue",
        "EV/EBITDA",
        "Revenue Growth",
        "Operating Margin",
        "ROE",
    ]
    if target_row:
        for h in metric_headers:
            c = headers.get(h)
            if c:
                out["target"][h] = _num(ws.cell(target_row, c).value)
        c = headers.get("Data Coverage %")
        if c:
            out["coverage"] = _num(ws.cell(target_row, c).value)

    def med(rows, h):
        c = headers.get(h)
        if not c:
            return None
        vals = [_num(ws.cell(r, c).value) for r in rows]
        vals = [x for x in vals if x is not None]
        return statistics.median(vals) if vals else None

    for h in metric_headers:
        out["medians"][h] = med(peer_rows, h)
        direct = med(direct_rows, h)
        out["direct_medians"][h] = direct if direct is not None else out["medians"][h]
    return out


def _next_year_consensus(wb):
    out = {"revenue_consensus": None, "revenue_model": None, "revenue_gap": None,
           "eps_consensus": None, "eps_model": None, "eps_gap": None}
    if "Expectations & Consensus" not in wb.sheetnames:
        return out
    ws = wb["Expectations & Consensus"]
    rows = []
    for r in range(7, min(ws.max_row, 50) + 1):
        metric = str(ws.cell(r, 1).value or "").strip()
        year = ws.cell(r, 2).value
        if metric in {"Revenue", "EPS"} and isinstance(year, (int, float)):
            rows.append((int(year), metric, r))
    if not rows:
        return out
    first_year = min(y for y, _, _ in rows)
    for y, metric, r in rows:
        if y != first_year:
            continue
        cons = _num(ws.cell(r, 3).value)
        model = _num(ws.cell(r, 4).value)
        gap = (model / cons - 1) if cons not in (None, 0) and model is not None else None
        key = metric.lower()
        out[f"{key}_consensus"] = cons
        out[f"{key}_model"] = model
        out[f"{key}_gap"] = gap
    return out


def _metric_snapshot(wb):
    company = wb["Company Data"] if "Company Data" in wb.sheetnames else None
    hist = wb["Historical Financials"] if "Historical Financials" in wb.sheetnames else None
    scenarios = wb["Three-Case Scenarios"] if "Three-Case Scenarios" in wb.sheetnames else None

    price = _num(company["B8"].value) if company else None
    market_cap = _num(company["B10"].value) if company else None
    enterprise_value = _num(company["B11"].value) if company else None
    cash = _num(company["B12"].value) if company else None
    debt = _num(company["B13"].value) if company else None
    net_debt = (debt - cash) if debt is not None and cash is not None else None
    net_debt_mc = (net_debt / market_cap) if net_debt is not None and market_cap not in (None, 0) else None
    sector = company["B6"].value if company else None
    industry = company["B7"].value if company else None

    base = _num(_label_value(wb, "Dashboard", "Base DCF Value / Share"))
    pw = _num(_label_value(wb, "Dashboard", "Probability-Weighted Value / Share"))
    severe = _num(_label_value(wb, "Three-Case Scenarios", "Severe Bear Value / Share", value_col=2))
    base_up = (base / price - 1) if base is not None and price else None
    pw_up = (pw / price - 1) if pw is not None and price else None
    severe_down = (severe / price - 1) if severe is not None and price else None

    revenue = _latest_num(hist, 4)
    op_income = _latest_num(hist, 9)
    net_income = _latest_num(hist, 11)
    ocf = _latest_num(hist, 14)
    capex = _latest_num(hist, 15)
    fcf = (ocf - abs(capex)) if ocf is not None and capex is not None else None
    op_margin = (op_income / revenue) if revenue not in (None, 0) and op_income is not None else None
    fcf_margin = (fcf / revenue) if revenue not in (None, 0) and fcf is not None else None
    rev_cagr = _history_cagr(hist, 4)
    eps_cagr = _history_cagr(hist, 12)

    peer = _peer_snapshot(wb)
    target = peer["target"]
    med = peer["direct_medians"]
    fpe = target.get("Forward P/E")
    evrev = target.get("EV/Revenue")
    evebitda = target.get("EV/EBITDA")
    latest_growth = target.get("Revenue Growth")
    peer_fpe = med.get("Forward P/E")
    peer_evrev = med.get("EV/Revenue")
    peer_evebitda = med.get("EV/EBITDA")
    peer_growth = med.get("Revenue Growth")
    peer_margin = med.get("Operating Margin")
    peer_roe = med.get("ROE")
    roe = target.get("ROE")

    fpe_gap = (fpe / peer_fpe - 1) if fpe is not None and peer_fpe not in (None, 0) else None
    evrev_gap = (evrev / peer_evrev - 1) if evrev is not None and peer_evrev not in (None, 0) else None
    evebitda_gap = (evebitda / peer_evebitda - 1) if evebitda is not None and peer_evebitda not in (None, 0) else None
    growth_gap = (latest_growth - peer_growth) if latest_growth is not None and peer_growth is not None else None
    margin_gap = (op_margin - peer_margin) if op_margin is not None and peer_margin is not None else None
    roe_gap = (roe - peer_roe) if roe is not None and peer_roe is not None else None

    wacc = _num(scenarios["C6"].value) if scenarios else None
    tgr = _num(scenarios["C7"].value) if scenarios else None
    base_10y_growth = _scenario_base_cagr(wb)

    market_base = _num(_label_value(wb, "Market Expectations", "10Y Revenue CAGR", 2))
    market_implied = _num(_label_value(wb, "Market Expectations", "10Y Revenue CAGR", 3))
    market_expectation_gap = (market_implied - market_base) if market_implied is not None and market_base is not None else None

    mc_median = _num(_any_label_value(wb, "Advanced Analytics", "Median Value / Share", search_cols=(9, 1)))
    mc_p10 = _num(_any_label_value(wb, "Advanced Analytics", "P10 Value / Share", search_cols=(9, 1)))
    mc_prob_gt_price = _num(_any_label_value(wb, "Advanced Analytics", "Probability > Current Price", search_cols=(9, 1)))
    implied_fcf_growth = _num(_any_label_value(wb, "Advanced Analytics", "Implied 10Y FCF CAGR", search_cols=(1, 9)))

    consensus = _next_year_consensus(wb)

    return {
        "price": price,
        "market_cap": market_cap,
        "enterprise_value": enterprise_value,
        "cash": cash,
        "debt": debt,
        "net_debt": net_debt,
        "net_debt_mc": net_debt_mc,
        "sector": sector,
        "industry": industry,
        "base": base,
        "pw": pw,
        "base_up": base_up,
        "pw_up": pw_up,
        "severe": severe,
        "severe_down": severe_down,
        "revenue": revenue,
        "net_income": net_income,
        "fcf": fcf,
        "rev_cagr": rev_cagr,
        "eps_cagr": eps_cagr,
        "latest_growth": latest_growth,
        "op_margin": op_margin,
        "fcf_margin": fcf_margin,
        "roe": roe,
        "fpe": fpe,
        "peer_fpe": peer_fpe,
        "fpe_gap": fpe_gap,
        "evrev": evrev,
        "peer_evrev": peer_evrev,
        "evrev_gap": evrev_gap,
        "evebitda": evebitda,
        "peer_evebitda": peer_evebitda,
        "evebitda_gap": evebitda_gap,
        "peer_growth": peer_growth,
        "growth_gap": growth_gap,
        "peer_margin": peer_margin,
        "margin_gap": margin_gap,
        "peer_roe": peer_roe,
        "roe_gap": roe_gap,
        "peer_coverage": peer.get("coverage"),
        "growth_score": _advanced_score(wb, "Growth"),
        "profit_score": _advanced_score(wb, "Profitability"),
        "fcf_score": _advanced_score(wb, "FCF Quality"),
        "balance_score": _advanced_score(wb, "Balance Sheet"),
        "stress_score": _advanced_score(wb, "Stress Robustness"),
        "bear_prob": _bayes_prob(wb, "Bear"),
        "base_prob": _bayes_prob(wb, "Base"),
        "bull_prob": _bayes_prob(wb, "Bull"),
        "wacc": wacc,
        "tgr": tgr,
        "base_10y_growth": base_10y_growth,
        "market_base_growth": market_base,
        "market_implied_growth": market_implied,
        "market_expectation_gap": market_expectation_gap,
        "mc_median": mc_median,
        "mc_p10": mc_p10,
        "mc_prob_gt_price": mc_prob_gt_price,
        "implied_fcf_growth": implied_fcf_growth,
        **consensus,
    }


def _score_from_premium(premium):
    if premium is None:
        return None
    return _clamp(50 - 50 * premium)


def _fmt_pct(v):
    return "N/A" if v is None else f"{v:.1%}"


def _fmt_ppt(v):
    return "N/A" if v is None else f"{v * 100:.1f} ppt"


def _fmt_mult(v):
    return "N/A" if v is None else f"{v:.1f}x"


def _fmt_price(v):
    return "N/A" if v is None else f"${v:,.2f}"


def _fmt_bn(v):
    return "N/A" if v is None else f"{v:,.1f}"


def _dimension_rows(m):
    rows = []

    ups = [x for x in (m["base_up"], m["pw_up"]) if x is not None]
    avg_up = sum(ups) / len(ups) if ups else None
    abs_score = _clamp(50 + 100 * avg_up) if avg_up is not None else None
    rows.append({
        "name": "Absolute valuation",
        "base_weight": 30,
        "score": abs_score,
        "actual": f"Base {_fmt_pct(m['base_up'])}; PW {_fmt_pct(m['pw_up'])}",
        "benchmark": "15% margin-of-safety reference",
        "gap": "N/A" if avg_up is None else f"{(avg_up - .15):+.1%} vs 15%",
        "status": "Complete" if len(ups) == 2 else ("Partial" if ups else "Missing"),
        "why": "Model value relative to the current share price; both Base DCF and probability-weighted value are used when available.",
    })

    gaps = [x for x in (m["fpe_gap"], m["evrev_gap"], m["evebitda_gap"]) if x is not None]
    relative_score = statistics.mean(_score_from_premium(x) for x in gaps) if gaps else None
    rows.append({
        "name": "Relative valuation",
        "base_weight": 15,
        "score": relative_score,
        "actual": f"P/E {_fmt_mult(m['fpe'])}; EV/Rev {_fmt_mult(m['evrev'])}; EV/EBITDA {_fmt_mult(m['evebitda'])}",
        "benchmark": f"Peer {_fmt_mult(m['peer_fpe'])}; {_fmt_mult(m['peer_evrev'])}; {_fmt_mult(m['peer_evebitda'])}",
        "gap": f"P/E {_fmt_pct(m['fpe_gap'])}; EV/Rev {_fmt_pct(m['evrev_gap'])}; EV/EBITDA {_fmt_pct(m['evebitda_gap'])}",
        "status": "Complete" if len(gaps) == 3 else ("Partial" if gaps else "Missing"),
        "why": "Direct/exact peer medians are preferred; broader sector peers are used only when the direct set lacks enough observations.",
    })

    growth_score = m["growth_score"]
    if growth_score is None and m["rev_cagr"] is not None:
        growth_score = _clamp(50 + (m["rev_cagr"] - .05) / .15 * 50)
    rows.append({
        "name": "Growth",
        "base_weight": 10,
        "score": growth_score,
        "actual": f"5Y rev CAGR {_fmt_pct(m['rev_cagr'])}; latest growth {_fmt_pct(m['latest_growth'])}",
        "benchmark": f"Peer latest growth {_fmt_pct(m['peer_growth'])}",
        "gap": _fmt_ppt(m["growth_gap"]),
        "status": "Complete" if m["rev_cagr"] is not None and m["latest_growth"] is not None else ("Partial" if growth_score is not None else "Missing"),
        "why": "Score is anchored to the workbook growth score / historical CAGR; the latest public growth rate is shown against peers for context.",
    })

    profit_score = m["profit_score"]
    if profit_score is None and m["op_margin"] is not None:
        profit_score = _clamp(50 + (m["op_margin"] - .08) / .20 * 50)
    rows.append({
        "name": "Profitability",
        "base_weight": 10,
        "score": profit_score,
        "actual": f"Op margin {_fmt_pct(m['op_margin'])}; ROE {_fmt_pct(m['roe'])}",
        "benchmark": f"Peer op margin {_fmt_pct(m['peer_margin'])}; ROE {_fmt_pct(m['peer_roe'])}",
        "gap": f"Margin {_fmt_ppt(m['margin_gap'])}; ROE {_fmt_ppt(m['roe_gap'])}",
        "status": "Complete" if m["op_margin"] is not None and m["roe"] is not None else ("Partial" if profit_score is not None else "Missing"),
        "why": "Operating profitability drives the score; ROE and peer profitability are shown as corroborating evidence.",
    })

    fcf_score = m["fcf_score"]
    if fcf_score is None and m["fcf_margin"] is not None:
        fcf_score = _clamp(50 + (m["fcf_margin"] - .05) / .15 * 50)
    rows.append({
        "name": "FCF quality",
        "base_weight": 10,
        "score": fcf_score,
        "actual": f"FCF {_fmt_bn(m['fcf'])}bn; margin {_fmt_pct(m['fcf_margin'])}",
        "benchmark": "10% FCF-margin reference",
        "gap": "N/A" if m["fcf_margin"] is None else f"{(m['fcf_margin'] - .10):+.1%} vs 10%",
        "status": "Complete" if m["fcf_margin"] is not None else ("Partial" if fcf_score is not None else "Missing"),
        "why": "Cash conversion matters more than accounting earnings for intrinsic value; latest reported OCF less capex is used.",
    })

    balance_score = m["balance_score"]
    if balance_score is None and m["net_debt"] is not None:
        balance_score = 85 if m["net_debt"] < 0 else 55
    rows.append({
        "name": "Balance sheet",
        "base_weight": 10,
        "score": balance_score,
        "actual": f"Net debt/(cash) {_fmt_bn(m['net_debt'])}bn; {_fmt_pct(m['net_debt_mc'])} of market cap",
        "benchmark": "Net debt = 0 reference",
        "gap": "Net cash" if m["net_debt"] is not None and m["net_debt"] < 0 else ("Net debt" if m["net_debt"] is not None else "N/A"),
        "status": "Complete" if m["net_debt"] is not None else ("Partial" if balance_score is not None else "Missing"),
        "why": "Net cash increases financing flexibility; material net debt reduces resilience and valuation optionality.",
    })

    stress_score = None
    if m["severe_down"] is not None:
        stress_score = _clamp(100 * (1 + m["severe_down"]))
    elif m["stress_score"] is not None:
        stress_score = m["stress_score"]
    rows.append({
        "name": "Stress resilience",
        "base_weight": 10,
        "score": stress_score,
        "actual": f"Severe bear {_fmt_pct(m['severe_down'])}; MC P10 {_fmt_price(m['mc_p10'])}",
        "benchmark": "-30% severe-downside reference",
        "gap": "N/A" if m["severe_down"] is None else f"{(m['severe_down'] + .30):+.1%} vs -30%",
        "status": "Complete" if m["severe_down"] is not None else ("Partial" if stress_score is not None else "Missing"),
        "why": f"Downside case plus Monte Carlo tail context. MC probability above current price: {_fmt_pct(m['mc_prob_gt_price'])}.",
    })

    bayes_score = None
    if m["bull_prob"] is not None and m["bear_prob"] is not None:
        bayes_score = _clamp(50 + (m["bull_prob"] - m["bear_prob"]) * 100)
    rows.append({
        "name": "Bayesian skew",
        "base_weight": 5,
        "score": bayes_score,
        "actual": f"Bull {_fmt_pct(m['bull_prob'])}; Base {_fmt_pct(m['base_prob'])}; Bear {_fmt_pct(m['bear_prob'])}",
        "benchmark": "Bull - Bear = 0 ppt",
        "gap": _fmt_ppt((m["bull_prob"] - m["bear_prob"]) if m["bull_prob"] is not None and m["bear_prob"] is not None else None),
        "status": "Complete" if bayes_score is not None else "Missing",
        "why": "Scenario posterior skew is a small overlay, not a substitute for valuation or fundamental evidence.",
    })
    return rows


def _investment_score(m):
    rows = _dimension_rows(m)
    available_weight = sum(r["base_weight"] for r in rows if r["score"] is not None)
    for r in rows:
        if r["score"] is None or not available_weight:
            r["effective_weight"] = 0.0
            r["contribution"] = 0.0
            r["impact"] = 0.0
        else:
            ew = r["base_weight"] / available_weight
            r["effective_weight"] = ew
            r["contribution"] = r["score"] * ew
            r["impact"] = (r["score"] - 50) * ew
    score = sum(r["contribution"] for r in rows) if available_weight else None
    coverage = available_weight / 100 if available_weight else 0
    return score, rows, coverage


def _verdict(score, m, failed, score_coverage):
    if failed:
        return "REVIEW — DATA QUALITY", "Resolve failed data-quality checks before relying on the investment conclusion."
    if score is None or score_coverage < .60:
        return "REVIEW — INSUFFICIENT DATA", f"Only {score_coverage:.0%} of the intended score weight has validated inputs."
    ups = [x for x in (m.get("base_up"), m.get("pw_up")) if x is not None]
    avg_up = sum(ups) / len(ups) if ups else None
    if avg_up is not None and avg_up >= .20 and score >= 65:
        return "ATTRACTIVE", "Valuation offers a meaningful margin of safety and the broader quantitative evidence is favorable."
    if avg_up is not None and avg_up >= .10 and score >= 55:
        return "POTENTIALLY ATTRACTIVE", "Valuation is favorable, but the score still reflects material execution, quality or downside risks."
    if (avg_up is None or avg_up > -.10) and score >= 45:
        return "NEUTRAL / WATCHLIST", "The evidence is mixed or the current margin of safety is not yet large enough."
    return "UNATTRACTIVE AT CURRENT PRICE", "The modeled return/risk trade-off is not favorable enough at the current price."


def _why_line(score, m, score_coverage):
    parts = []
    if m["pw_up"] is not None:
        parts.append(f"PW value implies {m['pw_up']:.1%} upside/downside")
    elif m["base_up"] is not None:
        parts.append(f"Base DCF implies {m['base_up']:.1%} upside/downside")
    if score is not None:
        parts.append(f"quant score {score:.1f}/100 with {score_coverage:.0%} score-weight coverage")
    if m["severe_down"] is not None:
        parts.append(f"severe-bear downside {m['severe_down']:.1%}")
    if m["mc_prob_gt_price"] is not None:
        parts.append(f"Monte Carlo P(value > price) {m['mc_prob_gt_price']:.0%}")
    return "; ".join(parts) + "." if parts else "Not enough validated numeric evidence for a concise investment conclusion."


def _strengths_risks(m, rows):
    strengths = []
    risks = []

    if m["pw_up"] is not None:
        (strengths if m["pw_up"] >= .15 else risks if m["pw_up"] < 0 else strengths).append(
            f"Probability-weighted value implies {m['pw_up']:.1%} upside/downside."
        )
    if m["rev_cagr"] is not None:
        text = f"Revenue CAGR is {m['rev_cagr']:.1%}"
        if m["latest_growth"] is not None and m["peer_growth"] is not None:
            text += f"; latest growth is {m['latest_growth']:.1%} vs {m['peer_growth']:.1%} peer median."
        else:
            text += "."
        (strengths if m["rev_cagr"] >= .08 else risks if m["rev_cagr"] < .03 else strengths).append(text)

    if m["op_margin"] is not None:
        if m["peer_margin"] is not None:
            text = f"Operating margin is {m['op_margin']:.1%} vs {m['peer_margin']:.1%} peer median ({(m['op_margin']-m['peer_margin'])*100:+.1f} ppt)."
        else:
            text = f"Operating margin is {m['op_margin']:.1%}."
        (strengths if m["op_margin"] >= .15 else risks if m["op_margin"] < .05 else strengths).append(text)

    if m["fcf_margin"] is not None:
        (strengths if m["fcf_margin"] >= .10 else risks if m["fcf_margin"] < .04 else strengths).append(
            f"Latest FCF margin is {m['fcf_margin']:.1%}."
        )

    if m["fpe_gap"] is not None:
        (strengths if m["fpe_gap"] <= -.10 else risks if m["fpe_gap"] >= .20 else strengths).append(
            f"Forward P/E is {abs(m['fpe_gap']):.1%} {'below' if m['fpe_gap'] < 0 else 'above'} the direct/exact peer median."
        )

    if m["net_debt"] is not None:
        if m["net_debt"] < 0:
            strengths.append(f"Balance sheet carries net cash of {abs(m['net_debt']):,.1f}bn.")
        elif m["net_debt_mc"] is not None and m["net_debt_mc"] > .20:
            risks.append(f"Net debt is {m['net_debt']:,.1f}bn, equal to {m['net_debt_mc']:.1%} of market capitalization.")

    if m["severe_down"] is not None and m["severe_down"] <= -.50:
        risks.append(f"Severe-bear stress implies {m['severe_down']:.1%} downside, indicating substantial tail risk.")

    if m["market_implied_growth"] is not None and m["market_base_growth"] is not None:
        gap = m["market_implied_growth"] - m["market_base_growth"]
        if gap > .05:
            risks.append(f"Market-implied 10Y revenue growth exceeds the model Base case by {gap*100:.1f} ppt.")
        elif gap < -.03:
            strengths.append(f"Market-implied 10Y revenue growth is {abs(gap)*100:.1f} ppt below the model Base case.")

    if m["revenue_gap"] is not None:
        if m["revenue_gap"] > .05:
            strengths.append(f"Next-year model revenue is {m['revenue_gap']:.1%} above public consensus — a clear positive variant view if justified.")
        elif m["revenue_gap"] < -.05:
            risks.append(f"Next-year model revenue is {abs(m['revenue_gap']):.1%} below public consensus.")

    if m["mc_prob_gt_price"] is not None:
        (strengths if m["mc_prob_gt_price"] >= .60 else risks if m["mc_prob_gt_price"] < .40 else strengths).append(
            f"Monte Carlo probability of intrinsic value exceeding the current price is {m['mc_prob_gt_price']:.0%}."
        )

    ranked = [r for r in rows if r["score"] is not None]
    if ranked:
        best = max(ranked, key=lambda x: x["impact"])
        worst = min(ranked, key=lambda x: x["impact"])
        if best["impact"] > 1:
            strengths.append(f"Largest score support: {best['name']} adds {best['impact']:+.1f} points versus a neutral 50.")
        if worst["impact"] < -1:
            risks.append(f"Largest score drag: {worst['name']} subtracts {abs(worst['impact']):.1f} points versus a neutral 50.")

    return strengths[:7], risks[:7]


def _write_pair(ws, label_cell, value_cell, label, value, fmt=None, fill=None, bold=False):
    ws[label_cell] = label
    ws[label_cell].font = Font(bold=True, color=GREY)
    ws[value_cell] = value
    if fmt:
        ws[value_cell].number_format = fmt
    if fill:
        ws[value_cell].fill = _fill(fill)
    if bold:
        ws[value_cell].font = Font(bold=True)


def ensure_investment_summary(wb, ticker):
    name = "Investment Summary"
    if name in wb.sheetnames:
        wb.remove(wb[name])
    ws = wb.create_sheet(name, 1)
    _style_title(ws, f"{ticker} — Investment Summary & Decision View")
    ws["A3"] = (
        "Decision-focused synthesis of valuation, peer, operating, stress, expectations and scenario outputs. "
        "All score dimensions remain visible even when data is missing; missing weights are reallocated transparently."
    )
    ws["A3"].font = Font(italic=True, color=GREY)
    ws["A3"].alignment = Alignment(wrap_text=True)
    try:
        ws.merge_cells("A3:J3")
    except Exception:
        pass

    m = _metric_snapshot(wb)
    passed, review, failed = _data_quality(wb)
    score, score_rows, score_coverage = _investment_score(m)
    verdict, why = _verdict(score, m, failed, score_coverage)
    verdict_fill = (
        PALE_GREEN if verdict == "ATTRACTIVE"
        else PALE_YELLOW if "POTENTIALLY" in verdict or "NEUTRAL" in verdict or "REVIEW" in verdict
        else PALE_RED
    )

    _section(ws, 5, "Investment Decision")
    ws["A6"] = "Model View"
    ws["B6"] = verdict
    ws["B6"].fill = _fill(verdict_fill)
    ws["B6"].font = Font(bold=True, size=14)
    try:
        ws.merge_cells("B6:C6")
    except Exception:
        pass

    score_color = PALE_GREEN if score is not None and score >= 65 else PALE_RED if score is not None and score < 40 else PALE_YELLOW if score is not None else LIGHT
    _write_pair(ws, "D6", "E6", "Quant Score / 100", score, FMT_SCORE, score_color, True)
    _write_pair(ws, "F6", "G6", "Score Coverage", score_coverage, FMT_PCT, PALE_GREEN if score_coverage >= .85 else PALE_YELLOW if score_coverage >= .60 else PALE_RED, True)
    ws["H6"] = "Data Quality"
    ws["I6"] = f"{passed} PASS / {review} REVIEW / {failed} FAIL"
    try:
        ws.merge_cells("I6:J6")
    except Exception:
        pass

    ws["A7"] = "Why"
    ws["B7"] = _why_line(score, m, score_coverage)
    ws["B7"].alignment = Alignment(wrap_text=True)
    try:
        ws.merge_cells("B7:J7")
    except Exception:
        pass

    _write_pair(ws, "A8", "B8", "Current Price", m["price"], FMT_PRICE)
    _write_pair(ws, "C8", "D8", "Base DCF / Share", m["base"], FMT_PRICE)
    _write_pair(ws, "E8", "F8", "Base Upside", m["base_up"], FMT_PCT)
    _write_pair(ws, "G8", "H8", "PW Value / Share", m["pw"], FMT_PRICE)
    _write_pair(ws, "I8", "J8", "PW Upside", m["pw_up"], FMT_PCT)

    _write_pair(ws, "A9", "B9", "Severe Bear / Share", m["severe"], FMT_PRICE)
    _write_pair(ws, "C9", "D9", "Severe Downside", m["severe_down"], FMT_PCT)
    _write_pair(ws, "E9", "F9", "MC Median / Share", m["mc_median"], FMT_PRICE)
    _write_pair(ws, "G9", "H9", "MC P(Value > Price)", m["mc_prob_gt_price"], FMT_PCT)
    _write_pair(ws, "I9", "J9", "Market Cap (bn)", m["market_cap"], FMT_BN)

    _write_pair(ws, "A10", "B10", "Net Debt / (Cash) (bn)", m["net_debt"], FMT_BN)
    _write_pair(ws, "C10", "D10", "WACC", m["wacc"], FMT_PCT)
    _write_pair(ws, "E10", "F10", "Terminal Growth", m["tgr"], FMT_PCT)
    ws["G10"] = "Sector / Industry"
    ws["G10"].font = Font(bold=True, color=GREY)
    ws["H10"] = f"{m['sector'] or 'Unknown'} / {m['industry'] or 'Unknown'}"
    try:
        ws.merge_cells("H10:J10")
    except Exception:
        pass

    _section(ws, 12, "Key Numbers Behind the View")
    _header(ws, 13, ["Metric", "Company / Model", "Peer / Reference", "Gap / Skew", "Signal", "Context", "", "", "", ""])

    key_rows = [
        ("5Y Revenue CAGR", m["rev_cagr"], None, None, "Historical growth", "Compounded across available annual history", FMT_PCT),
        ("Latest Revenue Growth", m["latest_growth"], m["peer_growth"], m["growth_gap"], "vs direct/exact peers", "Public latest growth from Peer Comps", FMT_PCT),
        ("Operating Margin", m["op_margin"], m["peer_margin"], m["margin_gap"], "vs direct/exact peers", "Latest reported operating margin", FMT_PCT),
        ("ROE", m["roe"], m["peer_roe"], m["roe_gap"], "vs direct/exact peers", "Public ROE / statement fallback", FMT_PCT),
        ("FCF Margin", m["fcf_margin"], .10, (m["fcf_margin"] - .10) if m["fcf_margin"] is not None else None, "10% reference", "OCF less capex / revenue", FMT_PCT),
        ("Forward P/E", m["fpe"], m["peer_fpe"], m["fpe_gap"], "vs direct/exact peers", "Lower relative multiple is favorable, all else equal", FMT_MULT),
        ("EV / Revenue", m["evrev"], m["peer_evrev"], m["evrev_gap"], "vs direct/exact peers", "Useful when margin structures are comparable", FMT_MULT),
        ("EV / EBITDA", m["evebitda"], m["peer_evebitda"], m["evebitda_gap"], "vs direct/exact peers", "Capital-structure-neutral valuation multiple", FMT_MULT),
        ("Base 10Y Revenue CAGR", m["base_10y_growth"], m["market_implied_growth"], (m["base_10y_growth"] - m["market_implied_growth"]) if m["base_10y_growth"] is not None and m["market_implied_growth"] is not None else None, "model vs market-implied", "Expectations gap, not a historical metric", FMT_PCT),
        ("Implied 10Y FCF CAGR", m["implied_fcf_growth"], None, None, "reverse DCF", "Growth required by the current enterprise value", FMT_PCT),
        ("Next-Year Revenue vs Consensus", m["revenue_model"], m["revenue_consensus"], m["revenue_gap"], "variant perception", "Public consensus where available", FMT_BN),
        ("Monte Carlo P10 / Median", m["mc_p10"], m["mc_median"], None, "valuation distribution", f"P(value > current) {_fmt_pct(m['mc_prob_gt_price'])}", FMT_PRICE),
    ]

    for r, (label, val, bench, gap, signal, context, fmt) in enumerate(key_rows, 14):
        ws.cell(r, 1, label)
        ws.cell(r, 2, val)
        ws.cell(r, 3, bench)
        ws.cell(r, 4, gap)
        ws.cell(r, 5, signal)
        ws.cell(r, 6, context)
        ws.cell(r, 2).number_format = fmt
        ws.cell(r, 3).number_format = fmt
        if label in {"Latest Revenue Growth", "Operating Margin", "ROE", "FCF Margin", "Base 10Y Revenue CAGR", "Next-Year Revenue vs Consensus"}:
            ws.cell(r, 4).number_format = FMT_PCT
        elif label in {"Forward P/E", "EV / Revenue", "EV / EBITDA"}:
            ws.cell(r, 4).number_format = FMT_PCT
        ws.cell(r, 6).alignment = Alignment(wrap_text=True)

    bridge_row = 28
    _section(ws, bridge_row, "Quantitative Score Bridge — Neutral 50 to Final Score")
    ws.cell(bridge_row + 1, 1, "Start at neutral")
    ws.cell(bridge_row + 1, 2, 50.0)
    ws.cell(bridge_row + 1, 2).number_format = FMT_SCORE
    ws.cell(bridge_row + 1, 3, "Each available dimension adds or subtracts points versus a neutral score of 50. Missing dimensions stay visible and receive 0% effective weight.")
    try:
        ws.merge_cells(start_row=bridge_row + 1, start_column=3, end_row=bridge_row + 1, end_column=10)
    except Exception:
        pass
    ws.cell(bridge_row + 1, 3).font = Font(italic=True, color=GREY)
    ws.cell(bridge_row + 1, 3).alignment = Alignment(wrap_text=True)

    hdr = bridge_row + 2
    _header(ws, hdr, [
        "Dimension", "Score / 100", "Base Weight", "Effective Weight", "Contribution",
        "Impact vs Neutral", "Key Actuals", "Peer / Reference", "Gap / Skew", "Data Status"
    ])

    first = hdr + 1
    for r, item in enumerate(score_rows, first):
        ws.cell(r, 1, item["name"])
        ws.cell(r, 2, item["score"])
        ws.cell(r, 3, item["base_weight"] / 100)
        ws.cell(r, 4, item["effective_weight"])
        ws.cell(r, 5, item["contribution"])
        ws.cell(r, 6, item["impact"])
        ws.cell(r, 7, item["actual"])
        ws.cell(r, 8, item["benchmark"])
        ws.cell(r, 9, item["gap"])
        ws.cell(r, 10, item["status"])
        ws.cell(r, 2).number_format = FMT_SCORE
        ws.cell(r, 3).number_format = FMT_PCT
        ws.cell(r, 4).number_format = FMT_PCT
        ws.cell(r, 5).number_format = FMT_SCORE
        ws.cell(r, 6).number_format = '+0.0;[Red]-0.0;-'
        ws.cell(r, 2).fill = _score_fill(item["score"])
        if item["status"] == "Missing":
            ws.cell(r, 10).fill = _fill(PALE_RED)
        elif item["status"] == "Partial":
            ws.cell(r, 10).fill = _fill(PALE_YELLOW)
        else:
            ws.cell(r, 10).fill = _fill(PALE_GREEN)
        for c in (7, 8, 9, 10):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
        for c in range(1, 11):
            ws.cell(r, c).border = Border(bottom=THIN)

    total_row = first + len(score_rows)
    ws.cell(total_row, 1, "Final Quantitative Score")
    ws.cell(total_row, 2, score)
    ws.cell(total_row, 3, 1.0)
    ws.cell(total_row, 4, sum(x["effective_weight"] for x in score_rows))
    ws.cell(total_row, 5, sum(x["contribution"] for x in score_rows))
    ws.cell(total_row, 6, sum(x["impact"] for x in score_rows))
    ws.cell(total_row, 7, f"50 neutral + {sum(x['impact'] for x in score_rows):+.1f} dimension impact")
    ws.cell(total_row, 10, f"{score_coverage:.0%} score-weight coverage")
    for c in range(1, 11):
        ws.cell(total_row, c).font = Font(bold=True)
        ws.cell(total_row, c).border = Border(top=MEDIUM)
    ws.cell(total_row, 2).number_format = FMT_SCORE
    ws.cell(total_row, 3).number_format = FMT_PCT
    ws.cell(total_row, 4).number_format = FMT_PCT
    ws.cell(total_row, 5).number_format = FMT_SCORE
    ws.cell(total_row, 6).number_format = '+0.0;[Red]-0.0;-'
    ws.cell(total_row, 2).fill = _score_fill(score)

    note_row = total_row + 1
    ws.cell(note_row, 1, "Score interpretation")
    ws.cell(note_row, 2, "80–100 very strong | 65–79 favorable | 50–64 mixed-positive | 35–49 weak | below 35 poor. The score is a decision aid; valuation and thesis evidence still require analyst judgment.")
    try:
        ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=10)
    except Exception:
        pass
    ws.cell(note_row, 2).font = Font(italic=True, color=GREY)
    ws.cell(note_row, 2).alignment = Alignment(wrap_text=True)

    strength_row = note_row + 2
    _section(ws, strength_row, "What the Numbers Like")
    strengths, risks = _strengths_risks(m, score_rows)
    for i, text in enumerate(strengths or ["No major quantitative strength passed the current rule thresholds."], strength_row + 1):
        ws.cell(i, 1, "+")
        ws.cell(i, 2, text)
        ws.cell(i, 2).alignment = Alignment(wrap_text=True)
        try:
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=10)
        except Exception:
            pass

    risk_row = strength_row + max(4, len(strengths) + 1) + 1
    _section(ws, risk_row, "Biggest Quantitative Risks / What Must Go Right")
    for i, text in enumerate(risks or ["No major quantitative risk passed the current rule thresholds; qualitative risks still require analyst review."], risk_row + 1):
        ws.cell(i, 1, "!")
        ws.cell(i, 2, text)
        ws.cell(i, 2).alignment = Alignment(wrap_text=True)
        try:
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=10)
        except Exception:
            pass

    foot = risk_row + max(4, len(risks) + 1) + 1
    ws.cell(foot, 1, "Interpretation")
    ws.cell(
        foot,
        2,
        "A favorable model view does not guarantee a favorable return. Re-underwrite the thesis when earnings, consensus, "
        "management guidance, capital allocation, sector conditions or the share price change. Missing data should be filled "
        "at the source rather than overridden in this summary."
    )
    ws.cell(foot, 2).font = Font(italic=True, color=GREY)
    ws.cell(foot, 2).alignment = Alignment(wrap_text=True)
    try:
        ws.merge_cells(start_row=foot, start_column=2, end_row=foot, end_column=10)
    except Exception:
        pass

    widths = {
        "A": 29, "B": 17, "C": 17, "D": 17, "E": 18,
        "F": 18, "G": 35, "H": 31, "I": 27, "J": 18
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 20
    ws.row_dimensions[3].height = 34
    ws.row_dimensions[7].height = 34
    for r in range(first, total_row):
        ws.row_dimensions[r].height = 46

    ws.freeze_panes = "A12"
    return ws
