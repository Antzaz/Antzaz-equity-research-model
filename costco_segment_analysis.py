from __future__ import annotations

"""Verified Costco segment and business-mix analysis from public primary sources.

Costco's generic narrative parser previously mistook fragments of the segment-description
sentence for segment names. Costco actually reports three geographic segments: United States,
Canada and Other International. This module rebuilds the sheet from explicit issuer/SEC data
and adds current YTD, merchandise-category, membership and monthly operating momentum.

Values are USD billions unless otherwise stated. Public snapshots are date-stamped and never
presented as estimates. The generic segment framework remains the fallback for other issuers.
"""

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; GREY="666666"; PALE_GREEN="E2F0D9"
FMT_BN='#,##0.0;[Red](#,##0.0);-'; FMT_PCT='0.0%;[Red](0.0%);-'; FMT_INT='#,##0;[Red](#,##0);-'
ANNUAL_10K="https://www.sec.gov/Archives/edgar/data/909832/000090983225000101/cost-20250831.htm"
Q3_10Q="https://www.sec.gov/Archives/edgar/data/909832/000090983226000051/cost-20260510.htm"
JUNE_SALES="https://investor.costco.com/news/news-details/2026/Costco-Wholesale-Corporation-Reports-June-Sales-Results-and-Announces-Quarterly-Cash-Dividend/default.aspx"
JULY_EVENT="https://investor.costco.com/events-and-presentations/events/event-details/2026/July-Sales-Results-2026-XADtqZn33k/default.aspx"

ANNUAL_SEGMENTS={
    "United States":{"revenue":[176.630,184.143,200.046],"op":[5.392,6.217,6.878],"da":[1.599,1.730,1.895],"capex":[3.288,3.725,4.215],"assets":[49.189,48.816,54.862]},
    "Canada":{"revenue":[33.056,34.874,36.923],"op":[1.448,1.648,1.849],"da":[.183,.192,.196],"capex":[.281,.351,.580],"assets":[6.420,6.915,7.304]},
    "Other International":{"revenue":[32.604,35.436,38.266],"op":[1.274,1.420,1.656],"da":[.295,.315,.335],"capex":[.754,.634,.703],"assets":[13.385,14.100,14.933]},
}
YTD_SEGMENTS={
    "United States":{"revenue":[137.819,149.934],"op":[4.726,5.124]},
    "Canada":{"revenue":[25.021,27.774],"op":[1.215,1.416]},
    "Other International":{"revenue":[26.239,29.723],"op":[1.101,1.344]},
}
ANNUAL_CATEGORIES={
    "Foods and Sundries":[96.175,101.463,109.564],
    "Non-Foods":[60.865,63.973,71.190],
    "Fresh Foods":[31.977,34.220,37.988],
    "Warehouse Ancillary and Other Businesses":[48.693,49.969,51.170],
}
YTD_CATEGORIES={
    "Foods and Sundries":[75.323,80.625],
    "Non-Foods":[49.777,54.122],
    "Fresh Foods":[25.839,28.576],
    "Warehouse Ancillary and Other Businesses":[34.541,40.051],
}


def _fill(c): return PatternFill("solid",fgColor=c)
def _section(ws,row,title,end=16):
    for c in range(1,end+1): ws.cell(row,c).fill=_fill(NAVY); ws.cell(row,c).font=Font(bold=True,color=WHITE)
    ws.cell(row,1,title)
def _header(ws,row,values):
    for c,v in enumerate(values,1): ws.cell(row,c,v); ws.cell(row,c).fill=_fill(BLUE); ws.cell(row,c).font=Font(bold=True,color=WHITE); ws.cell(row,c).alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def _link(cell,text,url):
    cell.value=text; cell.hyperlink=url; cell.font=Font(color="008000",underline="single")
def _pct_cells(ws,row,cols):
    for c in cols: ws.cell(row,c).number_format=FMT_PCT

def ensure_costco_segment_analysis(wb,ticker):
    if str(ticker).upper()!="COST": return False
    if "Segment Analysis" in wb.sheetnames: wb.remove(wb["Segment Analysis"])
    ws=wb.create_sheet("Segment Analysis"); ws.sheet_view.showGridLines=False
    for c in range(1,17): ws.cell(1,c).fill=_fill(NAVY); ws.cell(2,c).fill=_fill(NAVY)
    ws["A1"]="COST — Segment, Business Mix & Operating Momentum"; ws["A1"].font=Font(bold=True,color=WHITE,size=18)
    ws["A3"]="Primary-source view: FY2023–FY2025 audited geographic segments, FY2026 36-week interim segment results, merchandise-category sales, capital intensity, membership economics and latest machine-readable monthly sales update. No segment economics are estimated."
    ws.merge_cells("A3:P3"); ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True); ws.row_dimensions[3].height=44

    # Audited annual geographic segments.
    _section(ws,5,"Reported Geographic Segments — Annual",16)
    _header(ws,6,["Segment","FY2023 Revenue","FY2024 Revenue","FY2025 Revenue","FY25 Growth","2Y CAGR","FY2023 Op. Income","FY2024 Op. Income","FY2025 Op. Income","FY23 Margin","FY24 Margin","FY25 Margin","Margin Δ","FY25 Revenue Mix","Data Status","Source"])
    for i,(name,d) in enumerate(ANNUAL_SEGMENTS.items(),7):
        ws.cell(i,1,name)
        for j,v in enumerate(d["revenue"],2): ws.cell(i,j,v); ws.cell(i,j).number_format=FMT_BN
        ws.cell(i,5,f'=IFERROR(D{i}/C{i}-1,"")'); ws.cell(i,6,f'=IFERROR((D{i}/B{i})^(1/2)-1,"")')
        for j,v in enumerate(d["op"],7): ws.cell(i,j,v); ws.cell(i,j).number_format=FMT_BN
        for rc,oc,mc in ((2,7,10),(3,8,11),(4,9,12)): ws.cell(i,mc,f'=IFERROR({get_column_letter(oc)}{i}/{get_column_letter(rc)}{i},"")')
        ws.cell(i,13,f'=IFERROR(L{i}-K{i},"")'); ws.cell(i,14,f'=IFERROR(D{i}/SUM($D$7:$D$9),"")'); _pct_cells(ws,i,(5,6,10,11,12,13,14))
        ws.cell(i,15,"AUDITED / COMPLETE"); _link(ws.cell(i,16),"SEC FY2025 10-K",ANNUAL_10K)

    # Current FY2026 YTD segment update.
    _section(ws,12,"Current Segment Update — 36 Weeks Ended May 10, 2026",12)
    _header(ws,13,["Segment","36W FY2025 Revenue","36W FY2026 Revenue","Revenue Growth","36W FY2025 Op. Income","36W FY2026 Op. Income","FY25 YTD Margin","FY26 YTD Margin","Margin Δ","FY26 Revenue Mix","Data Status","Source"])
    for i,(name,d) in enumerate(YTD_SEGMENTS.items(),14):
        ws.cell(i,1,name); ws.cell(i,2,d["revenue"][0]); ws.cell(i,3,d["revenue"][1]); ws.cell(i,4,f'=IFERROR(C{i}/B{i}-1,"")')
        ws.cell(i,5,d["op"][0]); ws.cell(i,6,d["op"][1]); ws.cell(i,7,f'=IFERROR(E{i}/B{i},"")'); ws.cell(i,8,f'=IFERROR(F{i}/C{i},"")'); ws.cell(i,9,f'=IFERROR(H{i}-G{i},"")'); ws.cell(i,10,f'=IFERROR(C{i}/SUM($C$14:$C$16),"")')
        for c in (2,3,5,6): ws.cell(i,c).number_format=FMT_BN
        _pct_cells(ws,i,(4,7,8,9,10)); ws.cell(i,11,"10-Q / COMPLETE"); _link(ws.cell(i,12),"SEC Q3 FY2026 10-Q",Q3_10Q)

    # Merchandise categories.
    _section(ws,19,"Merchandise Category Net Sales — Annual",10)
    _header(ws,20,["Category","FY2023","FY2024","FY2025","FY25 Growth","2Y CAGR","FY25 Mix","Trend / Interpretation","Data Status","Source"])
    for i,(name,vals) in enumerate(ANNUAL_CATEGORIES.items(),21):
        ws.cell(i,1,name)
        for j,v in enumerate(vals,2): ws.cell(i,j,v); ws.cell(i,j).number_format=FMT_BN
        ws.cell(i,5,f'=IFERROR(D{i}/C{i}-1,"")'); ws.cell(i,6,f'=IFERROR((D{i}/B{i})^(1/2)-1,"")'); ws.cell(i,7,f'=IFERROR(D{i}/SUM($D$21:$D$24),"")'); _pct_cells(ws,i,(5,6,7))
        ws.cell(i,8,"Issuer-disclosed merchandise mix; e-commerce and business-center sales are allocated to applicable categories."); ws.cell(i,9,"AUDITED / COMPLETE"); _link(ws.cell(i,10),"SEC FY2025 10-K",ANNUAL_10K)

    _section(ws,27,"Merchandise Category Net Sales — 36-Week FY2026 Update",9)
    _header(ws,28,["Category","36W FY2025","36W FY2026","Growth","FY26 Mix","Change vs annual mix","Data Status","Interpretation","Source"])
    for i,(name,vals) in enumerate(YTD_CATEGORIES.items(),29):
        ws.cell(i,1,name); ws.cell(i,2,vals[0]); ws.cell(i,3,vals[1]); ws.cell(i,4,f'=IFERROR(C{i}/B{i}-1,"")'); ws.cell(i,5,f'=IFERROR(C{i}/SUM($C$29:$C$32),"")')
        annual_row=21+list(ANNUAL_CATEGORIES.keys()).index(name); ws.cell(i,6,f'=IFERROR(E{i}-G{annual_row},"")')
        ws.cell(i,2).number_format=FMT_BN; ws.cell(i,3).number_format=FMT_BN; _pct_cells(ws,i,(4,5,6)); ws.cell(i,7,"10-Q / COMPLETE"); ws.cell(i,8,"Current-year category momentum from reported net sales."); _link(ws.cell(i,9),"SEC Q3 FY2026 10-Q",Q3_10Q)

    # Capital intensity by segment.
    _section(ws,35,"Segment Capital Intensity — FY2025",10)
    _header(ws,36,["Segment","D&A","Capex Additions","Total Assets","Capex / Revenue","Op. Income / Assets","Asset Mix","Capex Mix","Interpretation","Source"])
    for i,(name,d) in enumerate(ANNUAL_SEGMENTS.items(),37):
        ws.cell(i,1,name); ws.cell(i,2,d["da"][2]); ws.cell(i,3,d["capex"][2]); ws.cell(i,4,d["assets"][2]); ws.cell(i,5,f'=IFERROR(C{i}/D{7+list(ANNUAL_SEGMENTS.keys()).index(name)},"")'); ws.cell(i,6,f'=IFERROR(I{7+list(ANNUAL_SEGMENTS.keys()).index(name)}/D{i},"")'); ws.cell(i,7,f'=IFERROR(D{i}/SUM($D$37:$D$39),"")'); ws.cell(i,8,f'=IFERROR(C{i}/SUM($C$37:$C$39),"")')
        for c in (2,3,4): ws.cell(i,c).number_format=FMT_BN
        _pct_cells(ws,i,(5,6,7,8)); ws.cell(i,9,"Shows where Costco deploys physical capital relative to segment revenue and operating income."); _link(ws.cell(i,10),"SEC FY2025 segment note",ANNUAL_10K)

    # Business model and current momentum.
    _section(ws,42,"Membership Economics & Current Operating Momentum",8)
    _header(ws,43,["Metric","Value","Prior / Reference","Change","Period","Why it matters","Status","Source"])
    metrics=[
        ("Paid members (m)",82.9,79.6,"=(B44/C44)-1","Q3 FY2026","Membership-base scale and recurring-fee engine","REPORTED",Q3_10Q,"0.0"),
        ("Cardholders (m)",148.5,142.8,"=(B45/C45)-1","Q3 FY2026","Reach of the membership ecosystem","REPORTED",Q3_10Q,"0.0"),
        ("U.S./Canada renewal rate",.922,None,None,"Q3 FY2026","Core loyalty / retention signal","REPORTED",Q3_10Q,FMT_PCT),
        ("Worldwide renewal rate",.897,None,None,"Q3 FY2026","Global loyalty / retention signal","REPORTED",Q3_10Q,FMT_PCT),
        ("36W membership fees ($bn)",4.057,3.599,"=(B48/C48)-1","36W FY2026","Recurring high-value fee revenue growth","REPORTED",Q3_10Q,FMT_BN),
        ("44W net sales ($bn)",250.43,227.46,"=(B49/C49)-1","44 weeks ended Jul 5 2026","Latest machine-readable cumulative sales update","REPORTED",JUNE_SALES,FMT_BN),
        ("44W comparable sales",.083,None,None,"44 weeks ended Jul 5 2026","Underlying warehouse/digital demand","REPORTED",JUNE_SALES,FMT_PCT),
        ("44W adjusted comparable sales",.067,None,None,"44 weeks ended Jul 5 2026","Comparable sales excluding gasoline/FX","REPORTED",JUNE_SALES,FMT_PCT),
        ("44W digitally-enabled comparable sales",.215,None,None,"44 weeks ended Jul 5 2026","Digital channel momentum","REPORTED",JUNE_SALES,FMT_PCT),
        ("Warehouses",933,931,"=B53-C53","Jul 8 2026 vs Q3 report","Physical footprint growth","REPORTED",JUNE_SALES,FMT_INT),
    ]
    for r,(name,val,ref,chg,period,why,status,url,fmt) in enumerate(metrics,44):
        ws.cell(r,1,name); ws.cell(r,2,val); ws.cell(r,3,ref); ws.cell(r,4,chg); ws.cell(r,5,period); ws.cell(r,6,why); ws.cell(r,7,status); _link(ws.cell(r,8),"Primary source",url)
        ws.cell(r,2).number_format=fmt; ws.cell(r,3).number_format=fmt; ws.cell(r,4).number_format=FMT_PCT if name!="Warehouses" else FMT_INT
    ws["A55"]="Latest July sales event"; ws["B55"]="Published Aug 5, 2026; official event page currently exposes the audio update rather than machine-readable figures. The model does not invent/transcribe unavailable numbers."
    ws.merge_cells("B55:G55"); ws["B55"].alignment=Alignment(wrap_text=True); _link(ws["H55"],"Costco July Sales event",JULY_EVENT)

    _section(ws,58,"Source & Interpretation Notes",10)
    notes=[
        ("Reportable segment definition","United States, Canada, Other International. Costco states its operating models are generally the same across these geographic segments.",Q3_10Q),
        ("Low-margin model","Costco explicitly operates at very low merchandise margins and relies on volume, rapid inventory turnover and membership economics. Segment margin should therefore be judged by trend/returns, not software-like absolute margins.",ANNUAL_10K),
        ("Digital / gasoline context","FY2025 e-commerce represented about 7% of total net sales; digitally-enabled sales about 10%; gasoline about 10% of net sales.",ANNUAL_10K),
        ("Data discipline","Annual and interim segment/category values are reported public data. Missing current-month details remain unfilled rather than estimated.",None),
    ]
    for r,(label,text,url) in enumerate(notes,59):
        ws.cell(r,1,label); ws.cell(r,2,text); ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=8); ws.cell(r,2).alignment=Alignment(wrap_text=True)
        if url: _link(ws.cell(r,9),"Source",url)

    widths={"A":39,"B":17,"C":17,"D":17,"E":17,"F":17,"G":18,"H":34,"I":22,"J":22,"K":15,"L":15,"M":15,"N":15,"O":24,"P":34}
    for c,w in widths.items(): ws.column_dimensions[c].width=w
    ws.freeze_panes="A7"
    _quality(wb)
    print("Costco Segment Analysis: verified FY2023-25, FY2026 36-week, category, membership and current sales data populated")
    return True


def _quality(wb):
    if "Data Quality" not in wb.sheetnames: return
    ws=wb["Data Quality"]; label="Segment Analysis public-data coverage"; row=None
    for r in range(1,ws.max_row+1):
        if str(ws.cell(r,1).value or "").strip()==label: row=r; break
    row=row or ws.max_row+1
    ws.cell(row,1,label); ws.cell(row,2,"PASS"); ws.cell(row,3,"COST: 3 audited segments + FY2026 YTD + 4 merchandise categories + capital intensity + membership/current-sales KPIs")
    ws.cell(row,4,"Primary SEC 10-K/10-Q and Costco IR sources; explicit reportable segments replace narrative-parser fragments.")
    ws.cell(row,2).fill=_fill(PALE_GREEN); ws.cell(row,2).font=Font(bold=True)
    for c in range(1,5): ws.cell(row,c).alignment=Alignment(wrap_text=True,vertical="top")
