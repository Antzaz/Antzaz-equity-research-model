from __future__ import annotations

"""Second-pass full-statement renderer.

This module reuses the standardized mappings/helpers from full_financial_statements.py but
removes ambiguous mappings that can double count non-operating items, labels share counts in
billions, and derives accounting identities only when source statements do not provide them.
"""

from openpyxl.styles import Alignment, Border, Font
from openpyxl.utils import get_column_letter

import full_financial_statements as base


def _replace(rows,label,new_row):
    out=[]
    for row in rows:
        out.append(new_row if row[0]==label else row)
    return out


INCOME=_replace(base.INCOME,"Other Non-Operating Income / (Expense)",(
    "Other Non-Operating Income / (Expense)",
    ["Other Non Operating Income Expenses"],
    ["OtherNonoperatingIncomeExpense"],"money",(),
))
# The condensed legacy row represents total non-operating income, so keep it separate rather
# than aliasing it into "Other" while interest income/expense is also shown.
_idx=next(i for i,r in enumerate(INCOME) if r[0]=="Pre-Tax Income")
INCOME.insert(_idx,(
    "Total Non-Operating Income / (Expense)",
    ["Total Other Finance Cost","Net Non Operating Interest Income Expense"],
    ["NonoperatingIncomeExpense"],"money",("Other Income / (Expense), Net",),
))
INCOME=_replace(INCOME,"Basic Weighted Average Shares",(
    "Basic Weighted Average Shares (bn)",["Basic Average Shares"],["WeightedAverageNumberOfSharesOutstandingBasic"],"shares",("Basic Weighted Average Shares",),
))
INCOME=_replace(INCOME,"Diluted Weighted Average Shares",(
    "Diluted Weighted Average Shares (bn)",["Diluted Average Shares"],["WeightedAverageNumberOfDilutedSharesOutstanding"],"shares",("Diluted Weighted Average Shares",),
))

BALANCE=_replace(base.BALANCE,"Total Liabilities & Equity",(
    "Total Liabilities & Equity",[],["LiabilitiesAndStockholdersEquity"],"money",(),
))

# Beginning cash is a period-start balance. An instant-period SEC cash fact without context is not
# a safe mapping for this row; use a statement-specific provider if available or derive from the
# prior year's ending balance below.
CASH=_replace(base.CASH,"Beginning Cash",(
    "Beginning Cash",["Beginning Cash Position"],[],"money",(),
))


def _write_formula_if_blank(ws,row,col,formula,number_format=base.FMT_BN):
    if ws.cell(row,col).value in (None,""):
        ws.cell(row,col,formula); ws.cell(row,col).number_format=number_format


def expand_financial_statements(wb,ticker,facts=None):
    existing=base._capture_existing(wb)
    frames=base._yf_frames(ticker)
    years=base._available_years(existing,frames,facts)
    if not years:
        return {"expanded":False,"income_rows":0,"balance_rows":0,"cash_rows":0,"years":[]}

    idx=wb.sheetnames.index("Financial Statements") if "Financial Statements" in wb.sheetnames else len(wb.sheetnames)
    if "Financial Statements" in wb.sheetnames:
        wb.remove(wb["Financial Statements"])
    ws=wb.create_sheet("Financial Statements",idx); ws.sheet_view.showGridLines=False
    last_col=2+len(years); base._title(ws,ticker,last_col)

    row,imap,source_col=base._write_section(ws,5,"Income Statement",INCOME,years,existing,frames,facts,"income")
    for label,numer in (("Operating Margin","Operating Income"),("Net Margin","Net Income")):
        ws.cell(row,1,label); ws.cell(row,source_col,"Derived: selected income ÷ revenue")
        for j,_ in enumerate(years,2):
            c=get_column_letter(j)
            ws.cell(row,j,f'=IFERROR({c}{imap[numer]}/{c}{imap["Revenue"]},"")')
            ws.cell(row,j).number_format=base.FMT_PCT
        row+=1

    row+=2
    row,bmap,source_col=base._write_section(ws,row,"Balance Sheet",BALANCE,years,existing,frames,facts,"balance")
    # Accounting identity only fills gaps; reported SEC/provider values remain untouched.
    tle=bmap["Total Liabilities & Equity"]
    for j,_ in enumerate(years,2):
        c=get_column_letter(j)
        _write_formula_if_blank(ws,tle,j,f'=IFERROR({c}{bmap["Total Liabilities"]}+{c}{bmap["Total Equity"]},"")')
    if all(ws.cell(tle,j).data_type=="f" for j in range(2,2+len(years))):
        ws.cell(tle,source_col,"Derived only when reported Liabilities & Equity is unavailable: Total Liabilities + Total Equity")

    for label,(a,b) in (("Net Debt",("Long-Term Debt","Cash & Cash Equivalents")),("Working Capital",("Total Current Assets","Total Current Liabilities"))):
        ws.cell(row,1,label); ws.cell(row,source_col,"Derived balance-sheet metric")
        for j,_ in enumerate(years,2):
            c=get_column_letter(j)
            ws.cell(row,j,f'=IFERROR({c}{bmap[a]}-{c}{bmap[b]},"")'); ws.cell(row,j).number_format=base.FMT_BN
        row+=1

    row+=2
    row,cmap,source_col=base._write_section(ws,row,"Cash Flow Statement",CASH,years,existing,frames,facts,"cash")
    # If beginning cash is unavailable, prior-year ending cash is the cleanest accounting bridge.
    begin=cmap["Beginning Cash"]; end=cmap["Ending Cash"]
    for i,j in enumerate(range(2,2+len(years))):
        if i==0: continue
        prev=get_column_letter(j-1)
        _write_formula_if_blank(ws,begin,j,f'={prev}{end}')
    ws.cell(begin,source_col,"Reported provider value where available; otherwise prior-year Ending Cash")

    ws.cell(row,1,"Free Cash Flow"); ws.cell(row,1).font=Font(bold=True)
    ws.cell(row,source_col,"Derived: Operating Cash Flow + Capital Expenditures")
    for j,_ in enumerate(years,2):
        c=get_column_letter(j)
        ws.cell(row,j,f'=IFERROR({c}{cmap["Operating Cash Flow"]}+{c}{cmap["Capital Expenditures"]},"")')
        ws.cell(row,j).number_format=base.FMT_BN
    for c in range(1,source_col+1): ws.cell(row,c).border=Border(top=base.THIN)
    row+=3

    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=source_col)
    ws.cell(row,1,"Full Statement Coverage"); ws.cell(row,1).fill=base._fill(base.NAVY); ws.cell(row,1).font=Font(bold=True,color=base.WHITE)
    row+=1
    for c,v in enumerate(["Section","Mapped rows","Total standardized rows","Coverage"],1):
        ws.cell(row,c,v); ws.cell(row,c).fill=base._fill(base.BLUE); ws.cell(row,c).font=Font(bold=True,color=base.WHITE)
    stats={}
    for key,name,row_map in (("income","Income Statement",imap),("balance","Balance Sheet",bmap),("cash","Cash Flow Statement",cmap)):
        mapped=sum(any(base._num(ws.cell(r,j).value) is not None for j in range(2,2+len(years))) for r in row_map.values())
        stats[key]=(mapped,len(row_map)); row+=1
        ws.cell(row,1,name); ws.cell(row,2,mapped); ws.cell(row,3,len(row_map)); ws.cell(row,4,mapped/len(row_map) if row_map else 0); ws.cell(row,4).number_format=base.FMT_PCT

    ws.column_dimensions["A"].width=46
    for c in range(2,2+len(years)): ws.column_dimensions[get_column_letter(c)].width=14
    ws.column_dimensions[get_column_letter(source_col)].width=48
    for r in range(1,ws.max_row+1): ws.cell(r,source_col).alignment=Alignment(wrap_text=True,vertical="top")
    ws.freeze_panes="B7"
    try:
        wb.calculation.fullCalcOnLoad=True
        wb.calculation.forceFullCalc=True
    except Exception:
        pass
    return {
        "expanded":True,
        "income_rows":stats["income"][0],"income_total":stats["income"][1],
        "balance_rows":stats["balance"][0],"balance_total":stats["balance"][1],
        "cash_rows":stats["cash"][0],"cash_total":stats["cash"][1],
        "years":years,
    }
