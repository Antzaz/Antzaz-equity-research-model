from __future__ import annotations

"""Enrich Company Data with an investor-useful business description and main products/services.

The product table is intentionally descriptive rather than financial: revenue is shown only in
Segment Analysis when the issuer discloses it. Curated issuer-backed rows are used for companies
where the project has verified current product taxonomies; all other tickers fall back to reported
segment/business-line names plus the company profile summary.
"""

from openpyxl.styles import Alignment, Font, PatternFill

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; LIGHT="F5F9FC"; GREY="666666"

CURATED_PRODUCTS={
    "GEV":{
        "description":"GE Vernova supplies equipment, services and software used to generate, transfer, orchestrate, convert and store electricity across Power, Wind and Electrification.",
        "rows":[
            ("Power","Heavy-duty and aeroderivative gas turbines","Gas-fired generation equipment plus installed-base services and parts.","https://www.gevernova.com/investors"),
            ("Power","Steam, nuclear and hydro power equipment & services","Generation technologies and lifecycle services, including nuclear and hydro platforms.","https://www.gevernova.com/investors/annual-report"),
            ("Wind","Onshore wind turbines & services","Utility-scale onshore turbines, installed-base service and lifecycle support.","https://www.gevernova.com/investors"),
            ("Wind","Offshore wind turbines","Large-scale offshore wind generation equipment and related services.","https://www.gevernova.com/investors"),
            ("Electrification","Transformers, switchgear, HVDC/substations & synchronous condensers","Grid equipment used to transmit, stabilize and modernize electricity networks.","https://www.gevernova.com/investors/annual-report/ceo-letter"),
            ("Electrification","Power conversion, storage and grid software","Power electronics, storage and software/digital orchestration used across modern grids.","https://www.gevernova.com/investors"),
        ],
    },
    "CEG":{
        "description":"Constellation combines large-scale electricity generation with wholesale/retail energy supply and customer energy solutions. Its fleet includes nuclear and, following Calpine, substantial natural-gas and geothermal generation.",
        "rows":[
            ("Generation","Nuclear electricity generation","Large-scale 24/7 emissions-free generation and plant operations.","https://www.constellationenergy.com/work/generation/nuclear.html"),
            ("Generation / Calpine","Natural-gas generation","Dispatchable gas-fired generation, including the Calpine fleet acquired in 2026.","https://www.constellationenergy.com/our-work/what-we-do/generation.html"),
            ("Generation","Hydro, wind, solar and geothermal generation","Diversified zero-/low-carbon generation portfolio alongside nuclear and gas.","https://www.constellationenergy.com/our-work/what-we-do/generation.html"),
            ("Retail / Commercial","Electricity supply","Fixed, index and managed electricity procurement solutions for residential, commercial and public-sector customers.","https://www.constellationenergy.com/work/commercial.html"),
            ("Retail / Commercial","Natural-gas supply","Customized natural-gas procurement and risk-management strategies.","https://www.constellationenergy.com/work/commercial.html"),
            ("Customer Solutions","Energy efficiency, offsite renewables & hourly carbon-free matching","Energy-management and clean-energy products designed to lower cost and support sustainability goals.","https://www.constellationenergy.com/work/commercial.html"),
            ("Generation Services","Constellation Generation Solutions & PowerLabs","Nuclear outage/maintenance/technical services plus calibration, testing and quality services.","https://www.constellationenergy.com/our-work/what-we-do/generation.html"),
        ],
    },
    "GOOGL":{
        "description":"Alphabet's economics are led by Google Services and Google Cloud, with Other Bets housing earlier-stage businesses. Company Data separates the major products/services without assigning revenue that Alphabet does not disclose at product level.",
        "rows":[
            ("Google Services","Search & advertising","Search and other Google properties monetize user intent primarily through advertising.","https://abc.xyz/investor/faqs-and-general-information/default.aspx"),
            ("Google Services","YouTube ads & subscriptions","Video advertising plus paid offerings such as YouTube Premium, Music, TV and related subscriptions.","https://abc.xyz/investor/faqs-and-general-information/default.aspx"),
            ("Google Services","Android, Chrome, Maps, Play & devices","Consumer platforms and devices that expand distribution, engagement and ecosystem monetization.","https://abc.xyz/investor/faqs-and-general-information/default.aspx"),
            ("Google Services","Google One & AI subscriptions","Consumer subscriptions including Google One and paid AI plans.","https://abc.xyz/investor/faqs-and-general-information/default.aspx"),
            ("Google Cloud","GCP infrastructure, AI, data & cybersecurity","Enterprise cloud infrastructure/platform services including TPUs/GPUs, Vertex AI, data analytics and cybersecurity.","https://abc.xyz/investor/faqs-and-general-information/default.aspx"),
            ("Google Cloud","Workspace & Gemini enterprise applications","Subscription productivity/collaboration and enterprise AI applications.","https://abc.xyz/investor/faqs-and-general-information/default.aspx"),
            ("Other Bets","Waymo, Verily, GFiber and other ventures","Earlier-stage businesses including autonomous mobility, health-related and connectivity services.","https://abc.xyz/investor/faqs-and-general-information/default.aspx"),
        ],
    },
    "AMZN":{
        "description":"Amazon combines global online/physical stores and third-party seller services with subscriptions, advertising and AWS cloud infrastructure. The product table is descriptive; reported segment economics remain in Segment Analysis.",
        "rows":[
            ("Stores","Online & physical retail","First-party merchandise sold through Amazon's online and physical-store channels.","https://ir.aboutamazon.com/annual-reports-proxies-and-shareholder-letters/default.aspx"),
            ("Stores","Third-party seller services","Marketplace services including commissions, fulfillment and related seller offerings.","https://ir.aboutamazon.com/annual-reports-proxies-and-shareholder-letters/default.aspx"),
            ("Stores","Prime & other subscriptions","Membership and digital subscription services that deepen customer engagement and recurring revenue.","https://ir.aboutamazon.com/annual-reports-proxies-and-shareholder-letters/default.aspx"),
            ("Advertising","Sponsored ads & advertising services","Performance and brand advertising sold across Amazon properties and services.","https://ir.aboutamazon.com/annual-reports-proxies-and-shareholder-letters/default.aspx"),
            ("AWS","Cloud compute, storage, databases & networking","Consumption-based enterprise infrastructure services.","https://aws.amazon.com/what-is-aws/"),
            ("AWS","AI/ML, analytics, security & application services","Higher-layer cloud services including AI/ML, data, security and developer/application tools.","https://aws.amazon.com/what-is-aws/"),
        ],
    },
    "NVDA":{
        "description":"NVIDIA is an accelerated-computing platform company spanning data-center AI, networking, gaming/graphics, professional visualization and automotive computing.",
        "rows":[
            ("Data Center","AI accelerators & accelerated-computing platforms","GPUs and systems used to train and infer AI models and accelerate scientific/enterprise workloads.","https://investor.nvidia.com/financial-info/annual-reports-and-proxies/default.aspx"),
            ("Data Center","Networking","High-performance interconnect, switching and networking products that connect accelerated-computing clusters.","https://investor.nvidia.com/financial-info/annual-reports-and-proxies/default.aspx"),
            ("Gaming","GeForce GPUs & gaming platforms","Consumer graphics processors and software/platform services for PC gaming.","https://investor.nvidia.com/financial-info/annual-reports-and-proxies/default.aspx"),
            ("Professional Visualization","RTX professional graphics","Workstation and visualization platforms for creators, designers and technical users.","https://investor.nvidia.com/financial-info/annual-reports-and-proxies/default.aspx"),
            ("Automotive","NVIDIA DRIVE & autonomous-driving compute","Compute, software and platform technology for assisted and autonomous vehicles.","https://investor.nvidia.com/financial-info/annual-reports-and-proxies/default.aspx"),
            ("Software / Platforms","CUDA and enterprise AI software","Programming platforms, libraries and enterprise software that reinforce NVIDIA's accelerated-computing ecosystem.","https://investor.nvidia.com/financial-info/annual-reports-and-proxies/default.aspx"),
        ],
    },
    "JPM":{
        "description":"JPMorgan Chase is a diversified global financial institution. Its product economics are better understood through banking/markets/asset-management business lines than through an industrial product taxonomy.",
        "rows":[
            ("Consumer & Community Banking","Deposits, payments & consumer banking","Checking, savings, payments and branch/digital banking services for consumers and small businesses.","https://www.jpmorganchase.com/ir/annual-report"),
            ("Consumer & Community Banking","Credit cards, auto & home lending","Consumer credit products including cards and major secured-lending categories.","https://www.jpmorganchase.com/ir/annual-report"),
            ("Commercial & Investment Bank","Investment banking & capital markets","Advisory, underwriting, financing and markets services for corporate and institutional clients.","https://www.jpmorganchase.com/ir/annual-report"),
            ("Commercial & Investment Bank","Payments & securities services","Treasury/payments, custody, clearing and related institutional transaction services.","https://www.jpmorganchase.com/ir/annual-report"),
            ("Asset & Wealth Management","Asset management & private banking","Investment management, wealth advice, lending and related services for institutions and private clients.","https://www.jpmorganchase.com/ir/annual-report"),
        ],
    },
    "BRK.B":{
        "description":"Berkshire Hathaway is an insurance-led operating conglomerate with major railroad, utility/energy, manufacturing, service and retail businesses plus a large investment portfolio.",
        "rows":[
            ("Insurance","GEICO and primary insurance","Property/casualty and other primary insurance operations that also generate investable insurance float.","https://www.berkshirehathaway.com/reports.html"),
            ("Insurance","Reinsurance","Property/casualty and life/health reinsurance written through Berkshire insurance subsidiaries.","https://www.berkshirehathaway.com/reports.html"),
            ("BNSF","Freight rail transportation","North American railroad transportation and logistics infrastructure.","https://www.berkshirehathaway.com/reports.html"),
            ("Berkshire Hathaway Energy","Electric utilities, renewables, pipelines & energy infrastructure","Regulated and non-regulated energy assets including utility, transmission and pipeline operations.","https://www.berkshirehathaway.com/reports.html"),
            ("Manufacturing","Industrial, building & consumer products","A broad portfolio of manufacturing businesses across industrial and consumer end markets.","https://www.berkshirehathaway.com/reports.html"),
            ("Service & Retailing","Distribution, services and retail businesses","Operating subsidiaries spanning distribution, aviation/services, automotive retail and other consumer/business services.","https://www.berkshirehathaway.com/reports.html"),
        ],
    },
    "TSM":{
        "description":"TSMC is a pure-play semiconductor foundry manufacturing customer-designed integrated circuits across advanced and specialty process technologies, complemented by advanced packaging and design enablement.",
        "rows":[
            ("Foundry","Leading-edge logic process technologies","Advanced semiconductor manufacturing nodes used in high-performance computing, smartphones and other demanding applications.","https://investor.tsmc.com/english/annual-reports"),
            ("Foundry","Mainstream & specialty technologies","Mature and specialty processes serving automotive, IoT, RF, power-management and other differentiated applications.","https://investor.tsmc.com/english/annual-reports"),
            ("Advanced Packaging","3DFabric / advanced packaging & testing","Advanced integration and packaging technologies that combine dies and memory for high-performance systems.","https://www.tsmc.com/english/dedicatedFoundry/technology/3DFabric"),
            ("Design Enablement","IP, design platforms & ecosystem services","Design-technology co-optimization, IP and ecosystem support that help customers move designs into production.","https://www.tsmc.com/english/dedicatedFoundry/services"),
        ],
    },
    "SIE.DE":{
        "description":"Siemens is an industrial technology group spanning automation/software, electrification and smart infrastructure, rail mobility, healthcare technology and financial services.",
        "rows":[
            ("Digital Industries","Factory automation, drives, industrial software & digital twins","Automation hardware/software and industrial digitalization tools used to design and operate factories and production systems.","https://www.siemens.com/global/en/company/about/businesses/digital-industries.html"),
            ("Smart Infrastructure","Electrification, grid & building technologies","Electrical products, grid/building systems and software supporting electrification, energy efficiency and resilient infrastructure.","https://www.siemens.com/global/en/company/about/businesses/smart-infrastructure.html"),
            ("Mobility","Rail vehicles, signaling, electrification & digital rail services","Passenger/freight rail equipment plus automation, signaling, electrification and lifecycle services.","https://www.siemens.com/global/en/company/about/businesses/mobility.html"),
            ("Siemens Healthineers","Imaging, diagnostics & therapeutic technology","Medical diagnostic and therapeutic equipment, software and related healthcare services.","https://www.siemens-healthineers.com/"),
            ("Siemens Financial Services","Equipment/project financing & investment solutions","Financing capabilities that support Siemens customers, infrastructure and industrial projects.","https://www.siemens.com/global/en/company/about/businesses/financial-services.html"),
        ],
    },
}

# Analyst-facing aliases share the same verified product taxonomy.
CURATED_PRODUCTS["GOOG"]=CURATED_PRODUCTS["GOOGL"]
CURATED_PRODUCTS["BRK-B"]=CURATED_PRODUCTS["BRK.B"]
CURATED_PRODUCTS["BRK.A"]=CURATED_PRODUCTS["BRK.B"]
CURATED_PRODUCTS["BRK-A"]=CURATED_PRODUCTS["BRK.B"]

IGNORE_SEGMENT_LABELS={
    "segment","business line","metric","total","company total","income statement","balance sheet",
    "cash flow statement","reported segments","segment analysis","manual segment input",
}


def _fill(color): return PatternFill("solid",fgColor=color)


def _segment_fallback(wb,limit=6):
    if "Segment Analysis" not in wb.sheetnames: return []
    ws=wb["Segment Analysis"]; out=[]; seen=set()
    for r in range(1,ws.max_row+1):
        name=str(ws.cell(r,1).value or "").strip()
        if not name or name.lower() in IGNORE_SEGMENT_LABELS or name.lower().startswith(("source","note","method")):
            continue
        numeric=sum(isinstance(ws.cell(r,c).value,(int,float)) for c in range(2,min(ws.max_column,9)+1))
        if numeric==0: continue
        key=name.lower()
        if key in seen: continue
        seen.add(key)
        source=""
        for c in range(ws.max_column,1,-1):
            v=ws.cell(r,c).value
            if isinstance(v,str) and v.startswith("http"):
                source=v; break
        out.append(("Reported business / segment",name,"Issuer-disclosed business line; see Segment Analysis for available economics.",source))
        if len(out)>=limit: break
    return out


def _summary(info):
    text=str((info or {}).get("longBusinessSummary") or "").strip()
    if not text: return ""
    # Two sentences is enough context for Company Data; detailed narrative belongs elsewhere.
    parts=[x.strip() for x in text.replace("\n"," ").split(". ") if x.strip()]
    return ". ".join(parts[:2])[:900].rstrip(".")+"."


def _clear_area(ws):
    for r in range(17,46):
        for c in range(1,5):
            ws.cell(r,c).value=None
            ws.cell(r,c).fill=PatternFill(fill_type=None)
            ws.cell(r,c).font=Font()
            ws.cell(r,c).alignment=Alignment()


def enrich_company_data(wb,ticker,info=None):
    if "Company Data" not in wb.sheetnames: return {"products":0,"source":"missing"}
    ws=wb["Company Data"]; t=str(ticker).upper().strip(); _clear_area(ws)
    curated=CURATED_PRODUCTS.get(t)
    description=(curated or {}).get("description") or _summary(info)
    rows=list((curated or {}).get("rows") or [])
    if not rows: rows=_segment_fallback(wb)
    website=str((info or {}).get("website") or "")

    for c in range(1,5):
        ws.cell(17,c).fill=_fill(NAVY); ws.cell(17,c).font=Font(bold=True,color=WHITE)
    ws.cell(17,1,"Business Overview & Main Products / Services")
    ws.merge_cells(start_row=17,start_column=1,end_row=17,end_column=4)

    ws.cell(18,1,"Business Description"); ws.cell(18,1).font=Font(bold=True)
    ws.cell(18,2,description or "No concise business description was verified automatically.")
    ws.merge_cells(start_row=18,start_column=2,end_row=18,end_column=3)
    ws.cell(18,2).alignment=Alignment(wrap_text=True,vertical="top")
    ws.cell(18,4,(curated or {}).get("rows",[[None,None,None,website]])[0][3] if curated else website)
    ws.cell(18,4).font=Font(color="008000",underline="single")

    headers=["Business / Segment","Main Product or Service","What It Does / Why It Matters","Primary Source"]
    for c,v in enumerate(headers,1):
        ws.cell(20,c,v); ws.cell(20,c).fill=_fill(BLUE); ws.cell(20,c).font=Font(bold=True,color=WHITE); ws.cell(20,c).alignment=Alignment(horizontal="center",wrap_text=True)
    if not rows:
        rows=[("Business profile","See business description above","Product-level taxonomy was not reliably available; Segment Analysis remains the financial source of truth.",website)]
    for r_idx,item in enumerate(rows[:10],21):
        for c,v in enumerate(item,1):
            ws.cell(r_idx,c,v); ws.cell(r_idx,c).alignment=Alignment(wrap_text=True,vertical="top")
        if item[3]: ws.cell(r_idx,4).font=Font(color="008000",underline="single")

    ws.cell(32,1,"Research rule"); ws.cell(32,1).font=Font(bold=True,color=GREY)
    ws.cell(32,2,"Products/services are descriptive. Do not infer standalone revenue unless the issuer discloses it; use Segment Analysis for reported economics.")
    ws.merge_cells(start_row=32,start_column=2,end_row=32,end_column=4); ws.cell(32,2).font=Font(italic=True,color=GREY); ws.cell(32,2).alignment=Alignment(wrap_text=True)
    ws.column_dimensions["A"].width=max(ws.column_dimensions["A"].width or 0,28)
    ws.column_dimensions["B"].width=max(ws.column_dimensions["B"].width or 0,42)
    ws.column_dimensions["C"].width=max(ws.column_dimensions["C"].width or 0,62)
    ws.column_dimensions["D"].width=max(ws.column_dimensions["D"].width or 0,55)
    ws.row_dimensions[18].height=48
    return {"products":len(rows),"source":"curated" if curated else "segment/profile fallback"}
