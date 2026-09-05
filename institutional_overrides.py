"""Production overrides for consensus and market-expectations research.

This module fixes three structural research problems:
1. Street consensus must remain independent from the model forecast.
2. Reverse-DCF sheets must use the same enterprise-value and starting-FCF definitions.
3. One-variable market-implied operating hurdles must stay inside economically feasible bounds.

The functions are installed at runtime by commodity_safe_runner so the existing workbook pipeline
can keep its public API while the saved research output uses the stronger methodology.
"""

from __future__ import annotations

import math
from typing import Any

from openpyxl.chart import BarChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import institutional_layers as il
from advanced_analytics import _solve_reverse_dcf
from consensus_sources import collect_consensus

try:
    import yfinance as yf
except Exception:
    yf = None

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; LIGHT="F5F9FC"; GOLD="FFF2CC"
PALE_GREEN="E2F0D9"; PALE_RED="FCE4D6"; GREY="666666"; LINK_GREEN="008000"
FMT_PCT='0.0%;[Red](0.0%);-'; FMT_PRICE='$#,##0.00;[Red]($#,##0.00);-'; FMT_BN='#,##0.0;[Red](#,##0.0);-'


def _num(value: Any, default=None):
    try:
        if isinstance(value, bool) or value in (None, ""):
            return default
        x=float(value)
        return x if math.isfinite(x) else default
    except Exception:
        return default


def _fill(color):
    return PatternFill("solid", fgColor=color)


def _create(wb,name):
    if name in wb.sheetnames:
        wb.remove(wb[name])
    ws=wb.create_sheet(name)
    ws.sheet_view.showGridLines=False
    return ws


def _title(ws,text,end=12):
    for c in range(1,end+1):
        ws.cell(1,c).fill=_fill(NAVY); ws.cell(2,c).fill=_fill(NAVY)
    ws["A1"]=text; ws["A1"].font=Font(bold=True,color=WHITE,size=18)


def _section(ws,row,title,end=12):
    for c in range(1,end+1):
        ws.cell(row,c).fill=_fill(NAVY); ws.cell(row,c).font=Font(bold=True,color=WHITE,size=11)
    ws.cell(row,1,title)


def _header(ws,row,labels):
    for c,label in enumerate(labels,1):
        cell=ws.cell(row,c,label); cell.fill=_fill(BLUE); cell.font=Font(bold=True,color=WHITE)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)


def _source_cell(cell,url=None):
    cell.font=Font(color=LINK_GREEN)
    if url:
        cell.hyperlink=url; cell.font=Font(color=LINK_GREEN,underline="single")


def _set_text_categories(chart,ws,col,first,last,titles):
    sheet_name=ws.title.replace("'","''")
    letter=get_column_letter(col)
    formula=f"'{sheet_name}'!${letter}${first}:${letter}${last}"
    for i,series in enumerate(chart.series):
        series.cat=AxDataSource(strRef=StrRef(f=formula))
        if i<len(titles): series.tx=SeriesLabel(v=str(titles[i]))


def _latest_year(wb):
    h=wb["Historical Financials"]
    years=[]
    for c in range(2,8):
        v=h.cell(3,c).value
        if isinstance(v,(int,float)) and 1900<=int(v)<=2200:
            years.append(int(v))
    return max(years) if years else None


def _company_inputs(wb):
    price,shares,net_debt=il._company(wb)
    model_ev=price*shares+net_debt if price and shares else None
    return price,shares,net_debt,model_ev


def _fcf_inputs(wb):
    h=wb["Historical Financials"]
    ocf=_num(h["G14"].value); capex=_num(h["G15"].value); sbc=_num(h["G21"].value,0.0)
    reported=(ocf-capex) if ocf is not None and capex is not None else None
    sbc_adjusted=(reported-sbc) if reported is not None else None
    normalized=sbc_adjusted if sbc_adjusted is not None and sbc_adjusted>0 else reported
    return {"ocf":ocf,"capex":capex,"sbc":sbc,"reported":reported,"normalized":normalized}


def _implied_growth(start_fcf,target_ev,wacc,tgr):
    if start_fcf is None or start_fcf<=0 or target_ev is None or target_ev<=0 or wacc<=tgr:
        return None
    return _solve_reverse_dcf(start_fcf,target_ev,wacc,tgr)


def _solve_shift(fn,target,lo,hi):
    """Bisection with explicit feasibility bounds; no extrapolation outside the range."""
    if lo is None or hi is None or lo>=hi:
        return None
    return il._bisect(fn,target,lo,hi)


def _operating_hurdles(wb):
    price,_,_,_= _company_inputs(wb)
    proj=il._base_projection(wb)
    if not proj or not price:
        return {}
    latest_rev=_num(wb["Historical Financials"]["G4"].value)
    base_margin=_num(proj[-1].get("margin")); base_cap=_num(proj[-1].get("capex_pct"))

    g_shift=_solve_shift(lambda x:il._model_value(wb,g_shift=x),price,-0.10,0.15)
    # Mature margin and capex are constrained to plausible economic ranges. If current price
    # cannot be matched inside the range, report N/A rather than an impossible hurdle.
    m_lo=(0.05-base_margin) if base_margin is not None else None
    m_hi=(0.55-base_margin) if base_margin is not None else None
    c_lo=(0.005-base_cap) if base_cap is not None else None
    c_hi=(0.30-base_cap) if base_cap is not None else None
    m_shift=_solve_shift(lambda x:il._model_value(wb,m_shift=x),price,m_lo,m_hi)
    cap_shift=_solve_shift(lambda x:il._model_value(wb,cap_shift=x),price,c_lo,c_hi)

    base_rev_cagr=None; implied_rev_cagr=None
    if latest_rev and latest_rev>0:
        base_rev_cagr=(proj[-1]["revenue"]/latest_rev)**(1/len(proj))-1 if proj[-1]["revenue"]>0 else None
        if g_shift is not None:
            rev=latest_rev
            for row in proj:
                rev*=1+max(-.5,min(.6,_num(row.get("growth"),0)+g_shift))
            implied_rev_cagr=(rev/latest_rev)**(1/len(proj))-1 if rev>0 else None

    return {
        "base_rev_cagr":base_rev_cagr,
        "implied_rev_cagr":implied_rev_cagr,
        "base_margin":base_margin,
        "implied_margin":base_margin+m_shift if base_margin is not None and m_shift is not None else None,
        "base_capex":base_cap,
        "implied_capex":base_cap+cap_shift if base_cap is not None and cap_shift is not None else None,
        "growth_shift":g_shift,"margin_shift":m_shift,"capex_shift":cap_shift,
        "duration":il._implied_duration(wb),
    }


def _yahoo_price_target(ticker):
    if yf is None:
        return None
    try:
        t=yf.Ticker(ticker)
        data=getattr(t,"analyst_price_targets",None)
        if not isinstance(data,dict):
            return None
        out={k:_num(data.get(k)) for k in ("mean","median","low","high","current")}
        if any(out.get(k) is not None for k in ("mean","median","low","high")):
            out.update({"provider":"Yahoo Finance","source_url":"https://finance.yahoo.com/"})
            return out
    except Exception:
        return None
    return None


def _consensus_bundle(ticker,latest_year,info=None):
    try:
        bundle=collect_consensus(ticker,latest_year,info or {}) or {}
    except Exception:
        bundle={"observations":[],"statuses":[],"blended":{},"price_targets":[]}
    if not bundle.get("price_targets"):
        pt=_yahoo_price_target(ticker)
        if pt: bundle["price_targets"]=[pt]
    return bundle


def _revision_map(bundle):
    out={}
    for rec in bundle.get("observations") or []:
        if rec.get("metric")!="EPS Revision": continue
        cur=_num(rec.get("mean")); d30=_num(rec.get("low")); d90=_num(rec.get("high"))
        fy=rec.get("fiscal_year")
        out[fy]={
            "30d":(cur/d30-1) if cur not in (None,0) and d30 not in (None,0) else None,
            "90d":(cur/d90-1) if cur not in (None,0) and d90 not in (None,0) else None,
        }
    return out


def ensure_expectations_consensus_v2(wb,ticker):
    """Street-versus-model sheet where the two forecasts are genuinely independent."""
    ws=_create(wb,"Expectations & Consensus"); _title(ws,f"{ticker} — Street Consensus vs Independent Model",13)
    ws["A3"]=(
        "Consensus is external evidence only; it does not overwrite Bear/Base/Bull assumptions. "
        "Provider counts are shown because a single public aggregator is weaker evidence than a reconciled multi-source estimate."
    )
    ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True)

    ly=_latest_year(wb); proj=il._base_projection(wb); model={}
    for x in proj[:3]:
        y=int(x["year"]) if isinstance(x.get("year"),(int,float)) else None
        if y:
            model[("Revenue",y)]=x.get("revenue")
            model[("EBIT Margin",y)]=x.get("margin")
            model[("FCF",y)]=x.get("fcf")
            model[("Capex / Revenue",y)]=x.get("capex_pct")
    info=getattr(wb,"_wacc_info",{}) or {}
    bundle=_consensus_bundle(ticker,ly,info); blended=bundle.get("blended") or {}; revs=_revision_map(bundle)

    _section(ws,5,"Consensus vs Independent Base Forecast",13)
    heads=["Metric","FY","Street Consensus","Independent Model","Gap","Gap % / ppt","30d Rev.","90d Rev.","Street Range","Analysts","Providers","Evidence Quality","Source / Note"]
    _header(ws,6,heads)
    years=[ly+1,ly+2,ly+3] if ly else []
    rows=[]
    for y in years:
        rows.extend([("Revenue",y,FMT_BN),("EPS",y,FMT_PRICE),("EBIT Margin",y,FMT_PCT),("FCF",y,FMT_BN),("Capex / Revenue",y,FMT_PCT)])
    r=7
    for metric,y,fmt in rows:
        dat=blended.get((metric,y)) or {}; cons=_num(dat.get("mean")); mv=model.get((metric,y))
        ws.cell(r,1,metric); ws.cell(r,2,y); ws.cell(r,3,cons); ws.cell(r,4,mv)
        ws.cell(r,3).number_format=fmt; ws.cell(r,4).number_format=fmt
        if metric in {"EBIT Margin","Capex / Revenue"}:
            ws.cell(r,5,f'=IF(OR(C{r}="",D{r}=""),"",D{r}-C{r})'); ws.cell(r,6,f'=E{r}')
            ws.cell(r,5).number_format=FMT_PCT; ws.cell(r,6).number_format=FMT_PCT
        else:
            ws.cell(r,5,f'=IF(OR(C{r}="",D{r}=""),"",D{r}-C{r})'); ws.cell(r,6,f'=IFERROR(D{r}/C{r}-1,"")')
            ws.cell(r,5).number_format=fmt; ws.cell(r,6).number_format=FMT_PCT
        rv=revs.get(y) or {}; ws.cell(r,7,rv.get("30d") if metric=="EPS" else None); ws.cell(r,8,rv.get("90d") if metric=="EPS" else None)
        ws.cell(r,7).number_format=FMT_PCT; ws.cell(r,8).number_format=FMT_PCT
        lo=_num(dat.get("low")); hi=_num(dat.get("high"))
        if lo is not None and hi is not None: ws.cell(r,9,f"{lo:,.1f} – {hi:,.1f}" if metric!="EPS" else f"${lo:.2f} – ${hi:.2f}")
        ws.cell(r,10,_num(dat.get("analysts"))); providers=int(dat.get("provider_count") or 0); ws.cell(r,11,providers)
        quality="MULTI-SOURCE" if providers>=2 else "SINGLE-SOURCE" if providers==1 else "NO PUBLIC CONSENSUS"
        ws.cell(r,12,quality); ws.cell(r,12).fill=_fill(PALE_GREEN if providers>=2 else GOLD)
        if metric=="EPS" and mv is None:
            note="Independent model does not manufacture an EPS forecast; compare EPS to normalized earnings separately."
        else:
            note=str(dat.get("providers") or "Manual / institutional input")
        ws.cell(r,13,note); ws.cell(r,13).alignment=Alignment(wrap_text=True,vertical="top")
        r+=1

    # Flag abnormal annual EPS transitions instead of silently treating the lower/higher year as a bad mapping.
    if ly:
        e0=_num((blended.get(("EPS",ly+1)) or {}).get("mean")); e1=_num((blended.get(("EPS",ly+2)) or {}).get("mean"))
        if e0 and e1 and abs(e1/e0-1)>=0.20:
            _section(ws,r+1,"Consensus Quality Flag",13)
            ws.cell(r+2,1,"EPS normalization"); ws.cell(r+2,2,"REVIEW NON-OPERATING ITEMS")
            ws.cell(r+2,3,(f"Street EPS changes from ${e0:.2f} to ${e1:.2f}. A >20% annual move can reflect one-off/non-operating items; "
                           "do not infer normalized earnings growth or a comparable forward P/E from the unusual year without reconciliation."))
            ws.merge_cells(start_row=r+2,start_column=3,end_row=r+2,end_column=13); ws.cell(r+2,3).alignment=Alignment(wrap_text=True)
            r+=4

    _section(ws,r+1,"Analyst Price Targets — Secondary Sentiment Input",13)
    _header(ws,r+2,["Provider","Mean","Median","Low","High","Upside vs Current","Source"])
    price,_,_,_= _company_inputs(wb); rr=r+3
    for pt in (bundle.get("price_targets") or [])[:4]:
        mean=_num(pt.get("mean")); vals=[pt.get("provider"),mean,_num(pt.get("median")),_num(pt.get("low")),_num(pt.get("high")),(mean/price-1) if mean and price else None,pt.get("source_url")]
        for c,v in enumerate(vals,1): ws.cell(rr,c,v)
        for c in range(2,6): ws.cell(rr,c).number_format=FMT_PRICE
        ws.cell(rr,6).number_format=FMT_PCT; _source_cell(ws.cell(rr,7),pt.get("source_url")); rr+=1
    if rr==r+3:
        ws.cell(rr,1,"No reliable public price-target snapshot returned; leave this secondary signal blank rather than estimate it.")
        rr+=1

    _section(ws,rr+1,"Variant Perception Questions",13)
    questions=[
        "Where is the independent revenue path furthest from Street consensus?",
        "Is the gap caused by volume/usage, pricing, mix, margin or capital intensity?",
        "Is consensus dispersion wide enough that the apparent variant view is actually common?",
        "Which estimate revision would matter most to intrinsic value?",
        "What evidence would force the model back toward consensus?",
    ]
    for i,q in enumerate(questions,rr+2):
        ws.cell(i,1,q); ws.merge_cells(start_row=i,start_column=2,end_row=i,end_column=8); ws.cell(i,2).fill=_fill(GOLD); ws.cell(i,2).alignment=Alignment(wrap_text=True)

    # Small, readable chart: two years, two series. No sentence-length category labels.
    chart_years=[]
    for y in years[:2]:
        c=_num((blended.get(("Revenue",y)) or {}).get("mean")); m=_num(model.get(("Revenue",y)))
        if c is not None and m is not None: chart_years.append((y,c,m))
    if chart_years:
        base=45; ws.cell(base,10,"FY"); ws.cell(base,11,"Street"); ws.cell(base,12,"Independent Model")
        for i,(y,c,m) in enumerate(chart_years,base+1): ws.cell(i,10,y); ws.cell(i,11,c); ws.cell(i,12,m)
        ch=BarChart(); ch.type="col"; ch.style=10; ch.title="Revenue: Model vs Street"; ch.y_axis.title="Revenue ($bn)"; ch.height=7.0; ch.width=12.5
        ch.add_data(Reference(ws,min_col=11,max_col=12,min_row=base+1,max_row=base+len(chart_years)),titles_from_data=False)
        ch.set_categories(Reference(ws,min_col=10,min_row=base+1,max_row=base+len(chart_years)))
        for i,series in enumerate(ch.series): series.tx=SeriesLabel(v=("Street" if i==0 else "Independent Model"))
        ch.legend.position="b"; ch.visible_cells_only=False; ws.add_chart(ch,"I7")

    widths={"A":25,"B":10,"C":18,"D":18,"E":18,"F":15,"G":13,"H":13,"I":24,"J":11,"K":11,"L":18,"M":55}
    for col,w in widths.items(): ws.column_dimensions[col].width=w
    return ws


def ensure_market_expectations_v2(wb,ticker):
    """Canonical reverse-DCF and feasible operating-hurdle sheet."""
    ws=_create(wb,"Market Expectations"); _title(ws,f"{ticker} — Market-Implied Expectations",10)
    ws["A3"]=(
        "One canonical reverse DCF is used across the workbook. The primary hurdle starts from SBC-adjusted FCF when positive; "
        "reported FCF is shown as a sensitivity. Operating hurdles are constrained to feasible ranges and are marked N/A when a one-variable solution does not exist."
    )
    ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True)

    price,shares,net_debt,target_ev=_company_inputs(wb); fcf=_fcf_inputs(wb); s=wb["Three-Case Scenarios"]; h=wb["Historical Financials"]
    wacc=_num(s["C6"].value,.09); tgr=_num(s["C7"].value,.03)
    implied_norm=_implied_growth(fcf["normalized"],target_ev,wacc,tgr)
    implied_reported=_implied_growth(fcf["reported"],target_ev,wacc,tgr)
    hurdles=_operating_hurdles(wb); proj=il._base_projection(wb)

    _section(ws,5,"Canonical Price-Implied Cash-Flow Hurdle",10)
    _header(ws,6,["Input / Output","Value","Definition / Why It Matters","Status"])
    inputs=[
        ("Current Price",price,FMT_PRICE,"Market close used by the model","FACT / MARKET"),
        ("Model Enterprise Value",target_ev,FMT_BN,"Price × valuation-equivalent shares + net debt; internally reproducible","CALCULATED"),
        ("Reported FCF",fcf["reported"],FMT_BN,"Operating cash flow − capex","REPORTED / CALCULATED"),
        ("Stock-Based Compensation",fcf["sbc"],FMT_BN,"Reported SBC; treated as an economic cost in the normalized hurdle","REPORTED"),
        ("Normalized Starting FCF",fcf["normalized"],FMT_BN,"Reported FCF − SBC when positive; otherwise reported FCF","CANONICAL INPUT"),
        ("WACC",wacc,FMT_PCT,"Base discount rate from Cost of Capital","MODEL INPUT"),
        ("Terminal Growth",tgr,FMT_PCT,"Base perpetual-growth assumption","MODEL INPUT"),
        ("Implied 10Y Normalized FCF CAGR",implied_norm,FMT_PCT,"Constant normalized-FCF growth required to match current enterprise value","PRIMARY HURDLE"),
        ("Implied 10Y Reported-FCF CAGR",implied_reported,FMT_PCT,"Sensitivity using OCF − capex without subtracting SBC","SENSITIVITY"),
    ]
    for r,(lab,val,fmt,note,status) in enumerate(inputs,7):
        ws.cell(r,1,lab); ws.cell(r,2,val); ws.cell(r,2).number_format=fmt; ws.cell(r,3,note); ws.cell(r,4,status)
        ws.cell(r,3).alignment=Alignment(wrap_text=True); ws.cell(r,4).fill=_fill(PALE_GREEN if status in {"PRIMARY HURDLE","FACT / MARKET","REPORTED"} else LIGHT)

    _section(ws,18,"Feasible One-Variable Operating Hurdles",10)
    _header(ws,19,["Metric","Independent Base","Market-Implied Hurdle","Gap","Feasible Range / Interpretation"])
    base_fcf_cagr=None
    if proj and fcf["normalized"] and fcf["normalized"]>0 and _num(proj[-1].get("fcf"),0)>0:
        base_fcf_cagr=(proj[-1]["fcf"]/fcf["normalized"])**(1/len(proj))-1
    rows=[
        ("10Y Revenue CAGR",hurdles.get("base_rev_cagr"),hurdles.get("implied_rev_cagr"),FMT_PCT,"Growth-only solve; parallel shift limited to -10ppt to +15ppt/year"),
        ("10Y Normalized FCF CAGR",base_fcf_cagr,implied_norm,FMT_PCT,"Canonical reverse-DCF hurdle; same definition as Advanced Analytics"),
        ("Terminal EBIT Margin",hurdles.get("base_margin"),hurdles.get("implied_margin"),FMT_PCT,"Margin-only solve constrained to 5%–55% mature EBIT margin"),
        ("Terminal Capex / Revenue",hurdles.get("base_capex"),hurdles.get("implied_capex"),FMT_PCT,"Capex-only solve constrained to 0.5%–30% of revenue; impossible negative capex is never displayed"),
        ("Competitive Advantage Duration",15,hurdles.get("duration"),'0 "yrs"',"Simplified duration cross-check; compare with qualitative moat/CAP evidence"),
    ]
    for r,(lab,base,imp,fmt,note) in enumerate(rows,20):
        ws.cell(r,1,lab); ws.cell(r,2,base); ws.cell(r,2).number_format=fmt
        if imp is None and lab not in {"10Y Normalized FCF CAGR"}:
            ws.cell(r,3,"N/A — no feasible one-variable solution"); ws.cell(r,4,None)
        else:
            ws.cell(r,3,imp); ws.cell(r,3).number_format=fmt; ws.cell(r,4,(imp-base) if imp is not None and base is not None else None); ws.cell(r,4).number_format=fmt
        ws.cell(r,5,note); ws.cell(r,5).alignment=Alignment(wrap_text=True)

    # MEROI/PVGO are retained, but explicitly as proxies rather than factual market expectations.
    op=_num(h["G9"].value); da=_num(h["G18"].value,0); capex=_num(h["G15"].value,0); tax=_num(s["C8"].value,.21)
    growth_cap=max(0,capex-da); nopat=op*(1-tax) if op is not None else None; reinvest=(growth_cap/nopat) if nopat and nopat>0 else None
    meroi=(implied_norm/reinvest) if implied_norm is not None and reinvest and reinvest>0 else None
    steady_ev=(fcf["normalized"]*(1+tgr)/(wacc-tgr)) if fcf["normalized"] is not None and wacc>tgr else None
    steady_share=((steady_ev-net_debt)/shares) if steady_ev is not None and shares else None
    pvgo=(price-steady_share) if steady_share is not None else None; pvgo_pct=(pvgo/price) if pvgo is not None and price else None

    _section(ws,28,"Incremental-Capital & PVGO Proxies",10)
    _header(ws,29,["Metric","Value","Interpretation / Limitation"])
    proxies=[
        ("Latest NOPAT Proxy",nopat,FMT_BN,"EBIT × (1 − normalized tax)"),
        ("Growth Capex Proxy",growth_cap,FMT_BN,"Capex − D&A; not an issuer-disclosed maintenance/growth split"),
        ("Reinvestment Rate Proxy",reinvest,FMT_PCT,"Growth-capex proxy ÷ NOPAT"),
        ("MEROI Proxy",meroi,FMT_PCT,"Implied normalized-FCF growth ÷ reinvestment proxy; use as a rough hurdle only"),
        ("MEROI − WACC",(meroi-wacc) if meroi is not None else None,FMT_PCT,"Proxy spread, not a directly observable market fact"),
        ("Steady-State Value / Share",steady_share,FMT_PRICE,"No-explicit-high-growth benchmark using normalized FCF"),
        ("PVGO / Share",pvgo,FMT_PRICE,"Price minus steady-state value"),
        ("PVGO as % of Price",pvgo_pct,FMT_PCT,"Share of price dependent on future growth opportunities under this simplified benchmark"),
    ]
    for r,(lab,val,fmt,note) in enumerate(proxies,30): ws.cell(r,1,lab); ws.cell(r,2,val); ws.cell(r,2).number_format=fmt; ws.cell(r,3,note); ws.cell(r,3).alignment=Alignment(wrap_text=True)

    _section(ws,40,"What Would Have to Be True?",10)
    ws["A41"]="These are isolated one-variable explanations, not simultaneous forecasts. A missing result means that lever alone cannot justify the current price inside the stated feasible range."
    ws.merge_cells("A41:E41"); ws["A41"].alignment=Alignment(wrap_text=True); ws["A41"].font=Font(italic=True,color=GREY)
    _header(ws,43,["Explanation","Required Revenue CAGR","Required Mature EBIT Margin","Required Mature Capex / Revenue","Assessment"])
    cases=[
        ("Growth-led",hurdles.get("implied_rev_cagr"),hurdles.get("base_margin"),hurdles.get("base_capex"),"Feasible" if hurdles.get("implied_rev_cagr") is not None else "Not feasible alone"),
        ("Margin-led",hurdles.get("base_rev_cagr"),hurdles.get("implied_margin"),hurdles.get("base_capex"),"Feasible" if hurdles.get("implied_margin") is not None else "Not feasible alone"),
        ("Capital-efficiency-led",hurdles.get("base_rev_cagr"),hurdles.get("base_margin"),hurdles.get("implied_capex"),"Feasible" if hurdles.get("implied_capex") is not None else "Not feasible alone"),
    ]
    for r,row in enumerate(cases,44):
        for c,v in enumerate(row,1): ws.cell(r,c,v)
        for c in (2,3,4): ws.cell(r,c).number_format=FMT_PCT

    # Compact chart with only three short labels; no paragraph-length categories.
    valid=[("Rev CAGR",hurdles.get("base_rev_cagr"),hurdles.get("implied_rev_cagr")),("EBIT Margin",hurdles.get("base_margin"),hurdles.get("implied_margin")),("Capex / Rev",hurdles.get("base_capex"),hurdles.get("implied_capex"))]
    valid=[x for x in valid if x[1] is not None and x[2] is not None]
    if valid:
        base=52; ws.cell(base,7,"Metric"); ws.cell(base,8,"Base"); ws.cell(base,9,"Market Implied")
        for i,(name,b,m) in enumerate(valid,base+1): ws.cell(i,7,name); ws.cell(i,8,b*100); ws.cell(i,9,m*100)
        ch=BarChart(); ch.type="col"; ch.style=10; ch.title="Base vs Price-Implied Hurdles"; ch.y_axis.title="Percent (%)"; ch.height=7; ch.width=12.5
        ch.add_data(Reference(ws,min_col=8,max_col=9,min_row=base+1,max_row=base+len(valid)),titles_from_data=False)
        _set_text_categories(ch,ws,7,base+1,base+len(valid),["Base","Market Implied"]); ch.legend.position="b"; ch.visible_cells_only=False
        ws.add_chart(ch,"F18")

    for col,w in {"A":32,"B":20,"C":28,"D":18,"E":62,"F":3,"G":16,"H":16,"I":16,"J":16}.items(): ws.column_dimensions[col].width=w
    return ws


def install_institutional_overrides():
    """Install v2 functions into institutional_layers without changing its public API."""
    il.ensure_expectations_consensus=ensure_expectations_consensus_v2
    il.ensure_market_expectations=ensure_market_expectations_v2
    return True
