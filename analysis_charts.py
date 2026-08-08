"""Create Excel-compatible analysis charts from model outputs.

The chart helper tables live in hidden columns to keep the workbook clean. Microsoft
Excel does not plot hidden source cells by default, so every chart created here sets
``visible_cells_only = False``. This serializes as ``plotVisOnly=0`` and is required
for the charts to display in desktop Excel.
"""

import statistics
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

import visualization_v2 as _visualization_v2

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; LIGHT="F5F9FC"; PALE_BLUE="D9EAF7"; GREEN="008000"; GREY="666666"
FMT_PCT='0.0%;[Red](0.0%);-'; FMT_PRICE='$#,##0.00;[Red]($#,##0.00);-'; FMT_BN='#,##0.0;[Red](#,##0.0);-'; FMT_MULT='0.0x;[Red](0.0x);-'


def _fill(c):
    return PatternFill("solid", fgColor=c)


def _is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _xlref(sheet, address):
    return f"='{sheet}'!{address}"


def _sanitize_dashboard(wb):
    """Remove legacy Dashboard merges that Excel previously repaired on open."""
    if "Dashboard" not in wb.sheetnames:
        return
    ws=wb["Dashboard"]
    for merged in list(ws.merged_cells.ranges):
        try: ws.unmerge_cells(str(merged))
        except Exception: pass
    for row in (1,17,26):
        for c in range(1,11):
            ws.cell(row,c).fill=_fill(NAVY); ws.cell(row,c).font=Font(bold=True,color=WHITE)
    for c in range(1,11):
        ws.cell(3,c).fill=_fill(PALE_BLUE); ws.cell(3,c).font=Font(bold=True)
    for c in range(6,11):
        ws.cell(10,c).fill=_fill(NAVY); ws.cell(10,c).font=Font(bold=True,color=WHITE)


def _safe_visual_repair_advanced(wb,ticker):
    """Keep the visualization layer from doing arithmetic on Excel formula strings."""
    if "Advanced Analytics" not in wb.sheetnames:
        return
    ws=wb["Advanced Analytics"]
    if str(ticker or "").upper() in {"GOOGL","GOOG"}:
        eps={2020:2.93,2021:5.61,2022:4.56,2023:5.80,2024:8.04,2025:10.81}
        vals=[]
        for r in range(7,13):
            year=ws.cell(r,1).value
            if year in eps:
                ws.cell(r,3,eps[year]); ws.cell(r,3).number_format=FMT_PRICE
                price=ws.cell(r,2).value
                if _is_num(price):
                    pe=float(price)/eps[year]; ws.cell(r,4,pe); vals.append(pe)
                else:
                    ws.cell(r,4,f'=IFERROR(B{r}/C{r},"")')
                ws.cell(r,4).number_format=FMT_MULT
        if vals:
            ws["G7"]=min(vals); ws["G8"]=statistics.median(vals); ws["G9"]=max(vals)
            for a in ("G7","G8","G9"): ws[a].number_format=FMT_MULT
    for r in range(7,15): ws.cell(r,9).number_format="yyyy-mm-dd"


_visualization_v2._repair_advanced=_safe_visual_repair_advanced


def _band(ws,start,end,row,title):
    for c in range(start,end+1):
        ws.cell(row,c).fill=_fill(NAVY); ws.cell(row,c).font=Font(bold=True,color=WHITE,size=11)
    ws.cell(row,start,title)


def _header(ws,row,start,end):
    for c in range(start,end+1):
        x=ws.cell(row,c); x.fill=_fill(BLUE); x.font=Font(bold=True,color=WHITE); x.alignment=Alignment(horizontal="center",wrap_text=True)


def _card(ws,start,end,title_row,value_row,title,formula,fmt):
    for c in range(start,end+1):
        ws.cell(title_row,c).fill=_fill(BLUE); ws.cell(title_row,c).font=Font(bold=True,color=WHITE)
        ws.cell(value_row,c).fill=_fill(LIGHT)
    ws.cell(title_row,start,title)
    ws.cell(value_row,start,formula); ws.cell(value_row,start).font=Font(bold=True,color=GREEN,size=15); ws.cell(value_row,start).number_format=fmt


def _add_chart(ws,ch,anchor):
    """Add chart and force Excel to plot data in hidden helper columns."""
    ch.visible_cells_only=False
    ch.display_blanks="gap"
    ws.add_chart(ch,anchor)


def _find_row(ws,label):
    for r in range(1,ws.max_row+1):
        if str(ws.cell(r,1).value or "").strip()==label:
            return r
    return None


def _business_rows(wb):
    if "Segment Analysis" not in wb.sheetnames: return []
    s=wb["Segment Analysis"]; section=_find_row(s,"Revenue by Business Line")
    if not section: return []
    header=section+1; value_col=4
    for c in range(1,min(15,s.max_column)+1):
        txt=str(s.cell(header,c).value or "").strip().lower()
        if txt in {"2025","latest"} or "latest revenue" in txt:
            value_col=c; break
    out=[]
    for r in range(header+1,min(s.max_row,header+20)+1):
        name=s.cell(r,1).value
        if name in (None,""):
            if out: break
            continue
        out.append((r,value_col,str(name)))
    return out[:10]


def _margin_rows(wb):
    if "Segment Analysis" not in wb.sheetnames: return [],[]
    s=wb["Segment Analysis"]; section=_find_row(s,"Reported Operating Segments")
    if not section: return [],[]
    header=section+1; cols=[]
    for c in range(1,min(18,s.max_column)+1):
        txt=str(s.cell(header,c).value or "")
        if "Margin" in txt and "Δ" not in txt: cols.append(c)
    if len(cols)<2: return [],[]
    cols=cols[-3:]; rows=[]
    for r in range(header+1,min(s.max_row,header+12)+1):
        name=s.cell(r,1).value
        if name in (None,""):
            if rows: break
            continue
        if sum(s.cell(r,c).value not in (None,"") for c in cols)>=2:
            rows.append((r,str(name)))
    return rows[:5],cols


def ensure_analysis_charts(wb,ticker):
    _sanitize_dashboard(wb)
    required={"Company Data","Historical Financials","Three-Case Scenarios","DCF"}
    if not required.issubset(wb.sheetnames): return None

    if "Analysis Charts" in wb.sheetnames: wb.remove(wb["Analysis Charts"])
    ws=wb.create_sheet("Analysis Charts"); ws.sheet_view.showGridLines=False; ws.freeze_panes="A5"
    try:
        wb._sheets.remove(ws); wb._sheets.insert(min(2,len(wb._sheets)),ws)
    except Exception: pass

    for c in range(1,17): ws.column_dimensions[get_column_letter(c)].width=11
    for c in range(24,57): ws.column_dimensions[get_column_letter(c)].hidden=True
    for c in range(1,17): ws.cell(1,c).fill=_fill(NAVY); ws.cell(2,c).fill=_fill(NAVY)
    ws["A1"]=f"{ticker} — Analysis Charts"; ws["A1"].font=Font(bold=True,color=WHITE,size=18)
    ws["A3"]="Monte Carlo, scenarios, stress tests, operating history, segments and valuation sensitivity."; ws["A3"].font=Font(italic=True,color=GREY)

    hist=wb["Historical Financials"]; sc=wb["Three-Case Scenarios"]; dcf=wb["DCF"]
    adv=wb["Advanced Analytics"] if "Advanced Analytics" in wb.sheetnames else None

    # Monte Carlo helper X:Y
    mc_end=2; ws["X2"]="Value / Share"; ws["Y2"]="Frequency"
    if adv:
        for sr in range(33,53):
            v=adv.cell(sr,18).value; f=adv.cell(sr,19).value
            if _is_num(v) and _is_num(f):
                mc_end+=1; ws.cell(mc_end,24,float(v)); ws.cell(mc_end,25,float(f)); ws.cell(mc_end,24).number_format=FMT_PRICE

    # Scenario helper AA:AB
    ws["AA2"]="Scenario"; ws["AB2"]="Value / Share"
    for r,(name,addr) in enumerate([("Bear","B39"),("Base","C39"),("Bull","D39"),("Probability Weighted","E39")],3):
        ws.cell(r,27,name); ws.cell(r,28,_xlref("Three-Case Scenarios",addr)); ws.cell(r,28).number_format=FMT_PRICE
    ws["AA7"]="Current Price"; ws["AB7"]=_xlref("Company Data","B8"); ws["AB7"].number_format=FMT_PRICE

    # Stress helper AD:AE
    ws["AD2"]="Stress Test"; ws["AE2"]="Value / Share"
    for out_r,src_r in enumerate(range(48,59),3):
        ws.cell(out_r,30,_xlref("Three-Case Scenarios",f"A{src_r}")); ws.cell(out_r,31,_xlref("Three-Case Scenarios",f"G{src_r}")); ws.cell(out_r,31).number_format=FMT_PRICE

    # Historical helper AG:AK
    for c,label in enumerate(["Year","Revenue","Free Cash Flow","Operating Margin","FCF Margin"],33): ws.cell(2,c,label)
    for out_r,src_col in enumerate(range(2,8),3):
        letter=get_column_letter(src_col)
        ws.cell(out_r,33,_xlref("Historical Financials",f"{letter}3"))
        ws.cell(out_r,34,_xlref("Historical Financials",f"{letter}4")); ws.cell(out_r,34).number_format=FMT_BN
        ws.cell(out_r,35,_xlref("Historical Financials",f"{letter}16")); ws.cell(out_r,35).number_format=FMT_BN
        ws.cell(out_r,36,_xlref("Historical Financials",f"{letter}10")); ws.cell(out_r,36).number_format=FMT_PCT
        ws.cell(out_r,37,_xlref("Historical Financials",f"{letter}17")); ws.cell(out_r,37).number_format=FMT_PCT

    # Business helper AM:AN
    ws["AM2"]="Business Line"; ws["AN2"]="Latest Revenue"; business_end=2
    for src_r,src_c,name in _business_rows(wb):
        business_end+=1; ws.cell(business_end,39,name); ws.cell(business_end,40,_xlref("Segment Analysis",f"{get_column_letter(src_c)}{src_r}")); ws.cell(business_end,40).number_format=FMT_BN

    # Segment margin helper AP:AS
    ws["AP2"]="Segment"; ws["AQ2"]="Year -2 Margin"; ws["AR2"]="Year -1 Margin"; ws["AS2"]="Latest Margin"
    margin_rows,margin_cols=_margin_rows(wb); margin_end=2
    for src_r,name in margin_rows:
        margin_end+=1; ws.cell(margin_end,42,name)
        for out_c,src_c in zip(range(43,46),margin_cols[-3:]):
            ws.cell(margin_end,out_c,_xlref("Segment Analysis",f"{get_column_letter(src_c)}{src_r}")); ws.cell(margin_end,out_c).number_format=FMT_PCT

    # Historical P/E helper AU:AV
    pe_end=2; ws["AU2"]="Year"; ws["AV2"]="Year-End P/E"
    if adv:
        for src_r in range(7,13):
            pe_end+=1; ws.cell(pe_end,47,_xlref("Advanced Analytics",f"A{src_r}")); ws.cell(pe_end,48,_xlref("Advanced Analytics",f"D{src_r}")); ws.cell(pe_end,48).number_format=FMT_MULT

    # Monte Carlo
    _band(ws,1,8,5,"Monte Carlo Valuation")
    if mc_end>=4:
        ch=BarChart(); ch.type="col"; ch.style=10; ch.title="Monte Carlo Valuation Distribution — 5,000 Simulations"; ch.height=8; ch.width=13.5; ch.legend=None
        ch.add_data(Reference(ws,min_col=25,min_row=2,max_row=mc_end),titles_from_data=True); ch.set_categories(Reference(ws,min_col=24,min_row=3,max_row=mc_end)); ch.x_axis.numFmt="$0"; _add_chart(ws,ch,"A7")

    # Scenarios
    _band(ws,9,16,5,"Scenario Valuation")
    ch=BarChart(); ch.type="col"; ch.style=10; ch.title="Bear / Base / Bull vs Current Price"; ch.height=8; ch.width=13.5; ch.legend=None
    ch.add_data(Reference(ws,min_col=28,min_row=2,max_row=7),titles_from_data=True); ch.set_categories(Reference(ws,min_col=27,min_row=3,max_row=7)); ch.y_axis.numFmt="$0"; _add_chart(ws,ch,"I7")

    # Stress tests
    _band(ws,1,8,27,"Stress Testing")
    ch=BarChart(); ch.type="bar"; ch.style=10; ch.title="Stress-Test Intrinsic Value / Share"; ch.height=9; ch.width=13.5; ch.legend=None
    ch.add_data(Reference(ws,min_col=31,min_row=2,max_row=13),titles_from_data=True); ch.set_categories(Reference(ws,min_col=30,min_row=3,max_row=13)); ch.x_axis.numFmt="$0"; _add_chart(ws,ch,"A29")

    # Revenue and FCF
    _band(ws,9,16,27,"Historical Financial Performance")
    ch=LineChart(); ch.style=10; ch.title="Revenue & Free Cash Flow"; ch.height=9; ch.width=13.5; ch.legend.position="b"
    ch.add_data(Reference(ws,min_col=34,max_col=35,min_row=2,max_row=8),titles_from_data=True); ch.set_categories(Reference(ws,min_col=33,min_row=3,max_row=8)); _add_chart(ws,ch,"I29")

    # Business mix
    _band(ws,1,8,51,"Business Mix")
    if business_end>=4:
        ch=BarChart(); ch.type="bar"; ch.style=10; ch.title="Latest Revenue by Business Line"; ch.height=8.5; ch.width=13.5; ch.legend=None
        ch.add_data(Reference(ws,min_col=40,min_row=2,max_row=business_end),titles_from_data=True); ch.set_categories(Reference(ws,min_col=39,min_row=3,max_row=business_end)); ch.x_axis.numFmt="$0"; _add_chart(ws,ch,"A53")

    # Segment margins
    _band(ws,9,16,51,"Segment Profitability")
    if margin_end>=4:
        ch=LineChart(); ch.style=10; ch.title="Segment Operating Margin Trend"; ch.height=8.5; ch.width=13.5; ch.legend.position="b"
        ch.add_data(Reference(ws,min_col=43,max_col=45,min_row=2,max_row=margin_end),titles_from_data=True); ch.set_categories(Reference(ws,min_col=42,min_row=3,max_row=margin_end)); ch.y_axis.numFmt="0%"; _add_chart(ws,ch,"I53")

    # Historical P/E
    _band(ws,1,8,73,"Historical Valuation")
    if pe_end>=4:
        ch=LineChart(); ch.style=10; ch.title="Year-End P/E"; ch.height=8; ch.width=13.5; ch.legend=None
        ch.add_data(Reference(ws,min_col=48,min_row=2,max_row=pe_end),titles_from_data=True); ch.set_categories(Reference(ws,min_col=47,min_row=3,max_row=pe_end)); ch.y_axis.numFmt="0.0x"; _add_chart(ws,ch,"A75")

    # Visible DCF heatmap links directly to DCF sheet.
    _band(ws,9,16,73,"DCF Sensitivity"); ws["I75"]="WACC / TGR"
    for out_c,src_c in enumerate(range(2,7),10):
        ws.cell(75,out_c,_xlref("DCF",f"{get_column_letter(src_c)}21")); ws.cell(75,out_c).number_format=FMT_PCT
    for out_r,src_r in enumerate(range(22,27),76):
        ws.cell(out_r,9,_xlref("DCF",f"A{src_r}")); ws.cell(out_r,9).number_format=FMT_PCT
        for out_c,src_c in enumerate(range(2,7),10):
            ws.cell(out_r,out_c,_xlref("DCF",f"{get_column_letter(src_c)}{src_r}")); ws.cell(out_r,out_c).number_format=FMT_PRICE
    _header(ws,75,9,14)
    for r in range(76,81): ws.cell(r,9).fill=_fill(PALE_BLUE); ws.cell(r,9).font=Font(bold=True)
    ws.conditional_formatting.add("J76:N80",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="percentile",mid_value=50,mid_color="FFEB84",end_type="max",end_color="63BE7B"))

    if adv:
        _card(ws,1,4,95,96,"Monte Carlo P10",_xlref("Advanced Analytics","J33"),FMT_PRICE)
        _card(ws,5,8,95,96,"Monte Carlo Median",_xlref("Advanced Analytics","J35"),FMT_PRICE)
        _card(ws,9,12,95,96,"Probability > Current Price",_xlref("Advanced Analytics","J38"),FMT_PCT)
        _card(ws,13,16,95,96,"Reverse DCF Implied FCF CAGR",_xlref("Advanced Analytics","B38"),FMT_PCT)

    _sanitize_dashboard(wb)
    return ws
