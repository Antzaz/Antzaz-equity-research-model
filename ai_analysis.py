"""Alphabet-specific AI operating and valuation analysis.

Creates two separate sheets:
- AI Analysis: reported AI adoption, Search, Cloud, backlog, capex/FCF and KPI watchlist.
- AI Valuation: AI Search economics, Cloud backlog conversion, AI capex ROI,
  AI-aware SOTP cross-check and market-vs-model bridge.

Reported inputs are blue and source-cited. User-editable assumptions are blue on yellow.
Formulas are black; internal workbook links are green.
"""

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; LIGHT="F5F9FC"; PALE_BLUE="D9EAF7"
GOLD="FFF2CC"; INPUT_BLUE="0000FF"; LINK_GREEN="008000"; GREY="666666"; BLACK="000000"
FMT_PCT='0.0%;[Red](0.0%);-'; FMT_PRICE='$#,##0.00;[Red]($#,##0.00);-'
FMT_BN='$#,##0.0;[Red]($#,##0.0);-'; FMT_NUM='#,##0.0;[Red](#,##0.0);-'; FMT_MULT='0.0x;[Red](0.0x);-'
THIN=Side(style="thin",color="D9E1F2")

Q2="https://blog.google/company-news/inside-google/message-ceo/alphabet-earnings-q2-2026/"
Q1_SEC="https://www.sec.gov/Archives/edgar/data/1652044/000165204426000048/goog-20260331.htm"
Q2_REUTERS="https://www.reuters.com/business/google-quarterly-cloud-revenue-growth-beats-expectations-2026-07-22/"
FCF_REUTERS="https://www.reuters.com/commentary/reuters-open-interest/ai-could-end-wall-streets-eps-obsession-marty-fridson-2026-08-07/"
K10="https://www.sec.gov/Archives/edgar/data/1652044/000165204426000018/goog-20251231.htm"


def _fill(c): return PatternFill("solid",fgColor=c)

def _merge(ws,rng):
    if rng not in {str(x) for x in ws.merged_cells.ranges}: ws.merge_cells(rng)

def _section(ws,rng,title):
    _merge(ws,rng); c=ws[rng.split(":")[0]]; c.value=title; c.fill=_fill(NAVY); c.font=Font(bold=True,color=WHITE,size=11)

def _header(ws,row,start,end):
    for c in range(start,end+1):
        x=ws.cell(row,c); x.fill=_fill(BLUE); x.font=Font(bold=True,color=WHITE); x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); x.border=Border(bottom=THIN)

def _reported(cell,value,source,fmt=None):
    cell.value=value; cell.font=Font(color=INPUT_BLUE); cell.comment=Comment(f"Reported/source input. Source: {source}","OpenAI")
    if fmt: cell.number_format=fmt

def _assumption(cell,value,fmt=None,note="User-editable AI scenario assumption; not separately disclosed by Alphabet."):
    cell.value=value; cell.font=Font(color=INPUT_BLUE); cell.fill=_fill(GOLD); cell.comment=Comment(note,"OpenAI")
    if fmt: cell.number_format=fmt

def _formula(cell,formula,fmt=None,internal=False):
    cell.value=formula; cell.font=Font(color=LINK_GREEN if internal else BLACK)
    if fmt: cell.number_format=fmt

def _card(ws,tr,vr,title,value,fmt,source=None,formula=False):
    _merge(ws,tr); _merge(ws,vr); t=ws[tr.split(":")[0]]; v=ws[vr.split(":")[0]]
    t.value=title; t.fill=_fill(BLUE); t.font=Font(bold=True,color=WHITE); t.alignment=Alignment(horizontal="center")
    v.value=value; v.fill=_fill(LIGHT); v.font=Font(bold=True,color=LINK_GREEN if formula else INPUT_BLUE,size=15); v.alignment=Alignment(horizontal="center",vertical="center"); v.number_format=fmt
    if source: v.comment=Comment(f"Source: {source}","OpenAI")

def _add_chart(ws,ch,anchor):
    ch.visible_cells_only=False; ch.display_blanks="gap"; ws.add_chart(ch,anchor)

def _base_style(ws):
    ws.sheet_view.showGridLines=False; ws.freeze_panes="A5"
    for c in range(1,15): ws.column_dimensions[get_column_letter(c)].width=14
    ws.column_dimensions["A"].width=31; ws.column_dimensions["F"].width=38; ws.column_dimensions["H"].width=40

def _title(ws,ticker,name,subtitle):
    _merge(ws,"A1:N2"); ws["A1"]=f"{ticker} — {name}"; ws["A1"].fill=_fill(NAVY); ws["A1"].font=Font(bold=True,color=WHITE,size=18)
    _merge(ws,"A3:N3"); ws["A3"]=subtitle; ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True)


def _operating_sheet(wb,ticker):
    if "AI Analysis" in wb.sheetnames: wb.remove(wb["AI Analysis"])
    ws=wb.create_sheet("AI Analysis"); _base_style(ws)
    _title(ws,ticker,"AI Operating Analysis","Separate AI evidence layer: adoption, Search effects, Cloud demand, backlog, capital intensity and quarterly thesis checks. As of Q2 2026.")

    _card(ws,"A5:C5","A6:C8","Q2 Cloud Growth",0.82,FMT_PCT,Q2)
    _card(ws,"D5:F5","D6:F8","Cloud Backlog ($bn)",514.0,FMT_BN,Q2)
    _card(ws,"G5:I5","G6:I8","Gemini MAU (m)",950.0,FMT_NUM,Q2)
    _card(ws,"J5:L5","J6:L8","AI Mode MAU",1.0,'0.0"bn"',Q2)
    _card(ws,"M5:N5","M6:N8","2026 Capex Midpoint",200.0,FMT_BN,Q2_REUTERS)

    _section(ws,"A10:N10","Latest AI Operating Evidence")
    heads=["KPI","Prior / Q1","Q2 / Latest","Change","Signal","Investment Read-Through","Source"]
    for c,h in enumerate(heads,1): ws.cell(11,c,h)
    _header(ws,11,1,7)
    rows=[
        ("Alphabet revenue ($bn)",109.9,119.8,"=(C12/B12)-1","Positive","AI adoption is occurring alongside broad company growth.",Q2),
        ("Search & Other growth",0.19,0.17,"=C13-B13","Positive / watch","Search remains double-digit while AI features expand; monitor monetization/query economics.",Q2),
        ("Cloud revenue ($bn)",20.0,24.8,"=(C14/B14)-1","Very strong","82% YoY growth supports AI infrastructure and enterprise-solutions demand.",Q2),
        ("Cloud YoY growth",0.63,0.82,"=C15-B15","Very strong","Acceleration suggests AI demand has not normalized yet.",Q2),
        ("Cloud backlog ($bn)",462.3,514.0,"=(C16/B16)-1","Very strong","Forward demand is large, though recognition depends on delivery and customer utilization.",Q1_SEC),
        ("Model API tokens/min (bn)",16.0,22.0,"=(C17/B17)-1","Very strong","Direct usage indicator across developers and enterprises.",Q2),
        ("Gemini app MAU (m)",750.0,950.0,"=(C18/B18)-1","Strong","Near-billion-user distribution; monetization per user is still the key economic question.",Q2),
        ("AI Mode MAU (bn)",None,1.0,None,"Strong","Management says AI Mode is driving incremental Search queries.",Q2),
        ("Fortune 100 using Gemini Enterprise",None,0.90,None,"Strong","Broad enterprise penetration; paid usage and margins matter next.",Q2),
        ("Cloud operating margin",None,8.8/24.8,None,"Strong","Q2 Cloud operating income of ~$8.8bn implies ~35.5% margin, supporting improving AI economics.",Q2_REUTERS),
        ("Q2 capex ($bn)",None,44.92,None,"High investment","AI infrastructure spending is large enough to dominate near-term cash flow.",FCF_REUTERS),
        ("Q2 free cash flow ($bn)",None,-5.9,None,"Risk","Negative FCF makes return on AI capital the central valuation question.",Q2_REUTERS),
        ("2026 capex midpoint ($bn)",185.0,200.0,"=(C24/B24)-1","Risk / opportunity","Higher guidance means stronger demand and higher return-on-capital risk at the same time.",Q2_REUTERS),
    ]
    for r,row in enumerate(rows,12):
        for c,v in enumerate(row,1): ws.cell(r,c,v)
        for c in (2,3):
            if isinstance(ws.cell(r,c).value,(int,float)): ws.cell(r,c).font=Font(color=INPUT_BLUE)
        if isinstance(ws.cell(r,4).value,str): ws.cell(r,4).font=Font(color=BLACK)
        ws.cell(r,6).alignment=Alignment(wrap_text=True,vertical="top"); ws.cell(r,7).font=Font(color=LINK_GREEN); ws.cell(r,7).alignment=Alignment(wrap_text=True)
    for r in [12,14,16,17,18,22,23,24]: ws.cell(r,2).number_format=FMT_BN if r not in [17,18] else FMT_NUM; ws.cell(r,3).number_format=FMT_BN if r not in [17,18] else FMT_NUM
    for r in [13,15,20,21]: ws.cell(r,2).number_format=FMT_PCT; ws.cell(r,3).number_format=FMT_PCT
    ws["C19"].number_format='0.0"bn"'
    for r in [12,13,14,15,16,17,18,24]: ws.cell(r,4).number_format=FMT_PCT
    ws.conditional_formatting.add("D12:D24",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="num",mid_value=0,mid_color="FFEB84",end_type="max",end_color="63BE7B"))

    ch=BarChart(); ch.type="col"; ch.style=10; ch.title="Cloud Revenue: Q1 vs Q2 2026"; ch.height=7; ch.width=11; ch.legend=None; ch.add_data(Reference(ws,min_col=2,max_col=3,min_row=14,max_row=14)); ch.set_categories(Reference(ws,min_col=2,max_col=3,min_row=11,max_row=11)); ch.y_axis.title="$bn"; ch.dLbls=DataLabelList(); ch.dLbls.showVal=True; _add_chart(ws,ch,"J11")
    ch=BarChart(); ch.type="col"; ch.style=10; ch.title="Cloud Backlog: Q1 vs Q2 2026"; ch.height=7; ch.width=11; ch.legend=None; ch.add_data(Reference(ws,min_col=2,max_col=3,min_row=16,max_row=16)); ch.set_categories(Reference(ws,min_col=2,max_col=3,min_row=11,max_row=11)); ch.y_axis.title="$bn"; ch.dLbls=DataLabelList(); ch.dLbls.showVal=True; _add_chart(ws,ch,"J25")

    _section(ws,"A27:H27","AI Thesis Scorecard — Evidence, Not a Price Target")
    score=[
        ["Dimension","Current read","Evidence","What would strengthen it","Main valuation linkage"],
        ["AI demand","Strong","Cloud +82%; backlog $514bn; supply constrained","Backlog and paid usage continue to compound","Cloud growth / duration"],
        ["Consumer adoption","Strong","Gemini 950m MAU; AI Mode >1bn MAU","Usage converts to subscriptions and Search economics","Search / subscriptions"],
        ["Search defense / expansion","Positive","Search +17%; AI Mode reportedly adds queries","Revenue/query and contribution margin remain resilient","Services multiple / margin"],
        ["Enterprise monetization","Strong but incomplete","~90% Fortune 100 using Gemini Enterprise","Paid token usage and Cloud margin expand together","Cloud margin / multiple"],
        ["Capital efficiency","Unproven","~$200bn capex midpoint; Q2 FCF negative","FCF rebounds as utilization rises","Highest AI valuation risk"],
        ["Execution / competition","Mixed-positive","Usage and breadth are strong; frontier competition remains intense","Share gains and cost/query keep improving","AI regime probabilities"],
    ]
    for r,row in enumerate(score,28):
        for c,v in enumerate(row,1): ws.cell(r,c,v); ws.cell(r,c).alignment=Alignment(wrap_text=True,vertical="top")
    _header(ws,28,1,5)
    for r in range(29,35): ws.row_dimensions[r].height=42

    _section(ws,"A37:N37","Quarterly AI KPI Watchlist")
    watch=[
        ["KPI","Bullish evidence","Bearish evidence","Current observation","Why it matters","Source"],
        ["Search growth",">15% with resilient monetization","<10% or falling commercial economics","17% Q2","Tests AI expansion vs cannibalization","Google Q2"],
        ["Cloud growth",">40%","<25% while capex stays high","82% Q2","Direct AI infrastructure / enterprise signal","Google Q2"],
        ["Cloud backlog","Outgrows revenue","Flat/down or cancellations","$514bn Q2","Forward demand / capacity visibility","Google Q2"],
        ["Cloud margin","Expands with AI growth","Falls materially despite scale","~35.5% Q2","Tests unit economics","Q2 results"],
        ["Gemini / AI Mode","Users + engagement + monetization rise","Usage stalls or monetization fails","950m / >1bn MAU","Distribution only matters if economics follow","Google Q2"],
        ["Capex / FCF","Capex growth slows before AI growth","Capex stays high as growth/backlog slow","~$200bn midpoint; Q2 FCF negative","Core return-on-capital test","Reuters"],
        ["Cost per AI response","Falls faster than usage rises","Stays structurally high","Management says AI Mode cost hit a post-launch low","Search contribution margin","Google Q2"],
    ]
    for r,row in enumerate(watch,38):
        for c,v in enumerate(row,1): ws.cell(r,c,v); ws.cell(r,c).alignment=Alignment(wrap_text=True,vertical="top")
    _header(ws,38,1,6)
    for r in range(39,46): ws.row_dimensions[r].height=43

    _section(ws,"A48:N48","Sources")
    sources=[["Google Q2 2026 CEO remarks",Q2],["Alphabet Q1 2026 10-Q",Q1_SEC],["Reuters Q2 results / capex",Q2_REUTERS],["Reuters AI capex / FCF analysis",FCF_REUTERS],["Alphabet 2025 10-K",K10]]
    for r,row in enumerate(sources,49): ws.cell(r,1,row[0]); ws.cell(r,2,row[1]); ws.cell(r,2).font=Font(color=LINK_GREEN); ws.cell(r,2).alignment=Alignment(wrap_text=True)
    return ws


def _valuation_sheet(wb,ticker):
    if "AI Valuation" in wb.sheetnames: wb.remove(wb["AI Valuation"])
    ws=wb.create_sheet("AI Valuation"); _base_style(ws)
    _title(ws,ticker,"AI Economics & Valuation","Alternative AI-specific cross-checks. Yellow cells are explicit assumptions. Do not add SOTP values on top of the DCF; these are alternative methods.")

    _section(ws,"A5:N5","Market vs Traditional Model")
    _card(ws,"A6:C6","A7:C9","Current Price","='Company Data'!B8",FMT_PRICE,formula=True)
    _card(ws,"D6:F6","D7:F9","Base DCF","='Three-Case Scenarios'!C39",FMT_PRICE,formula=True)
    _card(ws,"G6:I6","G7:I9","Market Premium vs Base","=IFERROR('Company Data'!B8/'Three-Case Scenarios'!C39-1,0)",FMT_PCT,formula=True)
    _card(ws,"J6:L6","J7:L9","Reverse DCF FCF CAGR","='Advanced Analytics'!B38",FMT_PCT,formula=True)
    _card(ws,"M6:N6","M7:N9","Shares (bn)","=IFERROR('Company Data'!B10/'Company Data'!B8,0)",FMT_NUM,formula=True)
    _merge(ws,"A10:N11"); ws["A10"]="The market–DCF gap is not automatically 'AI value.' It is the combined effect of higher growth, margins, duration, multiples and optionality that must be explained."; ws["A10"].fill=_fill(LIGHT); ws["A10"].font=Font(italic=True,color=GREY); ws["A10"].alignment=Alignment(wrap_text=True)

    _section(ws,"A13:E13","AI Search Economics")
    for c,h in enumerate(["Metric","Bear","Base","Bull","Definition"],1): ws.cell(14,c,h)
    _header(ws,14,1,5)
    rows=[
        ("AI-enabled query share",0.45,0.60,0.75,"Share of Search activity meaningfully using AI features."),
        ("Incremental query volume from AI",0.03,0.08,0.15,"Additional query volume attributable to AI experiences."),
        ("Monetization vs classic Search",0.80,1.00,1.10,"Revenue efficiency of AI-enabled queries versus classic Search."),
        ("Incremental AI serving-cost drag",0.12,0.08,0.05,"Incremental cost drag as a share of AI-enabled Search baseline."),
        ("2025 Search revenue ($bn)",224.532,224.532,224.532,"Reported Search & Other revenue."),
        ("Net AI Search economics uplift",None,None,None,"AI share × (query uplift × monetization − cost drag)."),
        ("Net Search economics impact ($bn)",None,None,None,"Search revenue × net AI economics uplift."),
    ]
    for r,(lab,b,c,d,note) in enumerate(rows,15):
        ws.cell(r,1,lab); ws.cell(r,5,note); ws.cell(r,5).alignment=Alignment(wrap_text=True)
        if r<=18:
            for col,val in zip(range(2,5),[b,c,d]): _assumption(ws.cell(r,col),val,FMT_PCT)
        elif r==19:
            for col,val in zip(range(2,5),[b,c,d]): _reported(ws.cell(r,col),val,K10,FMT_BN)
        elif r==20:
            for col in range(2,5): L=get_column_letter(col); _formula(ws.cell(r,col),f"={L}15*({L}16*{L}17-{L}18)",FMT_PCT)
        elif r==21:
            for col in range(2,5): L=get_column_letter(col); _formula(ws.cell(r,col),f"={L}19*{L}20",FMT_BN)
    ws.conditional_formatting.add("B20:D21",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="num",mid_value=0,mid_color="FFEB84",end_type="max",end_color="63BE7B"))
    ch=BarChart(); ch.type="col"; ch.style=10; ch.title="AI Search Economics Impact"; ch.height=7; ch.width=12; ch.legend=None; ch.add_data(Reference(ws,min_col=2,max_col=4,min_row=21,max_row=21)); ch.set_categories(Reference(ws,min_col=2,max_col=4,min_row=14,max_row=14)); ch.y_axis.title="$bn"; ch.dLbls=DataLabelList(); ch.dLbls.showVal=True; _add_chart(ws,ch,"G13")

    _section(ws,"A24:E24","Cloud Backlog Conversion")
    cloud=[["Metric","Value","Type","Interpretation"],["Q2 Cloud backlog ($bn)",514.0,"Reported","Contracted demand not yet recognized as revenue."],["24-month conversion assumption",0.52,"Assumption","Editable shorthand for 'just over 50%' recognition over 24 months."],["Backlog expected within 24m ($bn)","=B26*B27","Formula","Backlog × conversion assumption."],["Annualized backlog conversion ($bn)","=B28/2","Formula","Simple annualized run-rate from existing backlog."],["Q2 Cloud revenue annualized ($bn)","=24.8*4","Formula","Q2 revenue × 4; not a forecast."],["Coverage ratio","=B29/B30","Formula","Annualized backlog conversion ÷ current Cloud run-rate."]]
    for r,row in enumerate(cloud,25):
        for c,v in enumerate(row,1): ws.cell(r,c,v); ws.cell(r,c).alignment=Alignment(wrap_text=True,vertical="top")
    _header(ws,25,1,4); _reported(ws["B26"],514.0,Q2,FMT_BN); _assumption(ws["B27"],0.52,FMT_PCT,"Editable shorthand based on Alphabet's 'just over 50%' backlog recognition disclosure."); _formula(ws["B28"],"=B26*B27",FMT_BN); _formula(ws["B29"],"=B28/2",FMT_BN); _formula(ws["B30"],"=24.8*4",FMT_BN); _formula(ws["B31"],"=B29/B30",FMT_MULT)

    _section(ws,"A34:E34","AI Infrastructure / Capex Return")
    for c,h in enumerate(["Metric","Bear","Base","Bull","Definition"],1): ws.cell(35,c,h)
    _header(ws,35,1,5)
    cap=[("2026 total capex midpoint ($bn)",200,200,200,"Reported midpoint of $195–205bn guidance."),("AI / compute share of capex",0.50,0.60,0.70,"Model assumption; no clean standalone AI capex disclosure."),("Implied AI / compute capex ($bn)",None,None,None,"Total capex × AI share."),("Mature incremental annual AI revenue ($bn)",35,55,85,"Scenario revenue enabled by incremental infrastructure."),("Incremental contribution margin",0.30,0.40,0.50,"Revenue less direct serving / operating costs before tax."),("Pre-tax cash return",None,None,None,"Incremental contribution profit ÷ AI capex."),("After-tax cash return",None,None,None,"Pre-tax return × (1 − 21%)."),("Simple payback (years)",None,None,None,"AI capex ÷ annual contribution profit.")]
    for r,(lab,b,c,d,note) in enumerate(cap,36):
        ws.cell(r,1,lab); ws.cell(r,5,note); ws.cell(r,5).alignment=Alignment(wrap_text=True)
        if r==36:
            for col in range(2,5): _reported(ws.cell(r,col),200,Q2_REUTERS,FMT_BN)
        elif r in [37,39,40]:
            fmt=FMT_PCT if r in [37,40] else FMT_BN
            for col,val in zip(range(2,5),[b,c,d]): _assumption(ws.cell(r,col),val,fmt)
        elif r==38:
            for col in range(2,5): L=get_column_letter(col); _formula(ws.cell(r,col),f"={L}36*{L}37",FMT_BN)
        elif r==41:
            for col in range(2,5): L=get_column_letter(col); _formula(ws.cell(r,col),f"=IFERROR({L}39*{L}40/{L}38,0)",FMT_PCT)
        elif r==42:
            for col in range(2,5): L=get_column_letter(col); _formula(ws.cell(r,col),f"={L}41*(1-21%)",FMT_PCT)
        elif r==43:
            for col in range(2,5): L=get_column_letter(col); _formula(ws.cell(r,col),f"=IFERROR({L}38/({L}39*{L}40),0)",'0.0"x"')
    ws.conditional_formatting.add("B42:D42",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="percentile",mid_value=50,mid_color="FFEB84",end_type="max",end_color="63BE7B"))
    ch=BarChart(); ch.type="col"; ch.style=10; ch.title="After-Tax Cash Return on AI Capex"; ch.height=7; ch.width=12; ch.legend=None; ch.add_data(Reference(ws,min_col=2,max_col=4,min_row=42,max_row=42)); ch.set_categories(Reference(ws,min_col=2,max_col=4,min_row=35,max_row=35)); ch.y_axis.numFmt="0%"; ch.dLbls=DataLabelList(); ch.dLbls.showVal=True; _add_chart(ws,ch,"G34")

    _section(ws,"A46:E46","AI-Aware SOTP Cross-Check — Alternative to DCF")
    for c,h in enumerate(["Metric","Bear","Base","Bull","Method / note"],1): ws.cell(47,c,h)
    _header(ws,47,1,5)
    sotp=[("Google Services 2025 EBIT ($bn)",139.404,139.404,139.404,"Reported segment operating income."),("Services EV / EBIT",16,20,25,"Scenario multiple for Search/YouTube durability and AI defense."),("Google Services EV ($bn)",None,None,None,"EBIT × multiple."),("Q2 Cloud annualized revenue ($bn)",99.2,99.2,99.2,"$24.8bn × 4; run-rate, not forecast."),("Cloud EV / Revenue",6,9,12,"Scenario multiple for high-growth AI Cloud mix."),("Google Cloud EV ($bn)",None,None,None,"Run-rate revenue × multiple."),("Other Bets / optionality ($bn)",25,50,100,"Scenario value; not reported fair value."),("Shared AI / corporate haircut ($bn)",-150,-100,-75,"Capitalized drag for shared AI R&D / corporate costs."),("Net cash ($bn)",None,None,None,"Cash less debt."),("Equity value ($bn)",None,None,None,"Services + Cloud + optionality + haircut + net cash."),("Shares (bn)",None,None,None,"Market cap ÷ current price."),("SOTP value / share",None,None,None,"Equity value ÷ shares."),("Upside / (downside)",None,None,None,"SOTP value/share ÷ current price − 1.")]
    for r,(lab,b,c,d,note) in enumerate(sotp,48):
        ws.cell(r,1,lab); ws.cell(r,5,note); ws.cell(r,5).alignment=Alignment(wrap_text=True)
        if r in [48,51]:
            src=K10 if r==48 else Q2
            for col,val in zip(range(2,5),[b,c,d]): _reported(ws.cell(r,col),val,src,FMT_BN)
        elif r in [49,52,54,55]:
            fmt=FMT_MULT if r in [49,52] else FMT_BN
            for col,val in zip(range(2,5),[b,c,d]): _assumption(ws.cell(r,col),val,fmt,"User-editable SOTP assumption; cross-check only.")
        elif r==50:
            for col in range(2,5): L=get_column_letter(col); _formula(ws.cell(r,col),f"={L}48*{L}49",FMT_BN)
        elif r==53:
            for col in range(2,5): L=get_column_letter(col); _formula(ws.cell(r,col),f"={L}51*{L}52",FMT_BN)
        elif r==56:
            for col in range(2,5): _formula(ws.cell(r,col),"='Company Data'!B12-'Company Data'!B13",FMT_BN,True)
        elif r==57:
            for col in range(2,5): L=get_column_letter(col); _formula(ws.cell(r,col),f"={L}50+{L}53+{L}54+{L}55+{L}56",FMT_BN)
        elif r==58:
            for col in range(2,5): _formula(ws.cell(r,col),"=IFERROR('Company Data'!B10/'Company Data'!B8,0)",FMT_NUM,True)
        elif r==59:
            for col in range(2,5): L=get_column_letter(col); _formula(ws.cell(r,col),f"=IFERROR({L}57/{L}58,0)",FMT_PRICE)
        elif r==60:
            for col in range(2,5): L=get_column_letter(col); _formula(ws.cell(r,col),f"=IFERROR({L}59/'Company Data'!B8-1,0)",FMT_PCT,True)
    ws.conditional_formatting.add("B59:D60",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="percentile",mid_value=50,mid_color="FFEB84",end_type="max",end_color="63BE7B"))

    # Hidden helper table for SOTP chart.
    ws.column_dimensions["L"].hidden=True; ws.column_dimensions["M"].hidden=True
    helper=[["Case","Value / Share"],["Bear","=B59"],["Base","=C59"],["Bull","=D59"],["Current Price","='Company Data'!B8"]]
    for r,row in enumerate(helper,48):
        for c,v in enumerate(row,12): ws.cell(r,c,v)
    ch=BarChart(); ch.type="col"; ch.style=10; ch.title="AI-Aware SOTP vs Current Price"; ch.height=8; ch.width=12; ch.legend=None; ch.add_data(Reference(ws,min_col=13,min_row=48,max_row=52),titles_from_data=True); ch.set_categories(Reference(ws,min_col=12,min_row=49,max_row=52)); ch.y_axis.numFmt="$0"; ch.dLbls=DataLabelList(); ch.dLbls.showVal=True; _add_chart(ws,ch,"G49")

    _section(ws,"G24:N24","Market vs AI Cross-Check")
    bridge=[["Metric","Value","Interpretation"],["Current price","='Company Data'!B8","Market reference."],["Base DCF","='Three-Case Scenarios'!C39","Traditional consolidated DCF."],["Base AI-aware SOTP","=C59","Alternative segment / multiple cross-check."],["Premium vs DCF ($/share)","=H26-H27","Amount market price exceeds Base DCF."],["Premium vs SOTP ($/share)","=H26-H28","Residual not explained by Base SOTP."],["Residual value ($bn)","=H30*$N$7","Rough residual for extra duration / optionality / multiple premium."]]
    for r,row in enumerate(bridge,25):
        for c,v in enumerate(row,7): ws.cell(r,c,v); ws.cell(r,c).alignment=Alignment(wrap_text=True,vertical="top")
    _header(ws,25,7,9)
    for r in range(26,32): ws.cell(r,8).number_format=FMT_BN if r==31 else FMT_PRICE

    _section(ws,"A63:N63","Interpretation & Guardrails")
    guard=[("Do not add SOTP on top of DCF.","They are alternative valuation methods for the same businesses."),("Do not treat Search AI assumptions as reported facts.","Alphabet does not disclose AI query monetization or contribution profit separately."),("Do not add growth capex back to FCF.","The correct question is whether future incremental profit earns an adequate return on today's cash investment."),("Backlog is not immediate guaranteed revenue.","Timing depends on contracts, utilization, delivery and capacity; Q1 also changed the backlog definition."),("A professional price target can differ because of duration and multiples.","Use the bridge to identify exactly which assumptions explain the gap rather than forcing the DCF upward.")]
    for r,(a,b) in enumerate(guard,64): ws.cell(r,1,a); ws.cell(r,2,b); ws.cell(r,1).font=Font(bold=True,color=NAVY); ws.cell(r,2).alignment=Alignment(wrap_text=True); ws.row_dimensions[r].height=34

    _section(ws,"A71:N71","Primary Sources")
    for r,(name,url) in enumerate([("Google Q2 2026 CEO remarks",Q2),("Alphabet Q1 2026 10-Q",Q1_SEC),("Reuters Q2 results / capex",Q2_REUTERS),("Reuters AI capex / FCF",FCF_REUTERS),("Alphabet 2025 Form 10-K",K10)],72): ws.cell(r,1,name); ws.cell(r,2,url); ws.cell(r,2).font=Font(color=LINK_GREEN)
    return ws


def ensure_ai_analysis(wb,ticker):
    ticker=(ticker or "").upper()
    if ticker not in {"GOOGL","GOOG"}:
        for name in ["AI Analysis","AI Valuation"]:
            if name in wb.sheetnames: wb.remove(wb[name])
        return None
    a=_operating_sheet(wb,ticker); v=_valuation_sheet(wb,ticker)
    try:
        for sh in [v,a]: wb._sheets.remove(sh)
        idx=min(14,len(wb._sheets)); wb._sheets.insert(idx,a); wb._sheets.insert(idx+1,v)
    except Exception: pass
    return a,v
