"""Cross-company cleanup for legacy template tabs.

The clean template originated as an Alphabet model. This module removes stale template
labels/sources, creates a generic Business Portfolio Map from the current Segment Analysis,
adds an automatic context panel beside the manual Research Notes workbench, and repairs
cross-company data-quality checks after all company-specific modules have run.
"""

from openpyxl.styles import PatternFill, Font, Alignment

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; LIGHT="F5F9FC"; GOLD="FFF2CC"
PALE_GREEN="E2F0D9"; PALE_RED="FCE4D6"; GREY="666666"; LINK_GREEN="008000"; INPUT_BLUE="0000FF"
FMT_BN='#,##0.0;[Red](#,##0.0);-'; FMT_PCT='0.0%;[Red](0.0%);-'; FMT_PRICE='$#,##0.00;[Red]($#,##0.00);-'


def _fill(c): return PatternFill("solid",fgColor=c)
def _num(v):
    try:
        if isinstance(v,bool): return None
        return float(v)
    except Exception:
        return None

def _section(ws,row,title,end=8):
    for c in range(1,end+1):
        ws.cell(row,c).fill=_fill(NAVY); ws.cell(row,c).font=Font(bold=True,color=WHITE,size=11)
    ws.cell(row,1,title)

def _header(ws,row,start,end):
    for c in range(start,end+1):
        x=ws.cell(row,c); x.fill=_fill(BLUE); x.font=Font(bold=True,color=WHITE)
        x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

def _segment_status(seg):
    if seg is None: return "Missing"
    for r in range(1,seg.max_row+1):
        if str(seg.cell(r,1).value or "").strip()=="Extraction Status":
            return str(seg.cell(r,2).value or "")
    return "Unknown"

def _business_start(seg):
    if seg is None: return None
    for r in range(1,seg.max_row+1):
        if str(seg.cell(r,1).value or "").strip()=="Revenue by Business Line": return r
    return None

def _segment_counts(seg):
    if seg is None: return 0,0
    segments=0
    for r in range(7,min(seg.max_row,30)+1):
        name=seg.cell(r,1).value; latest=_num(seg.cell(r,4).value)
        if name and latest is not None: segments+=1
        if str(name or "").strip()=="Revenue by Business Line": break
    business=0; bs=_business_start(seg)
    if bs:
        for r in range(bs+2,min(seg.max_row,bs+20)+1):
            name=seg.cell(r,1).value; latest=_num(seg.cell(r,4).value)
            if name and latest is not None: business+=1
            if str(name or "").strip()=="Source & Data Quality": break
    return segments,business

def _refresh_labels(wb,ticker):
    if "Comparative Analysis" in wb.sheetnames:
        ws=wb["Comparative Analysis"]
        ws["A1"]=f"Comparative Analysis — {ticker} vs Selected Peers"
        ws["B3"]=ticker
    if "Dashboard" in wb.sheetnames:
        ws=wb["Dashboard"]
        if ws["A1"].value: ws["A1"]=f"{ticker} Long-Term Value Investing Dashboard"
        if ws["A18"].value=="Metric": ws["B18"]=ticker

def _refresh_historical_sources(wb):
    if "Historical Financials" not in wb.sheetnames: return
    h=wb["Historical Financials"]
    for r in range(24,29):
        for c in range(1,8): h.cell(r,c).value=None
    h["A24"]="Sources"
    filings=wb["Filings"] if "Filings" in wb.sheetnames else None
    rows=[]
    if filings:
        for r in range(4,min(filings.max_row,30)+1):
            if str(filings.cell(r,1).value or "").strip()=="10-K":
                rows.append((filings.cell(r,2).value,filings.cell(r,4).value))
    if rows:
        for i,(period,url) in enumerate(rows[:3],25):
            h.cell(i,1).value=f"10-K {period or ''}".strip(); h.cell(i,2).value=url
            h.cell(i,2).font=Font(color=LINK_GREEN,underline="single")
    else:
        h["A25"]="SEC filings"; h["B25"]="See Filings tab"

def ensure_business_portfolio_map(wb,ticker):
    if "Business Portfolio Map" in wb.sheetnames: wb.remove(wb["Business Portfolio Map"])
    ws=wb.create_sheet("Business Portfolio Map"); ws.sheet_view.showGridLines=False
    for c in range(1,13): ws.cell(1,c).fill=_fill(NAVY); ws.cell(2,c).fill=_fill(NAVY)
    ws["A1"]=f"{ticker} — Business Portfolio Map"; ws["A1"].font=Font(bold=True,color=WHITE,size=18)
    ws["A3"]="Cross-company operating map built from Company Data, Historical Financials and issuer-disclosed Segment Analysis. Manual interpretation cells are yellow."
    ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True)

    d=wb["Company Data"] if "Company Data" in wb.sheetnames else None
    h=wb["Historical Financials"] if "Historical Financials" in wb.sheetnames else None
    seg=wb["Segment Analysis"] if "Segment Analysis" in wb.sheetnames else None

    _section(ws,5,"Company Overview",8); _header(ws,6,1,4)
    for c,v in enumerate(["Metric","Value","Metric","Value"],1): ws.cell(6,c,v)
    overview=[
        ("Company",d["B5"].value if d else None,"Ticker",ticker),
        ("Sector",d["B6"].value if d else None,"Industry",d["B7"].value if d else None),
        ("Current Price",d["B8"].value if d else None,"Market Cap ($bn)",d["B10"].value if d else None),
        ("Latest Revenue ($bn)",h["G4"].value if h else None,"Operating Margin",h["G10"].value if h else None),
        ("FCF Margin",h["G17"].value if h else None,"Forward P/E",d["B15"].value if d else None),
        ("Segment Extraction",_segment_status(seg),"Latest 10-K",None),
    ]
    filings=wb["Filings"] if "Filings" in wb.sheetnames else None
    if filings:
        for r in range(4,min(filings.max_row,30)+1):
            if str(filings.cell(r,1).value or "")=="10-K":
                overview[-1]=(overview[-1][0],overview[-1][1],"Latest 10-K",filings.cell(r,4).value); break
    for r,row in enumerate(overview,7):
        for c,v in enumerate(row,1): ws.cell(r,c,v)
    ws["B9"].number_format=FMT_PRICE; ws["D9"].number_format=FMT_BN; ws["B10"].number_format=FMT_BN; ws["D10"].number_format=FMT_PCT; ws["B11"].number_format=FMT_PCT

    _section(ws,15,"Reported Operating Segments",14)
    heads=["Segment","Latest Revenue ($bn)","Revenue Mix","Latest Op. Income ($bn)","Operating Margin","Growth","2Y CAGR","Margin Δ","Source"]
    for c,v in enumerate(heads,1): ws.cell(16,c,v)
    _header(ws,16,1,9)
    seg_rows=[]
    if seg:
        for r in range(7,17):
            name=seg.cell(r,1).value; latest=_num(seg.cell(r,4).value)
            if name and latest is not None:
                seg_rows.append((name,latest,seg.cell(r,14).value,seg.cell(r,9).value,seg.cell(r,12).value,seg.cell(r,5).value,seg.cell(r,6).value,seg.cell(r,13).value,"Segment Analysis / SEC 10-K"))
    for i in range(10):
        r=17+i
        if i<len(seg_rows):
            for c,v in enumerate(seg_rows[i],1): ws.cell(r,c,v)
        else:
            for c in range(1,10): ws.cell(r,c).fill=_fill(LIGHT)
        for c in (3,5,6,7,8): ws.cell(r,c).number_format=FMT_PCT
        for c in (2,4): ws.cell(r,c).number_format=FMT_BN

    _section(ws,29,"Revenue Groups / Business Mix",10)
    heads=["Revenue Group","Latest Revenue ($bn)","Mix","Growth","2Y CAGR","Source"]
    for c,v in enumerate(heads,1): ws.cell(30,c,v)
    _header(ws,30,1,6)
    business=[]; bs=_business_start(seg)
    if seg and bs:
        for r in range(bs+2,min(seg.max_row,bs+14)+1):
            name=seg.cell(r,1).value; latest=_num(seg.cell(r,4).value)
            if name and latest is not None:
                business.append((name,latest,seg.cell(r,7).value,seg.cell(r,5).value,seg.cell(r,6).value,"Segment Analysis / SEC 10-K"))
    for i in range(12):
        r=31+i
        if i<len(business):
            for c,v in enumerate(business[i],1): ws.cell(r,c,v)
        else:
            for c in range(1,7): ws.cell(r,c).fill=_fill(LIGHT)
        ws.cell(r,2).number_format=FMT_BN
        for c in (3,4,5): ws.cell(r,c).number_format=FMT_PCT

    _section(ws,46,"Your Business-Mix Interpretation",10)
    prompts=["Which segment creates the most incremental value?","Which segment deserves the highest multiple and why?","Where is margin expansion / compression occurring?","Which revenue group is gaining mix?","What segment-level KPI would invalidate the thesis?","Does the market value the segments differently from your view?"]
    for r,prompt in enumerate(prompts,47):
        ws.cell(r,1,prompt); ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=8)
        ws.cell(r,2).fill=_fill(GOLD); ws.cell(r,2).font=Font(color=INPUT_BLUE); ws.cell(r,2).alignment=Alignment(wrap_text=True,vertical="top"); ws.row_dimensions[r].height=28

    for c,w in {"A":40,"B":19,"C":16,"D":20,"E":18,"F":16,"G":16,"H":16,"I":30}.items(): ws.column_dimensions[c].width=w
    ws.freeze_panes="A16"
    return ws

def seed_research_context(wb,ticker):
    if "Research Notes" not in wb.sheetnames: return
    ws=wb["Research Notes"]; d=wb["Company Data"] if "Company Data" in wb.sheetnames else None; h=wb["Historical Financials"] if "Historical Financials" in wb.sheetnames else None; seg=wb["Segment Analysis"] if "Segment Analysis" in wb.sheetnames else None
    for c in range(10,14): ws.cell(1,c).fill=_fill(NAVY); ws.cell(2,c).fill=_fill(NAVY)
    ws["J1"]="Auto Context"; ws["J1"].font=Font(bold=True,color=WHITE,size=14)
    ws["J3"]="Automatically populated facts for orientation; keep your thesis and judgments in the yellow manual fields on the left."; ws["J3"].font=Font(italic=True,color=GREY); ws["J3"].alignment=Alignment(wrap_text=True)
    metrics=[
        ("Company",d["B5"].value if d else None,"General"),("Sector",d["B6"].value if d else None,"General"),("Industry",d["B7"].value if d else None,"General"),("Current Price",d["B8"].value if d else None,FMT_PRICE),("Forward P/E",d["B15"].value if d else None,'0.0x'),
        ("Latest Revenue",h["G4"].value if h else None,FMT_BN),("Operating Margin",h["G10"].value if h else None,FMT_PCT),("FCF Margin",h["G17"].value if h else None,FMT_PCT),("Segment Status",_segment_status(seg),"General")
    ]
    _header(ws,5,10,11); ws["J5"]="Metric"; ws["K5"]="Value"
    for r,(lab,val,fmt) in enumerate(metrics,6):
        ws.cell(r,10,lab); ws.cell(r,11,val); ws.cell(r,11).number_format=fmt; ws.cell(r,11).font=Font(color=LINK_GREEN)
    ws.column_dimensions["J"].width=24; ws.column_dimensions["K"].width=38; ws.column_dimensions["L"].width=4; ws.column_dimensions["M"].width=4

def repair_data_quality(wb,ticker):
    if "Data Quality" not in wb.sheetnames: return
    ws=wb["Data Quality"]; seg=wb["Segment Analysis"] if "Segment Analysis" in wb.sheetnames else None
    status=_segment_status(seg); seg_count,bus_count=_segment_counts(seg)
    if status.startswith("AUTO") and (seg_count>0 or bus_count>0): dq_status="PASS"
    elif status.startswith("MANUAL"): dq_status="MANUAL"
    else: dq_status="REVIEW"
    for r in range(6,ws.max_row+1):
        if str(ws.cell(r,1).value or "").strip()=="Segment analysis":
            ws.cell(r,2).value=dq_status
            ws.cell(r,3).value=f"{seg_count} segment(s), {bus_count} revenue group(s); {status}"
            ws.cell(r,4).value="PASS requires actual issuer-disclosed rows, not merely the existence of a Segment Analysis tab."
            ws.cell(r,2).fill=_fill(PALE_GREEN if dq_status=="PASS" else GOLD)
            break
    # Add/refresh a template-leak check.
    row=None
    for r in range(6,ws.max_row+1):
        if str(ws.cell(r,1).value or "").strip()=="Legacy template labels": row=r; break
    if row is None: row=ws.max_row+1
    comp_ok=("Comparative Analysis" not in wb.sheetnames or wb["Comparative Analysis"]["B3"].value==ticker)
    dash_ok=("Dashboard" not in wb.sheetnames or wb["Dashboard"]["B18"].value in (None,ticker))
    leak_status="PASS" if comp_ok and dash_ok else "FAIL"
    vals=["Legacy template labels",leak_status,f"Comparative={comp_ok}; Dashboard={dash_ok}","Detects stale template-company labels after switching tickers."]
    for c,v in enumerate(vals,1): ws.cell(row,c).value=v
    ws.cell(row,2).fill=_fill(PALE_GREEN if leak_status=="PASS" else PALE_RED)

def refresh_cross_company_tabs(wb,ticker):
    _refresh_labels(wb,ticker)
    _refresh_historical_sources(wb)
    ensure_business_portfolio_map(wb,ticker)
    seed_research_context(wb,ticker)
    repair_data_quality(wb,ticker)
