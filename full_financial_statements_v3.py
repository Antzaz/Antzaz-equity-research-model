from __future__ import annotations

"""Profile-aware full financial-statement renderer.

This is the institutional statement layer.  It keeps the existing source hierarchy, but selects
an economically appropriate line-item structure for technology companies, IFRS foreign issuers,
banks and insurance/operating conglomerates.  Reported values are never synthesized merely to
make a statement look complete.
"""

import math
from collections import defaultdict

from openpyxl.styles import Alignment, Border, Font
from openpyxl.utils import get_column_letter

import full_financial_statements as base
from issuer_statement_profiles import get_statement_profile


ANNUAL_FORMS = {"10-K", "10-K/A", "20-F", "20-F/A"}


def _fact_candidates(facts, tag_spec):
    """Yield (namespace, tag, fact) for plain or namespace-qualified XBRL tags."""
    if not facts:
        return
    root = facts.get("facts") or {}
    if ":" in str(tag_spec):
        ns, tag = str(tag_spec).split(":", 1)
        fact = (root.get(ns) or {}).get(tag)
        if fact:
            yield ns, tag, fact
        return
    for ns in ("us-gaap", "ifrs-full"):
        fact = (root.get(ns) or {}).get(str(tag_spec))
        if fact:
            yield ns, str(tag_spec), fact


def _pick_units(units, unit_kind):
    if not units:
        return []
    keys = list(units)
    if unit_kind == "eps":
        preferred = [k for k in keys if "/shares" in str(k).lower() or "pershare" in str(k).lower()]
    elif unit_kind == "shares":
        preferred = [k for k in keys if str(k).lower() in {"shares", "share"}]
    else:
        preferred = [k for k in keys if "/shares" not in str(k).lower() and str(k).lower() not in {"shares", "share", "pure"}]
        # Monetary units are normally ISO currency codes. Prefer them to dimensionless units.
        currency = [k for k in preferred if len(str(k)) == 3 and str(k).isalpha()]
        if currency:
            preferred = currency + [k for k in preferred if k not in currency]
    return preferred or keys


def _sec_series_any(facts, tags, unit_kind="money"):
    best = {}
    source_ns = {}
    for priority, tag_spec in enumerate(tags or []):
        for ns, _tag, fact in _fact_candidates(facts, tag_spec) or []:
            units = fact.get("units") or {}
            for unit_priority, unit_name in enumerate(_pick_units(units, unit_kind)):
                for item in units.get(unit_name) or []:
                    if item.get("form") not in ANNUAL_FORMS:
                        continue
                    start = item.get("start")
                    end = item.get("end")
                    # Duration facts should represent roughly a fiscal year. Instant facts have no start.
                    if start and end:
                        try:
                            import datetime as _dt
                            days = (_dt.date.fromisoformat(str(end)) - _dt.date.fromisoformat(str(start))).days
                            if days < 250 or days > 450:
                                continue
                        except Exception:
                            pass
                    year = item.get("fy")
                    if year in (None, "") and end:
                        year = str(end)[:4]
                    try:
                        year = int(year)
                        value = float(item.get("val"))
                    except Exception:
                        continue
                    if not math.isfinite(value) or not (1900 <= year <= 2100):
                        continue
                    stamp = (
                        str(item.get("filed") or ""),
                        str(end or ""),
                        -priority,
                        -unit_priority,
                    )
                    if year not in best or stamp > best[year][0]:
                        best[year] = (stamp, value)
                        source_ns[year] = ns
    return {y: v for y, (_stamp, v) in best.items()}, source_ns


def _profile_series(existing, frames, facts, key, label, yf_names, sec_tags, unit, aliases, years):
    old = {}
    for name in (label,) + tuple(aliases or ()):
        old.update((existing.get(key) or {}).get(name, {}) or {})
    sec, sec_ns = _sec_series_any(facts, sec_tags, unit)
    yf = base._yf_series(frames.get(key), yf_names)
    values = {}
    sources = {}
    scale = 1.0 if unit == "eps" else 1e9
    for year in years:
        if year in old:
            values[year] = old[year]
            sources[year] = "Canonical issuer / reconciled statement"
        elif year in sec:
            values[year] = sec[year] if unit == "eps" else sec[year] / scale
            ns = sec_ns.get(year)
            sources[year] = "SEC Company Facts — IFRS / 20-F" if ns == "ifrs-full" else "SEC Company Facts — US GAAP"
        elif year in yf:
            values[year] = yf[year] if unit == "eps" else yf[year] / scale
            sources[year] = "Yahoo annual statement fallback"
    if label in base.OUTFLOW_LABELS:
        values = {y: -abs(v) for y, v in values.items()}
    return values, sources


def _available_years(existing, frames, facts, profile):
    years = set()
    for section in ("income", "balance", "cash"):
        for vals in (existing.get(section) or {}).values():
            years.update(vals)
    for section in ("income", "balance", "cash"):
        df = frames.get(section)
        if df is None or getattr(df, "empty", True):
            continue
        for col in df.columns:
            try:
                years.add(int(getattr(col, "year", str(col)[:4])))
            except Exception:
                pass
    # Facts can add a completed fiscal year that Yahoo has not exposed yet.
    revenue_row = next((r for r in profile["income"] if r[0] == profile["canonical_revenue"]), None)
    if revenue_row:
        sec, _ = _sec_series_any(facts, revenue_row[2], revenue_row[3])
        years.update(sec)
    valid = sorted(y for y in years if 2000 <= int(y) <= 2100)
    return valid[-6:]


def _write_section(ws, start, title, rows, years, existing, frames, facts, key):
    source_col = 2 + len(years)
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=source_col)
    ws.cell(start, 1, title)
    ws.cell(start, 1).fill = base._fill(base.NAVY)
    ws.cell(start, 1).font = Font(bold=True, color=base.WHITE, size=12)
    header = start + 1
    ws.cell(header, 1, "Metric")
    for j, year in enumerate(years, 2):
        ws.cell(header, j, year)
    ws.cell(header, source_col, "Source / Definition")
    for c in range(1, source_col + 1):
        ws.cell(header, c).fill = base._fill(base.BLUE)
        ws.cell(header, c).font = Font(bold=True, color=base.WHITE)
        ws.cell(header, c).alignment = Alignment(horizontal="center", wrap_text=True)

    row = header + 1
    row_map = {}
    for label, yf_names, sec_tags, unit, aliases in rows:
        row_map[label] = row
        ws.cell(row, 1, label)
        values, sources = _profile_series(existing, frames, facts, key, label, yf_names, sec_tags, unit, aliases, years)
        for j, year in enumerate(years, 2):
            ws.cell(row, j, values.get(year))
            ws.cell(row, j).number_format = base.FMT_EPS if unit == "eps" else base.FMT_SHARES if unit == "shares" else base.FMT_BN
        lineage = []
        for year in years:
            src = sources.get(year)
            if src and src not in lineage:
                lineage.append(src)
        ws.cell(row, source_col, " → ".join(lineage) if lineage else "Not reliably mapped")
        ws.cell(row, source_col).alignment = Alignment(wrap_text=True, vertical="top")
        if label in base.TOTAL_LABELS or label in {
            "Total Net Revenue", "Total Noninterest Expense", "Net Earnings", "Net Earnings Attributable to Berkshire",
            "Total Revenues", "Berkshire Shareholders' Equity", "Total Stockholders' Equity",
        }:
            ws.cell(row, 1).font = Font(bold=True)
            for c in range(1, source_col + 1):
                ws.cell(row, c).border = Border(top=base.THIN)
        row += 1
    return row, row_map, source_col


def _formula_if_rows(ws, row, source_col, years, row_map, numerator, denominator, label, definition, fmt=base.FMT_PCT):
    if numerator not in row_map or denominator not in row_map:
        return row
    ws.cell(row, 1, label)
    ws.cell(row, source_col, definition)
    for j, _year in enumerate(years, 2):
        col = get_column_letter(j)
        ws.cell(row, j, f'=IFERROR({col}{row_map[numerator]}/{col}{row_map[denominator]},"")')
        ws.cell(row, j).number_format = fmt
    return row + 1


def _coverage(ws, row, source_col, years, maps):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=source_col)
    ws.cell(row, 1, "Full Statement Coverage")
    ws.cell(row, 1).fill = base._fill(base.NAVY)
    ws.cell(row, 1).font = Font(bold=True, color=base.WHITE)
    row += 1
    for c, value in enumerate(["Section", "Mapped rows", "Total profile rows", "Coverage"], 1):
        ws.cell(row, c, value)
        ws.cell(row, c).fill = base._fill(base.BLUE)
        ws.cell(row, c).font = Font(bold=True, color=base.WHITE)
    stats = {}
    for key, name, row_map in maps:
        mapped = sum(
            any(base._num(ws.cell(r, j).value) is not None for j in range(2, 2 + len(years)))
            for r in row_map.values()
        )
        stats[key] = (mapped, len(row_map))
        row += 1
        ws.cell(row, 1, name)
        ws.cell(row, 2, mapped)
        ws.cell(row, 3, len(row_map))
        ws.cell(row, 4, mapped / len(row_map) if row_map else 0)
        ws.cell(row, 4).number_format = base.FMT_PCT
    return stats


def expand_financial_statements(wb, ticker, facts=None):
    profile = get_statement_profile(ticker)
    existing = base._capture_existing(wb)
    frames = base._yf_frames(ticker)
    years = _available_years(existing, frames, facts, profile)
    if not years:
        return {
            "expanded": False, "profile": profile["key"], "profile_name": profile["name"],
            "income_rows": 0, "balance_rows": 0, "cash_rows": 0, "years": [],
        }

    idx = wb.sheetnames.index("Financial Statements") if "Financial Statements" in wb.sheetnames else len(wb.sheetnames)
    if "Financial Statements" in wb.sheetnames:
        wb.remove(wb["Financial Statements"])
    ws = wb.create_sheet("Financial Statements", idx)
    ws.sheet_view.showGridLines = False
    last_col = 2 + len(years)
    base._title(ws, ticker, last_col)

    # The profile and currency basis are visible to the analyst. Currency normalization may later
    # rewrite A3 after cross-border conversion.
    info = getattr(wb, "_wacc_info", {}) or {}
    reporting = str(info.get("financialCurrency") or "unknown").upper()
    quote = str(info.get("currency") or "unknown").upper()
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=last_col)
    ws.cell(4, 1, f"Statement profile: {profile['name']} | reporting currency: {reporting} | traded/valuation currency: {quote}")
    ws.cell(4, 1).font = Font(italic=True, color=base.GREY)

    row, imap, source_col = _write_section(ws, 5, "Income Statement", profile["income"], years, existing, frames, facts, "income")
    if profile["key"] not in {"bank", "berkshire"}:
        row = _formula_if_rows(ws, row, source_col, years, imap, "Operating Income", profile["canonical_revenue"], "Operating Margin", "Derived: operating income ÷ revenue")
        net_label = "Net Income Attributable to Parent" if "Net Income Attributable to Parent" in imap else "Net Income"
        row = _formula_if_rows(ws, row, source_col, years, imap, net_label, profile["canonical_revenue"], "Net Margin", "Derived: net income ÷ revenue")
    elif profile["key"] == "bank":
        row = _formula_if_rows(ws, row, source_col, years, imap, "Total Noninterest Expense", "Total Net Revenue", "Efficiency Ratio (simple)", "Derived screening metric: noninterest expense ÷ total net revenue")
        row = _formula_if_rows(ws, row, source_col, years, imap, "Provision for Credit Losses", "Total Net Revenue", "Credit Cost / Revenue", "Derived screening metric: provision for credit losses ÷ total net revenue")

    row += 2
    row, bmap, source_col = _write_section(ws, row, "Balance Sheet", profile["balance"], years, existing, frames, facts, "balance")
    if profile.get("balance_net_debt"):
        for label, a, b, definition in (
            ("Net Debt", "Long-Term Debt", "Cash & Cash Equivalents", "Derived balance-sheet metric"),
            ("Working Capital", "Total Current Assets", "Total Current Liabilities", "Derived balance-sheet metric"),
        ):
            if a not in bmap or b not in bmap:
                continue
            ws.cell(row, 1, label)
            ws.cell(row, source_col, definition)
            for j, _year in enumerate(years, 2):
                col = get_column_letter(j)
                ws.cell(row, j, f'=IFERROR({col}{bmap[a]}-{col}{bmap[b]},"")')
                ws.cell(row, j).number_format = base.FMT_BN
            row += 1
    elif profile["key"] == "bank" and "Loans, Net" in bmap and "Deposits" in bmap:
        row = _formula_if_rows(ws, row, source_col, years, bmap, "Loans, Net", "Deposits", "Loans / Deposits", "Derived bank balance-sheet metric")

    # Reported liabilities-and-equity values are preferred. Only corporate/IFRS profiles get an
    # accounting-identity fallback; for bank/conglomerate profiles we keep a blank if unavailable.
    tle = bmap.get("Total Liabilities & Equity")
    if tle and profile["key"] not in {"bank", "berkshire"} and "Total Liabilities" in bmap and "Total Equity" in bmap:
        for j, _year in enumerate(years, 2):
            if ws.cell(tle, j).value in (None, ""):
                col = get_column_letter(j)
                ws.cell(tle, j, f'=IFERROR({col}{bmap["Total Liabilities"]}+{col}{bmap["Total Equity"]},"")')
                ws.cell(tle, j).number_format = base.FMT_BN

    row += 2
    row, cmap, source_col = _write_section(ws, row, "Cash Flow Statement", profile["cash"], years, existing, frames, facts, "cash")

    # Beginning cash may be bridged from prior-year ending cash.
    if "Beginning Cash" in cmap and "Ending Cash" in cmap:
        for offset, j in enumerate(range(2, 2 + len(years))):
            if offset == 0 or ws.cell(cmap["Beginning Cash"], j).value not in (None, ""):
                continue
            prev = get_column_letter(j - 1)
            ws.cell(cmap["Beginning Cash"], j, f'={prev}{cmap["Ending Cash"]}')
        ws.cell(cmap["Beginning Cash"], source_col, "Reported where available; otherwise prior-year Ending Cash")

    if profile.get("derive_fcf") and "Operating Cash Flow" in cmap and "Capital Expenditures" in cmap:
        ws.cell(row, 1, "Free Cash Flow")
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, source_col, "Derived: Operating Cash Flow + Capital Expenditures")
        for j, _year in enumerate(years, 2):
            col = get_column_letter(j)
            ws.cell(row, j, f'=IFERROR({col}{cmap["Operating Cash Flow"]}+{col}{cmap["Capital Expenditures"]},"")')
            ws.cell(row, j).number_format = base.FMT_BN
        for c in range(1, source_col + 1):
            ws.cell(row, c).border = Border(top=base.THIN)
        row += 1
    else:
        ws.cell(row, 1, "Free Cash Flow")
        ws.cell(row, source_col, "Not used as a primary decision metric for this business model; banking/insurance balance-sheet flows make industrial FCF misleading.")
        row += 1

    row += 2
    stats = _coverage(
        ws, row, source_col, years,
        (("income", "Income Statement", imap), ("balance", "Balance Sheet", bmap), ("cash", "Cash Flow Statement", cmap)),
    )

    ws.column_dimensions["A"].width = 48
    for c in range(2, 2 + len(years)):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.column_dimensions[get_column_letter(source_col)].width = 52
    for r in range(1, ws.max_row + 1):
        ws.cell(r, source_col).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "B7"
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass
    return {
        "expanded": True,
        "profile": profile["key"], "profile_name": profile["name"],
        "canonical_revenue": profile["canonical_revenue"],
        "income_rows": stats["income"][0], "income_total": stats["income"][1],
        "balance_rows": stats["balance"][0], "balance_total": stats["balance"][1],
        "cash_rows": stats["cash"][0], "cash_total": stats["cash"][1],
        "years": years,
        "min_structure": profile["min_structure"], "min_mapped": profile["min_mapped"],
    }
