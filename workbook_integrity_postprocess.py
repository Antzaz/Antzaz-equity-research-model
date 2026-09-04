from __future__ import annotations

"""Final Excel compatibility and visible-data quality pass.

This runs after optional ML / AI sheets are written. The deterministic generator already repairs
its charts before save, but ML/AI charts are added later in separate processes. Excel defaults to
plotting visible cells only, so charts sourced from hidden helper columns can appear blank even
when Python renderers show them correctly. This postprocessor explicitly allows hidden source
cells for every chart and replaces unexplained visible blanks in decision-facing ML/AI tables
with N/M text.
"""

import argparse
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

GREEN='E2F0D9'; GOLD='FFF2CC'; RED='FCE4D6'


def _find(ws, label: str, start: int = 1, end: int | None = None) -> int | None:
    needle = str(label).strip().lower(); end = end or ws.max_row
    for r in range(start, min(end, ws.max_row) + 1):
        if str(ws.cell(r, 1).value or '').strip().lower() == needle:
            return r
    return None


def _chart_series_count(chart) -> int:
    try:
        return len(chart.ser)
    except Exception:
        return 0


def _repair_charts(wb) -> dict[str, Any]:
    charts = 0; empty_series = 0; sheets = []
    for ws in wb.worksheets:
        local = 0
        for chart in getattr(ws, '_charts', []) or []:
            charts += 1; local += 1
            try:
                chart.visible_cells_only = False
            except Exception:
                pass
            try:
                chart.display_blanks = 'gap'
            except Exception:
                pass
            if _chart_series_count(chart) == 0:
                empty_series += 1
        if local:
            sheets.append(f'{ws.title}:{local}')
    return {'charts': charts, 'empty_series': empty_series, 'chart_sheets': sheets}


def _fill_ml_visible_states(wb) -> int:
    if 'ML & Quantitative Research' not in wb.sheetnames:
        return 0
    ws = wb['ML & Quantitative Research']; changed = 0
    # Main six-model dashboard is deliberately decision-facing; blank predictions or drivers are
    # ambiguous to non-technical readers. Detailed sections below may retain normal blank spacing.
    for r in range(7, min(ws.max_row, 12) + 1):
        status = str(ws.cell(r, 2).value or '')
        if ws.cell(r, 3).value in (None, ''):
            ws.cell(r, 3).value = 'N/M — insufficient data' if status == 'INSUFFICIENT_DATA' else 'N/M — no reliable numeric signal'
            changed += 1
        if ws.cell(r, 5).value in (None, ''):
            ws.cell(r, 5).value = 'See model detail below'; changed += 1
        if ws.cell(r, 6).value in (None, ''):
            ws.cell(r, 6).value = 'N/M — no reliable ranked drivers'; changed += 1
        for c in range(1, 10):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical='top')
    return changed


def _fill_ai_visible_states(wb) -> int:
    if 'AI Growth Forecast' not in wb.sheetnames:
        return 0
    ws = wb['AI Growth Forecast']; changed = 0
    # Revenue is not directly reverse-DCF-implied in this model; make that explicit instead of
    # leaving a visually unexplained blank.
    if ws['E16'].value in (None, ''):
        ws['E16'] = 'N/M — reverse DCF hurdle is FCF-based'; changed += 1
    # Fill decision cells only; evidence-table spacing remains intentionally blank.
    for row in (16, 17):
        for col in (2, 3, 4, 5, 6):
            cell = ws.cell(row, col)
            if cell.value in (None, ''):
                cell.value = 'N/M'; changed += 1
    return changed


def _decorate_quality(wb, result: dict[str, Any]) -> None:
    if 'Data Quality' not in wb.sheetnames:
        return
    ws = wb['Data Quality']; label = 'Excel chart / visible-state integrity'
    row = next((r for r in range(1, ws.max_row + 1) if str(ws.cell(r, 1).value or '').strip() == label), ws.max_row + 1)
    status = 'PASS' if int(result.get('empty_series') or 0) == 0 else 'REVIEW'
    observed = (
        f"{result.get('charts', 0)} chart(s) forced to include hidden helper cells; "
        f"{result.get('empty_series', 0)} chart(s) have zero series; "
        f"{result.get('visible_cells_filled', 0)} decision-facing blank cell(s) made explicit."
    )
    why = 'Prevents charts that render in Python but appear blank in desktop Excel, and prevents N/M states from looking like missing pipeline data.'
    for c, value in enumerate((label, status, observed, why), 1):
        ws.cell(row, c).value = value; ws.cell(row, c).alignment = Alignment(wrap_text=True, vertical='top')
    ws.cell(row, 2).fill = PatternFill('solid', fgColor=GREEN if status == 'PASS' else GOLD)
    ws.cell(row, 2).font = Font(bold=True)


def finalize_workbook_integrity(path: str | Path) -> dict[str, Any]:
    path = Path(path)
    wb = load_workbook(path)
    result = _repair_charts(wb)
    result['visible_cells_filled'] = _fill_ml_visible_states(wb) + _fill_ai_visible_states(wb)
    _decorate_quality(wb, result)
    try:
        wb.calculation.calcMode = 'auto'; wb.calculation.fullCalcOnLoad = True; wb.calculation.forceFullCalc = True
    except Exception:
        pass
    wb.save(path)
    return result


def main() -> int:
    p = argparse.ArgumentParser(description='Repair final workbook chart compatibility and explicit N/M states')
    p.add_argument('workbook')
    args = p.parse_args()
    result = finalize_workbook_integrity(args.workbook)
    print(
        f"[workbook-integrity] charts={result['charts']}; zero-series={result['empty_series']}; "
        f"visible blanks clarified={result['visible_cells_filled']}"
    )
    return 0 if result['empty_series'] == 0 else 2


if __name__ == '__main__':
    raise SystemExit(main())
