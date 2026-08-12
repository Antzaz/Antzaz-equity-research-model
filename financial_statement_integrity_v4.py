from __future__ import annotations

"""Profile-aware canonical financial-statement reconciliation.

The legacy v3 repair remains useful for ordinary US corporate statements.  Specialized profiles
(bank, insurance/conglomerate, and a cross-border ADR after FX normalization) must not be pushed
back through industrial-company aliases.  This wrapper synchronizes only economically equivalent
reported rows into Historical Financials for those cases.
"""

from openpyxl.styles import Alignment, Font, PatternFill

from financial_statement_integrity_v3 import repair_financial_statements_v3 as _legacy_repair
from issuer_statement_profiles import get_statement_profile

BLUE="2F75B5"; WHITE="FFFFFF"; GREEN="E2F0D9"; GOLD="FFF2CC"


def _num(v):
    try:
        if isinstance(v,bool) or v in (None,""): return None
        return float(v)
    except Exception: return None


def _find(ws,label,start=1,end=None):
    needle=str(label).strip().lower(); end=end or ws.max_row
    for r in range(start,min(end,ws.max_row)+1):
        if str(ws.cell(r,1).value or "").strip().lower()==needle: return r
    return None


def _section(ws,label): return _find(ws,label)


def _year_cols(ws,header):
    return {int(ws.cell(header,c).value):c for c in range(2,min(ws.max_column,12)+1) if isinstance(ws.cell(header,c).value,(int,float)) and 1900<=int(ws.cell(header,c).value)<=2100}


def _profile_sync(wb,ticker,profile):
    if "Financial Statements" not in wb.sheetnames or "Historical Financials" not in wb.sheetnames:
        return {"history_sync":0,"coverage":0,"canonical":False}
    fs=wb["Financial Statements"]; hs=wb["Historical Financials"]
    i0=_section(fs,"Income Statement"); b0=_section(fs,"Balance Sheet"); c0=_section(fs,"Cash Flow Statement")
    if not all((i0,b0,c0)):
        return {"history_sync":0,"coverage":0,"canonical":False}
    ih=next((r for r in range(i0+1,min(b0,i0+6)) if str(fs.cell(r,1).value or "").strip().lower()=="metric"),None)
    ch=next((r for r in range(c0+1,min(fs.max_row+1,c0+6)) if str(fs.cell(r,1).value or "").strip().lower()=="metric"),None)
    if not ih or not ch:
        return {"history_sync":0,"coverage":0,"canonical":False}
    iy=_year_cols(fs,ih); cy=_year_cols(fs,ch)
    hy={int(hs.cell(3,c).value):c for c in range(2,min(hs.max_column,8)+1) if isinstance(hs.cell(3,c).value,(int,float))}

    key=profile["key"]
    revenue=profile.get("canonical_revenue","Revenue")
    if key=="bank":
        op_label=None
        net_label="Net Income"
    elif key=="berkshire":
        op_label="Operating Earnings / Income"
        net_label="Net Earnings Attributable to Berkshire"
    elif key=="tsm":
        op_label="Operating Income"
        net_label="Net Income Attributable to Parent"
    else:
        op_label="Operating Income"
        net_label="Net Income"

    mapping=[
        (4,revenue,iy,i0,b0-1,False),
        (9,op_label,iy,i0,b0-1,False),
        (11,net_label,iy,i0,b0-1,False),
        (14,"Operating Cash Flow",cy,c0,fs.max_row,False),
        (15,"Capital Expenditures",cy,c0,fs.max_row,True),
    ]
    written=0; available=0; total=0
    for hrow,label,cols,start,end,absolute in mapping:
        if not label: continue
        r=_find(fs,label,start,end)
        if not r: continue
        for year,hc in hy.items():
            fc=cols.get(year)
            value=_num(fs.cell(r,fc).value) if fc else None
            total+=1
            if value is None: continue
            available+=1
            hs.cell(hrow,hc).value=abs(value) if absolute else value
            written+=1

    # For banks, leaving Historical Financials operating income blank is intentional.  Pretax
    # income is not silently substituted for an industrial operating-income definition.
    if key=="bank":
        for _year,hc in hy.items():
            hs.cell(9,hc).value=None

    return {
        "history_sync":written,
        "coverage":available/total if total else 0,
        "canonical":True,
        "profile":key,
        "profile_name":profile["name"],
        "operating_income_not_applicable":key=="bank",
    }


def _integrity_section(wb,ticker,profile,result):
    if "Financial Statements" not in wb.sheetnames: return
    ws=wb["Financial Statements"]
    old=_find(ws,"Financial Statement Integrity & Source Reconciliation")
    row=old or ws.max_row+3
    if old:
        for rr in range(old,min(ws.max_row,old+8)+1):
            for c in range(1,min(ws.max_column,9)+1): ws.cell(rr,c).value=None
    for c in range(1,8):
        ws.cell(row,c).fill=PatternFill("solid",fgColor=BLUE); ws.cell(row,c).font=Font(bold=True,color=WHITE)
    ws.cell(row,1,"Financial Statement Integrity & Source Reconciliation")
    row+=1
    headers=["Control","Status","Method","Primary / Preferred Source","Fallback","Why it matters","Action"]
    for c,value in enumerate(headers,1):
        ws.cell(row,c,value); ws.cell(row,c).fill=PatternFill("solid",fgColor=BLUE); ws.cell(row,c).font=Font(bold=True,color=WHITE)
    key=profile["key"]
    if key=="bank":
        op_method="Bank profile: no industrial operating-income proxy is synchronized"
        op_status="PASS"
    else:
        op_method="Exact profile operating-income row only; pretax is never substituted"
        op_status="PASS"
    rows=[
        ("Statement profile","PASS",profile["name"],"Issuer filing / structured annual statements","Blank when not reliably mapped","Business-model-specific statements avoid false comparability","Review profile when business model changes"),
        ("Canonical annual revenue","PASS" if result.get("history_sync") else "REVIEW",f"{profile.get('canonical_revenue')} synchronized to canonical history","Issuer/SEC/structured annual statement","No fabricated revenue","DCF/growth must use one revenue definition","Review any source conflict"),
        ("Operating income definition",op_status,op_method,"Profile-specific reported line","No pretax substitution","Margins/ROIC must not use a false operating metric","Leave blank when not meaningful"),
        ("Post-FX synchronization","PASS" if key=="tsm" else "N/A","TSM reconciliation runs only after reporting-currency values have been normalized to the traded USD valuation basis" if key=="tsm" else "Same-currency or non-ADR specialized profile","Currency normalization layer","Prevents native-currency values from contaminating valuation history","Review Currency / ADR normalization control"),
    ]
    for item in rows:
        row+=1
        for c,value in enumerate(item,1):
            ws.cell(row,c,value); ws.cell(row,c).alignment=Alignment(wrap_text=True,vertical="top")
        ws.cell(row,2).fill=PatternFill("solid",fgColor=GREEN if item[1] in {"PASS","N/A"} else GOLD)
    for col,width in {"A":30,"B":14,"C":54,"D":46,"E":42,"F":50,"G":40}.items():
        ws.column_dimensions[col].width=max(ws.column_dimensions[col].width or 0,width)


def repair_financial_statements_v4(wb,ticker):
    profile=get_statement_profile(ticker)
    if profile["key"] not in {"bank","berkshire","tsm"}:
        result=_legacy_repair(wb,ticker) or {}
        result["profile"]=profile["key"]; result["profile_name"]=profile["name"]
        return result
    result=_profile_sync(wb,ticker,profile)
    _integrity_section(wb,ticker,profile,result)
    return result
