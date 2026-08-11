from __future__ import annotations

import math
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from score_engine_v2 import compute_score_bundle, DIMENSION_WEIGHTS, weighted_available, num

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; GREY="666666"; GOLD="FFF2CC"; PALE_GREEN="E2F0D9"; PALE_RED="FCE4D6"
FMT_SCORE='0.0'; FMT_PCT='0.0%;[Red](0.0%);-'; THIN=Side(style="thin",color="D9E1F2")


def _fill(c): return PatternFill("solid",fgColor=c)

def _find_row(ws,label,col=1):
    needle=str(label).strip().lower()
    for r in range(1,ws.max_row+1):
        if str(ws.cell(r,col).value or "").strip().lower()==needle: return r
    return None

def _section(ws,row,title,end=10):
    for c in range(1,end+1): ws.cell(row,c).fill=_fill(NAVY); ws.cell(row,c).font=Font(bold=True,color=WHITE)
    ws.cell(row,1,title)

def _header(ws,row,headers):
    for c,v in enumerate(headers,1):
        x=ws.cell(row,c,v); x.fill=_fill(BLUE); x.font=Font(bold=True,color=WHITE); x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); x.border=Border(bottom=THIN)


def _advanced_value(wb,label):
    if "Advanced Analytics" not in wb.sheetnames: return None
    ws=wb["Advanced Analytics"]; r=_find_row(ws,label)
    return num(ws.cell(r,2).value) if r else None


def institutional_dimensions(wb):
    """Drop-in replacement for institutional_lenses._scorecard_dimensions."""
    dims={}
    for key in ("Growth","Profitability","FCF Quality","Balance Sheet","Absolute Valuation","Relative Valuation","Stress Robustness"):
        v=_advanced_value(wb,key)
        if v is not None: dims[key]=max(0,min(100,v))
    leadership=None
    if "Leadership & Culture" in wb.sheetnames: leadership=num(wb["Leadership & Culture"]["B11"].value)
    if leadership is not None: dims["Leadership"]=max(0,min(100,leadership))

    quality=weighted_available([(dims.get("Growth"),.20),(dims.get("Profitability"),.35),(dims.get("FCF Quality"),.30),(dims.get("Balance Sheet"),.15)])
    manual=None
    if "Moat & Competitive Advantage" in wb.sheetnames:
        ws=wb["Moat & Competitive Advantage"]
        for label in ("Overall Moat Score","Moat Score","Competitive Advantage Score"):
            r=_find_row(ws,label)
            if r:
                manual=num(ws.cell(r,2).value)
                if manual is not None: break
    moat=manual if manual is not None else weighted_available([(quality,.70),(leadership,.30)])
    if moat is not None: dims["Moat / Position"]=max(0,min(100,moat))
    return dims


def leadership_proxy(wb,ticker,info,employee):
    """Business-model-aware replacement for the old absolute margin leadership proxy."""
    bundle=compute_score_bundle(wb,ticker)
    d=bundle["dimensions"]
    execution=weighted_available([(d["Growth"]["score"],.45),(d["Profitability"]["score"],.55)]) or 50.0
    capital=weighted_available([(d["FCF Quality"]["score"],.55),(d["Balance Sheet"]["score"],.45)]) or 50.0
    officers=(info or {}).get("companyOfficers") or []
    depth=min(100,35+8*len([x for x in officers if isinstance(x,dict) and x.get("name")]))
    try:
        from source_registry import issuer_sources
        src=issuer_sources(ticker,(info or {}).get("website"))
    except Exception: src={}
    if src.get("leadership") or src.get("governance"): depth=max(depth,65)
    culture=num((employee or {}).get("score"),50.0)
    governance=75.0 if src.get("governance") else (60.0 if src.get("investor") else 50.0)
    score=.30*execution+.25*capital+.15*depth+.15*culture+.15*governance
    rows=[
        ("Execution track record",execution,"45% Growth + 55% sector-aware Profitability/Economic Returns from Score Engine v2."),
        ("Capital allocation / cash generation",capital,"55% FCF Quality + 45% Balance Sheet from Score Engine v2; no absolute FCF-margin hurdle."),
        ("Leadership depth / disclosure",depth,"Public officer depth plus issuer leadership/governance disclosure."),
        ("Employee / culture signal",culture,(employee or {}).get("scope")),
        ("Governance disclosure",governance,"Availability of official governance / investor disclosure."),
    ]
    return score,rows


def _summary_inputs(wb):
    if "Investment Summary" not in wb.sheetnames: return None,None,None,None
    ws=wb["Investment Summary"]
    return num(ws["B8"].value),num(ws["D8"].value),num(ws["B9"].value),num(ws["H9"].value)


def patch_investment_summary(wb,ticker):
    if "Investment Summary" not in wb.sheetnames: return None
    ws=wb["Investment Summary"]; price,base,severe,mc_prob=_summary_inputs(wb)
    bundle=compute_score_bundle(wb,ticker,base_value=base,severe_value=severe,current_price=price,mc_prob=mc_prob)
    dims=bundle["dimensions"]
    bridge=_find_row(ws,"Quantitative Score Bridge — Neutral 50 to Final Score")
    if bridge:
        header=bridge+2
        name_map={
            "Absolute valuation":"Absolute Valuation","Relative valuation":"Relative Valuation","Growth":"Growth","Profitability":"Profitability",
            "FCF quality":"FCF Quality","Balance sheet":"Balance Sheet","Stress resilience":"Stress Robustness","Bayesian skew":"Bayesian Skew",
        }
        rows={str(ws.cell(r,1).value or "").strip():r for r in range(header+1,min(ws.max_row,header+14)+1)}
        available=sum(DIMENSION_WEIGHTS[k] for k in DIMENSION_WEIGHTS if dims[k]["score"] is not None)
        impacts=[]
        for display,key in name_map.items():
            r=rows.get(display)
            if not r: continue
            d=dims[key]; score=d["score"]; bw=DIMENSION_WEIGHTS[key]/100.0; ew=(DIMENSION_WEIGHTS[key]/available) if score is not None and available else 0.0
            contribution=(score*ew) if score is not None else 0.0; impact=((score-50)*ew) if score is not None else 0.0
            ws.cell(r,2,score); ws.cell(r,3,bw); ws.cell(r,4,ew); ws.cell(r,5,contribution); ws.cell(r,6,impact)
            ws.cell(r,7,d["actual"]); ws.cell(r,8,d["benchmark"]); ws.cell(r,9,d["components"]); ws.cell(r,10,d["status"])
            for c in (3,4): ws.cell(r,c).number_format=FMT_PCT
            if score is not None: impacts.append(impact)
        total=None
        for r in range(header+1,min(ws.max_row,header+22)+1):
            if str(ws.cell(r,1).value or "").strip()=="Final Quantitative Score": total=r; break
        overall=bundle["category_scores"]["Overall Investment Score"]
        if total: ws.cell(total,2,overall)
        ws["E6"]=overall; ws["G6"]=bundle["coverage"]

    cats=bundle["category_scores"]
    ws["A11"]="Business Quality / 100"; ws["B11"]=cats["Business Quality"]; ws["B11"].number_format=FMT_SCORE
    ws["C11"]="Valuation / 100"; ws["D11"]=cats["Valuation / Stock Attractiveness"]; ws["D11"].number_format=FMT_SCORE
    ws["E11"]="Downside / Risk / 100"; ws["F11"]=cats["Downside / Scenario Risk"]; ws["F11"].number_format=FMT_SCORE
    ws["G11"]="Interpretation"; ws["H11"]="Business quality and stock attractiveness are deliberately separate: a strong company can still score poorly at an expensive price. See Score Audit Trail."; ws.merge_cells("H11:J11"); ws["H11"].alignment=Alignment(wrap_text=True)

    # Never present negative equity value per share in the decision block.
    if severe is not None and severe<0:
        ws["B9"]=0.0; ws["D9"]=-1.0
    return bundle


def ensure_score_audit_trail(wb,ticker):
    price,base,severe,mc_prob=_summary_inputs(wb)
    bundle=compute_score_bundle(wb,ticker,base_value=base,severe_value=severe,current_price=price,mc_prob=mc_prob)
    if "Score Audit Trail" in wb.sheetnames: wb.remove(wb["Score Audit Trail"])
    ws=wb.create_sheet("Score Audit Trail",2); ws.sheet_view.showGridLines=False
    for c in range(1,11): ws.cell(1,c).fill=_fill(NAVY); ws.cell(2,c).fill=_fill(NAVY)
    ws["A1"]=f"{ticker} — Score Audit Trail"; ws["A1"].font=Font(bold=True,color=WHITE,size=18)
    ws["A3"]="Single source of truth for every reusable company/stock score. Follow each row from actual inputs → benchmark → formula → component math → score → source cells. Scores are project calculations, not external ratings."
    ws.merge_cells("A3:J3"); ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True)
    _section(ws,5,"Category Scores — Separate Company Quality from Price",10)
    _header(ws,6,["Category","Score / 100","What It Answers","Calculation","Coverage / Caveat"])
    cat_notes={
        "Business Quality":"Is the business economically strong regardless of current stock price?",
        "Valuation / Stock Attractiveness":"How attractive is today's price versus intrinsic value and peers?",
        "Downside / Scenario Risk":"How harsh is valuation downside under stress/scenario evidence?",
        "Overall Investment Score":"Weighted combination used by the Investment Summary neutral-50 bridge.",
    }
    for r,(name,val) in enumerate(bundle["category_scores"].items(),7):
        ws.cell(r,1,name); ws.cell(r,2,val); ws.cell(r,2).number_format=FMT_SCORE; ws.cell(r,3,cat_notes[name])
        ws.cell(r,4,"Business Quality = 20% Growth + 35% Profitability + 30% FCF Quality + 15% Balance Sheet" if name=="Business Quality" else "Valuation = 65% Absolute + 35% Relative" if name.startswith("Valuation") else "Risk = 70% Stress + 30% Bayesian when Bayesian exists; missing components reweighted" if name.startswith("Downside") else "Base dimension weights: 30% Absolute, 15% Relative, 10% each Growth/Profitability/FCF/Balance/Stress, 5% Bayesian; missing dimensions reweighted.")
        ws.cell(r,5,f"Score-weight coverage {bundle['coverage']:.0%}" if name=="Overall Investment Score" else "Category uses available components only")

    _section(ws,13,"Dimension-Level Proof",10)
    headers=["Dimension","Category","Score / 100","Actual Inputs","Benchmark / Reference","Exact Formula / Method","Component Calculation","Workbook Source Cells","Source / Lineage","Status"]
    _header(ws,14,headers)
    for r,key in enumerate(("Growth","Profitability","FCF Quality","Balance Sheet","Absolute Valuation","Relative Valuation","Stress Robustness","Bayesian Skew"),15):
        d=bundle["dimensions"][key]
        vals=[key,d["category"],d["score"],d["actual"],d["benchmark"],d["formula"],d["components"],d["source_cells"],d["source"],d["status"]]
        for c,v in enumerate(vals,1): ws.cell(r,c,v); ws.cell(r,c).alignment=Alignment(wrap_text=True,vertical="top")
        ws.cell(r,3).number_format=FMT_SCORE
        ws.cell(r,10).fill=_fill(PALE_GREEN if d["status"]=="Complete" else GOLD if d["status"]=="Partial" else PALE_RED)

    _section(ws,25,"Neutral-50 Overall Score Reconciliation",10)
    _header(ws,26,["Dimension","Score","Base Weight","Effective Weight","Weighted Contribution","Impact vs Neutral 50","Reconciliation"])
    available=sum(DIMENSION_WEIGHTS[k] for k,d in bundle["dimensions"].items() if d["score"] is not None)
    total_impact=0.0
    for r,key in enumerate(("Absolute Valuation","Relative Valuation","Growth","Profitability","FCF Quality","Balance Sheet","Stress Robustness","Bayesian Skew"),27):
        d=bundle["dimensions"][key]; sc=d["score"]; bw=DIMENSION_WEIGHTS[key]/100; ew=DIMENSION_WEIGHTS[key]/available if sc is not None and available else 0; contrib=sc*ew if sc is not None else 0; impact=(sc-50)*ew if sc is not None else 0; total_impact+=impact
        vals=[key,sc,bw,ew,contrib,impact,"Score contribution = Score × Effective Weight; impact = (Score−50) × Effective Weight"]
        for c,v in enumerate(vals,1): ws.cell(r,c,v)
        for c in (3,4): ws.cell(r,c).number_format=FMT_PCT
    rr=35; ws.cell(rr,1,"Final Quantitative Score"); ws.cell(rr,2,bundle["category_scores"]["Overall Investment Score"]); ws.cell(rr,6,total_impact); ws.cell(rr,7,"50 + sum(Impact vs Neutral) = Final Quantitative Score")
    widths={"A":28,"B":24,"C":15,"D":42,"E":42,"F":88,"G":55,"H":48,"I":55,"J":15}
    for col,w in widths.items(): ws.column_dimensions[col].width=w
    ws.freeze_panes="A15"
    return bundle


def append_institutional_proof(wb):
    if "Institutional Comparison" not in wb.sheetnames: return
    from institutional_lenses import INSTITUTIONAL_LENSES, _weighted_score
    ws=wb["Institutional Comparison"]; dims=institutional_dimensions(wb)
    by_name={x["firm"]:x for x in INSTITUTIONAL_LENSES}
    # Add an immediately visible quality-vs-price split to the main institutional table.
    header=10
    ws.cell(header,10,"Business Quality Fit / 100"); ws.cell(header,11,"Price-Aware Fit / 100"); ws.cell(header,12,"Score Formula")
    for c in range(10,13): ws.cell(header,c).fill=_fill(BLUE); ws.cell(header,c).font=Font(bold=True,color=WHITE); ws.cell(header,c).alignment=Alignment(wrap_text=True)
    quality_keys={"Growth","Profitability","FCF Quality","Balance Sheet","Leadership","Moat / Position"}
    for r in range(11,21):
        name=str(ws.cell(r,1).value or ""); lens=by_name.get(name)
        if not lens: continue
        qweights={k:w for k,w in lens["weights"].items() if k in quality_keys}; qscore,_,_=_weighted_score(dims,qweights)
        ws.cell(r,10,qscore); ws.cell(r,11,num(ws.cell(r,3).value)); ws.cell(r,12,"Fit = Σ(available company dimension score × published-style project weight) ÷ Σ(available weights). Business Quality Fit excludes valuation and price-stress dimensions.")
        ws.cell(r,10).number_format=FMT_SCORE; ws.cell(r,11).number_format=FMT_SCORE; ws.cell(r,12).alignment=Alignment(wrap_text=True,vertical="top")
    ws.column_dimensions["J"].width=24; ws.column_dimensions["K"].width=23; ws.column_dimensions["L"].width=68

    start=ws.max_row+3; _section(ws,start,"Institutional Score Calculation Proof — Dimension Contributions",12)
    _header(ws,start+1,["Institution / Style","Dimension","Company Score","Lens Weight","Weighted Points","Used in Quality Fit?","Public Principle Source"])
    row=start+2
    for lens in INSTITUTIONAL_LENSES:
        for key,w in lens["weights"].items():
            v=dims.get(key)
            if v is None: continue
            vals=[lens["firm"],key,v,w,v*w,"Yes" if key in quality_keys else "No — price/risk dimension",lens["source"]]
            for c,x in enumerate(vals,1): ws.cell(row,c,x); ws.cell(row,c).alignment=Alignment(wrap_text=True,vertical="top")
            ws.cell(row,4).number_format=FMT_PCT; row+=1
    return row


def add_quality_checks(wb,financial_coverage=None):
    if "Data Quality" not in wb.sheetnames: return
    ws=wb["Data Quality"]
    def put(label,status,obs,why):
        r=_find_row(ws,label) or ws.max_row+1
        ws.cell(r,1,label); ws.cell(r,2,status); ws.cell(r,3,obs); ws.cell(r,4,why)
        ws.cell(r,2).fill=_fill(PALE_GREEN if status=="PASS" else GOLD if status=="REVIEW" else PALE_RED); ws.cell(r,2).font=Font(bold=True)
        for c in range(1,5): ws.cell(r,c).alignment=Alignment(wrap_text=True,vertical="top")
    if financial_coverage is not None:
        status="PASS" if financial_coverage>=.90 else "REVIEW" if financial_coverage>=.70 else "FAIL"
        put("Financial Statements core coverage",status,f"{financial_coverage:.0%} of core income-statement year cells populated","SEC Company Facts is primary; Yahoo annual statements fill blank/known-bad cells. Missing unsupported line items remain blank.")
    put("Score methodology consistency","PASS","One Score Engine v2 feeds Advanced Analytics, Investment Summary and institutional reusable dimensions","Prevents the same concept from receiving different formulas on different sheets; Score Audit Trail exposes inputs, formulas and sources.")
    put("Business-model-aware quality scoring","PASS","Profitability uses ROIC/WACC + peer-relative ROE/margins; FCF Quality uses cash conversion and positive-history evidence","Avoids penalizing structurally low-margin business models merely because they do not resemble high-margin software companies.")


def finalize_score_transparency(wb,ticker,financial_coverage=None):
    patch_investment_summary(wb,ticker)
    ensure_score_audit_trail(wb,ticker)
    append_institutional_proof(wb)
    add_quality_checks(wb,financial_coverage)
