from __future__ import annotations

"""Issuer-verified full-statement adapters for cases where generic structured providers can
select a narrower accounting definition or GitHub runners cannot reach SEC Company Facts.

Values are historical audited/public annual figures only. They never create forecasts.
"""

from runtime_data_guards import install_runtime_data_guards
install_runtime_data_guards()

from openpyxl.styles import Font

FMT_BN='#,##0.0;[Red](#,##0.0);-'; FMT_EPS='$0.00;[Red]($0.00);-'; FMT_SHARES='#,##0.0;[Red](#,##0.0);-'
CEG_SOURCE="Constellation investor annual financial tables / 2025 Form 10-K — verified annual values"

CEG_INCOME={
    "Revenue":{2022:24.440,2023:24.918,2024:23.568,2025:25.533},
    "Cost of Revenue":{2022:17.462,2023:16.001,2024:11.419,2025:14.681},
    "Gross Profit":{2022:6.978,2023:8.917,2024:12.149,2025:10.852},
    "Selling, General & Administrative":{2022:5.393,2023:6.238,2024:6.745,2025:6.781},
    "Total Operating Expenses":{2022:23.945,2023:23.308,2024:19.216,2025:22.447},
    "Operating Income":{2022:.495,2023:1.610,2024:4.352,2025:3.086},
    "Pre-Tax Income":{2022:-.542,2023:2.447,2024:4.516,2025:3.511},
    "Income Taxes":{2022:-.388,2023:.859,2024:.774,2025:1.187},
    "Income After Tax":{2022:-.154,2023:1.588,2024:3.742,2025:2.324},
    "Minority / Noncontrolling Interests":{2022:.007,2023:.046,2024:.011,2025:-.004},
    "Equity in Affiliates":{2022:-.013,2023:-.011,2024:-.004,2025:-.001},
    "Net Income":{2022:-.160,2023:1.623,2024:3.749,2025:2.319},
    "Net Income Available to Common":{2022:-.160,2023:1.623,2024:3.749,2025:2.319},
    "Basic EPS":{2022:-.49,2023:5.02,2024:11.90,2025:7.41},
    "Diluted EPS":{2022:-.49,2023:5.01,2024:11.90,2025:7.39},
    "Basic Weighted Average Shares (bn)":{2022:.328,2023:.323,2024:.315,2025:.313},
    "Diluted Weighted Average Shares (bn)":{2022:.328,2023:.324,2024:.315,2025:.314},
}

CEG_BALANCE={
    "Cash & Cash Equivalents":{2022:.422,2023:.368,2024:3.022,2025:3.641},
    "Accounts Receivable":{2022:3.316,2023:2.851,2024:3.718,2025:4.266},
    "Inventory":{2022:1.505,2023:1.500,2024:1.600,2025:1.736},
    "Total Current Assets":{2022:9.360,2023:8.299,2024:10.776,2025:12.119},
    "Gross Property, Plant & Equipment":{2022:37.093,2023:40.033,2024:39.759,2025:41.917},
    "Accumulated Depreciation":{2022:-16.726,2023:-17.423,2024:-18.088,2025:-19.072},
    "Property & Equipment, Net":{2022:20.367,2023:22.610,2024:21.671,2025:22.845},
    "Goodwill":{2022:0.0,2023:.425,2024:.420,2025:.420},
    "Total Assets":{2022:46.909,2023:50.758,2024:52.926,2025:57.249},
    "Total Current Liabilities":{2022:7.839,2023:6.319,2024:6.846,2025:7.944},
    "Long-Term Debt":{2022:4.466,2023:7.496,2024:7.384,2025:7.250},
    "Total Liabilities":{2022:35.891,2023:39.833,2024:39.760,2025:42.732},
    "Stockholders' Equity":{2022:11.018,2023:10.925,2024:13.166,2025:14.517},
    "Total Equity":{2022:11.018,2023:10.925,2024:13.166,2025:14.517},
    "Total Liabilities & Equity":{2022:46.909,2023:50.758,2024:52.926,2025:57.249},
}

CEG_CASH={
    "Net Income":{2022:-.160,2023:1.623,2024:3.749,2025:2.319},
    "Depreciation, Amortization & Accretion":{2022:2.430,2023:2.514,2024:2.700,2025:2.601},
    "Deferred Income Taxes":{2022:-.643,2023:.251,2024:.222,2025:.273},
    "Other Non-Cash Items":{2022:.249,2023:.018,2024:-.172,2025:-.223},
    "Change in Working Capital":{2022:-6.011,2023:-9.874,2024:-7.333,2025:-.953},
    "Operating Cash Flow":{2022:-2.353,2023:-5.301,2024:-2.464,2025:4.237},
    "Capital Expenditures":{2022:-1.689,2023:-2.422,2024:-2.565,2025:-2.949},
    "Purchases of Investments":{2022:-4.271,2023:-6.050,2024:-6.282,2025:-7.284},
    "Proceeds from Investment Sales / Maturities":{2022:9.010,2023:13.160,2024:16.220,2025:6.950},
    "Investing Cash Flow":{2022:3.100,2023:3.030,2024:7.430,2025:-3.198},
    "Financing Cash Flow":{2022:-.799,2023:2.200,2024:-2.289,2025:-.420},
    "Net Change in Cash":{2022:-.048,2023:-.074,2024:2.680,2025:.619},
    "Acquisitions":{2025:-.014},
    "Share Repurchases":{2025:-.400},
    "Dividends":{2025:-.486},
    "Debt Repayments":{2025:-1.076},
}


def _find(ws,label,start=1,end=None):
    needle=str(label).strip().lower(); end=end or ws.max_row
    for r in range(start,min(end,ws.max_row)+1):
        if str(ws.cell(r,1).value or "").strip().lower()==needle: return r
    return None


def _section(ws,title): return _find(ws,title)

def _year_cols(ws,header_row):
    return {int(ws.cell(header_row,c).value):c for c in range(2,min(ws.max_column,12)+1) if isinstance(ws.cell(header_row,c).value,(int,float)) and 1900<=int(ws.cell(header_row,c).value)<=2100}


def _apply_map(ws,start,end,data,year_cols,source_col):
    written=0
    for label,series in data.items():
        r=_find(ws,label,start,end)
        if not r: continue
        for y,v in series.items():
            c=year_cols.get(y)
            if not c: continue
            ws.cell(r,c,v)
            if "EPS" in label: ws.cell(r,c).number_format=FMT_EPS
            elif "Shares (bn)" in label: ws.cell(r,c).number_format=FMT_SHARES
            else: ws.cell(r,c).number_format=FMT_BN
            written+=1
        if source_col:
            ws.cell(r,source_col,CEG_SOURCE); ws.cell(r,source_col).font=Font(italic=True)
    return written


def apply_verified_full_statement_adapter(wb,ticker):
    if str(ticker).upper()!="CEG" or "Financial Statements" not in wb.sheetnames:
        return {"ticker":str(ticker).upper(),"written":0}
    ws=wb["Financial Statements"]
    i0=_section(ws,"Income Statement"); b0=_section(ws,"Balance Sheet"); c0=_section(ws,"Cash Flow Statement")
    if not all((i0,b0,c0)): return {"ticker":"CEG","written":0}
    ih=next((r for r in range(i0+1,min(b0,i0+5)) if str(ws.cell(r,1).value or "").strip().lower()=="metric"),None)
    bh=next((r for r in range(b0+1,min(c0,b0+5)) if str(ws.cell(r,1).value or "").strip().lower()=="metric"),None)
    ch=next((r for r in range(c0+1,min(ws.max_row+1,c0+5)) if str(ws.cell(r,1).value or "").strip().lower()=="metric"),None)
    if not all((ih,bh,ch)): return {"ticker":"CEG","written":0}
    iy=_year_cols(ws,ih); by=_year_cols(ws,bh); cy=_year_cols(ws,ch)
    source_col=max(max(iy.values(),default=1),max(by.values(),default=1),max(cy.values(),default=1))+1
    n=0
    n+=_apply_map(ws,ih+1,b0-1,CEG_INCOME,iy,source_col)
    n+=_apply_map(ws,bh+1,c0-1,CEG_BALANCE,by,source_col)
    n+=_apply_map(ws,ch+1,ws.max_row,CEG_CASH,cy,source_col)
    return {"ticker":"CEG","written":n,"source":CEG_SOURCE}
