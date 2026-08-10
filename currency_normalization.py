from __future__ import annotations

"""Currency/ADR normalization for cross-border equity research models.

The workbook is valued in the traded security's quote currency. Foreign issuers often
report financial statements in a different functional/reporting currency. Yahoo market
capitalization and price can therefore be in the quote currency while cash, debt and
annual statements remain in the reporting currency. This module converts statement
amounts and per-ADS/per-share EPS into the quote currency before valuation analytics run.

The conversion is deliberately conservative. If a reliable FX series cannot be obtained,
it leaves the values unchanged and writes a Data Quality REVIEW rather than silently
mixing currencies.
"""

from functools import lru_cache
import math

import pandas as pd
import yfinance as yf
from openpyxl.styles import Font, PatternFill, Alignment

PALE_GREEN = "E2F0D9"
GOLD = "FFF2CC"
PALE_RED = "FCE4D6"
GREY = "666666"


def _num(v, default=None):
    try:
        if isinstance(v, bool) or v in (None, ""):
            return default
        x = float(v)
        return x if math.isfinite(x) else default
    except Exception:
        return default


def _currency(v):
    s = str(v or "").upper().strip()
    return s if len(s) == 3 and s.isalpha() else None


def _close_series(df: pd.DataFrame) -> pd.Series:
    if df is None or df.empty or "Close" not in df:
        return pd.Series(dtype=float)
    close = df["Close"]
    if isinstance(close, pd.DataFrame):
        close = close.iloc[:, 0]
    close = pd.to_numeric(close, errors="coerce").dropna()
    close = close[close > 0]
    return close


@lru_cache(maxsize=64)
def _fx_table(financial_currency: str, quote_currency: str):
    """Return ({year: avg conversion rate}, latest rate, source symbol).

    Rate means quote-currency units per one financial-currency unit.
    """
    f = _currency(financial_currency)
    q = _currency(quote_currency)
    if not f or not q:
        return {}, None, None
    if f == q:
        return {}, 1.0, "same currency"

    candidates = [
        (f"{f}{q}=X", False),
        (f"{q}{f}=X", True),
    ]
    for symbol, invert in candidates:
        try:
            df = yf.download(
                symbol,
                period="10y",
                interval="1d",
                auto_adjust=False,
                progress=False,
                threads=False,
                timeout=10,
            )
            close = _close_series(df)
        except Exception:
            close = pd.Series(dtype=float)
        if close.empty:
            continue
        if invert:
            close = 1.0 / close
        annual = close.groupby(close.index.year).mean().to_dict()
        annual = {int(y): float(v) for y, v in annual.items() if _num(v) is not None}
        latest = _num(close.iloc[-1])
        if latest and latest > 0:
            return annual, latest, symbol + (" (inverted)" if invert else "")
    return {}, None, None


def fx_rate(financial_currency: str, quote_currency: str, year: int | None = None):
    annual, latest, source = _fx_table(financial_currency, quote_currency)
    if latest is None:
        return None, source
    if year is None:
        return latest, source
    if year in annual:
        return annual[year], source
    if annual:
        nearest = min(annual, key=lambda y: abs(int(y) - int(year)))
        return annual[nearest], source
    return latest, source


def convert_financial_amount_to_quote(value, info: dict | None, year: int | None = None):
    """Convert a raw monetary amount from reporting currency to quote currency."""
    v = _num(value)
    if v is None:
        return None
    info = info or {}
    financial = _currency(info.get("financialCurrency"))
    quote = _currency(info.get("currency"))
    if not financial or not quote or financial == quote:
        return v
    rate, _ = fx_rate(financial, quote, year)
    return v * rate if rate else None


def _data_quality_row(wb, status: str, observed: str):
    if "Data Quality" not in wb.sheetnames:
        return
    ws = wb["Data Quality"]
    row = None
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip() == "Currency / ADR normalization":
            row = r
            break
    row = row or ws.max_row + 1
    ws.cell(row, 1, "Currency / ADR normalization")
    ws.cell(row, 2, status)
    ws.cell(row, 3, observed)
    ws.cell(
        row,
        4,
        "Price/market cap, cash/debt, historical statements and EPS must use one valuation currency. ADR-equivalent shares are derived from market cap ÷ traded price.",
    )
    fill = PALE_GREEN if status == "PASS" else (PALE_RED if status == "FAIL" else GOLD)
    ws.cell(row, 2).fill = PatternFill("solid", fgColor=fill)
    ws.cell(row, 2).font = Font(bold=True)
    for c in range(1, 5):
        ws.cell(row, c).alignment = Alignment(wrap_text=True, vertical="top")


def _yahoo_operating_income(ticker: str) -> dict[int, float]:
    """Raw operating income in Yahoo's financial/reporting currency."""
    try:
        df = yf.Ticker(ticker).income_stmt
    except Exception:
        return {}
    if df is None or getattr(df, "empty", True):
        return {}
    row = None
    for label in ("Operating Income", "Operating Income Loss"):
        if label in df.index:
            row = label
            break
    if row is None:
        return {}
    out = {}
    for col, value in df.loc[row].items():
        try:
            year = int(pd.Timestamp(col).year)
        except Exception:
            continue
        v = _num(value)
        if v is not None:
            out[year] = v
    return out


def _normalize_historical(wb, ticker: str, info: dict, financial: str, quote: str, source: str):
    if "Historical Financials" not in wb.sheetnames:
        return
    ws = wb["Historical Financials"]
    marker = f"Currency normalized: {financial} -> {quote}"
    if str(ws["D2"].value or "").strip() == marker:
        return

    op_income = _yahoo_operating_income(ticker)
    monetary_rows = (4, 6, 9, 11, 14, 15, 18, 19, 21)
    for c in range(2, 8):
        y = ws.cell(3, c).value
        if not isinstance(y, (int, float)):
            continue
        year = int(y)
        rate, _ = fx_rate(financial, quote, year)
        if not rate:
            continue
        for r in monetary_rows:
            v = _num(ws.cell(r, c).value)
            if v is not None:
                ws.cell(r, c).value = v * rate
        eps = _num(ws.cell(12, c).value)
        if eps is not None:
            ws.cell(12, c).value = eps * rate
        # Yahoo's cross-border fallback can expose Pretax Income before Operating Income.
        # Replace that row with actual operating income where the provider supplies it.
        if year in op_income:
            ws.cell(9, c).value = (op_income[year] * rate) / 1e9

    ws["A2"] = f"{quote} in billions except per-share data"
    ws["D2"] = marker
    ws["E2"] = f"FX: annual average public market FX series ({source})"
    ws["D2"].font = Font(italic=True, color=GREY)
    ws["E2"].font = Font(italic=True, color=GREY)


def _normalize_financial_statements(wb, financial: str, quote: str, source: str):
    if "Financial Statements" not in wb.sheetnames or "Historical Financials" not in wb.sheetnames:
        return
    ws = wb["Financial Statements"]
    hist = wb["Historical Financials"]

    # Determine whether the statement sheet is still in local currency by comparing a
    # matching revenue period against the already-normalized Historical Financials sheet.
    hist_rev = {}
    for c in range(2, 8):
        y = hist.cell(3, c).value
        v = _num(hist.cell(4, c).value)
        if isinstance(y, (int, float)) and v is not None:
            hist_rev[int(y)] = v

    labels = {str(ws.cell(r, 1).value or "").strip().lower(): r for r in range(1, ws.max_row + 1)}
    revenue_row = next((r for label, r in labels.items() if label in {"revenue", "total revenue", "net revenue", "operating revenue"}), None)
    if revenue_row is None:
        return

    header_row = 6
    needs_conversion = False
    for c in range(2, min(8, ws.max_column + 1)):
        y = ws.cell(header_row, c).value
        v = _num(ws.cell(revenue_row, c).value)
        if not isinstance(y, (int, float)) or v is None or int(y) not in hist_rev or hist_rev[int(y)] == 0:
            continue
        ratio = abs(v / hist_rev[int(y)])
        rate, _ = fx_rate(financial, quote, int(y))
        if rate and ratio > max(3.0, 0.4 / rate):
            needs_conversion = True
            break
    if not needs_conversion:
        ws["A3"] = f"{quote} billions unless per-share data. Cross-border values normalized from {financial} to {quote}; source hierarchy remains issuer IR / SEC annual XBRL / Yahoo fallback."
        return

    eps_labels = {"diluted eps", "diluted earnings per share"}
    ratio_keywords = ("margin", "%", "ratio", "growth", "per share")
    for c in range(2, min(8, ws.max_column + 1)):
        y = ws.cell(header_row, c).value
        if not isinstance(y, (int, float)):
            continue
        rate, _ = fx_rate(financial, quote, int(y))
        if not rate:
            continue
        for r in range(7, ws.max_row + 1):
            label = str(ws.cell(r, 1).value or "").strip().lower()
            if not label or any(k in label for k in ratio_keywords):
                continue
            v = _num(ws.cell(r, c).value)
            if v is None:
                continue
            ws.cell(r, c).value = v * rate
        for label in eps_labels:
            r = labels.get(label)
            if r:
                v = _num(ws.cell(r, c).value)
                if v is not None:
                    ws.cell(r, c).value = v * rate
    ws["A3"] = f"{quote} billions unless per-share data. Cross-border values normalized from {financial} to {quote}; FX source {source}."


def normalize_workbook_currency(wb, ticker: str, info: dict | None = None) -> bool:
    """Normalize foreign financial statements to the traded security's quote currency.

    Returns True if currencies differ and a usable FX conversion was applied.
    """
    info = info or {}
    quote = _currency(info.get("currency"))
    financial = _currency(info.get("financialCurrency"))
    if not quote or not financial:
        _data_quality_row(wb, "REVIEW", f"Could not resolve quote/reporting currencies: quote={quote}, reporting={financial}")
        return False
    if quote == financial:
        _data_quality_row(wb, "PASS", f"Single currency: {quote}")
        return False

    current_rate, source = fx_rate(financial, quote, None)
    if not current_rate:
        _data_quality_row(wb, "FAIL", f"Reporting={financial}; quote={quote}; no FX series resolved")
        return False

    if "Company Data" in wb.sheetnames:
        ws = wb["Company Data"]
        price = _num(info.get("currentPrice")) or _num(info.get("regularMarketPrice")) or _num(ws["B8"].value)
        market_cap_raw = _num(info.get("marketCap"))
        cash_raw = _num(info.get("totalCash"))
        debt_raw = _num(info.get("totalDebt"))
        market_cap = market_cap_raw / 1e9 if market_cap_raw is not None else _num(ws["B10"].value)
        cash = cash_raw * current_rate / 1e9 if cash_raw is not None else None
        debt = debt_raw * current_rate / 1e9 if debt_raw is not None else None
        if price is not None:
            ws["B8"] = price
        if market_cap is not None:
            ws["B10"] = market_cap
        if price and market_cap:
            ws["B9"] = market_cap / price
        if cash is not None:
            ws["B12"] = cash
        if debt is not None:
            ws["B13"] = debt
        if cash is not None and debt is not None:
            net_debt = debt - cash
            ws["B14"] = net_debt
            if market_cap is not None:
                ws["B11"] = market_cap + net_debt
        ws["A10"] = f"Market Cap ({quote} bn)"
        ws["A11"] = f"Enterprise Value ({quote} bn)"
        ws["A12"] = f"Cash ({quote} bn)"
        ws["A13"] = f"Total Debt ({quote} bn)"
        ws["A14"] = f"Net Debt / (Cash) ({quote} bn)"
        ws["D8"] = f"Currency normalization: reporting {financial} -> traded quote {quote}"
        ws["D9"] = f"Current FX ≈ {current_rate:.6f} {quote}/{financial} via {source}"

    _normalize_historical(wb, ticker, info, financial, quote, source or "public FX")
    _normalize_financial_statements(wb, financial, quote, source or "public FX")
    _data_quality_row(
        wb,
        "PASS",
        f"Reporting={financial}; quote={quote}; current FX={current_rate:.6f}; source={source}. ADR-equivalent shares use market cap / traded price.",
    )
    return True
