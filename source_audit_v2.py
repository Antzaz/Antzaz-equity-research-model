"""Final source/provenance corrections for saved research workbooks.

Two auditability issues are handled here:
- the recent-filings table can fill up with 8-K/10-Q entries and accidentally omit the latest
  annual filing, even though the 10-K is essential for tracing historical facts;
- Yahoo summary fields such as totalDebt and marketCap/price are useful valuation inputs but are
  provider definitions, not necessarily identical to the issuer's statutory balance-sheet labels
  or legal common-share count.
"""

from __future__ import annotations

from openpyxl.styles import Font

try:
    import update_model as core
except Exception:
    core=None

LINK_GREEN="008000"; GREY="666666"
ANNUAL_FORMS={"10-K","10-K/A","20-F","20-F/A","40-F","40-F/A"}


def _has_annual(ws):
    for r in range(4,min(ws.max_row,40)+1):
        if str(ws.cell(r,1).value or "").strip().upper() in ANNUAL_FORMS:
            return True
    return False


def ensure_latest_annual_filing(wb,ticker):
    """Guarantee that Filings retains at least the latest annual report URL when SEC data permits."""
    if "Filings" not in wb.sheetnames or _has_annual(wb["Filings"]):
        return False
    if core is None:
        return False
    try:
        cik=core.cik_for(ticker)
        if not cik: return False
        recent=core.sec_json(f"https://data.sec.gov/submissions/CIK{cik}.json").get("filings",{}).get("recent",{})
        chosen=None
        for form,period,filed,acc,doc in zip(
            recent.get("form",[]),recent.get("reportDate",[]),recent.get("filingDate",[]),
            recent.get("accessionNumber",[]),recent.get("primaryDocument",[]),
        ):
            if str(form).upper() not in ANNUAL_FORMS:
                continue
            url=f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{str(acc).replace('-','')}/{doc}"
            chosen=(form,period,filed,url,"SEC annual filing — retained for source audit")
            break
        if chosen is None: return False
    except Exception:
        return False

    ws=wb["Filings"]
    # Keep the newest filings but reserve the final visible slot for the annual source.
    target=15
    for r in range(4,16):
        if all(ws.cell(r,c).value in (None,"") for c in range(1,6)):
            target=r; break
    for c,v in enumerate(chosen,1): ws.cell(target,c).value=v
    ws.cell(target,4).hyperlink=chosen[3]; ws.cell(target,4).font=Font(color=LINK_GREEN,underline="single")
    return True


def clarify_company_data_definitions(wb):
    """Make provider/calculated market fields explicit without changing their values."""
    if "Company Data" not in wb.sheetnames:
        return False
    ws=wb["Company Data"]
    labels={
        8:"Current Price",
        9:"Valuation-Equivalent Shares (bn)",
        10:"Market Cap (USD bn)",
        11:"Provider Enterprise Value (USD bn)",
        12:"Provider Total Cash (USD bn)",
        13:"Provider Total Debt (USD bn)",
        14:"Provider Net Debt / (Cash) (USD bn)",
        15:"Provider Forward P/E",
    }
    notes={
        8:"Yahoo Finance market snapshot; exact timestamp/date should be read with the generated workbook date.",
        9:"Calculated as market cap ÷ traded share price. For multi-class/ADR issuers this is a valuation-equivalent share count, not necessarily the legal class share count.",
        10:"Yahoo Finance market-cap definition.",
        11:"Yahoo Finance enterprise-value field; may not exactly equal the model's EV because provider debt/cash definitions differ.",
        12:"Yahoo Finance totalCash provider field; reconcile with issuer cash/securities disclosure when material.",
        13:"Yahoo Finance totalDebt provider field; can include debt-like items beyond long-term notes shown in a filing footnote.",
        14:"Calculated from provider totalDebt − provider totalCash.",
        15:"Yahoo Finance forward P/E; use only when forward EPS is positive and normalized/comparable.",
    }
    for r,label in labels.items():
        ws.cell(r,1).value=label
        ws.cell(r,4).value=notes[r]
        ws.cell(r,4).font=Font(color=GREY,italic=True,size=9)
    ws.column_dimensions["A"].width=max(ws.column_dimensions["A"].width or 0,38)
    ws.column_dimensions["D"].width=max(ws.column_dimensions["D"].width or 0,72)
    return True


def apply_source_audit_fixes(wb,ticker):
    return {
        "annual_filing_added":ensure_latest_annual_filing(wb,ticker),
        "definitions_clarified":clarify_company_data_definitions(wb),
    }
