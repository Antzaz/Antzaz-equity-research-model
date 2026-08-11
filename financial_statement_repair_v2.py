from __future__ import annotations

"""Repair sparse Financial Statements sheets using structured Yahoo annual statements.

SEC Company Facts remains primary. This module only fills blank/known-bad cells after the
standard sheet has been generated, and records the fallback lineage. It also materializes
key derived values so a freshly generated workbook does not depend on Excel formula cache
for basic numbers such as operating margin and free cash flow.
"""

import math
import pandas as pd
import yfinance as yf
from openpyxl.styles import Font, Alignment

GREY="666666"; LINK_GREEN="008000"
FMT_BN='#,##0.0;[Red](#,##0.0);-'; FMT_PCT='0.0%;[Red](0.0%);-'; FMT_EPS='$0.00;[Red]($0.00);-'


def _num(v):
    try:
        if isinstance(v,bool) or v in (None,""): return None
        x=float(v); return x if math.isfinite(x) else None
    except Exception: return None


def _table(ticker,attr):
    try:
        t=yf.Ticker(ticker); df=getattr(t,attr)
        if callable(df): df=df()
        return df.copy() if isinstance(df,pd.DataFrame) else pd.DataFrame()
    except Exception: return pd.DataFrame()


def _row(df,aliases):
    if df is None or df.empty: return None
    for a in aliases:
        if a in df.index: return a
    low={str(x).strip().lower():x for x in df.index}
    for a in aliases:
        if a.lower() in low: return low[a.lower()]
    return None


def _year_values(df,aliases):
    r=_row(df,aliases); out={}
    if r is None: return out
    for c,v in df.loc[r].items():
        try: y=int(pd.Timestamp(c).year)
        except Exception: continue
        n=_num(v)
        if n is not None: out[y]=n
    return out


def _find_section(ws,title):
    for r in range(1,ws.max_row+1):
        if str(ws.cell(r,1).value or "").strip().lower()==title.lower(): return r
    return None


def _find_label(ws,label,start,end):
    for r in range(start,end+1):
        if str(ws.cell(r,1).value or "").strip().lower()==label.lower(): return r
    return None


def _headers(ws,row):
    return {int(ws.cell(row,c).value):c for c in range(2,min(8,ws.max_column+1)) if isinstance(ws.cell(row,c).value,(int,float))}


def _fill_series(ws,row,year_cols,values,per_share=False,overwrite=False):
    if row is None: return 0
    filled=0
    for y,c in year_cols.items():
        v=values.get(y)
        if v is None: continue
        current=_num(ws.cell(row,c).value)
        if current is None or overwrite:
            ws.cell(row,c).value=v if per_share else v/1e9
            ws.cell(row,c).number_format=FMT_EPS if per_share else FMT_BN
            filled+=1
    return filled


def repair_financial_statements(wb,ticker:str):
    if "Financial Statements" not in wb.sheetnames: return {"filled":0,"coverage":0}
    ws=wb["Financial Statements"]
    inc=_table(ticker,"income_stmt"); bs=_table(ticker,"balance_sheet"); cf=_table(ticker,"cashflow")
    if inc.empty and bs.empty and cf.empty: return {"filled":0,"coverage":0}

    income_start=_find_section(ws,"Income Statement") or 5
    balance_start=_find_section(ws,"Balance Sheet") or 22
    cash_start=_find_section(ws,"Cash Flow Statement") or 49
    income_header=income_start+1
    balance_header=next((r for r in range(balance_start+1,min(balance_start+4,ws.max_row)+1) if str(ws.cell(r,1).value or "").strip().lower()=="metric"),balance_start+1)
    cash_header=cash_start+1
    iy=_headers(ws,income_header); by=_headers(ws,balance_header); cy=_headers(ws,cash_header)

    income_map={
        "Revenue":["Total Revenue","Operating Revenue"],
        "Cost of Revenue":["Cost Of Revenue","Cost of Revenue"],
        "Gross Profit":["Gross Profit"],
        "Operating Income":["Operating Income"],
        "Pre-Tax Income":["Pretax Income","Pre Tax Income"],
        "Income Taxes":["Tax Provision","Income Tax Expense"],
        "Net Income":["Net Income","Net Income Common Stockholders"],
        "Diluted EPS":["Diluted EPS"],
    }
    filled=0; methods={}
    for label,aliases in income_map.items():
        r=_find_label(ws,label,income_start,balance_start-1); vals=_year_values(inc,aliases)
        overwrite=False
        if label=="Pre-Tax Income" and r is not None:
            op_r=_find_label(ws,"Operating Income",income_start,balance_start-1)
            # Repair the legacy fallback bug that mapped pre-tax income to operating income.
            pairs=[(c,_num(ws.cell(r,c).value),_num(ws.cell(op_r,c).value) if op_r else None) for c in iy.values()]
            overwrite=bool(pairs) and all(a is None or b is None or abs(a-b)<1e-9 for _,a,b in pairs)
        n=_fill_series(ws,r,iy,vals,per_share=(label=="Diluted EPS"),overwrite=overwrite)
        if n: filled+=n; methods[label]="Yahoo annual statements fallback"

    # Materialize other income and operating margin where the direct components exist.
    op_r=_find_label(ws,"Operating Income",income_start,balance_start-1); pt_r=_find_label(ws,"Pre-Tax Income",income_start,balance_start-1)
    other_r=_find_label(ws,"Other Income / (Expense), Net",income_start,balance_start-1); rev_r=_find_label(ws,"Revenue",income_start,balance_start-1)
    margin_r=_find_label(ws,"Operating Margin",income_start,balance_start-1)
    for y,c in iy.items():
        op=_num(ws.cell(op_r,c).value) if op_r else None; pt=_num(ws.cell(pt_r,c).value) if pt_r else None; rev=_num(ws.cell(rev_r,c).value) if rev_r else None
        if other_r and op is not None and pt is not None: ws.cell(other_r,c).value=pt-op; ws.cell(other_r,c).number_format=FMT_BN
        if margin_r and op is not None and rev not in (None,0): ws.cell(margin_r,c).value=op/rev; ws.cell(margin_r,c).number_format=FMT_PCT

    balance_map={
        "Cash & Cash Equivalents":["Cash And Cash Equivalents","Cash Cash Equivalents And Short Term Investments"],
        "Marketable / Short-Term Securities":["Other Short Term Investments","Investments And Other Financial Assets"],
        "Accounts Receivable":["Accounts Receivable"],
        "Other Current Assets":["Other Current Assets"],
        "Total Current Assets":["Current Assets","Total Current Assets"],
        "Property & Equipment, Net":["Net PPE","Property Plant And Equipment Net"],
        "Goodwill":["Goodwill"],
        "Total Assets":["Total Assets"],
        "Accounts Payable":["Accounts Payable"],
        "Deferred Revenue":["Current Deferred Revenue","Deferred Revenue"],
        "Total Current Liabilities":["Current Liabilities","Total Current Liabilities"],
        "Long-Term Debt":["Long Term Debt","Long Term Debt And Capital Lease Obligation"],
        "Total Liabilities":["Total Liabilities Net Minority Interest","Total Liabilities"],
        "Stockholders' Equity":["Stockholders Equity","Total Equity Gross Minority Interest"],
    }
    for label,aliases in balance_map.items():
        r=_find_label(ws,label,balance_start,cash_start-1); n=_fill_series(ws,r,by,_year_values(bs,aliases))
        if n: filled+=n; methods[label]="Yahoo annual balance-sheet fallback"

    cash_map={
        "Net Income":["Net Income"],
        "Depreciation & Amortization":["Depreciation And Amortization","Depreciation Amortization Depletion"],
        "Stock-Based Compensation":["Stock Based Compensation"],
        "Operating Cash Flow":["Operating Cash Flow","Total Cash From Operating Activities"],
        "Capital Expenditures":["Capital Expenditure","Capital Expenditures"],
        "Acquisitions":["Net Business Purchases"],
        "Share Repurchases":["Repurchase Of Capital Stock"],
        "Dividends":["Cash Dividends Paid"],
        "Debt Issuance":["Issuance Of Debt","Long Term Debt Issuance"],
        "Debt Repayments":["Repayment Of Debt","Long Term Debt Payments"],
        "Ending Cash":["End Cash Position","Changes In Cash"],
    }
    for label,aliases in cash_map.items():
        r=_find_label(ws,label,cash_start,ws.max_row); vals=_year_values(cf,aliases)
        if label in {"Capital Expenditures","Acquisitions","Share Repurchases","Dividends","Debt Repayments"}: vals={y:-abs(v) for y,v in vals.items()}
        n=_fill_series(ws,r,cy,vals)
        if n: filled+=n; methods[label]="Yahoo annual cash-flow fallback"

    # Materialize FCF instead of leaving a formula that has no cached value until Excel recalculates.
    ocf_r=_find_label(ws,"Operating Cash Flow",cash_start,ws.max_row); cap_r=_find_label(ws,"Capital Expenditures",cash_start,ws.max_row); fcf_r=_find_label(ws,"Free Cash Flow",cash_start,ws.max_row)
    if fcf_r and ocf_r and cap_r:
        for y,c in cy.items():
            o=_num(ws.cell(ocf_r,c).value); cap=_num(ws.cell(cap_r,c).value)
            if o is not None and cap is not None: ws.cell(fcf_r,c).value=o+cap; ws.cell(fcf_r,c).number_format=FMT_BN

    # Add transparent sheet-level lineage without changing the core statement layout.
    ws["H5"]="Source / Fallback Method"; ws["H5"].font=Font(bold=True)
    ws["H6"]="SEC Company Facts remains primary; blank or known-bad cells are filled from Yahoo annual statements. See Data Quality and Score Audit Trail."
    ws["H6"].alignment=Alignment(wrap_text=True,vertical="top")
    ws["H7"]="https://finance.yahoo.com/quote/"+ticker+"/financials/"
    ws["H7"].hyperlink=ws["H7"].value; ws["H7"].font=Font(color=LINK_GREEN,underline="single")
    ws.column_dimensions["H"].width=58

    core=("Revenue","Operating Income","Pre-Tax Income","Income Taxes","Net Income","Diluted EPS")
    available=0; total=0
    for label in core:
        r=_find_label(ws,label,income_start,balance_start-1)
        if not r: continue
        for c in iy.values(): total+=1; available+=1 if _num(ws.cell(r,c).value) is not None else 0
    coverage=available/total if total else 0
    return {"filled":filled,"coverage":coverage,"methods":methods}
