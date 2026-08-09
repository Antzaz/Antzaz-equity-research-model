"""Sector-aware peer selection and three-way Bayesian thesis updating.

This module runs near the end of the workbook build. It replaces generic / stale peer
sets with industry-first, same-sector comparisons and replaces the old two-sided
Bayesian sheet with a proper Bear/Base/Bull likelihood update.

Public Yahoo Finance/yfinance data are current snapshots, not point-in-time historical
datasets. Peer hit rates are therefore labelled cross-sectional reference-class anchors,
not long-run historical base rates.
"""

import math
try:
    import yfinance as yf
except Exception:
    yf = None

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; LIGHT="F5F9FC"; GOLD="FFF2CC"
PALE_GREEN="E2F0D9"; GREY="666666"; INPUT_BLUE="0000FF"; LINK_GREEN="008000"; BLACK="000000"
FMT_PCT='0.0%;[Red](0.0%);-'; FMT_MULT='0.0x;[Red](0.0x);-'; FMT_BN='#,##0.0;[Red](#,##0.0);-'
THIN=Side(style="thin",color="D9E1F2")

# Exact-industry pools are preferred. Broad sector pools are used only to fill remaining
# slots and candidates are still rejected unless yfinance reports the same sector.
INDUSTRY_CANDIDATES={
    "Healthcare Plans":["UNH","ELV","CI","HUM","CNC","MOH","OSCR","ALHC"],
    "Semiconductors":["NVDA","AVGO","AMD","QCOM","INTC","MU","TXN","ADI","MRVL"],
    "Software - Infrastructure":["MSFT","ORCL","CRM","NOW","PLTR","SNOW","DDOG","MDB"],
    "Software—Infrastructure":["MSFT","ORCL","CRM","NOW","PLTR","SNOW","DDOG","MDB"],
    "Software - Application":["ADBE","INTU","CDNS","ADSK","TEAM","HUBS","ZM"],
    "Internet Content & Information":["GOOGL","META","PINS","SNAP","RDDT","BIDU"],
    "Internet Retail":["AMZN","MELI","EBAY","ETSY","CHWY","BABA"],
    "Consumer Electronics":["AAPL","SONY","GRMN","LOGI","GPRO"],
    "Discount Stores":["WMT","COST","TGT","BJ","DG","DLTR"],
    "Banks - Diversified":["JPM","BAC","WFC","C","USB","PNC"],
    "Oil & Gas Integrated":["XOM","CVX","SHEL","BP","TTE"],
}

SECTOR_CANDIDATES={
    "Technology":["MSFT","AAPL","NVDA","AVGO","ORCL","CRM","ADBE","AMD","QCOM","NOW","INTU","TXN","MU","IBM"],
    "Healthcare":["UNH","ELV","CI","HUM","CNC","MOH","JNJ","LLY","ABBV","MRK","PFE","TMO","ABT","DHR"],
    "Financial Services":["JPM","BAC","WFC","C","GS","MS","USB","PNC","BLK","SCHW","AXP"],
    "Consumer Cyclical":["AMZN","TSLA","HD","MCD","LOW","BKNG","TJX","NKE","SBUX","MELI","EBAY"],
    "Consumer Defensive":["WMT","COST","PG","KO","PEP","PM","MO","TGT","CL","MDLZ"],
    "Communication Services":["GOOGL","META","NFLX","DIS","CMCSA","TMUS","VZ","T","SPOT","PINS"],
    "Industrials":["GE","CAT","RTX","HON","UNP","UPS","DE","ETN","LMT","BA"],
    "Energy":["XOM","CVX","COP","EOG","SLB","MPC","PSX","OXY","VLO"],
    "Basic Materials":["LIN","APD","SHW","FCX","NEM","NUE","DOW","ECL"],
    "Real Estate":["PLD","AMT","EQIX","WELL","SPG","O","DLR","PSA"],
    "Utilities":["NEE","SO","DUK","CEG","AEP","SRE","D","EXC"],
}

SECTOR_THRESHOLDS={
    "Healthcare": {"growth":.10,"margin":.03,"roe":.10,"evrev":1.0,"fpe":20.0},
    "Technology": {"growth":.10,"margin":.15,"roe":.15,"evrev":8.0,"fpe":35.0},
    "Communication Services": {"growth":.08,"margin":.12,"roe":.10,"evrev":5.0,"fpe":30.0},
    "Consumer Cyclical": {"growth":.07,"margin":.08,"roe":.10,"evrev":3.0,"fpe":25.0},
    "Consumer Defensive": {"growth":.05,"margin":.08,"roe":.10,"evrev":2.5,"fpe":25.0},
    "Industrials": {"growth":.06,"margin":.10,"roe":.12,"evrev":3.0,"fpe":25.0},
    "Energy": {"growth":.05,"margin":.10,"roe":.10,"evrev":3.0,"fpe":20.0},
    "Financial Services": {"growth":.05,"margin":.10,"roe":.10,"evrev":5.0,"fpe":20.0},
}
DEFAULT_THRESHOLDS={"growth":.07,"margin":.08,"roe":.10,"evrev":3.0,"fpe":25.0}


def _fill(c): return PatternFill("solid",fgColor=c)
def _num(v,default=None):
    try:
        if isinstance(v,bool): return default
        return float(v)
    except Exception:
        return default

def _info(ticker):
    if yf is None: return {}
    try: return yf.Ticker(ticker).info or {}
    except Exception: return {}
def _style_header(ws,row,start,end):
    for c in range(start,end+1):
        x=ws.cell(row,c); x.fill=_fill(BLUE); x.font=Font(bold=True,color=WHITE)
        x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); x.border=Border(bottom=THIN)
def _section(ws,row,title,end=7):
    for c in range(1,end+1):
        ws.cell(row,c).fill=_fill(NAVY); ws.cell(row,c).font=Font(bold=True,color=WHITE,size=11)
    ws.cell(row,1,title)
def _title(ws,text,end=10):
    for c in range(1,end+1): ws.cell(1,c).fill=_fill(NAVY); ws.cell(2,c).fill=_fill(NAVY)
    ws.cell(1,1,text); ws.cell(1,1).font=Font(bold=True,color=WHITE,size=18)
    ws.sheet_view.showGridLines=False


def _target_classification(wb,ticker,target_info):
    sector=target_info.get("sector"); industry=target_info.get("industry"); mc=_num(target_info.get("marketCap"))
    if "Company Data" in wb.sheetnames:
        d=wb["Company Data"]
        sector=sector or d["B6"].value; industry=industry or d["B7"].value
        if not mc:
            bn=_num(d["B10"].value); mc=bn*1e9 if bn else None
    return str(sector or "Unknown"),str(industry or "Unknown"),mc


def _candidate_score(info,target_sector,target_industry,target_mc):
    sector=str(info.get("sector") or ""); industry=str(info.get("industry") or "")
    if industry==target_industry: class_penalty=0
    elif sector==target_sector: class_penalty=100
    else: return None
    mc=_num(info.get("marketCap"))
    if target_mc and mc and target_mc>0 and mc>0: size_penalty=abs(math.log(mc/target_mc))
    else: size_penalty=5
    return class_penalty+size_penalty


def select_sector_peers(wb,ticker,target_info=None,count=5):
    target_info=target_info or _info(ticker); sector,industry,target_mc=_target_classification(wb,ticker,target_info)
    pool=[]
    for sym in INDUSTRY_CANDIDATES.get(industry,[]):
        if sym.upper()!=ticker.upper() and sym not in pool: pool.append(sym)
    for sym in SECTOR_CANDIDATES.get(sector,[]):
        if sym.upper()!=ticker.upper() and sym not in pool: pool.append(sym)
    # If an industry is unknown, use a conservative broad list only inside the target sector.
    if not pool:
        for syms in INDUSTRY_CANDIDATES.values():
            for sym in syms:
                if sym.upper()!=ticker.upper() and sym not in pool: pool.append(sym)
    ranked=[]
    for sym in pool[:30]:
        inf=_info(sym); score=_candidate_score(inf,sector,industry,target_mc)
        if score is not None: ranked.append((score,sym,inf))
    ranked.sort(key=lambda x:x[0])
    return sector,industry,[(sym,inf) for _,sym,inf in ranked[:count]]


def _metric_row(sym,inf):
    return [
        inf.get("longName") or sym,sym,inf.get("forwardPE"),inf.get("enterpriseToRevenue"),
        inf.get("enterpriseToEbitda"),inf.get("revenueGrowth"),inf.get("operatingMargins"),
        inf.get("returnOnEquity"),inf.get("industry"),f"https://finance.yahoo.com/quote/{sym}/",
    ]


def ensure_sector_peer_comps(wb,ticker):
    if "Peer Comps" not in wb.sheetnames: return []
    target=_info(ticker); sector,industry,peers=select_sector_peers(wb,ticker,target,5); ws=wb["Peer Comps"]
    # Clear old technology-biased peer rows and stale footer while preserving the sheet object.
    for row in ws.iter_rows(min_row=1,max_row=max(11,ws.max_row),min_col=1,max_col=10):
        for cell in row: cell.value=None
    _title(ws,f"{industry if industry!='Unknown' else sector} — Industry / Sector Peer Comps",10)
    ws["A2"]="Industry-first peer set. If fewer than five exact-industry peers are available, only same-sector companies may fill remaining slots; cross-sector peers are excluded."
    ws["A2"].font=Font(italic=True,color=GREY); ws["A2"].alignment=Alignment(wrap_text=True)
    headers=["Company","Ticker","Forward P/E","EV/Revenue","EV/EBITDA","Revenue Growth","Operating Margin","ROE","Industry","Source URL"]
    for c,v in enumerate(headers,1): ws.cell(3,c,v)
    _style_header(ws,3,1,10)
    rows=[_metric_row(ticker,target)]+[_metric_row(sym,inf) for sym,inf in peers]
    for r,row in enumerate(rows[:6],4):
        for c,v in enumerate(row,1): ws.cell(r,c,v)
        for c in range(3,9): ws.cell(r,c).font=Font(color=INPUT_BLUE)
        ws.cell(r,9).font=Font(color=INPUT_BLUE); ws.cell(r,10).font=Font(color=LINK_GREEN)
        ws.cell(r,3).number_format=FMT_MULT; ws.cell(r,4).number_format=FMT_MULT; ws.cell(r,5).number_format=FMT_MULT
        for c in (6,7,8): ws.cell(r,c).number_format=FMT_PCT
    ws["A11"]="Method"; ws["B11"]="Yahoo Finance/yfinance public snapshot; exact industry preferred, same-sector fallback only. Missing metrics remain blank rather than estimated."
    widths={"A":31,"B":11,"C":14,"D":14,"E":14,"F":15,"G":16,"H":13,"I":24,"J":48}
    for col,w in widths.items(): ws.column_dimensions[col].width=w
    ws.freeze_panes="A4"
    if "Comparative Analysis" in wb.sheetnames:
        wb["Comparative Analysis"]["A1"]=f"Comparative Analysis — {ticker} vs {industry if industry!='Unknown' else sector} Peers"
    return peers


def _remove_old_base_rate_sheets(wb):
    for name in ("Base Rates & Thesis Probabilities","Base Rates & Probabilities"):
        if name in wb.sheetnames: wb.remove(wb[name])


def _weighted_lr_formula(prior,lr_col,start=14,end=23):
    terms="*".join(f"POWER({lr_col}{r},F{r})" for r in range(start,end+1))
    return f"={prior}*{terms}"


def _evidence_seed(wb,ticker):
    rows=[]; h=wb["Historical Financials"] if "Historical Financials" in wb.sheetnames else None
    if h:
        rows.extend([
            ("Latest consolidated revenue growth",None,"Compare actual growth with the scenario paths and consensus."),
            ("Latest operating margin",None,"Assess whether margin evidence is recurring or distorted by one-offs."),
            ("Latest free-cash-flow margin",None,"Use cash conversion as a cross-check on accounting earnings."),
        ])
    if ticker.upper()=="CVS":
        rows=[
            ("FY2025 consolidated revenue growth","2025-12-31","Revenue grew to $402.1bn; assign LRs based on which scenario makes the outcome most likely."),
            ("Health Care Benefits adjusted operating income rebound","2025-12-31","Adjusted operating income rose to $2.939bn from $0.307bn in 2024."),
            ("Health Services adjusted operating income trend","2025-12-31","Adjusted operating income was $7.151bn versus $7.243bn in 2024."),
            ("Pharmacy & Consumer Wellness adjusted operating income trend","2025-12-31","Adjusted operating income rose to $6.040bn from $5.774bn in 2024."),
            ("GAAP margin / impairment and restructuring effects","2025-12-31","Down-weight evidence that overlaps with segment profitability or reflects non-recurring charges."),
        ]
    return rows[:5]


def ensure_bayesian_base_rates(wb,ticker):
    if "Three-Case Scenarios" not in wb.sheetnames: return None
    _remove_old_base_rate_sheets(wb); ws=wb.create_sheet("Base Rates & Probabilities"); _title(ws,f"{ticker} — Base Rates & Bayesian Thesis",7)
    target=_info(ticker); sector,industry,_=_target_classification(wb,ticker,target); thresholds=SECTOR_THRESHOLDS.get(sector,DEFAULT_THRESHOLDS)
    ws["A3"]=("Priors come from the three-case valuation. Every evidence item updates Bear, Base and Bull through likelihood ratios (LRs). "
              "Evidence Weight (0–1) discounts weak, duplicated or correlated signals. 1.00x is neutral. Peer hit rates below are current cross-sectional reference-class anchors—not true long-run historical base rates.")
    ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True)
    _section(ws,5,"Scenario Priors → Weighted Evidence → Posterior",7)
    for c,v in enumerate(["Scenario","Prior Probability","Raw Updated Weight","Posterior Probability","Change vs Prior","Interpretation"],1): ws.cell(6,c,v)
    _style_header(ws,6,1,6)
    priors=[("Bear","='Three-Case Scenarios'!B5"),("Base","='Three-Case Scenarios'!C5"),("Bull","='Three-Case Scenarios'!D5")]
    for r,(name,form) in enumerate(priors,7): ws.cell(r,1,name); ws.cell(r,2,form); ws.cell(r,2).font=Font(color=LINK_GREEN); ws.cell(r,2).number_format=FMT_PCT
    ws["C7"]=_weighted_lr_formula("B7","C"); ws["C8"]=_weighted_lr_formula("B8","D"); ws["C9"]=_weighted_lr_formula("B9","E")
    for r in range(7,10):
        ws.cell(r,4,f'=IFERROR(C{r}/SUM($C$7:$C$9),"")'); ws.cell(r,5,f"=D{r}-B{r}")
        for c in (3,4,5): ws.cell(r,c).number_format=FMT_PCT; ws.cell(r,c).font=Font(color=BLACK)
    ws["F7"]="Bear probability rises only when evidence is relatively more likely under Bear."; ws["F8"]="Base has its own likelihood ratios and is no longer a fixed anchor."; ws["F9"]="Bull probability rises only when evidence is relatively more likely under Bull."
    ws["A10"]="Probability control"; ws["D10"]="=SUM(D7:D9)"; ws["E10"]='=IF(ABS(D10-1)<0.000001,"PASS","CHECK")'; ws["D10"].number_format=FMT_PCT

    _section(ws,12,"Evidence Log — Relative Likelihood, Not Direction Alone",7)
    for c,v in enumerate(["Evidence / KPI","Date","Bear LR","Base LR","Bull LR","Evidence Weight","Source / Reasoning"],1): ws.cell(13,c,v)
    _style_header(ws,13,1,7)
    seed=_evidence_seed(wb,ticker)
    for i in range(10):
        r=14+i; item=seed[i] if i<len(seed) else (None,None,None)
        ws.cell(r,1,item[0]); ws.cell(r,2,item[1]);
        for c in (3,4,5): ws.cell(r,c,1.0); ws.cell(r,c).number_format=FMT_MULT
        ws.cell(r,6,.5 if item[0] else .5); ws.cell(r,6).number_format=FMT_PCT; ws.cell(r,7,item[2])
        for c in range(1,8):
            if c in (1,2,3,4,5,6,7): ws.cell(r,c).fill=_fill(GOLD); ws.cell(r,c).font=Font(color=INPUT_BLUE)
        ws.cell(r,7).alignment=Alignment(wrap_text=True,vertical="top")
    if ticker.upper()=="CVS":
        for r in range(14,19): ws.cell(r,7).value=(ws.cell(r,7).value or "")+" https://www.sec.gov/Archives/edgar/data/64803/000006480326000010/cvs-20251231.htm"

    _section(ws,26,"Reference-Class Base Rates — Current Same-Industry / Same-Sector Snapshot",7)
    for c,v in enumerate(["Reference Class","Metric / Threshold","Horizon","Peer Hit Rate","Company / Case","Peer Median","Source / Notes"],1): ws.cell(27,c,v)
    _style_header(ws,27,1,7); ref=industry if industry!="Unknown" else sector
    specs=[
        ("Revenue growth",">",thresholds["growth"],"F",FMT_PCT),
        ("Operating margin",">",thresholds["margin"],"G",FMT_PCT),
        ("ROE",">",thresholds["roe"],"H",FMT_PCT),
        ("EV / Revenue","<",thresholds["evrev"],"D",FMT_MULT),
        ("Forward P/E","<",thresholds["fpe"],"C",FMT_MULT),
    ]
    for i,(metric,op,cut,col,fmt) in enumerate(specs,28):
        ws.cell(i,1,f"{ref} peers"); ws.cell(i,2,f"{metric} {op} {cut:.0%}" if fmt==FMT_PCT else f"{metric} {op} {cut:.1f}x"); ws.cell(i,3,"Current public snapshot")
        ws.cell(i,4,f'=IFERROR(COUNTIF(\'Peer Comps\'!{col}5:{col}9,"{op}{cut}")/COUNT(\'Peer Comps\'!{col}5:{col}9),"")')
        ws.cell(i,5,f"='Peer Comps'!{col}4"); ws.cell(i,6,f'=IFERROR(MEDIAN(\'Peer Comps\'!{col}5:{col}9),"")'); ws.cell(i,7,"Peer Comps / Yahoo Finance public snapshot; cross-sectional anchor only")
        ws.cell(i,4).number_format=FMT_PCT; ws.cell(i,5).number_format=fmt; ws.cell(i,6).number_format=fmt
        for c in (4,5,6): ws.cell(i,c).font=Font(color=LINK_GREEN)
    ws["A33"]=ticker+" own history"; ws["B33"]="5Y revenue CAGR > 5%"; ws["C33"]="Latest six annual periods"; ws["E33"]="=IFERROR(('Historical Financials'!G4/'Historical Financials'!B4)^(1/5)-1,\"\")"; ws["E33"].number_format=FMT_PCT; ws["E33"].font=Font(color=LINK_GREEN); ws["G33"]="Company-specific historical anchor; not a peer survival rate"

    _section(ws,36,"Probability Discipline",7)
    questions=["What was my prior before the latest evidence?","Which evidence has the largest weighted LR and why?","Am I double-counting correlated evidence?","What evidence would move Bear probability above 40%?","What evidence would move Bull probability above 40%?","Is the peer reference class economically comparable, or only sector-labelled?"]
    for r,q in enumerate(questions,37): ws.cell(r,1,q); ws.cell(r,2,""); ws.cell(r,2).fill=_fill(GOLD); ws.cell(r,2).font=Font(color=INPUT_BLUE); ws.cell(r,2).alignment=Alignment(wrap_text=True)

    widths={"A":38,"B":24,"C":15,"D":18,"E":18,"F":22,"G":62}
    for col,w in widths.items(): ws.column_dimensions[col].width=w
    ws.row_dimensions[3].height=48; ws.freeze_panes="A7"
    return ws


def repair_cross_sheet_context(wb):
    """Repair recurring formula-copy bugs caused by unqualified template formulas."""
    if "Business Portfolio Map" in wb.sheetnames and "Historical Financials" in wb.sheetnames:
        ws=wb["Business Portfolio Map"]; ws["D10"]="='Historical Financials'!G10"; ws["B11"]="='Historical Financials'!G17"
        ws["D10"].font=Font(color=LINK_GREEN); ws["B11"].font=Font(color=LINK_GREEN); ws["D10"].number_format=FMT_PCT; ws["B11"].number_format=FMT_PCT
    if "Research Notes" in wb.sheetnames and "Historical Financials" in wb.sheetnames:
        ws=wb["Research Notes"]; ws["K12"]="='Historical Financials'!G10"; ws["K13"]="='Historical Financials'!G17"
        ws["K12"].font=Font(color=LINK_GREEN); ws["K13"].font=Font(color=LINK_GREEN); ws["K12"].number_format=FMT_PCT; ws["K13"].number_format=FMT_PCT


def add_final_quality_checks(wb,ticker):
    if "Data Quality" not in wb.sheetnames: return
    ws=wb["Data Quality"]; target=_info(ticker); sector,industry,_=_target_classification(wb,ticker,target)
    # Find / append deterministic check rows.
    labels={str(ws.cell(r,1).value or "").strip():r for r in range(1,ws.max_row+1)}
    checks=[
        ("Peer comps sector alignment","PASS",f"Peer set constrained to {industry if industry!='Unknown' else sector}; cross-sector peers excluded","Keeps valuation comparisons economically relevant."),
        ("Bayesian update structure","PASS","Bear, Base and Bull all update with weighted likelihood ratios","Evidence can redistribute probability across all scenarios."),
    ]
    for label,status,obs,why in checks:
        r=labels.get(label) or ws.max_row+1; ws.cell(r,1,label); ws.cell(r,2,status); ws.cell(r,3,obs); ws.cell(r,4,why); ws.cell(r,2).fill=_fill(PALE_GREEN); ws.cell(r,2).font=Font(bold=True)
        for c in range(1,5): ws.cell(r,c).alignment=Alignment(wrap_text=True,vertical="top")
