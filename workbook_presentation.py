from __future__ import annotations

"""Light-touch workbook consolidation and presentation polish.

The research workbook contains a lot of useful detail, but the analyst should land on a small
set of decision-facing sheets first. This module keeps the underlying model intact while:
- ordering sheets into a logical research workflow;
- hiding only clearly redundant/technical support tabs by default;
- applying consistent tab colors, gridline/zoom settings and a cleaner dashboard title;
- adding internal navigation links to the visual dashboard;
- preserving every underlying calculation/source sheet for auditability.

The pass is deliberately idempotent because it runs once after deterministic generation and
again after optional ML/AI sheets are appended.
"""

from typing import Any

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.cell import range_boundaries

NAVY = "17365D"; BLUE = "2F75B5"; GREEN = "70AD47"; GOLD = "FFC000"; PURPLE = "8064A2"; GREY = "A5A5A5"; WHITE = "FFFFFF"; LIGHT = "F5F9FC"

PRIMARY_ORDER = [
    "Visual Dashboard",
    "Investment Summary",
    "Decision View",
    "Company Data",
    "Historical Financials",
    "Financial Statements",
    "Segment Analysis",
    "Three-Case Scenarios",
    "DCF",
    "Peer Comps",
    "Advanced Analytics",
    "ML & Quantitative Research",
    "AI Growth Forecast",
    "Deals & Transactions",
    "Ownership & Holders",
    "Research Notes",
    "Data Quality",
]

TAB_GROUPS = {
    NAVY: {"Visual Dashboard", "Investment Summary", "Decision View"},
    GREEN: {"Company Data", "Historical Financials", "Financial Statements", "Segment Analysis"},
    GOLD: {"Three-Case Scenarios", "DCF", "Peer Comps", "Advanced Analytics", "Stress Test"},
    PURPLE: {"ML & Quantitative Research", "AI Growth Forecast", "AI Impact Analysis"},
    GREY: {"Deals & Transactions", "Ownership & Holders", "Research Notes", "Filings", "Data Quality", "Score Audit Trail"},
}

# These tabs are retained for formula compatibility / auditability, but duplicate a cleaner primary
# view. Hiding is intentionally conservative; no source/data sheet is deleted here.
HIDE_IF_REDUNDANT = {
    "Dashboard": "Visual Dashboard",
    "Comparative Analysis": "Peer Comps",
    "Score Audit Trail": "Data Quality",
}


def _tab_color(sheet_name: str) -> str | None:
    for color, names in TAB_GROUPS.items():
        if sheet_name in names:
            return color
    return None


def _reorder(wb) -> None:
    priority = {name: i for i, name in enumerate(PRIMARY_ORDER)}
    # Capture original positions *before* sorting. Calling list.index() from inside a list.sort
    # key is unsafe because CPython temporarily mutates the target list during sorting.
    original = list(wb._sheets)
    original_pos = {id(ws): i for i, ws in enumerate(original)}
    wb._sheets[:] = sorted(
        original,
        key=lambda ws: (0, priority[ws.title]) if ws.title in priority else (1, original_pos[id(ws)]),
    )


def _style_tabs_and_views(wb) -> None:
    for ws in wb.worksheets:
        color = _tab_color(ws.title)
        if color:
            ws.sheet_properties.tabColor = color
        if ws.title in set(PRIMARY_ORDER) | {"Filings", "Score Audit Trail"}:
            ws.sheet_view.showGridLines = False
            try:
                ws.sheet_view.zoomScale = 90 if ws.title in {"Financial Statements", "Segment Analysis", "ML & Quantitative Research", "AI Growth Forecast"} else 95
            except Exception:
                pass


def _hide_redundant_tabs(wb) -> list[str]:
    hidden = []
    for old, replacement in HIDE_IF_REDUNDANT.items():
        if old in wb.sheetnames and replacement in wb.sheetnames:
            wb[old].sheet_state = "hidden"
            hidden.append(old)
    return hidden


def _unmerge_intersecting(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        a, b, c, d = range_boundaries(str(merged))
        if not (c < min_col or a > max_col or d < min_row or b > max_row):
            ws.unmerge_cells(str(merged))


def _dashboard_navigation(wb, ticker: str) -> int:
    if "Visual Dashboard" not in wb.sheetnames:
        return 0
    ws = wb["Visual Dashboard"]
    company = ""
    try:
        company = str(wb["Company Data"]["B5"].value or "").strip()
    except Exception:
        pass
    ws["A1"] = f"{ticker} — Equity Research Dashboard"
    ws["A1"].font = Font(bold=True, color=WHITE, size=18)
    if company:
        ws["A3"] = f"{company} | valuation, operating quality, scenarios, peers, ML/AI and source-quality controls"
        ws["A3"].font = Font(italic=True, color="666666")

    row = 98
    # This area is rebuilt on every pass. Unmerge first so a second run never writes into read-only
    # MergedCell objects left by the first run.
    _unmerge_intersecting(ws, row, row + 3, 1, 16)
    for r in range(row, row + 4):
        for c in range(1, 17):
            cell = ws.cell(r, c)
            cell.value = None
            cell.hyperlink = None
            cell.fill = PatternFill(fill_type=None)
            cell.font = Font()
            cell.alignment = Alignment()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=16)
    ws.cell(row, 1, "Workbook Navigation")
    ws.cell(row, 1).fill = PatternFill("solid", fgColor=NAVY); ws.cell(row, 1).font = Font(bold=True, color=WHITE)

    links = [
        ("Financials", "Financial Statements"),
        ("Segments", "Segment Analysis"),
        ("Scenarios", "Three-Case Scenarios"),
        ("Peers", "Peer Comps"),
        ("Advanced", "Advanced Analytics"),
        ("ML", "ML & Quantitative Research"),
        ("AI Growth", "AI Growth Forecast"),
        ("Data Quality", "Data Quality"),
    ]
    written = 0
    col = 1
    for label, target in links:
        if target not in wb.sheetnames:
            continue
        cell = ws.cell(row + 1, col)
        cell.value = label
        cell.hyperlink = f"#'{target}'!A1"
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(bold=True, color=WHITE, underline="single")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=min(16, col + 1))
        col += 2; written += 1
        if col > 16:
            break
    ws.row_dimensions[row + 1].height = 24
    return written


def _decorate_quality(wb, hidden: list[str], navigation_links: int) -> None:
    if "Data Quality" not in wb.sheetnames:
        return
    ws = wb["Data Quality"]; label = "Workbook presentation / consolidation"
    row = next((r for r in range(1, ws.max_row + 1) if str(ws.cell(r, 1).value or "").strip() == label), ws.max_row + 1)
    visible = sum(1 for s in wb.worksheets if s.sheet_state == "visible")
    detail = (
        f"{visible} sheet(s) visible in a decision-first order; redundant support tabs hidden by default: "
        + (", ".join(hidden) if hidden else "none")
        + f". Visual Dashboard contains {navigation_links} internal quick link(s)."
    )
    for c, value in enumerate((label, "PASS", detail, "No research/source sheet is deleted by the presentation pass; hidden tabs remain available for auditability."), 1):
        ws.cell(row, c).value = value; ws.cell(row, c).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row, 2).fill = PatternFill("solid", fgColor="E2F0D9"); ws.cell(row, 2).font = Font(bold=True)


def apply_workbook_presentation(wb, ticker: str) -> dict[str, Any]:
    ticker = str(ticker or "").upper().strip()
    _reorder(wb)
    if "Visual Dashboard" in wb.sheetnames:
        try:
            wb.active = wb.index(wb["Visual Dashboard"])
        except Exception:
            pass
    _style_tabs_and_views(wb)
    hidden = _hide_redundant_tabs(wb)
    navigation_links = _dashboard_navigation(wb, ticker)
    _decorate_quality(wb, hidden, navigation_links)
    return {"hidden_tabs": hidden, "navigation_links": navigation_links, "visible_sheets": sum(1 for s in wb.worksheets if s.sheet_state == "visible")}
