"""Institutional-style AI impact analysis for the equity research workbook.

Creates a company-agnostic AI Impact Analysis sheet that separates reported AI
evidence, segment exposure, capital intensity, disruption risk, and AI-related
valuation surprise versus the existing Base DCF. The 0-delta Base / Embedded
case is designed to reconcile to the authoritative Base model rather than adding
a second AI premium on top.
"""

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; LIGHT="F5F9FC"; PALE="D9EAF7"
GOLD="FFF2CC"; INPUT_BLUE="0000FF"; LINK_GREEN="008000"; GREY="666666"; BLACK="000000"
FMT_PCT='0.0%;[Red](0.0%);-'; FMT_BN='$#,##0.0;[Red]($#,##0.0);-'
FMT_NUM='#,##0.0;[Red](#,##0.0);-'; FMT_PRICE='$#,##0.00;[Red]($#,##0.00);-'
FMT_SCORE='0.00;[Red](0.00);-'; FMT_DF='0.000x;[Red](0.000x);-'
THIN=Side(style="thin",color="D9E1F2")

AMZN_Q2="https://ir.aboutamazon.com/news-release/news-release-details/2026/Amazon-com-Announces-Second-Quarter-Results/default.aspx"
AMZN_Q4="https://ir.aboutamazon.com/news-release/news-release-details/2026/Amazon-com-Announces-Fourth-Quarter-Results/"
GOOG_Q2="https://blog.google/company-news/inside-google/message-ceo/alphabet-earnings-q2-2026/"

EVIDENCE_PACKS={
    "AMZN":[
        ("AWS Q2 net sales",42.2,"$bn / +36.7% YoY","Very strong","Fastest AWS growth in 18 quarters; AI demand is translating into reported cloud revenue.","2026-07-30",AMZN_Q2,"Company reported"),
        ("AWS Q2 operating income",16.6,"$bn / ~39.3% margin","Very strong","Shows high cloud profitability despite heavy AI infrastructure build-out.","2026-07-30",AMZN_Q2,"Company reported"),
        ("AWS AI business annual run-rate",25.0,"$bn+ / triple-digit YoY","Very strong","Direct monetization indicator for AI infrastructure and services; management-defined run-rate, not a GAAP segment.","2026-07-30",AMZN_Q2,"Management KPI"),
        ("Amazon chips business annual run-rate",25.0,"$bn+ / triple-digit YoY","Very strong","Trainium/Graviton/Nitro can improve economics, reduce vendor dependence and deepen AWS differentiation.","2026-07-30",AMZN_Q2,"Management KPI"),
        ("TTM free cash flow",-7.6,"$bn","Key risk","FCF moved into outflow as property/equipment purchases rose; Amazon says the increase primarily reflects AI investment.","2026-07-30",AMZN_Q2,"Company reported"),
        ("2026 capital expenditure guide",200.0,"$bn approx.","Key risk / opportunity","AI demand is capital intensive; the investment case depends on utilization, pricing and incremental returns on invested capital.","2026-02-05",AMZN_Q4,"Management guidance"),
        ("Bedrock customer/spend momentum",None,"Hundreds of thousands of customers; Q2 spend > all prior quarters","Strong","Broad model choice and enterprise adoption support a platform thesis, but revenue attribution remains partly undisclosed.","2026-07-30",AMZN_Q2,"Qualitative KPI"),
        ("AI shopping engagement",None,"Active users ~2x; interactions >5x YoY","Strong","AI can raise conversion, basket size and Prime engagement in Stores, not just AWS.","2026-07-30",AMZN_Q2,"Qualitative KPI"),
        ("Ads Agent economics",None,"8% lower CPM; 6% lower CPA","Positive","Early evidence that AI tools may improve advertiser ROI and support ad monetization.","2026-07-30",AMZN_Q2,"Product KPI"),
    ],
    "GOOGL":[
        ("Google Cloud Q2 revenue",24.8,"$bn / +82% YoY","Very strong","Cloud acceleration is a direct enterprise AI demand and monetization signal.","2026-Q2",GOOG_Q2,"Company reported"),
        ("Google Cloud backlog",514.0,"$bn","Very strong","Large contracted demand supports duration, but conversion depends on capacity and customer utilization.","2026-Q2",GOOG_Q2,"Company reported"),
        ("Gemini app monthly active users",950.0,"million","Strong","Near-billion-user distribution creates monetization optionality; revenue/user and engagement quality remain key.","2026-Q2",GOOG_Q2,"Company reported"),
        ("AI Mode monthly active users",1.0,"billion+","Strong","Tests whether AI expands Search usage without degrading commercial query economics.","2026-Q2",GOOG_Q2,"Company reported"),
    ],
}
EVIDENCE_PACKS["GOOG"]=EVIDENCE_PACKS["GOOGL"]
GENERIC_KPIS=["AI-attributed / AI-assisted revenue","AI product users / seats / paid workloads","AI price realization / ARPU / attach rate","AI gross or contribution margin","AI-related backlog / commitments","Inference / training unit-cost trend","AI capex / revenue and utilization","AI productivity / cost savings","Cannibalization / displacement indicator"]


def _fill(c): return PatternFill("solid",fgColor=c)
def _merge(ws,rng):
    if rng not in {str(x) for x in ws.merged_cells.ranges}: ws.merge_cells(rng)
def _section(ws,row,title,end=14):
    _merge(ws,f"A{row}:{get_column_letter(end)}{row}"); c=ws.cell(row,1); c.value=title; c.fill=_fill(NAVY); c.font=Font(bold=True,color=WHITE,size=11)
def _header(ws,row,start,end):
    for c in range(start,end+1):
        x=ws.cell(row,c); x.fill=_fill(BLUE); x.font=Font(bold=True,color=WHITE); x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); x.border=Border(bottom=THIN)
def _formula(cell,formula,fmt=None,internal=False):
    cell.value=formula; cell.font=Font(color=LINK_GREEN if internal else BLACK)
    if fmt: cell.number_format=fmt
def _reported(cell,value,source,fmt=None):
    cell.value=value; cell.font=Font(color=INPUT_BLUE); cell.comment=Comment(f"Company-reported / external input. Source: {source}","OpenAI")
    if fmt: cell.number_format=fmt
def _assumption(cell,value,fmt=FMT_PCT,note=None):
    cell.value=value; cell.fill=_fill(GOLD); cell.font=Font(color=INPUT_BLUE); cell.comment=Comment(note or "User-editable analyst AI scenario assumption; not a company disclosure.","OpenAI")
    if fmt: cell.number_format=fmt


def _base_style(ws):
    ws.sheet_view.showGridLines=False; ws.freeze_panes="A5"
    widths={"A":32,"B":16,"C":25,"D":18,"E":48,"F":16,"G":46,"H":18,"I":15,"J":34,"K":15,"L":15,"M":18,"N":20}
    for col,width in widths.items(): ws.column_dimensions[col].width=width

def _title(ws,ticker):
    _merge(ws,"A1:N2"); ws["A1"]=f"{ticker} — AI Impact Analysis"; ws["A1"].fill=_fill(NAVY); ws["A1"].font=Font(bold=True,color=WHITE,size=18)
    _merge(ws,"A3:N3"); ws["A3"]="Institutional-style AI investment lens: reported adoption and monetization, segment exposure, capital intensity, cannibalization risk, and AI surprise scenarios valued relative to the existing Base DCF. The key question is not whether AI grows, but whether outcomes exceed what is already embedded in the Base case and market price."; ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True,vertical="top"); ws.row_dimensions[3].height=40

def _card(ws,tr,vr,title,value=None,formula=None,fmt=FMT_NUM,source=None,internal=False):
    _merge(ws,tr); _merge(ws,vr); t=ws[tr.split(":")[0]]; v=ws[vr.split(":")[0]]
    t.value=title; t.fill=_fill(BLUE); t.font=Font(bold=True,color=WHITE); t.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    if formula: _formula(v,formula,fmt,internal=internal); color=LINK_GREEN if internal else BLACK
    else: _reported(v,value,source or "Analyst input",fmt); color=INPUT_BLUE
    v.fill=_fill(LIGHT); v.font=Font(bold=True,color=color,size=14); v.alignment=Alignment(horizontal="center",vertical="center")

def _cards(ws,ticker):
    t=ticker.upper()
    if t=="AMZN":
        _card(ws,"A6:B6","A7:B8","AWS AI Run-Rate",25.0,fmt=FMT_BN,source=AMZN_Q2); _card(ws,"C6:D6","C7:D8","AWS Q2 YoY Growth",.367,fmt=FMT_PCT,source=AMZN_Q2); _card(ws,"E6:F6","E7:F8","AWS Q2 Op. Margin",16.6/42.2,fmt=FMT_PCT,source=AMZN_Q2); _card(ws,"G6:H6","G7:H8","Chips Run-Rate",25.0,fmt=FMT_BN,source=AMZN_Q2); _card(ws,"I6:J6","I7:J8","TTM Free Cash Flow",-7.6,fmt=FMT_BN,source=AMZN_Q2); _card(ws,"K6:L6","K7:L8","2026 Capex Guide",200.0,fmt=FMT_BN,source=AMZN_Q4); _card(ws,"M6:N6","M7:N8","Current Share Price",formula="='Company Data'!B8",fmt=FMT_PRICE,internal=True)
    elif t in {"GOOGL","GOOG"}:
        _card(ws,"A6:B6","A7:B8","Cloud Q2 Growth",.82,fmt=FMT_PCT,source=GOOG_Q2); _card(ws,"C6:D6","C7:D8","Cloud Backlog",514.0,fmt=FMT_BN,source=GOOG_Q2); _card(ws,"E6:F6","E7:F8","Gemini MAU",950.0,fmt=FMT_NUM,source=GOOG_Q2); _card(ws,"G6:H6","G7:H8","AI Mode MAU",1.0,fmt='0.0"bn"',source=GOOG_Q2); _card(ws,"I6:J6","I7:J8","Base DCF / Share",formula="='DCF'!B9",fmt=FMT_PRICE,internal=True); _card(ws,"K6:L6","K7:L8","Latest Capex / Revenue",formula="='Historical Financials'!G15/'Historical Financials'!G4",fmt=FMT_PCT,internal=True); _card(ws,"M6:N6","M7:N8","Current Share Price",formula="='Company Data'!B8",fmt=FMT_PRICE,internal=True)
    else:
        _card(ws,"A6:B6","A7:B8","Current Share Price",formula="='Company Data'!B8",fmt=FMT_PRICE,internal=True); _card(ws,"C6:D6","C7:D8","Base DCF / Share",formula="='DCF'!B9",fmt=FMT_PRICE,internal=True); _card(ws,"E6:F6","E7:F8","Latest Revenue Growth",formula="='Historical Financials'!G5",fmt=FMT_PCT,internal=True); _card(ws,"G6:H6","G7:H8","Latest Op. Margin",formula="='Historical Financials'!G10",fmt=FMT_PCT,internal=True); _card(ws,"I6:J6","I7:J8","Latest Capex / Revenue",formula="='Historical Financials'!G15/'Historical Financials'!G4",fmt=FMT_PCT,internal=True); _card(ws,"K6:L6","K7:L8","Latest FCF Margin",formula="='Historical Financials'!G17",fmt=FMT_PCT,internal=True); _card(ws,"M6:N6","M7:N8","Forward P/E",formula="='Company Data'!B15",fmt='0.0x',internal=True)

def _evidence(ws,ticker):
    _section(ws,10,"Reported AI Evidence — Separate Demand, Monetization and Capital Burden"); heads=["KPI / Evidence","Current","Unit / Comparison","Signal","Investment Read-Through","As of","Source URL","Data Type"]
    for c,v in enumerate(heads,1): ws.cell(11,c,v)
    _header(ws,11,1,8); rows=EVIDENCE_PACKS.get(ticker.upper())
    if rows:
        for rr,row in enumerate(rows,12):
            for c,v in enumerate(row,1): ws.cell(rr,c,v)
            if isinstance(row[1],(int,float)): _reported(ws.cell(rr,2),row[1],row[6],FMT_BN if abs(row[1])>=2 else FMT_NUM)
            ws.cell(rr,7).font=Font(color=LINK_GREEN)
            for c in (3,4,5,6,7,8): ws.cell(rr,c).alignment=Alignment(wrap_text=True,vertical="top")
            ws.row_dimensions[rr].height=56
        return
    for i,kpi in enumerate(GENERIC_KPIS,12):
        ws.cell(i,1,kpi)
        for c in range(2,9): ws.cell(i,c).fill=_fill(GOLD); ws.cell(i,c).font=Font(color=INPUT_BLUE)
        ws.cell(i,8,"Analyst input"); ws.row_dimensions[i].height=34
    ws["A21"]="No standardized company-specific AI KPI pack is coded for this ticker. Populate from the latest earnings release, 10-Q/10-K and investor presentation."; ws["A21"].font=Font(italic=True,color=GREY); ws["A21"].alignment=Alignment(wrap_text=True)

def _score_defaults(ticker,name):
    rules={("AMZN","north america"):(4,4,5,2,2),("AMZN","international"):(3,3,4,2,2),("AMZN","aws"):(5,5,5,2,5),("GOOGL","google services"):(5,4,5,3,3),("GOOGL","google cloud"):(5,5,5,2,5),("GOOGL","other bets"):(4,2,3,3,4),("GOOG","google services"):(5,4,5,3,3),("GOOG","google cloud"):(5,5,5,2,5),("GOOG","other bets"):(4,2,3,3,4),("MSFT","productivity and business processes"):(5,5,5,2,2),("MSFT","intelligent cloud"):(5,5,5,2,5),("MSFT","more personal computing"):(3,3,4,3,2),("META","family of apps"):(5,4,5,3,4),("META","reality labs"):(4,2,3,4,4),("NVDA","compute & networking"):(5,5,5,2,5),("NVDA","graphics"):(4,4,4,3,3)}
    return rules.get((ticker.upper(),str(name or "").strip().lower()),(3,3,3,3,3))

def _segment_rows(wb):
    if "Segment Analysis" not in wb.sheetnames: return []
    seg=wb["Segment Analysis"]; start=None; out=[]
    for r in range(1,min(seg.max_row,40)+1):
        if str(seg.cell(r,1).value or "").strip().lower()=="reported operating segments": start=r+2; break
    if not start: return []
    for r in range(start,min(seg.max_row,start+12)+1):
        name=seg.cell(r,1).value
        if name in (None,""):
            if out: break
            continue
        out.append((r,str(name)))
    return out[:8]

def _segment_exposure(ws,wb,ticker):
    _section(ws,23,"AI Transmission Channels & Segment Exposure — Analyst Scoring"); heads=["Segment","Latest Revenue Mix","AI Opportunity\n1–5","Monetization\n1–5","Moat / Data /\nDistribution 1–5","Disruption /\nCannibalization Risk 1–5","AI Capex\nIntensity 1–5","Net AI Score\n1–5","Weighted AI\nScore","Analyst Note"]
    for c,v in enumerate(heads,1): ws.cell(24,c,v)
    _header(ws,24,1,10); segs=_segment_rows(wb) or [(None,f"Manual Segment {i}") for i in range(1,6)]; dv=DataValidation(type="list",formula1='"1,2,3,4,5"',allow_blank=False); ws.add_data_validation(dv)
    for out,(src,name) in enumerate(segs,25):
        if src: _formula(ws.cell(out,1),f"='Segment Analysis'!A{src}",internal=True); _formula(ws.cell(out,2),f"='Segment Analysis'!N{src}",FMT_PCT,internal=True)
        else: ws.cell(out,1,name); ws.cell(out,1).fill=_fill(GOLD); ws.cell(out,1).font=Font(color=INPUT_BLUE); ws.cell(out,2).fill=_fill(GOLD); ws.cell(out,2).font=Font(color=INPUT_BLUE); ws.cell(out,2).number_format=FMT_PCT
        for j,val in enumerate(_score_defaults(ticker,name),3): _assumption(ws.cell(out,j),val,"0","Analyst scoring input (1=low, 5=high). This is not a company disclosure."); dv.add(ws.cell(out,j))
        _formula(ws.cell(out,8),f"=C{out}*30%+D{out}*25%+E{out}*25%+(6-F{out})*10%+(6-G{out})*10%",FMT_SCORE); _formula(ws.cell(out,9),f"=B{out}*H{out}",FMT_SCORE); ws.cell(out,10,"Update scores and note after each earnings cycle."); ws.cell(out,10).alignment=Alignment(wrap_text=True,vertical="top"); ws.row_dimensions[out].height=42
    last=24+len(segs); ws.conditional_formatting.add(f"H25:H{last}",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="num",mid_value=3,end_type="max",end_color="63BE7B")); total=last+2; ws.cell(total,1,"Revenue-weighted AI quality score (/5)"); ws.cell(total,1).font=Font(bold=True); _formula(ws.cell(total,8),f"=SUM(I25:I{last})",FMT_SCORE); ws.cell(total+1,1,"AI opportunity intensity (rev-weighted, % of max)"); ws.cell(total+1,1).font=Font(bold=True); _formula(ws.cell(total+1,8),f"=SUMPRODUCT(B25:B{last},C25:C{last})/5",FMT_PCT)
    for r in (total,total+1):
        for c in range(1,9): ws.cell(r,c).fill=_fill(LIGHT)

def _assumptions(ws):
    _section(ws,33,"AI Surprise Scenarios vs Existing Base Case — Avoid Double Counting"); heads=["Assumption","Bear AI Outcome","Base / Embedded","Bull AI Outcome","Interpretation","Status"]
    for c,v in enumerate(heads,1): ws.cell(34,c,v)
    _header(ws,34,1,6); rows=[("Probability",.25,.5,.25,"Scenario probabilities; should sum to 100%."),("Revenue growth Δ vs Base",-.02,0,.02,"Annual consolidated growth surprise from weaker/stronger AI monetization and second-order effects."),("EBIT margin Δ vs Base",-.02,0,.015,"Inference cost, pricing, labor productivity, ads/conversion and mix effects."),("Capex / revenue Δ vs Base",.03,0,-.01,"Downside assumes heavier AI capital burden; upside assumes better utilization/capital efficiency."),("WACC Δ vs Base",.005,0,-.0025,"Optional risk-premium adjustment for higher/lower uncertainty and durability."),("Terminal growth Δ vs Base",-.0025,0,.0025,"Long-run AI effect on growth durability; keep conservative.")]
    for i,(label,b,c,d,note) in enumerate(rows,35):
        ws.cell(i,1,label)
        for col,val in zip((2,3,4),(b,c,d)): _assumption(ws.cell(i,col),val,FMT_PCT)
        ws.cell(i,5,note); ws.cell(i,6,"Analyst input"); ws.cell(i,5).alignment=Alignment(wrap_text=True,vertical="top"); ws.row_dimensions[i].height=38
    _merge(ws,"G35:N40"); ws["G35"]="Model consistency: scenario FCF retains the Base model's 'Other cash conversion % revenue' calibration from Three-Case Scenarios!AL4. This preserves the authoritative Base cash-flow logic and keeps the 0Δ Base / Embedded AI case reconciled to DCF!B9."; ws["G35"].fill=_fill(LIGHT); ws["G35"].font=Font(italic=True,color=GREY); ws["G35"].alignment=Alignment(wrap_text=True,vertical="top")
    ws["A42"]="Control principle"; ws["E42"]="Base is deliberately 0Δ. Bear/Bull measure AI-related surprise around the current forecast—not total AI value."; ws["F42"]="Model design"; ws["E42"].alignment=Alignment(wrap_text=True)
    for c in range(1,7): ws.cell(42,c).fill=_fill(LIGHT)

def _scenario_block(ws,start,name,acol):
    _merge(ws,f"A{start}:N{start}"); ws.cell(start,1,f"{name} AI Outcome — 10-Year Scenario DCF"); ws.cell(start,1).fill=_fill(PALE); ws.cell(start,1).font=Font(bold=True,color=NAVY); years=list(range(2026,2036))
    for c,v in enumerate(["Metric"]+years,1): ws.cell(start+1,c,v)
    _header(ws,start+1,1,11); labels=["Revenue Growth","Revenue","EBIT Margin","EBIT","Tax Rate","NOPAT","D&A % Revenue","D&A","Capex % Revenue","Capex","NWC Investment","FCF","Discount Factor","PV of FCF"]
    for r,lab in enumerate(labels,start+2): ws.cell(r,1,lab)
    base_cols=["N","O","P","Q","R","S","T","U","V","W"]; rg=start+2; rev=start+3; margin=start+4; ebit=start+5; tax=start+6; nopat=start+7; dap=start+8; da=start+9; capp=start+10; cap=start+11; nwc=start+12; fcf=start+13; df=start+14; pv=start+15
    for j,(c,bc) in enumerate(zip(range(2,12),base_cols)):
        L=get_column_letter(c); _formula(ws.cell(rg,c),f"='Three-Case Scenarios'!{bc}12+${acol}$36",FMT_PCT,True)
        if c==2: _formula(ws.cell(rev,c),f"='Historical Financials'!G4*(1+{L}{rg})",FMT_NUM)
        else: P=get_column_letter(c-1); _formula(ws.cell(rev,c),f"={P}{rev}*(1+{L}{rg})",FMT_NUM)
        _formula(ws.cell(margin,c),f"='Three-Case Scenarios'!{bc}14+${acol}$37",FMT_PCT,True); _formula(ws.cell(ebit,c),f"={L}{rev}*{L}{margin}",FMT_NUM); _formula(ws.cell(tax,c),"='Three-Case Scenarios'!$C$8",FMT_PCT,True); _formula(ws.cell(nopat,c),f"={L}{ebit}*(1-{L}{tax})",FMT_NUM); _formula(ws.cell(dap,c),f"='Three-Case Scenarios'!{bc}18",FMT_PCT,True); _formula(ws.cell(da,c),f"={L}{rev}*{L}{dap}",FMT_NUM); _formula(ws.cell(capp,c),f"='Three-Case Scenarios'!{bc}20+${acol}$38",FMT_PCT,True); _formula(ws.cell(cap,c),f"={L}{rev}*{L}{capp}",FMT_NUM)
        if c==2: _formula(ws.cell(nwc,c),f"=1%*({L}{rev}-'Historical Financials'!G4)",FMT_NUM)
        else: P=get_column_letter(c-1); _formula(ws.cell(nwc,c),f"=1%*({L}{rev}-{P}{rev})",FMT_NUM)
        _formula(ws.cell(fcf,c),f"={L}{nopat}+{L}{da}-{L}{cap}-{L}{nwc}+{L}{rev}*'Three-Case Scenarios'!$AL$4",FMT_NUM); _formula(ws.cell(df,c),f"=1/(1+'Three-Case Scenarios'!$C$6+${acol}$39)^{j+1}",FMT_DF); _formula(ws.cell(pv,c),f"={L}{fcf}*{L}{df}",FMT_NUM)
    return {"fcf":fcf,"df":df,"pv":pv}

def _valuation(ws,bear,base,bull):
    _section(ws,97,"AI Scenario Valuation Summary — What AI Outcomes Would Need to Do to the Existing DCF"); heads=["Output","Bear","Base / Embedded","Bull","Probability Weighted","Current / Control"]
    for c,v in enumerate(heads,1): ws.cell(98,c,v)
    _header(ws,98,1,6); labels=["PV Explicit FCF","Terminal FCF","Terminal Value","PV Terminal Value","Enterprise Value","Net Debt / (Cash)","Equity Value","Shares Outstanding (bn)","Intrinsic Value / Share","Value vs Existing Base DCF","Upside / (Downside) to Current Price"]
    for r,lab in enumerate(labels,99): ws.cell(r,1,lab)
    for outcol,acol,blk in [("B","B",bear),("C","C",base),("D","D",bull)]:
        _formula(ws[f"{outcol}99"],f"=SUM(B{blk['pv']}:K{blk['pv']})",FMT_NUM); _formula(ws[f"{outcol}100"],f"=K{blk['fcf']}*(1+'Three-Case Scenarios'!$C$7+${acol}$40)",FMT_NUM); _formula(ws[f"{outcol}101"],f"={outcol}100/(('Three-Case Scenarios'!$C$6+${acol}$39)-('Three-Case Scenarios'!$C$7+${acol}$40))",FMT_NUM); _formula(ws[f"{outcol}102"],f"={outcol}101*K{blk['df']}",FMT_NUM); _formula(ws[f"{outcol}103"],f"={outcol}99+{outcol}102",FMT_NUM); _formula(ws[f"{outcol}104"],"='DCF'!$B$8",FMT_NUM,True); _formula(ws[f"{outcol}105"],f"={outcol}103-{outcol}104",FMT_NUM); _formula(ws[f"{outcol}106"],"='DCF'!$B$5",FMT_NUM,True); _formula(ws[f"{outcol}107"],f"={outcol}105/{outcol}106",FMT_PRICE); _formula(ws[f"{outcol}108"],f"={outcol}107/'DCF'!$B$9-1",FMT_PCT); _formula(ws[f"{outcol}109"],f"={outcol}107/'Company Data'!$B$8-1",FMT_PCT)
    for r in range(99,110): _formula(ws[f"E{r}"],f"=B{r}*$B$35+C{r}*$C$35+D{r}*$D$35",FMT_PRICE if r==107 else (FMT_PCT if r>=108 else FMT_NUM))
    ws["F99"]="Existing Base DCF"; _formula(ws["F107"],"='DCF'!$B$9",FMT_PRICE,True); ws["F108"]=0; ws["F108"].number_format=FMT_PCT; _formula(ws["F109"],"='DCF'!$B$9/'Company Data'!$B$8-1",FMT_PCT,True)
    for r in range(107,110):
        for c in range(1,7): ws.cell(r,c).fill=_fill(LIGHT)
    ws.conditional_formatting.add("B107:D107",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="percentile",mid_value=50,end_type="max",end_color="63BE7B"))
    ws["M122"]="Scenario"; ws["N122"]="Value / Share"
    for c in (13,14): ws.cell(122,c).fill=_fill(BLUE); ws.cell(122,c).font=Font(bold=True,color=WHITE)
    for r,(name,formula) in enumerate([("Bear","=B107"),("Base","=C107"),("Bull","=D107"),("Current Price","='Company Data'!B8")],123): ws.cell(r,13,name); _formula(ws.cell(r,14),formula,FMT_PRICE,True)
    ch=BarChart(); ch.type="col"; ch.style=10; ch.title="AI Scenario Value / Share vs Current Price"; ch.height=7.5; ch.width=12.5; ch.legend=None; ch.add_data(Reference(ws,min_col=14,min_row=122,max_row=126),titles_from_data=True); ch.set_categories(Reference(ws,min_col=13,min_row=123,max_row=126)); ch.visible_cells_only=False; ch.display_blanks="gap"; ch.dLbls=DataLabelList(); ch.dLbls.showVal=True; ws.add_chart(ch,"H98")

def _finish(ws,ticker):
    _section(ws,112,"Professional Interpretation — AI Does Not Equal Automatic Upside"); table=[["Question","What to Monitor","Bullish Confirmation","Bearish Warning"],["Is AI creating incremental revenue?","AI revenue, attach rates, seats/tokens, backlog, conversion","AI revenue outgrows infrastructure spend and broadens across products","Usage rises but paid monetization or revenue attribution stalls"],["Are unit economics improving?","AI gross/contribution margin, inference cost, pricing, utilization","Margins hold/expand while AI volumes scale","AI revenue growth requires structurally lower margins"],["Is capex earning its cost of capital?","FCF recovery, utilization, incremental ROIC","FCF inflects before AI demand slows","Capex stays elevated while growth/backlog decelerates"],["Does AI strengthen the moat?","Customer lock-in, data/distribution, silicon, ecosystem, switching costs","Durable share gains and lower unit costs","Commoditization shifts value to suppliers or lowers pricing power"],["Is the market already pricing the bull case?","Current price vs Base/Bull AI scenario value","Price leaves room for upside without heroic assumptions","Valuation requires bull growth + margin + capital efficiency simultaneously"]]
    for r,row in enumerate(table,113):
        for c,v in enumerate(row,1): ws.cell(r,c,v); ws.cell(r,c).alignment=Alignment(wrap_text=True,vertical="top")
    _header(ws,113,1,4)
    for r in range(114,119): ws.row_dimensions[r].height=44
    _section(ws,121,"Sources & Methodology"); sources=[]
    for row in EVIDENCE_PACKS.get(ticker.upper(),[]):
        if row[6] and row[6] not in [u for _,u in sources]: sources.append((row[0],row[6]))
    if not sources: sources=[("Company filings / investor relations","Populate latest issuer source URLs in Reported AI Evidence")]
    sources.extend([("Existing Base DCF","Internal workbook: DCF + Three-Case Scenarios"),("Segment exposure","Internal workbook: Segment Analysis")])
    for r,(name,url) in enumerate(sources,122): ws.cell(r,1,name); ws.cell(r,2,url); ws.cell(r,2).font=Font(color=LINK_GREEN); ws.cell(r,2).alignment=Alignment(wrap_text=True)
    note=122+len(sources)+2; _merge(ws,f"A{note}:N{note+2}"); ws.cell(note,1,"Methodology note: reported AI KPIs may be management-defined and may not map directly to GAAP revenue. The scenario DCF measures AI-related surprise versus the existing Base forecast, reducing double counting of AI already embedded in revenue, margins and capex assumptions. Segment scores are analyst judgments, not company disclosures, and should be updated after each earnings cycle."); ws.cell(note,1).fill=_fill(LIGHT); ws.cell(note,1).font=Font(italic=True,color=GREY); ws.cell(note,1).alignment=Alignment(wrap_text=True,vertical="top")

def _documentation(wb):
    if "Data Dictionary" in wb.sheetnames:
        ws=wb["Data Dictionary"]; row=None
        for r in range(4,ws.max_row+2):
            if str(ws.cell(r,1).value or "").strip()=="AI Impact Analysis": row=r; break
            if row is None and ws.cell(r,1).value in (None,""): row=r
        row=row or ws.max_row+1; ws.cell(row,1,"AI Impact Analysis"); ws.cell(row,2,"Company-reported AI KPIs + Segment Analysis + Base scenario / DCF + analyst-editable AI surprise assumptions"); ws.cell(row,3,"Separates AI demand, monetization, capex and disruption risk; values AI outcomes without double counting the Base case")
    if "Research Checklist" in wb.sheetnames:
        ws=wb["Research Checklist"]; existing={str(ws.cell(r,2).value or "").strip() for r in range(1,ws.max_row+1)}
        for item in ["AI monetization vs capex / FCF burden assessed","AI disruption / cannibalization and scenario valuation reviewed"]:
            if item not in existing:
                r=ws.max_row+1; ws.cell(r,1,"Technology / AI"); ws.cell(r,2,item); ws.cell(r,3,"Not Started"); ws.cell(r,5,"AI Impact Analysis")

def ensure_ai_impact_analysis(wb,ticker):
    ticker=ticker.upper()
    if "AI Impact Analysis" in wb.sheetnames: wb.remove(wb["AI Impact Analysis"])
    ws=wb.create_sheet("AI Impact Analysis"); _base_style(ws); _title(ws,ticker); _section(ws,5,"Executive AI Investment Lens — Latest Reported Evidence"); _cards(ws,ticker); _evidence(ws,ticker); _segment_exposure(ws,wb,ticker); _assumptions(ws)
    bear=_scenario_block(ws,45,"Bear","B"); base=_scenario_block(ws,62,"Base / Embedded","C"); bull=_scenario_block(ws,79,"Bull","D"); _valuation(ws,bear,base,bull); _finish(ws,ticker); _documentation(wb); return ws
