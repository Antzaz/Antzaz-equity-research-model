from __future__ import annotations

from pathlib import Path
from typing import Iterable
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

from .common import MLResult

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; GREY="666666"; LIGHT="F5F9FC"; GOLD="FFF2CC"; RED="FCE4D6"; GREEN="E2F0D9"
THIN=Side(style="thin",color="D9E1F2")

FEATURE_LABELS={
    "revenue_growth":"Revenue growth",
    "operating_margin":"Operating profit margin",
    "net_margin":"Net profit margin",
    "fcf_margin":"Free-cash-flow margin",
    "capex_to_revenue":"Capital spending / revenue",
    "rd_to_revenue":"R&D / revenue",
    "sbc_to_revenue":"Stock compensation / revenue",
    "roe":"Return on equity",
    "net_debt_to_revenue":"Net debt / revenue",
    "earnings_yield":"Earnings yield",
    "fcf_yield":"Free-cash-flow yield",
    "book_to_market":"Book value / market value",
    "ev_to_sales":"Enterprise value / sales",
    "momentum_12m":"12-month price trend",
    "momentum_6m":"6-month price trend",
    "price_momentum_3m":"3-month price trend",
    "price_vol_3m":"3-month price volatility",
    "volatility_6m":"6-month volatility",
    "drawdown_12m":"Recent drawdown",
    "prior_surprise_1q":"Last earnings surprise",
    "prior_surprise_2q_avg":"Average surprise, last 2 quarters",
    "prior_surprise_4q_avg":"Average surprise, last 4 quarters",
    "surprise_vol_4q":"Variability of recent surprises",
    "eps_estimate":"Current EPS estimate",
}


def _label(value):
    return FEATURE_LABELS.get(str(value),str(value).replace("_"," ").title())

def _fill(c): return PatternFill("solid",fgColor=c)
def _header(ws,row,headers):
    for c,v in enumerate(headers,1):
        x=ws.cell(row,c,v); x.fill=_fill(BLUE); x.font=Font(bold=True,color=WHITE); x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); x.border=Border(bottom=THIN)
def _section(ws,row,title,end=9):
    for c in range(1,end+1): ws.cell(row,c).fill=_fill(NAVY); ws.cell(row,c).font=Font(bold=True,color=WHITE,size=11)
    ws.cell(row,1,title)
def _wf(res):
    m=res.metrics or {}
    return m.get("walk_forward") or m.get("hgb_walk_forward") or {}
def _pct(v):
    try: return float(v)
    except Exception: return None


def _validation_summary(res: MLResult) -> str:
    """Compact validation copy written for a non-technical reader first."""
    wf=_wf(res)
    parts=[]
    n=wf.get("n")
    da=_pct(wf.get("directional_accuracy"))
    base_da=_pct(wf.get("baseline_directional_accuracy"))
    r2=_pct(wf.get("r2"))
    mae=_pct(wf.get("mae"))
    base_mae=_pct(wf.get("baseline_mae"))
    if n is not None: parts.append(f"{int(n):,} walk-forward tests")
    if da is not None:
        parts.append(f"direction {da:.1%}" + (f" vs baseline {base_da:.1%}" if base_da is not None else ""))
    if mae is not None:
        parts.append(f"typical error {mae:.1%}" + (f" vs baseline {base_mae:.1%}" if base_mae is not None else ""))
    if r2 is not None: parts.append(f"R² {r2:.2f}")
    if parts: return " | ".join(parts)

    metrics=res.metrics or {}
    if res.name=="Financial Anomaly Detection":
        years=metrics.get("years") or []
        return f"{len(years)} annual observations" if years else "No robust annual validation sample"
    if res.name=="Market Regime Classifier":
        probs=metrics.get("regime_probabilities") or {}
        total=sum((_pct(v) or 0) for v in probs.values())
        top=max([_pct(v) or 0 for v in probs.values()] or [0])
        rows=int(metrics.get("monthly_rows") or 0)
        suffix=f" | weights total {total:.1%}" if total and abs(total-1.0)>.01 else ""
        return f"{rows:,} months | strongest regime weight {top:.1%}{suffix}" if rows else f"Strongest regime weight {top:.1%}{suffix}"
    return "See detail section"


def _how_to_read_box(ws):
    ws.merge_cells("K2:N2"); ws["K2"]="How to read the ML page"
    ws["K2"].fill=_fill(NAVY); ws["K2"].font=Font(bold=True,color=WHITE)
    ws.merge_cells("K3:N6")
    ws["K3"]=(
        "PASS = usable second-opinion evidence. REVIEW = context only. INS UFFICIENT_DATA = do not infer a signal. "
        "A forecast is useful only if its historical error is small enough and it beats a simple historical baseline. "
        "Driver charts show what the model pays attention to, not what caused the stock return."
    ).replace("INS UFFICIENT_DATA","INSUFFICIENT_DATA")
    ws["K3"].alignment=Alignment(wrap_text=True,vertical="top")
    ws["K3"].fill=_fill(LIGHT)


def _add_ml_charts(ws, result_list):
    """Add plain-English charts. Helper tables live in hidden columns X:AA."""
    for col in ("X","Y","Z","AA"):
        ws.column_dimensions[col].hidden=True
    _how_to_read_box(ws)

    # 1) Forecast magnitude versus walk-forward typical error.
    start=2
    ws.cell(start,24,"Model"); ws.cell(start,25,"Forecast magnitude"); ws.cell(start,26,"Typical historical error")
    rr=start+1
    for res in result_list:
        wf=_wf(res); mae=_pct(wf.get("mae")); pred=_pct(res.prediction) if isinstance(res.prediction,(int,float)) else None
        if mae is None or pred is None or res.name not in {"Expected 12M Excess Return","Consensus / Earnings Surprise"}:
            continue
        name="12M excess return" if res.name=="Expected 12M Excess Return" else "Next earnings surprise"
        ws.cell(rr,24,name); ws.cell(rr,25,abs(pred)*100); ws.cell(rr,26,mae*100); rr+=1
    if rr>start+1:
        ch=BarChart(); ch.type="col"; ch.style=10
        ch.title="Forecast size vs typical historical error"
        ch.y_axis.title="Percent (%)"; ch.y_axis.numFmt="0.0"
        ch.height=7.0; ch.width=13.0
        ch.add_data(Reference(ws,min_col=25,max_col=26,min_row=start,max_row=rr-1),titles_from_data=True)
        ch.set_categories(Reference(ws,min_col=24,min_row=start+1,max_row=rr-1))
        ch.legend.position="b"
        ch.dLbls=DataLabelList(); ch.dLbls.showVal=True; ch.dLbls.numFmt="0.0"
        ws.add_chart(ch,"K8")

    # 2) Directional accuracy versus a leakage-safe simple historical baseline.
    start=9
    ws.cell(start,24,"Model"); ws.cell(start,25,"Model accuracy"); ws.cell(start,26,"Simple baseline")
    rr=start+1
    for res in result_list:
        wf=_wf(res); da=_pct(wf.get("directional_accuracy")); base=_pct(wf.get("baseline_directional_accuracy"))
        if da is None: continue
        name="12M excess return" if res.name=="Expected 12M Excess Return" else "Next earnings surprise" if res.name=="Consensus / Earnings Surprise" else None
        if not name: continue
        ws.cell(rr,24,name); ws.cell(rr,25,da*100); ws.cell(rr,26,(base if base is not None else .50)*100); rr+=1
    if rr>start+1:
        ch=BarChart(); ch.type="col"; ch.style=11
        ch.title="Did the model beat a simple historical rule?"
        ch.y_axis.title="Directional accuracy (%)"; ch.y_axis.scaling.min=0; ch.y_axis.scaling.max=100; ch.y_axis.numFmt="0.0"
        ch.height=7.0; ch.width=13.0
        ch.add_data(Reference(ws,min_col=25,max_col=26,min_row=start,max_row=rr-1),titles_from_data=True)
        ch.set_categories(Reference(ws,min_col=24,min_row=start+1,max_row=rr-1))
        ch.legend.position="b"
        ch.dLbls=DataLabelList(); ch.dLbls.showVal=True; ch.dLbls.numFmt="0.0"
        ws.add_chart(ch,"K24")

    # 3) Expected-return model drivers, normalized so the chart is intuitive.
    expected=next((r for r in result_list if r.name=="Expected 12M Excess Return" and r.drivers),None)
    if expected:
        s=16; ws.cell(s,24,"Feature"); ws.cell(s,25,"Share of top-driver influence")
        drivers=expected.drivers[:6]
        raw=[abs(_pct(d.get("importance")) or 0) for d in drivers]
        total=sum(raw)
        shares=[v/total if total>0 else 0 for v in raw]
        for i,(d,share) in enumerate(zip(drivers,shares),s+1):
            ws.cell(i,24,_label(d.get("feature"))); ws.cell(i,25,share*100); ws.cell(i,25).number_format="0.0"
        ch=BarChart(); ch.type="bar"; ch.style=12
        ch.title="What the 12M return model pays attention to"
        ch.y_axis.title="Input"; ch.x_axis.title="Share of top-driver influence (%)"; ch.x_axis.numFmt="0.0"
        ch.height=7.5; ch.width=13.0
        ch.add_data(Reference(ws,min_col=25,min_row=s,max_row=s+len(drivers)),titles_from_data=True)
        ch.set_categories(Reference(ws,min_col=24,min_row=s+1,max_row=s+len(drivers))); ch.legend=None
        ch.dLbls=DataLabelList(); ch.dLbls.showVal=True; ch.dLbls.numFmt="0.0"
        ws.add_chart(ch,"K40")

    # 4) Market-regime weights. If an old run lost a duplicate label, expose the residual instead of hiding it.
    regime=next((r for r in result_list if r.name=="Market Regime Classifier"),None)
    probs=(regime.metrics or {}).get("regime_probabilities") if regime else None
    if probs:
        items=sorted(((str(k),max(0.0,_pct(v) or 0.0)) for k,v in probs.items()),key=lambda x:x[1],reverse=True)
        total=sum(v for _,v in items)
        if total<.99:
            items.append(("Unassigned / duplicate cluster",max(0.0,1.0-total)))
        elif total>1.01 and total>0:
            items=[(k,v/total) for k,v in items]
        s=25; ws.cell(s,24,"Market state"); ws.cell(s,25,"Weight")
        for i,(name,val) in enumerate(items,s+1):
            ws.cell(i,24,name); ws.cell(i,25,val*100); ws.cell(i,25).number_format='0.0'
        ch=BarChart(); ch.type="bar"; ch.style=10
        ch.title="What kind of market does the model think we are in?"
        ch.y_axis.title="Market state"; ch.x_axis.title="Distance-based weight (%)"
        ch.x_axis.scaling.min=0; ch.x_axis.scaling.max=100; ch.x_axis.numFmt="0.0"
        ch.height=7.0; ch.width=13.0
        ch.add_data(Reference(ws,min_col=25,min_row=s,max_row=s+len(items)),titles_from_data=True)
        ch.set_categories(Reference(ws,min_col=24,min_row=s+1,max_row=s+len(items))); ch.legend=None
        ch.dLbls=DataLabelList(); ch.dLbls.showVal=True; ch.dLbls.numFmt="0.0"
        ws.add_chart(ch,"K58")


def write_ml_sheet(workbook_path: Path, ticker: str, results: Iterable[MLResult]) -> Path:
    wb=load_workbook(workbook_path)
    if "ML & Quantitative Research" in wb.sheetnames: wb.remove(wb["ML & Quantitative Research"])
    ws=wb.create_sheet("ML & Quantitative Research")
    ws.sheet_view.showGridLines=False; ws.freeze_panes="A6"
    for c in range(1,10): ws.cell(1,c).fill=_fill(NAVY); ws.cell(2,c).fill=_fill(NAVY)
    ws["A1"]=f"{ticker} — Machine Learning & Quantitative Research"; ws["A1"].font=Font(bold=True,color=WHITE,size=18)
    ws.merge_cells("A3:I3"); ws["A3"]="ML is a second-opinion layer. The page emphasizes whether each model actually worked out of sample, not just whether it produced a number. ML never overwrites DCF assumptions, reported financials or thesis decisions."; ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True)
    _section(ws,5,"Six-Model Decision Dashboard",9)
    _header(ws,6,["Model","Evidence status","What the model says","Confidence","How much evidence","Main influences","How to use it","Important limitation","Decision rule"])
    row=7
    result_list=list(results)
    for res in result_list:
        drivers=", ".join(_label(d.get("feature")) for d in (res.drivers or [])[:4])
        use={
            "Expected 12M Excess Return":"Cross-check valuation only when walk-forward skill clearly beats a simple baseline.",
            "Consensus / Earnings Surprise":"Challenge near-term consensus only if the model adds value over the company's normal beat/miss pattern.",
            "Financial Anomaly Detection":"Flag unusual accounting/operating patterns for diligence; never proof of wrongdoing.",
            "Market Regime Classifier":"Describe the current macro backdrop; not a timing signal.",
            "AI Impact ML":"Test whether accumulating AI KPIs predict monetization/economic outcomes.",
            "Portfolio ML / Position Sizing":"Translate validated expected-return evidence into constrained risk-aware sizing.",
        }.get(res.name,"Research cross-check")
        usable=res.status=="PASS"
        values=[
            res.name,res.status,res.prediction,res.confidence,_validation_summary(res),drivers,use,
            "Historical relationships can fail out of sample. A high hit-rate is not enough unless it beats a simple baseline.",
            "USE AS SECOND OPINION" if usable else "DO NOT USE IN SCORE / REVIEW",
        ]
        for c,v in enumerate(values,1): ws.cell(row,c,v); ws.cell(row,c).alignment=Alignment(wrap_text=True,vertical="top")
        if isinstance(res.prediction,float): ws.cell(row,3).number_format='0.0%;[Red](0.0%);-'
        ws.cell(row,2).fill=_fill(GREEN if res.status=="PASS" else GOLD if res.status in {"REVIEW","WEAK_SIGNAL"} else RED); ws.cell(row,2).font=Font(bold=True)
        ws.row_dimensions[row].height=64; row+=1

    for res in result_list:
        row+=1; _section(ws,row,res.name,9); row+=1
        ws.cell(row,1,"Summary"); ws.cell(row,2,res.summary); ws.merge_cells(start_row=row,start_column=2,end_row=row,end_column=9); ws.cell(row,2).alignment=Alignment(wrap_text=True); row+=1
        notes=(res.details or {}).get("quality_gate_notes") or []
        if notes:
            ws.cell(row,1,"Evidence gate"); ws.cell(row,2," ".join(notes)); ws.merge_cells(start_row=row,start_column=2,end_row=row,end_column=9); ws.cell(row,2).alignment=Alignment(wrap_text=True); row+=1
        if res.drivers:
            _header(ws,row,["Driver","Model influence","Plain-English note"]); row+=1
            for d in res.drivers[:8]:
                key=d.get("feature"); val=d.get("importance",d.get("coefficient",d.get("standardized_deviation")))
                ws.cell(row,1,_label(key)); ws.cell(row,2,val); ws.cell(row,3,"Relative model signal; it does not prove that this factor caused the outcome."); row+=1
        if res.name=="Portfolio ML / Position Sizing" and (res.details or {}).get("weights"):
            _header(ws,row,["Ticker","Suggested Weight","Current Weight","Change","Expected Return Input"]); row+=1
            for rec in res.details["weights"]:
                for c,key in enumerate(["ticker","suggested_weight","current_weight","weight_change","expected_return_input"],1): ws.cell(row,c,rec.get(key))
                for c in range(2,6): ws.cell(row,c).number_format='0.0%;[Red](0.0%);-'
                row+=1

    _add_ml_charts(ws,result_list)
    widths={"A":32,"B":24,"C":22,"D":15,"E":34,"F":34,"G":42,"H":42,"I":26,"J":3,"K":22,"L":22,"M":22,"N":22}
    for col,w in widths.items(): ws.column_dimensions[col].width=w
    wb.save(workbook_path)
    return workbook_path
