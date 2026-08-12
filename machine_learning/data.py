from __future__ import annotations

from pathlib import Path
from typing import Iterable
import json
import re

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook

from .common import num, max_drawdown_from_prices

DEFAULT_UNIVERSE = [
    "AAPL","MSFT","GOOGL","AMZN","META","NVDA","AVGO","TSM","AMD","QCOM",
    "JPM","BAC","V","MA","UNH","LLY","JNJ","XOM","CVX","CAT","GE","COST","WMT","HD","NEE",
]
BENCHMARK = "SPY"


def latest_workbook(repo_root: Path, ticker: str) -> Path:
    files = sorted((repo_root / "updated_models").glob(f"{ticker}_Equity_Research_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"No generated workbook found for {ticker}")
    return files[0]


def peer_tickers_from_workbook(path: Path, target: str) -> list[str]:
    """Read only ticker-like symbols; explanatory Peer Comps notes are never market-data inputs."""
    wb = load_workbook(path, data_only=True, read_only=True)
    if "Peer Comps" not in wb.sheetnames:
        return [target]
    ws = wb["Peer Comps"]
    out = []
    valid = re.compile(r"^[A-Z0-9][A-Z0-9.\-]{0,11}$")
    for r in range(4, min(ws.max_row, 40) + 1):
        sym = str(ws.cell(r, 2).value or "").strip().upper()
        if not valid.fullmatch(sym):
            continue
        if sym not in out:
            out.append(sym)
    t = str(target).upper().strip()
    if valid.fullmatch(t) and t not in out:
        out.insert(0, t)
    return out or [target]


def _latest_history_columns(ws):
    cols=[c for c in range(2,min(ws.max_column,8)+1) if isinstance(ws.cell(3,c).value,(int,float))]
    return cols[-1] if cols else None, cols[-2] if len(cols)>=2 else None


def workbook_current_snapshot(path: Path, ticker: str) -> dict[str, float | str | None]:
    wb = load_workbook(path, data_only=True, read_only=True)
    out: dict[str, float | str | None] = {"ticker": ticker}
    if "Historical Financials" in wb.sheetnames:
        h = wb["Historical Financials"]
        c,pc=_latest_history_columns(h)
        rev = num(h.cell(4,c).value) if c else None; op = num(h.cell(9,c).value) if c else None; ni = num(h.cell(11,c).value) if c else None
        ocf = num(h.cell(14,c).value) if c else None; cap = num(h.cell(15,c).value) if c else None; rd = num(h.cell(19,c).value) if c else None; sbc = num(h.cell(21,c).value) if c else None
        prior_rev = num(h.cell(4,pc).value) if pc else None
        out.update({
            "revenue_growth": (rev / prior_rev - 1) if rev and prior_rev else None,
            "operating_margin": op / rev if rev and op is not None else None,
            "net_margin": ni / rev if rev and ni is not None else None,
            "fcf_margin": (ocf - abs(cap)) / rev if rev and ocf is not None and cap is not None else None,
            "capex_to_revenue": abs(cap) / rev if rev and cap is not None else None,
            "rd_to_revenue": rd / rev if rev and rd is not None else None,
            "sbc_to_revenue": sbc / rev if rev and sbc is not None else None,
        })
    if "Company Data" in wb.sheetnames:
        d = wb["Company Data"]
        out["forward_pe"] = num(d["B15"].value)
        mc = num(d["B10"].value); net_debt = num(d["B14"].value)
        out["net_debt_to_market_cap"] = net_debt / mc if mc and net_debt is not None else None
        if "Historical Financials" in wb.sheetnames:
            h=wb["Historical Financials"]; c,_=_latest_history_columns(h); rev=num(h.cell(4,c).value) if c else None
            out["net_debt_to_revenue"] = net_debt / rev if rev and net_debt is not None else None
    if "Peer Comps" in wb.sheetnames:
        peer = wb["Peer Comps"]
        for r in range(4, min(peer.max_row, 30) + 1):
            if str(peer.cell(r, 2).value or "").strip().upper() == ticker.upper():
                out["roe"] = num(peer.cell(r, 8).value)
                break
    return out


def _series_value(df: pd.DataFrame, labels: list[str], col) -> float | None:
    if df is None or df.empty or col not in df.columns:
        return None
    for label in labels:
        if label in df.index:
            return num(df.at[label, col])
    return None


def annual_fundamental_rows(ticker: str, benchmark_prices: pd.Series | None = None) -> list[dict]:
    try:
        t = yf.Ticker(ticker)
        inc = t.income_stmt
        cf = t.cashflow
        bs = t.balance_sheet
        hist = t.history(period="10y", auto_adjust=True)
    except Exception:
        return []
    if inc is None or inc.empty or hist is None or hist.empty:
        return []
    close = hist["Close"].dropna()
    close.index = pd.to_datetime(close.index, utc=True).tz_convert(None)
    rows = []
    cols = sorted(pd.to_datetime(inc.columns))
    for i, col in enumerate(cols):
        rev = _series_value(inc,["Total Revenue","Operating Revenue"],col)
        op = _series_value(inc,["Operating Income"],col)
        ni = _series_value(inc,["Net Income","Net Income Common Stockholders"],col)
        equity = _series_value(bs,["Stockholders Equity","Total Equity Gross Minority Interest"],col)
        cash = _series_value(bs,["Cash Cash Equivalents And Short Term Investments","Cash And Cash Equivalents"],col)
        debt = _series_value(bs,["Total Debt"],col)
        ocf = _series_value(cf,["Operating Cash Flow","Total Cash From Operating Activities"],col)
        capex = _series_value(cf,["Capital Expenditure","Capital Expenditures"],col)
        rd = _series_value(inc,["Research And Development"],col)
        if capex is not None:
            capex = abs(capex)
        if rev in (None, 0):
            continue
        prev_rev = None
        if i > 0:
            prev_col = cols[i-1]
            prev_rev = _series_value(inc,["Total Revenue","Operating Revenue"],prev_col)
        as_of = pd.Timestamp(col).tz_localize(None) + pd.Timedelta(days=120)
        start = close.index.searchsorted(as_of)
        if start >= len(close):
            continue
        p0 = float(close.iloc[start])
        target_date = as_of + pd.DateOffset(years=1)
        end = close.index.searchsorted(target_date)
        if end >= len(close):
            continue
        p1 = float(close.iloc[end])
        stock_ret = p1 / p0 - 1.0
        bench_ret = 0.0
        if benchmark_prices is not None and not benchmark_prices.empty:
            bp = benchmark_prices.dropna().copy()
            bp.index = pd.to_datetime(bp.index, utc=True).tz_convert(None)
            b0i = bp.index.searchsorted(as_of); b1i = bp.index.searchsorted(target_date)
            if b0i < len(bp) and b1i < len(bp):
                bench_ret = float(bp.iloc[b1i] / bp.iloc[b0i] - 1.0)
        past = close.loc[:as_of].tail(252)
        mom12 = float(past.iloc[-1] / past.iloc[0] - 1.0) if len(past) >= 126 else None
        past6 = close.loc[:as_of].tail(126)
        mom6 = float(past6.iloc[-1] / past6.iloc[0] - 1.0) if len(past6) >= 63 else None
        vol6 = float(past6.pct_change().std() * np.sqrt(252)) if len(past6) >= 63 else None
        rows.append({
            "ticker": ticker,
            "as_of": as_of,
            "target_date": target_date,
            "revenue_growth": (rev / prev_rev - 1.0) if prev_rev not in (None,0) else None,
            "operating_margin": op / rev if op is not None else None,
            "net_margin": ni / rev if ni is not None else None,
            "fcf_margin": (ocf - capex) / rev if ocf is not None and capex is not None else None,
            "capex_to_revenue": capex / rev if capex is not None else None,
            "rd_to_revenue": rd / rev if rd is not None else None,
            "roe": ni / equity if ni is not None and equity not in (None,0) else None,
            "net_debt_to_revenue": ((debt or 0) - (cash or 0)) / rev,
            "momentum_12m": mom12,
            "momentum_6m": mom6,
            "volatility_6m": vol6,
            "drawdown_12m": max_drawdown_from_prices(past),
            "target_excess_return_12m": stock_ret - bench_ret,
        })
    return rows


def build_expected_return_dataset(universe: Iterable[str], benchmark: str = BENCHMARK) -> pd.DataFrame:
    try:
        bp = yf.Ticker(benchmark).history(period="10y", auto_adjust=True)["Close"]
        bp.index = pd.to_datetime(bp.index, utc=True).tz_convert(None)
    except Exception:
        bp = pd.Series(dtype=float)
    rows = []
    for ticker in dict.fromkeys(str(x).upper() for x in universe):
        rows.extend(annual_fundamental_rows(ticker, bp))
    return pd.DataFrame(rows)


def current_market_features(ticker: str) -> dict[str, float | None]:
    try:
        hist = yf.Ticker(ticker).history(period="2y", auto_adjust=True)["Close"].dropna()
        hist.index = pd.to_datetime(hist.index, utc=True).tz_convert(None)
    except Exception:
        hist = pd.Series(dtype=float)
    out = {"momentum_12m":None,"momentum_6m":None,"volatility_6m":None,"drawdown_12m":None}
    if len(hist) >= 63:
        p6 = hist.tail(126)
        out["momentum_6m"] = float(p6.iloc[-1]/p6.iloc[0]-1) if len(p6)>=63 else None
        out["volatility_6m"] = float(p6.pct_change().std()*np.sqrt(252))
    if len(hist) >= 126:
        p12 = hist.tail(252)
        out["momentum_12m"] = float(p12.iloc[-1]/p12.iloc[0]-1)
        out["drawdown_12m"] = max_drawdown_from_prices(p12)
    return out


def earnings_history(ticker: str, limit: int = 40) -> pd.DataFrame:
    try:
        df = yf.Ticker(ticker).get_earnings_dates(limit=limit)
    except Exception:
        return pd.DataFrame()
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy().reset_index()
    date_col = out.columns[0]
    out = out.rename(columns={date_col:"earnings_date","EPS Estimate":"eps_estimate","Reported EPS":"reported_eps","Surprise(%)":"surprise_pct"})
    out["earnings_date"] = pd.to_datetime(out["earnings_date"], errors="coerce", utc=True).dt.tz_convert(None)
    out["surprise_pct"] = pd.to_numeric(out.get("surprise_pct"), errors="coerce")
    med = out["surprise_pct"].abs().median(skipna=True)
    if pd.notna(med) and med > 1:
        out["surprise_pct"] = out["surprise_pct"] / 100.0
    return out.sort_values("earnings_date").reset_index(drop=True)


def earnings_model_frame(ticker: str) -> pd.DataFrame:
    columns=["as_of","target_date","prior_surprise_1q","prior_surprise_2q_avg","prior_surprise_4q_avg","surprise_vol_4q","price_momentum_3m","price_vol_3m","eps_estimate","target_surprise"]
    e = earnings_history(ticker)
    if e.empty:
        return pd.DataFrame(columns=columns)
    try:
        prices = yf.Ticker(ticker).history(period="10y", auto_adjust=True)["Close"].dropna()
        prices.index = pd.to_datetime(prices.index, utc=True).tz_convert(None)
    except Exception:
        prices = pd.Series(dtype=float)
    rows=[]
    surprises = e["surprise_pct"]
    for i,row in e.iterrows():
        if i < 4 or pd.isna(row.get("surprise_pct")):
            continue
        event=pd.Timestamp(row["earnings_date"])
        as_of=event-pd.Timedelta(days=1)
        prior=prices.loc[:as_of].tail(63)
        mom=float(prior.iloc[-1]/prior.iloc[0]-1) if len(prior)>=30 else None
        vol=float(prior.pct_change().std()*np.sqrt(252)) if len(prior)>=30 else None
        rows.append({
            "as_of":as_of,
            "target_date":event,
            "prior_surprise_1q":num(surprises.iloc[i-1]),
            "prior_surprise_2q_avg":num(surprises.iloc[max(0,i-2):i].mean()),
            "prior_surprise_4q_avg":num(surprises.iloc[max(0,i-4):i].mean()),
            "surprise_vol_4q":num(surprises.iloc[max(0,i-4):i].std()),
            "price_momentum_3m":mom,
            "price_vol_3m":vol,
            "eps_estimate":num(row.get("eps_estimate")),
            "target_surprise":num(row.get("surprise_pct")),
        })
    return pd.DataFrame(rows,columns=columns)


def regime_feature_frame(period: str = "20y") -> pd.DataFrame:
    tickers=["SPY","TLT","HYG","LQD","DBC","UUP"]
    try:
        data=yf.download(tickers,period=period,auto_adjust=True,progress=False)["Close"].dropna(how="all")
    except Exception:
        return pd.DataFrame()
    monthly=data.resample("ME").last().dropna()
    r6=monthly.pct_change(6); vol=monthly["SPY"].pct_change().rolling(3).std()*np.sqrt(12)
    return pd.DataFrame({
        "equity_6m":r6["SPY"],"bond_6m":r6["TLT"],"credit_6m":r6["HYG"]-r6["LQD"],
        "commodity_6m":r6["DBC"],"dollar_6m":r6["UUP"],"equity_vol_3m":vol,
    }).dropna()


def historical_financial_anomaly_frame(path: Path) -> pd.DataFrame:
    wb=load_workbook(path,data_only=True,read_only=True)
    if "Historical Financials" not in wb.sheetnames: return pd.DataFrame()
    ws=wb["Historical Financials"]
    rows=[]; prev_rev=None
    for c in range(2,8):
        year=ws.cell(3,c).value
        if not isinstance(year,(int,float)): continue
        rev=num(ws.cell(4,c).value); op=num(ws.cell(9,c).value); ni=num(ws.cell(11,c).value); ocf=num(ws.cell(14,c).value); cap=num(ws.cell(15,c).value); rd=num(ws.cell(19,c).value); sbc=num(ws.cell(21,c).value)
        if rev in (None,0): continue
        rows.append({"year":int(year),"revenue_growth":rev/prev_rev-1 if prev_rev else None,"operating_margin":op/rev if op is not None else None,"net_margin":ni/rev if ni is not None else None,"fcf_margin":(ocf-abs(cap))/rev if ocf is not None and cap is not None else None,"capex_to_revenue":abs(cap)/rev if cap is not None else None,"rd_to_revenue":rd/rev if rd is not None else None,"sbc_to_revenue":sbc/rev if sbc is not None else None})
        prev_rev=rev
    return pd.DataFrame(rows)


def load_ai_kpi_snapshots(repo_root: Path, ticker: str) -> pd.DataFrame:
    path=repo_root/"research_data"/ticker/"kpi_history.json"
    if not path.exists(): return pd.DataFrame()
    try: data=json.loads(path.read_text(encoding="utf-8"))
    except Exception: return pd.DataFrame()
    rows=[]
    for snap in data.get("snapshots",[]):
        row={"as_of":pd.to_datetime(snap.get("as_of"),errors="coerce")}
        for item in snap.get("kpis",[]):
            val=num(item.get("value"))
            if val is not None: row[str(item.get("name"))]=val
        rows.append(row)
    return pd.DataFrame(rows).sort_values("as_of") if rows else pd.DataFrame()


def load_portfolio(repo_root: Path) -> pd.DataFrame:
    path=repo_root/"institutional_research"/"portfolio.csv"
    if not path.exists(): return pd.DataFrame(columns=["ticker","current_weight"])
    try:
        df=pd.read_csv(path); df.columns=[str(c).strip().lower() for c in df.columns]
    except Exception: return pd.DataFrame(columns=["ticker","current_weight"])
    if "ticker" not in df.columns: return pd.DataFrame(columns=["ticker","current_weight"])
    if "current_weight" not in df.columns: df["current_weight"]=0.0
    df["ticker"]=df["ticker"].astype(str).str.upper().str.strip(); df["current_weight"]=pd.to_numeric(df["current_weight"],errors="coerce").fillna(0.0)
    if df["current_weight"].sum()>1.5: df["current_weight"]/=100.0
    return df[["ticker","current_weight"]].drop_duplicates("ticker")