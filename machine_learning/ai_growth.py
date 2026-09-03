from __future__ import annotations

"""AI-growth forecasting layer for equity research.

Architecture:
1) Evidence extraction: convert issuer / KPI evidence into normalized AI demand,
   monetization, adoption, efficiency, capex-burden, and risk signals. OpenAI
   Structured Outputs are optional; deterministic extraction is always available.
2) LightGBM: forecast next-fiscal-year revenue and FCF growth from point-in-time
   fundamental / market features stored by the existing ML history layer.
3) Expectations gap: compare the AI-adjusted FCF-growth forecast with a reverse-DCF
   implied long-run FCF growth rate.

The AI overlay is intentionally bounded. Until the project accumulates enough dated
AI KPI history to train AI features directly, it acts as an evidence-based adjustment
to the independently trained LightGBM fundamental forecast rather than pretending a
sparse AI dataset can support a fully supervised model.
"""

from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
import json
import math
import os
import re
import sqlite3

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from lightgbm import LGBMRegressor
except Exception:  # pragma: no cover - dependency gate is reported at runtime
    LGBMRegressor = None

from sklearn.impute import SimpleImputer
from sklearn.linear_model import ElasticNet
from sklearn.metrics import mean_absolute_error
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler

from .common import RANDOM_STATE, num


GROWTH_FEATURES = [
    "revenue_growth",
    "operating_margin",
    "net_margin",
    "fcf_margin",
    "capex_to_revenue",
    "rd_to_revenue",
    "roe",
    "net_debt_to_revenue",
    "momentum_12m",
    "momentum_6m",
    "volatility_6m",
    "drawdown_12m",
]

AI_SCORE_FIELDS = [
    "demand_score",
    "monetization_score",
    "adoption_score",
    "efficiency_score",
    "capex_burden_score",
    "risk_score",
]

SIGNAL_WORDS = {
    "very strong": 0.92,
    "strong": 0.80,
    "positive": 0.70,
    "moderate": 0.55,
    "neutral": 0.50,
    "mixed": 0.45,
    "weak": 0.30,
    "key risk": 0.20,
    "risk": 0.25,
}

CATEGORY_TERMS = {
    "demand_score": (
        "demand", "backlog", "booking", "rpo", "orders", "pipeline", "capacity",
        "workload", "cloud growth", "data center", "datacenter", "inference",
    ),
    "monetization_score": (
        "revenue", "arr", "run-rate", "run rate", "pricing", "paid", "sales",
        "monetization", "arpu", "attach rate", "spend", "contract",
    ),
    "adoption_score": (
        "user", "customer", "seat", "developer", "deployment", "adoption", "usage",
        "monthly active", "mau", "enterprise", "engagement", "workload",
    ),
    "efficiency_score": (
        "margin", "roi", "roic", "productivity", "cost saving", "cost savings",
        "unit cost", "efficiency", "conversion", "cpm", "cpa",
    ),
    "capex_burden_score": (
        "capex", "capital expenditure", "property and equipment", "depreciation",
        "power", "gpu capacity", "infrastructure build", "cash outflow",
    ),
    "risk_score": (
        "risk", "constraint", "cannibal", "competition", "regulation", "shortage",
        "bottleneck", "outflow", "uncertain", "pressure", "dependency",
    ),
}


@dataclass
class AISignalSnapshot:
    demand_score: float = 0.5
    monetization_score: float = 0.5
    adoption_score: float = 0.5
    efficiency_score: float = 0.5
    capex_burden_score: float = 0.5
    risk_score: float = 0.5
    confidence: float = 0.2
    extraction_mode: str = "deterministic"
    model: str | None = None
    evidence_count: int = 0
    source_characters: int = 0
    summary: str = "No material AI evidence was available; scores remain neutral."
    evidence: list[str] | None = None

    def to_dict(self) -> dict[str, Any]:
        out = asdict(self)
        out["evidence"] = list(self.evidence or [])
        return out


@dataclass
class GrowthForecast:
    target: str
    status: str
    prediction: float | None = None
    confidence: str = "Low"
    metrics: dict[str, Any] | None = None
    drivers: list[dict[str, Any]] | None = None
    training_rows: int = 0

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def _finite(v: Any) -> float | None:
    try:
        x = float(v)
        return x if math.isfinite(x) else None
    except Exception:
        return None


def _clip01(v: Any, default: float = 0.5) -> float:
    x = _finite(v)
    if x is None:
        return default
    return float(np.clip(x, 0.0, 1.0))


def _signal_value(text: Any) -> float:
    raw = str(text or "").lower().strip()
    for key, value in SIGNAL_WORDS.items():
        if key in raw:
            return value
    return 0.5


def _growth_mentions(text: str) -> list[float]:
    values: list[float] = []
    # Captures +82% YoY, 36.7% growth, grew 24%, etc. Avoid arbitrary percentages
    # unless the nearby phrase suggests growth / demand / spend / usage.
    for m in re.finditer(
        r"(?i)(?:growth|grew|increase|increased|up|yoy|year[- ]over[- ]year|spend|usage|revenue)"
        r"[^%\n]{0,32}?([+-]?\d{1,3}(?:\.\d+)?)\s*%",
        text,
    ):
        try:
            values.append(float(m.group(1)) / 100.0)
        except Exception:
            pass
    return values


def load_kpi_evidence(repo_root: Path, ticker: str) -> tuple[list[dict[str, Any]], str]:
    """Load the latest dated KPI snapshot plus optional analyst-supplied AI source files."""
    base = Path(repo_root) / "research_data" / ticker.upper()
    rows: list[dict[str, Any]] = []
    corpus_parts: list[str] = []

    history = base / "kpi_history.json"
    if history.exists():
        try:
            payload = json.loads(history.read_text(encoding="utf-8"))
            snapshots = payload.get("snapshots") or []
            if snapshots:
                latest = sorted(
                    snapshots,
                    key=lambda x: str(x.get("captured_at") or ""),
                )[-1]
                for row in latest.get("kpis") or []:
                    if isinstance(row, dict):
                        rows.append(dict(row))
                        corpus_parts.append(
                            " | ".join(
                                str(row.get(k) or "")
                                for k in (
                                    "kpi", "current", "unit_comparison", "signal",
                                    "investment_read_through", "as_of", "data_type",
                                )
                            )
                        )
        except Exception:
            pass

    source_dir = base / "ai_sources"
    if source_dir.exists():
        for path in sorted(source_dir.rglob("*")):
            if not path.is_file() or path.suffix.lower() not in {".txt", ".md", ".json"}:
                continue
            try:
                text = path.read_text(encoding="utf-8", errors="ignore")
            except Exception:
                continue
            if text.strip():
                corpus_parts.append(f"\nSOURCE FILE: {path.name}\n{text[:50000]}")

    return rows, "\n".join(corpus_parts)


def deterministic_ai_signals(rows: list[dict[str, Any]], corpus: str) -> AISignalSnapshot:
    text = corpus.lower()
    buckets: dict[str, list[float]] = {k: [] for k in AI_SCORE_FIELDS}
    evidence: list[str] = []

    for row in rows:
        line = " ".join(
            str(row.get(k) or "")
            for k in ("kpi", "unit_comparison", "signal", "investment_read_through", "data_type")
        )
        low = line.lower()
        base_score = _signal_value(row.get("signal"))
        for field, terms in CATEGORY_TERMS.items():
            if any(term in low for term in terms):
                score = base_score
                if field in {"capex_burden_score", "risk_score"}:
                    # A "key risk" label means burden/risk is high, while a very strong
                    # demand label attached to capex does not automatically make burden low.
                    score = 1.0 - base_score if "risk" not in low and "outflow" not in low else max(0.65, 1.0 - base_score)
                buckets[field].append(float(np.clip(score, 0.0, 1.0)))
        if line.strip():
            evidence.append(line[:240])

    growths = _growth_mentions(corpus)
    if growths:
        positive = np.clip(np.nanmedian(growths), -0.5, 1.5)
        growth_score = float(np.clip(0.5 + positive / 1.5 * 0.45, 0.05, 0.95))
        buckets["demand_score"].append(growth_score)
        buckets["monetization_score"].append(growth_score)

    # Corpus-only keyword evidence provides a conservative nudge where structured KPI
    # labels are missing.
    for field, terms in CATEGORY_TERMS.items():
        hits = sum(text.count(term) for term in terms)
        if hits:
            if field in {"capex_burden_score", "risk_score"}:
                buckets[field].append(min(0.75, 0.50 + 0.025 * hits))
            else:
                buckets[field].append(min(0.75, 0.50 + 0.025 * hits))

    scores = {
        field: float(np.mean(vals)) if vals else 0.5
        for field, vals in buckets.items()
    }
    evidence_count = len(rows)
    category_coverage = sum(bool(buckets[f]) for f in AI_SCORE_FIELDS) / len(AI_SCORE_FIELDS)
    confidence = float(np.clip(0.15 + 0.08 * min(evidence_count, 6) + 0.25 * category_coverage, 0.15, 0.72))
    if not rows and not corpus.strip():
        confidence = 0.12
    return AISignalSnapshot(
        **scores,
        confidence=confidence,
        extraction_mode="deterministic",
        evidence_count=evidence_count,
        source_characters=len(corpus),
        summary=(
            f"Deterministic evidence extraction from {evidence_count} KPI row(s). "
            "Use --llm for structured LLM classification when an API key is available."
            if evidence_count or corpus.strip()
            else "No AI KPI/source evidence found; neutral scores are retained."
        ),
        evidence=evidence[:8],
    )


def llm_ai_signals(
    rows: list[dict[str, Any]],
    corpus: str,
    *,
    model: str = "gpt-5.6-luna",
) -> AISignalSnapshot:
    """Extract normalized AI signals with OpenAI Structured Outputs.

    Falls back to deterministic extraction on missing credentials, SDK issues, refusal,
    schema mismatch, or transient API errors.
    """
    fallback = deterministic_ai_signals(rows, corpus)
    if not os.getenv("OPENAI_API_KEY") or not corpus.strip():
        return fallback

    schema = {
        "type": "object",
        "properties": {
            "demand_score": {"type": "number", "minimum": 0, "maximum": 1},
            "monetization_score": {"type": "number", "minimum": 0, "maximum": 1},
            "adoption_score": {"type": "number", "minimum": 0, "maximum": 1},
            "efficiency_score": {"type": "number", "minimum": 0, "maximum": 1},
            "capex_burden_score": {"type": "number", "minimum": 0, "maximum": 1},
            "risk_score": {"type": "number", "minimum": 0, "maximum": 1},
            "confidence": {"type": "number", "minimum": 0, "maximum": 1},
            "summary": {"type": "string"},
            "evidence": {
                "type": "array",
                "items": {"type": "string"},
                "maxItems": 8,
            },
        },
        "required": [
            "demand_score", "monetization_score", "adoption_score",
            "efficiency_score", "capex_burden_score", "risk_score",
            "confidence", "summary", "evidence",
        ],
        "additionalProperties": False,
    }
    instructions = (
        "You are an equity-research data extraction engine. Score only evidence present "
        "in the supplied issuer/KPI corpus. Scores are 0 to 1. 0.5 is neutral/unknown. "
        "Demand measures AI demand/backlog/workload momentum; monetization measures actual "
        "AI revenue/pricing/paid usage; adoption measures customers/users/deployments; "
        "efficiency measures margins/unit economics/productivity; capex_burden measures "
        "cash/capital intensity (higher is more burdensome); risk measures competition, "
        "capacity, regulation, cannibalization and execution risk (higher is worse). "
        "Do not infer undisclosed AI revenue. Keep evidence strings short and factual."
    )
    try:
        from openai import OpenAI

        client = OpenAI()
        response = client.responses.create(
            model=model,
            instructions=instructions,
            input=corpus[:30000],
            reasoning={"effort": "low"},
            text={
                "verbosity": "low",
                "format": {
                    "type": "json_schema",
                    "name": "ai_growth_signals",
                    "strict": True,
                    "schema": schema,
                },
            },
        )
        raw = getattr(response, "output_text", "") or ""
        data = json.loads(raw)
        return AISignalSnapshot(
            demand_score=_clip01(data.get("demand_score")),
            monetization_score=_clip01(data.get("monetization_score")),
            adoption_score=_clip01(data.get("adoption_score")),
            efficiency_score=_clip01(data.get("efficiency_score")),
            capex_burden_score=_clip01(data.get("capex_burden_score")),
            risk_score=_clip01(data.get("risk_score")),
            confidence=_clip01(data.get("confidence"), fallback.confidence),
            extraction_mode="openai_structured",
            model=model,
            evidence_count=len(rows),
            source_characters=len(corpus),
            summary=str(data.get("summary") or fallback.summary),
            evidence=[str(x)[:260] for x in (data.get("evidence") or [])[:8]],
        )
    except Exception as exc:
        fallback.summary += f" LLM extraction fell back safely: {type(exc).__name__}."
        return fallback


def growth_training_frame_from_history(db_path: Path) -> pd.DataFrame:
    """Create realized company-year growth targets from the persistent PIT store."""
    db_path = Path(db_path)
    if not db_path.exists():
        return pd.DataFrame()
    query = """
        SELECT symbol,fiscal_year,fiscal_date,available_date,revenue,operating_income,
               net_income,fcf,capex,rd,cash,debt,equity,source
        FROM fundamentals
        WHERE period='annual' AND revenue IS NOT NULL
    """
    try:
        with sqlite3.connect(db_path) as con:
            raw = pd.read_sql_query(query, con)
    except Exception:
        return pd.DataFrame()
    if raw.empty:
        return raw

    raw["available_date"] = pd.to_datetime(raw["available_date"], errors="coerce")
    raw["fiscal_date"] = pd.to_datetime(raw["fiscal_date"], errors="coerce")
    def source_rank(s: Any) -> int:
        s = str(s or "").lower()
        if "alpha vantage" in s:
            return 0
        if "sec" in s:
            return 1
        return 2
    raw["_source_rank"] = raw["source"].map(source_rank)
    raw = (
        raw.sort_values(["symbol", "fiscal_year", "_source_rank", "available_date"])
        .drop_duplicates(["symbol", "fiscal_year"], keep="first")
        .sort_values(["symbol", "fiscal_year"])
        .reset_index(drop=True)
    )
    g = raw.groupby("symbol", group_keys=False)
    prev_rev = g["revenue"].shift(1)
    next_rev = g["revenue"].shift(-1)
    next_fcf = g["fcf"].shift(-1)
    next_available = g["available_date"].shift(-1)
    next_year = g["fiscal_year"].shift(-1)
    contiguous_next_year = next_year.eq(raw["fiscal_year"] + 1)

    rev = pd.to_numeric(raw["revenue"], errors="coerce")
    op = pd.to_numeric(raw["operating_income"], errors="coerce")
    ni = pd.to_numeric(raw["net_income"], errors="coerce")
    fcf = pd.to_numeric(raw["fcf"], errors="coerce")
    capex = pd.to_numeric(raw["capex"], errors="coerce").abs()
    rd = pd.to_numeric(raw["rd"], errors="coerce")
    equity = pd.to_numeric(raw["equity"], errors="coerce")
    cash = pd.to_numeric(raw["cash"], errors="coerce").fillna(0.0)
    debt = pd.to_numeric(raw["debt"], errors="coerce").fillna(0.0)

    out = pd.DataFrame({
        "ticker": raw["symbol"],
        "as_of": raw["available_date"],
        "target_date": next_available,
        "revenue_growth": np.where(prev_rev.notna() & (prev_rev != 0), rev / prev_rev - 1.0, np.nan),
        "operating_margin": np.where(rev != 0, op / rev, np.nan),
        "net_margin": np.where(rev != 0, ni / rev, np.nan),
        "fcf_margin": np.where(rev != 0, fcf / rev, np.nan),
        "capex_to_revenue": np.where(rev != 0, capex / rev, np.nan),
        "rd_to_revenue": np.where(rev != 0, rd / rev, np.nan),
        "roe": np.where(equity != 0, ni / equity, np.nan),
        "net_debt_to_revenue": np.where(rev != 0, (debt - cash) / rev, np.nan),
        "target_next_revenue_growth": np.where(
            contiguous_next_year & next_rev.notna() & (rev != 0), next_rev / rev - 1.0, np.nan
        ),
        "target_next_fcf_growth": np.where(
            contiguous_next_year & next_fcf.notna() & (fcf > 0) & (next_fcf > 0),
            next_fcf / fcf - 1.0,
            np.nan,
        ),
    })

    # Join market features already built point-in-time by the existing history layer.
    try:
        with sqlite3.connect(db_path) as con:
            features = pd.read_sql_query(
                """SELECT symbol,as_of,momentum_12m,momentum_6m,volatility_6m,drawdown_12m
                   FROM features""",
                con,
            )
        if not features.empty:
            features["as_of"] = pd.to_datetime(features["as_of"], errors="coerce")
            features = features.sort_values(["symbol", "as_of"])
            out = out.sort_values(["ticker", "as_of"])
            joined = []
            for ticker, grp in out.groupby("ticker"):
                f = features[features["symbol"] == ticker].drop(columns=["symbol"])
                if f.empty:
                    joined.append(grp)
                    continue
                merged = pd.merge_asof(
                    grp.sort_values("as_of"),
                    f.sort_values("as_of"),
                    on="as_of",
                    direction="backward",
                    tolerance=pd.Timedelta(days=400),
                )
                joined.append(merged)
            out = pd.concat(joined, ignore_index=True) if joined else out
    except Exception:
        pass

    for col in ("momentum_12m", "momentum_6m", "volatility_6m", "drawdown_12m"):
        if col not in out:
            out[col] = np.nan
    return out.replace([np.inf, -np.inf], np.nan)


class LightGBMGrowthForecaster:
    """Time-aware LightGBM growth forecaster with Elastic Net benchmark and SHAP."""

    def __init__(self, feature_cols: list[str] | None = None):
        self.feature_cols = list(feature_cols or GROWTH_FEATURES)

    def fit_predict(
        self,
        frame: pd.DataFrame,
        current: dict[str, Any],
        *,
        target_col: str,
        target_label: str,
        min_rows: int = 45,
    ) -> GrowthForecast:
        if LGBMRegressor is None:
            return GrowthForecast(
                target_label, "DEPENDENCY_MISSING",
                metrics={"required_dependency": "lightgbm"},
            )
        if frame is None or frame.empty or target_col not in frame:
            return GrowthForecast(target_label, "INSUFFICIENT_DATA")
        df = frame.copy()
        df["as_of"] = pd.to_datetime(df.get("as_of"), errors="coerce")
        df["target_date"] = pd.to_datetime(df.get("target_date"), errors="coerce")
        df[target_col] = pd.to_numeric(df[target_col], errors="coerce")
        df = df.dropna(subset=["as_of", "target_date", target_col]).sort_values("as_of")
        if len(df) < min_rows:
            return GrowthForecast(
                target_label, "INSUFFICIENT_DATA", training_rows=len(df),
                metrics={"minimum_rows": min_rows},
            )
        for col in self.feature_cols:
            if col not in df:
                df[col] = np.nan
            df[col] = pd.to_numeric(df[col], errors="coerce")

        # Last ~20% of observations are the temporal holdout. Training observations are
        # purged unless their target was knowable before the first holdout as-of date.
        split_idx = max(int(len(df) * 0.80), min_rows - 1)
        split_idx = min(split_idx, len(df) - 12)
        cutoff = df.iloc[split_idx]["as_of"]
        train = df[df["target_date"] < cutoff].copy()
        test = df[df["as_of"] >= cutoff].copy()
        if len(train) < 30 or len(test) < 8:
            train = df.iloc[:split_idx].copy()
            test = df.iloc[split_idx:].copy()

        imputer = SimpleImputer(strategy="median")
        X_train = imputer.fit_transform(train[self.feature_cols])
        X_test = imputer.transform(test[self.feature_cols])
        y_train = train[target_col].to_numpy(dtype=float)
        y_test = test[target_col].to_numpy(dtype=float)

        model = LGBMRegressor(
            n_estimators=350,
            learning_rate=0.03,
            num_leaves=15,
            max_depth=4,
            min_child_samples=12,
            subsample=0.85,
            colsample_bytree=0.85,
            reg_alpha=0.15,
            reg_lambda=0.8,
            random_state=RANDOM_STATE,
            verbosity=-1,
        )
        model.fit(X_train, y_train)
        p_test = model.predict(X_test)
        mae = float(mean_absolute_error(y_test, p_test))

        elastic = Pipeline([
            ("imputer", SimpleImputer(strategy="median")),
            ("scale", StandardScaler()),
            ("model", ElasticNet(alpha=0.02, l1_ratio=0.25, max_iter=10000, random_state=RANDOM_STATE)),
        ])
        elastic.fit(train[self.feature_cols], y_train)
        elastic_mae = float(mean_absolute_error(y_test, elastic.predict(test[self.feature_cols])))

        # Refit on all realized rows for the current forecast.
        full_imputer = SimpleImputer(strategy="median")
        X_full = full_imputer.fit_transform(df[self.feature_cols])
        model.fit(X_full, df[target_col].to_numpy(dtype=float))
        x_current_df = pd.DataFrame([{c: num(current.get(c)) for c in self.feature_cols}])
        x_current = full_imputer.transform(x_current_df)
        pred = float(model.predict(x_current)[0])

        drivers: list[dict[str, Any]] = []
        try:
            import shap
            explainer = shap.TreeExplainer(model)
            values = explainer.shap_values(x_current)
            arr = np.asarray(values).reshape(-1)
            drivers = sorted(
                [
                    {
                        "feature": feature,
                        "shap_value": float(value),
                        "direction": "positive" if value >= 0 else "negative",
                        "current_value": _finite(x_current_df.iloc[0][feature]),
                    }
                    for feature, value in zip(self.feature_cols, arr)
                ],
                key=lambda x: abs(x["shap_value"]),
                reverse=True,
            )[:8]
        except Exception:
            drivers = sorted(
                [
                    {
                        "feature": feature,
                        "importance": float(value),
                        "direction": "unknown",
                        "current_value": _finite(x_current_df.iloc[0][feature]),
                    }
                    for feature, value in zip(self.feature_cols, model.feature_importances_)
                ],
                key=lambda x: x["importance"],
                reverse=True,
            )[:8]

        improvement = (elastic_mae - mae) / elastic_mae if elastic_mae > 0 else None
        # A nonlinear model does not earn higher confidence merely because the sample is large.
        # If it fails to match the simple Elastic Net holdout benchmark, confidence is downgraded.
        if elastic_mae > 0 and mae > elastic_mae * 1.05:
            confidence = "Low"
        elif len(df) >= 250 and mae <= elastic_mae:
            confidence = "High"
        elif len(df) >= 90:
            confidence = "Moderate"
        else:
            confidence = "Low"
        return GrowthForecast(
            target=target_label,
            status="PASS",
            prediction=pred,
            confidence=confidence,
            training_rows=len(df),
            metrics={
                "time_purged_holdout_mae": mae,
                "elastic_net_holdout_mae": elastic_mae,
                "mae_improvement_vs_elastic_net": improvement,
                "train_rows": len(train),
                "holdout_rows": len(test),
                "holdout_start": str(pd.Timestamp(cutoff).date()),
                "feature_columns": self.feature_cols,
            },
            drivers=drivers,
        )


def ai_adjustments(signals: AISignalSnapshot) -> dict[str, float]:
    """Return bounded percentage-point adjustments to fundamental growth forecasts."""
    positive = (
        0.30 * signals.demand_score
        + 0.30 * signals.monetization_score
        + 0.20 * signals.adoption_score
        + 0.20 * signals.efficiency_score
    )
    confidence_scale = 0.45 + 0.55 * signals.confidence
    revenue_adjustment = float(np.clip((positive - 0.5) * 0.20 * confidence_scale, -0.08, 0.08))
    fcf_adjustment = revenue_adjustment
    fcf_adjustment += (signals.efficiency_score - 0.5) * 0.06 * confidence_scale
    fcf_adjustment -= (signals.capex_burden_score - 0.5) * 0.06 * confidence_scale
    fcf_adjustment -= (signals.risk_score - 0.5) * 0.03 * confidence_scale
    return {
        "revenue_growth_adjustment": float(np.clip(revenue_adjustment, -0.10, 0.10)),
        "fcf_growth_adjustment": float(np.clip(fcf_adjustment, -0.12, 0.12)),
    }


def apply_ai_overlay(
    forecast: GrowthForecast,
    adjustment: float,
    *,
    lower: float = -0.60,
    upper: float = 1.50,
) -> float | None:
    if forecast.status != "PASS" or forecast.prediction is None:
        return None
    return float(np.clip(forecast.prediction + adjustment, lower, upper))


def workbook_current_growth_features(path: Path, ticker: str) -> dict[str, Any]:
    from .data import current_market_features, workbook_current_snapshot
    from runtime_data_guards import provider_symbol

    out = workbook_current_snapshot(path, ticker)
    try:
        out.update(current_market_features(provider_symbol(ticker)))
    except Exception:
        pass
    return out


def _latest_history_col(ws) -> int | None:
    cols = [
        c for c in range(2, min(ws.max_column, 12) + 1)
        if isinstance(ws.cell(3, c).value, (int, float))
    ]
    return cols[-1] if cols else None


def _equity_value_from_fcf(
    fcf0: float,
    growth: float,
    years: int,
    wacc: float,
    terminal_growth: float,
    net_debt: float,
) -> float:
    if wacc <= terminal_growth:
        return float("nan")
    fcf = fcf0
    pv = 0.0
    for year in range(1, years + 1):
        fcf *= 1.0 + growth
        pv += fcf / ((1.0 + wacc) ** year)
    terminal = fcf * (1.0 + terminal_growth) / (wacc - terminal_growth)
    return pv + terminal / ((1.0 + wacc) ** years) - net_debt


def reverse_dcf_from_workbook(
    path: Path,
    *,
    wacc: float = 0.09,
    terminal_growth: float = 0.03,
    years: int = 10,
    min_growth: float = -0.20,
    max_growth: float = 0.50,
) -> dict[str, Any]:
    """Solve a recruiter-safe FCF growth hurdle from the existing workbook."""
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        company = wb["Company Data"]
        hist = wb["Historical Financials"]
        market_cap = _finite(company["B10"].value)
        net_debt = _finite(company["B14"].value) or 0.0
        c = _latest_history_col(hist)
        if c is None:
            raise ValueError("Historical Financials has no dated columns")
        ocf = _finite(hist.cell(14, c).value)
        capex = _finite(hist.cell(15, c).value)
        current_fcf = (ocf - abs(capex)) if ocf is not None and capex is not None else None
    except Exception as exc:
        return {"status": "INSUFFICIENT_DATA", "error": repr(exc)}
    if market_cap is None or current_fcf is None or current_fcf <= 0:
        return {
            "status": "INSUFFICIENT_DATA",
            "market_cap": market_cap,
            "current_fcf": current_fcf,
            "net_debt": net_debt,
        }
    lo, hi = min_growth, max_growth
    vlo = _equity_value_from_fcf(current_fcf, lo, years, wacc, terminal_growth, net_debt)
    vhi = _equity_value_from_fcf(current_fcf, hi, years, wacc, terminal_growth, net_debt)
    if not np.isfinite(vlo) or not np.isfinite(vhi) or not (vlo <= market_cap <= vhi):
        return {
            "status": "OUTSIDE_SEARCH_RANGE",
            "implied_annual_fcf_growth": None,
            "wacc": wacc,
            "terminal_growth": terminal_growth,
            "forecast_years": years,
            "market_cap": market_cap,
            "current_fcf": current_fcf,
            "net_debt": net_debt,
        }
    for _ in range(100):
        mid = (lo + hi) / 2.0
        value = _equity_value_from_fcf(current_fcf, mid, years, wacc, terminal_growth, net_debt)
        if value < market_cap:
            lo = mid
        else:
            hi = mid
    implied = (lo + hi) / 2.0
    return {
        "status": "PASS",
        "implied_annual_fcf_growth": float(implied),
        "wacc": float(wacc),
        "terminal_growth": float(terminal_growth),
        "forecast_years": int(years),
        "market_cap": float(market_cap),
        "current_fcf": float(current_fcf),
        "net_debt": float(net_debt),
    }


def expectations_gap(
    ai_adjusted_fcf_growth: float | None,
    reverse_dcf: dict[str, Any],
) -> dict[str, Any]:
    implied = _finite(reverse_dcf.get("implied_annual_fcf_growth"))
    forecast = _finite(ai_adjusted_fcf_growth)
    if forecast is None or implied is None:
        return {
            "status": "INSUFFICIENT_DATA",
            "fcf_growth_gap": None,
            "interpretation": "A comparable FCF-growth forecast and reverse-DCF hurdle are both required.",
        }
    gap = forecast - implied
    if gap >= 0.05:
        interpretation = "Forecast growth exceeds the market-implied FCF hurdle by a meaningful margin."
    elif gap <= -0.05:
        interpretation = "Market-implied FCF growth is materially above the model forecast; expectations risk is elevated."
    else:
        interpretation = "Model forecast and reverse-DCF hurdle are broadly aligned."
    return {
        "status": "PASS",
        "fcf_growth_gap": float(gap),
        "forecast_ai_adjusted_fcf_growth": float(forecast),
        "reverse_dcf_implied_fcf_growth": float(implied),
        "interpretation": interpretation,
    }


def _fmt_pct(v: Any) -> str:
    x = _finite(v)
    return "N/M" if x is None else f"{x:.1%}"


def _fmt_num(v: Any, digits: int = 2) -> str:
    x = _finite(v)
    return "N/M" if x is None else f"{x:.{digits}f}"


def write_ai_growth_sheet(
    workbook_path: Path,
    ticker: str,
    payload: dict[str, Any],
    *,
    sheet_name: str = "AI Growth Forecast",
) -> None:
    """Write a separate research sheet; never overwrite authoritative DCF assumptions."""
    path = Path(workbook_path)
    wb = load_workbook(path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    for col, width in {"A": 31, "B": 18, "C": 18, "D": 18, "E": 50, "F": 22}.items():
        ws.column_dimensions[col].width = width

    navy = "17365D"; blue = "2F75B5"; white = "FFFFFF"; light = "F5F9FC"
    gold = "FFF2CC"; green = "008000"; grey = "666666"
    ws.merge_cells("A1:F2")
    ws["A1"] = f"{ticker} — AI Growth Forecast"
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws["A1"].font = Font(bold=True, color=white, size=18)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells("A3:F3")
    ws["A3"] = (
        "LLM/deterministic AI evidence extraction + LightGBM fundamental growth + reverse-DCF expectations gap. "
        "AI is a bounded evidence overlay until enough dated AI observations exist for supervised training."
    )
    ws["A3"].font = Font(italic=True, color=grey)
    ws["A3"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[3].height = 34

    def section(row: int, title: str):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row, 1, title)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(bold=True, color=white)

    def header(row: int, labels: list[str]):
        for col, label in enumerate(labels, 1):
            c = ws.cell(row, col, label)
            c.fill = PatternFill("solid", fgColor=blue)
            c.font = Font(bold=True, color=white)
            c.alignment = Alignment(horizontal="center", wrap_text=True)

    signals = payload.get("ai_signals") or {}
    section(5, "AI Evidence Signals — normalized 0 to 1")
    header(6, ["Signal", "Score", "Interpretation", "Extraction", "Evidence", "Notes"])
    signal_rows = [
        ("AI demand", "demand_score", "Higher = stronger demand/backlog/workload evidence"),
        ("AI monetization", "monetization_score", "Higher = stronger revenue/pricing/paid-usage evidence"),
        ("AI adoption", "adoption_score", "Higher = stronger users/customers/deployment evidence"),
        ("AI efficiency", "efficiency_score", "Higher = stronger margins/unit-economics/productivity evidence"),
        ("AI capex burden", "capex_burden_score", "Higher = heavier cash/capital burden"),
        ("AI risk", "risk_score", "Higher = greater competition/capacity/regulatory/execution risk"),
    ]
    for r, (label, field, note) in enumerate(signal_rows, 7):
        ws.cell(r, 1, label)
        ws.cell(r, 2, _finite(signals.get(field)))
        ws.cell(r, 2).number_format = "0%"
        ws.cell(r, 3, note)
        ws.cell(r, 4, signals.get("extraction_mode"))
        ws.cell(r, 5, signals.get("evidence_count"))
        ws.cell(r, 6, signals.get("summary") if r == 7 else "")
        for c in range(1, 7):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[7].height = 44

    section(14, "Growth Forecast & Market Expectations")
    header(15, ["Metric", "Fundamental ML", "AI adjustment", "AI-adjusted", "Market implied", "Gap / validation"])
    rev = payload.get("revenue_forecast") or {}
    fcf = payload.get("fcf_forecast") or {}
    adj = payload.get("ai_adjustments") or {}
    reverse = payload.get("reverse_dcf") or {}
    gap = payload.get("expectations_gap") or {}
    rows = [
        (
            "Next FY revenue growth",
            rev.get("prediction"),
            adj.get("revenue_growth_adjustment"),
            payload.get("ai_adjusted_revenue_growth"),
            None,
            f"LightGBM holdout MAE {_fmt_pct((rev.get('metrics') or {}).get('time_purged_holdout_mae'))}; "
            f"Elastic Net {_fmt_pct((rev.get('metrics') or {}).get('elastic_net_holdout_mae'))}",
        ),
        (
            "Next FY FCF growth",
            fcf.get("prediction"),
            adj.get("fcf_growth_adjustment"),
            payload.get("ai_adjusted_fcf_growth"),
            reverse.get("implied_annual_fcf_growth"),
            gap.get("fcf_growth_gap"),
        ),
    ]
    for r, item in enumerate(rows, 16):
        for c, value in enumerate(item, 1):
            ws.cell(r, c, value)
        for c in (2, 3, 4, 5):
            ws.cell(r, c).number_format = "0.0%"
        if r == 17 and _finite(ws.cell(r, 6).value) is not None:
            ws.cell(r, 6).number_format = "0.0%"
        ws.cell(r, 4).fill = PatternFill("solid", fgColor=gold)
        ws.cell(r, 4).font = Font(bold=True)

    ws.merge_cells("A19:F19")
    ws["A19"] = gap.get("interpretation") or "Expectations gap unavailable."
    ws["A19"].font = Font(bold=True, color=green if (_finite(gap.get("fcf_growth_gap")) or 0) >= 0 else "C00000")
    ws["A19"].alignment = Alignment(wrap_text=True)

    section(21, "LightGBM Explainability — current forecast drivers")
    header(22, ["Target", "Feature", "Direction", "SHAP / importance", "Current value", "Model confidence"])
    rr = 23
    for label, forecast in (("Revenue", rev), ("FCF", fcf)):
        for driver in (forecast.get("drivers") or [])[:6]:
            ws.cell(rr, 1, label)
            ws.cell(rr, 2, driver.get("feature"))
            ws.cell(rr, 3, driver.get("direction"))
            impact = driver.get("shap_value", driver.get("importance"))
            ws.cell(rr, 4, impact)
            ws.cell(rr, 5, driver.get("current_value"))
            ws.cell(rr, 6, forecast.get("confidence"))
            rr += 1
    if rr == 23:
        ws.cell(rr, 1, "No explainability output; growth history may still be insufficient.")
        rr += 1

    section(rr + 1, "Evidence & Governance")
    header(rr + 2, ["Evidence", "", "", "", "", ""])
    cursor = rr + 3
    for item in (signals.get("evidence") or [])[:8]:
        ws.merge_cells(start_row=cursor, start_column=1, end_row=cursor, end_column=6)
        ws.cell(cursor, 1, str(item))
        ws.cell(cursor, 1).alignment = Alignment(wrap_text=True)
        cursor += 1
    notes = [
        "LightGBM is benchmarked against Elastic Net on a chronological holdout; lower holdout MAE is better.",
        "If LightGBM fails to match the Elastic Net benchmark, model confidence is downgraded.",
        "The AI overlay is bounded and confidence-scaled because current AI KPI history is sparse relative to financial history.",
        "The LLM extracts and scores evidence only; it does not directly set the target price or DCF assumptions.",
        "Reverse DCF is a simplified FCF-growth hurdle and is not meaningful for every business model.",
        "No trades are executed and no private portfolio economics are exposed by this sheet.",
    ]
    for note in notes:
        ws.merge_cells(start_row=cursor, start_column=1, end_row=cursor, end_column=6)
        ws.cell(cursor, 1, "• " + note)
        ws.cell(cursor, 1).font = Font(color=grey)
        ws.cell(cursor, 1).alignment = Alignment(wrap_text=True)
        cursor += 1

    ws["F1"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
    wb.save(path)
