from __future__ import annotations

"""Workbook output for the AI Growth Forecast layer.

The presentation is intentionally plain-English first. Detailed model values remain in the
visible tables, while charts focus on three questions a non-technical reader can answer:
1) Is the AI evidence actually supportive or mostly neutral?
2) What growth does the model expect versus what the current price appears to require?
3) Which inputs influence the growth model most?
"""

from datetime import datetime, timezone
from pathlib import Path
from typing import Any
import math

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter


FEATURE_LABELS={
    "revenue_growth":"Revenue growth",
    "operating_margin":"Operating profit margin",
    "net_margin":"Net profit margin",
    "fcf_margin":"Free-cash-flow margin",
    "capex_to_revenue":"Capital spending / revenue",
    "rd_to_revenue":"R&D / revenue",
    "roe":"Return on equity",
    "net_debt_to_revenue":"Net debt / revenue",
    "momentum_12m":"12-month price trend",
    "momentum_6m":"6-month price trend",
    "volatility_6m":"6-month volatility",
    "drawdown_12m":"Recent drawdown",
}


def _label(value: Any) -> str:
    return FEATURE_LABELS.get(str(value),str(value).replace("_"," ").title())


def _finite(value: Any) -> float | None:
    try:
        x = float(value)
        return x if math.isfinite(x) else None
    except Exception:
        return None


def _fmt_pct(value: Any) -> str:
    x = _finite(value)
    return "N/M" if x is None else f"{x:.1%}"


def _set_chart_text_categories(chart, ws, category_col, first_row, last_row, series_titles=None):
    """Use explicit string categories/literal series titles so Excel need not repair chart OOXML."""
    sheet_name=ws.title.replace("'","''")
    col=get_column_letter(category_col)
    category_formula=f"'{sheet_name}'!${col}${first_row}:${col}${last_row}"
    titles=list(series_titles or [])
    for idx,series in enumerate(chart.series):
        series.cat=AxDataSource(strRef=StrRef(f=category_formula))
        if idx<len(titles):
            series.tx=SeriesLabel(v=str(titles[idx]))


def _placeholder_evidence(signals: dict[str, Any]) -> bool:
    evidence=[str(x or "").strip().lower() for x in (signals.get("evidence") or []) if str(x or "").strip()]
    if not evidence:
        return True
    placeholder_terms=("analyst input","to be updated","placeholder","not disclosed","n/a")
    substantive=0
    for line in evidence:
        if not any(term in line for term in placeholder_terms):
            substantive+=1
    return substantive==0


def _add_read_me(ws, signals: dict[str, Any]) -> None:
    navy="17365D"; white="FFFFFF"; light="F5F9FC"; gold="FFF2CC"
    ws.merge_cells("H2:N2"); ws["H2"]="How to read the AI page"
    ws["H2"].fill=PatternFill("solid",fgColor=navy); ws["H2"].font=Font(bold=True,color=white)
    ws.merge_cells("H3:N4")
    ws["H3"]=(
        "AI evidence scores use 50 as neutral. In the chart, risk and capital burden are inverted so higher always means more supportive. "
        "The valuation chart compares model growth with the growth today's price appears to require. Driver shares show model influence, not probability or causality."
    )
    ws["H3"].alignment=Alignment(wrap_text=True,vertical="top"); ws["H3"].fill=PatternFill("solid",fgColor=light)
    if _placeholder_evidence(signals) or str(signals.get("extraction_mode") or "").lower()=="deterministic":
        ws.merge_cells("H50:N54")
        ws["H50"]=(
            "Evidence caution: this run used deterministic extraction and/or mostly placeholder AI rows. Small differences around 50 should be treated as neutral. "
            "Do not treat the AI adjustment as a high-confidence company-specific forecast unless substantive AI evidence is present."
        )
        ws["H50"].fill=PatternFill("solid",fgColor=gold); ws["H50"].font=Font(bold=True); ws["H50"].alignment=Alignment(wrap_text=True,vertical="top")


def _add_ai_growth_charts(ws, signals: dict[str, Any], revenue: dict[str, Any], fcf: dict[str, Any],
                          payload: dict[str, Any], reverse: dict[str, Any]) -> None:
    """Create decision-oriented charts using hidden helper tables."""
    ws.column_dimensions["P"].hidden=True
    ws.column_dimensions["Q"].hidden=True
    _add_read_me(ws,signals)

    signal_rows=[
        ("Demand",_finite(signals.get("demand_score"))),
        ("Monetization",_finite(signals.get("monetization_score"))),
        ("Adoption",_finite(signals.get("adoption_score"))),
        ("Efficiency",_finite(signals.get("efficiency_score"))),
        ("Capital-light score",None if _finite(signals.get("capex_burden_score")) is None else 1.0-_finite(signals.get("capex_burden_score"))),
        ("Risk-adjusted score",None if _finite(signals.get("risk_score")) is None else 1.0-_finite(signals.get("risk_score"))),
    ]
    ws["P2"]="AI evidence factor"; ws["Q2"]="Supportive score"
    for i,(name,val) in enumerate(signal_rows,3):
        ws.cell(i,16,name); ws.cell(i,17,val*100 if val is not None else None); ws.cell(i,17).number_format="0.0"
    ch=BarChart(); ch.type="col"; ch.style=10
    ch.title="AI evidence: 50 = neutral, higher = more supportive"
    ch.y_axis.title="Supportive score"; ch.y_axis.scaling.min=0; ch.y_axis.scaling.max=100; ch.y_axis.numFmt="0.0"
    ch.height=7.2; ch.width=13.5; ch.legend=None
    ch.add_data(Reference(ws,min_col=17,min_row=3,max_row=8),titles_from_data=False)
    _set_chart_text_categories(ch,ws,16,3,8,["Supportive score"])
    ch.dLbls=DataLabelList(); ch.dLbls.showVal=True; ch.dLbls.numFmt="0.0"
    ws.add_chart(ch,"H6")

    ws["P11"]="Growth view"; ws["Q11"]="Annual FCF growth"
    fcf_rows=[
        ("Fundamental ML forecast",_finite(fcf.get("prediction"))),
        ("AI-adjusted forecast",_finite(payload.get("ai_adjusted_fcf_growth"))),
        ("Growth today's price appears to require",_finite(reverse.get("implied_annual_fcf_growth"))),
    ]
    end=11
    for i,(name,val) in enumerate(fcf_rows,12):
        if val is None: continue
        ws.cell(i,16,name); ws.cell(i,17,val*100); ws.cell(i,17).number_format="0.0"; end=max(end,i)
    if end>=12:
        ch=BarChart(); ch.type="col"; ch.style=11
        ch.title="What the model expects vs what today's price seems to require"
        ch.y_axis.title="Annual FCF growth (%)"; ch.y_axis.numFmt="0.0"
        vals=[v for _,v in fcf_rows if v is not None]
        if vals and min(vals)>=0: ch.y_axis.scaling.min=0
        ch.height=7.0; ch.width=13.5; ch.legend=None
        ch.add_data(Reference(ws,min_col=17,min_row=12,max_row=end),titles_from_data=False)
        _set_chart_text_categories(ch,ws,16,12,end,["Annual FCF growth"])
        ch.dLbls=DataLabelList(); ch.dLbls.showVal=True; ch.dLbls.numFmt="0.0"
        ws.add_chart(ch,"H20")

    driver_rows=[]
    for label,forecast in (("Revenue",revenue),("FCF",fcf)):
        candidates=[]
        for d in (forecast.get("drivers") or [])[:6]:
            val=_finite(d.get("shap_value",d.get("importance")))
            if val is None: continue
            direction="↑" if val>=0 else "↓"
            candidates.append((f"{label}: {_label(d.get('feature'))} {direction}",abs(val)))
        driver_rows.extend(sorted(candidates,key=lambda x:x[1],reverse=True)[:3])
    if driver_rows:
        total=sum(v for _,v in driver_rows)
        ws["P17"]="Driver"; ws["Q17"]="Relative influence"
        for i,(name,val) in enumerate(driver_rows,18):
            ws.cell(i,16,name); ws.cell(i,17,(val/total*100 if total>0 else 0)); ws.cell(i,17).number_format="0.0"
        ch=BarChart(); ch.type="bar"; ch.style=12
        ch.title="What influences the growth forecast most"
        ch.y_axis.title="Driver"; ch.x_axis.title="Share of top-driver influence (%)"; ch.x_axis.numFmt="0.0"
        ch.height=8.0; ch.width=14.5; ch.legend=None
        ch.add_data(Reference(ws,min_col=17,min_row=18,max_row=17+len(driver_rows)),titles_from_data=False)
        _set_chart_text_categories(ch,ws,16,18,17+len(driver_rows),["Relative influence"])
        ch.dLbls=DataLabelList(); ch.dLbls.showVal=True; ch.dLbls.numFmt="0.0"
        ws.add_chart(ch,"H34")


def write_ai_growth_sheet(
    workbook_path: Path,
    ticker: str,
    payload: dict[str, Any],
    *,
    sheet_name: str = "AI Growth Forecast",
) -> None:
    path = Path(workbook_path)
    wb = load_workbook(path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    for col, width in {
        "A":31,"B":18,"C":22,"D":20,"E":50,"F":28,"G":3,
        "H":15,"I":15,"J":15,"K":15,"L":15,"M":15,"N":15
    }.items():
        ws.column_dimensions[col].width = width

    navy = "17365D"; blue = "2F75B5"; white = "FFFFFF"; gold = "FFF2CC"
    green = "008000"; red = "C00000"; grey = "666666"; pale_red="FCE4D6"; pale_green="E2F0D9"

    # Keep the title/subtitle in the left analysis pane so the right pane is available for guidance/charts.
    ws.merge_cells("A1:F2")
    ws["A1"] = f"{ticker} — AI Growth Forecast"
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws["A1"].font = Font(bold=True, color=white, size=18)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells("A3:F3")
    ws["A3"] = (
        "AI evidence extraction + LightGBM fundamental growth + a reverse-DCF expectations check. "
        "AI remains a bounded second-opinion overlay until the company has enough dated AI evidence for supervised training."
    )
    ws["A3"].font = Font(italic=True, color=grey)
    ws["A3"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[3].height = 34
    ws["F4"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
    ws["F4"].font = Font(color=grey, italic=True, size=9)

    def section(row: int, title: str) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row, 1, title)
        c.fill = PatternFill("solid", fgColor=navy)
        c.font = Font(bold=True, color=white)

    def header(row: int, labels: list[str]) -> None:
        for col, label in enumerate(labels, 1):
            c = ws.cell(row, col, label)
            c.fill = PatternFill("solid", fgColor=blue)
            c.font = Font(bold=True, color=white)
            c.alignment = Alignment(horizontal="center", wrap_text=True)

    signals = payload.get("ai_signals") or {}
    section(5, "AI Evidence Signals — 50% means neutral / unknown")
    header(6, ["Signal", "Score", "What a higher score means", "Extraction", "Evidence rows", "Notes"])
    signal_rows = [
        ("AI demand", "demand_score", "More AI demand, backlog or workload evidence"),
        ("AI monetization", "monetization_score", "More paid usage, pricing or AI-revenue evidence"),
        ("AI adoption", "adoption_score", "More users, customers or deployments"),
        ("AI efficiency", "efficiency_score", "Better margins, unit economics or productivity"),
        ("AI capital burden", "capex_burden_score", "More cash/capital required — this is adverse"),
        ("AI risk", "risk_score", "More competition, regulation, capacity or execution risk — this is adverse"),
    ]
    for r, (label, field, note) in enumerate(signal_rows, 7):
        ws.cell(r, 1, label)
        ws.cell(r, 2, _finite(signals.get(field)))
        ws.cell(r, 2).number_format = "0%"
        ws.cell(r, 3, note)
        ws.cell(r, 4, signals.get("extraction_mode"))
        ws.cell(r, 5, signals.get("evidence_count"))
        ws.cell(r, 6, signals.get("summary") if r == 7 else "")
        for c in range(1, 7):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[7].height = 48

    section(14, "Growth Forecast & Market Expectations")
    header(15, ["Metric", "Fundamental model", "AI adjustment", "AI-adjusted forecast", "Growth implied by today's price", "Gap / model validation"])
    revenue = payload.get("revenue_forecast") or {}
    fcf = payload.get("fcf_forecast") or {}
    adj = payload.get("ai_adjustments") or {}
    reverse = payload.get("reverse_dcf") or {}
    gap = payload.get("expectations_gap") or {}
    rows = [
        (
            "Next FY revenue growth", revenue.get("prediction"), adj.get("revenue_growth_adjustment"),
            payload.get("ai_adjusted_revenue_growth"), None,
            f"Typical holdout error: LightGBM {_fmt_pct((revenue.get('metrics') or {}).get('time_purged_holdout_mae'))}; "
            f"simple Elastic Net {_fmt_pct((revenue.get('metrics') or {}).get('elastic_net_holdout_mae'))}; "
            f"confidence {revenue.get('confidence') or 'N/M'}",
        ),
        (
            "Next FY FCF growth", fcf.get("prediction"), adj.get("fcf_growth_adjustment"),
            payload.get("ai_adjusted_fcf_growth"), reverse.get("implied_annual_fcf_growth"),
            gap.get("fcf_growth_gap"),
        ),
    ]
    for r, item in enumerate(rows, 16):
        for c, value in enumerate(item, 1):
            ws.cell(r, c, value)
        for c in (2, 3, 4, 5):
            ws.cell(r, c).number_format = "0.0%"
        if r == 17 and _finite(ws.cell(r, 6).value) is not None:
            ws.cell(r, 6).number_format = "0.0%"
        ws.cell(r, 4).fill = PatternFill("solid", fgColor=gold)
        ws.cell(r, 4).font = Font(bold=True)

    ws.merge_cells("A19:F19")
    ws["A19"] = gap.get("interpretation") or "Expectations gap unavailable."
    gap_value = _finite(gap.get("fcf_growth_gap"))
    ws["A19"].font = Font(bold=True, color=green if gap_value is not None and gap_value >= 0 else red)
    ws["A19"].alignment = Alignment(wrap_text=True)

    section(21, "LightGBM Explainability — current forecast drivers")
    header(22, ["Forecast", "Feature", "Effect on forecast", "Model impact", "Current company value", "Confidence"])
    rr = 23
    for label, forecast in (("Revenue", revenue), ("FCF", fcf)):
        for driver in (forecast.get("drivers") or [])[:6]:
            ws.cell(rr, 1, label)
            ws.cell(rr, 2, _label(driver.get("feature")))
            ws.cell(rr, 3, "Raises forecast" if str(driver.get("direction"))=="positive" else "Lowers forecast" if str(driver.get("direction"))=="negative" else "Influences forecast")
            ws.cell(rr, 4, driver.get("shap_value", driver.get("importance")))
            ws.cell(rr, 5, driver.get("current_value"))
            ws.cell(rr, 6, forecast.get("confidence"))
            ws.cell(rr,4).number_format='0.0%;[Red](0.0%);-'
            ws.cell(rr,5).number_format='0.0%;[Red](0.0%);-'
            if str(forecast.get("confidence") or "").lower()=="low":
                ws.cell(rr,6).fill=PatternFill("solid",fgColor=pale_red)
            elif str(forecast.get("confidence") or "").lower()=="high":
                ws.cell(rr,6).fill=PatternFill("solid",fgColor=pale_green)
            rr += 1
    if rr == 23:
        ws.cell(rr, 1, "No explainability output; growth history may still be insufficient.")
        rr += 1

    section(rr + 1, "Evidence & Governance")
    header(rr + 2, ["Evidence", "", "", "", "", ""])
    cursor = rr + 3
    for item in (signals.get("evidence") or [])[:8]:
        ws.merge_cells(start_row=cursor, start_column=1, end_row=cursor, end_column=6)
        ws.cell(cursor, 1, str(item))
        ws.cell(cursor, 1).alignment = Alignment(wrap_text=True)
        cursor += 1
    for note in [
        "LightGBM is benchmarked against Elastic Net on a chronological holdout; lower holdout error is better.",
        "If the nonlinear model does not add value over the simpler benchmark, confidence is reduced.",
        "The AI overlay is bounded because company-specific AI evidence is usually much sparser than financial history.",
        "The LLM/deterministic extractor scores evidence only; it does not directly set target price or DCF assumptions.",
        "A SHAP/model-impact value is relative to the model's baseline forecast and can be larger than the final forecast; use the chart for relative influence.",
        "Reverse DCF is a simplified growth hurdle and is not meaningful for every business model.",
        "No trades are executed and no private portfolio economics are exposed by this sheet.",
    ]:
        ws.merge_cells(start_row=cursor, start_column=1, end_row=cursor, end_column=6)
        ws.cell(cursor, 1, "• " + note)
        ws.cell(cursor, 1).font = Font(color=grey)
        ws.cell(cursor, 1).alignment = Alignment(wrap_text=True)
        cursor += 1

    _add_ai_growth_charts(ws,signals,revenue,fcf,payload,reverse)
    wb.save(path)
