"""Consensus normalization, provider cross-checks and market-expectation integration.

This layer runs after the workbook's core institutional sheets have been built. It replaces
single-provider consensus cells with a transparent multi-source reference when additional
API providers are configured, preserves provider-level provenance, derives EBIT margin and
capex intensity only when the underlying consensus components exist, and adds external
expectations beside the reverse-DCF hurdles.
"""

from __future__ import annotations

import math
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from consensus_sources import collect_consensus

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; GOLD="FFF2CC"; PALE_GREEN="E2F0D9"; PALE_RED="FCE4D6"; GREY="666666"; LINK_GREEN="008000"
FMT_PCT='0.0%;[Red](0.0%);-'; FMT_BN='#,##0.0;[Red](#,##0.0);-'; FMT_PRICE='$#,##0.00;[Red]($#,##0.00);-'; FMT_INT='0'
THIN=Side(style="thin",color="D9E1F2")


def _fill(c): return PatternFill("solid",fgColor=c)

def _num(v,default=None):
    try:
        if isinstance(v,bool) or v in (None,""): return default
        x=float(v); return x if math.isfinite(x) else default
    except Exception: return default

def _latest_year(wb):
    if "Historical Financials" not in wb.sheetnames: return None
    ws=wb["Historical Financials"]
    ys=[int(ws.cell(3,c).value) for c in range(2,8) if isinstance(ws.cell(3,c).value,(int,float))]
    return max(ys) if ys else None

def _section(ws,row,title,end=13):
    for c in range(1,end+1): ws.cell(row,c).fill=_fill(NAVY); ws.cell(row,c).font=Font(bold=True,color=WHITE,size=11)
    ws.cell(row,1,title)

def _header(ws,row,headers):
    for c,v in enumerate(headers,1):
        x=ws.cell(row,c,v); x.fill=_fill(BLUE); x.font=Font(bold=True,color=WHITE); x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); x.border=Border(bottom=THIN)

def _find_row(ws,label,col=1):
    target=str(label).strip().lower()
    for r in range(1,ws.max_row+1):
        if str(ws.cell(r,col).value or "").strip().lower()==target: return r
    return None

def _row_map(ws):
    out={}
    for r in range(7,min(ws.max_row,80)+1):
        metric=str(ws.cell(r,1).value or "").strip(); year=ws.cell(r,2).value
        if metric and isinstance(year,(int,float)): out[(metric,int(year))]=r
    return out

def _range_dispersion(item):
    if not item: return None
    mean=_num(item.get("mean")); low=_num(item.get("low")); high=_num(item.get("high"))
    if mean in (None,0) or low is None or high is None: return item.get("dispersion")
    return (high-low)/abs(mean)

def _derived(blended,metric,year):
    if metric=="EBIT Margin":
        rev=blended.get(("Revenue",year)); ebit=blended.get(("EBIT",year))
        if rev and ebit and _num(rev.get("mean")) not in (None,0) and _num(ebit.get("mean")) is not None:
            rm=_num(rev["mean"]); em=_num(ebit["mean"])
            lo=None; hi=None
            if _num(ebit.get("low")) is not None and _num(rev.get("high")) not in (None,0): lo=_num(ebit["low"])/_num(rev["high"])
            if _num(ebit.get("high")) is not None and _num(rev.get("low")) not in (None,0): hi=_num(ebit["high"])/_num(rev["low"])
            return {"mean":em/rm,"low":lo,"high":hi,"analysts":max(_num(rev.get("analysts"),0),_num(ebit.get("analysts"),0)) or None,"provider_count":min(rev.get("provider_count",1),ebit.get("provider_count",1)),"providers":f"Derived from {ebit.get('providers','')} EBIT / {rev.get('providers','')} Revenue"}
    if metric=="Capex / Revenue":
        rev=blended.get(("Revenue",year)); cap=blended.get(("Capex",year))
        if rev and cap and _num(rev.get("mean")) not in (None,0) and _num(cap.get("mean")) is not None:
            return {"mean":abs(_num(cap["mean"]))/_num(rev["mean"]),"low":None,"high":None,"analysts":max(_num(rev.get("analysts"),0),_num(cap.get("analysts"),0)) or None,"provider_count":min(rev.get("provider_count",1),cap.get("provider_count",1)),"providers":f"Derived from {cap.get('providers','')} Capex / {rev.get('providers','')} Revenue"}
    return blended.get((metric,year))

def _eps_revisions(result,year):
    rows=[r for r in result["observations"] if r.get("metric")=="EPS Revision" and r.get("fiscal_year")==year]
    if not rows: return None,None
    r=rows[0]; cur=_num(r.get("mean")); d30=_num(r.get("low")); d90=_num(r.get("high"))
    rev30=(cur/d30-1) if cur not in (None,0) and d30 not in (None,0) else None
    rev90=(cur/d90-1) if cur not in (None,0) and d90 not in (None,0) else None
    return rev30,rev90


def _write_expectations(wb,result):
    if "Expectations & Consensus" not in wb.sheetnames: return {"updated":0}
    ws=wb["Expectations & Consensus"]; rows=_row_map(ws); blended=result["blended"]
    ws["A3"]="Consensus is a best-effort multi-source stack. Yahoo works without API keys; FMP, Alpha Vantage and Finnhub activate when keys/plan access are configured. Aggregator universes can overlap, so provider analyst counts are not summed."
    ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True)
    ws.cell(6,12,"Provider Count"); ws.cell(6,13,"Providers / Provenance")
    for c in (12,13): ws.cell(6,c).fill=_fill(BLUE); ws.cell(6,c).font=Font(bold=True,color=WHITE); ws.cell(6,c).alignment=Alignment(horizontal="center",wrap_text=True)
    updated=0
    for (metric,year),r in rows.items():
        item=_derived(blended,metric,year)
        if not item or _num(item.get("mean")) is None: continue
        ws.cell(r,3,item["mean"])
        fmt=FMT_PCT if metric in {"EBIT Margin","Capex / Revenue"} else FMT_PRICE if metric=="EPS" else FMT_BN
        ws.cell(r,3).number_format=fmt
        ws.cell(r,9,_range_dispersion(item)); ws.cell(r,9).number_format=FMT_PCT
        ws.cell(r,12,item.get("provider_count")); ws.cell(r,12).number_format=FMT_INT
        ws.cell(r,13,item.get("providers")); ws.cell(r,13).alignment=Alignment(wrap_text=True,vertical="top")
        ws.cell(r,11,f"Multi-source reference — {item.get('providers')}. Provider-level observations below; no analyst-count summation across aggregators.")
        ws.cell(r,11).alignment=Alignment(wrap_text=True,vertical="top")
        if metric=="EPS":
            r30,r90=_eps_revisions(result,year); ws.cell(r,7,r30); ws.cell(r,8,r90)
            ws.cell(r,7).number_format=FMT_PCT; ws.cell(r,8).number_format=FMT_PCT
        updated+=1

    start=ws.max_row+2
    _section(ws,start,"Consensus Provider Status & API Coverage",13)
    _header(ws,start+1,["Provider","Configured","Usable Now","Detail","Source / Documentation"])
    for rr,st in enumerate(result["statuses"],start+2):
        vals=[st.get("provider"),"YES" if st.get("configured") else "NO","YES" if st.get("ok") else "NO",st.get("detail"),st.get("source_url")]
        for c,v in enumerate(vals,1): ws.cell(rr,c,v); ws.cell(rr,c).alignment=Alignment(wrap_text=True,vertical="top")
        ws.cell(rr,3).fill=_fill(PALE_GREEN if st.get("ok") else GOLD)
        if st.get("source_url"): ws.cell(rr,5).hyperlink=st["source_url"]; ws.cell(rr,5).font=Font(color=LINK_GREEN,underline="single")

    prov=start+len(result["statuses"])+4
    _section(ws,prov,"Consensus Source Cross-Check — Provider-Level Observations",13)
    _header(ws,prov+1,["Fiscal Year","Metric","Provider","Mean","Low","High","Analysts","Provider Spread","Fetched At","Source URL","Notes"])
    for rr,rec in enumerate(sorted(result["observations"],key=lambda x:(x.get("fiscal_year") or 0,x.get("metric") or "",x.get("provider") or "")),prov+2):
        vals=[rec.get("fiscal_year"),rec.get("metric"),rec.get("provider"),rec.get("mean"),rec.get("low"),rec.get("high"),rec.get("analysts"),None,rec.get("fetched_at"),rec.get("source_url"),rec.get("note")]
        for c,v in enumerate(vals,1): ws.cell(rr,c,v)
        mean=_num(rec.get("mean")); low=_num(rec.get("low")); high=_num(rec.get("high"))
        ws.cell(rr,8,((high-low)/abs(mean)) if mean not in (None,0) and low is not None and high is not None else None); ws.cell(rr,8).number_format=FMT_PCT
        fmt=FMT_PRICE if rec.get("metric") in {"EPS","EPS Revision"} else FMT_BN
        for c in (4,5,6): ws.cell(rr,c).number_format=fmt
        if rec.get("source_url"): ws.cell(rr,10).hyperlink=rec["source_url"]; ws.cell(rr,10).font=Font(color=LINK_GREEN,underline="single")
        for c in (2,3,9,10,11): ws.cell(rr,c).alignment=Alignment(wrap_text=True,vertical="top")
    ws.column_dimensions["L"].width=16; ws.column_dimensions["M"].width=48
    return {"updated":updated,"provider_rows":len(result["observations"])}


def _write_market_expectations(wb,result):
    if "Market Expectations" not in wb.sheetnames: return False
    ws=wb["Market Expectations"]
    # Existing sheet is regenerated each run, so append once.
    start=ws.max_row+2
    _section(ws,start,"External Consensus Cross-Check — Street Estimates vs Market-Implied Hurdles",12)
    _header(ws,start+1,["Metric","Fiscal Year / Horizon","External Consensus","Your Model / Market Hurdle","Gap","Provider Count","Providers","Interpretation"])
    exp=wb["Expectations & Consensus"] if "Expectations & Consensus" in wb.sheetnames else None
    rr=start+2
    if exp:
        rows=_row_map(exp)
        years=sorted({y for _,y in rows})[:3]
        for year in years:
            for metric in ("Revenue","EPS","EBIT Margin","FCF","Capex / Revenue"):
                r=rows.get((metric,year))
                if not r: continue
                cons=_num(exp.cell(r,3).value); model=_num(exp.cell(r,4).value)
                if cons is None and model is None: continue
                gap=(model-cons) if metric in {"EBIT Margin","Capex / Revenue"} and cons is not None and model is not None else ((model/cons-1) if cons not in (None,0) and model is not None else None)
                vals=[metric,year,cons,model,gap,exp.cell(r,12).value,exp.cell(r,13).value,"Near-term analyst expectations are evidence; reverse DCF remains the price-implied long-duration hurdle."]
                for c,v in enumerate(vals,1): ws.cell(rr,c,v)
                fmt=FMT_PCT if metric in {"EBIT Margin","Capex / Revenue"} else FMT_PRICE if metric=="EPS" else FMT_BN
                ws.cell(rr,3).number_format=fmt; ws.cell(rr,4).number_format=fmt; ws.cell(rr,5).number_format=FMT_PCT
                for c in (7,8): ws.cell(rr,c).alignment=Alignment(wrap_text=True,vertical="top")
                rr+=1
    for pt in result.get("price_targets",[]):
        vals=["Analyst Price Target", "Current", pt.get("mean") or pt.get("median"), None, None, 1, pt.get("provider"), f"Provider range: low {pt.get('low')} / median {pt.get('median')} / high {pt.get('high')}. Price targets are sentiment/expectations context, not intrinsic value."]
        for c,v in enumerate(vals,1): ws.cell(rr,c,v)
        ws.cell(rr,3).number_format=FMT_PRICE; ws.cell(rr,8).alignment=Alignment(wrap_text=True); rr+=1
    ws.column_dimensions["G"].width=max(ws.column_dimensions["G"].width or 0,34); ws.column_dimensions["H"].width=max(ws.column_dimensions["H"].width or 0,58)
    return True


def _write_quality(wb,result):
    if "Data Quality" not in wb.sheetnames: return
    ws=wb["Data Quality"]
    good=[s for s in result["statuses"] if s.get("ok")]; configured=[s for s in result["statuses"] if s.get("configured")]
    status="PASS" if good else "REVIEW"
    detail=f"{len(good)}/{len(result['statuses'])} provider(s) returned usable consensus; {len(configured)} configured. Yahoo is zero-config; optional API sources improve coverage."
    existing=None
    for r in range(1,ws.max_row+1):
        if str(ws.cell(r,1).value or "").strip()=="Consensus provider coverage": existing=r; break
    r=existing or ws.max_row+1; ws.cell(r,1,"Consensus provider coverage"); ws.cell(r,2,status); ws.cell(r,3,detail); ws.cell(r,4,"Do not interpret several aggregators as independent analyst samples; compare provider levels/ranges and preserve provenance.")
    ws.cell(r,2).fill=_fill(PALE_GREEN if status=="PASS" else GOLD); ws.cell(r,2).font=Font(bold=True)
    for c in range(1,5): ws.cell(r,c).alignment=Alignment(wrap_text=True,vertical="top")


def normalize_expectations_consensus(wb,ticker,info=None):
    latest=_latest_year(wb)
    result=collect_consensus(ticker,latest,info or {})
    exp=_write_expectations(wb,result)
    market=_write_market_expectations(wb,result)
    _write_quality(wb,result)
    return {"consensus":result,"expectations":exp,"market_expectations":market}
