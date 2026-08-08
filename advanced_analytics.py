"""Advanced valuation, expectations and visualization analytics for the equity research workbook."""

import math
import random
import statistics
from datetime import datetime

import yfinance as yf
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import LineChart, BarChart, RadarChart, Reference
from openpyxl.formatting.rule import DataBarRule

NAVY = "17365D"
BLUE = "2F75B5"
WHITE = "FFFFFF"
LIGHT = "F5F9FC"
PALE_BLUE = "D9EAF7"
GOLD = "FFF2CC"
GREY = "666666"

FMT_PCT = '0.0%;[Red](0.0%);-'
FMT_PRICE = '$#,##0.00;[Red]($#,##0.00);-'
FMT_BN = '#,##0.0;[Red](#,##0.0);-'
FMT_MULT = '0.0x;[Red](0.0x);-'


def _fill(color):
    return PatternFill("solid", fgColor=color)


def _title(ws, text):
    ws.merge_cells("A1:P2")
    ws["A1"] = text
    ws["A1"].fill = _fill(NAVY)
    ws["A1"].font = Font(bold=True, color=WHITE, size=18)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.sheet_view.showGridLines = False


def _section(ws, cell_range, title):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = title
    cell.fill = _fill(NAVY)
    cell.font = Font(bold=True, color=WHITE, size=12)


def _header(ws, row, start_col, end_col):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row, c)
        cell.fill = _fill(BLUE)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _safe_float(v, default=None):
    try:
        return float(v)
    except Exception:
        return default


def _latest_value(ws, row, start_col=2, end_col=7):
    for c in range(end_col, start_col - 1, -1):
        v = ws.cell(row, c).value
        if isinstance(v, (int, float)):
            return float(v)
    return None


def _year_end_prices(ticker, years):
    try:
        hist = yf.Ticker(ticker).history(period="10y", interval="1d", auto_adjust=False)
        if hist is None or hist.empty:
            return {}
        out = {}
        idx_years = hist.index.year
        for y in years:
            rows = hist[idx_years == y]
            if not rows.empty:
                out[y] = float(rows["Close"].dropna().iloc[-1])
        return out
    except Exception:
        return {}


def _earnings_history(ticker, limit=8):
    try:
        df = yf.Ticker(ticker).get_earnings_dates(limit=limit)
        if df is None or df.empty:
            return []
        out = []
        for idx, row in df.sort_index().tail(limit).iterrows():
            est = _safe_float(row.get("EPS Estimate"))
            actual = _safe_float(row.get("Reported EPS"))
            surprise = _safe_float(row.get("Surprise(%)"))
            if surprise is not None and abs(surprise) > 2:
                surprise = surprise / 100.0
            out.append({
                "date": idx.to_pydatetime().date() if hasattr(idx, "to_pydatetime") else idx,
                "eps_est": est,
                "eps_actual": actual,
                "eps_surprise": surprise,
            })
        return out
    except Exception:
        return []


def _solve_reverse_dcf(latest_fcf, target_ev, wacc, tgr, years=10):
    if not latest_fcf or not target_ev or wacc <= tgr:
        return None

    def ev_for(g):
        f = latest_fcf
        pv = 0.0
        for t in range(1, years + 1):
            f *= 1 + g
            pv += f / ((1 + wacc) ** t)
        tv = f * (1 + tgr) / (wacc - tgr)
        return pv + tv / ((1 + wacc) ** years)

    lo, hi = -0.50, 0.75
    for _ in range(120):
        mid = (lo + hi) / 2
        if ev_for(mid) < target_ev:
            lo = mid
        else:
            hi = mid
    return (lo + hi) / 2


def _monte_carlo(wb, n=5000, seed=42):
    scenarios = wb["Three-Case Scenarios"]
    hist = wb["Historical Financials"]
    company = wb["Company Data"]

    base_growth = [float(scenarios.cell(12, c).value) for c in range(14, 24)]
    base_margin = [float(scenarios.cell(14, c).value) for c in range(14, 24)]
    base_da = [float(scenarios.cell(18, c).value) for c in range(14, 24)]
    base_capex = [float(scenarios.cell(20, c).value) for c in range(14, 24)]
    tax = _safe_float(scenarios["C8"].value, 0.21)
    latest_revenue = _latest_value(hist, 4) or 1.0
    shares = _safe_float(company["B9"].value, 1.0)
    net_debt = _safe_float(company["B14"].value, 0.0)

    random.seed(seed)
    values = []
    for _ in range(n):
        gshock = max(-0.07, min(0.07, random.gauss(0, 0.025)))
        mshock = max(-0.06, min(0.06, random.gauss(0, 0.025)))
        cshock = max(-0.06, min(0.08, random.gauss(0, 0.03)))
        sim_w = max(0.07, min(0.12, random.gauss(_safe_float(scenarios["C6"].value, 0.09), 0.01)))
        sim_t = max(0.01, min(0.045, random.gauss(_safe_float(scenarios["C7"].value, 0.03), 0.005)))
        if sim_t >= sim_w - 0.002:
            sim_t = sim_w - 0.002

        rev = latest_revenue
        prev = rev
        pv = 0.0
        fcf = 0.0
        for t in range(10):
            rev *= 1 + max(-0.50, base_growth[t] + gshock)
            nopat = rev * max(0.05, base_margin[t] + mshock) * (1 - tax)
            da = rev * base_da[t]
            capex = rev * max(0.03, base_capex[t] + cshock)
            nwc = (rev - prev) * 0.01
            fcf = nopat + da - capex - nwc
            pv += fcf / ((1 + sim_w) ** (t + 1))
            prev = rev
        tv = fcf * (1 + sim_t) / (sim_w - sim_t)
        ev = pv + tv / ((1 + sim_w) ** 10)
        equity = ev - net_debt
        values.append(equity / shares)
    values.sort()
    return values


def _scorecard(wb, current_price, forward_pe):
    hist = wb["Historical Financials"]
    scenarios = wb["Three-Case Scenarios"]
    peers = wb["Peer Comps"]

    rev_first = _safe_float(hist["B4"].value)
    rev_last = _safe_float(hist["G4"].value)
    revenue_cagr = ((rev_last / rev_first) ** (1 / 5) - 1) if rev_first and rev_last else 0
    op_margin = _safe_float(hist["G10"].value, 0)
    fcf_margin = _safe_float(hist["G17"].value, 0)
    base_value = _safe_float(scenarios["C39"].value, 0)
    severe_bear = _safe_float(scenarios["G58"].value, 0)

    peer_pes = []
    for r in range(5, 10):
        v = _safe_float(peers.cell(r, 3).value)
        if v:
            peer_pes.append(v)
    peer_med = statistics.median(peer_pes) if peer_pes else forward_pe or 1

    scores = [
        ("Growth", min(100, max(0, revenue_cagr / 0.20 * 100)), "5Y revenue CAGR"),
        ("Profitability", min(100, max(0, op_margin / 0.40 * 100)), "Operating margin"),
        ("FCF Quality", min(100, max(0, fcf_margin / 0.25 * 100)), "FCF margin"),
        ("Balance Sheet", 85.0 if _safe_float(wb["Company Data"]["B14"].value, 0) < 0 else 55.0, "Net cash / debt profile"),
        ("Absolute Valuation", min(100, max(0, base_value / current_price * 100)) if current_price else 0, "Base DCF vs market price"),
        ("Relative Valuation", min(100, max(0, peer_med / forward_pe * 70)) if forward_pe else 50, "Forward P/E vs peer median"),
        ("Stress Robustness", min(100, max(0, severe_bear / current_price * 100)) if current_price else 0, "Severe-bear value vs market"),
    ]
    return scores


def ensure_advanced_analytics(wb, ticker, info=None):
    """Create price/valuation history, earnings surprises, reverse DCF, Monte Carlo, and a diagnostic scorecard."""
    required = {"Company Data", "Historical Financials", "Three-Case Scenarios", "Peer Comps"}
    if not required.issubset(set(wb.sheetnames)):
        return None
    if "Advanced Analytics" in wb.sheetnames:
        wb.remove(wb["Advanced Analytics"])
    ws = wb.create_sheet("Advanced Analytics")
    _title(ws, "Advanced Valuation & Expectations Analytics")
    ws.merge_cells("A3:P3")
    ws["A3"] = "Historical valuation, earnings surprises, reverse DCF, Monte Carlo valuation and investment scorecard."
    ws["A3"].font = Font(italic=True, color=GREY)

    company = wb["Company Data"]
    hist = wb["Historical Financials"]
    scenarios = wb["Three-Case Scenarios"]

    current_price = _safe_float(company["B8"].value, 0)
    shares = _safe_float(company["B9"].value, 1)
    net_debt = _safe_float(company["B14"].value, 0)
    forward_pe = _safe_float(company["B15"].value, _safe_float((info or {}).get("forwardPE"), 0))

    # Historical year-end price / P-E.
    _section(ws, "A5:H5", "Historical Price & Valuation")
    years = []
    eps_by_year = {}
    for c in range(2, 8):
        y = hist.cell(3, c).value
        eps = hist.cell(12, c).value
        if isinstance(y, (int, float)):
            y = int(y)
            years.append(y)
            if isinstance(eps, (int, float)):
                eps_by_year[y] = float(eps)
    prices = _year_end_prices(ticker, years)
    ws.append([])
    for c, v in enumerate(["Year", "Year-End Price", "Diluted EPS", "Year-End P/E"], 1):
        ws.cell(6, c, v)
    _header(ws, 6, 1, 4)
    pe_vals = []
    row = 7
    for y in years:
        px = prices.get(y)
        eps = eps_by_year.get(y)
        pe = px / eps if px and eps else None
        if pe:
            pe_vals.append(pe)
        ws.cell(row, 1, y)
        ws.cell(row, 2, px)
        ws.cell(row, 3, eps)
        ws.cell(row, 4, pe)
        ws.cell(row, 2).number_format = FMT_PRICE
        ws.cell(row, 3).number_format = FMT_PRICE
        ws.cell(row, 4).number_format = FMT_MULT
        row += 1

    stats = [
        ("Historical Min", min(pe_vals) if pe_vals else None),
        ("Historical Median", statistics.median(pe_vals) if pe_vals else None),
        ("Historical Max", max(pe_vals) if pe_vals else None),
        ("Current Forward P/E", forward_pe),
    ]
    ws["F6"], ws["G6"] = "Valuation Statistic", "P/E"
    _header(ws, 6, 6, 7)
    for r, (label, value) in enumerate(stats, 7):
        ws.cell(r, 6, label)
        ws.cell(r, 7, value)
        ws.cell(r, 7).number_format = FMT_MULT

    if row > 8:
        chart = LineChart()
        chart.style = 10
        chart.title = "Year-End Share Price"
        chart.height = 7
        chart.width = 13
        chart.legend = None
        chart.add_data(Reference(ws, min_col=2, min_row=6, max_row=row - 1), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=7, max_row=row - 1))
        ws.add_chart(chart, "A14")

    # Earnings surprise.
    _section(ws, "I5:P5", "Recent Earnings Surprises")
    headers = ["Reported", "EPS Estimate", "Actual EPS", "EPS Surprise"]
    for c, v in enumerate(headers, 9):
        ws.cell(6, c, v)
    _header(ws, 6, 9, 12)
    earnings = _earnings_history(ticker, limit=8)
    erow = 7
    for e in earnings:
        ws.cell(erow, 9, e["date"])
        ws.cell(erow, 10, e["eps_est"])
        ws.cell(erow, 11, e["eps_actual"])
        ws.cell(erow, 12, e["eps_surprise"])
        ws.cell(erow, 10).number_format = FMT_PRICE
        ws.cell(erow, 11).number_format = FMT_PRICE
        ws.cell(erow, 12).number_format = FMT_PCT
        erow += 1
    if erow == 7:
        ws["I7"] = "No earnings-surprise history returned by data provider."
    else:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "EPS Surprise %"
        chart.height = 7
        chart.width = 13
        chart.legend = None
        chart.add_data(Reference(ws, min_col=12, min_row=6, max_row=erow - 1), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=9, min_row=7, max_row=erow - 1))
        chart.y_axis.numFmt = "0%"
        ws.add_chart(chart, "I14")

    # Reverse DCF.
    _section(ws, "A31:H31", "Reverse DCF — Market-Implied Expectations")
    latest_fcf = _latest_value(hist, 16)
    wacc = _safe_float(scenarios["C6"].value, 0.09)
    tgr = _safe_float(scenarios["C7"].value, 0.03)
    target_ev = current_price * shares + net_debt
    implied = _solve_reverse_dcf(latest_fcf, target_ev, wacc, tgr)
    reverse_rows = [
        ("Current Price", current_price, FMT_PRICE, "Market price"),
        ("Current Enterprise Value ($bn)", target_ev, FMT_BN, "Equity value plus net debt"),
        ("Latest FCF ($bn)", latest_fcf, FMT_BN, "Latest annual free cash flow"),
        ("WACC", wacc, FMT_PCT, "Base discount rate"),
        ("Terminal Growth", tgr, FMT_PCT, "Base perpetual growth"),
        ("Implied 10Y FCF CAGR", implied, FMT_PCT, "Constant FCF growth required to justify current price"),
    ]
    for c, v in enumerate(["Input / Output", "Value", "Interpretation"], 1):
        ws.cell(32, c, v)
    _header(ws, 32, 1, 3)
    for r, (label, value, fmt, note) in enumerate(reverse_rows, 33):
        ws.cell(r, 1, label)
        ws.cell(r, 2, value)
        ws.cell(r, 2).number_format = fmt
        ws.cell(r, 3, note)
    ws["B38"].fill = _fill(GOLD)
    ws["B38"].font = Font(bold=True)

    # Monte Carlo valuation.
    _section(ws, "I31:P31", "Monte Carlo Valuation — 5,000 Simulations")
    values = _monte_carlo(wb, n=5000)
    if values:
        def pct(p):
            return values[int((len(values) - 1) * p)]
        summary = [
            ("P10 Value / Share", pct(0.10), FMT_PRICE),
            ("P25 Value / Share", pct(0.25), FMT_PRICE),
            ("Median Value / Share", pct(0.50), FMT_PRICE),
            ("P75 Value / Share", pct(0.75), FMT_PRICE),
            ("P90 Value / Share", pct(0.90), FMT_PRICE),
            ("Probability > Current Price", sum(v > current_price for v in values) / len(values), FMT_PCT),
            ("Probability > Base DCF", sum(v > _safe_float(scenarios["C39"].value, 0) for v in values) / len(values), FMT_PCT),
        ]
        ws["I32"], ws["J32"] = "Metric", "Result"
        _header(ws, 32, 9, 10)
        for r, (label, val, fmt) in enumerate(summary, 33):
            ws.cell(r, 9, label)
            ws.cell(r, 10, val)
            ws.cell(r, 10).number_format = fmt

        lo, hi = min(values), max(values)
        bins = 20
        width = (hi - lo) / bins if hi > lo else 1
        ws["R32"], ws["S32"] = "Value / Share", "Frequency"
        for i in range(bins):
            left = lo + i * width
            right = left + width
            count = sum(1 for v in values if v >= left and (v < right or i == bins - 1))
            ws.cell(33 + i, 18, (left + right) / 2)
            ws.cell(33 + i, 19, count)
            ws.cell(33 + i, 18).number_format = FMT_PRICE
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Monte Carlo Valuation Distribution"
        chart.height = 8
        chart.width = 13
        chart.legend = None
        chart.add_data(Reference(ws, min_col=19, min_row=32, max_row=52), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=18, min_row=33, max_row=52))
        ws.add_chart(chart, "I41")

    # Investment scorecard + radar.
    _section(ws, "A41:H41", "Investment Scorecard")
    scores = _scorecard(wb, current_price, forward_pe)
    for c, v in enumerate(["Dimension", "Score (0–100)", "Comment"], 1):
        ws.cell(42, c, v)
    _header(ws, 42, 1, 3)
    for r, (name, score, note) in enumerate(scores, 43):
        ws.cell(r, 1, name)
        ws.cell(r, 2, score)
        ws.cell(r, 3, note)
    if scores:
        composite = sum(s for _, s, _ in scores) / len(scores)
        ws["E42"] = "Composite Score"
        ws["F42"] = composite
        ws["F42"].number_format = "0.0"
        ws["E42"].font = Font(bold=True)
        ws.conditional_formatting.add("B43:B49", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color="5B9BD5"))

        radar = RadarChart()
        radar.type = "filled"
        radar.style = 26
        radar.title = "Investment Scorecard Radar"
        radar.height = 8
        radar.width = 11
        radar.add_data(Reference(ws, min_col=2, min_row=42, max_row=42 + len(scores)), titles_from_data=True)
        radar.set_categories(Reference(ws, min_col=1, min_row=43, max_row=42 + len(scores)))
        ws.add_chart(radar, "A52")

    ws.merge_cells("A70:P72")
    ws["A70"] = (
        "Live sources: Yahoo Finance for price history and earnings dates; SEC-based workbook inputs for reported financials. "
        "Monte Carlo and scorecard results are analytical diagnostics, not investment recommendations."
    )
    ws["A70"].fill = _fill(PALE_BLUE)
    ws["A70"].alignment = Alignment(wrap_text=True, vertical="center")

    ws.column_dimensions["A"].width = 26
    for col in "BCDEFGHIJKLMNOP":
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["C"].width = 33
    ws.column_dimensions["R"].hidden = True
    ws.column_dimensions["S"].hidden = True
    return ws
