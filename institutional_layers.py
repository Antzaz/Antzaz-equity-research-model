"""Institutional-style expectations, moat, base-rate and market-expectations layers.

These sheets are intentionally transparent. Public/free data cannot reproduce a Bloomberg,
FactSet or Visible Alpha consensus stack, so consensus/revision fields remain editable when
a reliable public snapshot is unavailable. The workbook distinguishes sourced observations,
internal model outputs and user assumptions.
"""

import math
import statistics
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

try:
    import yfinance as yf
except Exception:
    yf = None

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; LIGHT="F5F9FC"; GOLD="FFF2CC"
PALE_GREEN="E2F0D9"; PALE_RED="FCE4D6"; GREY="666666"; INPUT_BLUE="0000FF"; LINK_GREEN="008000"
FMT_PCT='0.0%;[Red](0.0%);-'; FMT_PRICE='$#,##0.00;[Red]($#,##0.00);-'; FMT_BN='#,##0.0;[Red](#,##0.0);-'
FMT_MULT='0.0x;[Red](0.0x);-'; THIN=Side(style="thin",color="D9E1F2")

def _fill(c): return PatternFill("solid",fgColor=c)
def _num(v,default=None):
    try:
        if isinstance(v,bool): return default
        return float(v)
    except Exception:
        return default
def _create(wb,name):
    if name in wb.sheetnames: wb.remove(wb[name])
    ws=wb.create_sheet(name); ws.sheet_view.showGridLines=False
    return ws
def _title(ws,text,end="L"):
    for c in range(1,ord(end)-64+1):
        ws.cell(1,c).fill=_fill(NAVY); ws.cell(2,c).fill=_fill(NAVY)
    ws["A1"]=text; ws["A1"].font=Font(bold=True,color=WHITE,size=18)
def _section(ws,row,title,end=12):
    for c in range(1,end+1):
        ws.cell(row,c).fill=_fill(NAVY); ws.cell(row,c).font=Font(bold=True,color=WHITE,size=11)
    ws.cell(row,1,title)
def _header(ws,row,start,end):
    for c in range(start,end+1):
        x=ws.cell(row,c); x.fill=_fill(BLUE); x.font=Font(bold=True,color=WHITE)
        x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); x.border=Border(bottom=THIN)
def _input(cell,fmt=None):
    cell.fill=_fill(GOLD); cell.font=Font(color=INPUT_BLUE)
    if fmt: cell.number_format=fmt
def _link(cell,fmt=None):
    cell.font=Font(color=LINK_GREEN)
    if fmt: cell.number_format=fmt
def _company(wb):
    d=wb["Company Data"]
    price=_num(d["B8"].value,0); mc=_num(d["B10"].value,0); cash=_num(d["B12"].value,0); debt=_num(d["B13"].value,0)
    shares=_num(d["B9"].value) or ((mc/price) if price and mc else 1)
    return price,shares,debt-cash
def _base_projection(wb):
    if not {"Historical Financials","Three-Case Scenarios"}.issubset(wb.sheetnames): return []
    h=wb["Historical Financials"]; s=wb["Three-Case Scenarios"]
    rev=_num(h["G4"].value)
    if not rev: return []
    tax=_num(s["C8"].value,.21); adj=_num(s["AL4"].value,0)
    out=[]; prev=rev
    for c in range(14,24):
        year=s.cell(11,c).value
        g=_num(s.cell(12,c).value,0); margin=_num(s.cell(14,c).value,0); da=_num(s.cell(18,c).value,0); cap=_num(s.cell(20,c).value,0)
        rev=rev*(1+g); nwc=(rev-prev)*.01
        fcf=rev*margin*(1-tax)+rev*da-rev*cap-nwc+rev*adj
        out.append({"year":year,"growth":g,"revenue":rev,"margin":margin,"da":da,"capex_pct":cap,"fcf":fcf})
        prev=rev
    return out
def _model_value(wb,g_shift=0,m_shift=0,cap_shift=0,wacc_override=None,tgr_override=None,years=10):
    s=wb["Three-Case Scenarios"]; h=wb["Historical Financials"]; _,shares,net_debt=_company(wb)
    rev=_num(h["G4"].value,1); prev=rev; tax=_num(s["C8"].value,.21); adj=_num(s["AL4"].value,0)
    wacc=_num(wacc_override,_num(s["C6"].value,.09)); tgr=_num(tgr_override,_num(s["C7"].value,.03))
    wacc=max(.04,min(.20,wacc)); tgr=max(-.01,min(wacc-.005,tgr))
    pv=0; fcf=None; n=max(1,min(10,int(years)))
    for i,c in enumerate(range(14,14+n)):
        g=max(-.5,min(.6,_num(s.cell(12,c).value,0)+g_shift))
        margin=max(-.2,min(.7,_num(s.cell(14,c).value,0)+m_shift))
        da=max(0,min(.25,_num(s.cell(18,c).value,0)))
        cap=max(.002,min(.6,_num(s.cell(20,c).value,0)+cap_shift))
        rev*=1+g; nwc=(rev-prev)*.01
        fcf=rev*margin*(1-tax)+rev*da-rev*cap-nwc+rev*adj
        pv+=fcf/((1+wacc)**(i+1)); prev=rev
    if fcf is None or shares<=0: return None
    tv=fcf*(1+tgr)/(wacc-tgr); ev=pv+tv/((1+wacc)**n)
    return (ev-net_debt)/shares
def _bisect(fn,target,lo,hi,iterations=80):
    flo=fn(lo); fhi=fn(hi)
    if flo is None or fhi is None: return None
    if (flo-target)*(fhi-target)>0: return None
    for _ in range(iterations):
        mid=(lo+hi)/2; fm=fn(mid)
        if fm is None: return None
        if abs(fm-target)<1e-6: return mid
        if (flo-target)*(fm-target)<=0: hi=mid; fhi=fm
        else: lo=mid; flo=fm
    return (lo+hi)/2
def _implied_fcf_growth(wb):
    h=wb["Historical Financials"]; s=wb["Three-Case Scenarios"]; price,shares,net_debt=_company(wb)
    ocf=_num(h["G14"].value); cap=_num(h["G15"].value)
    if ocf is None or cap is None or shares<=0: return None
    fcf0=ocf-cap; target_ev=price*shares+net_debt; wacc=_num(s["C6"].value,.09); tgr=_num(s["C7"].value,.03)
    def ev(g):
        f=fcf0; pv=0
        for t in range(1,11):
            f*=1+g; pv+=f/((1+wacc)**t)
        tv=f*(1+tgr)/(wacc-tgr)
        return pv+tv/((1+wacc)**10)
    return _bisect(ev,target_ev,-.2,.8)
def _implied_duration(wb):
    h=wb["Historical Financials"]; s=wb["Three-Case Scenarios"]; price,shares,net_debt=_company(wb)
    ocf=_num(h["G14"].value); cap=_num(h["G15"].value)
    if ocf is None or cap is None or shares<=0: return None
    fcf0=ocf-cap; target=price*shares+net_debt; wacc=_num(s["C6"].value,.09); tgr=_num(s["C7"].value,.03)
    proj=_base_projection(wb); g=statistics.mean([x["growth"] for x in proj[:5]]) if proj else .10
    g=max(tgr+.005,min(.30,g)); best=None
    for n in range(1,31):
        f=fcf0; pv=0
        for t in range(1,n+1):
            f*=1+g; pv+=f/((1+wacc)**t)
        tv=f*(1+tgr)/(wacc-tgr); ev=pv+tv/((1+wacc)**n); gap=abs(ev-target)
        if best is None or gap<best[0]: best=(gap,n,ev)
    return best[1] if best else None
def _consensus_snapshot(ticker):
    out={}
    if yf is None: return out
    try:
        t=yf.Ticker(ticker)
        for kind,attr in (("Revenue","revenue_estimate"),("EPS","earnings_estimate")):
            df=getattr(t,attr,None)
            if df is None or getattr(df,"empty",True): continue
            for idx in ("0y","+1y"):
                if idx not in df.index: continue
                row=df.loc[idx]
                out[(kind,idx)]={"avg":_num(row.get("avg")),"low":_num(row.get("low")),"high":_num(row.get("high")),"analysts":_num(row.get("numberOfAnalysts")),"growth":_num(row.get("growth"))}
        tr=getattr(t,"eps_trend",None)
        if tr is not None and not getattr(tr,"empty",True):
            for idx in ("0y","+1y"):
                if idx in tr.index:
                    row=tr.loc[idx]; out[("EPS Trend",idx)]={"current":_num(row.get("current")),"30d":_num(row.get("30daysAgo")),"90d":_num(row.get("90daysAgo"))}
    except Exception:
        pass
    return out
def _latest_year(wb):
    h=wb["Historical Financials"]; years=[h.cell(3,c).value for c in range(2,8)]; years=[int(y) for y in years if isinstance(y,(int,float))]
    return max(years) if years else None

def ensure_expectations_consensus(wb,ticker):
    ws=_create(wb,"Expectations & Consensus"); _title(ws,f"{ticker} — Expectations, Consensus & Variant Perception")
    ws["A3"]="Blue/yellow cells are editable. Public consensus is best-effort from Yahoo Finance; institutional users should replace it with point-in-time broker/FactSet/Bloomberg data."
    ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True)
    _section(ws,5,"Consensus vs Your Model"); heads=["Metric","Fiscal Year","Consensus","Your Model","Absolute Gap","Gap % / ppt","30d Revision","90d Revision","Consensus Range / Dispersion","Variant View","Source / Notes"]
    for c,v in enumerate(heads,1): ws.cell(6,c,v)
    _header(ws,6,1,11)
    proj=_base_projection(wb); ly=_latest_year(wb); snap=_consensus_snapshot(ticker); model={}
    for x in proj[:3]:
        y=int(x["year"]) if isinstance(x["year"],(int,float)) else None
        if y: model[("Revenue",y)]=x["revenue"]; model[("EBIT Margin",y)]=x["margin"]; model[("FCF",y)]=x["fcf"]; model[("Capex / Revenue",y)]=x["capex_pct"]
    years=[ly+1,ly+2,ly+3] if ly else [2026,2027,2028]; rows=[]
    for y in years: rows.extend([("Revenue",y,FMT_BN),("EPS",y,FMT_PRICE),("EBIT Margin",y,FMT_PCT),("FCF",y,FMT_BN),("Capex / Revenue",y,FMT_PCT)])
    r=7
    for metric,y,fmt in rows:
        ws.cell(r,1,metric); ws.cell(r,2,y); cons=None; rev30=None; rev90=None; dispersion=None; source="Manual / institutional data"
        idx="0y" if ly and y==ly+1 else ("+1y" if ly and y==ly+2 else None)
        if idx and metric in {"Revenue","EPS"} and (metric,idx) in snap:
            dat=snap[(metric,idx)]; cons=dat.get("avg"); lo=dat.get("low"); hi=dat.get("high")
            if metric=="Revenue" and cons is not None: cons/=1e9
            if metric=="Revenue" and lo is not None: lo/=1e9
            if metric=="Revenue" and hi is not None: hi/=1e9
            if cons not in (None,0) and lo is not None and hi is not None: dispersion=(hi-lo)/abs(cons)
            source="Yahoo Finance public consensus snapshot"
        if idx and metric=="EPS" and ("EPS Trend",idx) in snap:
            tr=snap[("EPS Trend",idx)]; cur=tr.get("current"); d30=tr.get("30d"); d90=tr.get("90d")
            if cur not in (None,0) and d30 is not None: rev30=cur/d30-1 if d30 else None
            if cur not in (None,0) and d90 is not None: rev90=cur/d90-1 if d90 else None
        ws.cell(r,3,cons); _input(ws.cell(r,3),fmt); mv=model.get((metric,y)); ws.cell(r,4,mv); _link(ws.cell(r,4),fmt)
        if metric in {"EBIT Margin","Capex / Revenue"}:
            ws.cell(r,5,f'=IF(OR(C{r}="",D{r}=""),"",D{r}-C{r})'); ws.cell(r,6,f'=E{r}'); ws.cell(r,5).number_format=FMT_PCT; ws.cell(r,6).number_format=FMT_PCT
        else:
            ws.cell(r,5,f'=IF(OR(C{r}="",D{r}=""),"",D{r}-C{r})'); ws.cell(r,6,f'=IFERROR(D{r}/C{r}-1,"")'); ws.cell(r,5).number_format=fmt; ws.cell(r,6).number_format=FMT_PCT
        ws.cell(r,7,rev30); ws.cell(r,8,rev90); ws.cell(r,9,dispersion)
        for c in (7,8,9): _input(ws.cell(r,c),FMT_PCT)
        ws.cell(r,10,""); _input(ws.cell(r,10)); ws.cell(r,11,source); r+=1
    ws.conditional_formatting.add(f"F7:F{r-1}",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="num",mid_value=0,mid_color="FFEB84",end_type="max",end_color="63BE7B"))
    _section(ws,r+1,"Variant Perception & Estimate Revision Thesis"); _header(ws,r+2,1,2); ws.cell(r+2,1,"Question"); ws.cell(r+2,2,"Your Analysis")
    questions=["Where is my forecast most different from consensus?","What operating driver explains that gap?","What does consensus appear to assume about margins/capital intensity?","Which estimate has the greatest positive revision potential?","Which estimate has the greatest negative revision risk?","What new evidence would force me back toward consensus?"]
    for rr,q in enumerate(questions,r+3):
        ws.cell(rr,1,q); _input(ws.cell(rr,2)); ws.cell(rr,2).alignment=Alignment(wrap_text=True,vertical="top"); ws.row_dimensions[rr].height=34
    ws.column_dimensions["A"].width=34; ws.column_dimensions["B"].width=13; ws.column_dimensions["C"].width=15; ws.column_dimensions["D"].width=15
    for c in range(5,10): ws.column_dimensions[get_column_letter(c)].width=14
    ws.column_dimensions["J"].width=25; ws.column_dimensions["K"].width=34; ws.freeze_panes="A7"; return ws

def _rd_adjustment(wb,useful_life=5):
    h=wb["Historical Financials"]; rd=[]; rev=[]; op=[]
    for c in range(2,8): rd.append(_num(h.cell(19,c).value)); rev.append(_num(h.cell(4,c).value)); op.append(_num(h.cell(9,c).value))
    valid=[x for x in rd if x is not None]
    if not valid: return None
    n=max(2,int(useful_life)); recent=valid[-n:]; asset=sum(v*(n-i)/n for i,v in enumerate(recent)); amort=sum(recent)/n
    latest_rd=valid[-1]; latest_op=next((x for x in reversed(op) if x is not None),None); latest_rev=next((x for x in reversed(rev) if x is not None),None)
    adj_ebit=latest_op+latest_rd-amort if latest_op is not None else None
    return asset,amort,adj_ebit,(adj_ebit/latest_rev if adj_ebit is not None and latest_rev else None)

def ensure_moat_cap(wb,ticker):
    ws=_create(wb,"Moat & Competitive Advantage"); _title(ws,f"{ticker} — Moat, ROIC & Competitive Advantage Period")
    ws["A3"]="Separate business quality from valuation. Moat scores are manual judgments; market-implied duration is a simplified cash-flow-duration cross-check, not a precise economic-moat forecast."
    ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True)
    _section(ws,5,"Moat Source Scorecard"); heads=["Moat Source","Score 1–5","Evidence","Durability / Trend","Key Threat","Confidence 1–5"]
    for c,v in enumerate(heads,1): ws.cell(6,c,v)
    _header(ws,6,1,6); sources=["Network effects","Switching costs","Cost advantage","Brand / IP / intangible assets","Efficient scale","Distribution / ecosystem / data"]
    dv=DataValidation(type="whole",operator="between",formula1="1",formula2="5",allow_blank=True); ws.add_data_validation(dv)
    for r,name in enumerate(sources,7):
        ws.cell(r,1,name)
        for c in (2,3,4,5,6): _input(ws.cell(r,c))
        dv.add(ws.cell(r,2)); dv.add(ws.cell(r,6)); ws.cell(r,3).alignment=Alignment(wrap_text=True); ws.cell(r,5).alignment=Alignment(wrap_text=True)
    ws["A14"]="Average Moat Score"; ws["B14"]='=IFERROR(AVERAGE(B7:B12),"")'; ws["B14"].number_format="0.0"; _link(ws["B14"])
    ws["D14"]="Average Confidence"; ws["E14"]='=IFERROR(AVERAGE(F7:F12),"")'; ws["E14"].number_format="0.0"; _link(ws["E14"])
    _section(ws,17,"Competitive Advantage Period / Fade"); s=wb["Three-Case Scenarios"]; h=wb["Historical Financials"]
    wacc=_num(s["C6"].value,.09); duration=_implied_duration(wb); rev0=_num(h["B4"].value); rev1=_num(h["G4"].value); hist_cagr=(rev1/rev0)**(1/5)-1 if rev0 and rev1 and rev0>0 else None
    metrics=[("WACC",wacc,FMT_PCT,"Discount-rate hurdle"),("5Y Revenue CAGR",hist_cagr,FMT_PCT,"Historical growth context"),("Market-Implied Growth Duration",duration,'0 "yrs"',"Years of elevated cash-flow growth that best reconcile current EV in a simplified duration model"),("Your Competitive Advantage Period",15,'0 "yrs"',"Editable estimate of how long excess returns can persist"),("Mature ROIC / RONIC Target",wacc,FMT_PCT,"Default assumes excess returns fade toward WACC")]
    _header(ws,18,1,4); ws["A18"]="Metric"; ws["B18"]="Value"; ws["C18"]="Type"; ws["D18"]="Interpretation"
    for r,(lab,val,fmt,note) in enumerate(metrics,19):
        ws.cell(r,1,lab); ws.cell(r,2,val); ws.cell(r,2).number_format=fmt; ws.cell(r,3,"Input" if r in (22,23) else "Model"); ws.cell(r,4,note)
        if r in (22,23): _input(ws.cell(r,2),fmt)
        else: _link(ws.cell(r,2),fmt)
    _section(ws,26,"Intangible / R&D Capitalization Cross-Check"); rd=_rd_adjustment(wb,5); ws["A27"]="R&D Useful Life (years)"; ws["B27"]=5; _input(ws["B27"],"0")
    labels=["Capitalized R&D Asset Proxy","Annual R&D Amortization Proxy","Adjusted EBIT","Adjusted Operating Margin"]; vals=list(rd) if rd else [None,None,None,None]
    for r,(lab,val) in enumerate(zip(labels,vals),28):
        fmt=FMT_PCT if "Margin" in lab else FMT_BN; ws.cell(r,1,lab); ws.cell(r,2,val); _link(ws.cell(r,2),fmt)
    ws["D27"]="Purpose"; ws["E27"]="Traditional accounting expenses R&D immediately. This section treats recent R&D as an economic investment to show how reported margins and capital can differ."; ws["E27"].alignment=Alignment(wrap_text=True)
    _section(ws,34,"Moat Thesis Questions"); qs=["Why should excess returns persist?","Which moat source is strengthening?","Which moat source is weakening?","What is the most credible disruption path?","What evidence would shorten my CAP estimate?","What evidence would extend it?"]
    for r,q in enumerate(qs,35): ws.cell(r,1,q); _input(ws.cell(r,2)); ws.cell(r,2).alignment=Alignment(wrap_text=True); ws.row_dimensions[r].height=30
    ws.column_dimensions["A"].width=38; ws.column_dimensions["B"].width=22; ws.column_dimensions["C"].width=18; ws.column_dimensions["D"].width=46; ws.column_dimensions["E"].width=46; ws.column_dimensions["F"].width=16; return ws

def ensure_base_rates(wb,ticker):
    ws=_create(wb,"Base Rates & Thesis Probabilities"); _title(ws,f"{ticker} — Base Rates & Bayesian Thesis Updating")
    ws["A3"]="Use reference-class evidence to anchor scenario probabilities. Likelihood-ratio updates assume evidence items are independent, which is rarely perfectly true; use this as a discipline tool, not false precision."
    ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True)
    _section(ws,5,"Scenario Priors and Evidence Updates"); heads=["Scenario","Prior Probability","Raw Updated Weight","Posterior Probability","Interpretation"]
    for c,v in enumerate(heads,1): ws.cell(6,c,v)
    _header(ws,6,1,5); scenarios=[("Bear","='Three-Case Scenarios'!B5"),("Base","='Three-Case Scenarios'!C5"),("Bull","='Three-Case Scenarios'!D5")]
    for r,(lab,form) in enumerate(scenarios,7): ws.cell(r,1,lab); ws.cell(r,2,form); _link(ws.cell(r,2),FMT_PCT)
    _section(ws,12,"Evidence Log"); heads=["Evidence / KPI","Date","Bull Likelihood Ratio","Bear Likelihood Ratio","Source","Reasoning / Notes"]
    for c,v in enumerate(heads,1): ws.cell(13,c,v)
    _header(ws,13,1,6)
    for r in range(14,24):
        for c in range(1,7): _input(ws.cell(r,c))
        ws.cell(r,3,1.0); ws.cell(r,4,1.0); ws.cell(r,3).number_format="0.00x"; ws.cell(r,4).number_format="0.00x"
    ws["C7"]="=B7*PRODUCT(D14:D23)"; ws["C8"]="=B8"; ws["C9"]="=B9*PRODUCT(C14:C23)"
    ws["D7"]='=IFERROR(C7/SUM($C$7:$C$9),"")'; ws["D8"]='=IFERROR(C8/SUM($C$7:$C$9),"")'; ws["D9"]='=IFERROR(C9/SUM($C$7:$C$9),"")'
    for r in range(7,10): _link(ws.cell(r,3),FMT_PCT); _link(ws.cell(r,4),FMT_PCT)
    ws["E7"]="Bear evidence raises Bear LR > 1"; ws["E8"]="Base is the neutral anchor"; ws["E9"]="Bull evidence raises Bull LR > 1"
    _section(ws,27,"Reference-Class Base Rates"); heads=["Reference Class","Metric / Threshold","Horizon","Historical Base Rate","Your Company / Case","Source","Notes"]
    for c,v in enumerate(heads,1): ws.cell(28,c,v)
    _header(ws,28,1,7); examples=[("Large-cap compounders",">10% revenue CAGR","5 years",None,"","Manual research",""),("Large-cap compounders",">15% revenue CAGR","5 years",None,"","Manual research",""),("High-ROIC firms","Maintain ROIC > WACC + 5ppt","10 years",None,"","Manual research",""),("Growth firms","Expand EBIT margin >300bps","5 years",None,"","Manual research",""),("Capital-intensive growth","Growth capex earns > WACC","5–10 years",None,"","Manual research","")]
    for r,row in enumerate(examples,29):
        for c,v in enumerate(row,1): ws.cell(r,c,v); _input(ws.cell(r,c))
        ws.cell(r,4).number_format=FMT_PCT
    for r in range(34,39):
        for c in range(1,8): _input(ws.cell(r,c))
        ws.cell(r,4).number_format=FMT_PCT
    _section(ws,41,"Probability Discipline"); qs=["What was my prior before the latest news?","Which evidence has the largest likelihood ratio and why?","Am I double-counting correlated evidence?","What evidence would move Bear probability above 40%?","What evidence would move Bull probability above 40%?"]
    for r,q in enumerate(qs,42): ws.cell(r,1,q); _input(ws.cell(r,2)); ws.cell(r,2).alignment=Alignment(wrap_text=True); ws.row_dimensions[r].height=30
    ws.column_dimensions["A"].width=34; ws.column_dimensions["B"].width=18; ws.column_dimensions["C"].width=18; ws.column_dimensions["D"].width=18; ws.column_dimensions["E"].width=30; ws.column_dimensions["F"].width=44; ws.column_dimensions["G"].width=32; return ws

def ensure_market_expectations(wb,ticker):
    ws=_create(wb,"Market Expectations"); _title(ws,f"{ticker} — Market-Implied Expectations, MEROI & PVGO")
    ws["A3"]="Reverse-engineers operating outcomes consistent with the current share price. Implied values are hurdle rates, not forecasts."; ws["A3"].font=Font(italic=True,color=GREY)
    price,shares,net_debt=_company(wb); s=wb["Three-Case Scenarios"]; h=wb["Historical Financials"]; proj=_base_projection(wb); wacc=_num(s["C6"].value,.09); tgr=_num(s["C7"].value,.03)
    implied_fcf=_implied_fcf_growth(wb); g_shift=_bisect(lambda x:_model_value(wb,g_shift=x),price,-.15,.25); m_shift=_bisect(lambda x:_model_value(wb,m_shift=x),price,-.25,.35); cap_shift=_bisect(lambda x:_model_value(wb,cap_shift=x),price,-.25,.15)
    latest_rev=_num(h["G4"].value); base_rev_cagr=((proj[-1]["revenue"]/latest_rev)**(1/10)-1) if proj and latest_rev else None; implied_rev_cagr=None
    if g_shift is not None and latest_rev:
        rev=latest_rev
        for x in proj: rev*=1+max(-.5,min(.6,x["growth"]+g_shift))
        implied_rev_cagr=(rev/latest_rev)**(1/10)-1
    base_margin=proj[-1]["margin"] if proj else None; implied_margin=(base_margin+m_shift) if base_margin is not None and m_shift is not None else None
    base_cap=proj[-1]["capex_pct"] if proj else None; implied_cap=(base_cap+cap_shift) if base_cap is not None and cap_shift is not None else None; duration=_implied_duration(wb)
    ocf=_num(h["G14"].value); cap=_num(h["G15"].value); fcf0=(ocf-cap) if ocf is not None and cap is not None else None
    steady_ev=(fcf0*(1+tgr)/(wacc-tgr)) if fcf0 is not None and wacc>tgr else None; steady_share=((steady_ev-net_debt)/shares) if steady_ev is not None and shares else None
    pvgo=(price-steady_share) if steady_share is not None else None; pvgo_pct=(pvgo/price) if pvgo is not None and price else None
    op=_num(h["G9"].value); da=_num(h["G18"].value,0); latest_cap=_num(h["G15"].value,0); tax=_num(s["C8"].value,.21); growth_cap=max(0,latest_cap-da); nopat=op*(1-tax) if op is not None else None
    reinvest=(growth_cap/nopat) if nopat and nopat>0 else None; meroi=(implied_fcf/reinvest) if implied_fcf is not None and reinvest and reinvest>0 else None
    _section(ws,5,"Market-Implied Operating Hurdles"); heads=["Metric","Your Base","Market Implied","Gap","Interpretation / Method"]
    for c,v in enumerate(heads,1): ws.cell(6,c,v)
    _header(ws,6,1,5); metrics=[("10Y Revenue CAGR",base_rev_cagr,implied_rev_cagr,FMT_PCT,"Parallel shift to the Base revenue-growth path that reconciles market price"),("10Y FCF CAGR",((proj[-1]["fcf"]/fcf0)**(1/10)-1) if proj and fcf0 and fcf0>0 and proj[-1]["fcf"]>0 else None,implied_fcf,FMT_PCT,"Constant FCF growth required by current enterprise value"),("2035 EBIT Margin",base_margin,implied_margin,FMT_PCT,"Parallel margin shift required to reconcile market price"),("2035 Capex / Revenue",base_cap,implied_cap,FMT_PCT,"Parallel capex-intensity shift required to reconcile market price"),("Competitive Advantage Duration",15,duration,'0 "yrs"',"Simplified elevated-growth duration; compare with your CAP judgment")]
    for r,(lab,base,imp,fmt,note) in enumerate(metrics,7):
        ws.cell(r,1,lab); ws.cell(r,2,base); ws.cell(r,3,imp); ws.cell(r,4,f'=IF(OR(B{r}="",C{r}=""),"",C{r}-B{r})'); ws.cell(r,5,note); ws.cell(r,2).number_format=fmt; ws.cell(r,3).number_format=fmt; ws.cell(r,4).number_format=fmt
        if r==11: _input(ws.cell(r,2),fmt)
        else: _link(ws.cell(r,2),fmt)
        _link(ws.cell(r,3),fmt)
    _section(ws,15,"MEROI / Incremental-Capital Expectations"); data=[("Latest NOPAT Proxy",nopat,FMT_BN,"EBIT × (1 − normalized tax)"),("Growth Capex Proxy",growth_cap,FMT_BN,"Capex − D&A; rough growth-investment proxy"),("Reinvestment Rate Proxy",reinvest,FMT_PCT,"Growth capex ÷ NOPAT"),("Market-Implied FCF Growth",implied_fcf,FMT_PCT,"Reverse-DCF hurdle"),("Market-Expected Return on Incremental Capital (MEROI Proxy)",meroi,FMT_PCT,"Implied growth ÷ reinvestment rate; use cautiously when reinvestment is small or volatile"),("WACC",wacc,FMT_PCT,"Economic hurdle rate"),("MEROI − WACC",(meroi-wacc) if meroi is not None else None,FMT_PCT,"Positive spread means the market prices value-creating reinvestment")]
    _header(ws,16,1,4); ws["A16"]="Metric"; ws["B16"]="Value"; ws["C16"]="Unit"; ws["D16"]="Interpretation"
    for r,(lab,val,fmt,note) in enumerate(data,17): ws.cell(r,1,lab); ws.cell(r,2,val); _link(ws.cell(r,2),fmt); ws.cell(r,3,fmt); ws.cell(r,4,note)
    _section(ws,26,"Steady-State Value & PVGO"); pvdata=[("Current Price",price,FMT_PRICE),("Steady-State Value / Share",steady_share,FMT_PRICE),("PVGO / Share",pvgo,FMT_PRICE),("PVGO as % of Price",pvgo_pct,FMT_PCT)]
    for r,(lab,val,fmt) in enumerate(pvdata,27): ws.cell(r,1,lab); ws.cell(r,2,val); _link(ws.cell(r,2),fmt)
    ws["D27"]="PVGO = present value of future growth opportunities above a no-explicit-growth steady-state benchmark. Large PVGO means more of today's price depends on future execution."; ws["D27"].alignment=Alignment(wrap_text=True)
    _section(ws,34,"Expectation Frontier — Manual Decision Grid"); ws["A35"]="Use this section to record combinations of operating outcomes that you believe would justify the current price."
    heads=["Case","Revenue CAGR","Mature EBIT Margin","Mature Capex / Revenue","Moat Duration","Assessment / Probability"]
    for c,v in enumerate(heads,1): ws.cell(37,c,v)
    _header(ws,37,1,6)
    for r in range(38,46):
        for c in range(1,7): _input(ws.cell(r,c))
        for c in (2,3,4): ws.cell(r,c).number_format=FMT_PCT
    ws.column_dimensions["A"].width=38; ws.column_dimensions["B"].width=18; ws.column_dimensions["C"].width=18; ws.column_dimensions["D"].width=22; ws.column_dimensions["E"].width=20; ws.column_dimensions["F"].width=44; return ws

def ensure_institutional_layers(wb,ticker):
    ensure_expectations_consensus(wb,ticker)
    ensure_moat_cap(wb,ticker)
    ensure_base_rates(wb,ticker)
    ensure_market_expectations(wb,ticker)
