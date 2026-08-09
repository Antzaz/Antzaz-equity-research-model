"""Verified GE Vernova fallback repairs and current-guidance anchors.

Runs only for GEV. It restores directly disclosed FY2023-FY2025 segment revenue and
Segment EBITDA when generic SEC HTML-table extraction fails, adds the consolidated
Equipment/Services revenue mix, supplies D&A history when Company Facts tags are
incomplete, and anchors near-term scenario margins to current company guidance without
mislabeling adjusted EBITDA as reported EBIT.
"""

from openpyxl.styles import Font, Alignment
from openpyxl.comments import Comment

SEC_SEG="https://www.sec.gov/Archives/edgar/data/1996810/000199681026000015/R132.htm"
SEC_10K="https://www.sec.gov/Archives/edgar/data/1996810/000199681026000015/gev-20251231.htm"
Q2_2026="https://www.gevernova.com/news/articles/ge-vernova-releases-second-quarter-2026-financial-results"
Q4_2025="https://www.sec.gov/Archives/edgar/data/1996810/000199681026000012/gevpressrelease4q25.htm"
BLUE="0000FF"; GREEN="008000"; BLACK="000000"; GREY="666666"
FMT_BN='#,##0.0;[Red](#,##0.0);-'; FMT_PCT='0.0%;[Red](0.0%);-'

SEGMENTS=[
    ("Power",[17.436,18.127,19.767],[1.722,2.268,2.902]),
    ("Wind",[9.826,9.701,9.110],[-1.033,-0.588,-0.598]),
    ("Electrification",[6.378,7.550,9.642],[0.234,0.679,1.433]),
]
REVENUE_GROUPS=[("Equipment",[18.258,18.952,20.934]),("Services",[14.981,15.983,17.134])]
D_AND_A={2022:1.797,2023:.964,2024:1.172,2025:.853}

def _num(v,default=None):
    try:
        if isinstance(v,bool): return default
        return float(v)
    except Exception: return default

def _input(cell,value,source=SEC_SEG,fmt=None):
    cell.value=value; cell.font=Font(color=BLUE); cell.comment=Comment(f"Issuer-reported / filing-derived input. Source: {source}","Model repair")
    if fmt: cell.number_format=fmt

def _formula(cell,formula,fmt=None):
    cell.value=formula; cell.font=Font(color=BLACK)
    if fmt: cell.number_format=fmt

def _segment_has_data(ws):
    for r in range(7,min(ws.max_row,16)+1):
        if str(ws.cell(r,1).value or "") in {"Power","Wind","Electrification"} and isinstance(ws.cell(r,4).value,(int,float)): return True
    return False

def _repair_history_d_and_a(wb):
    if "Historical Financials" not in wb.sheetnames: return
    h=wb["Historical Financials"]
    for c in range(2,8):
        y=h.cell(3,c).value
        if isinstance(y,(int,float)) and int(y) in D_AND_A and not isinstance(h.cell(18,c).value,(int,float)):
            _input(h.cell(18,c),D_AND_A[int(y)],SEC_10K,FMT_BN)
    if "Financial Statements" in wb.sheetnames:
        ws=wb["Financial Statements"]; da_row=None; header_row=None
        for r in range(1,ws.max_row+1):
            if str(ws.cell(r,1).value or "").strip()=="Depreciation & Amortization": da_row=r
            if str(ws.cell(r,1).value or "").strip()=="Metric" and r>40: header_row=r
        if da_row and header_row:
            for c in range(2,8):
                y=ws.cell(header_row,c).value
                if isinstance(y,(int,float)) and int(y) in D_AND_A and not isinstance(ws.cell(da_row,c).value,(int,float)):
                    _input(ws.cell(da_row,c),D_AND_A[int(y)],SEC_10K,FMT_BN)

def repair_gev_model(wb,ticker):
    if str(ticker).upper()!="GEV": return False
    _repair_history_d_and_a(wb)
    if "Segment Analysis" not in wb.sheetnames: return False
    ws=wb["Segment Analysis"]
    if not _segment_has_data(ws):
        ws["A3"]="Issuer-disclosed segment schema. Status: AUTO/FALLBACK — verified GE Vernova FY2025 10-K. GE Vernova principally evaluates segments using Segment EBITDA, so profitability uses Segment EBITDA rather than GAAP operating income."
        ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True)
        headers=["Segment","2023 Revenue","2024 Revenue","2025 Revenue","2025 Growth","2023–2025 CAGR","2023 Segment EBITDA","2024 Segment EBITDA","2025 Segment EBITDA","2023 EBITDA Margin","2024 EBITDA Margin","2025 EBITDA Margin","Margin Δ","2025 Revenue Mix"]
        for c,v in enumerate(headers,1): ws.cell(6,c,v)
        for r,(name,rev,ebitda) in enumerate(SEGMENTS,7):
            ws.cell(r,1,name)
            for c,v in enumerate(rev,2): _input(ws.cell(r,c),v,SEC_SEG,FMT_BN)
            for c,v in enumerate(ebitda,7): _input(ws.cell(r,c),v,SEC_SEG,FMT_BN)
            _formula(ws.cell(r,5),f'=IFERROR(D{r}/C{r}-1,"")',FMT_PCT); _formula(ws.cell(r,6),f'=IFERROR((D{r}/B{r})^(1/2)-1,"")',FMT_PCT); _formula(ws.cell(r,10),f'=IFERROR(G{r}/B{r},"")',FMT_PCT); _formula(ws.cell(r,11),f'=IFERROR(H{r}/C{r},"")',FMT_PCT); _formula(ws.cell(r,12),f'=IFERROR(I{r}/D{r},"")',FMT_PCT); _formula(ws.cell(r,13),f'=IFERROR(L{r}-K{r},"")',FMT_PCT); _formula(ws.cell(r,14),f'=IFERROR(D{r}/SUM($D$7:$D$9),"")',FMT_PCT)
        for r in range(10,17):
            for c in range(1,15): ws.cell(r,c).value=None
        start=None
        for r in range(1,ws.max_row+1):
            if str(ws.cell(r,1).value or "").strip()=="Revenue by Business Line": start=r; break
        if start:
            for i,(name,vals) in enumerate(REVENUE_GROUPS,start+2):
                ws.cell(i,1,name)
                for c,v in enumerate(vals,2): _input(ws.cell(i,c),v,SEC_SEG,FMT_BN)
                _formula(ws.cell(i,5),f'=IFERROR(D{i}/C{i}-1,"")',FMT_PCT); _formula(ws.cell(i,6),f'=IFERROR((D{i}/B{i})^(1/2)-1,"")',FMT_PCT); _formula(ws.cell(i,7),f'=IFERROR(D{i}/SUM($D${start+2}:$D${start+3}),"")',FMT_PCT); ws.cell(i,8,SEC_SEG); ws.cell(i,8).font=Font(color=GREEN)
            for r in range(start+4,min(ws.max_row,start+12)+1):
                if str(ws.cell(r,1).value or "").strip() not in {"Source & Data Quality","SEC 10-K Source","Extraction Status","Important"}:
                    for c in range(1,9): ws.cell(r,c).value=None
        for r in range(1,ws.max_row+1):
            label=str(ws.cell(r,1).value or "").strip()
            if label=="SEC 10-K Source": ws.cell(r,2,SEC_SEG); ws.cell(r,2).font=Font(color=GREEN)
            elif label=="Extraction Status": ws.cell(r,2,"AUTO/FALLBACK — verified FY2025 10-K: 3 operating segments; Equipment/Services revenue mix")
            elif label=="Important": ws.cell(r,2,"Segment revenues include intersegment activity; Segment EBITDA is GE Vernova's segment performance measure. Equipment/Services are consolidated revenue categories."); ws.cell(r,2).alignment=Alignment(wrap_text=True)
    return True

def apply_gev_guidance_assumptions(wb,ticker):
    """Use current management guidance as a near-term Base-case anchor.

    2026 adjusted EBITDA margin guidance is 12%-14%; 2028 outlook is 20%. The DCF uses
    EBIT, so we subtract the latest D&A/revenue intensity as an explicit analyst proxy.
    This conversion is documented and is not represented as company EBIT guidance.
    """
    if str(ticker).upper()!="GEV" or not {"Historical Financials","Three-Case Scenarios"}.issubset(wb.sheetnames): return False
    h=wb["Historical Financials"]; s=wb["Three-Case Scenarios"]; rev=_num(h["G4"].value); da=_num(h["G18"].value)
    if not rev: return False
    da_pct=(da/rev) if da is not None and da>=0 else .022
    # Revenue: official Q2 2026 midpoint; current consensus module may already have reset N/O.
    if not isinstance(s["N12"].value,(int,float)): s["N12"]=(46.0/rev)-1
    rev26=rev*(1+_num(s["N12"].value,46.0/rev-1)); rev27=rev26*(1+_num(s["O12"].value,.12))
    s["P12"]=(56.0/rev27)-1 if rev27>0 else .08  # company 2028 revenue outlook
    long_g=[.08,.075,.07,.065,.06,.055,.05]
    for c,v in zip(("Q","R","S","T","U","V","W"),long_g): s[f"{c}12"]=v
    base_m=[.13-da_pct,.165-da_pct,.20-da_pct,.18,.185,.19,.19,.19,.19,.19]
    base_cols=("N","O","P","Q","R","S","T","U","V","W"); bear_cols=("B","C","D","E","F","G","H","I","J","K"); bull_cols=("Z","AA","AB","AC","AD","AE","AF","AG","AH","AI")
    for bc,cc,xc,m in zip(bear_cols,base_cols,bull_cols,base_m):
        s[f"{cc}14"]=m; s[f"{bc}14"]=max(-.10,m-.03); s[f"{xc}14"]=min(.60,m+.02)
        s[f"{bc}14"].number_format=s[f"{cc}14"].number_format=s[f"{xc}14"].number_format=FMT_PCT
    # Keep Bear/Bull revenue spreads around the final Base growth path.
    for bc,cc,xc in zip(bear_cols,base_cols,bull_cols):
        g=_num(s[f"{cc}12"].value,0); s[f"{bc}12"]=max(-.20,g-.03); s[f"{xc}12"]=min(.50,g+.03); s[f"{bc}12"].number_format=s[f"{cc}12"].number_format=s[f"{xc}12"].number_format=FMT_PCT
    s["AK9"]="GEV margin guidance anchor"; s["AL9"]="2026 Base EBIT proxy uses midpoint of 12%-14% adjusted EBITDA guidance less latest D&A/revenue; 2028 uses 20% adjusted EBITDA outlook less D&A/revenue. Analyst conversion, not reported EBIT guidance."; s["AL9"].font=Font(italic=True,color=GREY); s["AL9"].alignment=Alignment(wrap_text=True)
    return True

def repair_gev_expectations(wb,ticker):
    if str(ticker).upper()!="GEV" or "Expectations & Consensus" not in wb.sheetnames: return
    ws=wb["Expectations & Consensus"]
    for r in range(7,min(ws.max_row,30)+1):
        metric=str(ws.cell(r,1).value or ""); year=ws.cell(r,2).value
        if metric=="FCF" and year==2026:
            _input(ws.cell(r,3),12.0,Q2_2026,FMT_BN); ws.cell(r,5,f'=IF(OR(C{r}="",D{r}=""),"",D{r}-C{r})'); ws.cell(r,6,f'=IF(OR(C{r}="",D{r}=""),"",IFERROR(D{r}/C{r}-1,""))'); ws.cell(r,5).number_format=FMT_BN; ws.cell(r,6).number_format=FMT_PCT; ws.cell(r,11,"GE Vernova Q2 2026 management guidance midpoint ($11.5–12.5bn FCF)")
