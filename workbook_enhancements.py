from __future__ import annotations

"""Public-evidence recovery and presentation polish for generated equity-research workbooks.

This is deliberately conservative:
- it prefers issuer / regulator evidence already present in the workbook;
- it uses public market-data only as a fallback;
- it never invents a numeric employee-satisfaction, market-share or management score;
- missing but non-essential metrics are marked N/A instead of being treated as research failure.

The module runs after the deterministic workbook is built, so it can improve existing and
future tickers without changing valuation inputs or model formulas.
"""

from datetime import datetime
from io import BytesIO
from pathlib import Path
import re
from urllib.parse import urljoin

import requests
import yfinance as yf
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from pypdf import PdfReader
except Exception:
    PdfReader = None

try:
    from source_registry import issuer_sources
except Exception:
    issuer_sources = None


NAVY = "17365D"
BLUE = "2F75B5"
WHITE = "FFFFFF"
LIGHT_BLUE = "D9EAF7"
GREEN = "E2F0D9"
GOLD = "FFF2CC"
GREY = "666666"
LINK_GREEN = "008000"

STATUS_FILL = {
    "PASS": GREEN,
    "PUBLIC-EVIDENCE": GREEN,
    "SCOPE-LIMITED": LIGHT_BLUE,
    "PARTIAL": GOLD,
    "REVIEW": GOLD,
    "N/A": LIGHT_BLUE,
}

CORE_ORDER = [
    "Dashboard",
    "Investment Summary",
    "Decision View",
    "Historical Financials",
    "Financial Statements",
    "Three-Case Scenarios",
    "Advanced Analytics",
    "Peer Comps",
    "Segment Analysis",
    "Expectations & Consensus",
    "Market Expectations",
    "AI Impact Analysis",
    "Leadership & Culture",
    "Institutional Comparison",
    "Deals & Transactions",
    "Ownership Analysis",
    "Research Workbench",
    "Data Quality",
]

BACKEND_SHEETS_TO_HIDE = {"Company Data", "Filings"}

TAB_COLORS = {
    "Dashboard": "17365D",
    "Investment Summary": "17365D",
    "Decision View": "17365D",
    "Historical Financials": "548235",
    "Financial Statements": "548235",
    "Three-Case Scenarios": "548235",
    "Segment Analysis": "548235",
    "Advanced Analytics": "C65911",
    "Peer Comps": "C65911",
    "Market Expectations": "C65911",
    "Expectations & Consensus": "8064A2",
    "AI Impact Analysis": "8064A2",
    "Leadership & Culture": "8064A2",
    "Institutional Comparison": "8064A2",
    "Deals & Transactions": "8064A2",
    "Research Workbench": "7F7F7F",
    "Data Quality": "7F7F7F",
}

WORKFORCE_PATTERNS = [
    r"\bemployee engagement\b",
    r"\bengagement survey\b",
    r"\bemployee survey\b",
    r"\bemployee satisfaction\b",
    r"\bemployee experience\b",
    r"\bhuman capital\b",
    r"\bworkforce\b",
    r"\bemployee retention\b",
    r"\bretention rate\b",
    r"\bemployee turnover\b",
    r"\bvoluntary turnover\b",
    r"\btraining and development\b",
    r"\blearning and development\b",
    r"\bpeople and culture\b",
]

GOVERNANCE_PATTERNS = [
    r"\bcorporate governance\b",
    r"\bboard of directors\b",
    r"\bindependent director",
    r"\baudit committee\b",
    r"\bcompensation committee\b",
]

LEADERSHIP_PATTERNS = [
    r"\bchief executive officer\b",
    r"\bexecutive officers\b",
    r"\bmanagement team\b",
    r"\bleadership team\b",
]


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _safe_info(ticker: str) -> dict:
    try:
        obj = yf.Ticker(ticker)
        info = obj.get_info() if hasattr(obj, "get_info") else obj.info
        return info or {}
    except Exception:
        return {}


def _find_row(ws, label: str, col: int = 1):
    needle = str(label).strip().lower()
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, col).value or "").strip().lower() == needle:
            return r
    return None


def _annual_filing_url(wb):
    if "Filings" not in wb.sheetnames:
        return None, None
    ws = wb["Filings"]
    forms = {"10-K", "10-K/A", "20-F", "20-F/A", "40-F", "40-F/A"}
    for r in range(1, min(ws.max_row, 60) + 1):
        form = str(ws.cell(r, 1).value or "").strip().upper()
        if form not in forms:
            continue
        for c in range(2, min(ws.max_column, 8) + 1):
            value = str(ws.cell(r, c).value or "").strip()
            if value.startswith("http"):
                return form, value
    return None, None


def _clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", text or " ").strip()


def _html_text(content: bytes) -> str:
    try:
        soup = BeautifulSoup(content, "lxml")
    except Exception:
        soup = BeautifulSoup(content, "html.parser")
    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()
    return _clean_text(soup.get_text(" ", strip=True))


def _pdf_text(content: bytes, max_pages: int = 80) -> str:
    if PdfReader is None:
        return ""
    try:
        reader = PdfReader(BytesIO(content))
        chunks = []
        for page in reader.pages[:max_pages]:
            try:
                chunks.append(page.extract_text() or "")
            except Exception:
                continue
        return _clean_text(" ".join(chunks))
    except Exception:
        return ""


def _fetch_public_text(url: str, timeout: int = 10) -> tuple[str, str]:
    if not url or not str(url).startswith("http"):
        return "", ""
    headers = {
        "User-Agent": "Antzaz Equity Research/1.0 public-source verification",
        "Accept-Language": "en-US,en;q=0.8",
    }
    try:
        response = requests.get(url, timeout=timeout, headers=headers, allow_redirects=True)
        response.raise_for_status()
    except Exception:
        return "", ""
    content_type = str(response.headers.get("content-type") or "").lower()
    final_url = response.url or url
    if "pdf" in content_type or final_url.lower().split("?")[0].endswith(".pdf"):
        return _pdf_text(response.content), final_url
    return _html_text(response.content), final_url


def _discover_link(page_url: str, keywords: tuple[str, ...]) -> str | None:
    if not page_url:
        return None
    headers = {"User-Agent": "Antzaz Equity Research/1.0 public-source verification"}
    try:
        response = requests.get(page_url, timeout=8, headers=headers, allow_redirects=True)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, "lxml")
    except Exception:
        return None
    ranked = []
    for a in soup.find_all("a", href=True):
        href = urljoin(response.url, a.get("href"))
        label = _clean_text(f"{a.get_text(' ', strip=True)} {href}").lower()
        score = sum(1 for k in keywords if k in label)
        if score:
            ranked.append((score, href))
    ranked.sort(reverse=True)
    return ranked[0][1] if ranked else None


def _issuer_candidate_pages(ticker: str, info: dict) -> dict[str, str]:
    out = {}
    if issuer_sources is not None:
        try:
            out.update(issuer_sources(ticker, info.get("website")) or {})
        except Exception:
            pass
    website = str(info.get("website") or "").strip().rstrip("/")
    if website:
        out.setdefault("company_website", website)
        candidates = {
            "investor_candidate": "/investors",
            "governance_candidate": "/governance",
            "sustainability_candidate": "/sustainability",
            "careers_candidate": "/careers",
            "leadership_candidate": "/leadership",
        }
        for key, suffix in candidates.items():
            out.setdefault(key, website + suffix)
    return out


def _first_evidence(text: str, patterns: list[str], max_chars: int = 320):
    if not text:
        return None
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I)
        if not match:
            continue
        start = max(0, match.start() - 120)
        end = min(len(text), match.end() + 180)
        snippet = _clean_text(text[start:end])
        if snippet:
            return snippet[:max_chars]
    return None


def _collect_public_evidence(wb, ticker: str, info: dict) -> dict:
    evidence = {
        "annual_form": None,
        "annual_filing": None,
        "workforce": None,
        "workforce_source": None,
        "governance": None,
        "governance_source": None,
        "leadership": None,
        "leadership_source": None,
        "headcount": info.get("fullTimeEmployees"),
        "officer_count": len([x for x in (info.get("companyOfficers") or []) if isinstance(x, dict) and x.get("name")]),
        "issuer_pages": _issuer_candidate_pages(ticker, info),
    }

    form, filing_url = _annual_filing_url(wb)
    evidence["annual_form"] = form
    evidence["annual_filing"] = filing_url
    if filing_url:
        filing_text, final_url = _fetch_public_text(filing_url)
        if final_url:
            evidence["annual_filing"] = final_url
        evidence["workforce"] = _first_evidence(filing_text, WORKFORCE_PATTERNS)
        evidence["governance"] = _first_evidence(filing_text, GOVERNANCE_PATTERNS)
        evidence["leadership"] = _first_evidence(filing_text, LEADERSHIP_PATTERNS)
        if evidence["workforce"]:
            evidence["workforce_source"] = evidence["annual_filing"]
        if evidence["governance"]:
            evidence["governance_source"] = evidence["annual_filing"]
        if evidence["leadership"]:
            evidence["leadership_source"] = evidence["annual_filing"]

    pages = evidence["issuer_pages"]
    probes = [
        ("workforce", ("workplace", "sustainability", "careers_candidate", "annual_reports", "investor"), WORKFORCE_PATTERNS),
        ("governance", ("governance", "workplace_governance", "governance_candidate", "investor"), GOVERNANCE_PATTERNS),
        ("leadership", ("leadership", "leadership_candidate", "governance", "investor"), LEADERSHIP_PATTERNS),
    ]
    for kind, keys, patterns in probes:
        if evidence.get(kind):
            continue
        seen = set()
        for key in keys:
            url = pages.get(key)
            if not url or url in seen:
                continue
            seen.add(url)
            text, final_url = _fetch_public_text(url, timeout=7)
            snippet = _first_evidence(text, patterns)
            if snippet:
                evidence[kind] = snippet
                evidence[f"{kind}_source"] = final_url or url
                break
            if key in {"annual_reports", "investor"}:
                discovered = _discover_link(url, ("annual report", "10-k", "20-f", "sustainability report", "human capital"))
                if discovered and discovered not in seen:
                    text, final_url = _fetch_public_text(discovered, timeout=8)
                    snippet = _first_evidence(text, patterns)
                    if snippet:
                        evidence[kind] = snippet
                        evidence[f"{kind}_source"] = final_url or discovered
                        break
    return evidence


def _patch_leadership_sheet(wb, evidence: dict):
    if "Leadership & Culture" not in wb.sheetnames:
        return
    ws = wb["Leadership & Culture"]
    section_row = _find_row(ws, "Worker Happiness / Employee Experience Evidence")
    if section_row:
        ws.cell(section_row, 1).value = "Workforce & Employee Experience Evidence"

    metric_row = _find_row(ws, "Worker happiness / satisfaction signal")
    if not metric_row:
        metric_row = _find_row(ws, "Workforce / employee experience evidence")
    if not metric_row:
        return

    existing_score = ws.cell(metric_row, 2).value
    source = evidence.get("workforce_source")
    snippet = evidence.get("workforce")
    headcount = evidence.get("headcount")

    if isinstance(existing_score, (int, float)):
        status = str(ws.cell(metric_row, 5).value or "SCOPE-LIMITED")
        scope = str(ws.cell(metric_row, 3).value or "Quantitative issuer signal; review exact scope.")
        value = existing_score
    elif snippet:
        status = "PUBLIC-EVIDENCE"
        scope = "Issuer/regulatory disclosure contains workforce or human-capital evidence; no comparable company-wide happiness score is inferred."
        value = "Qualitative"
    elif headcount:
        status = "PARTIAL"
        scope = f"Public market-data reports approximately {int(headcount):,} employees; no comparable company-wide engagement score was verified."
        value = "Headcount only"
    else:
        status = "REVIEW"
        scope = "No comparable employee-experience metric or qualitative workforce disclosure was automatically verified."
        value = "Not verified"

    ws.cell(metric_row, 1).value = "Workforce / employee experience evidence"
    ws.cell(metric_row, 2).value = value
    ws.cell(metric_row, 3).value = scope
    ws.cell(metric_row, 5).value = status
    if source:
        ws.cell(metric_row, 6).value = source
        ws.cell(metric_row, 6).hyperlink = source
        ws.cell(metric_row, 6).font = Font(color=LINK_GREEN, underline="single")

    ev_row = _find_row(ws, "Evidence")
    if ev_row:
        detail = snippet or scope
        if headcount and snippet:
            detail = f"{detail} Public employee count: {int(headcount):,}."
        ws.cell(ev_row, 2).value = detail
        ws.cell(ev_row, 2).alignment = Alignment(wrap_text=True, vertical="top")

    exec_header = _find_row(ws, "Executive Team — Public Snapshot")
    if exec_header:
        first_data = exec_header + 2
        if str(ws.cell(first_data, 1).value or "").startswith("Public officer data unavailable") and evidence.get("leadership"):
            ws.cell(first_data, 1).value = "Leadership disclosure located"
            ws.cell(first_data, 2).value = evidence.get("leadership")
            ws.merge_cells(start_row=first_data, start_column=2, end_row=first_data, end_column=4)
            ws.cell(first_data, 2).alignment = Alignment(wrap_text=True, vertical="top")
            src = evidence.get("leadership_source")
            if src:
                ws.cell(first_data, 5).value = src
                ws.cell(first_data, 5).hyperlink = src
                ws.cell(first_data, 5).font = Font(color=LINK_GREEN, underline="single")


def _patch_summary_sheets(wb, evidence: dict):
    status = "PUBLIC-EVIDENCE" if evidence.get("workforce") else ("PARTIAL" if evidence.get("headcount") else "REVIEW")
    scope = (
        "Issuer/regulatory workforce evidence located; no numeric happiness score inferred."
        if evidence.get("workforce")
        else (
            f"Public employee count available ({int(evidence['headcount']):,}); qualitative engagement evidence still limited."
            if evidence.get("headcount")
            else "No comparable workforce signal automatically verified."
        )
    )
    source = evidence.get("workforce_source") or evidence.get("annual_filing")

    if "Dashboard" in wb.sheetnames:
        ws = wb["Dashboard"]
        r = _find_row(ws, "Worker happiness / satisfaction signal") or _find_row(ws, "Workforce / employee evidence")
        if r:
            ws.cell(r, 1).value = "Workforce / employee evidence"
            if not isinstance(ws.cell(r, 2).value, (int, float)):
                ws.cell(r, 2).value = "Public evidence" if evidence.get("workforce") else status
            ws.cell(r, 3).value = scope
        r = _find_row(ws, "Workforce evidence status")
        if r:
            ws.cell(r, 2).value = status
            ws.cell(r, 3).value = source or scope

    if "Investment Summary" in wb.sheetnames:
        ws = wb["Investment Summary"]
        r = _find_row(ws, "Worker happiness / employee signal") or _find_row(ws, "Workforce / employee evidence")
        if r:
            ws.cell(r, 1).value = "Workforce / employee evidence"
            if not isinstance(ws.cell(r, 2).value, (int, float)):
                ws.cell(r, 2).value = "Public evidence" if evidence.get("workforce") else status
            ws.cell(r, 3).value = scope


def _write_public_evidence_quality(wb, evidence: dict):
    ws = wb["Data Quality"] if "Data Quality" in wb.sheetnames else wb.create_sheet("Data Quality")
    marker = "Public Evidence Recovery — automated, source-scoped"
    start = _find_row(ws, marker)
    if start:
        for r in range(start, min(ws.max_row, start + 12) + 1):
            for c in range(1, 6):
                ws.cell(r, c).value = None
    else:
        start = ws.max_row + 2

    for c in range(1, 6):
        ws.cell(start, c).fill = _fill(NAVY)
        ws.cell(start, c).font = Font(bold=True, color=WHITE, size=11)
    ws.cell(start, 1).value = marker

    headers = ["Evidence Area", "Status", "Observed Evidence", "Public Source", "Research Treatment"]
    for c, value in enumerate(headers, 1):
        cell = ws.cell(start + 1, c, value)
        cell.fill = _fill(BLUE)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    form = evidence.get("annual_form")
    filing = evidence.get("annual_filing")
    headcount = evidence.get("headcount")
    officer_count = evidence.get("officer_count") or 0
    rows = [
        ("Latest annual filing", "PASS" if filing else "REVIEW", f"{form or 'Annual filing'} located" if filing else "No annual filing URL was recovered from the workbook", filing, "Primary/regulatory evidence anchor for qualitative research."),
        ("Workforce / human capital", "PUBLIC-EVIDENCE" if evidence.get("workforce") else ("PARTIAL" if headcount else "REVIEW"), evidence.get("workforce") or (f"Public employee count: {int(headcount):,}" if headcount else "No workforce evidence automatically verified"), evidence.get("workforce_source") or filing, "Do not infer a numeric happiness score unless the source explicitly reports one."),
        ("Leadership disclosure", "PUBLIC-EVIDENCE" if evidence.get("leadership") or officer_count else "REVIEW", evidence.get("leadership") or (f"{officer_count} public officer record(s) available" if officer_count else "No leadership disclosure automatically verified"), evidence.get("leadership_source") or filing, "Use named public officers / issuer disclosure; keep management-quality score a model proxy."),
        ("Governance disclosure", "PUBLIC-EVIDENCE" if evidence.get("governance") else ("PARTIAL" if filing else "REVIEW"), evidence.get("governance") or ("Annual filing available for manual governance review" if filing else "No governance disclosure automatically verified"), evidence.get("governance_source") or filing, "Governance availability supports auditability; it is not itself a quality score."),
        ("Market-share comparability", "N/A", "No generic market-share estimate is created when no like-for-like public specialist source exists.", None, "Treat as not applicable to scoring rather than an evidence failure; use mapped specialist sources only."),
    ]

    for r, row in enumerate(rows, start + 2):
        for c, value in enumerate(row, 1):
            ws.cell(r, c).value = value
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
        status = str(ws.cell(r, 2).value or "")
        ws.cell(r, 2).fill = _fill(STATUS_FILL.get(status, GOLD))
        ws.cell(r, 2).font = Font(bold=True)
        src = ws.cell(r, 4).value
        if isinstance(src, str) and src.startswith("http"):
            ws.cell(r, 4).hyperlink = src
            ws.cell(r, 4).font = Font(color=LINK_GREEN, underline="single")

    r = _find_row(ws, "Market-share comparability")
    if r:
        ws.cell(r, 2).value = "N/A"
        ws.cell(r, 2).fill = _fill(LIGHT_BLUE)
        ws.cell(r, 3).value = "No like-for-like public market-share source mapped; metric is excluded rather than estimated."
        ws.cell(r, 4).value = "Not every company needs a market-share input. Only comparable specialist-source data should be used."

    r = _find_row(ws, "Employee sentiment scope")
    if r:
        if evidence.get("workforce"):
            ws.cell(r, 2).value = "PASS"
            ws.cell(r, 2).fill = _fill(GREEN)
            ws.cell(r, 3).value = "Public workforce/human-capital disclosure located and kept scope-limited; no unsupported numeric happiness score inferred."
        elif headcount:
            ws.cell(r, 2).value = "PARTIAL"
            ws.cell(r, 2).fill = _fill(GOLD)
            ws.cell(r, 3).value = f"Public employee count available ({int(headcount):,}); engagement/satisfaction evidence remains qualitative or unavailable."

    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 38)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 18)
    ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width or 0, 72)
    ws.column_dimensions["D"].width = max(ws.column_dimensions["D"].width or 0, 54)
    ws.column_dimensions["E"].width = max(ws.column_dimensions["E"].width or 0, 54)
    ws.freeze_panes = ws.freeze_panes or "A6"


def _style_title_band(ws):
    if not ws["A1"].value:
        return
    end = min(max(ws.max_column, 6), 14)
    for c in range(1, end + 1):
        cell = ws.cell(1, c)
        cell.fill = _fill(NAVY)
        cell.font = Font(name="Aptos Display", size=16 if c == 1 else 11, bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="center")
        cell2 = ws.cell(2, c)
        if cell2.value in (None, ""):
            cell2.fill = _fill(NAVY)
    ws.row_dimensions[1].height = max(ws.row_dimensions[1].height or 15, 26)
    if ws["A3"].value and isinstance(ws["A3"].value, str):
        ws["A3"].font = Font(name="Aptos", size=9, italic=True, color=GREY)
        ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")


def _polish_workbook(wb):
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        try:
            ws.sheet_view.zoomScale = 90
        except Exception:
            pass
        _style_title_band(ws)
        color = TAB_COLORS.get(ws.title)
        if color:
            ws.sheet_properties.tabColor = color

    visible_names = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
    for name in BACKEND_SHEETS_TO_HIDE:
        if name in wb.sheetnames and len(visible_names) > 1:
            wb[name].sheet_state = "hidden"

    order = {name: i for i, name in enumerate(CORE_ORDER)}
    indexed = list(enumerate(wb._sheets))
    wb._sheets = [sheet for _, sheet in sorted(indexed, key=lambda item: (order.get(item[1].title, len(CORE_ORDER) + item[0]), item[0]))]
    if "Dashboard" in wb.sheetnames:
        wb.active = wb.sheetnames.index("Dashboard")


def apply_workbook_enhancements(workbook_path: str | Path, ticker: str) -> dict:
    path = Path(workbook_path)
    wb = load_workbook(path)
    info = _safe_info(ticker)
    evidence = _collect_public_evidence(wb, ticker, info)
    _patch_leadership_sheet(wb, evidence)
    _patch_summary_sheets(wb, evidence)
    _write_public_evidence_quality(wb, evidence)
    _polish_workbook(wb)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(path)
    return {
        "workbook": str(path),
        "ticker": ticker,
        "workforce_status": "PUBLIC-EVIDENCE" if evidence.get("workforce") else ("PARTIAL" if evidence.get("headcount") else "REVIEW"),
        "annual_filing": evidence.get("annual_filing"),
        "officer_count": evidence.get("officer_count"),
        "hidden_support_sheets": [name for name in BACKEND_SHEETS_TO_HIDE if name in wb.sheetnames],
        "enhanced_at": datetime.now().astimezone().isoformat(timespec="seconds"),
    }


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Recover public evidence and polish an equity-research workbook.")
    parser.add_argument("workbook")
    parser.add_argument("ticker")
    args = parser.parse_args()
    print(apply_workbook_enhancements(args.workbook, args.ticker.upper().strip()))
