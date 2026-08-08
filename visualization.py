"""Automated visual dashboard for the equity research workbook."""

from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter

NAVY = "17365D"
BLUE = "2F75B5"
WHITE = "FFFFFF"
LIGHT = "F5F9FC"
PALE_BLUE = "D9EAF7"
PALE_GREEN = "E2F0D9"
PALE_RED = "FCE4D6"
LINK_GREEN = "008000"
GREY = "666666"
DARK = "1F1F1F"

FMT_PCT = "0.0%;[Red](0.0%);-"
FMT_PRICE = "$#,##0.00;[Red]($#,##0.00);-"
FMT_BN = "#,##0.0;[Red](#,##0.0);-"
FMT_MULT = "0.0x;[Red](0.0x);-"


def _fill(color):
    return PatternFill("solid", fgColor=color)


def _merge_once(ws, cell_range):
    if cell_range not in {str(r) for r in ws.merged_cells.ranges}:
        ws.merge_cells(cell_range)


def _style_section(ws, start_cell, end_col, title):
    row = ws[start_cell].row
    _merge_once(ws, f"{start_cell}:{end_col}{row}")
    cell = ws[start_cell]
    cell.value = title
    cell.fill = _fill(NAVY)
    cell.font = Font(bold=True, color=WHITE, size=11)
    cell.alignment = Alignment(horizontal="left")


def _style_card(ws, title_range, value_range, title, formula, number_format):
    _merge_once(ws, title_range)
    _merge_once(ws, value_range)
    t = ws[title_range.split(":")[0]]
    v = ws[value_range.split(":")[0]]
    t.value = title
    t.fill = _fill(BLUE)
    t.font = Font(bold=True, color=WHITE, size=10)
    t.alignment = Alignment(horizontal="center", vertical="center")
    v.value = formula
    v.fill = _fill(LIGHT)
    v.font = Font(bold=True, color=LINK_GREEN, size=16)
    v.alignment = Alignment(horizontal="center", vertical="center")
    v.number_format = number_format


def _chart_style(chart, title, width=13.5, height=7.2, legend=True):
    chart.title = title
    chart.width = width
    chart.height = height
    chart.style = 10
    chart.legend = chart.legend if legend else None
    if chart.legend is not None:
        chart.legend.position = "b"


def ensure_visual_dashboard(wb, ticker=None):
    """Create/rebuild a presentation dashboard linked to the core equity model."""
    required = {"Company Data", "Historical Financials", "Three-Case Scenarios", "DCF", "Peer Comps"}
    if not required.issubset(set(wb.sheetnames)):
        return

    if "Visual Dashboard" in wb.sheetnames:
        old = wb["Visual Dashboard"]
        wb.remove(old)
    ws = wb.create_sheet("Visual Dashboard")
    ws.sheet_view.showGridLines = False
    try:
        wb.active = wb.index(ws)
        if wb.views:
            wb.views[0].activeTab = wb.index(ws)
            wb.views[0].firstSheet = wb.index(ws)
    except Exception:
        pass

    hist = wb["Historical Financials"]
    scenarios = wb["Three-Case Scenarios"]
    peers = wb["Peer Comps"]

    for col in range(1, 17):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["I"].width = 14
    for col in range(24, 32):
        ws.column_dimensions[get_column_letter(col)].width = 16
        ws.column_dimensions[get_column_letter(col)].hidden = True

    _merge_once(ws, "A1:P2")
    ws["A1"] = "Equity Research — Visual Dashboard"
    ws["A1"].fill = _fill(NAVY)
    ws["A1"].font = Font(bold=True, color=WHITE, size=18)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25

    _merge_once(ws, "A3:P3")
    ws["A3"] = '=CONCAT("Automated visual summary for ",\'Company Data\'!B4," | All charts link to model outputs")'
    ws["A3"].font = Font(italic=True, color=GREY, size=10)

    cards = [
        ("A5:D5", "A6:D8", "Current Price", "='Company Data'!B8", FMT_PRICE),
        ("E5:H5", "E6:H8", "Base Intrinsic Value", "='Three-Case Scenarios'!C39", FMT_PRICE),
        ("I5:L5", "I6:L8", "Probability-Weighted Value", "='Three-Case Scenarios'!E39", FMT_PRICE),
        ("M5:P5", "M6:P8", "Base Upside / (Downside)", "='Three-Case Scenarios'!C40", FMT_PCT),
        ("A10:D10", "A11:D13", "Revenue CAGR (2020–25)", "=Dashboard!B8", FMT_PCT),
        ("E10:H10", "E11:H13", "2025 FCF Margin", "=Dashboard!B11", FMT_PCT),
        ("I10:L10", "I11:L13", "Severe Bear Value / Share", "='Three-Case Scenarios'!B63", FMT_PRICE),
        ("M10:P10", "M11:P13", "Robustness Status", "='Three-Case Scenarios'!B66", "@"),
    ]
    for args in cards:
        _style_card(ws, *args)

    ws.conditional_formatting.add("M6:P8", CellIsRule(operator="lessThan", formula=["0"], fill=_fill(PALE_RED)))
    ws.conditional_formatting.add("M6:P8", CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=_fill(PALE_GREEN)))

    _style_section(ws, "A15", "H", "Historical Financial Performance")
    _style_section(ws, "I15", "P", "Valuation & Stress")
    _style_section(ws, "A39", "H", "Peer Valuation")
    _style_section(ws, "I39", "P", "DCF Sensitivity")

    line1 = LineChart()
    _chart_style(line1, "Revenue & Free Cash Flow", width=13.0, height=7.2)
    cats = Reference(hist, min_col=2, max_col=7, min_row=3)
    line1.append(Series(Reference(hist, min_col=2, max_col=7, min_row=4), title="Revenue ($bn)"))
    line1.append(Series(Reference(hist, min_col=2, max_col=7, min_row=16), title="FCF ($bn)"))
    line1.set_categories(cats)
    line1.y_axis.title = "$bn"
    line1.x_axis.title = "Year"
    ws.add_chart(line1, "A17")

    line2 = LineChart()
    _chart_style(line2, "Operating & FCF Margins", width=13.0, height=4.2)
    line2.append(Series(Reference(hist, min_col=2, max_col=7, min_row=10), title="Operating Margin"))
    line2.append(Series(Reference(hist, min_col=2, max_col=7, min_row=17), title="FCF Margin"))
    line2.set_categories(cats)
    line2.y_axis.numFmt = "0%"
    ws.add_chart(line2, "A32")

    ws["X11"], ws["Y11"] = "Scenario", "Value / Share"
    scenario_labels = ["Bear", "Base", "Bull", "Probability Weighted", "Current Price"]
    scenario_formulas = [
        "='Three-Case Scenarios'!B39",
        "='Three-Case Scenarios'!C39",
        "='Three-Case Scenarios'!D39",
        "='Three-Case Scenarios'!E39",
        "='Company Data'!B8",
    ]
    for r, (label, formula) in enumerate(zip(scenario_labels, scenario_formulas), 12):
        ws.cell(r, 24, label)
        ws.cell(r, 25, formula)
        ws.cell(r, 25).number_format = FMT_PRICE

    bar1 = BarChart()
    _chart_style(bar1, "Scenario Value vs Current Price", width=13.0, height=6.8, legend=False)
    bar1.type = "col"
    bar1.add_data(Reference(ws, min_col=25, min_row=11, max_row=16), titles_from_data=True)
    bar1.set_categories(Reference(ws, min_col=24, min_row=12, max_row=16))
    bar1.y_axis.numFmt = "$0"
    ws.add_chart(bar1, "I17")

    stress = BarChart()
    _chart_style(stress, "Stress-Test Intrinsic Value / Share", width=13.0, height=5.4, legend=False)
    stress.type = "bar"
    stress.add_data(Reference(scenarios, min_col=7, min_row=47, max_row=58), titles_from_data=True)
    stress.set_categories(Reference(scenarios, min_col=1, min_row=48, max_row=58))
    stress.x_axis.numFmt = "$0"
    ws.add_chart(stress, "I31")

    peer_chart = BarChart()
    _chart_style(peer_chart, "Peer Valuation Multiples", width=13.0, height=7.5)
    peer_chart.type = "col"
    peer_cats = Reference(peers, min_col=2, min_row=4, max_row=9)
    peer_chart.append(Series(Reference(peers, min_col=3, min_row=3, max_row=9), title_from_data=True))
    peer_chart.append(Series(Reference(peers, min_col=5, min_row=3, max_row=9), title_from_data=True))
    peer_chart.set_categories(peer_cats)
    peer_chart.y_axis.numFmt = "0.0x"
    ws.add_chart(peer_chart, "A41")

    ws["I41"] = "WACC \\ TGR"
    for c, src_c in zip(range(10, 15), range(2, 7)):
        ws.cell(41, c, f"=DCF!{get_column_letter(src_c)}21")
        ws.cell(41, c).number_format = "0.0%"
    for r, src_r in zip(range(42, 47), range(22, 27)):
        ws.cell(r, 9, f"=DCF!A{src_r}")
        ws.cell(r, 9).number_format = "0.0%"
        for c, src_c in zip(range(10, 15), range(2, 7)):
            ws.cell(r, c, f"=DCF!{get_column_letter(src_c)}{src_r}")
            ws.cell(r, c).number_format = FMT_PRICE
    for cell in ws[41][8:14]:
        cell.fill = _fill(BLUE)
        cell.font = Font(bold=True, color=WHITE)
    for r in range(42, 47):
        ws.cell(r, 9).fill = _fill(PALE_BLUE)
        ws.cell(r, 9).font = Font(bold=True)
    ws.conditional_formatting.add(
        "J42:N46",
        ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50,
                       mid_color="FFEB84", end_type="max", end_color="63BE7B")
    )

    _merge_once(ws, "I48:P49")
    ws["I48"] = (
        "DCF heatmap: green cells indicate higher intrinsic values; red cells indicate lower values. "
        "The center of the grid corresponds closely to the Base WACC / terminal-growth assumptions."
    )
    ws["I48"].fill = _fill(LIGHT)
    ws["I48"].font = Font(italic=True, color=GREY)
    ws["I48"].alignment = Alignment(wrap_text=True, vertical="center")

    _merge_once(ws, "A59:P61")
    ws["A59"] = (
        "Use this sheet as the presentation layer. Assumptions remain editable in Three-Case Scenarios; "
        "historical data remains in Historical Financials; peer data remains in Peer Comps. "
        "Charts update when Excel recalculates after the Python updater refreshes the workbook."
    )
    ws["A59"].fill = _fill(PALE_BLUE)
    ws["A59"].font = Font(color="404040")
    ws["A59"].alignment = Alignment(wrap_text=True, vertical="center")

    for rng in ["A6", "E6", "I6", "M6", "A11", "E11", "I11", "M11"]:
        ws[rng].font = Font(bold=True, color=LINK_GREEN, size=16)

    return ws
