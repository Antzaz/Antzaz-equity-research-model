"""Professional-methodology extensions for AI, market expectations and institutional lenses.

The purpose is not to imitate proprietary sell-side or asset-manager models.  It translates
public professional research methods into auditable workbook diagnostics, while keeping
company facts, analyst assumptions and external framework anchors clearly separated.
"""

from __future__ import annotations

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment

from institutional_lenses import INSTITUTIONAL_LENSES

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; GOLD="FFF2CC"; LIGHT="F5F9FC"
PALE_GREEN="E2F0D9"; GREY="666666"; INPUT_BLUE="0000FF"; LINK_GREEN="008000"
THIN=Side(style="thin",color="D9E1F2")
FMT_PCT='0.0%;[Red](0.0%);-'; FMT_BN='#,##0.0;[Red](#,##0.0);-'; FMT_SCORE='0.0'; FMT_YEARS='0.0 "yrs"'

PROFESSIONAL_AI_FRAMEWORKS=[
    (
        "McKinsey Global Institute",
        "$2.6–4.4tn annual GenAI use-case value potential; ~75% concentrated in customer operations, marketing/sales, software engineering and R&D.",
        "Bottom-up use-case analysis: identify addressable business functions, adoption, measurable cost/revenue outcome and implementation capture.",
        "Use function exposure × adoption × realized benefit; avoid applying an economy-wide percentage directly to company revenue.",
        "Potential-value study, not a company forecast; realization depends on adoption and execution.",
        "https://www.mckinsey.com/capabilities/tech-and-ai/our-insights/the-economic-potential-of-generative-ai-the-next-productivity-frontier",
    ),
    (
        "Goldman Sachs Research",
        "AI automation exposure is assessed using labor costs as a share of sales and exposure of work/tasks to AI; case studies cited by GS showed ~25% average productivity improvement after adoption.",
        "Productivity-beneficiary framework: labor intensity × task exposure × adoption × realized productivity, then translate savings into earnings/FCF.",
        "Use as a sensitivity anchor for cost-side AI benefits, not as an automatic 25% cost cut.",
        "Company adoption, implementation quality and reinvestment determine how much productivity reaches margins.",
        "https://www.goldmansachs.com/insights/articles/why-ai-companies-may-invest-more-than-500-billion-in-2026",
    ),
    (
        "BlackRock Investment Institute",
        "Core investor questions: how costly AI is, whether capex and eventual revenues match in order of magnitude, and which layer of the AI value chain captures the value.",
        "Value-capture / capital-cycle analysis: revenue capture, capex burden, financing, scarcity bottlenecks, margins and returns on incremental capital.",
        "Track AI revenue-to-capex conversion, incremental ROIC, funding mix, and value-chain position rather than AI demand alone.",
        "Ultimate value capture remains uncertain and can migrate between model, cloud, chip, power and application layers.",
        "https://www.blackrock.com/corporate/insights/blackrock-investment-institute/publications/outlook",
    ),
    (
        "Goldman Sachs — market valuation lens",
        "AI productivity scenarios are translated into EPS-growth and valuation effects, while explicitly recognizing that high expectations can cause multiple compression if delivery disappoints.",
        "Translate operating benefit into EPS/FCF growth, then compare to what is already implied by price and valuation multiples.",
        "Use the existing reverse DCF and AI surprise scenarios to test whether AI is incremental to expectations, not merely positive in isolation.",
        "Macro/index estimates are not company-specific and timing is highly uncertain.",
        "https://www.goldmansachs.com/insights/articles/how-much-could-ai-boost-us-stocks",
    ),
]

PROFESSIONAL_TOOLKIT=[
    ("Incremental ROIC & capital allocation", "ROIC on new investment matters more than historic ROIC when valuing reinvestment.", "FCF & Capital Economics / Advanced Analytics", "https://www.morganstanley.com/im/en-us/institutional-investor/insights/series/consilient-observer.html"),
    ("Competitive Advantage Period (CAP)", "Value depends on ROIC spread, reinvestment and how long excess returns persist; infer the duration embedded in price.", "Market Expectations", "https://www.morganstanley.com/im/en-us/institutional-investor/insights/series/consilient-observer.html"),
    ("Bayes + reference-class base rates", "Start with historical base rates, then update the prior with company-specific evidence instead of treating a narrative as a probability.", "Base Rates & Probabilities", "https://www.morganstanley.com/im/en-fi/institutional-investor/insights/consilient-observer/bayes-and-base-rates-2.html"),
    ("Reverse DCF / expectations investing", "Solve for the growth, margin, capital intensity and duration required by the current price, then compare those hurdles with your forecast.", "Market Expectations / Advanced Analytics", "https://www.morganstanley.com/im/en-us/institutional-investor/insights/series/consilient-observer.html"),
    ("Consensus dispersion & revisions", "The level of consensus is less informative without the estimate range, analyst coverage and direction of revisions.", "Expectations & Consensus", "https://site.financialmodelingprep.com/developer/docs/stable/financial-estimates"),
    ("AI value-capture & capital-cycle analysis", "Separate demand from monetization, margin capture, capital burden and scarcity; test incremental AI ROIC and FCF.", "AI Impact Analysis", "https://www.blackrock.com/corporate/insights/blackrock-investment-institute/publications/outlook"),
    ("Segment / SOTP economics", "Underwrite business lines independently where growth, margins, capital intensity and competitive position differ materially.", "Segment Analysis / Business Portfolio Map", "https://www.mckinsey.com/capabilities/strategy-and-corporate-finance/our-insights"),
    ("Earnings quality / cash conversion", "Reconcile accounting profit to OCF and FCF, including SBC, capex and working capital, before capitalizing earnings.", "FCF & Capital Economics", "https://pages.stern.nyu.edu/~adamodar/"),
]


def _fill(c): return PatternFill("solid",fgColor=c)

def _section(ws,row,title,end):
    for c in range(1,end+1):
        ws.cell(row,c).fill=_fill(NAVY); ws.cell(row,c).font=Font(bold=True,color=WHITE,size=11)
    ws.cell(row,1,title)

def _header(ws,row,headers):
    for c,v in enumerate(headers,1):
        cell=ws.cell(row,c,v); cell.fill=_fill(BLUE); cell.font=Font(bold=True,color=WHITE)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=Border(bottom=THIN)

def _find_row(ws,label):
    for r in range(1,ws.max_row+1):
        if str(ws.cell(r,1).value or "").strip()==label: return r
    return None

def _input(cell,note=None,fmt=FMT_PCT):
    cell.fill=_fill(GOLD); cell.font=Font(color=INPUT_BLUE)
    if fmt: cell.number_format=fmt
    if note: cell.comment=Comment(note,"OpenAI")


def extend_ai_impact_analysis(wb,ticker):
    if "AI Impact Analysis" not in wb.sheetnames: return False
    ws=wb["AI Impact Analysis"]
    # Idempotently clear a prior appended methodology block.
    marker=_find_row(ws,"Professional AI Impact Benchmarks — How Industry Research Sizes AI")
    if marker:
        for r in range(marker,ws.max_row+1):
            for c in range(1,15): ws.cell(r,c).value=None
        start=marker
    else:
        start=ws.max_row+2

    _section(ws,start,"Professional AI Impact Benchmarks — How Industry Research Sizes AI",14)
    _header(ws,start+1,["Professional Source","Published Estimate / Question","Estimation Method","How This Model Should Use It","Important Caveat","Source URL"])
    for rr,row in enumerate(PROFESSIONAL_AI_FRAMEWORKS,start+2):
        for c,v in enumerate(row,1): ws.cell(rr,c,v)
        ws.cell(rr,6).hyperlink=row[5]; ws.cell(rr,6).font=Font(color=LINK_GREEN,underline="single")
        for c in range(1,7): ws.cell(rr,c).alignment=Alignment(wrap_text=True,vertical="top")
        ws.row_dimensions[rr].height=74

    eco=start+len(PROFESSIONAL_AI_FRAMEWORKS)+4
    _section(ws,eco,"AI Economics Bridge — From Adoption to Incremental FCF and ROIC",14)
    ws.cell(eco+1,1,"This bridge deliberately leaves company-specific adoption/value-capture assumptions editable. Professional research supplies reference anchors, not a universal company forecast.")
    ws.merge_cells(start_row=eco+1,start_column=1,end_row=eco+1,end_column=14); ws.cell(eco+1,1).font=Font(italic=True,color=GREY); ws.cell(eco+1,1).alignment=Alignment(wrap_text=True)
    _header(ws,eco+2,["Driver","Analyst Input","Professional Anchor / Interpretation","Calculated $bn Impact","Investment Question","Status"])
    inputs=[
        ("Revenue exposed to AI",None,"Company/segment-specific share of revenue where AI can change demand, price or share."),
        ("AI adoption / attach rate",None,"Use reported seats, workloads, MAU, paid users or customer penetration where possible."),
        ("Revenue uplift at adopted exposure",None,"Price/ARPU/conversion/share uplift. McKinsey sizes revenue use cases from measurable outcome improvement."),
        ("Incremental AI gross margin",None,"Reflect inference/serving cost and revenue mix; do not assume software-like margins automatically."),
        ("Opex / labor cost addressable by AI",None,"Goldman uses labor-cost intensity and automation exposure as key productivity-beneficiary inputs."),
        ("Realized productivity capture",None,"Case studies can show large task-level gains, but only the realized/redeployed share should flow to earnings."),
        ("Incremental AI capex / revenue",None,"BlackRock/Goldman emphasize front-loaded capex and the need to link investment to eventual revenues."),
        ("Incremental inference / operating cost as % AI revenue",None,"Captures ongoing compute, model, data and serving costs not included in capex."),
    ]
    first=eco+3
    for r,(label,default,note) in enumerate(inputs,first):
        ws.cell(r,1,label); ws.cell(r,2,default); _input(ws.cell(r,2),note); ws.cell(r,3,note); ws.cell(r,3).alignment=Alignment(wrap_text=True)
        ws.cell(r,6,"INPUT"); ws.cell(r,6).fill=_fill(GOLD)
    # Calculation block uses latest reported revenue/EBIT, so it remains company-scaled.
    out=first+len(inputs)+1
    _header(ws,out,["Output","Formula / Result","Interpretation","","","Status"])
    rev_ref="'Historical Financials'!G4"; op_ref="'Historical Financials'!G9"
    row_rev=first; row_adopt=first+1; row_uplift=first+2; row_gm=first+3; row_opex=first+4; row_prod=first+5; row_capex=first+6; row_infer=first+7
    outputs=[
        ("Incremental AI revenue",f'=IF(COUNT(B{row_rev}:B{row_uplift})<3,"",{rev_ref}*B{row_rev}*B{row_adopt}*B{row_uplift})',FMT_BN,"Monetization beyond what is already in the Base case should be compared with the existing AI surprise scenarios."),
        ("Incremental AI gross profit",f'=IF(B{out+1}="","",B{out+1}*B{row_gm})',FMT_BN,"Revenue value capture after AI-specific gross/inference economics."),
        ("Potential annual opex savings",f'=IF(COUNT(B{row_opex}:B{row_prod})<2,"",MAX(0,{rev_ref}-{op_ref})*B{row_opex}*B{row_prod})',FMT_BN,"A cost-side productivity scenario; review whether savings are retained or reinvested."),
        ("Incremental AI operating/inference cost",f'=IF(B{out+1}="","",B{out+1}*B{row_infer})',FMT_BN,"Recurring serving/model/data cost not captured by infrastructure capex."),
        ("Incremental AI EBIT before D&A effects",f'=IF(COUNT(B{out+2}:B{out+4})<3,"",B{out+2}+B{out+3}-B{out+4})',FMT_BN,"Operating benefit before incremental depreciation and financing effects."),
        ("Incremental AI capex",f'=IF(B{row_capex}="","",{rev_ref}*B{row_capex})',FMT_BN,"Annual incremental capital burden associated with the AI scenario."),
        ("Pre-tax AI ROIC proxy",f'=IFERROR(B{out+5}/B{out+6},"")',FMT_PCT,"Simple incremental EBIT / incremental capex proxy; replace capex with cumulative invested capital for a full ROIC study."),
        ("Pre-tax AI FCF bridge",f'=IF(OR(B{out+5}="",B{out+6}=""),"",B{out+5}-B{out+6})',FMT_BN,"Direct test of whether AI improves near-term owner cash generation after the capital burden."),
        ("Capex payback proxy",f'=IFERROR(B{out+6}/MAX(0.0001,B{out+5}),"")',FMT_YEARS,"Years of incremental EBIT required to recover one year of incremental capex; use cautiously for long-lived infrastructure."),
    ]
    for rr,(label,formula,fmt,note) in enumerate(outputs,out+1):
        ws.cell(rr,1,label); ws.cell(rr,2,formula); ws.cell(rr,2).number_format=fmt; ws.cell(rr,3,note); ws.cell(rr,3).alignment=Alignment(wrap_text=True); ws.cell(rr,6,"CALCULATED")
    ws.column_dimensions["A"].width=max(ws.column_dimensions["A"].width or 0,34)
    ws.column_dimensions["B"].width=max(ws.column_dimensions["B"].width or 0,18)
    ws.column_dimensions["C"].width=max(ws.column_dimensions["C"].width or 0,55)
    return True


def extend_institutional_comparison(wb):
    if "Institutional Comparison" not in wb.sheetnames: return False
    ws=wb["Institutional Comparison"]
    marker=_find_row(ws,"Score Construction & Provenance — Where the Institutional Scores Come From")
    if marker:
        for r in range(marker,ws.max_row+1):
            for c in range(1,16): ws.cell(r,c).value=None
        start=marker
    else:
        start=ws.max_row+2

    _section(ws,start,"Score Construction & Provenance — Where the Institutional Scores Come From",15)
    ws.cell(start+1,1,"Formula")
    ws.cell(start+1,2,"Fit Score = Σ(company dimension score × lens weight) / Σ(available lens weights). Missing company dimensions are excluded and Data Coverage shows the weight actually represented.")
    ws.merge_cells(start_row=start+1,start_column=2,end_row=start+1,end_column=15); ws.cell(start+1,2).alignment=Alignment(wrap_text=True)
    ws.cell(start+2,1,"Important")
    ws.cell(start+2,2,"The weights are this project's transparent interpretation of each firm's published philosophy. They are NOT the firm's proprietary weights, rating model, target price or recommendation. Fit thresholds are also this project's rubric.")
    ws.merge_cells(start_row=start+2,start_column=2,end_row=start+2,end_column=15); ws.cell(start+2,2).font=Font(italic=True,color=GREY); ws.cell(start+2,2).alignment=Alignment(wrap_text=True)

    dim=start+4
    _section(ws,dim,"Company Dimension Provenance",15)
    _header(ws,dim+1,["Dimension","Primary Workbook Evidence","Construction / Interpretation","Main Limitation"])
    dims=[
        ("Growth","Advanced Analytics / Historical Financials","0–100 growth score derived from historical revenue growth and model analytics.","Past growth may not represent future reinvestment runway."),
        ("Profitability","Advanced Analytics / Historical Financials","Operating profitability score; peer margin evidence is a cross-check.","Accounting margin alone does not measure incremental returns."),
        ("FCF Quality","Advanced Analytics / FCF & Capital Economics","Cash conversion / FCF economics rather than EPS alone.","Capex cycle and working-capital timing can distort one period."),
        ("Balance Sheet","Company Data / Advanced Analytics","Net cash/debt and financing resilience proxy.","Does not fully model off-balance-sheet obligations."),
        ("Absolute Valuation","Advanced Analytics / DCF","Intrinsic-value output relative to market price.","Highly sensitive to long-duration assumptions."),
        ("Relative Valuation","Peer Comps","Target multiples relative to direct/exact peer medians.","Peers can have structurally different growth/margins/capital intensity."),
        ("Stress Robustness","Advanced Analytics / Stress Test","Downside valuation resilience under combined severe assumptions.","Stress scenario is a model, not a probability forecast."),
        ("Leadership","Leadership & Culture","Execution, capital allocation, public leadership depth, culture evidence and governance disclosure proxy.","Not a factual management-quality rating."),
        ("Moat / Position","Leadership & Culture + profitability + source-scoped market position","Observable proxy combining economics, leadership evidence and comparable position/share where available.","Cannot reproduce a proprietary qualitative moat assessment."),
    ]
    for rr,row in enumerate(dims,dim+2):
        for c,v in enumerate(row,1): ws.cell(rr,c,v); ws.cell(rr,c).alignment=Alignment(wrap_text=True,vertical="top")

    weights=dim+len(dims)+4
    _section(ws,weights,"Institution Lens Weight Map — Explicit Project Assumptions",15)
    dimensions=["Growth","Profitability","FCF Quality","Balance Sheet","Absolute Valuation","Relative Valuation","Stress Robustness","Leadership","Moat / Position"]
    _header(ws,weights+1,["Institution / Style"]+dimensions+["Published Principle Source"])
    for rr,lens in enumerate(INSTITUTIONAL_LENSES,weights+2):
        ws.cell(rr,1,lens["firm"])
        for c,d in enumerate(dimensions,2):
            ws.cell(rr,c,lens["weights"].get(d,0)); ws.cell(rr,c).number_format=FMT_PCT
        ws.cell(rr,11,lens["source"]); ws.cell(rr,11).hyperlink=lens["source"]; ws.cell(rr,11).font=Font(color=LINK_GREEN,underline="single")
    tool=weights+len(INSTITUTIONAL_LENSES)+4
    _section(ws,tool,"Professional Analysis Toolkit — Additional Methods Worth Using",15)
    _header(ws,tool+1,["Analysis","Why Professionals Use It","Where It Fits in This Workbook","Public Methodology Source"])
    for rr,row in enumerate(PROFESSIONAL_TOOLKIT,tool+2):
        for c,v in enumerate(row,1): ws.cell(rr,c,v); ws.cell(rr,c).alignment=Alignment(wrap_text=True,vertical="top")
        ws.cell(rr,4).hyperlink=row[3]; ws.cell(rr,4).font=Font(color=LINK_GREEN,underline="single")
        ws.row_dimensions[rr].height=48
    for col,width in {"A":31,"B":52,"C":40,"D":58,"E":16,"F":16,"G":16,"H":16,"I":16,"J":16,"K":58}.items():
        ws.column_dimensions[col].width=max(ws.column_dimensions[col].width or 0,width)
    return True


def ensure_professional_analysis_extensions(wb,ticker):
    return {
        "ai": extend_ai_impact_analysis(wb,ticker),
        "institutional": extend_institutional_comparison(wb),
    }
