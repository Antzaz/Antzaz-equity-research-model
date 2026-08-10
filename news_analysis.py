"""Recent company news, sentiment and investment-impact worksheet.

Uses yfinance's ticker-specific news feed. Headlines and publisher summaries are kept short;
sentiment and company impact are transparent rule-based research aids rather than claims about
future stock performance. The original article URL is always retained for verification.
"""

import re
from datetime import datetime, timezone

try:
    import yfinance as yf
except Exception:
    yf=None

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; GREY="666666"
PALE_GREEN="E2F0D9"; PALE_RED="FCE4D6"; PALE_YELLOW="FFF2CC"; LINK_GREEN="008000"
THIN=Side(style="thin",color="D9E1F2")

POSITIVE={
    "beat":2,"beats":2,"surge":2,"record":2,"raises guidance":3,"raised guidance":3,"upgrade":2,
    "upgraded":2,"wins":2,"contract win":2,"approval":2,"approved":2,"expands":1,"growth":1,
    "strong":1,"profit rises":2,"revenue rises":2,"buyback":2,"repurchase":2,"dividend increase":2,
    "partnership":1,"launches":1,"outperform":2,"acquisition completed":1,"rebound":1,"accelerates":1,
}
NEGATIVE={
    "miss":-2,"misses":-2,"cuts guidance":-3,"cut guidance":-3,"downgrade":-2,"downgraded":-2,
    "lawsuit":-2,"sued":-2,"probe":-2,"investigation":-2,"recall":-2,"breach":-3,"cyberattack":-3,
    "fine":-2,"penalty":-2,"warning":-2,"weak":-1,"decline":-1,"falls":-1,"drops":-1,"layoffs":-1,
    "job cuts":-1,"impairment":-2,"charge":-1,"loss widens":-2,"default":-3,"fraud":-3,
    "regulatory scrutiny":-2,"antitrust":-2,"restructuring":-1,"slump":-2,"disappoint":-2,
}


def _fill(c): return PatternFill("solid",fgColor=c)
def _clean(v): return re.sub(r"\s+"," ",str(v or "")).strip()
def _clip(v,n=260):
    s=_clean(v)
    if len(s)<=n: return s
    cut=s[:n-1].rsplit(" ",1)[0]
    return (cut or s[:n-1])+"…"
def _nested(obj,*path):
    cur=obj
    for key in path:
        if not isinstance(cur,dict): return None
        cur=cur.get(key)
    return cur

def _date(v):
    if v in (None,""): return None
    if isinstance(v,(int,float)):
        try: return datetime.fromtimestamp(v,tz=timezone.utc).strftime("%Y-%m-%d")
        except Exception: return None
    s=str(v)
    try: return datetime.fromisoformat(s.replace("Z","+00:00")).strftime("%Y-%m-%d")
    except Exception: return s[:10]
def _url(item,content):
    candidates=[_nested(content,"canonicalUrl","url"),_nested(content,"clickThroughUrl","url"),content.get("link") if isinstance(content,dict) else None,item.get("link"),item.get("url")]
    return next((str(x) for x in candidates if x),None)
def _publisher(item,content):
    p=_nested(content,"provider","displayName") or item.get("publisher") or item.get("provider"); return _clean(p)
def _normalize_item(item):
    if not isinstance(item,dict): return None
    content=item.get("content") if isinstance(item.get("content"),dict) else item
    title=_clean(content.get("title") or item.get("title"))
    if not title: return None
    summary=_clean(content.get("summary") or content.get("description") or item.get("summary") or item.get("description"))
    published=_date(content.get("pubDate") or content.get("displayTime") or item.get("providerPublishTime") or item.get("pubDate"))
    return {"title":title,"summary":summary,"date":published,"publisher":_publisher(item,content),"url":_url(item,content)}
def _news(ticker,count=40):
    if yf is None: return []
    try:
        t=yf.Ticker(ticker)
        try: raw=t.get_news(count=count,tab="all") or []
        except Exception: raw=getattr(t,"news",[]) or []
    except Exception: return []
    out=[]; seen=set()
    for item in raw:
        x=_normalize_item(item)
        if not x: continue
        key=(x["url"] or "",x["title"].lower())
        if key in seen: continue
        seen.add(key); out.append(x)
    out.sort(key=lambda x:x.get("date") or "",reverse=True); return out

def _sentiment(title,summary):
    text=(title+" "+summary).lower(); score=0; drivers=[]
    for phrase,w in POSITIVE.items():
        if phrase in text: score+=w; drivers.append(phrase)
    for phrase,w in NEGATIVE.items():
        if phrase in text: score+=w; drivers.append(phrase)
    if score>=2: return "Positive",score,drivers
    if score<=-2: return "Negative",score,drivers
    return "Mixed / Neutral",score,drivers

def _impact(title,summary,tone):
    text=(title+" "+summary).lower(); sign="support" if tone=="Positive" else ("pressure" if tone=="Negative" else "affect")
    if any(k in text for k in ("earnings","revenue","profit","guidance","forecast","outlook","quarter","results")): return f"Could {sign} near-term earnings expectations and valuation if the development persists; compare it with the model's revenue, margin and FCF assumptions."
    if any(k in text for k in ("upgrade","downgrade","price target","rating","analyst")): return "Primarily affects market expectations and sentiment; the fundamental impact is limited unless estimates or the cost of capital also change."
    if any(k in text for k in ("acquisition","acquire","merger","deal","divest","sale of")): return "Affects strategic positioning, capital allocation and integration/execution risk; check whether expected returns exceed the company's cost of capital."
    if any(k in text for k in ("lawsuit","probe","investigation","regulatory","antitrust","fine","penalty","recall","fraud")): return f"Could {sign} cash flow through legal/regulatory costs, operational constraints or reputation; monitor probability, duration and potential financial exposure."
    if any(k in text for k in ("contract","partnership","launch","product","approval","customer","order","backlog")): return f"Could {sign} future revenue growth and competitive positioning; the key question is whether it is material enough to change segment growth or margins."
    if any(k in text for k in ("layoff","job cuts","restructuring","cost cut","efficiency")): return "May improve cost efficiency and margins, but can also signal weaker demand or execution pressure; watch revenue growth and restructuring charges together."
    if any(k in text for k in ("buyback","repurchase","dividend","capital return")): return f"Could {sign} per-share value through capital returns and share-count changes; assess the return on capital versus alternative reinvestment uses."
    if any(k in text for k in ("interest rate","rates","inflation","recession","credit","tariff","currency")): return "Changes the macro backdrop and may affect demand, investment income, credit losses, pricing or discount rates depending on the business model."
    return f"Potentially {sign}s the investment thesis through expectations, execution or competitive positioning; verify materiality against the company's financial model."

def _style_title(ws,text):
    for c in range(1,8): ws.cell(1,c).fill=_fill(NAVY); ws.cell(2,c).fill=_fill(NAVY)
    ws["A1"]=text; ws["A1"].font=Font(bold=True,color=WHITE,size=19); ws.sheet_view.showGridLines=False

def _section(ws,row,title,end=7):
    for c in range(1,end+1): ws.cell(row,c).fill=_fill(NAVY); ws.cell(row,c).font=Font(bold=True,color=WHITE,size=11)
    ws.cell(row,1,title)
def _header(ws,row):
    heads=["Date","Tone","Headline","Publisher","Short Summary","Potential Company Impact","Article Link"]
    for c,v in enumerate(heads,1):
        x=ws.cell(row,c,v); x.fill=_fill(BLUE); x.font=Font(bold=True,color=WHITE); x.border=Border(bottom=THIN); x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def _write_rows(ws,start,rows):
    _header(ws,start); r=start+1
    for item in rows:
        vals=[item["date"],item["tone"],item["title"],item["publisher"],_clip(item["summary"] or item["title"]),item["impact"],item["url"]]
        for c,v in enumerate(vals,1): ws.cell(r,c,v)
        ws.cell(r,2).fill=_fill(PALE_GREEN if item["tone"]=="Positive" else (PALE_RED if item["tone"]=="Negative" else PALE_YELLOW)); ws.cell(r,2).font=Font(bold=True)
        if item["url"]:
            ws.cell(r,7).hyperlink=item["url"]
            ws.cell(r,7).font=Font(color=LINK_GREEN,underline="single")
        for c in range(1,8): ws.cell(r,c).alignment=Alignment(wrap_text=True,vertical="top")
        ws.row_dimensions[r].height=48; r+=1
    if not rows:
        ws.cell(r,1,"No qualifying stories returned in this category from the current public news feed."); ws.cell(r,1).font=Font(italic=True,color=GREY); r+=1
    return r

def ensure_news_analysis(wb,ticker,max_per_tone=8):
    name="Recent News & Impact"
    if name in wb.sheetnames: wb.remove(wb[name])
    ws=wb.create_sheet(name); _style_title(ws,f"{ticker} — Recent News & Investment Impact")
    ws["A3"]=("Ticker-specific public news via Yahoo Finance/yfinance. Tone and impact are rule-based research aids, not investment recommendations. Open the linked original article before changing a thesis or forecast.")
    ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True); ws.row_dimensions[3].height=34
    items=[]
    for x in _news(ticker,40):
        tone,score,_=_sentiment(x["title"],x["summary"]); x["tone"]=tone; x["score"]=score; x["impact"]=_impact(x["title"],x["summary"],tone); items.append(x)
    groups=[("Positive Developments",[x for x in items if x["tone"]=="Positive"][:max_per_tone]),("Negative Developments",[x for x in items if x["tone"]=="Negative"][:max_per_tone]),("Mixed / Neutral Context",[x for x in items if x["tone"]=="Mixed / Neutral"][:max_per_tone])]
    row=5
    for title,rows in groups: _section(ws,row,title); row=_write_rows(ws,row+1,rows)+1
    _section(ws,row,"News Balance & Method"); pos=sum(x["tone"]=="Positive" for x in items); neg=sum(x["tone"]=="Negative" for x in items); neu=len(items)-pos-neg
    ws.cell(row+1,1,"Stories reviewed"); ws.cell(row+1,2,len(items)); ws.cell(row+2,1,"Positive / Negative / Mixed"); ws.cell(row+2,2,f"{pos} / {neg} / {neu}")
    ws.cell(row+3,1,"Method"); ws.cell(row+3,2,"Headline + available publisher/Yahoo summary keyword classification; no article text is invented. Missing summaries stay short and article links are retained for verification."); ws.cell(row+3,2).alignment=Alignment(wrap_text=True)
    widths={"A":13,"B":17,"C":48,"D":20,"E":55,"F":64,"G":55}
    for col,w in widths.items(): ws.column_dimensions[col].width=w
    ws.freeze_panes="A7"; return ws
