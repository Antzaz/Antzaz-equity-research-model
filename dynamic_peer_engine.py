"""Automatic, sector-aware comparable-company selection with fallback metric recovery.

The TARGET ticker determines the peer universe. Exact-industry peers are preferred, but
business-model peers may be included when the public classification is imperfect. The peer
sheet deliberately separates direct operating peers from broader strategic/sector comps.

Peer Comps keeps three concepts distinct:
- Industry Market Share %: only when a comparable external industry source exists.
- Peer-Set Market Cap %: market-cap weight inside the selected comparison set.
- Peer Type: direct business-model, exact-industry, or broader same-sector comparison.

Live provider fields are supplemented with statement-derived calculations when practical.
Deep statement fallback work is performed only for the selected peer set, not every
discovery candidate, so expanded peer analysis remains practical.
"""

from __future__ import annotations

import math

try:
    import yfinance as yf
    from yfinance import EquityQuery
except Exception:
    yf = None
    EquityQuery = None

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from market_context import market_share_record, preferred_peer_symbols

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; GREY="666666"
INPUT_BLUE="0000FF"; LINK_GREEN="008000"; GOLD="FFF2CC"
FMT_PCT='0.0%;[Red](0.0%);-'; FMT_MULT='0.0x;[Red](0.0x);-'
THIN=Side(style="thin",color="D9E1F2")

SECTOR_FALLBACK={
    "Industrials":["ETN","PH","CMI","EMR","ROK","ITW","AME","IR","CAT","HON","DE","GE"],
    "Healthcare":["UNH","ELV","CI","HUM","CNC","MOH","JNJ","LLY","ABBV","MRK","TMO","ABT"],
    "Technology":["MSFT","AAPL","NVDA","AVGO","ORCL","CRM","ADBE","AMD","QCOM","NOW","INTU","TXN","MU","INTC"],
    "Communication Services":["GOOGL","META","NFLX","DIS","CMCSA","TMUS","VZ","T","SPOT","PINS"],
    "Consumer Cyclical":["AMZN","TSLA","HD","MCD","LOW","BKNG","TJX","NKE","SBUX","MELI"],
    "Consumer Defensive":["WMT","COST","PG","KO","PEP","PM","TGT","CL","MDLZ"],
    "Financial Services":["JPM","BAC","WFC","C","GS","MS","USB","PNC","BLK","SCHW","AXP","PRU","AFL","PFG","GL","UNM","MFC","SLF"],
    "Energy":["XOM","CVX","COP","EOG","SLB","MPC","PSX","OXY","VLO"],
    "Basic Materials":["LIN","APD","SHW","FCX","NEM","NUE","DOW","ECL"],
    "Real Estate":["PLD","AMT","EQIX","WELL","SPG","O","DLR","PSA"],
    "Utilities":["NEE","SO","DUK","CEG","AEP","SRE","D","EXC"],
}

INDUSTRY_FALLBACK={
    "Insurance - Life":["PRU","AFL","PFG","GL","UNM","MFC","SLF"],
    "Insurance—Life":["PRU","AFL","PFG","GL","UNM","MFC","SLF"],
    "Healthcare Plans":["UNH","ELV","CI","HUM","CNC","MOH"],
    "Specialty Industrial Machinery":["ETN","PH","CMI","EMR","ROK","ITW","AME","IR"],
    "Semiconductors":["NVDA","AVGO","AMD","QCOM","INTC","MU","TXN","ADI","MRVL","UMC","GFS","TSEM","SMICY"],
    "Software - Infrastructure":["MSFT","ORCL","CRM","NOW","PLTR","SNOW","DDOG","MDB"],
    "Software—Infrastructure":["MSFT","ORCL","CRM","NOW","PLTR","SNOW","DDOG","MDB"],
    "Internet Content & Information":["GOOGL","META","PINS","SNAP","RDDT"],
    "Internet Retail":["AMZN","MELI","EBAY","ETSY","CHWY"],
}

COUNTRY_TO_REGION={
    "United States":"US","Canada":"CA","Finland":"FI","Sweden":"SE","Norway":"NO","Denmark":"DK",
    "Germany":"DE","France":"FR","United Kingdom":"GB","Switzerland":"CH","Netherlands":"NL","Japan":"JP",
    "Australia":"AU","China":"CN","Hong Kong":"HK","India":"IN","Singapore":"SG","Italy":"IT","Spain":"ES",
    "Taiwan":"TW","South Korea":"KR",
}

STRATEGIC_PEERS={
    "TSM":["INTC","MU","NVDA","AVGO","AMD","QCOM","TXN","ADI"],
    "GOOGL":["META","MSFT","AMZN","NFLX","ORCL","CRM","AAPL"],
    "GOOG":["META","MSFT","AMZN","NFLX","ORCL","CRM","AAPL"],
}

MAX_DISCOVERY_CANDIDATES=40


def _fill(c): return PatternFill("solid",fgColor=c)

def _num(v,default=None):
    try:
        if isinstance(v,bool) or v in (None,""): return default
        x=float(v)
        return x if math.isfinite(x) else default
    except Exception:
        return default


def _row_value(df,labels,col_index=0):
    if df is None or getattr(df,"empty",True): return None
    try:
        normalized={str(x).strip().lower():x for x in df.index}
        row=None
        for label in labels:
            key=str(label).strip().lower()
            if key in normalized:
                row=normalized[key]; break
        if row is None: return None
        values=list(df.loc[row].values)
        if col_index >= len(values): return None
        return _num(values[col_index])
    except Exception:
        return None


def _raw_info(symbol):
    if yf is None: return {}
    try: return yf.Ticker(symbol).info or {}
    except Exception: return {}


def _enrich_info(symbol,raw=None):
    """Fill missing public market metrics from yfinance statements when possible."""
    info=dict(raw or {})
    notes=[]
    if yf is None: return info,notes
    try: t=yf.Ticker(symbol)
    except Exception: return info,notes

    try:
        fast=getattr(t,"fast_info",None)
        if fast:
            if info.get("marketCap") in (None,""):
                mc=_num(getattr(fast,"market_cap",None) if not isinstance(fast,dict) else fast.get("market_cap"))
                if mc: info["marketCap"]=mc; notes.append("market cap from fast_info")
            if info.get("currentPrice") in (None,""):
                px=_num(getattr(fast,"last_price",None) if not isinstance(fast,dict) else fast.get("last_price"))
                if px: info["currentPrice"]=px; notes.append("price from fast_info")
    except Exception: pass

    try: income=t.income_stmt
    except Exception: income=None
    try: balance=t.balance_sheet
    except Exception: balance=None
    try: cashflow=t.cashflow
    except Exception: cashflow=None

    revenue=_row_value(income,["Total Revenue","Operating Revenue"],0)
    revenue_prev=_row_value(income,["Total Revenue","Operating Revenue"],1)
    op_income=_row_value(income,["Operating Income"],0)
    net_income=_row_value(income,["Net Income","Net Income Common Stockholders"],0)
    equity0=_row_value(balance,["Stockholders Equity","Total Equity Gross Minority Interest"],0)
    equity1=_row_value(balance,["Stockholders Equity","Total Equity Gross Minority Interest"],1)
    ebitda=_row_value(income,["EBITDA","Normalized EBITDA"],0)
    if ebitda is None:
        da=_row_value(cashflow,["Depreciation And Amortization","Depreciation Amortization Depletion"],0)
        if op_income is not None and da is not None: ebitda=op_income+abs(da)

    if info.get("revenueGrowth") in (None,"") and revenue and revenue_prev and revenue_prev>0:
        info["revenueGrowth"]=revenue/revenue_prev-1; notes.append("revenue growth calculated from annual statements")
    if info.get("operatingMargins") in (None,"") and revenue and op_income is not None and revenue!=0:
        info["operatingMargins"]=op_income/revenue; notes.append("operating margin calculated from annual statements")
    if info.get("returnOnEquity") in (None,"") and net_income is not None:
        avg_equity=(equity0+equity1)/2 if equity0 is not None and equity1 is not None else equity0
        if avg_equity and avg_equity>0:
            info["returnOnEquity"]=net_income/avg_equity; notes.append("ROE calculated from annual statements")
    if info.get("ebitda") in (None,"") and ebitda is not None:
        info["ebitda"]=ebitda; notes.append("EBITDA recovered/calculated from statements")
    if info.get("totalRevenue") in (None,"") and revenue is not None: info["totalRevenue"]=revenue

    ev=_num(info.get("enterpriseValue")); mc=_num(info.get("marketCap")); debt=_num(info.get("totalDebt")); cash=_num(info.get("totalCash"))
    if ev is None and mc is not None:
        ev=mc+(debt or 0)-(cash or 0); info["enterpriseValue"]=ev; notes.append("enterprise value calculated from market cap/debt/cash")
    if info.get("enterpriseToRevenue") in (None,"") and ev is not None and revenue and revenue>0:
        info["enterpriseToRevenue"]=ev/revenue; notes.append("EV/Revenue calculated")
    if info.get("enterpriseToEbitda") in (None,"") and ev is not None and ebitda and ebitda>0:
        info["enterpriseToEbitda"]=ev/ebitda; notes.append("EV/EBITDA calculated")

    if info.get("forwardPE") in (None,""):
        px=_num(info.get("currentPrice") or info.get("regularMarketPrice")); feps=_num(info.get("forwardEps"))
        if feps in (None,0):
            try:
                est=getattr(t,"earnings_estimate",None)
                if est is not None and not getattr(est,"empty",True) and "0y" in est.index: feps=_num(est.loc["0y"].get("avg"))
            except Exception: pass
        if px and feps and feps>0:
            info["forwardPE"]=px/feps; notes.append("forward P/E calculated from price and forward EPS")
    return info,list(dict.fromkeys(notes))


def _target_classification(target):
    sector=str(target.get("sector") or "").strip() or "Unknown"
    industry=str(target.get("industry") or "").strip() or "Unknown"
    sector_key=str(target.get("sectorKey") or "").strip(); industry_key=str(target.get("industryKey") or "").strip()
    mc=_num(target.get("marketCap")); region=COUNTRY_TO_REGION.get(str(target.get("country") or ""),"US")
    return sector,industry,sector_key,industry_key,mc,region


def _symbols_from_frame(obj):
    out=[]
    if obj is None: return out
    try:
        cols=[str(c).lower() for c in obj.columns]
        for key in ("symbol","ticker"):
            if key in cols:
                col=obj.columns[cols.index(key)]; out.extend(str(x).upper() for x in obj[col].tolist() if x)
        for x in list(obj.index):
            s=str(x).upper().strip()
            if 1<=len(s)<=12 and any(ch.isalpha() for ch in s): out.append(s)
    except Exception: pass
    return out


def _symbols_from_screen(result):
    if not isinstance(result,dict): return []
    rows=result.get("quotes") or result.get("results") or []; out=[]
    if isinstance(rows,list):
        for row in rows:
            if isinstance(row,dict):
                s=row.get("symbol") or row.get("ticker")
                if s: out.append(str(s).upper())
    return out


def _discover_candidates(ticker,sector,industry,sector_key,industry_key,region):
    if yf is None or sector=="Unknown": return []
    out=[]
    out.extend(preferred_peer_symbols(ticker,industry)); out.extend(STRATEGIC_PEERS.get(str(ticker).upper(),[]))
    if industry_key:
        try:
            dom=yf.Industry(industry_key,region=region)
            for attr in ("top_companies","top_performing_companies","top_growth_companies"): out.extend(_symbols_from_frame(getattr(dom,attr,None)))
        except Exception: pass
    if EquityQuery is not None and industry!="Unknown":
        try:
            q=EquityQuery('and',[EquityQuery('eq',['region',region.lower()]),EquityQuery('eq',['sector',sector]),EquityQuery('eq',['industry',industry])])
            out.extend(_symbols_from_screen(yf.screen(q,size=50,sortField='intradaymarketcap',sortAsc=False)))
        except Exception: pass
    out.extend(INDUSTRY_FALLBACK.get(industry,[]))
    if sector_key:
        try: out.extend(_symbols_from_frame(yf.Sector(sector_key,region=region).top_companies))
        except Exception: pass
    out.extend(SECTOR_FALLBACK.get(sector,[]))
    seen=[]
    for symbol in out:
        symbol=str(symbol).upper().strip()
        if symbol and symbol not in seen: seen.append(symbol)
    return seen[:MAX_DISCOVERY_CANDIDATES]


def _rank(target_ticker,target_sector,target_industry,target_mc,symbol,info):
    preferred=preferred_peer_symbols(target_ticker,target_industry)
    if symbol in preferred: return -100+preferred.index(symbol)*.05
    candidate_sector=str(info.get("sector") or "").strip()
    if candidate_sector != target_sector: return None
    candidate_industry=str(info.get("industry") or "").strip()
    industry_penalty=0 if candidate_industry==target_industry else 20
    if symbol in STRATEGIC_PEERS.get(str(target_ticker).upper(),[]): industry_penalty=max(8,industry_penalty-8)
    mc=_num(info.get("marketCap")); size_penalty=abs(math.log(mc/target_mc)) if target_mc and mc and target_mc>0 and mc>0 else 4
    coverage=sum(info.get(k) not in (None,"") for k in ("forwardPE","enterpriseToRevenue","enterpriseToEbitda","revenueGrowth","operatingMargins","returnOnEquity"))
    return industry_penalty+size_penalty+(6-coverage)*.50


def _peer_type(target_ticker,target_industry,symbol,info):
    if symbol in preferred_peer_symbols(target_ticker,target_industry): return "Direct business-model peer"
    if str(info.get("industry") or "").strip()==target_industry: return "Exact industry"
    if symbol in STRATEGIC_PEERS.get(str(target_ticker).upper(),[]): return "Strategic sector comp"
    return "Same-sector fallback"


def select_dynamic_peers(wb,ticker,count=9):
    target_raw=_raw_info(ticker); sector,industry,sector_key,industry_key,target_mc,region=_target_classification(target_raw)
    target,target_notes=_enrich_info(ticker,target_raw); target["_metric_notes"]=target_notes
    if sector=="Unknown": return target,sector,industry,[]
    ranked=[]
    preferred=set(preferred_peer_symbols(ticker,industry))
    for symbol in _discover_candidates(ticker,sector,industry,sector_key,industry_key,region):
        if symbol==ticker.upper(): continue
        raw=_raw_info(symbol); score=_rank(ticker,sector,industry,target_mc,symbol,raw)
        if score is not None: ranked.append((score,symbol,raw))
        elif symbol in preferred and market_share_record(symbol): ranked.append((-90,symbol,raw))
    ranked.sort(key=lambda x:x[0])
    selected=[]
    for _,symbol,raw in ranked[:count]:
        info,notes=_enrich_info(symbol,raw); info["_metric_notes"]=notes; selected.append((symbol,info))
    return target,sector,industry,selected


def _coverage(info):
    keys=("forwardPE","enterpriseToRevenue","enterpriseToEbitda","revenueGrowth","operatingMargins","returnOnEquity")
    return sum(info.get(k) not in (None,"") for k in keys)/len(keys)


def _metric_row(symbol,info,sector,industry,method):
    notes=info.get("_metric_notes") or []; source_note="Live Yahoo fields" if not notes else "Live Yahoo + calculated fallback: "+"; ".join(notes)
    return [info.get("longName") or info.get("shortName") or symbol,symbol,info.get("forwardPE"),info.get("enterpriseToRevenue"),info.get("enterpriseToEbitda"),info.get("revenueGrowth"),info.get("operatingMargins"),info.get("returnOnEquity"),info.get("sector") or sector,info.get("industry") or industry,method,f"https://finance.yahoo.com/quote/{symbol}/",None,None,None,method,_coverage(info),source_note]


def _header(ws,row,start,end):
    for c in range(start,end+1):
        x=ws.cell(row,c); x.fill=_fill(BLUE); x.font=Font(bold=True,color=WHITE); x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); x.border=Border(bottom=THIN)


def _unmerge_peer_area(ws):
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row<=30 and merged.max_row>=1 and merged.min_col<=18 and merged.max_col>=1: ws.unmerge_cells(str(merged))


def _repair_comparative(wb,ticker,sector,industry,peer_count):
    if "Comparative Analysis" not in wb.sheetnames: return
    ws=wb["Comparative Analysis"]; label=industry if industry!="Unknown" else sector; ws["A1"]=f"Comparative Analysis — {ticker} vs {label} Peers"
    specs=[("Forward P/E","C","Lower","Lower is better"),("EV/Revenue","D","Lower","Lower is better"),("EV/EBITDA","E","Lower","Lower is better"),("Revenue Growth","F","Higher","Higher is better"),("Operating Margin","G","Higher","Higher is better"),("ROE","H","Higher","Higher is better")]
    last=4+max(1,peer_count)
    for r,(metric,col,direction,note) in enumerate(specs,4):
        ws.cell(r,1,metric); ws.cell(r,2,f"='Peer Comps'!{col}4"); ws.cell(r,3,f'=IFERROR(MEDIAN(\'Peer Comps\'!{col}5:{col}{last}),"")'); ws.cell(r,4,f'=IFERROR(B{r}/C{r}-1,"")'); ws.cell(r,5,direction)
        op="<" if direction=="Lower" else ">"; good="Attractive" if direction=="Lower" else "Better"; bad="Premium" if direction=="Lower" else "Worse"
        ws.cell(r,6,f'=IF(D{r}="","",IF(D{r}{op}0,"{good}","{bad}"))'); ws.cell(r,7,note)
        for c in (2,3): ws.cell(r,c).font=Font(color=LINK_GREEN); ws.cell(r,c).number_format=FMT_MULT if r<=6 else FMT_PCT
        ws.cell(r,4).number_format=FMT_PCT


def ensure_dynamic_peer_comps(wb,ticker,count=9):
    if "Peer Comps" not in wb.sheetnames: return []
    target,sector,industry,peers=select_dynamic_peers(wb,ticker,count); ws=wb["Peer Comps"]
    _unmerge_peer_area(ws)
    for row in ws.iter_rows(min_row=1,max_row=max(30,ws.max_row),min_col=1,max_col=18):
        for cell in row: cell.value=None
    for c in range(1,19): ws.cell(1,c).fill=_fill(NAVY); ws.cell(2,c).fill=_fill(NAVY)
    label=industry if industry!="Unknown" else sector; ws["A1"]=f"{label} — Expanded Peer Comps"; ws["A1"].font=Font(bold=True,color=WHITE,size=18)
    ws["A2"]=("Direct business-model peers are prioritized, then exact-industry and strategic same-sector comps. Missing public metrics are calculated from statements when possible. Industry market share and peer-set market-cap weight remain separate.")
    ws["A2"].font=Font(italic=True,color=GREY); ws["A2"].alignment=Alignment(wrap_text=True)
    headers=["Company","Ticker","Forward P/E","EV/Revenue","EV/EBITDA","Revenue Growth","Operating Margin","ROE","Sector","Industry","Discovery","Source URL","Industry Market Share %","Peer-Set Market Cap %","Market Share Basis / Source","Peer Type","Data Coverage %","Metric Source / Fallback Notes"]
    for c,v in enumerate(headers,1): ws.cell(3,c,v)
    _header(ws,3,1,18)
    if sector!="Unknown":
        all_rows=[(ticker,target,"Target classification")]+[(symbol,info,_peer_type(ticker,industry,symbol,info)) for symbol,info in peers]
        total_mc=sum(_num(info.get("marketCap"),0) or 0 for _,info,_ in all_rows)
        for r,(symbol,info,method) in enumerate(all_rows,4):
            row=_metric_row(symbol,info,sector,industry,method); share=market_share_record(symbol); row[12]=share.get("share") if share else None
            mc=_num(info.get("marketCap")); row[13]=(mc/total_mc) if mc and total_mc>0 else None
            row[14]=(f"{share.get('period','')} {share.get('basis','')} — {share.get('method','')}; {share.get('source','')}".strip() if share else "Not populated: no like-for-like industry market-share source mapped")
            for c,v in enumerate(row,1): ws.cell(r,c,v)
            for c in range(3,11): ws.cell(r,c).font=Font(color=INPUT_BLUE)
            ws.cell(r,12).font=Font(color=LINK_GREEN)
            for c in (3,4,5): ws.cell(r,c).number_format=FMT_MULT
            for c in (6,7,8,13,14,17): ws.cell(r,c).number_format=FMT_PCT
            for c in (15,16,18): ws.cell(r,c).alignment=Alignment(wrap_text=True,vertical="top")
        method_row=4+len(all_rows)+1
    else:
        ws["A4"]=ticker; ws["B4"]=ticker; ws["A5"]="REVIEW — target sector could not be resolved from live ticker metadata. No peer set was generated."; ws["A5"].fill=_fill(GOLD); ws["A5"].font=Font(color=INPUT_BLUE,bold=True); ws["A5"].alignment=Alignment(wrap_text=True); method_row=8
    if sector!="Unknown" and not peers:
        ws["A5"]="REVIEW — no validated peers were returned. Stale template peers were removed rather than reused."; ws["A5"].fill=_fill(GOLD); ws["A5"].font=Font(color=INPUT_BLUE,bold=True); ws["A5"].alignment=Alignment(wrap_text=True)
    ws.cell(method_row,1,"Method"); ws.cell(method_row,2,"Expanded to target + up to 9 peers. Discovery uses lightweight live classification; statement-derived fallback metrics are calculated only for the final selected peer set. Unsupported fields remain blank."); ws.cell(method_row,2).alignment=Alignment(wrap_text=True)
    widths={"A":31,"B":10,"C":13,"D":13,"E":13,"F":15,"G":16,"H":13,"I":18,"J":28,"K":25,"L":45,"M":20,"N":20,"O":62,"P":26,"Q":18,"R":68}
    for col,w in widths.items(): ws.column_dimensions[col].width=w
    ws.freeze_panes="A4"; _repair_comparative(wb,ticker,sector,industry,len(peers)); return peers
