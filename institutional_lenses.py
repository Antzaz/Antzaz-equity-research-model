"""Institutional-style comparison and safe workbook tab consolidation.

The institutional scores are transparent public-style lenses, not claims about proprietary
models, current holdings, target prices or actual investment decisions by the named firms.
Each lens translates publicly described investment priorities into explicit weights applied
to the workbook's existing model-quality dimensions.
"""

from __future__ import annotations

import copy
import math

from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; GREY="666666"; GOLD="FFF2CC"
PALE_GREEN="E2F0D9"; PALE_RED="FCE4D6"; LINK_GREEN="008000"
THIN=Side(style="thin",color="D9E1F2")
FMT_SCORE='0.0'; FMT_PCT='0.0%;[Red](0.0%);-'


def _fill(c): return PatternFill("solid",fgColor=c)

def _num(v,default=None):
    try:
        if isinstance(v,bool) or v in (None,""): return default
        x=float(v)
        return x if math.isfinite(x) else default
    except Exception:
        return default


def _label_value(ws,label,value_col=2):
    if ws is None: return None
    target=str(label).strip().lower()
    for r in range(1,ws.max_row+1):
        if str(ws.cell(r,1).value or "").strip().lower()==target:
            return _num(ws.cell(r,value_col).value)
    return None


def _scorecard_dimensions(wb):
    """Recover reusable 0-100 dimensions from the existing workbook."""
    dims={}
    if "Advanced Analytics" in wb.sheetnames:
        ws=wb["Advanced Analytics"]
        aliases={
            "Growth":"Growth","Profitability":"Profitability","FCF Quality":"FCF Quality",
            "Balance Sheet":"Balance Sheet","Absolute Valuation":"Absolute Valuation",
            "Relative Valuation":"Relative Valuation","Stress Robustness":"Stress Robustness",
        }
        for key,label in aliases.items():
            v=_label_value(ws,label,2)
            if v is not None: dims[key]=max(0,min(100,v))
    if "Leadership & Culture" in wb.sheetnames:
        v=_num(wb["Leadership & Culture"]["B11"].value)
        if v is not None: dims["Leadership"]=max(0,min(100,v))
    # Moat/competitive-position proxy. It intentionally combines observable public model
    # outputs rather than pretending the workbook knows a firm's proprietary moat score.
    parts=[dims.get("Profitability"),dims.get("Leadership")]
    if "Leadership & Culture" in wb.sheetnames:
        share=None
        ws=wb["Leadership & Culture"]
        for r in range(35,min(ws.max_row,42)+1):
            if isinstance(ws.cell(r,2).value,(int,float)):
                share=max(share or 0,float(ws.cell(r,2).value))
        if share is not None: parts.append(max(0,min(100,share*140)))
    parts=[x for x in parts if x is not None]
    if parts: dims["Moat / Position"]=sum(parts)/len(parts)
    return dims


INSTITUTIONAL_LENSES=[
    {
        "firm":"Berkshire Hathaway / Buffett-style",
        "focus":"Durable economics, owner-oriented cash generation, capable management, conservative financing and price vs intrinsic value.",
        "weights":{"Profitability":.20,"FCF Quality":.20,"Balance Sheet":.15,"Absolute Valuation":.15,"Leadership":.15,"Moat / Position":.10,"Stress Robustness":.05},
        "source":"https://www.berkshirehathaway.com/letters/letters.html",
    },
    {
        "firm":"Fundsmith",
        "focus":"High and sustainable returns on capital, strong free cash flow, durable growth, low leverage and avoiding overpayment.",
        "weights":{"Profitability":.25,"FCF Quality":.20,"Growth":.15,"Balance Sheet":.15,"Moat / Position":.10,"Absolute Valuation":.10,"Leadership":.05},
        "source":"https://www.fundsmith.co.uk/news/2015/2036-fidelity-three-steps-to-heaven/",
    },
    {
        "firm":"Dodge & Cox",
        "focus":"Long-term fundamental value, franchise and financial strength, competitive position, management quality and valuation discipline.",
        "weights":{"Absolute Valuation":.22,"Relative Valuation":.13,"Balance Sheet":.15,"Stress Robustness":.15,"Profitability":.12,"Growth":.08,"Leadership":.08,"Moat / Position":.07},
        "source":"https://www.dodgeandcox.com/individual-investor/us/en/our-approach/our-philosophy-and-process.html",
    },
    {
        "firm":"Davis Advisors",
        "focus":"Durable, financially strong businesses, owner earnings, proven management, sustainable advantages and a margin of safety.",
        "weights":{"FCF Quality":.20,"Balance Sheet":.15,"Leadership":.15,"Moat / Position":.15,"Profitability":.12,"Absolute Valuation":.13,"Growth":.05,"Stress Robustness":.05},
        "source":"https://davisfunds.com/about/discipline/",
    },
    {
        "firm":"Pershing Square",
        "focus":"Concentrated high-quality businesses with predictable recurring cash flow, strong competitive position, limited downside and identifiable value-creation levers.",
        "weights":{"FCF Quality":.20,"Profitability":.18,"Moat / Position":.18,"Growth":.12,"Leadership":.12,"Stress Robustness":.10,"Balance Sheet":.05,"Absolute Valuation":.05},
        "source":"https://pershingsquareholdings.com/about-us/",
    },
    {
        "firm":"Baillie Gifford — Long Term Growth",
        "focus":"Exceptional companies with transformational long-duration growth, large market opportunities, adaptability, competitive advantage and asymmetric upside.",
        "weights":{"Growth":.30,"Moat / Position":.20,"Leadership":.15,"Profitability":.12,"FCF Quality":.08,"Balance Sheet":.05,"Absolute Valuation":.05,"Stress Robustness":.05},
        "source":"https://www.bailliegifford.com/en/uk/institutional-investor/investment-strategies/long-term-global-growth/",
    },
    {
        "firm":"Akre Capital",
        "focus":"Three-legged stool: extraordinary business economics, shareholder-aligned management and long runways for disciplined reinvestment.",
        "weights":{"Profitability":.23,"Moat / Position":.20,"Leadership":.20,"Growth":.15,"FCF Quality":.12,"Balance Sheet":.05,"Absolute Valuation":.05},
        "source":"https://www.akrefund.com/investment-process/",
    },
    {
        "firm":"Polen Capital",
        "focus":"Quality growth, sustainable earnings, high returns on capital, durable competitive advantages, financial strength and proven management.",
        "weights":{"Growth":.25,"Profitability":.20,"Moat / Position":.18,"Balance Sheet":.12,"Leadership":.10,"FCF Quality":.10,"Stress Robustness":.03,"Absolute Valuation":.02},
        "source":"https://www.polencapital.com/strategies/global-growth",
    },
    {
        "firm":"Capital Group",
        "focus":"Deep fundamental research, long-term perspective, valuation awareness, multiple viewpoints and resilience across market cycles.",
        "weights":{"Growth":.14,"Profitability":.14,"FCF Quality":.14,"Balance Sheet":.14,"Absolute Valuation":.14,"Relative Valuation":.08,"Stress Robustness":.12,"Leadership":.05,"Moat / Position":.05},
        "source":"https://www.capitalgroup.com/about-us/capital-system.html",
    },
    {
        "firm":"T. Rowe Price",
        "focus":"Rigorous fundamental research, long-term growth potential, competitive environment, valuation, risk management and continuous thesis testing.",
        "weights":{"Growth":.20,"Profitability":.15,"FCF Quality":.12,"Absolute Valuation":.13,"Relative Valuation":.07,"Stress Robustness":.13,"Moat / Position":.10,"Leadership":.05,"Balance Sheet":.05},
        "source":"https://www.troweprice.com/en/us/investment-capabilities/active-investing-approach",
    },
]


def _weighted_score(dims,weights):
    total=0.0; used=0.0; detail=[]
    for key,w in weights.items():
        v=dims.get(key)
        if v is None: continue
        total+=v*w; used+=w; detail.append((key,v,w))
    return (total/used if used else None),used,detail


def _fit(score):
    if score is None: return "REVIEW"
    if score>=82: return "STRONG FIT"
    if score>=72: return "GOOD FIT"
    if score>=60: return "MIXED / INVESTIGATE"
    return "WEAK FIT"


def _strength_challenge(detail):
    if not detail: return "Insufficient evidence","Insufficient evidence"
    ranked=sorted(detail,key=lambda x:x[1],reverse=True)
    high=", ".join(f"{k} ({v:.0f})" for k,v,_ in ranked[:2])
    low=", ".join(f"{k} ({v:.0f})" for k,v,_ in sorted(detail,key=lambda x:x[1])[:2])
    return high,low


def ensure_institutional_lens_comparison(wb,ticker):
    if "Institutional Comparison" in wb.sheetnames: wb.remove(wb["Institutional Comparison"])
    ws=wb.create_sheet("Institutional Comparison"); ws.sheet_view.showGridLines=False
    for c in range(1,10): ws.cell(1,c).fill=_fill(NAVY); ws.cell(2,c).fill=_fill(NAVY)
    ws["A1"]=f"{ticker} — 10 Institutional Investment Lenses"; ws["A1"].font=Font(bold=True,color=WHITE,size=18)
    ws["A3"]=("These are transparent public-style comparison lenses built from the firms' published investment principles. "
              "They are not proprietary firm models, current holdings, target prices, recommendations, or claims that the firms would buy the company.")
    ws.merge_cells("A3:I3"); ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True,vertical="top")

    dims=_scorecard_dimensions(wb)
    ws["A5"]="Reusable Company Dimension Scores"
    for c in range(1,10): ws.cell(5,c).fill=_fill(NAVY); ws.cell(5,c).font=Font(bold=True,color=WHITE)
    headers=list(dims.keys())
    for c,name in enumerate(headers,1): ws.cell(6,c,name); ws.cell(6,c).fill=_fill(BLUE); ws.cell(6,c).font=Font(bold=True,color=WHITE)
    for c,name in enumerate(headers,1): ws.cell(7,c,dims[name]); ws.cell(7,c).number_format=FMT_SCORE

    start=10
    cols=["Institution / Style","Public Investment Lens","Fit Score / 100","Fit","Data Coverage","What the Lens Likes","What It Would Challenge","Public Source","Important Caveat"]
    for c,v in enumerate(cols,1): ws.cell(start,c,v)
    for c in range(1,len(cols)+1):
        ws.cell(start,c).fill=_fill(BLUE); ws.cell(start,c).font=Font(bold=True,color=WHITE); ws.cell(start,c).alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); ws.cell(start,c).border=Border(bottom=THIN)
    for r,lens in enumerate(INSTITUTIONAL_LENSES,start+1):
        score,coverage,detail=_weighted_score(dims,lens["weights"]); strengths,challenges=_strength_challenge(detail)
        values=[lens["firm"],lens["focus"],score,_fit(score),coverage,strengths,challenges,lens["source"],"Public-style approximation only; validate sector-specific qualitative factors before drawing a conclusion."]
        for c,v in enumerate(values,1): ws.cell(r,c,v)
        ws.cell(r,3).number_format=FMT_SCORE; ws.cell(r,5).number_format=FMT_PCT
        ws.cell(r,8).hyperlink=lens["source"]; ws.cell(r,8).font=Font(color=LINK_GREEN,underline="single")
        for c in (1,2,4,6,7,8,9): ws.cell(r,c).alignment=Alignment(wrap_text=True,vertical="top")
        if score is not None:
            ws.cell(r,4).fill=_fill(PALE_GREEN if score>=72 else (GOLD if score>=60 else PALE_RED)); ws.cell(r,4).font=Font(bold=True)

    end=start+len(INSTITUTIONAL_LENSES)
    ws.cell(end+2,1,"How to use this")
    for c in range(1,10): ws.cell(end+2,c).fill=_fill(NAVY); ws.cell(end+2,c).font=Font(bold=True,color=WHITE)
    ws.cell(end+3,1,"1") ; ws.cell(end+3,2,"Use high-scoring lenses to identify why the company may appeal to a specific institutional style.")
    ws.cell(end+4,1,"2") ; ws.cell(end+4,2,"Use the challenge column as a diligence queue; it is often more useful than the fit score.")
    ws.cell(end+5,1,"3") ; ws.cell(end+5,2,"Do not infer that a firm owns, endorses or values the company from this sheet.")
    for r in range(end+3,end+6): ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=9); ws.cell(r,2).alignment=Alignment(wrap_text=True)

    widths={"A":29,"B":49,"C":16,"D":19,"E":16,"F":35,"G":35,"H":55,"I":46}
    for col,w in widths.items(): ws.column_dimensions[col].width=w
    ws.freeze_panes="A11"
    return {"dimensions":dims,"institutions":len(INSTITUTIONAL_LENSES)}


def _sheet_referenced(wb,name):
    needles=(f"'{name}'!",f"{name}!")
    for ws in wb.worksheets:
        if ws.title==name: continue
        for row in ws.iter_rows():
            for cell in row:
                v=cell.value
                if isinstance(v,str) and v.startswith("=") and any(n in v for n in needles):
                    return True
    return False


def _copy_block(source,target,start_row,max_col=None):
    max_col=max_col or source.max_column
    for r in range(1,source.max_row+1):
        target.row_dimensions[start_row+r-1].height=source.row_dimensions[r].height
        for c in range(1,max_col+1):
            s=source.cell(r,c); t=target.cell(start_row+r-1,c)
            t.value=s.value
            if s.has_style:
                t._style=copy.copy(s._style)
                t.font=copy.copy(s.font); t.fill=copy.copy(s.fill); t.border=copy.copy(s.border)
                t.alignment=copy.copy(s.alignment); t.protection=copy.copy(s.protection); t.number_format=s.number_format
            if s.hyperlink: t._hyperlink=copy.copy(s.hyperlink)
            if s.comment: t.comment=copy.copy(s.comment)
    for merged in source.merged_cells.ranges:
        target.merge_cells(start_row=start_row+merged.min_row-1,start_column=merged.min_col,end_row=start_row+merged.max_row-1,end_column=merged.max_col)


def consolidate_research_workbench(wb):
    """Reduce tab count by merging only low-risk overlapping research/admin sheets."""
    removed=[]
    if "Research Notes" in wb.sheetnames:
        notes=wb["Research Notes"]
        for name in ("Research Checklist","Research Journal"):
            if name not in wb.sheetnames or _sheet_referenced(wb,name): continue
            src=wb[name]; start=max(notes.max_row+3,82)
            _copy_block(src,notes,start)
            wb.remove(src); removed.append(name)
        notes.title="Research Workbench"
    elif "Research Notes" not in wb.sheetnames and "Research Workbench" in wb.sheetnames:
        notes=wb["Research Workbench"]

    if "Data Dictionary" in wb.sheetnames and "Data Quality" in wb.sheetnames and not _sheet_referenced(wb,"Data Dictionary"):
        dq=wb["Data Quality"]; src=wb["Data Dictionary"]; start=dq.max_row+3
        _copy_block(src,dq,start,max_col=min(4,src.max_column)); wb.remove(src); removed.append("Data Dictionary")

    return removed
