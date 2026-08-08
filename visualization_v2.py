"""Quality-control and visualization layer for the equity research workbook.

Called last by update_model.py. It repairs known source/cache issues, replaces the
Alphabet segment fallback with filing-backed data, and rebuilds the visual dashboard
from a clean sheet so stale chart anchors/merges cannot survive between runs.
"""

import statistics
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, DataBarRule
from openpyxl.utils import get_column_letter

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; LIGHT="F5F9FC"; PALE_BLUE="D9EAF7"; PALE_GREEN="E2F0D9"; PALE_RED="FCE4D6"; GOLD="FFF2CC"; INPUT_BLUE="0000FF"; LINK_GREEN="008000"; GREY="666666"
FMT_PCT='0.0%;[Red](0.0%);-'; FMT_PRICE='$#,##0.00;[Red]($#,##0.00);-'; FMT_BN='#,##0.0;[Red](#,##0.0);-'; FMT_MULT='0.0x;[Red](0.0x);-'


def _fill(c): return PatternFill("solid",fgColor=c)
def _merge(ws,rng):
    if rng not in {str(x) for x in ws.merged_cells.ranges}: ws.merge_cells(rng)
def _section(ws,rng,title):
    _merge(ws,rng); c=ws[rng.split(":")[0]]; c.value=title; c.fill=_fill(NAVY); c.font=Font(bold=True,color=WHITE,size=11)
def _header(ws,row,start,end):
    for c in range(start,end+1):
        x=ws.cell(row,c); x.fill=_fill(BLUE); x.font=Font(bold=True,color=WHITE); x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def _card(ws,tr,vr,title,value,fmt,formula=False):
    _merge(ws,tr); _merge(ws,vr); t=ws[tr.split(":")[0]]; v=ws[vr.split(":")[0]]
    t.value=title; t.fill=_fill(BLUE); t.font=Font(bold=True,color=WHITE); t.alignment=Alignment(horizontal="center")
    v.value=value; v.fill=_fill(LIGHT); v.font=Font(bold=True,color=LINK_GREEN if formula else INPUT_BLUE,size=15); v.alignment=Alignment(horizontal="center",vertical="center"); v.number_format=fmt


def _repair_alphabet(wb,ticker):
    if ticker.upper() not in {"GOOGL","GOOG"}: return
    if "Historical Financials" in wb.sheetnames:
        h=wb["Historical Financials"]
        years=[h.cell(3,c).value for c in range(2,8)]
        if years==[2020,2021,2022,2023,2024,2025]:
            eps=[2.93,5.61,4.56,5.80,8.04,10.81]; da=[12.9,11.6,15.3,11.946,15.311,21.136]
            for c,v in enumerate(eps,2): h.cell(12,c,v); h.cell(12,c).number_format=FMT_PRICE; ni=h.cell(11,c).value; h.cell(13,c,(ni/v) if isinstance(ni,(int,float)) else None)
            for c,v in enumerate(da,2): h.cell(18,c,v); h.cell(18,c).number_format=FMT_BN
    if "Financial Statements" in wb.sheetnames:
        f=wb["Financial Statements"]
        labels={f.cell(r,1).value:r for r in range(1,f.max_row+1)}
        eps_r=labels.get("Diluted EPS"); op_r=labels.get("Operating Income"); pretax_r=labels.get("Pre-Tax Income"); other_r=labels.get("Other Income / (Expense), Net"); da_r=labels.get("Depreciation & Amortization")
        if eps_r:
            for c,v in enumerate([2.93,5.61,4.56,5.80,8.04,10.81],2): f.cell(eps_r,c,v); f.cell(eps_r,c).number_format=FMT_PRICE
        if op_r and pretax_r and other_r:
            for c in range(2,8): f.cell(other_r,c,f"={get_column_letter(c)}{pretax_r}-{get_column_letter(c)}{op_r}"); f.cell(other_r,c).number_format=FMT_BN
        if da_r:
            for c,v in enumerate([12.9,11.6,15.3,11.946,15.311,21.136],2): f.cell(da_r,c,v); f.cell(da_r,c).number_format=FMT_BN
        ppe_r=labels.get("Property & Equipment, Net")
        if ppe_r:
            for c in range(2,7):
                if f.cell(23,c).value==2025 and f.cell(ppe_r,c).value is None: f.cell(ppe_r,c,246.597); f.cell(ppe_r,c).number_format=FMT_BN


def _alphabet_segment_sheet(wb,ticker):
    if "Segment Analysis" in wb.sheetnames: wb.remove(wb["Segment Analysis"])
    ws=wb.create_sheet("Segment Analysis"); ws.sheet_view.showGridLines=False
    _merge(ws,"A1:N2"); ws["A1"]=f"{ticker} — Business & Segment Analysis"; ws["A1"].fill=_fill(NAVY); ws["A1"].font=Font(bold=True,color=WHITE,size=18); ws["A1"].alignment=Alignment(vertical="center")
    _merge(ws,"A3:N3"); ws["A3"]="Reported operating segments, business-line revenue, segment profitability, geography, costs and key operating drivers. USD billions."; ws["A3"].font=Font(italic=True,color=GREY)
    cards=[("A5:C5","A6:C8","2025 Revenue",402.836,FMT_BN),("D5:F5","D6:F8","Google Services Mix",342.721/402.836,FMT_PCT),("G5:I5","G6:I8","Google Cloud Growth",58.705/43.229-1,FMT_PCT),("J5:L5","J6:L8","Cloud Op. Margin",13.910/58.705,FMT_PCT),("M5:N5","M6:N8","Cloud-Led Backlog",242.8,FMT_BN)]
    for x in cards: _card(ws,*x)

    _section(ws,"A10:N10","Reported Operating Segments"); heads=["Segment","2023 Revenue","2024 Revenue","2025 Revenue","2025 Growth","2023–25 CAGR","2023 Op. Income","2024 Op. Income","2025 Op. Income","2023 Margin","2024 Margin","2025 Margin","Margin Δ vs 2024","2025 Revenue Mix"]
    for c,x in enumerate(heads,1): ws.cell(11,c,x)
    _header(ws,11,1,14)
    segs=[("Google Services",272.543,304.930,342.721,95.858,121.263,139.404),("Google Cloud",33.088,43.229,58.705,1.716,6.112,13.910),("Other Bets",1.527,1.648,1.537,-4.095,-4.444,-7.515),("Alphabet-level activities",None,None,None,-9.186,-10.541,-16.760)]
    for r,(n,a,b,c,o1,o2,o3) in enumerate(segs,12):
        vals=[n,a,b,c,(c/b-1) if c and b else None,((c/a)**.5-1) if c and a else None,o1,o2,o3,(o1/a) if a else None,(o2/b) if b else None,(o3/c) if c else None,((o3/c)-(o2/b)) if c and b else None,(c/402.836) if c else None]
        for col,v in enumerate(vals,1): ws.cell(r,col,v)
        for col in range(2,5): ws.cell(r,col).number_format=FMT_BN
        for col in range(5,7): ws.cell(r,col).number_format=FMT_PCT
        for col in range(7,10): ws.cell(r,col).number_format=FMT_BN
        for col in range(10,15): ws.cell(r,col).number_format=FMT_PCT
    ws.conditional_formatting.add("E12:F15",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="percentile",mid_value=50,mid_color="FFEB84",end_type="max",end_color="63BE7B")); ws.conditional_formatting.add("L12:M15",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="percentile",mid_value=50,mid_color="FFEB84",end_type="max",end_color="63BE7B"))

    _section(ws,"A18:N18","Revenue by Business Line"); heads=["Business Line","2023","2024","2025","2025 Growth","2023–25 CAGR","2025 Mix","2025 Δ Revenue","Share of 2025 Revenue Growth","Trend / Investment Read-Through"]
    for c,x in enumerate(heads,1): ws.cell(19,c,x)
    _header(ws,19,1,10); growth=402.836-350.018
    lines=[("Google Search & other",175.033,198.084,224.532,"Core earnings engine; strong double-digit growth"),("YouTube ads",31.510,36.147,40.367,"Growing ad platform; smaller than Search"),("Google Network",31.312,30.359,29.792,"Structurally declining; lower-quality mix"),("Subscriptions, Platforms & Devices",34.688,40.340,48.030,"Fast growth from subscriptions / Google One / YouTube"),("Google Cloud",33.088,43.229,58.705,"Fastest major growth; margin expansion creates operating leverage"),("Other Bets",1.527,1.648,1.537,"Small revenue base; losses dominated by long-duration bets")]
    for r,(n,a,b,c,note) in enumerate(lines,20):
        vals=[n,a,b,c,c/b-1,(c/a)**.5-1,c/402.836,c-b,(c-b)/growth,note]
        for col,v in enumerate(vals,1): ws.cell(r,col,v)
        for col in range(2,5): ws.cell(r,col).number_format=FMT_BN
        for col in range(5,8): ws.cell(r,col).number_format=FMT_PCT
        ws.cell(r,8).number_format=FMT_BN; ws.cell(r,9).number_format=FMT_PCT
    ws.conditional_formatting.add("E20:F25",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="percentile",mid_value=50,mid_color="FFEB84",end_type="max",end_color="63BE7B")); ws.conditional_formatting.add("I20:I25",DataBarRule(start_type="min",end_type="max",color="9E4C9E"))

    _section(ws,"A28:H28","Segment Cost Structure"); heads=["Segment / Cost","2023","2024","2025","2025 Growth","% of Segment Revenue","Read-Through"]
    for c,x in enumerate(heads,1): ws.cell(29,c,x)
    _header(ws,29,1,7); costs=[("Services — employee compensation",46.224,44.560,45.124,"Compensation stable despite revenue growth",342.721),("Services — other costs & expenses",130.461,139.107,158.193,"Includes TAC, content, legal, infrastructure and device costs",342.721),("Cloud — employee compensation",19.054,20.519,22.078,"Revenue growth far exceeds employee-cost growth",58.705),("Cloud — other costs & expenses",12.318,16.598,22.717,"Infrastructure usage rises with scale",58.705)]
    for r,(n,a,b,c,note,den) in enumerate(costs,30):
        for col,v in enumerate([n,a,b,c,c/b-1,c/den,note],1): ws.cell(r,col,v)
        for col in range(2,5): ws.cell(r,col).number_format=FMT_BN
        for col in range(5,7): ws.cell(r,col).number_format=FMT_PCT

    _section(ws,"A36:H36","Revenue by Geography"); heads=["Region","2023 Revenue","2024 Revenue","2025 Revenue","2025 Growth","2025 Mix","2023 Mix","Trend"]
    for c,x in enumerate(heads,1): ws.cell(37,c,x)
    _header(ws,37,1,8); geo=[("United States",146.286,170.447,194.229,"Largest market; mix broadly stable"),("EMEA",91.038,102.127,117.152,"29% of revenue"),("APAC",51.514,56.815,67.680,"Mix increased to ~17%"),("Other Americas",18.320,20.418,23.902,"~6% of revenue")]
    for r,(n,a,b,c,note) in enumerate(geo,38):
        for col,v in enumerate([n,a,b,c,c/b-1,c/402.836,a/307.394,note],1): ws.cell(r,col,v)
        for col in range(2,5): ws.cell(r,col).number_format=FMT_BN
        for col in range(5,8): ws.cell(r,col).number_format=FMT_PCT

    _section(ws,"A44:N44","Key Segment Drivers & Investment Questions"); rows=[["Driver","Latest Evidence","Investment Relevance"],["Search / advertising","Google advertising revenue $294.7bn; Search & other $224.5bn","Search remains the dominant cash engine; assess AI Search monetization and query economics."],["TAC / distribution economics","2025 TAC $59.9bn; TAC rate fell from 20.7% to 20.3%","Lower TAC mix is supportive, but distribution agreements remain strategically important."],["Google Cloud","2025 revenue $58.7bn (+35.8%); operating margin 23.7%","Cloud is the strongest incremental growth/profit engine and key AI monetization channel."],["Cloud-led backlog","RPO / revenue backlog $242.8bn at 2025 year-end; just over 50% expected within 24 months","Provides forward visibility, although backlog timing depends on utilization and delivery."],["Other Bets / Waymo","2025 Other Bets operating loss $7.5bn; includes a $2.1bn Waymo compensation charge","Large optionality but meaningful current drag; separate strategic value from reported losses."],["Alphabet-level AI costs","2025 unallocated Alphabet-level operating costs $16.8bn","Shared AI R&D can depress consolidated margins while benefiting multiple segments."],["Geographic concentration","48% US / 29% EMEA / 17% APAC / 6% Other Americas","Diversified globally, but regulatory and FX exposure matter."]]
    for r,row in enumerate(rows,45):
        for c,v in enumerate(row,1): ws.cell(r,c,v)
    _header(ws,45,1,3)
    for r in range(46,53): ws.row_dimensions[r].height=36; ws.cell(r,2).alignment=Alignment(wrap_text=True,vertical="top"); ws.cell(r,3).alignment=Alignment(wrap_text=True,vertical="top")

    ws.cell(10,16,"Business Line"); ws.cell(10,17,"2025 Revenue")
    for r,(n,v) in enumerate([("Search & other",224.532),("Cloud",58.705),("Subscriptions / Platforms / Devices",48.030),("YouTube ads",40.367),("Network",29.792),("Other Bets",1.537)],11): ws.cell(r,16,n); ws.cell(r,17,v)
    ch=BarChart(); ch.type="bar"; ch.style=10; ch.title="2025 Revenue by Business Line"; ch.height=7; ch.width=12.5; ch.legend=None; ch.add_data(Reference(ws,min_col=17,min_row=10,max_row=16),titles_from_data=True); ch.set_categories(Reference(ws,min_col=16,min_row=11,max_row=16)); ws.add_chart(ch,"I28")
    ws.cell(19,16,"Segment"); ws.cell(19,17,"2023 Margin"); ws.cell(19,18,"2024 Margin"); ws.cell(19,19,"2025 Margin")
    for r,row in enumerate([("Google Services",95.858/272.543,121.263/304.930,139.404/342.721),("Google Cloud",1.716/33.088,6.112/43.229,13.910/58.705)],20):
        for c,v in enumerate(row,16): ws.cell(r,c,v)
    ch=LineChart(); ch.style=10; ch.title="Services & Cloud Operating Margin"; ch.height=7; ch.width=12.5; ch.add_data(Reference(ws,min_col=17,max_col=19,min_row=19,max_row=21),titles_from_data=True); ch.set_categories(Reference(ws,min_col=16,min_row=20,max_row=21)); ch.y_axis.numFmt="0%"; ch.legend.position="b"; ws.add_chart(ch,"I40")

    _merge(ws,"A55:N58"); ws["A55"]="Primary SEC source: https://www.sec.gov/Archives/edgar/data/1652044/000165204426000018/goog-20251231.htm\nSegment table: https://www.sec.gov/Archives/edgar/data/1652044/000165204426000018/R42.htm"; ws["A55"].fill=_fill(PALE_BLUE); ws["A55"].alignment=Alignment(wrap_text=True)
    ws.column_dimensions["A"].width=31
    for c in range(2,10): ws.column_dimensions[get_column_letter(c)].width=14
    ws.column_dimensions["J"].width=38
    for c in range(11,15): ws.column_dimensions[get_column_letter(c)].width=14
    for c in range(16,20): ws.column_dimensions[get_column_letter(c)].hidden=True
    return ws


def _repair_advanced(wb,ticker):
    if "Advanced Analytics" not in wb.sheetnames: return
    ws=wb["Advanced Analytics"]
    if ticker.upper() in {"GOOGL","GOOG"}:
        eps={2020:2.93,2021:5.61,2022:4.56,2023:5.80,2024:8.04,2025:10.81}
        for r in range(7,13):
            y=ws.cell(r,1).value
            if y in eps: ws.cell(r,3,eps[y]); ws.cell(r,3).number_format=FMT_PRICE; ws.cell(r,4,f'=IFERROR(B{r}/C{r},"")'); ws.cell(r,4).number_format=FMT_MULT
        vals=[ws.cell(r,2).value/eps.get(ws.cell(r,1).value,1) for r in range(7,13) if isinstance(ws.cell(r,2).value,(int,float)) and ws.cell(r,1).value in eps]
        if vals: ws["G7"]=min(vals); ws["G8"]=statistics.median(vals); ws["G9"]=max(vals)
    for r in range(7,15): ws.cell(r,9).number_format="yyyy-mm-dd"
    h=wb["Historical Financials"]; rev=h["G4"].value; op=h["G9"].value; ocf=h["G14"].value; cap=h["G15"].value
    if all(isinstance(x,(int,float)) for x in [rev,op,ocf,cap]) and rev:
        r0=h["B4"].value; cagr=(rev/r0)**.2-1 if isinstance(r0,(int,float)) and r0 else 0; opm=op/rev; fcfm=(ocf-cap)/rev; price=wb["Company Data"]["B8"].value or 0; base=wb["Three-Case Scenarios"]["C39"].value or 0; severe=wb["Three-Case Scenarios"]["G58"].value or 0; fpe=wb["Company Data"]["B15"].value or 0; pes=[wb["Peer Comps"].cell(r,3).value for r in range(5,10)]; pes=[x for x in pes if isinstance(x,(int,float))]; med=statistics.median(pes) if pes else fpe or 1
        scores=[min(100,max(0,cagr/.20*100)),min(100,max(0,opm/.40*100)),min(100,max(0,fcfm/.25*100)),85,min(100,max(0,base/price*100)) if price else 0,min(100,max(0,med/fpe*70)) if fpe else 50,min(100,max(0,severe/price*100)) if price else 0]
        for r,v in zip(range(43,50),scores): ws.cell(r,2,v)
        ws["F42"]=sum(scores)/len(scores)


def ensure_visual_dashboard(wb,ticker=None):
    ticker=(ticker or "").upper(); _repair_alphabet(wb,ticker)
    if ticker in {"GOOGL","GOOG"}: _alphabet_segment_sheet(wb,ticker)
    _repair_advanced(wb,ticker)
    required={"Company Data","Historical Financials","Three-Case Scenarios","DCF","Peer Comps"}
    if not required.issubset(wb.sheetnames): return None
    if "Visual Dashboard" in wb.sheetnames: wb.remove(wb["Visual Dashboard"])
    ws=wb.create_sheet("Visual Dashboard"); ws.sheet_view.showGridLines=False
    try: wb.active=wb.index(ws)
    except Exception: pass
    for c in range(1,17): ws.column_dimensions[get_column_letter(c)].width=11
    ws.column_dimensions["A"].width=23; ws.column_dimensions["I"].width=24
    for c in range(24,31): ws.column_dimensions[get_column_letter(c)].hidden=True
    _merge(ws,"A1:P2"); ws["A1"]="Equity Research — Visual Dashboard"; ws["A1"].fill=_fill(NAVY); ws["A1"].font=Font(bold=True,color=WHITE,size=18); ws["A1"].alignment=Alignment(vertical="center")
    _merge(ws,"A3:P3"); ws["A3"]='="Automated visual summary for "&\'Company Data\'!B4&" | valuation, risk, peers and business mix"'; ws["A3"].font=Font(italic=True,color=GREY)
    rich="Segment Analysis" in wb.sheetnames and wb["Segment Analysis"]["A10"].value=="Reported Operating Segments"
    cards=[("A5:D5","A6:D8","Current Price","='Company Data'!B8",FMT_PRICE,True),("E5:H5","E6:H8","Base Intrinsic Value","='Three-Case Scenarios'!C39",FMT_PRICE,True),("I5:L5","I6:L8","Probability-Weighted Value","='Three-Case Scenarios'!E39",FMT_PRICE,True),("M5:P5","M6:P8","Base Upside / (Downside)","='Three-Case Scenarios'!C40",FMT_PCT,True),("A10:D10","A11:D13","Revenue CAGR (2020–25)","=Dashboard!B8",FMT_PCT,True),("E10:H10","E11:H13","2025 FCF Margin","=Dashboard!B11",FMT_PCT,True)]
    cards += [("I10:L10","I11:L13","Cloud Growth","='Segment Analysis'!G6",FMT_PCT,True),("M10:P10","M11:P13","Cloud Op. Margin","='Segment Analysis'!J6",FMT_PCT,True)] if rich else [("I10:L10","I11:L13","Severe Bear Value / Share","='Three-Case Scenarios'!B63",FMT_PRICE,True),("M10:P10","M11:P13","Robustness Status","='Three-Case Scenarios'!B66","@",True)]
    for x in cards: _card(ws,*x)
    ws.conditional_formatting.add("M6:P8",CellIsRule(operator="lessThan",formula=["0"],fill=_fill(PALE_RED))); ws.conditional_formatting.add("M6:P8",CellIsRule(operator="greaterThanOrEqual",formula=["0"],fill=_fill(PALE_GREEN)))

    h=wb["Historical Financials"]
    ws["X2"],ws["Y2"],ws["Z2"]="Year","Revenue","FCF"
    for r,c in enumerate(range(2,8),3):
        rev=h.cell(4,c).value; ocf=h.cell(14,c).value; cap=h.cell(15,c).value; ws.cell(r,24,h.cell(3,c).value); ws.cell(r,25,rev); ws.cell(r,26,(ocf-cap) if isinstance(ocf,(int,float)) and isinstance(cap,(int,float)) else None)
    ws["X11"],ws["Y11"],ws["Z11"]="Year","Operating Margin","FCF Margin"
    for r,c in enumerate(range(2,8),12):
        rev=h.cell(4,c).value; op=h.cell(9,c).value; ocf=h.cell(14,c).value; cap=h.cell(15,c).value; ws.cell(r,24,h.cell(3,c).value); ws.cell(r,25,(op/rev) if isinstance(op,(int,float)) and isinstance(rev,(int,float)) and rev else None); ws.cell(r,26,((ocf-cap)/rev) if all(isinstance(x,(int,float)) for x in [ocf,cap,rev]) and rev else None)
    ws["X20"],ws["Y20"]="Scenario","Value / Share"
    for r,(lab,ref) in enumerate([("Bear","B39"),("Base","C39"),("Bull","D39"),("Prob.-weighted","E39")],21): ws.cell(r,24,lab); ws.cell(r,25,f"='Three-Case Scenarios'!{ref}")
    ws["X25"]="Current Price"; ws["Y25"]="='Company Data'!B8"; ws["X28"],ws["Y28"]="Stress Test","Value / Share"
    for r,src in enumerate(range(48,59),29): ws.cell(r,24,f"='Three-Case Scenarios'!A{src}"); ws.cell(r,25,f"='Three-Case Scenarios'!G{src}")
    ws["X42"],ws["Y42"],ws["Z42"]="Ticker","Forward P/E","EV/EBITDA"
    for r,src in enumerate(range(4,10),43): ws.cell(r,24,f"='Peer Comps'!B{src}"); ws.cell(r,25,f"='Peer Comps'!C{src}"); ws.cell(r,26,f"='Peer Comps'!E{src}")

    def line_chart(rng,title,anchor,pct=False):
        ch=LineChart(); ch.style=10; ch.title=title; ch.height=7; ch.width=13; ch.add_data(Reference(ws,min_col=25,max_col=26,min_row=rng[0],max_row=rng[1]),titles_from_data=True); ch.set_categories(Reference(ws,min_col=24,min_row=rng[0]+1,max_row=rng[1])); ch.legend.position="b"; ch.y_axis.numFmt="0%" if pct else '#,##0'; ws.add_chart(ch,anchor)
    _section(ws,"A15:H15","Historical Financial Performance"); line_chart((2,8),"Revenue & Free Cash Flow","A17")
    _section(ws,"I15:P15","Scenario Valuation"); ch=BarChart(); ch.type="col"; ch.style=10; ch.title="Scenario Value vs Current Price"; ch.height=7; ch.width=13; ch.legend=None; ch.add_data(Reference(ws,min_col=25,min_row=20,max_row=25),titles_from_data=True); ch.set_categories(Reference(ws,min_col=24,min_row=21,max_row=25)); ch.y_axis.numFmt="$0"; ws.add_chart(ch,"I17")
    _section(ws,"A32:H32","Margins"); line_chart((11,17),"Operating & FCF Margins","A34",True)
    _section(ws,"I32:P32","Stress Testing"); ch=BarChart(); ch.type="bar"; ch.style=10; ch.title="Stress-Test Intrinsic Value / Share"; ch.height=7; ch.width=13; ch.legend=None; ch.add_data(Reference(ws,min_col=25,min_row=28,max_row=39),titles_from_data=True); ch.set_categories(Reference(ws,min_col=24,min_row=29,max_row=39)); ch.x_axis.numFmt="$0"; ws.add_chart(ch,"I34")
    _section(ws,"A50:H50","Peer Valuation"); ch=BarChart(); ch.type="col"; ch.style=10; ch.title="Peer Valuation Multiples"; ch.height=7; ch.width=13; ch.add_data(Reference(ws,min_col=25,max_col=26,min_row=42,max_row=48),titles_from_data=True); ch.set_categories(Reference(ws,min_col=24,min_row=43,max_row=48)); ch.y_axis.numFmt="0.0x"; ch.legend.position="b"; ws.add_chart(ch,"A52")
    _section(ws,"I50:P50","DCF Sensitivity"); ws["I52"]="WACC \\ TGR"
    for c,src in zip(range(10,15),range(2,7)): ws.cell(52,c,f"=DCF!{get_column_letter(src)}21")
    for r,sr in zip(range(53,58),range(22,27)):
        ws.cell(r,9,f"=DCF!A{sr}")
        for c,sc_c in zip(range(10,15),range(2,7)): ws.cell(r,c,f"=DCF!{get_column_letter(sc_c)}{sr}"); ws.cell(r,c).number_format=FMT_PRICE
    _header(ws,52,9,14); ws.conditional_formatting.add("J53:N57",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="percentile",mid_value=50,mid_color="FFEB84",end_type="max",end_color="63BE7B")); _merge(ws,"I59:P61"); ws["I59"]="Green = higher modeled value; red = lower. Use this to see how much the thesis depends on WACC and terminal growth."; ws["I59"].fill=_fill(LIGHT); ws["I59"].font=Font(italic=True,color=GREY); ws["I59"].alignment=Alignment(wrap_text=True)

    _section(ws,"A69:H69","2025 Business Mix"); _section(ws,"I69:P69","Segment Economics")
    if rich:
        ws["X51"],ws["Y51"]="Business Line","2025 Revenue"
        for r,src in enumerate(range(20,26),52): ws.cell(r,24,f"='Segment Analysis'!A{src}"); ws.cell(r,25,f"='Segment Analysis'!D{src}")
        ch=BarChart(); ch.type="bar"; ch.style=10; ch.title="2025 Revenue by Business Line"; ch.height=7; ch.width=13; ch.legend=None; ch.add_data(Reference(ws,min_col=25,min_row=51,max_row=57),titles_from_data=True); ch.set_categories(Reference(ws,min_col=24,min_row=52,max_row=57)); ws.add_chart(ch,"A71")
        heads=["Segment","2025 Revenue","2025 Growth","2025 Op. Income","2025 Margin","2025 Mix","3Y CAGR","Margin Δ"]
        for c,x in enumerate(heads,9): ws.cell(71,c,x)
        _header(ws,71,9,16)
        for r,src in enumerate(range(12,16),72):
            for c,src_c in enumerate([1,4,5,9,12,14,6,13],9): ws.cell(r,c,f"='Segment Analysis'!{get_column_letter(src_c)}{src}")
        for r in range(72,76): ws.cell(r,10).number_format=FMT_BN; ws.cell(r,11).number_format=FMT_PCT; ws.cell(r,12).number_format=FMT_BN
        for r in range(72,76):
            for c in range(13,17): ws.cell(r,c).number_format=FMT_PCT
        ws.conditional_formatting.add("K72:P75",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="percentile",mid_value=50,mid_color="FFEB84",end_type="max",end_color="63BE7B"))

    _section(ws,"A88:H88","Advanced Valuation"); _section(ws,"I88:P88","Key Investment Read-Through")
    if "Advanced Analytics" in wb.sheetnames:
        rows=[("Reverse DCF implied FCF CAGR","='Advanced Analytics'!B38",FMT_PCT,"Growth required to justify current price"),("Monte Carlo P10","='Advanced Analytics'!J33",FMT_PRICE,"10th percentile valuation"),("Monte Carlo Median","='Advanced Analytics'!J35",FMT_PRICE,"Median simulated valuation"),("Probability > Current Price","='Advanced Analytics'!J38",FMT_PCT,"Share of simulations above market"),("Composite Investment Score","='Advanced Analytics'!F42","0.0","Diagnostic score, not recommendation"),("Forward P/E","='Company Data'!B15",FMT_MULT,"Current forward valuation multiple")]
        for c,x in enumerate(["Metric","Result","Interpretation","Status"],1): ws.cell(89,c,x)
        _header(ws,89,1,4)
        for r,(lab,form,fmt,note) in enumerate(rows,90): ws.cell(r,1,lab); ws.cell(r,2,form); ws.cell(r,2).number_format=fmt; ws.cell(r,3,note); ws.cell(r,4,"Review")
    _merge(ws,"I89:P95"); ws["I89"]="Alphabet remains dominated by Search, but Cloud is the most important incremental growth and margin-expansion engine. Monitor Search/AI monetization, Cloud profitability, capex normalization, backlog conversion and shared AI R&D costs. The valuation should be judged against reverse-DCF expectations and downside stress tests, not the Base DCF alone." if ticker in {"GOOGL","GOOG"} else "Use Financial Statements, Segment Analysis and Advanced Analytics together with this dashboard to identify the operating variables that matter most to valuation."; ws["I89"].fill=_fill(LIGHT); ws["I89"].alignment=Alignment(wrap_text=True,vertical="top")
    return ws
