from __future__ import annotations

"""Conservative cross-company public-data recovery.

The workbook should not stop at ``INSUFFICIENT_DATA`` when the same disclosed annual fact is
available from another structured public source. This module fills *blank* secondary statement
cells from annual Yahoo/yfinance statements using conservative label matching, then derives only
accounting identities that are mechanically safe. It never overwrites an exact SEC/issuer value.

The hierarchy remains:
1. issuer / regulator / exact canonical values already in the workbook;
2. exact structured-provider labels;
3. high-confidence fuzzy structured-provider labels for blanks only;
4. mechanically derived accounting identities;
5. explicit N/M when the issuer genuinely does not disclose the item.

This is deliberately not a web-scraping free-for-all. A plausible number is worse than a blank.
"""

from copy import deepcopy
from difflib import SequenceMatcher
import math
import re
from typing import Any

import yfinance as yf
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import full_financial_statements as base
import issuer_statement_profiles as statement_profiles
from business_model_registry import get_business_model_policy

GREEN = "E2F0D9"; GOLD = "FFF2CC"; BLUE = "2F75B5"; WHITE = "FFFFFF"; GREY = "666666"
FMT_BN = '#,##0.0;[Red](#,##0.0);-'; FMT_PCT = '0.0%;[Red](0.0%);-'

STOP = {
    "and", "the", "of", "to", "from", "for", "with", "including", "excluding", "other",
    "total", "net", "current", "non", "noncurrent", "expense", "expenses", "income", "loss",
}


def _num(v: Any) -> float | None:
    try:
        if isinstance(v, bool) or v in (None, ""):
            return None
        x = float(v)
        return x if math.isfinite(x) else None
    except Exception:
        return None


def _norm(value: Any) -> str:
    text = str(value or "").lower().replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def _tokens(value: Any) -> set[str]:
    return {x for x in _norm(value).split() if len(x) > 2 and x not in STOP}


def _conflict(target: str, candidate: str) -> bool:
    """Reject matches that are lexically similar but economically different."""
    t = f" {_norm(target)} "; c = f" {_norm(candidate)} "
    pairs = [
        ((" operating ",), (" pretax ", " pre tax ", " before tax ")),
        ((" pretax ", " pre tax ", " before tax "), (" operating ",)),
        ((" gross ",), (" net ",)),
        ((" diluted ",), (" basic ",)),
        ((" basic ",), (" diluted ",)),
        ((" current debt ", " short term debt "), (" long term debt ", " noncurrent debt ")),
        ((" long term debt ", " noncurrent debt "), (" current debt ", " short term debt ")),
        ((" current assets ",), (" non current assets ", " noncurrent assets ")),
        ((" current liabilities ",), (" non current liabilities ", " noncurrent liabilities ")),
        ((" accounts receivable ",), (" accounts payable ",)),
        ((" accounts payable ",), (" accounts receivable ",)),
        ((" stock based compensation ", " share based compensation "), (" compensation and benefits ", " labor and related ")),
    ]
    for left, right in pairs:
        if any(x in t for x in left) and any(x in c for x in right):
            return True
    return False


def _score(targets: list[str], candidate: str) -> float:
    best = 0.0
    c_norm = _norm(candidate); c_tokens = _tokens(candidate)
    if not c_norm:
        return 0.0
    for target in targets:
        if _conflict(target, candidate):
            continue
        t_norm = _norm(target); t_tokens = _tokens(target)
        if not t_norm:
            continue
        if t_norm == c_norm:
            return 1.0
        seq = SequenceMatcher(None, t_norm, c_norm).ratio()
        union = t_tokens | c_tokens
        jaccard = len(t_tokens & c_tokens) / len(union) if union else 0.0
        containment = len(t_tokens & c_tokens) / len(t_tokens) if t_tokens else 0.0
        combined = 0.45 * seq + 0.35 * jaccard + 0.20 * containment
        best = max(best, combined)
    return best


def _best_provider_row(df, label: str, aliases: list[str]) -> tuple[str | None, float]:
    if df is None or getattr(df, "empty", True):
        return None, 0.0
    targets = [label] + [str(x) for x in aliases if x]
    normalized = {_norm(str(x)): str(x) for x in df.index}
    for target in targets:
        exact = normalized.get(_norm(target))
        if exact is not None:
            return exact, 1.0
    ranked = sorted(((_score(targets, str(row)), str(row)) for row in df.index), reverse=True)
    if not ranked:
        return None, 0.0
    score, row = ranked[0]
    return (row, score) if score >= 0.86 else (None, score)


def _profile_for(wb, ticker: str):
    t = str(ticker or "").upper().strip()
    if t in statement_profiles.TICKER_PROFILE:
        return statement_profiles.get_statement_profile(t)
    try:
        cd = wb["Company Data"]
        policy = get_business_model_policy(t, cd["B6"].value, cd["B7"].value, cd["B5"].value)
    except Exception:
        policy = get_business_model_policy(t)
    if policy.key in {"bank", "capital_markets"}:
        p = deepcopy(statement_profiles.PROFILES["bank"]); p["key"] = "bank"; return p
    if policy.key in {"insurance", "insurance_conglomerate"}:
        p = deepcopy(statement_profiles.PROFILES["berkshire"]); p["key"] = "berkshire"; return p
    p = deepcopy(statement_profiles.PROFILES["default"]); p["key"] = "default"; return p


def _find(ws, label: str, start: int = 1, end: int | None = None) -> int | None:
    needle = _norm(label); end = end or ws.max_row
    for r in range(start, min(end, ws.max_row) + 1):
        if _norm(ws.cell(r, 1).value) == needle:
            return r
    return None


def _year_cols(ws, header: int) -> dict[int, int]:
    return {
        int(ws.cell(header, c).value): c
        for c in range(2, min(ws.max_column, 12) + 1)
        if isinstance(ws.cell(header, c).value, (int, float)) and 1900 <= int(ws.cell(header, c).value) <= 2100
    }


def _provider_frames(ticker: str):
    try:
        t = yf.Ticker(ticker)
        return {"income": t.income_stmt, "balance": t.balance_sheet, "cash": t.cashflow}
    except Exception:
        return {"income": None, "balance": None, "cash": None}


def _provider_values(df, row_name: str, unit: str) -> dict[int, float]:
    if df is None or getattr(df, "empty", True) or row_name not in df.index:
        return {}
    out: dict[int, float] = {}
    series = df.loc[row_name]
    for col, raw in series.items():
        try:
            year = int(getattr(col, "year", str(col)[:4])); value = float(raw)
        except Exception:
            continue
        if not math.isfinite(value):
            continue
        out[year] = value if unit == "eps" else value / 1e9
    return out


def _sections(ws):
    i0 = _find(ws, "Income Statement"); b0 = _find(ws, "Balance Sheet"); c0 = _find(ws, "Cash Flow Statement")
    if not all((i0, b0, c0)):
        return None
    ih = next((r for r in range(i0 + 1, min(b0, i0 + 6)) if _norm(ws.cell(r, 1).value) == "metric"), None)
    bh = next((r for r in range(b0 + 1, min(c0, b0 + 6)) if _norm(ws.cell(r, 1).value) == "metric"), None)
    ch = next((r for r in range(c0 + 1, min(ws.max_row + 1, c0 + 6)) if _norm(ws.cell(r, 1).value) == "metric"), None)
    return (i0, b0, c0, ih, bh, ch) if all((ih, bh, ch)) else None


def _fill_section(ws, rows, df, year_cols, start, end, source_col) -> tuple[int, list[dict]]:
    filled = 0; matches: list[dict] = []
    for label, yf_names, _sec_tags, unit, aliases in rows:
        row = _find(ws, label, start, end)
        if not row:
            continue
        if all(ws.cell(row, c).value not in (None, "") for c in year_cols.values()):
            continue
        candidate, confidence = _best_provider_row(df, label, list(yf_names or []) + list(aliases or []))
        if candidate is None:
            continue
        values = _provider_values(df, candidate, unit)
        local = 0
        for year, col in year_cols.items():
            if ws.cell(row, col).value not in (None, ""):
                continue
            value = values.get(year)
            if value is None:
                continue
            if label in base.OUTFLOW_LABELS:
                value = -abs(value)
            ws.cell(row, col).value = value
            ws.cell(row, col).number_format = base.FMT_EPS if unit == "eps" else base.FMT_SHARES if unit == "shares" else base.FMT_BN
            filled += 1; local += 1
        if local:
            note = f"Yahoo annual statement fallback — {'exact' if confidence == 1.0 else 'high-confidence'} label match: {candidate}"
            existing = str(ws.cell(row, source_col).value or "").strip()
            if not existing or "Not reliably mapped" in existing:
                ws.cell(row, source_col).value = note
            elif note not in existing:
                ws.cell(row, source_col).value = existing + " → " + note
            ws.cell(row, source_col).font = Font(italic=True)
            ws.cell(row, source_col).alignment = Alignment(wrap_text=True, vertical="top")
            matches.append({"label": label, "provider_row": candidate, "confidence": confidence, "cells": local})
    return filled, matches


def _formula_if_blank(ws, row: int | None, col: int, formula: str, fmt: str = FMT_BN) -> bool:
    if not row or ws.cell(row, col).value not in (None, ""):
        return False
    ws.cell(row, col).value = formula; ws.cell(row, col).number_format = fmt
    return True


def _derive_safe_identities(ws, sections, profile) -> int:
    i0, b0, c0, ih, bh, ch = sections
    iy = _year_cols(ws, ih); by = _year_cols(ws, bh); cy = _year_cols(ws, ch)
    n = 0
    if profile.get("key") not in {"bank", "berkshire"}:
        rev = _find(ws, "Revenue", ih + 1, b0 - 1); cost = _find(ws, "Cost of Revenue", ih + 1, b0 - 1); gross = _find(ws, "Gross Profit", ih + 1, b0 - 1)
        if rev and cost and gross:
            for c in iy.values():
                if ws.cell(rev, c).value not in (None, "") and ws.cell(cost, c).value not in (None, ""):
                    col = get_column_letter(c); n += int(_formula_if_blank(ws, gross, c, f'=IFERROR({col}{rev}-{col}{cost},"")'))

    asset_pairs = [
        ("Cash + Short-Term Investments", "Cash & Cash Equivalents", "Short-Term Investments / Marketable Securities", "+"),
        ("Total Non-Current Assets", "Total Assets", "Total Current Assets", "-"),
        ("Total Non-Current Liabilities", "Total Liabilities", "Total Current Liabilities", "-"),
        ("Total Liabilities & Equity", "Total Liabilities", "Total Equity", "+"),
    ]
    for target, a, b, op in asset_pairs:
        tr = _find(ws, target, bh + 1, c0 - 1); ar = _find(ws, a, bh + 1, c0 - 1); br = _find(ws, b, bh + 1, c0 - 1)
        if not all((tr, ar, br)):
            continue
        for c in by.values():
            if ws.cell(ar, c).value in (None, "") or ws.cell(br, c).value in (None, ""):
                continue
            col = get_column_letter(c); n += int(_formula_if_blank(ws, tr, c, f'=IFERROR({col}{ar}{op}{col}{br},"")'))

    ocf = _find(ws, "Operating Cash Flow", ch + 1, ws.max_row); capex = _find(ws, "Capital Expenditures", ch + 1, ws.max_row); fcf = _find(ws, "Free Cash Flow", ch + 1, ws.max_row)
    if ocf and capex and fcf:
        for c in cy.values():
            if ws.cell(ocf, c).value not in (None, "") and ws.cell(capex, c).value not in (None, ""):
                col = get_column_letter(c); n += int(_formula_if_blank(ws, fcf, c, f'=IFERROR({col}{ocf}+{col}{capex},"")'))

    begin = _find(ws, "Beginning Cash", ch + 1, ws.max_row); end = _find(ws, "Ending Cash", ch + 1, ws.max_row)
    if begin and end:
        ordered = sorted(cy.items())
        for idx, (_year, c) in enumerate(ordered):
            if idx == 0:
                continue
            prev_c = ordered[idx - 1][1]
            if ws.cell(end, prev_c).value not in (None, ""):
                n += int(_formula_if_blank(ws, begin, c, f'={get_column_letter(prev_c)}{end}'))
    return n


def _summary_sentence_rows(info: dict | None) -> list[tuple[str, str, str, str]]:
    info = info or {}; summary = str(info.get("longBusinessSummary") or "").strip(); website = str(info.get("website") or "").strip()
    if not summary:
        return []
    source = website or "https://finance.yahoo.com/"
    sentences = [x.strip() for x in re.split(r"(?<=[.!?])\s+", summary.replace("\n", " ")) if len(x.strip()) >= 35]
    return [("Public business profile", f"Business activity {idx}", sentence[:500], source) for idx, sentence in enumerate(sentences[:3], 1)]


def _backfill_company_profile(wb, ticker: str, info: dict | None) -> int:
    if "Company Data" not in wb.sheetnames:
        return 0
    ws = wb["Company Data"]; header = _find(ws, "Business / Segment")
    if not header:
        return 0
    populated = []
    for r in range(header + 1, min(ws.max_row, header + 10) + 1):
        product = str(ws.cell(r, 2).value or "").strip()
        if product:
            populated.append((r, product))
    if len(populated) >= 2 or (populated and "see business description" not in populated[0][1].lower()):
        return 0
    rows = _summary_sentence_rows(info)
    if not rows:
        return 0
    for r in range(header + 1, min(ws.max_row, header + 10) + 1):
        for c in range(1, 5):
            ws.cell(r, c).value = None
    for r, item in enumerate(rows, header + 1):
        for c, value in enumerate(item, 1):
            ws.cell(r, c).value = value; ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
        if item[3]:
            ws.cell(r, 4).font = Font(color="008000", underline="single")
    return len(rows)


def _decorate_quality(wb, result: dict[str, Any]) -> None:
    if "Data Quality" not in wb.sheetnames:
        return
    ws = wb["Data Quality"]; label = "Public-data recovery coverage"
    row = next((r for r in range(1, ws.max_row + 1) if str(ws.cell(r, 1).value or "").strip() == label), ws.max_row + 1)
    filled = int(result.get("provider_cells_filled") or 0); derived = int(result.get("derived_cells_filled") or 0); profile = int(result.get("profile_rows_filled") or 0)
    status = "PASS" if filled + derived + profile > 0 else "REVIEW"
    observed = (
        f"Recovered {filled} blank annual statement cell(s) from conservative structured public-source matches, "
        f"derived {derived} mechanically safe accounting identity cell(s), and added {profile} public business-profile row(s)."
    )
    why = (
        "The model exhausts issuer/regulator and structured public fallbacks before declaring insufficient evidence. "
        "Existing exact values are never overwritten by this layer."
    )
    for c, value in enumerate((label, status, observed, why), 1):
        ws.cell(row, c).value = value; ws.cell(row, c).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row, 2).fill = PatternFill("solid", fgColor=GREEN if status == "PASS" else GOLD); ws.cell(row, 2).font = Font(bold=True)


def decorate_public_data_quality(wb, result: dict[str, Any]) -> None:
    """Reapply the recovery disclosure after downstream modules rebuild Data Quality."""
    _decorate_quality(wb, result or {})


def backfill_public_data(wb, ticker: str, info: dict | None = None) -> dict[str, Any]:
    result: dict[str, Any] = {
        "ticker": str(ticker or "").upper().strip(), "provider_cells_filled": 0,
        "derived_cells_filled": 0, "profile_rows_filled": 0, "matches": [],
    }
    if "Financial Statements" not in wb.sheetnames:
        result["profile_rows_filled"] = _backfill_company_profile(wb, ticker, info)
        _decorate_quality(wb, result)
        return result

    ws = wb["Financial Statements"]; sections = _sections(ws)
    if not sections:
        result["profile_rows_filled"] = _backfill_company_profile(wb, ticker, info)
        _decorate_quality(wb, result)
        return result
    i0, b0, c0, ih, bh, ch = sections
    profile = _profile_for(wb, ticker); frames = _provider_frames(ticker)
    iy = _year_cols(ws, ih); by = _year_cols(ws, bh); cy = _year_cols(ws, ch)
    source_col = max(max(iy.values(), default=1), max(by.values(), default=1), max(cy.values(), default=1)) + 1

    for key, rows, frame, cols, start, end in (
        ("income", profile.get("income") or [], frames.get("income"), iy, ih + 1, b0 - 1),
        ("balance", profile.get("balance") or [], frames.get("balance"), by, bh + 1, c0 - 1),
        ("cash", profile.get("cash") or [], frames.get("cash"), cy, ch + 1, ws.max_row),
    ):
        filled, matches = _fill_section(ws, rows, frame, cols, start, end, source_col)
        result["provider_cells_filled"] += filled
        result["matches"].extend({"section": key, **m} for m in matches)

    result["derived_cells_filled"] = _derive_safe_identities(ws, sections, profile)
    result["profile_rows_filled"] = _backfill_company_profile(wb, ticker, info)
    _decorate_quality(wb, result)
    try:
        wb.calculation.calcMode = "auto"; wb.calculation.fullCalcOnLoad = True; wb.calculation.forceFullCalc = True
    except Exception:
        pass
    return result
