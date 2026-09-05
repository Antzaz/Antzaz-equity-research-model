from __future__ import annotations

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference

from chart_excel_compat import apply_chart_compatibility_fix
from machine_learning.ai_growth_output import (
    _add_ai_growth_charts,
    _set_chart_text_categories as set_ai_categories,
)
from machine_learning.common import MLResult
from machine_learning.workbook import _add_ml_charts, _set_chart_text_categories as set_ml_categories


def _sample_chart(ws):
    ws["X2"] = "Category"
    ws["Y2"] = "Model"
    ws["Z2"] = "Baseline"
    ws["X3"] = "Alpha"
    ws["X4"] = "Beta"
    ws["Y3"] = 10
    ws["Y4"] = 20
    ws["Z3"] = 12
    ws["Z4"] = 18
    chart = BarChart()
    chart.add_data(Reference(ws, min_col=25, max_col=26, min_row=3, max_row=4), titles_from_data=False)
    return chart


def _assert_chart_refs(chart, expected_titles):
    assert len(chart.series) == len(expected_titles)
    for series, title in zip(chart.series, expected_titles):
        assert series.cat.strRef is not None
        assert series.cat.numRef is None
        assert series.cat.strRef.f == "'Chart Test'!$X$3:$X$4"
        assert series.tx is not None
        assert series.tx.v == title
        assert series.tx.strRef is None


def test_ml_chart_categories_are_explicit_text_refs_with_literal_series_names():
    wb = Workbook()
    ws = wb.active
    ws.title = "Chart Test"
    chart = _sample_chart(ws)
    set_ml_categories(chart, ws, 24, 3, 4, ["Forecast magnitude", "Typical historical error"])
    _assert_chart_refs(chart, ["Forecast magnitude", "Typical historical error"])


def test_ai_growth_chart_categories_are_explicit_text_refs_with_literal_series_names():
    wb = Workbook()
    ws = wb.active
    ws.title = "Chart Test"
    chart = _sample_chart(ws)
    set_ai_categories(chart, ws, 24, 3, 4, ["Model", "Baseline"])
    _assert_chart_refs(chart, ["Model", "Baseline"])


def _add_orientation_chart(ws, chart_type: str, anchor: str):
    chart = BarChart()
    chart.type = chart_type
    chart.add_data(Reference(ws, min_col=25, min_row=3, max_row=4), titles_from_data=False)
    set_ml_categories(chart, ws, 24, 3, 4, ["Forecast"])
    ws.add_chart(chart, anchor)
    return chart


def test_excel_compat_postpass_unhides_sources_and_repairs_axis_positions(tmp_path):
    wb = Workbook()
    ml = wb.active
    ml.title = "ML & Quantitative Research"
    for col in ("X", "Y", "Z", "AA"):
        ml.column_dimensions[col].hidden = True
    ml["X2"] = "Model"
    ml["Y2"] = "Forecast"
    ml["X3"] = "A"
    ml["X4"] = "B"
    ml["Y3"] = 10
    ml["Y4"] = 20

    _add_orientation_chart(ml, "col", "K8")
    _add_orientation_chart(ml, "bar", "K40")

    path = tmp_path / "chart_compat.xlsx"
    wb.save(path)
    result = apply_chart_compatibility_fix(path)
    assert result["charts_repaired"] == 2

    fixed = load_workbook(path)
    ws = fixed["ML & Quantitative Research"]
    assert ws.column_dimensions["X"].hidden is False
    assert ws.column_dimensions["Y"].hidden is False

    c1, c2 = ws._charts
    assert c1.visible_cells_only is False
    assert c1.x_axis.axPos == "b"
    assert c1.y_axis.axPos == "l"
    assert c2.visible_cells_only is False
    assert c2.x_axis.axPos == "l"
    assert c2.y_axis.axPos == "b"


def test_excel_compat_uses_chart_type_not_chart_order(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "ML & Quantitative Research"
    ws["X2"] = "Model"
    ws["Y2"] = "Forecast"
    ws["X3"] = "A"
    ws["X4"] = "B"
    ws["Y3"] = 10
    ws["Y4"] = 20

    _add_orientation_chart(ws, "bar", "K8")
    _add_orientation_chart(ws, "col", "K40")

    path = tmp_path / "chart_order_independent.xlsx"
    wb.save(path)
    apply_chart_compatibility_fix(path)

    fixed = load_workbook(path)
    horizontal, column = fixed["ML & Quantitative Research"]._charts
    assert horizontal.x_axis.axPos == "l"
    assert horizontal.y_axis.axPos == "b"
    assert column.x_axis.axPos == "b"
    assert column.y_axis.axPos == "l"


def test_primary_ml_generator_is_excel_safe_without_postpass():
    wb = Workbook()
    ws = wb.active
    ws.title = "ML & Quantitative Research"
    results = [
        MLResult(
            name="Expected 12M Excess Return",
            status="REVIEW",
            summary="",
            prediction=0.05,
            metrics={"walk_forward": {"mae": 0.20, "directional_accuracy": 0.55, "baseline_directional_accuracy": 0.50}},
            drivers=[{"feature": "momentum_12m", "importance": 1.0}],
        ),
        MLResult(
            name="Consensus / Earnings Surprise",
            status="REVIEW",
            summary="",
            prediction=0.10,
            metrics={"walk_forward": {"mae": 0.15, "directional_accuracy": 0.60, "baseline_directional_accuracy": 0.52}},
        ),
        MLResult(
            name="Market Regime Classifier",
            status="REVIEW",
            summary="",
            metrics={"regime_probabilities": {"Growth / risk-on": 0.6, "Risk-off / crisis": 0.4}, "monthly_rows": 120},
        ),
    ]
    _add_ml_charts(ws, results)
    assert ws.column_dimensions["X"].hidden is False
    assert ws.column_dimensions["Y"].hidden is False
    assert len(ws._charts) == 4
    for chart in ws._charts:
        assert chart.visible_cells_only is False
    assert ws._charts[0].x_axis.axPos == "b"
    assert ws._charts[2].x_axis.axPos == "l"


def test_primary_ai_generator_is_excel_safe_without_postpass():
    wb = Workbook()
    ws = wb.active
    ws.title = "AI Growth Forecast"
    signals = {
        "demand_score": 0.7,
        "monetization_score": 0.65,
        "adoption_score": 0.75,
        "efficiency_score": 0.55,
        "capex_burden_score": 0.4,
        "risk_score": 0.35,
        "evidence": ["substantive AI evidence"],
        "extraction_mode": "test",
    }
    revenue = {"drivers": [{"feature": "momentum_12m", "shap_value": 0.02}]}
    fcf = {"prediction": 0.20, "drivers": [{"feature": "fcf_margin", "shap_value": -0.03}]}
    payload = {"ai_adjusted_fcf_growth": 0.23}
    reverse = {"implied_annual_fcf_growth": 0.15}
    _add_ai_growth_charts(ws, signals, revenue, fcf, payload, reverse)
    assert ws.column_dimensions["P"].hidden is False
    assert ws.column_dimensions["Q"].hidden is False
    assert len(ws._charts) == 3
    for chart in ws._charts:
        assert chart.visible_cells_only is False
    assert ws._charts[0].x_axis.axPos == "b"
    assert ws._charts[2].x_axis.axPos == "l"
