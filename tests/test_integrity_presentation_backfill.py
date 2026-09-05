from __future__ import annotations

from openpyxl import Workbook

from canonical_statement_guard import apply_canonical_statement_guard
from public_data_backfill import _best_provider_row, _conflict
from workbook_presentation import apply_workbook_presentation


def _fact(tag, value, fy=2025, unit="USD"):
    return {
        "facts": {
            "us-gaap": {
                tag: {
                    "units": {
                        unit: [{
                            "form": "10-K", "fy": fy, "start": f"{fy-1}-10-01", "end": f"{fy}-09-30",
                            "filed": f"{fy}-11-01", "val": value,
                        }]
                    }
                }
            }
        }
    }


def _merge_facts(*payloads):
    out = {"facts": {"us-gaap": {}}}
    for payload in payloads:
        out["facts"]["us-gaap"].update(payload["facts"]["us-gaap"])
    return out


def _visa_like_workbook():
    wb = Workbook(); wb.remove(wb.active)
    cd = wb.create_sheet("Company Data")
    cd["B4"] = "V"; cd["B5"] = "Visa Inc."; cd["B6"] = "Financial Services"; cd["B7"] = "Credit Services"

    fs = wb.create_sheet("Financial Statements")
    fs["A5"] = "Income Statement"; fs["A6"] = "Metric"; fs["B6"] = 2025
    labels = ["Revenue", "Total Operating Expenses", "Operating Income", "Pre-Tax Income", "Income Taxes", "Net Income", "Diluted EPS"]
    for r, label in enumerate(labels, 7): fs.cell(r, 1, label)
    fs["B7"] = 40.000; fs["B8"] = 13.444; fs["B9"] = 26.556  # normalized/provider-like wrong GAAP op income
    fs["A16"] = "Balance Sheet"; fs["A17"] = "Metric"; fs["B17"] = 2025
    fs["A20"] = "Cash Flow Statement"; fs["A21"] = "Metric"; fs["B21"] = 2025
    for r, label in enumerate(["Operating Cash Flow", "Capital Expenditures", "Depreciation, Amortization & Accretion", "Stock-Based Compensation"], 22):
        fs.cell(r, 1, label)

    hs = wb.create_sheet("Historical Financials")
    hs["B3"] = 2025; hs["B4"] = 40.0; hs["B9"] = 26.556; hs["B11"] = 20.0
    return wb


def test_sec_canonical_guard_overwrites_adjusted_operating_income():
    wb = _visa_like_workbook()
    facts = _merge_facts(
        _fact("RevenueFromContractWithCustomerExcludingAssessedTax", 40_000_000_000),
        _fact("OperatingExpenses", 16_006_000_000),
        _fact("OperatingIncomeLoss", 23_994_000_000),
        _fact("NetIncomeLoss", 20_000_000_000),
    )
    result = apply_canonical_statement_guard(wb, "V", facts)
    fs = wb["Financial Statements"]
    assert abs(fs["B9"].value - 23.994) < 1e-9
    assert abs(wb["Historical Financials"]["B9"].value - 23.994) < 1e-9
    assert any(x["label"] == "Operating Income" for x in result["material_corrections"])
    # Payments networks should not display a synthetic industrial gross-profit structure.
    assert result["business_model"] == "payments"


def test_provider_match_rejects_economically_wrong_alias():
    assert _conflict("Operating Income", "Pretax Income")


def test_provider_match_accepts_obvious_structured_alias():
    import pandas as pd
    frame = pd.DataFrame([[1.0]], index=["Selling General And Administration"], columns=[pd.Timestamp("2025-12-31")])
    row, score = _best_provider_row(frame, "Selling, General & Administrative", ["Selling General And Administration"])
    assert row == "Selling General And Administration"
    assert score == 1.0


def test_presentation_keeps_model_but_hides_redundant_tabs():
    wb = Workbook(); wb.active.title = "Dashboard"
    for name in [
        "Visual Dashboard", "Investment Summary", "Decision View", "Company Data", "Historical Financials",
        "Financial Statements", "Segment Analysis", "Three-Case Scenarios", "DCF", "Peer Comps",
        "Advanced Analytics", "ML & Quantitative Research", "AI Growth Forecast", "Data Quality",
        "Comparative Analysis", "Score Audit Trail",
    ]:
        wb.create_sheet(name)
    wb["Company Data"]["B4"] = "V"; wb["Company Data"]["B5"] = "Visa Inc."
    result = apply_workbook_presentation(wb, "V")
    assert wb.sheetnames[0] == "Visual Dashboard"
    assert wb["Dashboard"].sheet_state == "hidden"
    assert wb["Comparative Analysis"].sheet_state == "hidden"
    assert wb["Score Audit Trail"].sheet_state == "hidden"
    assert "Financial Statements" in wb.sheetnames
    assert result["navigation_links"] >= 6
    assert wb["Visual Dashboard"]["A1"].value == "V — Equity Research Dashboard"
