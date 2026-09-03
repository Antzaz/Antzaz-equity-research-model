from __future__ import annotations

from pathlib import Path
import math

from openpyxl import Workbook

from business_model_registry import get_business_model_policy
import cross_sector_runtime as runtime
from institutional_research.src.reverse_dcf import reverse_dcf_table
import sector_score_policy


def test_business_model_matrix_routes_by_economics_not_ticker_whitelist():
    cases = [
        ("BAC", "Financial Services", "Banks - Diversified", "Bank of America", "bank", False),
        ("WFC", "Financial Services", "Banks - Diversified", "Wells Fargo", "bank", False),
        ("PGR", "Financial Services", "Insurance - Property & Casualty", "Progressive", "insurance", False),
        ("CB", "Financial Services", "Insurance - Property & Casualty", "Chubb", "insurance", False),
        ("PLD", "Real Estate", "REIT - Industrial", "Prologis", "reit", False),
        ("O", "Real Estate", "REIT - Retail", "Realty Income", "reit", False),
        ("CVX", "Energy", "Oil & Gas Integrated", "Chevron", "commodity", True),
        ("NEM", "Basic Materials", "Gold", "Newmont", "commodity", True),
        ("NEE", "Utilities", "Utilities - Regulated Electric", "NextEra Energy", "utility", True),
        ("MSFT", "Technology", "Software - Infrastructure", "Microsoft", "software", True),
        ("AMD", "Technology", "Semiconductors", "Advanced Micro Devices", "semiconductor", True),
        ("LLY", "Healthcare", "Drug Manufacturers - General", "Eli Lilly", "pharma", True),
        ("CRSP", "Healthcare", "Biotechnology", "CRISPR Therapeutics", "biotech", True),
        ("CAT", "Industrials", "Farm & Heavy Construction Machinery", "Caterpillar", "industrial", True),
        ("WMT", "Consumer Defensive", "Discount Stores", "Walmart", "consumer", True),
        ("V", "Financial Services", "Credit Services", "Visa", "payments", True),
        ("GOOGL", "Communication Services", "Internet Content & Information", "Alphabet", "digital_platform", True),
    ]
    for ticker, sector, industry, name, expected_key, reverse_allowed in cases:
        policy = get_business_model_policy(ticker, sector, industry, name)
        assert policy.key == expected_key, (ticker, policy.key, expected_key)
        assert policy.reverse_dcf_allowed is reverse_allowed, (ticker, policy)
        if policy.key == "commodity":
            assert policy.commodity_normalization is True


def test_generic_financial_statement_profiles_use_sector_context():
    runtime._CONTEXT.clear()
    runtime._CONTEXT["BAC"] = {"ticker": "BAC", "sector": "Financial Services", "industry": "Banks - Diversified", "name": "Bank of America"}
    runtime._CONTEXT["PGR"] = {"ticker": "PGR", "sector": "Financial Services", "industry": "Insurance - Property & Casualty", "name": "Progressive"}
    runtime._CONTEXT["PLD"] = {"ticker": "PLD", "sector": "Real Estate", "industry": "REIT - Industrial", "name": "Prologis"}

    bank = runtime.contextual_statement_profile("BAC")
    insurer = runtime.contextual_statement_profile("PGR")
    reit = runtime.contextual_statement_profile("PLD")

    assert bank["key"] == "bank" and not bank["derive_fcf"] and not bank["balance_net_debt"]
    assert "Total Net Revenue" in {row[0] for row in bank["income"]}
    assert insurer["key"] == "berkshire" and not insurer["derive_fcf"]
    assert "Insurance Premiums Earned" in {row[0] for row in insurer["income"]}
    assert reit["derive_fcf"] is False
    assert reit["name"].startswith("REIT")


def test_public_reverse_dcf_is_nm_for_bank_insurer_and_reit():
    info = {
        "BAC": {"sector": "Financial Services", "industry": "Banks - Diversified", "longName": "Bank of America", "marketCap": 300e9, "freeCashflow": 25e9, "totalCash": 100e9, "totalDebt": 300e9},
        "PGR": {"sector": "Financial Services", "industry": "Insurance - Property & Casualty", "longName": "Progressive", "marketCap": 150e9, "freeCashflow": 10e9},
        "PLD": {"sector": "Real Estate", "industry": "REIT - Industrial", "longName": "Prologis", "marketCap": 120e9, "freeCashflow": 6e9},
        "MSFT": {"sector": "Technology", "industry": "Software - Infrastructure", "longName": "Microsoft", "marketCap": 3000e9, "freeCashflow": 120e9, "totalCash": 80e9, "totalDebt": 70e9},
    }
    table = reverse_dcf_table(
        list(info), info,
        {"years": 10, "wacc": 0.09, "terminal_growth": 0.03, "min_growth": -0.20, "max_growth": 0.50},
    ).set_index("Ticker")
    for ticker in ("BAC", "PGR", "PLD"):
        assert math.isnan(float(table.loc[ticker, "ImpliedAnnualFCFGrowth"]))
        assert "N/M" in str(table.loc[ticker, "Status"])
    assert "N/M" not in str(table.loc["MSFT", "Status"])


def test_descriptive_segment_fallback_uses_company_taxonomy_without_inventing_numbers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Company Data"
    ws["B4"] = "TEST"
    ws["B5"] = "Test Industrial"
    ws["B6"] = "Industrials"
    ws["B7"] = "Specialty Industrial Machinery"
    ws["A20"] = "Business / Segment"
    rows = [
        ("Automation", "Factory automation", "Issuer product", "https://issuer.example/investors"),
        ("Automation", "Industrial software", "Issuer product", "https://issuer.example/investors"),
        ("Electrification", "Grid equipment", "Issuer product", "https://issuer.example/investors"),
        ("Services", "Lifecycle services", "Issuer product", "https://issuer.example/investors"),
    ]
    for r, values in enumerate(rows, 21):
        for c, value in enumerate(values, 1):
            ws.cell(r, c, value)
    h = wb.create_sheet("Historical Financials")
    for c, year in enumerate((2023, 2024, 2025), 5):
        h.cell(3, c, year)
    result = runtime.ensure_descriptive_segment_contract(wb, "TEST")
    assert result["changed"] is True
    seg = wb["Segment Analysis"]
    names = {str(seg.cell(r, 1).value or "") for r in range(7, 17)}
    assert {"Automation", "Electrification", "Services"}.issubset(names)
    for r in range(7, 10):
        # Descriptive fallback must not manufacture revenue/profit values.
        assert all(seg.cell(r, c).value is None for c in (2, 3, 4, 7, 8, 9))


def test_sector_score_gate_excludes_industrial_dcf_for_bank(monkeypatch):
    wb = Workbook()
    ws = wb.active
    ws.title = "Company Data"
    ws["B4"] = "BAC"; ws["B5"] = "Bank of America"; ws["B6"] = "Financial Services"; ws["B7"] = "Banks - Diversified"

    dims = {
        key: {"score": 70.0, "status": "Complete", "formula": "legacy", "components": "legacy"}
        for key in ("Growth", "Profitability", "FCF Quality", "Balance Sheet", "Absolute Valuation", "Relative Valuation", "Stress Robustness", "Bayesian Skew")
    }
    fake = {
        "dimensions": dims,
        "valuation_model_reliability": {"status": "PASS", "reasons": []},
        "category_scores": {},
        "coverage": 1.0,
    }
    monkeypatch.setattr(sector_score_policy.base, "compute_score_bundle", lambda *a, **k: fake)
    monkeypatch.setattr(sector_score_policy.base, "_recompute", lambda bundle: bundle)
    result = sector_score_policy.compute_score_bundle(wb, ticker="BAC")
    assert result["dimensions"]["FCF Quality"]["score"] is None
    assert result["dimensions"]["Balance Sheet"]["score"] is None
    assert result["dimensions"]["Absolute Valuation"]["score"] is None
    assert result["dimensions"]["Stress Robustness"]["score"] is None
    assert result["valuation_model_reliability"]["status"] == "REVIEW"
