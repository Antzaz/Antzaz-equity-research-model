from __future__ import annotations

from openpyxl import Workbook
from score_engine_v2 import compute_score_bundle


def _costco_like_workbook():
    wb=Workbook(); d=wb.active; d.title="Company Data"
    for cell,val in {"B8":945.01,"B10":419.09,"B12":14.161,"B13":13.263,"B14":-0.898}.items(): d[cell]=val
    h=wb.create_sheet("Historical Financials")
    years=[2020,2021,2022,2023,2024,2025]
    rev=[166.761,195.929,226.954,242.290,254.453,275.235]
    op=[5.435,6.708,7.793,8.114,9.285,10.383]
    ni=[4.002,5.007,5.844,6.292,7.367,8.099]
    ocf=[8.861,8.958,7.392,11.068,11.339,13.335]
    cap=[2.810,3.588,3.891,4.323,4.710,5.498]
    for i,y in enumerate(years,2):
        h.cell(3,i,y); h.cell(4,i,rev[i-2]); h.cell(9,i,op[i-2]); h.cell(11,i,ni[i-2]); h.cell(14,i,ocf[i-2]); h.cell(15,i,cap[i-2])
    h["A24"]="Sources"; h["B24"]="SEC Company Facts / issuer annual report"

    p=wb.create_sheet("Peer Comps")
    headers=["Company","Ticker","Forward P/E","EV/Revenue","EV/EBITDA","Revenue Growth","Operating Margin","ROE","Sector","Industry","Discovery","Source URL","Industry Market Share %","Peer-Set Market Cap %","Market Share Basis / Source","Peer Type"]
    for c,v in enumerate(headers,1): p.cell(3,c,v)
    target=["Costco","COST",41.7,1.433,30.52,.0817,.0377,.2915,"Consumer Defensive","Discount Stores",None,None,None,None,None,"Target classification"]
    peers=[
        ["Walmart","WMT",38,.95,18,.07,.043,.24,"Consumer Defensive","Discount Stores",None,None,None,None,None,"Direct business-model peer"],
        ["Target","TGT",12,.55,9,.03,.055,.25,"Consumer Defensive","Discount Stores",None,None,None,None,None,"Direct business-model peer"],
        ["BJ's","BJ",23,1.0,14,.08,.04,.19,"Consumer Defensive","Discount Stores",None,None,None,None,None,"Direct business-model peer"],
    ]
    for r,row in enumerate([target]+peers,4):
        for c,v in enumerate(row,1): p.cell(r,c,v)

    s=wb.create_sheet("Three-Case Scenarios"); s["C6"]=.09

    fs=wb.create_sheet("Financial Statements")
    fs["A5"]="Income Statement"; fs["A6"]="Metric"; fs["B6"]=2024; fs["C6"]=2025
    labels={"Revenue":7,"Operating Income":8,"Pre-Tax Income":9,"Income Taxes":10,"Net Income":11}
    for name,r in labels.items(): fs.cell(r,1,name)
    fs["B7"]=254.453; fs["C7"]=275.235; fs["B8"]=9.285; fs["C8"]=10.383; fs["B9"]=9.65; fs["C9"]=10.818; fs["B10"]=2.283; fs["C10"]=2.719; fs["B11"]=7.367; fs["C11"]=8.099
    fs["A14"]="Balance Sheet"; fs["A15"]="Metric"; fs["B15"]=2024; fs["C15"]=2025
    for name,r in {"Cash & Cash Equivalents":16,"Long-Term Debt":17,"Stockholders' Equity":18}.items(): fs.cell(r,1,name)
    fs["B16"]=9.906; fs["C16"]=14.161; fs["B17"]=8.169; fs["C17"]=8.173; fs["B18"]=23.622; fs["C18"]=29.164
    fs["A21"]="Cash Flow Statement"
    return wb


def test_costco_like_business_quality_not_destroyed_by_low_margin():
    wb=_costco_like_workbook()
    b=compute_score_bundle(wb,"COST",base_value=483.27,severe_value=0,current_price=945.01,mc_prob=.183)
    q=b["category_scores"]["Business Quality"]
    v=b["category_scores"]["Valuation / Stock Attractiveness"]
    assert q is not None and q>65, q
    assert b["dimensions"]["Profitability"]["score"]>60
    assert b["dimensions"]["FCF Quality"]["score"]>65
    assert v is not None and v<25, v
    assert q>v+35, (q,v)
    assert b["dimensions"]["Stress Robustness"]["score"]>=0


if __name__=="__main__":
    test_costco_like_business_quality_not_destroyed_by_low_margin()
    print("score-engine-v2 tests passed")
