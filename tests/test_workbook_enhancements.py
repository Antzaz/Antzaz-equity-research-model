from openpyxl import Workbook

from workbook_enhancements import (
    _patch_leadership_sheet,
    _polish_workbook,
    _write_public_evidence_quality,
)


def _sample_workbook():
    wb = Workbook()
    wb.active.title = "Company Data"
    wb.create_sheet("Filings")
    wb.create_sheet("Dashboard")
    wb.create_sheet("Investment Summary")
    wb.create_sheet("Decision View")
    leadership = wb.create_sheet("Leadership & Culture")
    quality = wb.create_sheet("Data Quality")

    leadership["A5"] = "Worker Happiness / Employee Experience Evidence"
    headers = ["Metric", "Value", "Scope", "Period", "Status", "Source"]
    for c, value in enumerate(headers, 1):
        leadership.cell(6, c, value)
    leadership["A7"] = "Worker happiness / satisfaction signal"
    leadership["B7"] = None
    leadership["C7"] = "No comparable company-wide employee happiness/engagement score was automatically verified"
    leadership["E7"] = "REVIEW"
    leadership["A8"] = "Evidence"
    leadership["B8"] = "Use issuer human-capital disclosures first."

    quality["A1"] = "Market-share comparability"
    quality["B1"] = "REVIEW"
    quality["C1"] = "No comparable public industry-share source mapped"
    quality["A2"] = "Employee sentiment scope"
    quality["B2"] = "REVIEW"
    return wb


def test_public_workforce_evidence_replaces_avoidable_review_without_inventing_score():
    wb = _sample_workbook()
    evidence = {
        "workforce": "The annual report discusses human capital, workforce development and retention.",
        "workforce_source": "https://example.com/annual-report",
        "annual_form": "10-K",
        "annual_filing": "https://example.com/annual-report",
        "governance": "The filing describes the board of directors and audit committee.",
        "governance_source": "https://example.com/annual-report",
        "leadership": "The filing identifies the chief executive officer and executive officers.",
        "leadership_source": "https://example.com/annual-report",
        "headcount": 125000,
        "officer_count": 8,
    }

    _patch_leadership_sheet(wb, evidence)
    ws = wb["Leadership & Culture"]
    assert ws["A7"].value == "Workforce / employee experience evidence"
    assert ws["B7"].value == "Qualitative"
    assert ws["E7"].value == "PUBLIC-EVIDENCE"
    assert ws["F7"].value == evidence["workforce_source"]
    assert "125,000" in ws["B8"].value


def test_data_quality_uses_na_for_non_comparable_market_share_and_pass_for_public_workforce_evidence():
    wb = _sample_workbook()
    evidence = {
        "workforce": "Human-capital and retention disclosure is available.",
        "workforce_source": "https://example.com/10k",
        "annual_form": "10-K",
        "annual_filing": "https://example.com/10k",
        "governance": None,
        "governance_source": None,
        "leadership": None,
        "leadership_source": None,
        "headcount": 10000,
        "officer_count": 5,
    }

    _write_public_evidence_quality(wb, evidence)
    ws = wb["Data Quality"]
    assert ws["B1"].value == "N/A"
    assert ws["B2"].value == "PASS"
    assert any(
        str(ws.cell(r, 1).value or "") == "Public Evidence Recovery — automated, source-scoped"
        for r in range(1, ws.max_row + 1)
    )


def test_workbook_polish_frontloads_decision_sheets_and_hides_support_tabs_only():
    wb = _sample_workbook()
    wb.create_sheet("Peer Comps")
    wb.create_sheet("Advanced Analytics")

    _polish_workbook(wb)

    assert wb.sheetnames[:3] == ["Dashboard", "Investment Summary", "Decision View"]
    assert wb["Company Data"].sheet_state == "hidden"
    assert wb["Filings"].sheet_state == "hidden"
    assert wb["Data Quality"].sheet_state == "visible"
    assert wb["Leadership & Culture"].sheet_state == "visible"
    assert wb.active.title == "Dashboard"
