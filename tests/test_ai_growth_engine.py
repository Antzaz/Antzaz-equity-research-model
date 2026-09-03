from __future__ import annotations

from pathlib import Path
import sqlite3

import numpy as np
import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from machine_learning.ai_growth import (
    LightGBMGrowthForecaster,
    ai_adjustments,
    deterministic_ai_signals,
    expectations_gap,
    growth_training_frame_from_history,
    reverse_dcf_from_workbook,
)
from machine_learning.ai_growth_output import write_ai_growth_sheet


def test_deterministic_ai_signal_extraction_is_bounded():
    rows = [
        {
            "kpi": "AI revenue",
            "unit_comparison": "+60% YoY",
            "signal": "Very strong",
            "investment_read_through": "AI demand and paid enterprise adoption are accelerating.",
            "data_type": "Company reported",
        },
        {
            "kpi": "AI capex",
            "unit_comparison": "$100bn",
            "signal": "Key risk / opportunity",
            "investment_read_through": "Heavy infrastructure spending raises utilization and return hurdles.",
            "data_type": "Management guidance",
        },
    ]
    signal = deterministic_ai_signals(rows, "AI customers increased 50%; backlog expanded; capex is elevated.")
    for value in (
        signal.demand_score, signal.monetization_score, signal.adoption_score,
        signal.efficiency_score, signal.capex_burden_score, signal.risk_score,
        signal.confidence,
    ):
        assert 0.0 <= value <= 1.0
    adjustments = ai_adjustments(signal)
    assert -0.10 <= adjustments["revenue_growth_adjustment"] <= 0.10
    assert -0.12 <= adjustments["fcf_growth_adjustment"] <= 0.12


def test_history_frame_requires_contiguous_next_fiscal_year(tmp_path: Path):
    db = tmp_path / "history.sqlite"
    with sqlite3.connect(db) as con:
        con.execute("""CREATE TABLE fundamentals(
            symbol TEXT,fiscal_year INTEGER,fiscal_date TEXT,available_date TEXT,period TEXT,
            revenue REAL,operating_income REAL,net_income REAL,fcf REAL,capex REAL,rd REAL,
            cash REAL,debt REAL,equity REAL,source TEXT
        )""")
        con.execute("""CREATE TABLE features(
            symbol TEXT,as_of TEXT,momentum_12m REAL,momentum_6m REAL,
            volatility_6m REAL,drawdown_12m REAL
        )""")
        con.executemany(
            "INSERT INTO fundamentals VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            [
                ("AAA", 2020, "2020-12-31", "2021-03-01", "annual", 100, 20, 15, 10, 5, 10, 20, 5, 80, "SEC"),
                ("AAA", 2021, "2021-12-31", "2022-03-01", "annual", 110, 22, 16, 11, 5, 11, 20, 5, 82, "SEC"),
                ("AAA", 2023, "2023-12-31", "2024-03-01", "annual", 150, 30, 22, 15, 7, 15, 25, 5, 90, "SEC"),
            ],
        )
        con.executemany(
            "INSERT INTO features VALUES(?,?,?,?,?,?)",
            [
                ("AAA", "2021-03-01", 0.1, 0.05, 0.2, -0.1),
                ("AAA", "2022-03-01", 0.1, 0.05, 0.2, -0.1),
                ("AAA", "2024-03-01", 0.1, 0.05, 0.2, -0.1),
            ],
        )
    frame = growth_training_frame_from_history(db)
    row_2020 = frame.loc[frame["as_of"] == pd.Timestamp("2021-03-01")].iloc[0]
    row_2021 = frame.loc[frame["as_of"] == pd.Timestamp("2022-03-01")].iloc[0]
    assert row_2020["target_next_revenue_growth"] == pytest.approx(0.10)
    assert pd.isna(row_2021["target_next_revenue_growth"])


def test_lightgbm_has_temporal_holdout_and_baseline():
    pytest.importorskip("lightgbm")
    rng = np.random.default_rng(7)
    n = 180
    dates = pd.date_range("2008-01-01", periods=n, freq="MS")
    frame = pd.DataFrame({"as_of": dates, "target_date": dates + pd.DateOffset(months=12)})
    features = [
        "revenue_growth", "operating_margin", "net_margin", "fcf_margin",
        "capex_to_revenue", "rd_to_revenue", "roe", "net_debt_to_revenue",
        "momentum_12m", "momentum_6m", "volatility_6m", "drawdown_12m",
    ]
    for col in features:
        frame[col] = rng.normal(0.05, 0.12, n)
    frame["target_next_revenue_growth"] = (
        0.06 + 0.45 * frame["revenue_growth"] + 0.15 * frame["rd_to_revenue"]
        - 0.10 * frame["capex_to_revenue"] + rng.normal(0, 0.03, n)
    )
    current = {c: float(frame[c].iloc[-1]) for c in features}
    result = LightGBMGrowthForecaster().fit_predict(
        frame, current,
        target_col="target_next_revenue_growth",
        target_label="Next FY revenue growth",
    )
    assert result.status == "PASS"
    assert result.prediction is not None
    assert result.metrics["holdout_rows"] >= 8
    assert "elastic_net_holdout_mae" in result.metrics
    assert result.drivers


def test_reverse_dcf_and_workbook_sheet(tmp_path: Path):
    path = tmp_path / "model.xlsx"
    wb = Workbook()
    company = wb.active
    company.title = "Company Data"
    company["B10"] = 3_000_000_000_000
    company["B14"] = -50_000_000_000
    hist = wb.create_sheet("Historical Financials")
    for col, year in enumerate(range(2020, 2026), start=2):
        hist.cell(3, col, year)
    hist["G14"] = 130_000_000_000
    hist["G15"] = -50_000_000_000
    wb.save(path)

    reverse = reverse_dcf_from_workbook(path)
    assert reverse["status"] == "PASS"
    assert 0 < reverse["implied_annual_fcf_growth"] < 0.5
    gap = expectations_gap(0.18, reverse)
    assert gap["status"] == "PASS"

    signals = deterministic_ai_signals(
        [{"kpi":"AI revenue","signal":"Very strong","investment_read_through":"Demand and monetization are improving.","data_type":"Company reported"}],
        "AI backlog and adoption are rising.",
    )
    payload = {
        "generated_at": "2026-09-03T00:00:00+00:00",
        "ai_signals": signals.to_dict(),
        "revenue_forecast": {
            "status":"PASS","prediction":0.20,"confidence":"Moderate",
            "metrics":{"time_purged_holdout_mae":0.08,"elastic_net_holdout_mae":0.09},
            "drivers":[{"feature":"revenue_growth","shap_value":0.03,"current_value":0.15,"direction":"positive"}],
        },
        "fcf_forecast": {
            "status":"PASS","prediction":0.14,"confidence":"Moderate",
            "drivers":[{"feature":"fcf_margin","shap_value":0.04,"current_value":0.18,"direction":"positive"}],
        },
        "ai_adjustments": {"revenue_growth_adjustment": 0.02, "fcf_growth_adjustment": 0.03},
        "ai_adjusted_revenue_growth": 0.22,
        "ai_adjusted_fcf_growth": 0.17,
        "reverse_dcf": reverse,
        "expectations_gap": expectations_gap(0.17, reverse),
    }
    write_ai_growth_sheet(path, "TEST", payload)
    reopened = load_workbook(path, data_only=False)
    sheet = reopened["AI Growth Forecast"]
    assert sheet["A1"].value == "TEST — AI Growth Forecast"
    assert len(sheet._charts) >= 3
    assert sheet["D23"].number_format.startswith("0.0%")
