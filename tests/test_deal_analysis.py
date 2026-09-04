from openpyxl import Workbook
import deal_analysis as da


def _item(title, summary="", date="2026-08-01", publisher="Reuters", kind="news", url="https://example.com/a"):
    return {"title": title, "summary": summary, "date": date, "publisher": publisher, "source_kind": kind, "url": url}


def test_big_acquisition_extracts_deal_value_and_materiality():
    items = [_item(
        "Alphabet agrees to acquire Wiz for $32.0 billion",
        "The definitive agreement expands Google Cloud cybersecurity and is subject to regulatory approval.",
    )]
    deals = da.analyze_items("GOOGL", items, {"longName": "Alphabet Inc.", "marketCap": 2_000e9, "currency": "USD"})
    assert len(deals) == 1
    d = deals[0]
    assert d["type"] == "Acquisition / M&A"
    assert d["amount"]["currency"] == "USD"
    assert d["amount"]["value"] == 32e9
    assert d["status"] in {"Announced / Pending", "Approved / Clearing"}
    assert d["materiality"] in {"Notable", "Material", "Major"}
    assert "Cloud / Cybersecurity" in d["strategic_area"]


def test_long_term_power_contract_without_price_is_kept():
    items = [_item(
        "Google signs long-term power purchase agreement with Fervo Energy",
        "The multi-year contract will supply 396 MW of geothermal power for data center operations.",
    )]
    deals = da.analyze_items("GOOGL", items, {"longName": "Alphabet Inc."})
    assert len(deals) == 1
    d = deals[0]
    assert d["type"] == "Major Commercial Contract"
    assert d["deal_value"] is None
    assert d["materiality"] == "Strategically material / value not normalized"
    assert "396 MW" in d["terms"]
    assert "Power / Energy Infrastructure" in d["strategic_area"]


def test_divestiture_terms_preserve_cash_note_and_equity_interest():
    items = [_item(
        "Alphabet agrees sale of GFiber",
        "Alphabet expects $1.5 billion cash, a $2.0 billion note receivable and a 49.99% equity interest when the divestiture closes.",
    )]
    deals = da.analyze_items("GOOGL", items, {"longName": "Alphabet Inc.", "marketCap": 2_000e9, "currency": "USD"})
    assert len(deals) == 1
    d = deals[0]
    assert d["type"] == "Divestiture / Asset Sale"
    assert d["amount"]["currency"] == "USD"
    assert d["amount"]["value"] in {1.5e9, 2.0e9}
    assert "Note receivable" in d["terms"]
    assert "Equity interest" in d["terms"]
    assert "49.99%" in d["terms"]


def test_small_generic_partnership_is_filtered_out():
    items = [_item(
        "Example Corp partners with Tiny Startup",
        "The companies will explore a pilot marketing collaboration.",
        publisher="Small Blog",
    )]
    assert da.analyze_items("EXM", items, {"longName": "Example Corp"}) == []


def test_duplicate_announcement_and_close_merge_to_one_record():
    items = [
        _item("Alphabet to acquire Wiz for $32 billion", "Definitive agreement for cloud cybersecurity.", date="2025-03-18"),
        _item("Alphabet completes acquisition of Wiz for $29.5 billion", "Google Cloud completed the acquisition after purchase price adjustments.", date="2026-03-11", url="https://example.com/b"),
    ]
    deals = da.analyze_items("GOOGL", items, {"longName": "Alphabet Inc.", "marketCap": 2_000e9, "currency": "USD"})
    assert len(deals) == 1
    assert deals[0]["status"] == "Completed / Closed"
    assert deals[0]["amount"]["value"] in {29.5e9, 32e9}


def test_primary_amount_prefers_transaction_context_over_unrelated_larger_number():
    text = "The company had $100 billion of annual revenue. It completed the acquisition for a purchase price of $4.8 billion in cash and assumed debt."
    assert da._amount(text)["value"] == 4.8e9


def test_workbook_sheet_has_professional_headers(monkeypatch):
    sample = {
        "date": "2026-03-11", "status": "Completed / Closed", "type": "Acquisition / M&A", "counterparty": "Wiz",
        "title": "Alphabet completes acquisition of Wiz", "amount": {"raw": "$29.5 billion", "value": 29.5e9, "currency": "USD"},
        "deal_value": "$29.5 billion", "materiality": "Major", "market_cap_ratio": 0.015, "strategic_area": "Cloud / Cybersecurity",
        "terms": "All cash", "rationale": "Strategic rationale", "financial_impact": "Financial impact", "risks": "integration / execution",
        "regulatory": "Cleared", "next_milestone": "Track integration", "evidence": "Primary filing evidence",
        "source_quality": "High — primary filing", "publisher": "SEC / EDGAR", "url": "https://sec.example/filing", "score": 90,
    }
    monkeypatch.setattr(da, "collect_deals", lambda *args, **kwargs: ([sample], {"yahoo_items": 2, "sec_items": 1, "qualified_deals": 1, "sources": "test sources"}))
    wb = Workbook(); wb.remove(wb.active)
    ws = da.ensure_deal_analysis(wb, "GOOGL", {"marketCap": 2_000e9, "currency": "USD"})
    assert ws.title == "Deals & Transactions"
    assert ws["A14"].value == "Date"
    assert ws["J14"].value == "Strategic Rationale — Analyst Lens"
    assert ws["R14"].value == "Source Link"
    assert ws["C15"].value == "Acquisition / M&A"
    assert ws["R15"].hyperlink.target == "https://sec.example/filing"
