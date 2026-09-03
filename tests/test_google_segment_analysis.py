from openpyxl import Workbook

from google_segment_analysis import ensure_google_segment_analysis


def test_google_segment_analysis_populates_verified_2023_2025_values():
    wb = Workbook()
    ws = ensure_google_segment_analysis(wb, "GOOGL")

    assert ws.title == "Segment Analysis"
    assert ws["A7"].value == "Google Services"
    assert ws["A8"].value == "Google Cloud"
    assert ws["A9"].value == "Other Bets"

    assert ws["B7"].value == 272.543
    assert ws["D7"].value == 342.721
    assert ws["G7"].value == 95.858
    assert ws["I7"].value == 139.404

    assert ws["B8"].value == 33.088
    assert ws["D8"].value == 58.705
    assert ws["I8"].value == 13.910

    assert ws["B9"].value == 1.527
    assert ws["D9"].value == 1.537
    assert ws["I9"].value == -7.515

    assert ws["A21"].value == "Google Search & other"
    assert ws["D21"].value == 224.532
    assert ws["A22"].value == "YouTube ads"
    assert ws["D22"].value == 40.367
    assert ws["A25"].value == "Google Cloud"
    assert ws["D25"].value == 58.705

    # Derived analytics remain formula-driven rather than hardcoded.
    assert str(ws["E7"].value).startswith("=IFERROR")
    assert str(ws["J7"].value).startswith("=IFERROR")
    assert "VERIFIED OFFICIAL" in str(ws["A3"].value)
    assert ws["P7"].hyperlink is not None


def test_google_segment_adapter_is_ticker_scoped():
    wb = Workbook()
    assert ensure_google_segment_analysis(wb, "MSFT") is None
