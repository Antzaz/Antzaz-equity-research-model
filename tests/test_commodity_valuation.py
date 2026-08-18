from __future__ import annotations

import statistics

from openpyxl import Workbook

from advanced_analytics_v2 import _base_value as generic_base_value
from commodity_valuation_v3 import (
    apply_commodity_normalization,
    commodity_base_value,
    commodity_monte_carlo,
    equity_fcf_value,
    is_commodity_workbook,
)


def _cvx_fixture():
    wb=Workbook(); wb.remove(wb.active)
    cd=wb.create_sheet('Company Data'); hf=wb.create_sheet('Historical Financials'); sc=wb.create_sheet('Three-Case Scenarios')
    cd['B4']='CVX'; cd['B6']='Energy'; cd['B7']='Oil & Gas Integrated'; cd['B8']=197.70
    cd['B9']=1.961603364; cd['B10']=387.809; cd['B12']=8.530; cd['B13']=37.075; cd['B14']=28.545
    years=[2020,2021,2022,2023,2024,2025]
    rev=[94.471,155.606,235.717,196.913,193.414,184.432]
    op=[-6.942,16.104,39.950,26.229,18.917,16.674]
    ocf=[10.577,29.187,49.618,35.609,31.493,33.939]
    cap=[8.922,8.654,12.314,15.829,16.375,17.347]
    da=[17.181,17.251,18.990,19.207,20.127,20.132]
    for i,y in enumerate(years,2):
        hf.cell(3,i,y); hf.cell(4,i,rev[i-2]); hf.cell(9,i,op[i-2])
        hf.cell(14,i,ocf[i-2]); hf.cell(15,i,cap[i-2]); hf.cell(18,i,da[i-2]); hf.cell(21,i,0)
    # Reproduce the old over-aggressive generic scenario from the uploaded Chevron workbook.
    generic_growth=[.2384,-.0848,.1236,.1138,.1040,.0942,.0844,.0746,.0648,.0550]
    for block in (range(2,12),range(14,24),range(26,36)):
        for c,g in zip(block,generic_growth):
            sc.cell(12,c,g); sc.cell(14,c,.22); sc.cell(18,c,.10916); sc.cell(20,c,.09406)
    sc['C6']=.06426; sc['C7']=.03; sc['C8']=.21; sc['AL4']=-.0038
    return wb


def test_cvx_v3_replaces_extreme_generic_dcf_with_triangulation():
    wb=_cvx_fixture()
    assert is_commodity_workbook(wb)
    before=generic_base_value(wb)
    meta=apply_commodity_normalization(wb,'CVX',{})
    operating=generic_base_value(wb)
    direct=equity_fcf_value(wb,'base')
    primary=commodity_base_value(wb)
    sc=wb['Three-Case Scenarios']

    assert meta['applied'] is True
    assert 'Commodity Valuation' in wb.sheetnames
    assert before>500, before

    # v2 normalization remains active.
    assert float(sc['C6'].value)>=.0799, sc['C6'].value
    assert abs(float(sc['C7'].value)-.015)<1e-12, sc['C7'].value
    assert abs(float(sc['N12'].value)-.15)<1e-12, sc['N12'].value
    assert abs(float(sc['O12'].value)+.08)<1e-12, sc['O12'].value
    assert float(sc['W12'].value)<=.0201, sc['W12'].value
    assert float(sc['N14'].value)<.19, sc['N14'].value

    # Independent methods must both be plausible and the primary value is their disclosed median.
    assert 75<operating<250, operating
    assert 140<direct<240, direct
    assert abs(primary-statistics.median([operating,direct]))<1e-9
    assert 110<primary<220, primary
    assert primary<before*.45, (before,primary)
    assert primary/197.70-1<.25, primary

    meta=getattr(wb,'_commodity_valuation')
    assert abs(meta['normalized_equity_fcf_start']-16.592)<.15, meta['normalized_equity_fcf_start']
    assert meta['base_equity_discount_rate']>=.0849
    assert meta['method_divergence'] is not None
    text=' '.join(str(wb['Commodity Valuation'].cell(r,c).value or '') for r in range(1,wb['Commodity Valuation'].max_row+1) for c in range(1,min(8,wb['Commodity Valuation'].max_column)+1))
    assert 'Authoritative Commodity Valuation' in text
    assert 'Normalized Equity FCF DCF' in text
    assert 'No net-debt subtraction' in text
    assert 'Calculated CAPM/WACC before overlay' in text


def test_cvx_commodity_monte_carlo_uses_both_methods():
    wb=_cvx_fixture()
    apply_commodity_normalization(wb,'CVX',{})
    values=commodity_monte_carlo(wb,n=500,seed=7)
    assert len(values)==500
    assert values==sorted(values)
    median=values[len(values)//2]
    assert 80<median<280, median


def test_noncommodity_company_is_untouched():
    wb=_cvx_fixture(); wb['Company Data']['B4']='GOOGL'
    wb['Company Data']['B6']='Communication Services'; wb['Company Data']['B7']='Internet Content & Information'
    original_tgr=wb['Three-Case Scenarios']['C7'].value
    original_value=generic_base_value(wb)
    meta=apply_commodity_normalization(wb,'GOOGL',{})
    assert meta['applied'] is False
    assert wb['Three-Case Scenarios']['C7'].value==original_tgr
    assert 'Commodity Valuation' not in wb.sheetnames
    assert abs(commodity_base_value(wb)-original_value)<1e-9


if __name__=='__main__':
    test_cvx_v3_replaces_extreme_generic_dcf_with_triangulation()
    test_cvx_commodity_monte_carlo_uses_both_methods()
    test_noncommodity_company_is_untouched()
    print('commodity valuation v3 tests passed')
