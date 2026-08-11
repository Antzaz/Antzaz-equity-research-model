from __future__ import annotations

from datetime import datetime, timezone

from openpyxl import Workbook

import data_integrity


def test_corrupt_foreign_history_is_replaced_and_future_fy_removed():
    original=data_integrity.yahoo_reference_history
    try:
        data_integrity.yahoo_reference_history=lambda ticker: {
            2024:{"revenue":75.9e9,"ni":9.0e9,"eps":11.15,"ocf":12.0e9,"capex":2.1e9},
            2025:{"revenue":78.9e9,"ni":10.4e9,"eps":12.25,"ocf":13.2e9,"capex":2.4e9},
        }
        info={
            "lastFiscalYearEnd":int(datetime(2025,9,30,tzinfo=timezone.utc).timestamp()),
            "totalRevenue":78.9e9,
            "currency":"EUR","financialCurrency":"EUR",
        }
        bad={
            2024:{"revenue":75.9e9,"ni":9.0e9,"eps":11.15},
            2025:{"revenue":2.025193772e18,"ni":2.02522432e17,"eps":2.025202532e9},
            2026:{"revenue":2.026207942e18,"ni":2.02625832e17,"eps":2.026202632e9},
        }
        hist,meta=data_integrity.sanitize_crossborder_history("SIE.DE",info,bad,{})
        assert 2026 not in hist
        assert abs(hist[2025]["revenue"]-78.9e9)<1
        assert abs(hist[2025]["ni"]-10.4e9)<1
        assert abs(hist[2025]["eps"]-12.25)<1e-9
        assert meta["last_completed_fiscal_year"]==2025
        assert meta["integrity_warnings"]
    finally:
        data_integrity.yahoo_reference_history=original


def test_template_sources_and_currency_labels_are_cleaned():
    wb=Workbook(); ws=wb.active; ws.title="Historical Financials"
    wb.create_sheet("Filings"); wb.create_sheet("Company Data"); wb.create_sheet("Data Quality")
    ws["A2"]="$ in billions except per-share data"
    ws["A30"]="https://www.sec.gov/Archives/edgar/data/1652044/goog-20251231.htm"
    wb["Filings"]["D4"]="https://www.sec.gov/Archives/edgar/data/1652044/goog-20251231.htm"
    data_integrity.apply_workbook_integrity_controls(
        wb,"SIE.DE",{"currency":"EUR","financialCurrency":"EUR","lastFiscalYearEnd":datetime(2025,9,30)},
        {"last_completed_fiscal_year":2025,"integrity_warnings":[]},
    )
    assert wb["Historical Financials"]["A30"].value is None
    assert wb["Filings"]["D4"].value is None
    assert wb["Historical Financials"]["A2"].value.startswith("EUR")
    assert wb["Company Data"]["A10"].value=="Market Cap (EUR bn)"


if __name__=="__main__":
    test_corrupt_foreign_history_is_replaced_and_future_fy_removed()
    test_template_sources_and_currency_labels_are_cleaned()
    print("data-integrity tests passed")
