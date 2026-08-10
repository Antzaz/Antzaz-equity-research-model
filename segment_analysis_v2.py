"""Company-agnostic SEC segment analysis with resilient HTML + narrative parsing.

The module extracts issuer-disclosed operating-segment and business-line data from the
latest 10-K. It uses lxml table parsing first, pandas as a fallback, and also reads narrative
filing text for statements such as "organized into the following six segments: ...".

Reliability rule: a disclosed segment name is valuable even when reliable segment financials
cannot be extracted. Every discovered/configured reportable segment is therefore preserved in
the worksheet and missing financials remain blank/yellow rather than causing the segment to
disappear or being estimated.
"""

import re
import statistics
from io import StringIO

import requests
try:
    import pandas as pd
except Exception:
    pd = None
try:
    from lxml import html as lxml_html
except Exception:
    lxml_html = None

from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; GOLD="FFF2CC"
INPUT_BLUE="0000FF"; GREY="666666"; LINK_GREEN="008000"
FMT_BN='#,##0.0;[Red](#,##0.0);-'; FMT_PCT='0.0%;[Red](0.0%);-'

CONFIGS={
    "GOOGL":{"segments":["Google Services","Google Cloud","Other Bets"],"business":["Google Search & other","YouTube ads","Google Network","Google subscriptions, platforms, and devices","Google Cloud","Other Bets"]},
    "GOOG":{"segments":["Google Services","Google Cloud","Other Bets"],"business":["Google Search & other","YouTube ads","Google Network","Google subscriptions, platforms, and devices","Google Cloud","Other Bets"]},
    "MSFT":{"segments":["Productivity and Business Processes","Intelligent Cloud","More Personal Computing"]},
    "AMZN":{"segments":["North America","International","AWS"],"business":["Online stores","Physical stores","Third-party seller services","Advertising services","Subscription services","AWS","Other"]},
    "META":{"segments":["Family of Apps","Reality Labs"]},
    "NVDA":{"segments":["Compute & Networking","Graphics"]},
    "AAPL":{"business":["iPhone","Mac","iPad","Wearables, Home and Accessories","Services"]},
    "NFLX":{"segments":["United States and Canada","Europe, Middle East and Africa","Latin America","Asia-Pacific"]},
    "ORCL":{"business":["Cloud services and license support","Cloud license and on-premise license","Hardware","Services"]},
    "CRM":{"business":["Sales","Service","Platform and Other","Marketing and Commerce","Integration and Analytics"]},
    "ADBE":{"segments":["Digital Media","Digital Experience","Publishing and Advertising"]},
    "AMD":{"segments":["Data Center","Client","Gaming","Embedded"]},
    "INTC":{"segments":["Client Computing Group","Data Center and AI","Network and Edge","Mobileye"]},
    "QCOM":{"segments":["QCT","QTL"]},
    "WMT":{"segments":["Walmart U.S.","Walmart International","Sam's Club"]},
    "MET":{"segments":["Group Benefits","RIS","Asia","Latin America","EMEA","MIM"]},
}

BLACKLIST=("total","consolidated","revenue","net sales","operating income","operating loss","segment profit","year ended","three months","six months","nine months","cost of","income from","eliminations","corporate","other income","depreciation","assets","liabilities","capital expenditures","segment information","reportable segment")


def _fill(c): return PatternFill("solid",fgColor=c)
def _num(v):
    try:
        if isinstance(v,bool): return None
        return float(v)
    except Exception:
        return None
def _label_key(text):
    s=_clean_text(text).replace("“",'"').replace("”",'"').strip()
    m=re.search(r"\(([A-Z]{2,10})\)\s*$",s)
    if m: return m.group(1).lower()
    if re.fullmatch(r"[A-Z]{2,10}",s): return s.lower()
    return re.sub(r"[^a-z0-9]","",s.lower())

def _title(ws,text,end=16):
    for c in range(1,end+1):
        ws.cell(1,c).fill=_fill(NAVY); ws.cell(2,c).fill=_fill(NAVY)
    ws["A1"]=text; ws["A1"].font=Font(bold=True,color=WHITE,size=18); ws.sheet_view.showGridLines=False

def _section(ws,row,title,end=16):
    for c in range(1,end+1):
        ws.cell(row,c).fill=_fill(NAVY); ws.cell(row,c).font=Font(bold=True,color=WHITE,size=11)
    ws.cell(row,1,title)

def _header(ws,row,start,end):
    for c in range(start,end+1):
        ws.cell(row,c).fill=_fill(BLUE); ws.cell(row,c).font=Font(bold=True,color=WHITE)
        ws.cell(row,c).alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

def _latest_10k_html(ticker,headers):
    try:
        tickers=requests.get("https://www.sec.gov/files/company_tickers.json",headers=headers,timeout=30).json(); cik=None
        for item in tickers.values():
            if str(item.get("ticker","")).upper()==ticker.upper():
                cik=str(item["cik_str"]).zfill(10); break
        if not cik: return None,None
        subs=requests.get(f"https://data.sec.gov/submissions/CIK{cik}.json",headers=headers,timeout=30).json()
        recent=subs.get("filings",{}).get("recent",{})
        for form,acc,doc in zip(recent.get("form",[]),recent.get("accessionNumber",[]),recent.get("primaryDocument",[])):
            if form=="10-K":
                url=f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{acc.replace('-','')}/{doc}"
                r=requests.get(url,headers=headers,timeout=45); r.raise_for_status(); return r.text,url
    except Exception:
        return None,None
    return None,None

def _clean_text(value): return re.sub(r"\s+"," ",str(value or "").replace("\xa0"," ")).strip()
def _document_text(raw_html):
    if not raw_html: return ""
    if lxml_html is not None:
        try:
            root=lxml_html.fromstring(raw_html)
            for node in root.xpath("//script|//style|//noscript"):
                try: node.drop_tree()
                except Exception: pass
            return _clean_text(" ".join(root.itertext()))
        except Exception: pass
    return _clean_text(re.sub(r"<[^>]+>"," ",raw_html))
def _lxml_tables(raw_html):
    if not raw_html or lxml_html is None: return []
    try:
        root=lxml_html.fromstring(raw_html); out=[]
        for table in root.xpath("//table"):
            rows=[]
            for tr in table.xpath(".//tr"):
                cells=tr.xpath("./th|./td")
                if not cells: continue
                vals=[_clean_text(" ".join(cell.itertext())) for cell in cells]
                if any(vals): rows.append(vals)
            if len(rows)>=2: out.append(rows)
        return out
    except Exception: return []
def _pandas_tables(raw_html):
    if not raw_html or pd is None: return []
    try:
        dfs=pd.read_html(StringIO(raw_html)); out=[]
        for df in dfs:
            rows=[]; cols=[_clean_text(x) for x in getattr(df,"columns",[])]
            if any(cols): rows.append(cols)
            for _,row in df.iterrows(): rows.append([_clean_text(v) for v in row.tolist()])
            if len(rows)>=2: out.append(rows)
        return out
    except Exception: return []
def _tables(raw_html):
    rows=_lxml_tables(raw_html)
    if rows: return rows,f"lxml ({len(rows)} tables)"
    rows=_pandas_tables(raw_html)
    if rows: return rows,f"pandas ({len(rows)} tables)"
    if lxml_html is None and pd is None: return [],"no HTML parser available"
    return [],"HTML table parsing failed"
def _row_text(values): return _clean_text(" | ".join(_clean_text(v) for v in values))
def _numbers(values):
    out=[]
    for val in values:
        text=_clean_text(val).replace("−","-").replace("—","")
        for token in re.findall(r"\(?-?\$?\s*\d[\d,]*(?:\.\d+)?\)?",text):
            raw=token.strip(); neg=raw.startswith("(") and raw.endswith(")")
            clean=raw.replace("$","").replace(",","").replace(" ","").strip("()")
            try: n=float(clean)
            except Exception: continue
            if 1900<=n<=2100: continue
            out.append(-n if neg else n)
    return out
def _text_name(values):
    for val in values:
        s=_clean_text(val)
        if s and s.lower() not in {"nan","none"} and re.search(r"[A-Za-z]",s) and len(s)<=100: return s
    return None
def _row_metric(text):
    low=text.lower()
    if any(k in low for k in ("operating income","operating loss","income (loss) from operations","segment profit","segment income","profit (loss)","adjusted operating income","adjusted earnings","segment ebitda","adjusted ebitda","operating profit")): return "op"
    if any(k in low for k in ("net sales","revenue","revenues","external sales","sales to external customers")): return "revenue"
    return None
def _metric_context(rows,row_idx): return _row_metric(" ".join(_row_text(rows[rr]) for rr in range(max(0,row_idx-12),row_idx+1)))
def _normalize_label_text(text):
    s=_clean_text(text).strip(":"); s=re.sub(r"\s*\(\d+\)\s*$","",s); return s.strip()
def _find_label(values,labels):
    cells=[_normalize_label_text(v) for v in values if _clean_text(v)]
    if not cells: return None
    first=cells[0].lower(); first_key=_label_key(cells[0])
    exact=[lab for lab in labels if first==lab.lower() or (first_key and first_key==_label_key(lab))]
    if exact: return max(exact,key=len)
    pref=[lab for lab in labels if re.match(rf"^[^A-Za-z0-9]*{re.escape(lab.lower())}(?:\s|$)",first)]
    return max(pref,key=len) if pref else None
def _extract_known(tables,labels):
    out={lab:{"revenue":[],"op":[]} for lab in labels}
    for rows in tables:
        current_label=None
        for pos,values in enumerate(rows):
            text=_row_text(values); nums=_numbers(values); lab=_find_label(values,labels)
            if lab and len(nums)<2: current_label=lab; continue
            if lab and len(nums)>=2:
                metric=_row_metric(text) or _metric_context(rows,pos)
                if metric: out[lab][metric].append(nums[-3:])
            if current_label and len(nums)>=2:
                metric=_row_metric(text)
                if metric: out[current_label][metric].append(nums[-3:])
            low=text.lower()
            if current_label and any(x in low for x in ("consolidated","total reportable segments","corporate and other")): current_label=None
    return out
def _table_segment_candidates(tables):
    candidates={}
    for rows in tables:
        whole=" ".join(_row_text(r) for r in rows).lower()
        if "segment" not in whole: continue
        for pos,vals in enumerate(rows):
            name=_text_name(vals); nums=_numbers(vals)
            if not name: continue
            low=name.lower()
            if any(b in low for b in BLACKLIST) or len(low)<2: continue
            if len(nums)>=2 and _metric_context(rows,pos) in ("revenue","op"):
                candidates.setdefault(_normalize_label_text(name),0); candidates[_normalize_label_text(name)]+=3; continue
            if len(nums)<2 and len(name)<=70:
                nearby=" ".join(_row_text(rows[j]) for j in range(pos+1,min(len(rows),pos+5))).lower()
                if any(k in nearby for k in ("revenue","operating income","segment profit","adjusted earnings","segment ebitda")):
                    candidates.setdefault(_normalize_label_text(name),0); candidates[_normalize_label_text(name)]+=1
    return [name for name,_ in sorted(candidates.items(),key=lambda kv:(-kv[1],len(kv[0])))[:15]]
def _parse_list(text):
    s=_clean_text(text).replace("“",'"').replace("”",'"'); s=re.split(r"\.\s+(?:In addition|See |These |The Company|We )",s,maxsplit=1)[0]
    if ";" in s: raw=[x.strip() for x in s.split(";")]
    else: raw=[x.strip() for x in re.split(r",\s+(?=[A-Z][A-Za-z0-9&' -]{1,45}(?:,| and |$))",s)]
    out=[]
    for x in raw:
        x=re.sub(r"^(?:and|the)\s+","",x,flags=re.I).strip(" .:,-")
        if not x or len(x)>90 or not re.search(r"[A-Za-z]",x): continue
        low=x.lower()
        if any(b==low or low.startswith(b+" ") for b in BLACKLIST): continue
        if x not in out: out.append(x)
    return out if 2<=len(out)<=15 else []
def _narrative_segment_candidates(raw_html):
    text=_document_text(raw_html)
    if not text: return []
    patterns=[
        r"(?:is|are|was|were)\s+organized\s+into\s+(?:the\s+following\s+)?(?:six|seven|eight|nine|ten|five|four|three|two|\d+)?\s*(?:reportable\s+|operating\s+)?segments?\s*:\s*(.{8,650}?)\.(?:\s+In addition|\s+See |\s+The |\s+We |$)",
        r"(?:reportable|operating)\s+segments?\s+(?:are|include|consist of)\s*:?\s*(.{8,650}?)\.(?:\s+In addition|\s+See |\s+The |\s+We |$)",
        r"(?:business|company)\s+(?:is|are)\s+(?:managed|organized)\s+(?:through|in|into)\s+(?:the\s+following\s+)?(?:six|seven|eight|nine|ten|five|four|three|two|\d+)?\s*segments?\s*:\s*(.{8,650}?)\.(?:\s+In addition|\s+See |\s+The |\s+We |$)",
    ]
    best=[]
    for pattern in patterns:
        for m in re.finditer(pattern,text,flags=re.I):
            vals=_parse_list(m.group(1))
            if len(vals)>len(best): best=vals
    return best
def _discover_segments(tables,raw_html=None):
    table_names=_table_segment_candidates(tables); narrative=_narrative_segment_candidates(raw_html); out=[]; keys=set()
    for name in narrative+table_names:
        key=_label_key(name)
        if key and key not in keys: out.append(name); keys.add(key)
    return out[:15]
def _pick(series):
    if not series: return None
    vals=max(series,key=lambda x:(len(x),abs(x[-1]) if x else 0)); vals=list(vals[-3:])
    while len(vals)<3: vals.insert(0,None)
    return vals
def _years(wb):
    h=wb["Historical Financials"]; ys=[h.cell(3,c).value for c in range(2,8)]; ys=[int(x) for x in ys if isinstance(x,(int,float))]
    if len(ys)>=3: return ys[-3:]
    latest=max(ys) if ys else 2025; return [latest-2,latest-1,latest]
def _company_revenue(wb): return _num(wb["Historical Financials"]["G4"].value) if "Historical Financials" in wb.sheetnames else None

def _write_sheet(wb,ticker,years,segments,business,url,auto_status):
    if "Segment Analysis" in wb.sheetnames: wb.remove(wb["Segment Analysis"])
    ws=wb.create_sheet("Segment Analysis"); _title(ws,f"{ticker} — Business & Segment Analysis")
    ws["A3"]=(f"Standardized segment schema. Status: {auto_status}. All reportable segment names discovered in the filing are retained even when reliable numbers are unavailable. Blank yellow financial cells mean disclosed economics were not reliably extracted; they are not estimates.")
    ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True); ws.row_dimensions[3].height=42
    y0,y1,y2=years; _section(ws,5,"Reported Operating / Reportable Segments")
    heads=["Segment",f"{y0} Revenue",f"{y1} Revenue",f"{y2} Revenue",f"{y2} Growth",f"{y0}–{y2} CAGR",f"{y0} Segment Profit",f"{y1} Segment Profit",f"{y2} Segment Profit",f"{y0} Margin",f"{y1} Margin",f"{y2} Margin","Margin Δ",f"{y2} Revenue Mix","Data Status","Source / Notes"]
    for c,v in enumerate(heads,1): ws.cell(6,c,v)
    _header(ws,6,1,16); max_rows=max(10,len(segments))
    for i in range(max_rows):
        r=7+i; item=segments[i] if i<len(segments) else None
        if item:
            name,rev,op=item; rev=(rev or [None,None,None]); op=(op or [None,None,None]); ws.cell(r,1,name)
            for j,v in enumerate(rev,2):
                ws.cell(r,j,v); ws.cell(r,j).number_format=FMT_BN
                if v is None: ws.cell(r,j).fill=_fill(GOLD); ws.cell(r,j).font=Font(color=INPUT_BLUE)
            ws.cell(r,5,f'=IFERROR(D{r}/C{r}-1,"")'); ws.cell(r,6,f'=IFERROR((D{r}/B{r})^(1/2)-1,"")')
            for j,v in enumerate(op,7):
                ws.cell(r,j,v); ws.cell(r,j).number_format=FMT_BN
                if v is None: ws.cell(r,j).fill=_fill(GOLD); ws.cell(r,j).font=Font(color=INPUT_BLUE)
            for c_rev,c_op,c_margin in ((2,7,10),(3,8,11),(4,9,12)):
                ws.cell(r,c_margin,f'=IFERROR({get_column_letter(c_op)}{r}/{get_column_letter(c_rev)}{r},"")'); ws.cell(r,c_margin).number_format=FMT_PCT
            ws.cell(r,13,f'=IFERROR(L{r}-K{r},"")'); ws.cell(r,14,f'=IFERROR(D{r}/SUM($D$7:$D${6+max_rows}),"")')
            for c in (5,6,10,11,12,13,14): ws.cell(r,c).number_format=FMT_PCT
            have_rev=any(_num(v) is not None for v in rev); have_op=any(_num(v) is not None for v in op)
            ws.cell(r,15,"Revenue + profitability" if have_rev and have_op else ("Revenue only" if have_rev else ("Profitability only" if have_op else "Name disclosed; numbers unavailable")))
            ws.cell(r,16,"SEC 10-K"); ws.cell(r,16).font=Font(color=LINK_GREEN)
        else:
            for c in range(1,10): ws.cell(r,c).fill=_fill(GOLD); ws.cell(r,c).font=Font(color=INPUT_BLUE); ws.cell(r,c).number_format=FMT_BN if c>=2 else "General"
            for c in (5,6,10,11,12,13,14): ws.cell(r,c).number_format=FMT_PCT
    ws.conditional_formatting.add(f"E7:F{6+max_rows}",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="num",mid_value=0,mid_color="FFEB84",end_type="max",end_color="63BE7B"))
    start=9+max_rows; _section(ws,start,"Revenue by Business Line / Product Group",10)
    bh=["Business Line / Revenue Group",str(y0),str(y1),str(y2),f"{y2} Growth",f"{y0}–{y2} CAGR",f"{y2} Mix","Source / Notes"]
    for c,v in enumerate(bh,1): ws.cell(start+1,c,v)
    _header(ws,start+1,1,8); bmax=max(10,len(business))
    for i in range(bmax):
        r=start+2+i; item=business[i] if i<len(business) else None
        if item:
            name,rev=item; rev=rev or [None,None,None]; ws.cell(r,1,name)
            for j,v in enumerate(rev,2):
                ws.cell(r,j,v); ws.cell(r,j).number_format=FMT_BN
                if v is None: ws.cell(r,j).fill=_fill(GOLD); ws.cell(r,j).font=Font(color=INPUT_BLUE)
            ws.cell(r,5,f'=IFERROR(D{r}/C{r}-1,"")'); ws.cell(r,6,f'=IFERROR((D{r}/B{r})^(1/2)-1,"")'); ws.cell(r,7,f'=IFERROR(D{r}/SUM($D${start+2}:$D${start+1+bmax}),"")'); ws.cell(r,8,"SEC 10-K"); ws.cell(r,8).font=Font(color=LINK_GREEN)
            for c in (5,6,7): ws.cell(r,c).number_format=FMT_PCT
        else:
            for c in range(1,9): ws.cell(r,c).fill=_fill(GOLD); ws.cell(r,c).font=Font(color=INPUT_BLUE); ws.cell(r,c).number_format=FMT_BN if c in (2,3,4) else "General"
            for c in (5,6,7): ws.cell(r,c).number_format=FMT_PCT
    source_row=start+bmax+4; _section(ws,source_row,"Source & Data Quality",10)
    ws.cell(source_row+1,1,"SEC 10-K Source"); ws.cell(source_row+1,2,url or ""); ws.cell(source_row+1,2).font=Font(color=LINK_GREEN)
    ws.cell(source_row+2,1,"Extraction Status"); ws.cell(source_row+2,2,auto_status)
    ws.cell(source_row+3,1,"Important"); ws.cell(source_row+3,2,"Segment disclosure differs by issuer. Segment Profit can represent operating income, adjusted operating income, adjusted earnings, segment EBITDA or another issuer-defined CODM measure. The model preserves the issuer disclosure and leaves missing economics blank rather than estimating them."); ws.cell(source_row+3,2).alignment=Alignment(wrap_text=True)
    ws.column_dimensions["A"].width=38
    for c in range(2,15): ws.column_dimensions[get_column_letter(c)].width=15
    ws.column_dimensions["O"].width=27; ws.column_dimensions["P"].width=34; ws.column_dimensions["H"].width=20; ws.freeze_panes="A7"; return ws

def ensure_segment_analysis_v2(wb,ticker,headers):
    ticker=ticker.upper(); raw_html,url=_latest_10k_html(ticker,headers); tables,parser_status=_tables(raw_html); cfg=CONFIGS.get(ticker,{})
    configured=list(cfg.get("segments",[])); discovered=_discover_segments(tables,raw_html); seg_labels=[]; seg_keys=set()
    for lab in configured+discovered:
        key=_label_key(lab)
        if lab and key and key not in seg_keys: seg_labels.append(lab); seg_keys.add(key)
    bus_labels=list(dict.fromkeys(cfg.get("business",[]))); labels=list(dict.fromkeys(seg_labels+bus_labels)); extracted=_extract_known(tables,labels) if labels else {}; total=_company_revenue(wb); raw_rev={}
    for lab in labels:
        raw=_pick(extracted.get(lab,{}).get("revenue",[]))
        if raw: raw_rev[lab]=raw
    scale=1.0
    if raw_rev:
        latest=[abs(v[-1]) for v in raw_rev.values() if v and _num(v[-1]) is not None]; med=statistics.median(latest) if latest else 0
        if total and med>total*10: scale=.001
        elif med>10000: scale=.001
    segments=[]
    for lab in seg_labels:
        rev=raw_rev.get(lab); rev=[_num(x)*scale if _num(x) is not None else None for x in rev] if rev else [None,None,None]
        op_raw=_pick(extracted.get(lab,{}).get("op",[])); op=[_num(x)*scale if _num(x) is not None else None for x in op_raw] if op_raw else [None,None,None]
        segments.append((lab,rev,op))
    business=[]
    for lab in bus_labels:
        rev=raw_rev.get(lab); vals=[_num(x)*scale if _num(x) is not None else None for x in rev] if rev else [None,None,None]; business.append((lab,vals))
    numeric_segments=sum(any(_num(v) is not None for v in (rev+op)) for _,rev,op in segments); narrative_count=len(discovered)
    if seg_labels: status=f"AUTO/BEST EFFORT — {len(seg_labels)} segment name(s) retained, {numeric_segments} with extracted financials; filing discovery={narrative_count}; parser={parser_status}"
    elif business: status=f"AUTO/BEST EFFORT — no reportable segment names identified, {len(business)} business/revenue group(s) retained; parser={parser_status}"
    else: status=f"MANUAL REQUIRED — no reliable segment names or tables extracted; parser={parser_status}"
    return _write_sheet(wb,ticker,_years(wb),segments,business,url,status)
