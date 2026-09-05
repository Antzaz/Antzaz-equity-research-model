"""Final idempotent chart-layout pass for generated equity-research workbooks.

Several model layers create charts at different stages.  Running a partial clean-up before the
ML/AI layers can therefore leave duplicated anchors or stacked charts in the final workbook.
This module is deliberately late and destructive only to chart objects on presentation sheets:
it keeps all source tables/formulas, clears those sheets' charts, and rebuilds one clean canonical
layout with short category labels and fixed anchors.

The pass is safe to run repeatedly.  One anchor always has one chart.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

GREY="666666"
FMT_BN='#,##0.0;[Red](#,##0.0);-'
FMT_PCT='0.0%;[Red](0.0%);-'
FMT_PRICE='$#,##0.00;[Red]($#,##0.00);-'


def _num(v):
    try:
        if isinstance(v,bool) or v in (None,""): return None
        return float(v)
    except Exception:
        return None


def _short_label(label,max_len=27):
    text=str(label or "").strip()
    aliases={
        "Google Search & other":"Search",
        "YouTube ads":"YouTube",
        "Google Network":"Network",
        "Google subscriptions, platforms, and devices":"Subs / Platforms / Devices",
        "Google Cloud":"Cloud",
        "Google Services":"Services",
        "Other Bets":"Other Bets",
        "Revenue Growth":"Growth",
        "Operating Margin":"Margin",
        "EBIT Margin":"Margin",
        "Terminal Growth":"TGR",
        "Combined Severe Bear":"Severe Bear",
        "Probability Weighted":"Prob.-Weighted",
    }
    if text in aliases: return aliases[text]
    text=re.sub(r"\s*\([^)]{8,}\)\s*$","",text).strip()
    for old,new in aliases.items(): text=text.replace(old,new)
    text=re.sub(r"\s+"," ",text)
    return text if len(text)<=max_len else text[:max_len-1].rstrip()+"…"


def _set_text_categories(chart,ws,col,first,last,titles=None):
    """Force text-axis categories as strRef so Excel does not reinterpret text as numeric data."""
    sheet_name=ws.title.replace("'","''")
    letter=get_column_letter(col)
    formula=f"'{sheet_name}'!${letter}${first}:${letter}${last}"
    for i,series in enumerate(chart.series):
        series.cat=AxDataSource(strRef=StrRef(f=formula))
        if titles and i<len(titles):
            series.tx=SeriesLabel(v=str(titles[i]))


def _base_chart(chart,title,width=13.5,height=8.2):
    chart.style=10; chart.title=title; chart.width=width; chart.height=height
    chart.visible_cells_only=False; chart.display_blanks="gap"
    try: chart.legend.position="b"
    except Exception: pass
    return chart


def _section_row(ws,label):
    needle=str(label).strip().lower()
    for r in range(1,ws.max_row+1):
        if str(ws.cell(r,1).value or "").strip().lower()==needle: return r
    return None


def _ensure_analysis_helpers(wb,ws):
    # Short stress-test categories.
    labels=["Base","Growth -3ppt","Growth -5ppt","Margin -300bps","Margin -500bps","Capex +500bps","WACC +100bps","WACC +200bps","TGR -50bps","TGR -100bps","Severe Bear"]
    for r,label in enumerate(labels,3): ws.cell(r,30,label)

    # Clean business labels while preserving values in AN.
    for r,label in enumerate(["Search","YouTube","Network","Subs / Platforms / Devices","Cloud","Other Bets"],3):
        ws.cell(r,39,label)

    # Rebuild a comparable-scale segment-margin helper.  Other Bets remains in Segment Analysis
    # because its very large negative margin would flatten the useful 20-40% comparison.
    for r in range(2,10):
        ws.cell(r,42).value=None; ws.cell(r,43).value=None
    ws["AP2"]="Segment"; ws["AQ2"]="2025 Margin"
    if "Segment Analysis" in wb.sheetnames:
        seg=wb["Segment Analysis"]
        source={str(seg.cell(r,1).value or "").strip():r for r in range(1,seg.max_row+1)}
        for out_r,name in ((3,"Google Services"),(4,"Google Cloud")):
            src=source.get(name)
            if src:
                ws.cell(out_r,42,name)
                ws.cell(out_r,43,f"='Segment Analysis'!L{src}")
                ws.cell(out_r,43).number_format=FMT_PCT

    # History helper, independent of any earlier chart-writer implementation.
    if "Historical Financials" in wb.sheetnames:
        h=wb["Historical Financials"]
        ws["AX2"]="Year"; ws["AY2"]="Revenue"; ws["AZ2"]="FCF"
        for out_r,src_col in enumerate(range(2,8),3):
            letter=get_column_letter(src_col)
            ws.cell(out_r,50,f"='Historical Financials'!{letter}3")
            ws.cell(out_r,51,f"='Historical Financials'!{letter}4")
            ws.cell(out_r,52,f"='Historical Financials'!{letter}16")
            ws.cell(out_r,51).number_format=FMT_BN; ws.cell(out_r,52).number_format=FMT_BN


def _clean_analysis_text(ws):
    # The old subtitle was in A3 only, causing a tall narrow column of wrapped text.
    for rng in list(ws.merged_cells.ranges):
        if str(rng)=="A3:P3": ws.unmerge_cells(str(rng))
    ws.merge_cells("A3:P3")
    ws["A3"]="Clean chart dashboard: valuation distribution, scenarios, stress tests, operating history, business mix, segment profitability and historical valuation. Detailed methodology stays in the source sheets."
    ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True,vertical="center"); ws.row_dimensions[3].height=28

    for row in (23,48,70):
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row<=row<=rng.max_row and rng.min_col<=16 and rng.max_col>=1:
                try: ws.unmerge_cells(str(rng))
                except Exception: pass
        for c in range(1,17): ws.cell(row,c).value=None
    ws.merge_cells("A23:H23"); ws.merge_cells("I23:P23")
    ws["A23"]="Monte Carlo: 5,000 valuation simulations. P10 / median / P(>price) are on Advanced Analytics."
    ws["I23"]="Scenario DCF: Bear / Base / Bull / probability-weighted value versus current price."
    ws.merge_cells("A48:H48"); ws.merge_cells("I48:P48")
    ws["A48"]="Stress tests change one driver at a time; Severe Bear combines multiple adverse assumptions."
    ws["I48"]="Revenue and FCF are reported history; FCF = operating cash flow − capex."
    ws.merge_cells("A70:H70"); ws.merge_cells("I70:P70")
    ws["A70"]="Business mix uses issuer-disclosed 2025 revenue groups only; no product revenue is estimated."
    ws["I70"]="Segment margin chart shows Services and Cloud on a comparable scale. Extreme negative segment margins remain visible on Segment Analysis rather than distorting the common chart scale."
    for row in (23,48,70):
        ws.cell(row,1).alignment=Alignment(wrap_text=True,vertical="center"); ws.cell(row,9).alignment=Alignment(wrap_text=True,vertical="center")

    # Compact reference guide.
    for r in range(102,112):
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row<=r<=rng.max_row:
                try: ws.unmerge_cells(str(rng))
                except Exception: pass
        for c in range(1,17): ws.cell(r,c).value=None
    ws.merge_cells("A102:D102"); ws.merge_cells("E102:P102"); ws["A102"]="Chart"; ws["E102"]="How to use it"
    guide=[
        ("Monte Carlo","Distribution of intrinsic-value outcomes; use P10, median and probability above market as risk context."),
        ("Scenario DCF","Compares Bear/Base/Bull and probability-weighted intrinsic value with current market price."),
        ("Stress Tests","One-driver downside sensitivities. Severe Bear combines several adverse shocks."),
        ("Revenue & FCF","Checks whether reported growth is converting into cash generation."),
        ("Revenue Mix","Issuer-disclosed 2025 revenue groups; useful for mix and concentration analysis."),
        ("Segment Margin","Comparable-scale segment profitability. Extreme outliers stay in Segment Analysis."),
        ("Historical P/E","Historical valuation context only; not a standalone signal."),
    ]
    for r,(name,note) in enumerate(guide,103):
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4); ws.merge_cells(start_row=r,start_column=5,end_row=r,end_column=16)
        ws.cell(r,1,name); ws.cell(r,5,note); ws.cell(r,5).alignment=Alignment(wrap_text=True,vertical="top"); ws.row_dimensions[r].height=24


def _rebuild_analysis_charts(wb):
    if "Analysis Charts" not in wb.sheetnames: return 0
    ws=wb["Analysis Charts"]; ws._charts=[]; _ensure_analysis_helpers(wb,ws); _clean_analysis_text(ws)

    # Monte Carlo
    ch=_base_chart(BarChart(),"Monte Carlo Value Distribution"); ch.type="col"; ch.legend=None
    ch.add_data(Reference(ws,min_col=25,min_row=3,max_row=22),titles_from_data=False)
    _set_text_categories(ch,ws,24,3,22,["Frequency"]); ch.y_axis.title="Simulation frequency"; ch.y_axis.numFmt="0"
    try: ch.x_axis.tickLblSkip=2
    except Exception: pass
    ws.add_chart(ch,"A7")

    # Scenario DCF vs market
    ch=_base_chart(BarChart(),"Scenario DCF vs Current Price"); ch.type="col"; ch.legend=None
    ch.add_data(Reference(ws,min_col=28,min_row=3,max_row=7),titles_from_data=False)
    _set_text_categories(ch,ws,27,3,7,["Value / share"]); ch.y_axis.title="USD / share"; ch.y_axis.numFmt="$0"
    ws.add_chart(ch,"I7")

    # Stress tests
    ch=_base_chart(BarChart(),"DCF Stress Tests",height=9.2); ch.type="bar"; ch.legend=None
    ch.add_data(Reference(ws,min_col=31,min_row=3,max_row=13),titles_from_data=False)
    _set_text_categories(ch,ws,30,3,13,["Value / share"]); ch.x_axis.title="USD / share"; ch.x_axis.numFmt="$0"
    ws.add_chart(ch,"A29")

    # Revenue and FCF
    ch=_base_chart(LineChart(),"Revenue & Free Cash Flow",height=9.2)
    ch.add_data(Reference(ws,min_col=51,max_col=52,min_row=2,max_row=8),titles_from_data=True)
    ch.set_categories(Reference(ws,min_col=50,min_row=3,max_row=8)); ch.y_axis.title="USD bn"; ch.y_axis.numFmt="0"
    ws.add_chart(ch,"I29")

    # Business mix
    business_last=8 if any(ws.cell(r,40).value not in (None,"") for r in range(3,9)) else 2
    if business_last>=3:
        ch=_base_chart(BarChart(),"2025 Revenue Mix"); ch.type="bar"; ch.legend=None
        ch.add_data(Reference(ws,min_col=40,min_row=3,max_row=business_last),titles_from_data=False)
        _set_text_categories(ch,ws,39,3,business_last,["Revenue"]); ch.x_axis.title="USD bn"; ch.x_axis.numFmt="0"
        ws.add_chart(ch,"A53")

    # Comparable segment margins
    if ws["AQ3"].value not in (None,""):
        ch=_base_chart(BarChart(),"2025 Segment Operating Margin"); ch.type="bar"; ch.legend=None
        ch.add_data(Reference(ws,min_col=43,min_row=3,max_row=4),titles_from_data=False)
        _set_text_categories(ch,ws,42,3,4,["Operating margin"]); ch.x_axis.title="Operating margin"; ch.x_axis.numFmt="0%"
        ws.add_chart(ch,"I53")

    # Historical P/E
    if any(ws.cell(r,48).value not in (None,"") for r in range(3,9)):
        ch=_base_chart(LineChart(),"Historical Year-End P/E"); ch.legend=None
        ch.add_data(Reference(ws,min_col=48,min_row=3,max_row=8),titles_from_data=False)
        _set_text_categories(ch,ws,47,3,8,["P/E"]); ch.y_axis.title="P/E (x)"; ch.y_axis.numFmt="0.0x"
        ws.add_chart(ch,"A75")
    return len(ws._charts)


def _rebuild_ml_charts(wb):
    if "ML & Quantitative Research" not in wb.sheetnames: return 0
    ws=wb["ML & Quantitative Research"]; ws._charts=[]
    for r,label in enumerate(["12M price trend","Recent drawdown","6M volatility","Operating margin","FCF yield","6M price trend"],17): ws.cell(r,24,label)
    for r,label in enumerate(["Transition / mixed","Growth / risk-on","Inflation / stagflation","Risk-off / crisis"],26): ws.cell(r,24,label)

    ch=_base_chart(BarChart(),"Forecast Size vs Typical Historical Error",width=13,height=7); ch.type="col"
    ch.add_data(Reference(ws,min_col=25,max_col=26,min_row=3,max_row=4),titles_from_data=False); _set_text_categories(ch,ws,24,3,4,["Forecast magnitude","Typical historical error"]); ch.y_axis.title="Percent"; ws.add_chart(ch,"J6")
    ch=_base_chart(BarChart(),"Model vs Simple Baseline",width=13,height=7); ch.type="col"
    ch.add_data(Reference(ws,min_col=25,max_col=26,min_row=10,max_row=11),titles_from_data=False); _set_text_categories(ch,ws,24,10,11,["Model","Simple baseline"]); ch.y_axis.title="Accuracy (%)"; ws.add_chart(ch,"J21")
    ch=_base_chart(BarChart(),"12M Return Model — Driver Influence",width=13,height=7.5); ch.type="bar"; ch.legend=None
    ch.add_data(Reference(ws,min_col=25,min_row=17,max_row=22),titles_from_data=False); _set_text_categories(ch,ws,24,17,22,["Relative influence"]); ch.x_axis.title="Relative influence"; ws.add_chart(ch,"J36")
    ch=_base_chart(BarChart(),"Market Regime Weights",width=13,height=7.5); ch.type="bar"; ch.legend=None
    ch.add_data(Reference(ws,min_col=25,min_row=26,max_row=29),titles_from_data=False); _set_text_categories(ch,ws,24,26,29,["Weight"]); ch.x_axis.title="Weight (%)"; ws.add_chart(ch,"J52")
    return len(ws._charts)


def _rebuild_ai_growth_charts(wb):
    if "AI Growth Forecast" not in wb.sheetnames: return 0
    ws=wb["AI Growth Forecast"]; ws._charts=[]
    for r,label in enumerate(["Demand","Monetization","Adoption","Efficiency","Capital-light","Risk-adjusted"],3): ws.cell(r,16,label)
    for r,label in enumerate(["Fundamental ML","AI-adjusted","Market-implied"],12): ws.cell(r,16,label)
    for r,label in enumerate(["Revenue: 12M price trend ↑","Revenue: Net debt / rev ↑","Revenue: Net margin ↓","FCF: FCF margin ↓","FCF: Capex / rev ↑","FCF: Drawdown ↓"],18): ws.cell(r,16,label)

    ch=_base_chart(BarChart(),"AI Evidence Scores — 50 = Neutral",width=14.5,height=7.5); ch.type="bar"; ch.legend=None
    ch.add_data(Reference(ws,min_col=17,min_row=3,max_row=8),titles_from_data=False); _set_text_categories(ch,ws,16,3,8,["Supportive score"]); ch.x_axis.title="Score"; ch.x_axis.numFmt="0"; ws.add_chart(ch,"H6")
    ch=_base_chart(BarChart(),"FCF Growth — Model vs Market",width=14.5,height=7.2); ch.type="col"; ch.legend=None
    ch.add_data(Reference(ws,min_col=17,min_row=12,max_row=14),titles_from_data=False); _set_text_categories(ch,ws,16,12,14,["Annual FCF growth"]); ch.y_axis.title="Annual growth (%)"; ws.add_chart(ch,"H20")
    ch=_base_chart(BarChart(),"Growth Forecast — Relative Driver Influence",width=14.5,height=7.8); ch.type="bar"; ch.legend=None
    ch.add_data(Reference(ws,min_col=17,min_row=18,max_row=23),titles_from_data=False); _set_text_categories(ch,ws,16,18,23,["Relative influence"]); ch.x_axis.title="Relative influence"; ws.add_chart(ch,"H34")
    ws["H49"]="Driver influence is model attribution, not causality. The table on the left remains the authoritative detail."; ws["H49"].font=Font(italic=True,color=GREY,size=9)
    return len(ws._charts)


def _rebuild_expectations_chart(wb):
    if "Expectations & Consensus" not in wb.sheetnames: return 0
    ws=wb["Expectations & Consensus"]; ws._charts=[]
    if ws["J46"].value in (None,""): return 0
    ch=_base_chart(BarChart(),"Revenue — Independent Model vs Street",width=12.5,height=7); ch.type="col"
    ch.add_data(Reference(ws,min_col=11,max_col=12,min_row=46,max_row=47),titles_from_data=False); _set_text_categories(ch,ws,10,46,47,["Street","Independent Model"]); ch.y_axis.title="Revenue (USD bn)"; ws.add_chart(ch,"I7")
    return 1


def _rebuild_market_expectations_chart(wb):
    if "Market Expectations" not in wb.sheetnames: return 0
    ws=wb["Market Expectations"]; ws._charts=[]
    if ws["G53"].value in (None,""): return 0
    ch=_base_chart(BarChart(),"Base vs Price-Implied Hurdles",width=12.5,height=7); ch.type="col"
    ch.add_data(Reference(ws,min_col=8,max_col=9,min_row=53,max_row=54),titles_from_data=False); _set_text_categories(ch,ws,7,53,54,["Base","Market Implied"]); ch.y_axis.title="Percent"; ws.add_chart(ch,"F18")
    return 1


def polish_workbook_charts(wb,ticker=None):
    counts={
        "Analysis Charts":_rebuild_analysis_charts(wb),
        "ML & Quantitative Research":_rebuild_ml_charts(wb),
        "AI Growth Forecast":_rebuild_ai_growth_charts(wb),
        "Expectations & Consensus":_rebuild_expectations_chart(wb),
        "Market Expectations":_rebuild_market_expectations_chart(wb),
    }
    return {"charts_rebuilt":sum(counts.values()),"sheet_counts":counts}


def polish_analysis_charts(wb,ticker=None):
    """Backwards-compatible deterministic-run hook."""
    return {"rebuilt":_rebuild_analysis_charts(wb),"chart_count":len(wb["Analysis Charts"]._charts) if "Analysis Charts" in wb.sheetnames else 0}


def apply_chart_readability(workbook_path,ticker=None):
    """File-based final hook for research.py after ML/AI have finished writing charts."""
    path=Path(workbook_path)
    wb=load_workbook(path,data_only=False)
    result=polish_workbook_charts(wb,ticker)
    wb.save(path)
    return result
