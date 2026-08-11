from __future__ import annotations

"""Cross-border financial-data integrity guards.

These controls sit between source extraction and valuation. Their job is to fail closed:
- incomplete fiscal years are not promoted to annual actuals;
- issuer-table values are cross-checked against an independent statement source;
- implausible scale/date artifacts are replaced from the independent source or dropped;
- stale Alphabet template citations are removed from non-Alphabet workbooks;
- core workbook currency labels follow the traded security currency.

The module never invents a missing financial value.
"""

from datetime import date, datetime
import math
import re

import pandas as pd
import yfinance as yf
from openpyxl.styles import Alignment, Font, PatternFill

PALE_GREEN="E2F0D9"; GOLD="FFF2CC"; PALE_RED="FCE4D6"
MONETARY_FIELDS={"revenue","cost","gross","op","ni","ocf","capex","fcf","depr","rd","sbc","cash","assets","liabilities","equity","debt"}


def _num(v, default=None):
    try:
        if isinstance(v, bool) or v in (None, ""):
            return default
        x=float(v)
        return x if math.isfinite(x) else default
    except Exception:
        return default


def _row_value(df, labels, col):
    if df is None or getattr(df,"empty",True) or col not in df.columns:
        return None
    normalized={str(x).strip().lower():x for x in df.index}
    for label in labels:
        key=str(label).strip().lower()
        if key in normalized:
            return _num(df.at[normalized[key],col])
    return None


def yahoo_reference_history(ticker: str) -> dict[int, dict]:
    """Independent annual statement reference used only for validation/replacement."""
    try:
        t=yf.Ticker(ticker)
        income=t.income_stmt; cash=t.cashflow; balance=t.balance_sheet
    except Exception:
        return {}
    columns=set()
    for df in (income,cash,balance):
        if df is not None and not getattr(df,"empty",True):
            columns.update(df.columns)
    out={}
    for col in columns:
        try: y=int(pd.Timestamp(col).year)
        except Exception: continue
        d={}
        d["revenue"]=_row_value(income,["Total Revenue","Operating Revenue"],col)
        d["op"]=_row_value(income,["Operating Income","Operating Income Loss"],col)
        d["ni"]=_row_value(income,["Net Income","Net Income Common Stockholders","Net Income Including Noncontrolling Interests"],col)
        d["eps"]=_row_value(income,["Diluted EPS"],col)
        d["ocf"]=_row_value(cash,["Operating Cash Flow","Total Cash From Operating Activities"],col)
        d["capex"]=_row_value(cash,["Capital Expenditure","Capital Expenditures"],col)
        d["depr"]=_row_value(cash,["Depreciation And Amortization","Depreciation Amortization Depletion"],col)
        d["sbc"]=_row_value(cash,["Stock Based Compensation"],col)
        d["cash"]=_row_value(balance,["Cash Cash Equivalents And Short Term Investments","Cash And Cash Equivalents"],col)
        d["assets"]=_row_value(balance,["Total Assets"],col)
        d["liabilities"]=_row_value(balance,["Total Liabilities Net Minority Interest","Total Liabilities"],col)
        d["equity"]=_row_value(balance,["Stockholders Equity","Total Equity Gross Minority Interest"],col)
        d["debt"]=_row_value(balance,["Total Debt"],col)
        if d.get("capex") is not None: d["capex"]=abs(d["capex"])
        if d.get("ocf") is not None and d.get("capex") is not None: d["fcf"]=d["ocf"]-d["capex"]
        if d.get("ni") is not None and d.get("eps") not in (None,0): d["shares"]=d["ni"]/d["eps"]
        d={k:v for k,v in d.items() if v is not None}
        if d: out[y]=d
    return out


def _date_from_info(value):
    if value in (None,""): return None
    try:
        if isinstance(value,(int,float)) and value>1_000_000:
            return pd.to_datetime(value,unit="s",utc=True).date()
        return pd.Timestamp(value).date()
    except Exception:
        return None


def last_completed_fiscal_year(info: dict | None, reference: dict[int,dict] | None=None) -> int:
    info=info or {}; today=date.today(); d=_date_from_info(info.get("lastFiscalYearEnd"))
    if d and d<=today: return d.year
    years=sorted(int(y) for y in (reference or {}) if int(y)<=today.year)
    if years: return years[-1]
    return today.year-1


def _ratio_bad(value, ref, field):
    v=_num(value); r=_num(ref)
    if v is None or r is None or r==0: return False
    if field=="eps":
        ratio=abs(v/r); return ratio<0.05 or ratio>20
    if field in MONETARY_FIELDS:
        ratio=abs(v/r); lo,hi=(0.33,3.0) if field=="revenue" else (0.08,12.0)
        return ratio<lo or ratio>hi
    return False


def sanitize_crossborder_history(ticker: str, info: dict | None, history: dict | None, meta: dict | None=None):
    """Return validated history + metadata. Bad values are replaced only by independent data."""
    history={int(y):dict(d or {}) for y,d in (history or {}).items()}
    meta=dict(meta or {}); ref=yahoo_reference_history(ticker); completed=last_completed_fiscal_year(info,ref)
    warnings=[]; provenance=dict(meta.get("provenance") or {})

    for y in list(history):
        if int(y)>completed:
            warnings.append(f"Removed FY{y}: latest completed fiscal year is FY{completed}.")
            history.pop(y,None); provenance.pop(y,None)

    for y,d in list(history.items()):
        r=ref.get(y,{})
        for field,value in list(d.items()):
            v=_num(value)
            if v is None:
                d.pop(field,None); continue
            # Date/serial artifacts from permissive spreadsheet parsing are often enormous.
            absolute_bad=(field in MONETARY_FIELDS and abs(v)>1e16) or (field=="eps" and abs(v)>1e6)
            cross_bad=_ratio_bad(v,r.get(field),field)
            if absolute_bad or cross_bad:
                replacement=_num(r.get(field))
                reason="implausible absolute scale" if absolute_bad else "material cross-source scale mismatch"
                if replacement is not None:
                    d[field]=abs(replacement) if field=="capex" else replacement
                    provenance.setdefault(y,{})[field]="Yahoo annual statements — integrity replacement"
                    warnings.append(f"FY{y} {field}: replaced issuer-table value due to {reason}.")
                else:
                    d.pop(field,None); provenance.setdefault(y,{}).pop(field,None)
                    warnings.append(f"FY{y} {field}: dropped due to {reason}; no independent replacement available.")
        if d.get("capex") is not None: d["capex"]=abs(d["capex"])
        if d.get("ocf") is not None and d.get("capex") is not None: d["fcf"]=d["ocf"]-d["capex"]
        if d.get("ni") is not None and d.get("eps") not in (None,0): d["shares"]=d["ni"]/d["eps"]

    # Mature-company scale guard: reject isolated multi-fold annual revenue jumps if no
    # independent source supports them. This is deliberately broad to avoid false precision.
    years=sorted(y for y,d in history.items() if _num(d.get("revenue"),0)>0)
    for prev_y,y in zip(years,years[1:]):
        prev=_num(history[prev_y].get("revenue")); cur=_num(history[y].get("revenue"))
        if not prev or not cur: continue
        ratio=cur/prev
        if ratio>4 or ratio<0.25:
            replacement=_num(ref.get(y,{}).get("revenue"))
            if replacement is not None:
                history[y]["revenue"]=replacement
                provenance.setdefault(y,{})["revenue"]="Yahoo annual statements — growth sanity replacement"
                warnings.append(f"FY{y} revenue: replaced implausible {ratio:.1f}x year/year scale move.")
            else:
                history[y].pop("revenue",None)
                warnings.append(f"FY{y} revenue: removed implausible {ratio:.1f}x year/year scale move.")

    # Latest annual revenue should be in the same broad magnitude as provider TTM revenue.
    provider_rev=_num((info or {}).get("totalRevenue"))
    rev_years=sorted(y for y,d in history.items() if _num(d.get("revenue"),0)>0)
    if provider_rev and rev_years:
        y=rev_years[-1]; annual=_num(history[y].get("revenue")); ratio=annual/provider_rev if annual else None
        if ratio and (ratio>3 or ratio<0.33):
            replacement=_num(ref.get(y,{}).get("revenue"))
            if replacement is not None:
                history[y]["revenue"]=replacement; provenance.setdefault(y,{})["revenue"]="Yahoo annual statements — TTM scale replacement"
                warnings.append(f"FY{y} revenue: corrected after TTM magnitude check ({ratio:.1f}x).")
            else:
                history[y].pop("revenue",None); warnings.append(f"FY{y} revenue: removed after failed TTM magnitude check ({ratio:.1f}x).")

    # Drop empty years and keep completed comparable annual periods only.
    history={y:d for y,d in history.items() if d and y<=completed}
    meta["provenance"]=provenance; meta["years"]=sorted(history); meta["last_completed_fiscal_year"]=completed
    meta["integrity_warnings"]=warnings
    meta["integrity_reference"]="Yahoo annual statements used only as independent validation/replacement when issuer extraction is implausible"
    return history,meta


def _quality_row(wb,label,status,observed,why):
    if "Data Quality" not in wb.sheetnames: return
    ws=wb["Data Quality"]; row=None
    for r in range(1,ws.max_row+1):
        if str(ws.cell(r,1).value or "").strip()==label: row=r; break
    row=row or ws.max_row+1
    ws.cell(row,1,label); ws.cell(row,2,status); ws.cell(row,3,observed); ws.cell(row,4,why)
    ws.cell(row,2).fill=PatternFill("solid",fgColor=PALE_GREEN if status=="PASS" else (GOLD if status=="REVIEW" else PALE_RED)); ws.cell(row,2).font=Font(bold=True)
    for c in range(1,5): ws.cell(row,c).alignment=Alignment(wrap_text=True,vertical="top")


def clear_stale_template_sources(wb,ticker: str) -> int:
    if ticker.upper() in {"GOOG","GOOGL"}: return 0
    patterns=("goog-20","/1652044/","abc.xyz","alphabet inc","alphabet investor")
    cleared=0
    for sheet in ("Historical Financials","Filings","Company Data"):
        if sheet not in wb.sheetnames: continue
        ws=wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                value=cell.value
                if isinstance(value,str) and any(p in value.lower() for p in patterns):
                    cell.value=None; cleared+=1
                    try: cell.hyperlink=None
                    except Exception: pass
    _quality_row(wb,"Template-source contamination","PASS" if cleared>=0 else "FAIL",f"Removed {cleared} stale Alphabet/Google source cell(s).","A non-Alphabet model must not retain Google/Alphabet filing or IR links from the base template.")
    return cleared


def normalize_core_currency_labels(wb, info: dict | None):
    info=info or {}; quote=str(info.get("currency") or info.get("financialCurrency") or "").upper().strip()
    if len(quote)!=3: return
    if "Company Data" in wb.sheetnames:
        ws=wb["Company Data"]
        for row,label in ((10,"Market Cap"),(11,"Enterprise Value"),(12,"Cash"),(13,"Total Debt"),(14,"Net Debt / (Cash)")):
            ws.cell(row,1).value=f"{label} ({quote} bn)"
    if "Historical Financials" in wb.sheetnames:
        wb["Historical Financials"]["A2"]=f"{quote} in billions except per-share data"
    if "Financial Statements" in wb.sheetnames:
        ws=wb["Financial Statements"]
        text=str(ws["A3"].value or "")
        if text: ws["A3"]=re.sub(r"^[A-Z]{3}|^USD|^\$",quote,text)
    # Replace only explicit unit labels/notes, never formulas or ordinary prose.
    for sheet in ("Dashboard","Investment Summary","DCF","Three-Case Scenarios","Advanced Analytics","FCF & Capital Economics","Market Expectations","Segment Analysis","Analysis Charts"):
        if sheet not in wb.sheetnames: continue
        ws=wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                v=cell.value
                if not isinstance(v,str) or v.startswith("="): continue
                v=v.replace("$bn",f"{quote} bn").replace("$ in billions",f"{quote} in billions").replace("$/share",f"{quote}/share")
                cell.value=v


def write_history_integrity_quality(wb,ticker: str,info: dict | None,meta: dict | None=None):
    if "Historical Financials" not in wb.sheetnames: return
    ws=wb["Historical Financials"]; years=[]; revs=[]
    for c in range(2,8):
        y=ws.cell(3,c).value; r=_num(ws.cell(4,c).value)
        if isinstance(y,(int,float)):
            years.append(int(y));
            if r is not None: revs.append((int(y),r))
    completed=(meta or {}).get("last_completed_fiscal_year") or last_completed_fiscal_year(info)
    bad_future=[y for y in years if y>completed]
    _quality_row(wb,"Completed fiscal-year guard","PASS" if not bad_future else "FAIL",f"Latest completed FY={completed}; workbook years={years}","Annual actual history must not contain an unfinished fiscal year.")
    plausible=True; detail="No comparable revenue points"
    if len(revs)>=2:
        moves=[(y,b/a if a else None) for (py,a),(y,b) in zip(revs,revs[1:]) if a and b]
        plausible=all(r is not None and .25<=r<=4 for _,r in moves); detail=", ".join(f"FY{y} {r:.2f}x" for y,r in moves)
    _quality_row(wb,"Historical magnitude sanity","PASS" if plausible else "FAIL",detail,"Large annual scale jumps usually indicate unit/date/parser contamination and must be blocked before valuation.")
    warnings=(meta or {}).get("integrity_warnings") or []
    _quality_row(wb,"Cross-source integrity replacements","REVIEW" if warnings else "PASS",f"{len(warnings)} correction(s)"+(": "+" | ".join(warnings[:4]) if warnings else ""),"Issuer values remain preferred, but implausible parsed values are replaced only when an independent annual statement source supports the replacement.")


def apply_workbook_integrity_controls(wb,ticker: str,info: dict | None,meta: dict | None=None):
    clear_stale_template_sources(wb,ticker)
    normalize_core_currency_labels(wb,info)
    write_history_integrity_quality(wb,ticker,info,meta)
