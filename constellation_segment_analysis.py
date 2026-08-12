from __future__ import annotations

"""Primary-source Constellation Energy segment analysis.

CEG's generic narrative parser previously confused prose fragments with segment names.
This adapter uses the company's SEC XBRL segment disclosures and makes RNF explicit as a
non-GAAP operating metric rather than calling it operating income.
"""

from openpyxl.styles import Alignment, Font, PatternFill

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; GREY="666666"; LIGHT="F5F9FC"
FMT_BN='#,##0.0;[Red](#,##0.0);-'; FMT_PCT='0.0%;[Red](0.0%);-'
SOURCE_10K="https://www.sec.gov/Archives/edgar/data/1868275/000186827526000032/R18.htm"
SOURCE_Q1="https://www.sec.gov/Archives/edgar/data/1868275/000186827526000067/R17.htm"

ANNUAL={
"Mid-Atlantic":{2023:(5.138,2.924),2024:(5.522,3.080),2025:(6.487,3.411)},
"Midwest":{2023:(4.658,3.255),2024:(4.805,3.202),2025:(5.804,3.702)},
"New York":{2023:(2.021,1.251),2024:(2.050,1.453),2025:(2.190,1.600)},
"ERCOT":{2023:(1.346,.582),2024:(1.550,1.047),2025:(1.904,1.137)},
"Other Power Regions":{2023:(5.851,1.240),2024:(5.506,1.268),2025:(5.583,.819)},
}
Q1_2026={"Mid-Atlantic":(1.847,.812),"Midwest":(1.732,.854),"New York":(.569,.409),"ERCOT":(.370,.209),"Other Power Regions":(1.487,.267),"Calpine":(2.395,1.126)}


def _fill(c): return PatternFill("solid",fgColor=c)
def _section(ws,row,title,end=10):
    for c in range(1,end+1): ws.cell(row,c).fill=_fill(NAVY); ws.cell(row,c).font=Font(bold=True,color=WHITE)
    ws.cell(row,1,title)
def _header(ws,row,values):
    for c,v in enumerate(values,1):
        ws.cell(row,c,v); ws.cell(row,c).fill=_fill(BLUE); ws.cell(row,c).font=Font(bold=True,color=WHITE); ws.cell(row,c).alignment=Alignment(wrap_text=True,horizontal="center")


def ensure_constellation_segment_analysis(wb,ticker="CEG"):
    if "Segment Analysis" in wb.sheetnames: wb.remove(wb["Segment Analysis"])
    ws=wb.create_sheet("Segment Analysis"); ws.sheet_view.showGridLines=False
    for c in range(1,11): ws.cell(1,c).fill=_fill(NAVY); ws.cell(2,c).fill=_fill(NAVY)
    ws["A1"]="CEG — Segment Analysis"; ws["A1"].font=Font(bold=True,color=WHITE,size=18)
    ws.merge_cells("A3:J3"); ws["A3"]="Primary-source SEC segment data. RNF = operating revenues less purchased power and fuel expense; it is management's segment performance metric and is not GAAP operating income."; ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True)

    _section(ws,5,"FY2023–FY2025 Reportable Segments",10)
    _header(ws,6,["Segment","FY23 Revenue","FY24 Revenue","FY25 Revenue","FY25 Growth","FY23 RNF","FY24 RNF","FY25 RNF","FY25 RNF Margin","Source"])
    r=7
    for name,data in ANNUAL.items():
        rev23,rnf23=data[2023]; rev24,rnf24=data[2024]; rev25,rnf25=data[2025]
        vals=[name,rev23,rev24,rev25,rev25/rev24-1,rnf23,rnf24,rnf25,rnf25/rev25,SOURCE_10K]
        for c,v in enumerate(vals,1): ws.cell(r,c,v); ws.cell(r,c).alignment=Alignment(wrap_text=True,vertical="top")
        for c in (2,3,4,6,7,8): ws.cell(r,c).number_format=FMT_BN
        for c in (5,9): ws.cell(r,c).number_format=FMT_PCT
        ws.cell(r,10).hyperlink=SOURCE_10K; r+=1
    total23=sum(v[2023][0] for v in ANNUAL.values()); total24=sum(v[2024][0] for v in ANNUAL.values()); total25=sum(v[2025][0] for v in ANNUAL.values())
    rnfs=[sum(v[y][1] for v in ANNUAL.values()) for y in (2023,2024,2025)]
    ws.cell(r,1,"Total Reportable Segments"); ws.cell(r,2,total23); ws.cell(r,3,total24); ws.cell(r,4,total25); ws.cell(r,5,total25/total24-1); ws.cell(r,6,rnfs[0]); ws.cell(r,7,rnfs[1]); ws.cell(r,8,rnfs[2]); ws.cell(r,9,rnfs[2]/total25); ws.cell(r,10,SOURCE_10K); ws.cell(r,10).hyperlink=SOURCE_10K
    for c in range(1,11): ws.cell(r,c).font=Font(bold=True)
    for c in (2,3,4,6,7,8): ws.cell(r,c).number_format=FMT_BN
    for c in (5,9): ws.cell(r,c).number_format=FMT_PCT

    r+=3; _section(ws,r,"Q1 2026 — Calpine Becomes a Sixth Reportable Segment",10); r+=1
    _header(ws,r,["Segment","Q1 2026 Revenue","Q1 2026 RNF","RNF Margin","Change in Reporting","Source"]); r+=1
    for name,(rev,rnf) in Q1_2026.items():
        vals=[name,rev,rnf,rnf/rev,"New reportable segment in 2026" if name=="Calpine" else "Continuing segment",SOURCE_Q1]
        for c,v in enumerate(vals,1): ws.cell(r,c,v); ws.cell(r,c).alignment=Alignment(wrap_text=True)
        ws.cell(r,2).number_format=FMT_BN; ws.cell(r,3).number_format=FMT_BN; ws.cell(r,4).number_format=FMT_PCT; ws.cell(r,6).hyperlink=SOURCE_Q1; r+=1

    r+=2; _section(ws,r,"Interpretation & Reconciliation",10); r+=1
    notes=[
        ("2025 consolidated operating revenue","$25.533bn","Includes reportable segments plus activities not allocated to a region; do not substitute contracts-with-customers revenue of $22.663bn for consolidated revenue."),
        ("2025 consolidated RNF","$10.852bn","RNF is revenue less purchased power and fuel; it is not GAAP operating income."),
        ("2026 structural break","Calpine acquired Jan. 7, 2026","Historical segment comparisons must flag the new Calpine segment rather than treating 2026 as directly comparable."),
    ]
    _header(ws,r,["Item","Value","Research interpretation"]); r+=1
    for a,b,c in notes:
        ws.cell(r,1,a); ws.cell(r,2,b); ws.cell(r,3,c); ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=10); ws.cell(r,3).alignment=Alignment(wrap_text=True); r+=1

    widths={"A":27,"B":16,"C":16,"D":16,"E":14,"F":16,"G":16,"H":16,"I":16,"J":55}
    for col,w in widths.items(): ws.column_dimensions[col].width=w
    ws.freeze_panes="A7"
    return ws
