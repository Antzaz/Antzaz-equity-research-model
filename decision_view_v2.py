from __future__ import annotations

"""Clear stock-attractiveness decision view and cross-sheet consistency repair.

The workbook intentionally distinguishes business quality from stock price. This module turns
that into a one-page decision surface: what the company quality is, what valuation says, how
large the margin of safety is, what stress/Monte Carlo say, and what price would satisfy common
upside hurdles. It also repairs stale narrative values in Investment Summary by rebuilding the
headline from the single-source Score Engine v2.
"""

import math
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from advanced_analytics_v2 import _base_value
from score_engine_v2 import compute_score_bundle

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; GREY="666666"; GREEN="E2F0D9"; YELLOW="FFF2CC"; RED="FCE4D6"; LIGHT="F5F9FC"
FMT_PCT='0.0%;[Red](0.0%);-'; FMT_PRICE='$#,##0.00;[Red]($#,##0.00);-'; FMT_SCORE='0.0'; FMT_MULT='0.0x;[Red](0.0x);-'
THIN=Side(style="thin",color="D9E1F2")


def _fill(c): return PatternFill("solid",fgColor=c)
def _num(v,default=None):
    try:
        if isinstance(v,bool) or v in (None,""): return default
        x=float(v); return x if math.isfinite(x) else default
    except Exception: return default

def _find(ws,label,col=1):
    needle=str(label).strip().lower()
    for r in range(1,ws.max_row+1):
        if str(ws.cell(r,col).value or "").strip().lower()==needle: return r
    return None

def _value(wb,sheet,label,value_col=2,search_col=1):
    if sheet not in wb.sheetnames: return None
    ws=wb[sheet]; r=_find(ws,label,search_col); return ws.cell(r,value_col).value if r else None

def _any_value(wb,sheet,label,search_cols=(1,9)):
    if sheet not in wb.sheetnames: return None
    ws=wb[sheet]; needle=str(label).strip().lower()
    for c in search_cols:
        for r in range(1,ws.max_row+1):
            if str(ws.cell(r,c).value or "").strip().lower()==needle: return ws.cell(r,c+1).value
    return None

def _section(ws,row,title,end=8):
    for c in range(1,end+1): ws.cell(row,c).fill=_fill(NAVY); ws.cell(row,c).font=Font(bold=True,color=WHITE)
    ws.cell(row,1,title)
def _header(ws,row,values):
    for c,v in enumerate(values,1): ws.cell(row,c,v); ws.cell(row,c).fill=_fill(BLUE); ws.cell(row,c).font=Font(bold=True,color=WHITE); ws.cell(row,c).alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def _score_fill(x):
    if x is None: return _fill(LIGHT)
    return _fill(GREEN if x>=65 else RED if x<40 else YELLOW)
def _signal(score,good=65,bad=40):
    if score is None: return "REVIEW"
    return "POSITIVE" if score>=good else "NEGATIVE" if score<bad else "MIXED"


def _mc(wb):
    return {
        "median":_num(_any_value(wb,"Advanced Analytics","Median Value / Share")),
        "p10":_num(_any_value(wb,"Advanced Analytics","P10 Value / Share")),
        "prob":_num(_any_value(wb,"Advanced Analytics","Probability > Current Price")),
        "implied_fcf":_num(_any_value(wb,"Advanced Analytics","Implied 10Y FCF CAGR",(1,9))),
    }


def _peer_metrics(wb):
    out={}
    if "Peer Comps" not in wb.sheetnames: return out
    ws=wb["Peer Comps"]; headers={str(ws.cell(3,c).value or "").strip():c for c in range(1,ws.max_column+1)}
    target=None; peers=[]
    for r in range(4,min(ws.max_row,40)+1):
        typ=str(ws.cell(r,headers.get("Peer Type",16)).value or "").strip()
        if typ=="Target classification": target=r
        elif str(ws.cell(r,headers.get("Ticker",2)).value or "").strip() and typ in {"Direct business-model peer","Exact industry"}: peers.append(r)
    if target is None: target=4
    import statistics
    for key in ("Forward P/E","EV/Revenue","EV/EBITDA","Revenue Growth","Operating Margin","ROE"):
        c=headers.get(key)
        if not c: continue
        v=_num(ws.cell(target,c).value); vals=[_num(ws.cell(r,c).value) for r in peers]; vals=[x for x in vals if x is not None]
        out[key]=(v,statistics.median(vals) if vals else None)
    return out


def _classification(quality,valuation,overall,base_up,mc_prob,failed=False):
    if failed: return "REVIEW — DATA QUALITY","Resolve failed data-quality checks before relying on valuation."
    if quality is not None and quality>=70 and valuation is not None and valuation<30:
        return "HIGH-QUALITY BUSINESS / EXPENSIVE STOCK","Business economics screen strongly, but valuation offers little or no margin of safety at the current price."
    if overall is not None and overall>=70 and (base_up is None or base_up>=.15) and (mc_prob is None or mc_prob>=.60):
        return "ATTRACTIVE","Quality, valuation and scenario evidence align positively."
    if overall is not None and overall>=58 and (base_up is None or base_up>=.05):
        return "POTENTIALLY ATTRACTIVE","Several pillars are favorable, but the margin of safety is not yet strong enough for a high-conviction model view."
    if overall is not None and overall>=45:
        return "WATCHLIST / NEUTRAL","The analysis is mixed; price, expectations or risk need to improve before the model becomes attractive."
    return "UNATTRACTIVE AT CURRENT PRICE","The current price is not supported by enough of the model's valuation/risk evidence."


def _data_quality(wb):
    p=r=f=0
    if "Data Quality" not in wb.sheetnames: return p,r,f
    ws=wb["Data Quality"]
    for rr in range(1,ws.max_row+1):
        s=str(ws.cell(rr,2).value or "").upper().strip()
        if s=="PASS": p+=1
        elif s=="REVIEW": r+=1
        elif s=="FAIL": f+=1
    return p,r,f


def _repair_investment_summary(wb,bundle,base,severe,mc):
    if "Investment Summary" not in wb.sheetnames: return
    ws=wb["Investment Summary"]; price=_num(wb["Company Data"]["B8"].value) if "Company Data" in wb.sheetnames else None
    quality=bundle["category_scores"].get("Business Quality"); valuation=bundle["category_scores"].get("Valuation / Stock Attractiveness"); overall=bundle["category_scores"].get("Overall Investment Score")
    base_up=base/price-1 if base is not None and price else None; severe=max(0.0,severe) if severe is not None else None; severe_down=severe/price-1 if severe is not None and price else None
    _,_,failed=_data_quality(wb); view,why=_classification(quality,valuation,overall,base_up,mc.get("prob"),failed>0)
    # Replace headline fields by label rather than fixed coordinates.
    for label,value in (("Model View",view),("Current Price",price),("Base DCF / Share",base),("Base Upside",base_up),("Severe Bear / Share",severe),("Severe Downside",severe_down),("WACC",_num(wb["Three-Case Scenarios"]["C6"].value) if "Three-Case Scenarios" in wb.sheetnames else None)):
        r=_find(ws,label)
        if r: ws.cell(r,2,value)
    # Quantitative score may be stored in a second value column in the decision header.
    r=_find(ws,"Quant Score / 100")
    if r:
        for c in range(2,min(10,ws.max_column)+1):
            if isinstance(ws.cell(r,c).value,(int,float)): ws.cell(r,c,overall); break
    r=_find(ws,"Why")
    if r:
        pieces=[why]
        if base_up is not None: pieces.append(f"Base DCF {base_up:+.1%} vs current price")
        if overall is not None: pieces.append(f"overall score {overall:.1f}/100")
        if quality is not None and valuation is not None: pieces.append(f"business quality {quality:.1f}/100 vs valuation {valuation:.1f}/100")
        if severe_down is not None: pieces.append(f"severe-bear downside {severe_down:.1%}")
        if mc.get("prob") is not None: pieces.append(f"Monte Carlo P(value > price) {mc['prob']:.1%}")
        ws.cell(r,2,"; ".join(pieces)+".")
        ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=min(10,ws.max_column)); ws.cell(r,2).alignment=Alignment(wrap_text=True)
    # Disambiguate provider/TTM growth from annual growth used in Score Engine.
    r=_find(ws,"Latest Revenue Growth")
    if r: ws.cell(r,1,"Current / TTM Revenue Growth (market-data field)")
    ws["J1"]="Decision View"; ws["J1"].hyperlink="#'Decision View'!A1"; ws["J1"].font=Font(color="008000",underline="single",bold=True)


def ensure_decision_view(wb,ticker):
    ticker=str(ticker).upper(); price=_num(wb["Company Data"]["B8"].value) if "Company Data" in wb.sheetnames else None
    try: base=_num(_base_value(wb))
    except Exception: base=None
    try: severe=_num(_base_value(wb,-.05,-.05,.05,.02,-.01))
    except Exception: severe=None
    severe_floor=max(0.0,severe) if severe is not None else None
    mc=_mc(wb); bundle=compute_score_bundle(wb,ticker=ticker,base_value=base,severe_value=severe_floor,current_price=price,mc_prob=mc.get("prob"))
    scores=bundle["category_scores"]; quality=scores.get("Business Quality"); valuation=scores.get("Valuation / Stock Attractiveness"); risk=scores.get("Downside / Scenario Risk"); overall=scores.get("Overall Investment Score")
    base_up=base/price-1 if base is not None and price else None; _,reviews,failed=_data_quality(wb); view,why=_classification(quality,valuation,overall,base_up,mc.get("prob"),failed>0)
    _repair_investment_summary(wb,bundle,base,severe_floor,mc)

    if "Decision View" in wb.sheetnames: wb.remove(wb["Decision View"])
    ws=wb.create_sheet("Decision View",0); ws.sheet_view.showGridLines=False
    for c in range(1,9): ws.cell(1,c).fill=_fill(NAVY); ws.cell(2,c).fill=_fill(NAVY)
    ws["A1"]=f"{ticker} — Is the Stock Attractive?"; ws["A1"].font=Font(bold=True,color=WHITE,size=19)
    ws["A3"]="Model decision, not a recommendation. A good company and a good stock price are evaluated separately. Read the score/audit and source sheets before relying on the conclusion."
    ws.merge_cells("A3:H3"); ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True)

    _section(ws,5,"Headline Decision",8)
    ws["A6"]="MODEL VIEW"; ws["B6"]=view; ws.merge_cells("B6:H6"); ws["B6"].font=Font(bold=True,size=15); ws["B6"].fill=_score_fill(overall)
    ws["A7"]="Interpretation"; ws["B7"]=why; ws.merge_cells("B7:H8"); ws["B7"].alignment=Alignment(wrap_text=True,vertical="top")

    _section(ws,10,"Decision Scorecard",8)
    _header(ws,11,["Pillar","Score / Value","Signal","What it measures","Current Evidence","Reference / Threshold","Proof","Status"])
    dims=bundle["dimensions"]
    rows=[
        ("Business Quality",quality,_signal(quality),"Growth, returns, cash conversion and balance sheet",f"Growth {dims['Growth']['score']:.1f}; Profitability {dims['Profitability']['score']:.1f}; FCF {dims['FCF Quality']['score']:.1f}" if all(dims[x]['score'] is not None for x in ('Growth','Profitability','FCF Quality')) else "See audit trail","≥65 positive","Score Audit Trail",dims['Profitability']['status']),
        ("Valuation / Price",valuation,_signal(valuation),"Absolute DCF plus peer-relative valuation",f"Base value {base:,.2f}; current price {price:,.2f}" if base is not None and price is not None else "N/A","≥60 positive","Score Audit Trail",dims['Absolute Valuation']['status']),
        ("Overall Investment Score",overall,_signal(overall),"Weighted quality + valuation + risk bridge",f"Coverage {bundle['coverage']:.0%}","≥60 increasingly attractive","Investment Summary",f"{bundle['coverage']:.0%} coverage"),
        ("Base DCF Upside",base_up,"POSITIVE" if base_up is not None and base_up>=.15 else "NEGATIVE" if base_up is not None and base_up<0 else "MIXED","Margin of safety versus deterministic base value",base_up,"+15% preferred margin-of-safety reference","DCF / Cost of Capital","Complete" if base_up is not None else "Missing"),
        ("Monte Carlo P(Value > Price)",mc.get("prob"),"POSITIVE" if mc.get("prob") is not None and mc['prob']>=.60 else "NEGATIVE" if mc.get("prob") is not None and mc['prob']<.40 else "MIXED","How often simulated value exceeds current price",mc.get("prob"),">60% positive; <40% negative","Advanced Analytics","Complete" if mc.get("prob") is not None else "Missing"),
        ("Downside / Scenario Risk",risk,_signal(risk),"Severe-bear valuation and scenario skew",f"Severe floor {severe_floor:,.2f}" if severe_floor is not None else "N/A","Higher score = better downside protection","Score Audit Trail",dims['Stress Robustness']['status']),
    ]
    for r,row in enumerate(rows,12):
        for c,v in enumerate(row,1): ws.cell(r,c,v); ws.cell(r,c).alignment=Alignment(wrap_text=True,vertical="top")
        v=row[1]
        if isinstance(v,(int,float)):
            if row[0] in {"Base DCF Upside","Monte Carlo P(Value > Price)"}: ws.cell(r,2).number_format=FMT_PCT
            else: ws.cell(r,2).number_format=FMT_SCORE
        ws.cell(r,2).fill=_score_fill(v*100 if row[0] in {"Base DCF Upside","Monte Carlo P(Value > Price)"} and isinstance(v,(int,float)) else v)

    _section(ws,20,"Price & Margin-of-Safety Map",8)
    _header(ws,21,["Price Level","Value / Share","Gap vs Current","Meaning","Formula","Signal","Source","Notes"])
    levels=[
        ("Current Market Price",price,0.0,"Price the model must justify","Market price","Reference","Company Data","Refreshes each build"),
        ("Base DCF Fair Value",base,base/price-1 if base is not None and price else None,"Deterministic base-case present value","DCF intrinsic value","Positive" if base_up is not None and base_up>0 else "Negative","DCF","Uses calculated company-specific WACC"),
        ("15% Upside Entry",base/1.15 if base is not None else None,(base/1.15)/price-1 if base is not None and price else None,"Maximum price that leaves 15% upside to base fair value","Base value / 1.15","Entry reference","Calculated","Not a target; margin-of-safety reference"),
        ("25% Upside Entry",base/1.25 if base is not None else None,(base/1.25)/price-1 if base is not None and price else None,"Maximum price that leaves 25% upside","Base value / 1.25","Entry reference","Calculated","Stricter margin-of-safety reference"),
        ("Monte Carlo Median",mc.get("median"),mc['median']/price-1 if mc.get('median') is not None and price else None,"Median simulated intrinsic value","50th percentile","Scenario reference","Advanced Analytics","Regime Monte Carlo"),
        ("Monte Carlo P10",mc.get("p10"),mc['p10']/price-1 if mc.get('p10') is not None and price else None,"Downside valuation percentile","10th percentile","Risk reference","Advanced Analytics","Not a price target"),
    ]
    for r,row in enumerate(levels,22):
        for c,v in enumerate(row,1): ws.cell(r,c,v); ws.cell(r,c).alignment=Alignment(wrap_text=True,vertical="top")
        ws.cell(r,2).number_format=FMT_PRICE; ws.cell(r,3).number_format=FMT_PCT

    _section(ws,30,"What the Market Price Requires",8)
    peer=_peer_metrics(wb); wacc=_num(wb["Three-Case Scenarios"]["C6"].value) if "Three-Case Scenarios" in wb.sheetnames else None
    items=[
        ("Calculated Base WACC",wacc,FMT_PCT,"See Cost of Capital for risk-free rate, ERP, beta, debt cost and capital weights."),
        ("Market-Implied 10Y FCF CAGR",mc.get("implied_fcf"),FMT_PCT,"Reverse DCF growth required to support current enterprise value under the model's discount/terminal assumptions."),
        ("Forward P/E",peer.get("Forward P/E",(None,None))[0],FMT_MULT,f"Direct/exact peer median {peer.get('Forward P/E',(None,None))[1]}"),
        ("EV / Revenue",peer.get("EV/Revenue",(None,None))[0],FMT_MULT,f"Direct/exact peer median {peer.get('EV/Revenue',(None,None))[1]}"),
        ("EV / EBITDA",peer.get("EV/EBITDA",(None,None))[0],FMT_MULT,f"Direct/exact peer median {peer.get('EV/EBITDA',(None,None))[1]}"),
        ("Latest Annual Revenue Growth",_annual_growth(wb),FMT_PCT,"Actual latest fiscal-year growth used by Score Engine; separate from current/TTM provider fields."),
        ("Current / TTM Revenue Growth",peer.get("Revenue Growth",(None,None))[0],FMT_PCT,f"Market-data field; direct/exact peer median {peer.get('Revenue Growth',(None,None))[1]}"),
    ]
    _header(ws,31,["Metric","Current","Format","Interpretation / Reference","Source","Status","Decision Use","Caveat"])
    for r,(label,val,fmt,note) in enumerate(items,32):
        ws.cell(r,1,label); ws.cell(r,2,val); ws.cell(r,2).number_format=fmt; ws.cell(r,4,note); ws.cell(r,5,"Cost of Capital / Advanced Analytics / Peer Comps / Historical Financials"); ws.cell(r,6,"Complete" if val is not None else "Missing"); ws.cell(r,7,"Expectations / valuation proof"); ws.cell(r,8,"Compare like-for-like periods and business models.")
        for c in range(1,9): ws.cell(r,c).alignment=Alignment(wrap_text=True,vertical="top")

    _section(ws,41,"Data Confidence & Next Decision",8)
    p,rev,fail=_data_quality(wb)
    ws["A42"]="Data Quality"; ws["B42"]=f"{p} PASS / {rev} REVIEW / {fail} FAIL"; ws["C42"]="A FAIL overrides an attractive score until resolved. REVIEW items lower confidence but do not fabricate missing data."; ws.merge_cells("C42:H42"); ws["C42"].alignment=Alignment(wrap_text=True)
    ws["A43"]="Decision Rule"; ws["B43"]="Quality + price + downside"; ws["C43"]="High business quality alone is not enough. The model becomes attractive only when valuation/margin-of-safety and scenario evidence also improve."; ws.merge_cells("C43:H44"); ws["C43"].alignment=Alignment(wrap_text=True)
    ws["A46"]="Proof links"; ws["B46"]="Score Audit Trail | Cost of Capital | DCF | Advanced Analytics | Peer Comps | Segment Analysis | Data Quality"
    ws.merge_cells("B46:H46"); ws["B46"].font=Font(color="008000",bold=True)

    widths={"A":31,"B":20,"C":18,"D":44,"E":38,"F":23,"G":25,"H":36}
    for c,w in widths.items(): ws.column_dimensions[c].width=w
    ws.freeze_panes="A11"
    _quality_row(wb,view,overall)
    print(f"Decision View: {ticker} = {view}; overall={overall}; quality={quality}; valuation={valuation}; base_up={base_up}")
    return {"view":view,"overall":overall,"quality":quality,"valuation":valuation,"risk":risk,"base_upside":base_up}


def _annual_growth(wb):
    if "Historical Financials" not in wb.sheetnames: return None
    ws=wb["Historical Financials"]; vals=[]
    for c in range(2,8):
        v=_num(ws.cell(4,c).value)
        if v is not None: vals.append(v)
    return vals[-1]/vals[-2]-1 if len(vals)>=2 and vals[-2] else None


def _quality_row(wb,view,overall):
    if "Data Quality" not in wb.sheetnames: return
    ws=wb["Data Quality"]; label="Decision-view score reconciliation"; row=None
    for r in range(1,ws.max_row+1):
        if str(ws.cell(r,1).value or "").strip()==label: row=r; break
    row=row or ws.max_row+1
    ws.cell(row,1,label); ws.cell(row,2,"PASS"); ws.cell(row,3,f"{view}; overall score {overall:.1f}" if overall is not None else view); ws.cell(row,4,"Decision View and Investment Summary use Score Engine v2; severe-bear equity value is floored at zero consistently and annual vs TTM growth labels are separated.")
    ws.cell(row,2).fill=_fill(GREEN); ws.cell(row,2).font=Font(bold=True)
    for c in range(1,5): ws.cell(row,c).alignment=Alignment(wrap_text=True,vertical="top")
