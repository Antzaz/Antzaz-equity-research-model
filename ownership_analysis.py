"""Ownership and shareholder analysis using yfinance's public holdings endpoints.

The workbook separates ownership percentages, largest institutions/funds, insider roster
holdings, and recent insider transactions. Public holdings data can lag regulatory filings,
so dates and a source note are always shown and missing fields remain blank.
"""

import re
from datetime import datetime

try:
    import yfinance as yf
except Exception:
    yf=None

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; GREY="666666"
FMT_PCT='0.0%;[Red](0.0%);-'; FMT_INT='#,##0;[Red](#,##0);-'; FMT_M='$#,##0.0,,"m";[Red]($#,##0.0,,"m");-'
THIN=Side(style="thin",color="D9E1F2")


def _fill(c): return PatternFill("solid",fgColor=c)
def _num(v,default=None):
    try:
        if isinstance(v,bool) or v in (None,""): return default
        return float(v)
    except Exception: return default

def _norm(x): return re.sub(r"[^a-z0-9]","",str(x or "").lower())
def _style_title(ws,text):
    for c in range(1,9): ws.cell(1,c).fill=_fill(NAVY); ws.cell(2,c).fill=_fill(NAVY)
    ws["A1"]=text; ws["A1"].font=Font(bold=True,color=WHITE,size=19); ws.sheet_view.showGridLines=False

def _section(ws,row,title,end=8):
    for c in range(1,end+1): ws.cell(row,c).fill=_fill(NAVY); ws.cell(row,c).font=Font(bold=True,color=WHITE,size=11)
    ws.cell(row,1,title)

def _header(ws,row,headers):
    for c,v in enumerate(headers,1):
        cell=ws.cell(row,c,v); cell.fill=_fill(BLUE); cell.font=Font(bold=True,color=WHITE)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=Border(bottom=THIN)

def _frame(ticker,attr):
    if yf is None: return None
    try:
        obj=getattr(ticker,attr,None)
        return obj if obj is not None and not getattr(obj,"empty",False) else None
    except Exception: return None

def _records(df):
    if df is None: return []
    try:
        x=df.reset_index()
        x.columns=[" | ".join(map(str,c)) if isinstance(c,tuple) else str(c) for c in x.columns]
        return x.to_dict("records")
    except Exception: return []
def _pick(rec,*aliases):
    lookup={_norm(k):v for k,v in rec.items()}
    for a in aliases:
        na=_norm(a)
        if na in lookup: return lookup[na]
    for a in aliases:
        na=_norm(a)
        for k,v in lookup.items():
            if na and (na in k or k in na): return v
    return None

def _major_value(df,*keywords):
    for rec in _records(df):
        text=" ".join(str(v) for v in rec.values() if v is not None).lower()
        if all(k.lower() in text for k in keywords):
            nums=[_num(v) for v in rec.values()]
            nums=[v for v in nums if v is not None]
            if nums: return nums[-1]
    return None

def _pct(v):
    v=_num(v)
    if v is None: return None
    return v/100 if v>1.5 else v

def _date(v):
    if v in (None,""): return None
    try:
        if hasattr(v,"strftime"): return v.strftime("%Y-%m-%d")
    except Exception: pass
    return str(v)[:10]

def _institution_rows(df,shares_out,limit=15):
    out=[]
    for rec in _records(df):
        holder=_pick(rec,"Holder","Name")
        shares=_num(_pick(rec,"Shares","Shares Held"))
        pct=_pct(_pick(rec,"% Out","Pct Out","Percent Out","Percentage Held"))
        if pct is None and shares is not None and shares_out: pct=shares/shares_out
        value=_num(_pick(rec,"Value","Market Value"))
        date=_date(_pick(rec,"Date Reported","Report Date","Date"))
        if holder: out.append([str(holder),shares,pct,value,date])
    out.sort(key=lambda r:(r[1] is not None,r[1] or 0),reverse=True)
    return out[:limit]

def _insider_rows(df,shares_out,limit=15):
    out=[]
    for rec in _records(df):
        name=_pick(rec,"Name","Insider")
        position=_pick(rec,"Position","Title","Relation")
        direct=_num(_pick(rec,"Shares Owned Directly","Position Direct","Direct Shares","Shares"))
        indirect=_num(_pick(rec,"Shares Owned Indirectly","Position Indirect","Indirect Shares"))
        total=(direct or 0)+(indirect or 0) if direct is not None or indirect is not None else None
        pct=total/shares_out if total is not None and shares_out else None
        recent=_pick(rec,"Most Recent Transaction","Latest Transaction","Transaction")
        date=_date(_pick(rec,"Latest Transaction Date","Position Direct Date","Date"))
        if name: out.append([str(name),position,direct,indirect,total,pct,recent,date])
    out.sort(key=lambda r:(r[4] is not None,r[4] or 0),reverse=True)
    return out[:limit]

def _transaction_rows(df,limit=15):
    out=[]
    for rec in _records(df):
        insider=_pick(rec,"Insider","Name")
        text=_pick(rec,"Text","Transaction","Transaction Type")
        shares=_num(_pick(rec,"Shares","Shares Traded"))
        value=_num(_pick(rec,"Value","Transaction Value"))
        start=_date(_pick(rec,"Start Date","Date","Transaction Date"))
        ownership=_pick(rec,"Ownership","Ownership Type")
        if insider or text: out.append([start,insider,text,shares,value,ownership])
    return out[:limit]

def ensure_ownership_analysis(wb,ticker):
    name="Ownership & Holders"
    if name in wb.sheetnames: wb.remove(wb[name])
    index=2 if len(wb.worksheets)>=2 else len(wb.worksheets)
    ws=wb.create_sheet(name,index)
    _style_title(ws,f"{ticker} — Ownership & Largest Shareholders")
    ws["A3"]="Ownership snapshot from Yahoo Finance via yfinance. Institutional/fund positions can lag regulatory filings; insider roster figures may not equal full beneficial ownership. Always use the reported dates when interpreting changes."
    ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True)
    try: ws.merge_cells("A3:H3")
    except Exception: pass

    t=yf.Ticker(ticker) if yf is not None else None
    info={}
    try: info=(t.info or {}) if t is not None else {}
    except Exception: info={}
    shares_out=_num(info.get("sharesOutstanding"))
    insider_pct=_pct(info.get("heldPercentInsiders")); inst_pct=_pct(info.get("heldPercentInstitutions"))
    major=_frame(t,"major_holders") if t is not None else None
    insider_pct=insider_pct if insider_pct is not None else _pct(_major_value(major,"insider"))
    inst_pct=inst_pct if inst_pct is not None else _pct(_major_value(major,"institution"))

    inst=_frame(t,"institutional_holders") if t is not None else None
    funds=_frame(t,"mutualfund_holders") if t is not None else None
    roster=_frame(t,"insider_roster_holders") if t is not None else None
    trans=_frame(t,"insider_transactions") if t is not None else None
    purchases=_frame(t,"insider_purchases") if t is not None else None

    _section(ws,5,"Ownership Snapshot")
    summary=[
        ("Shares Outstanding",shares_out,FMT_INT,"Company / Yahoo Finance"),
        ("Held by Insiders",insider_pct,FMT_PCT,"Yahoo Finance ownership summary"),
        ("Held by Institutions",inst_pct,FMT_PCT,"Yahoo Finance ownership summary"),
        ("Institutional Float Held",_pct(_major_value(major,"institution","float")),FMT_PCT,"Yahoo Finance major holders"),
        ("Reporting Snapshot Generated",datetime.now().strftime("%Y-%m-%d %H:%M"),None,"Workbook generation time"),
    ]
    for r,(lab,val,fmt,src) in enumerate(summary,6):
        ws.cell(r,1,lab); ws.cell(r,2,val); ws.cell(r,3,src)
        if fmt: ws.cell(r,2).number_format=fmt

    inst_rows=_institution_rows(inst,shares_out,15)
    r=13; _section(ws,r,"Largest Institutional Holders")
    _header(ws,r+1,["Holder","Shares","% Outstanding","Reported Value","Date Reported","","",""])
    for rr,row in enumerate(inst_rows or [["No institutional-holder table returned",None,None,None,None]],r+2):
        for c,v in enumerate(row,1): ws.cell(rr,c,v)
        ws.cell(rr,2).number_format=FMT_INT; ws.cell(rr,3).number_format=FMT_PCT; ws.cell(rr,4).number_format=FMT_M

    r=r+2+max(3,len(inst_rows))+2; _section(ws,r,"Largest Mutual Fund / ETF Holders")
    _header(ws,r+1,["Fund / ETF","Shares","% Outstanding","Reported Value","Date Reported","","",""])
    fund_rows=_institution_rows(funds,shares_out,15)
    for rr,row in enumerate(fund_rows or [["No mutual-fund holder table returned",None,None,None,None]],r+2):
        for c,v in enumerate(row,1): ws.cell(rr,c,v)
        ws.cell(rr,2).number_format=FMT_INT; ws.cell(rr,3).number_format=FMT_PCT; ws.cell(rr,4).number_format=FMT_M

    r=r+2+max(3,len(fund_rows))+2; _section(ws,r,"Insider Ownership Roster")
    _header(ws,r+1,["Insider","Position","Direct Shares","Indirect Shares","Total Shown","% Outstanding","Most Recent Activity","Date"])
    insider_rows=_insider_rows(roster,shares_out,15)
    for rr,row in enumerate(insider_rows or [["No insider-roster table returned",None,None,None,None,None,None,None]],r+2):
        for c,v in enumerate(row,1): ws.cell(rr,c,v)
        for c in (3,4,5): ws.cell(rr,c).number_format=FMT_INT
        ws.cell(rr,6).number_format=FMT_PCT

    r=r+2+max(3,len(insider_rows))+2; _section(ws,r,"Recent Insider Transactions")
    _header(ws,r+1,["Date","Insider","Transaction / Description","Shares","Value","Ownership","",""])
    trans_rows=_transaction_rows(trans,15)
    for rr,row in enumerate(trans_rows or [[None,"No recent insider-transaction table returned",None,None,None,None]],r+2):
        for c,v in enumerate(row,1): ws.cell(rr,c,v)
        ws.cell(rr,4).number_format=FMT_INT; ws.cell(rr,5).number_format=FMT_M

    r=r+2+max(3,len(trans_rows))+2; _section(ws,r,"Insider Purchase / Sale Summary")
    if purchases is not None:
        recs=_records(purchases)[:12]
        if recs:
            keys=list(recs[0].keys())[:6]
            _header(ws,r+1,[str(k) for k in keys]+[""]*(8-len(keys)))
            for rr,rec in enumerate(recs,r+2):
                for c,k in enumerate(keys,1): ws.cell(rr,c,rec.get(k))
        else:
            ws.cell(r+1,1,"No purchase summary returned")
    else:
        ws.cell(r+1,1,"No purchase summary returned")

    widths={"A":31,"B":30,"C":19,"D":19,"E":20,"F":18,"G":42,"H":17}
    for col,w in widths.items(): ws.column_dimensions[col].width=w
    for row in ws.iter_rows():
        for cell in row: cell.alignment=Alignment(vertical="top",wrap_text=True)
    ws.freeze_panes="A5"
    return ws
