from __future__ import annotations

"""Excel chart-compatibility post-pass for generated research workbooks.

The ML and AI chart writers keep compact helper tables outside the main analysis area. Excel
and third-party renderers are sensitive to two details in openpyxl-generated bar charts:

1. hidden source cells are skipped when ``plotVisOnly`` is left at its default; and
2. openpyxl can serialize both category/value axes on the same side unless their positions are
   made explicit, while horizontal bar charts also need their category/value titles and scales
   attached to the correct OOXML axes.

This post-pass repairs both problems after every research run. Helper columns remain technically
visible but extremely narrow/white so the workbook stays clean while chart labels remain valid.
"""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font


HELPER_RANGES = {
    "ML & Quantitative Research": ("X", "AA", "X1:AA40"),
    "AI Growth Forecast": ("P", "Q", "P1:Q30"),
}


def _make_helpers_excel_visible(ws, first_col: str, last_col: str, helper_range: str) -> None:
    start = ws[first_col + "1"].column
    end = ws[last_col + "1"].column
    for col_idx in range(start, end + 1):
        letter = ws.cell(1, col_idx).column_letter
        dim = ws.column_dimensions[letter]
        dim.hidden = False
        dim.width = 1.5
    for row in ws[helper_range]:
        for cell in row:
            cell.font = Font(name="Aptos", size=1, color="FFFFFF")


def _repair_axis_positions(chart, *, horizontal: bool) -> None:
    """Write unambiguous Excel axis positions and include helper cells in the plot."""
    chart.visible_cells_only = False
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.tickLblPos = "nextTo"
    chart.y_axis.tickLblPos = "nextTo"
    if horizontal:
        # openpyxl's x_axis object is the category axis for BarChart even when barDir='bar'.
        chart.x_axis.axPos = "l"
        chart.y_axis.axPos = "b"
    else:
        chart.x_axis.axPos = "b"
        chart.y_axis.axPos = "l"


def _repair_ml_charts(ws) -> int:
    charts = list(getattr(ws, "_charts", []) or [])
    for idx, chart in enumerate(charts):
        horizontal = idx >= 2
        _repair_axis_positions(chart, horizontal=horizontal)
        if idx == 0:
            chart.x_axis.title = None
            chart.y_axis.title = "Percent (%)"
            if chart.legend is not None:
                chart.legend.position = "b"
        elif idx == 1:
            chart.x_axis.title = None
            chart.y_axis.title = "Directional accuracy (%)"
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 100
            chart.y_axis.numFmt = "0.0"
            if chart.legend is not None:
                chart.legend.position = "b"
        elif idx == 2:
            chart.x_axis.title = "Input"
            chart.y_axis.title = "Share of top-driver influence (%)"
            chart.x_axis.scaling.min = None
            chart.x_axis.scaling.max = None
            chart.y_axis.scaling.min = 0
            chart.y_axis.numFmt = "0.0"
        elif idx == 3:
            chart.x_axis.title = "Market state"
            chart.y_axis.title = "Distance-based weight (%)"
            chart.x_axis.scaling.min = None
            chart.x_axis.scaling.max = None
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 100
            chart.y_axis.numFmt = "0.0"
    return len(charts)


def _repair_ai_charts(ws) -> int:
    charts = list(getattr(ws, "_charts", []) or [])
    for idx, chart in enumerate(charts):
        horizontal = idx == 2
        _repair_axis_positions(chart, horizontal=horizontal)
        if idx == 0:
            chart.x_axis.title = None
            chart.y_axis.title = "Supportive score"
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 100
            chart.y_axis.numFmt = "0.0"
        elif idx == 1:
            chart.x_axis.title = None
            chart.y_axis.title = "Annual FCF growth (%)"
            chart.y_axis.numFmt = "0.0"
        elif idx == 2:
            chart.x_axis.title = "Driver"
            chart.y_axis.title = "Share of top-driver influence (%)"
            chart.x_axis.scaling.min = None
            chart.x_axis.scaling.max = None
            chart.y_axis.scaling.min = 0
            chart.y_axis.numFmt = "0.0"
    return len(charts)


def apply_chart_compatibility_fix(workbook_path: str | Path) -> dict:
    path = Path(workbook_path)
    wb = load_workbook(path)
    touched: list[str] = []
    chart_count = 0

    for sheet_name, (first_col, last_col, helper_range) in HELPER_RANGES.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        _make_helpers_excel_visible(ws, first_col, last_col, helper_range)
        if sheet_name == "ML & Quantitative Research":
            chart_count += _repair_ml_charts(ws)
        elif sheet_name == "AI Growth Forecast":
            chart_count += _repair_ai_charts(ws)
        touched.append(sheet_name)

    if touched:
        wb.save(path)
    return {
        "workbook": str(path),
        "chart_helper_sheets": touched,
        "charts_repaired": chart_count,
    }


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Repair ML/AI chart sources and axes for Excel compatibility.")
    parser.add_argument("workbook")
    args = parser.parse_args()
    print(apply_chart_compatibility_fix(args.workbook))
