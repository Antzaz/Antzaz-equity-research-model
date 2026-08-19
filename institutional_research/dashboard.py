from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook


BASE = Path(__file__).resolve().parent
OUT = BASE / "outputs" / "latest"
THESIS_FILE = BASE / "portfolio_thesis.xlsx"
BENCHMARK_FILE = BASE / "benchmark_reference.json"

st.set_page_config(page_title="Institutional Portfolio Research Dashboard", layout="wide")
st.title("Institutional Portfolio Research Dashboard")
st.caption(
    "Private/local portfolio workstation · professional reporting views plus deeper institutional diagnostics."
)

if not OUT.exists() or not (OUT / "summary.json").exists():
    st.error("No analysis outputs found. Run: python run_research.py")
    st.stop()

with open(OUT / "summary.json", "r", encoding="utf-8") as f:
    summary = json.load(f)

portfolio = summary.get("portfolio", {})
mc = summary.get("monte_carlo", {})


def read(name: str) -> pd.DataFrame:
    path = OUT / f"{name}.csv"
    return pd.read_csv(path) if path.exists() else pd.DataFrame()


def pct(value, digits=1):
    return "—" if value is None or pd.isna(value) else f"{float(value):.{digits}%}"


def num(value, digits=2):
    return "—" if value is None or pd.isna(value) else f"{float(value):.{digits}f}"


def multiple(value, digits=1):
    return "—" if value is None or pd.isna(value) else f"{float(value):.{digits}f}x"


def _number(value):
    if value in (None, ""):
        return None
    try:
        if isinstance(value, str) and value.strip().endswith("%"):
            return float(value.strip()[:-1]) / 100.0
        return float(value)
    except Exception:
        return None


def _text(value):
    if value is None:
        return None
    out = str(value).strip()
    return out or None


def load_benchmark_reference() -> dict:
    if not BENCHMARK_FILE.exists():
        return {}
    try:
        return json.loads(BENCHMARK_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def load_local_theses() -> list[dict]:
    if not THESIS_FILE.exists():
        return []
    try:
        wb = load_workbook(THESIS_FILE, data_only=True, read_only=True)
    except Exception:
        return []
    if "Company Theses" not in wb.sheetnames:
        return []

    ws = wb["Company Theses"]
    headers = {
        str(cell.value).strip(): idx + 1
        for idx, cell in enumerate(ws[3])
        if cell.value is not None
    }
    if "Ticker" not in headers:
        return []

    score_fields = {
        "Business Quality (1-5)": "Business quality",
        "Moat (1-5)": "Moat",
        "Management & Capital Allocation (1-5)": "Management / capital allocation",
        "Balance Sheet (1-5)": "Balance sheet",
        "Growth (1-5)": "Growth",
        "Valuation (1-5)": "Valuation",
        "Risk/Resilience (1-5)": "Risk / resilience",
    }
    text_fields = {
        "Status": "status",
        "Time Horizon": "time_horizon",
        "Investment Thesis": "investment_thesis",
        "Why I Own It": "why_owned",
        "Competitive Advantage / Moat": "competitive_advantage",
        "Growth Drivers": "growth_drivers",
        "Valuation Rationale": "valuation_rationale",
        "Catalysts": "catalysts",
        "Key Risks": "key_risks",
        "Falsification / Sell Condition": "sell_condition",
        "Monitoring KPI": "monitoring_kpi",
        "Public Notes": "public_notes",
        "Private Notes": "private_notes",
    }

    out = []
    for row_idx in range(4, ws.max_row + 1):
        ticker = _text(ws.cell(row_idx, headers["Ticker"]).value)
        if not ticker:
            continue
        record = {
            "ticker": ticker.upper(),
            "company": _text(ws.cell(row_idx, headers.get("Company", 0)).value) if headers.get("Company") else ticker,
        }
        if "Conviction (1-5)" in headers:
            record["conviction"] = _number(ws.cell(row_idx, headers["Conviction (1-5)"]).value)
        if "Expected Annual Return %" in headers:
            record["expected_return"] = _number(ws.cell(row_idx, headers["Expected Annual Return %"]).value)
        if "Review Date" in headers:
            v = ws.cell(row_idx, headers["Review Date"]).value
            if isinstance(v, datetime):
                record["review_date"] = v.date().isoformat()
            elif v not in (None, ""):
                record["review_date"] = str(v)

        for excel_name, key in text_fields.items():
            if excel_name in headers:
                record[key] = _text(ws.cell(row_idx, headers[excel_name]).value)

        scores = {}
        for excel_name, label in score_fields.items():
            if excel_name in headers:
                val = _number(ws.cell(row_idx, headers[excel_name]).value)
                if val is not None:
                    scores[label] = val
        record["scores"] = scores
        record["composite_score"] = float(np.mean(list(scores.values()))) if scores else None
        out.append(record)
    return out


def canonical_sector(value: str | None) -> str:
    aliases = {
        "Technology": "Information Technology",
        "Financial Services": "Financials",
        "Consumer Cyclical": "Consumer Discretionary",
        "Healthcare": "Health Care",
        "Consumer Defensive": "Consumer Staples",
        "Basic Materials": "Materials",
    }
    return aliases.get(str(value), str(value) if value not in (None, "", "nan") else "Unknown")


def weighted_average(df: pd.DataFrame, value_col: str, weight_col: str = "Weight"):
    if df.empty or value_col not in df or weight_col not in df:
        return None
    x = pd.to_numeric(df[value_col], errors="coerce")
    w = pd.to_numeric(df[weight_col], errors="coerce")
    ok = x.notna() & w.notna() & (w > 0)
    if not ok.any() or w[ok].sum() <= 0:
        return None
    return float(np.average(x[ok], weights=w[ok]))


def weighted_expected_return(holdings: pd.DataFrame, theses: list[dict]):
    if holdings.empty or not theses or "Ticker" not in holdings:
        return None
    lookup = {x["ticker"]: x.get("expected_return") for x in theses}
    vals = []
    weights = []
    for _, row in holdings.iterrows():
        er = lookup.get(str(row.get("Ticker", "")).upper())
        wt = _number(row.get("Weight"))
        if er is not None and wt is not None and wt > 0:
            vals.append(er)
            weights.append(wt)
    return float(np.average(vals, weights=weights)) if weights else None


def prep_timeseries() -> pd.DataFrame:
    ts = read("portfolio_timeseries")
    if ts.empty or "Date" not in ts:
        return ts
    ts = ts.copy()
    ts["Date"] = pd.to_datetime(ts["Date"], errors="coerce", utc=True).dt.tz_localize(None)
    return ts.dropna(subset=["Date"]).sort_values("Date")


def compounded_return(ts: pd.DataFrame, return_col: str, start=None, end=None):
    if ts.empty or return_col not in ts:
        return None
    view = ts
    if start is not None:
        view = view[view["Date"] >= pd.Timestamp(start)]
    if end is not None:
        view = view[view["Date"] <= pd.Timestamp(end)]
    r = pd.to_numeric(view[return_col], errors="coerce").dropna()
    if r.empty:
        return None
    return float((1 + r).prod() - 1)


def annualized_period_return(ts: pd.DataFrame, return_col: str, start=None):
    if ts.empty:
        return None
    view = ts if start is None else ts[ts["Date"] >= pd.Timestamp(start)]
    r = pd.to_numeric(view.get(return_col), errors="coerce").dropna()
    if r.empty or len(view) < 2:
        return None
    total = float((1 + r).prod() - 1)
    years = (view["Date"].max() - view["Date"].min()).days / 365.25
    return float((1 + total) ** (1 / years) - 1) if years > 1.25 else total


def performance_periods(ts: pd.DataFrame):
    if ts.empty:
        return []
    end = ts["Date"].max()
    quarter_month = ((end.month - 1) // 3) * 3 + 1
    starts = {
        "MTD": end.replace(day=1),
        "QTD": pd.Timestamp(end.year, quarter_month, 1),
        "YTD": pd.Timestamp(end.year, 1, 1),
        "1Y": end - pd.DateOffset(years=1),
        "3Y": end - pd.DateOffset(years=3),
        "5Y": end - pd.DateOffset(years=5),
    }
    rows = []
    for label in ["MTD", "QTD", "YTD", "1Y", "3Y", "5Y", "Since analysis start"]:
        start = None if label == "Since analysis start" else starts[label]
        annualized = label in {"3Y", "5Y", "Since analysis start"}
        fn = annualized_period_return if annualized else compounded_return
        p = fn(ts, "PortfolioReturn", start)
        b = fn(ts, "BenchmarkReturn", start)
        rows.append({
            "Period": label + (" (ann.)" if annualized else ""),
            "Portfolio": p,
            "Benchmark": b,
            "Excess": p - b if p is not None and b is not None else None,
        })
    return rows


def calendar_returns(ts: pd.DataFrame) -> pd.DataFrame:
    if ts.empty:
        return pd.DataFrame()
    rows = []
    for year, group in ts.groupby(ts["Date"].dt.year):
        p = compounded_return(group, "PortfolioReturn")
        b = compounded_return(group, "BenchmarkReturn")
        rows.append({"Year": int(year), "Portfolio": p, "Benchmark": b, "Excess": p - b if p is not None and b is not None else None})
    return pd.DataFrame(rows).sort_values("Year", ascending=False)


def monthly_return_matrix(ts: pd.DataFrame) -> pd.DataFrame:
    if ts.empty:
        return pd.DataFrame()
    rows = []
    work = ts.copy()
    work["Year"] = work["Date"].dt.year
    work["Month"] = work["Date"].dt.month
    for (year, month), group in work.groupby(["Year", "Month"]):
        rows.append({"Year": year, "Month": month, "Return": compounded_return(group, "PortfolioReturn")})
    if not rows:
        return pd.DataFrame()
    m = pd.DataFrame(rows).pivot(index="Year", columns="Month", values="Return")
    m = m.reindex(columns=range(1, 13))
    m.columns = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    return m.tail(8)


def drawdown_frame(ts: pd.DataFrame) -> pd.DataFrame:
    if ts.empty:
        return pd.DataFrame()
    out = pd.DataFrame({"Date": ts["Date"]})
    for growth_col, name in [("PortfolioGrowth", "Portfolio"), ("BenchmarkGrowth", "Benchmark")]:
        if growth_col in ts:
            g = pd.to_numeric(ts[growth_col], errors="coerce")
            out[name] = g / g.cummax() - 1
    return out


def format_percent_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    out = df.copy()
    for col in columns:
        if col in out:
            out[col] = out[col].map(lambda x: pct(x) if pd.notna(x) else "—")
    return out


holdings = read("holdings_analysis")
sector = read("sector_exposure")
ts = prep_timeseries()
theses = load_local_theses()
benchmark_ref = load_benchmark_reference()
expected_portfolio_return = weighted_expected_return(holdings, theses)

# A professional summary strip that remains useful across all tabs.
periods = performance_periods(ts)
ytd = next((x for x in periods if x["Period"] == "YTD"), {})
summary_cols = st.columns(8)
summary_cols[0].metric("YTD model", pct(ytd.get("Portfolio")))
summary_cols[1].metric("YTD benchmark", pct(ytd.get("Benchmark")))
summary_cols[2].metric("YTD excess", pct(ytd.get("Excess")))
summary_cols[3].metric("Ann. volatility", pct(portfolio.get("annualized_volatility")))
summary_cols[4].metric("Sharpe", num(portfolio.get("sharpe")))
summary_cols[5].metric("Max drawdown", pct(portfolio.get("max_drawdown")))
summary_cols[6].metric("Expected return", pct(expected_portfolio_return))
summary_cols[7].metric("Tracking error", pct(portfolio.get("tracking_error")))

st.info(
    "Performance note: historical analytics apply the current portfolio weights across available adjusted price history. "
    "They are research/model diagnostics, not a transaction-weighted realized client track record. "
    "Listing-currency returns are not fully FX-normalized to a single reporting currency."
)

TABS = st.tabs([
    "Overview",
    "Performance",
    "Portfolio",
    "Company Theses",
    "Risk & Attribution",
    "Equity Research",
    "Construction",
    "Factors",
    "Liquidity",
    "Monte Carlo",
    "Forecasts",
    "Methodology",
    "Power BI",
])

with TABS[0]:
    st.subheader("Portfolio overview")
    left, right = st.columns([1.5, 1])
    with left:
        if not ts.empty:
            growth_cols = [c for c in ["PortfolioGrowth", "BenchmarkGrowth"] if c in ts]
            growth = ts[["Date"] + growth_cols].melt("Date", var_name="Series", value_name="Growth")
            growth["Series"] = growth["Series"].replace({"PortfolioGrowth": "Model portfolio", "BenchmarkGrowth": "Benchmark"})
            fig = px.line(growth, x="Date", y="Growth", color="Series", title="Model portfolio vs benchmark")
            st.plotly_chart(fig, use_container_width=True)
    with right:
        c1, c2 = st.columns(2)
        c1.metric("Holdings", int(len(holdings)))
        c2.metric("Top 5 weight", pct(portfolio.get("top_5_weight")))
        c3, c4 = st.columns(2)
        c3.metric("Effective holdings", num(portfolio.get("effective_number_of_holdings"), 1))
        c4.metric("Largest risk share", pct(portfolio.get("largest_risk_contribution")))
        c5, c6 = st.columns(2)
        c5.metric("Information ratio", num(portfolio.get("information_ratio")))
        c6.metric("Beta", num(portfolio.get("beta")))

    left, right = st.columns(2)
    with left:
        if not holdings.empty:
            fig = px.bar(
                holdings.sort_values("Weight"),
                x="Weight",
                y="Company" if "Company" in holdings else "Ticker",
                orientation="h",
                title="Portfolio weights",
            )
            fig.update_xaxes(tickformat=".0%")
            st.plotly_chart(fig, use_container_width=True)
    with right:
        if not sector.empty:
            fig = px.bar(sector, x="Sector", y="Weight", title="Sector allocation")
            fig.update_yaxes(tickformat=".0%")
            st.plotly_chart(fig, use_container_width=True)

    st.subheader("Portfolio characteristics")
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Forward P/E", multiple(weighted_average(holdings, "ForwardPE")))
    c2.metric("Revenue growth", pct(weighted_average(holdings, "RevenueGrowth")))
    c3.metric("Operating margin", pct(weighted_average(holdings, "OperatingMargin")))
    c4.metric("ROE", pct(weighted_average(holdings, "ROE")))
    thesis_scores = [x.get("composite_score") for x in theses if x.get("composite_score") is not None]
    thesis_weights = []
    thesis_values = []
    if thesis_scores and not holdings.empty:
        weight_lookup = holdings.set_index("Ticker")["Weight"].to_dict() if "Ticker" in holdings else {}
        for x in theses:
            if x.get("composite_score") is not None and x["ticker"] in weight_lookup:
                thesis_values.append(x["composite_score"])
                thesis_weights.append(weight_lookup[x["ticker"]])
    weighted_score = float(np.average(thesis_values, weights=thesis_weights)) if thesis_weights else None
    c5.metric("Research score", "—" if weighted_score is None else f"{weighted_score:.2f} / 5")

with TABS[1]:
    st.subheader("Standardized performance")
    if periods:
        ptable = pd.DataFrame(periods)
        st.dataframe(format_percent_columns(ptable, ["Portfolio", "Benchmark", "Excess"]), use_container_width=True, hide_index=True)

    left, right = st.columns(2)
    with left:
        if not ts.empty:
            growth_cols = [c for c in ["PortfolioGrowth", "BenchmarkGrowth"] if c in ts]
            growth = ts[["Date"] + growth_cols].melt("Date", var_name="Series", value_name="Growth")
            growth["Series"] = growth["Series"].replace({"PortfolioGrowth": "Model portfolio", "BenchmarkGrowth": "Benchmark"})
            st.plotly_chart(px.line(growth, x="Date", y="Growth", color="Series", title="Cumulative growth"), use_container_width=True)
    with right:
        dd = drawdown_frame(ts)
        if not dd.empty:
            melted = dd.melt("Date", var_name="Series", value_name="Drawdown")
            fig = px.area(melted, x="Date", y="Drawdown", color="Series", title="Drawdown history")
            fig.update_yaxes(tickformat=".0%")
            st.plotly_chart(fig, use_container_width=True)

    left, right = st.columns(2)
    with left:
        cal = calendar_returns(ts)
        st.subheader("Calendar-year returns")
        if not cal.empty:
            st.dataframe(format_percent_columns(cal, ["Portfolio", "Benchmark", "Excess"]), use_container_width=True, hide_index=True)
    with right:
        st.subheader("Monthly return heatmap")
        monthly = monthly_return_matrix(ts)
        if not monthly.empty:
            fig = px.imshow(monthly, text_auto=".1%", aspect="auto", color_continuous_midpoint=0)
            st.plotly_chart(fig, use_container_width=True)

    rolling = read("rolling_risk")
    if not rolling.empty:
        st.subheader("Rolling risk")
        st.dataframe(rolling, use_container_width=True, hide_index=True)

with TABS[2]:
    st.subheader("Portfolio positioning")
    if not holdings.empty:
        display_cols = [
            "Ticker", "Company", "Sector", "Country", "Currency", "Weight", "Shares",
            "CurrentPrice", "AverageCost", "MarketValue", "UnrealizedPnL", "UnrealizedPnLPct",
            "ForwardPE", "RevenueGrowth", "OperatingMargin", "ROE", "RiskContributionPct",
        ]
        st.dataframe(holdings[[c for c in display_cols if c in holdings.columns]], use_container_width=True, hide_index=True)

    left, right = st.columns(2)
    with left:
        st.subheader("Sector vs benchmark")
        if not holdings.empty and benchmark_ref.get("sector_weights"):
            p_sector = holdings.assign(_sector=holdings["Sector"].map(canonical_sector)).groupby("_sector")["Weight"].sum().to_dict()
            bench = benchmark_ref.get("sector_weights", {})
            names = sorted(set(p_sector) | set(bench))
            rows = [{"Sector": s, "Portfolio": p_sector.get(s, 0.0), "Benchmark": bench.get(s, 0.0), "Active": p_sector.get(s, 0.0) - bench.get(s, 0.0)} for s in names if s != "Unknown"]
            sec_df = pd.DataFrame(rows)
            chart = sec_df.melt("Sector", value_vars=["Portfolio", "Benchmark"], var_name="Series", value_name="Weight")
            fig = px.bar(chart, x="Sector", y="Weight", color="Series", barmode="group", title="Portfolio vs S&P 500 sector weights")
            fig.update_yaxes(tickformat=".0%")
            st.plotly_chart(fig, use_container_width=True)
            st.dataframe(format_percent_columns(sec_df, ["Portfolio", "Benchmark", "Active"]), use_container_width=True, hide_index=True)
            st.caption(f"Benchmark sector reference: {benchmark_ref.get('benchmark', 'S&P 500')} as of {benchmark_ref.get('as_of', '—')}.")
        else:
            st.info("Benchmark sector comparison will appear when benchmark_reference.json and refreshed holdings are available.")
    with right:
        st.subheader("Geographic exposure")
        if not holdings.empty and "Country" in holdings and holdings["Country"].notna().any():
            geo = holdings.assign(Country=holdings["Country"].fillna("Unknown")).groupby("Country", as_index=False)["Weight"].sum().sort_values("Weight", ascending=False)
            fig = px.bar(geo, x="Country", y="Weight", title="Country allocation")
            fig.update_yaxes(tickformat=".0%")
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Country exposure will populate after you rerun the updated portfolio research engine.")

    risk_budget = read("risk_budget")
    if not risk_budget.empty:
        fig = px.bar(risk_budget, x="Ticker", y=["Weight", "RiskContributionPct"], barmode="group", title="Capital weight vs risk contribution")
        fig.update_yaxes(tickformat=".0%")
        st.plotly_chart(fig, use_container_width=True)

with TABS[3]:
    st.subheader("Company investment theses")
    if not theses:
        st.info("No local portfolio_thesis.xlsx found. Save the thesis workbook under institutional_research/portfolio_thesis.xlsx.")
    else:
        labels = [f"{x['ticker']} — {x.get('company') or x['ticker']}" for x in theses]
        selected_label = st.selectbox("Company", labels, key="local_thesis_company")
        thesis = theses[labels.index(selected_label)]
        ticker = thesis["ticker"]
        hrow = holdings[holdings["Ticker"].astype(str).str.upper() == ticker].iloc[0] if not holdings.empty and "Ticker" in holdings and (holdings["Ticker"].astype(str).str.upper() == ticker).any() else None

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Portfolio weight", pct(hrow.get("Weight")) if hrow is not None else "—")
        c2.metric("Expected return", pct(thesis.get("expected_return")))
        c3.metric("Conviction", "—" if thesis.get("conviction") is None else f"{thesis['conviction']:.1f} / 5")
        c4.metric("Research score", "—" if thesis.get("composite_score") is None else f"{thesis['composite_score']:.2f} / 5")
        c5.metric("Review date", thesis.get("review_date") or "—")

        scores = thesis.get("scores", {})
        if scores:
            s_df = pd.DataFrame({"Dimension": list(scores.keys()), "Score": list(scores.values())})
            fig = px.bar(s_df, x="Dimension", y="Score", range_y=[0, 5.5], text="Score", title="Investment scorecard")
            st.plotly_chart(fig, use_container_width=True)

        decision1, decision2, decision3 = st.columns(3)
        with decision1:
            st.markdown("**Why I own it**")
            st.write(thesis.get("why_owned") or "—")
        with decision2:
            st.markdown("**What changes my mind**")
            st.write(thesis.get("sell_condition") or "—")
        with decision3:
            st.markdown("**Monitoring KPI**")
            st.write(thesis.get("monitoring_kpi") or "—")

        fields = [
            ("Investment Thesis", "investment_thesis"),
            ("Competitive Advantage / Moat", "competitive_advantage"),
            ("Growth Drivers", "growth_drivers"),
            ("Valuation Rationale", "valuation_rationale"),
            ("Catalysts", "catalysts"),
            ("Key Risks", "key_risks"),
            ("Falsification / Sell Condition", "sell_condition"),
            ("Public Notes", "public_notes"),
            ("Private Notes", "private_notes"),
        ]
        for title, key in fields:
            value = thesis.get(key)
            if value:
                with st.expander(title, expanded=key in {"investment_thesis", "key_risks"}):
                    st.write(value)

with TABS[4]:
    st.subheader("Risk & attribution")
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Daily VaR 95%", pct(portfolio.get("daily_var_95"), 2))
    c2.metric("Expected Shortfall 95%", pct(portfolio.get("daily_expected_shortfall_95"), 2))
    c3.metric("Beta", num(portfolio.get("beta")))
    c4.metric("Tracking error", pct(portfolio.get("tracking_error")))
    c5.metric("Info ratio", num(portfolio.get("information_ratio")))
    c6.metric("Active return", pct(portfolio.get("active_annualized_return")))

    left, right = st.columns(2)
    with left:
        attribution = read("return_attribution")
        if not attribution.empty:
            fig = px.bar(attribution, x="Ticker", y="StaticArithmeticContribution", title="Static-weight return contribution")
            fig.update_yaxes(tickformat=".1%")
            st.plotly_chart(fig, use_container_width=True)
            st.dataframe(attribution, use_container_width=True, hide_index=True)
            st.caption("Current weights applied to historical returns; diagnostic attribution, not transaction-level realized attribution.")
    with right:
        risk = read("risk_contribution")
        if not risk.empty:
            fig = px.bar(risk, x="Ticker", y="RiskContributionPct", title="Risk contribution")
            fig.update_yaxes(tickformat=".0%")
            st.plotly_chart(fig, use_container_width=True)

    alpha = read("alpha_summary")
    if not alpha.empty:
        st.subheader("Factor-adjusted alpha")
        alpha_cols = [c for c in ["Model", "AnnualizedAlpha", "AlphaTStat", "AlphaPValue", "R2", "Significant5Pct", "Interpretation"] if c in alpha]
        st.dataframe(alpha[alpha_cols], use_container_width=True, hide_index=True)

    rel = read("benchmark_relative")
    if not rel.empty:
        st.subheader("Benchmark-relative metrics")
        st.dataframe(rel, use_container_width=True, hide_index=True)

    left, right = st.columns(2)
    with left:
        stress = read("stress_tests")
        if not stress.empty and "Ticker" in stress:
            ps = stress[stress["Ticker"] == "PORTFOLIO"]
            if not ps.empty:
                fig = px.bar(ps, x="Scenario", y="EstimatedHoldingReturn", title="Forward stress scenarios")
                fig.update_yaxes(tickformat=".0%")
                st.plotly_chart(fig, use_container_width=True)
    with right:
        historical = read("historical_stress_windows")
        if not historical.empty:
            st.markdown("**Historical stress windows**")
            st.dataframe(historical, use_container_width=True, hide_index=True)

with TABS[5]:
    st.subheader("Equity research library")
    reverse_dcf = read("reverse_dcf")
    factor_scores = read("factor_scores")
    if holdings.empty:
        st.info("No holdings output available.")
    else:
        tickers = holdings["Ticker"].astype(str).tolist()
        selected = st.selectbox("Research company", tickers, key="equity_research_company")
        row = holdings[holdings["Ticker"].astype(str) == selected].iloc[0]
        thesis = next((x for x in theses if x["ticker"] == selected.upper()), {})
        dcf_row = reverse_dcf[reverse_dcf["Ticker"].astype(str) == selected].iloc[0] if not reverse_dcf.empty and "Ticker" in reverse_dcf and (reverse_dcf["Ticker"].astype(str) == selected).any() else None
        frow = factor_scores[factor_scores["Ticker"].astype(str) == selected].iloc[0] if not factor_scores.empty and "Ticker" in factor_scores and (factor_scores["Ticker"].astype(str) == selected).any() else None

        st.markdown(f"### {row.get('Company', selected)}")
        c1, c2, c3, c4, c5, c6 = st.columns(6)
        c1.metric("Expected return", pct(thesis.get("expected_return")))
        c2.metric("Forward P/E", multiple(row.get("ForwardPE")))
        c3.metric("Revenue growth", pct(row.get("RevenueGrowth")))
        c4.metric("Operating margin", pct(row.get("OperatingMargin")))
        c5.metric("ROE", pct(row.get("ROE")))
        c6.metric("Beta", num(row.get("Beta")))

        left, right = st.columns(2)
        with left:
            st.markdown("**Investment case**")
            st.write(thesis.get("investment_thesis") or thesis.get("why_owned") or "Add the thesis in portfolio_thesis.xlsx.")
            st.markdown("**Valuation rationale**")
            st.write(thesis.get("valuation_rationale") or "—")
            st.markdown("**Key risks**")
            st.write(thesis.get("key_risks") or "—")
        with right:
            if dcf_row is not None:
                st.markdown("**Reverse DCF / market-implied expectations**")
                dcf_fields = ["ImpliedAnnualFCFGrowth", "WACC", "TerminalGrowth", "ForecastYears", "Status"]
                st.dataframe(pd.DataFrame([{c: dcf_row.get(c) for c in dcf_fields if c in dcf_row.index}]), use_container_width=True, hide_index=True)
            if frow is not None:
                st.markdown("**Factor scorecard**")
                factor_cols = [c for c in ["ValueScore", "QualityScore", "GrowthScore", "MomentumScore", "LowVolatilityScore", "CompositeScore"] if c in frow.index]
                fdf = pd.DataFrame({"Factor": factor_cols, "Score": [frow.get(c) for c in factor_cols]})
                st.plotly_chart(px.bar(fdf, x="Factor", y="Score", title="Research factor scores"), use_container_width=True)

        st.caption(
            "For full filing/segment/forecast/valuation workbooks, use the Company Research view in the private portal when the encrypted live bundle is available."
        )

with TABS[6]:
    st.subheader("Portfolio construction")
    constraints = read("constraint_report")
    optimization = read("portfolio_optimizations")
    expected = read("expected_returns_inputs")
    active = read("active_share_detail")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Active return", pct(portfolio.get("active_annualized_return")))
    c2.metric("Tracking error", pct(portfolio.get("tracking_error")))
    c3.metric("Information ratio", num(portfolio.get("information_ratio")))
    c4.metric("Daily active hit rate", pct(portfolio.get("daily_active_hit_rate")))

    if not constraints.empty:
        st.markdown("**Policy / constraint monitor**")
        st.dataframe(constraints, use_container_width=True, hide_index=True)
    if not optimization.empty:
        choices = optimization["Portfolio"].dropna().unique().tolist()
        selected = st.selectbox("Portfolio construction method", choices, key="construction_method")
        view = optimization[optimization["Portfolio"] == selected].copy()
        fig = px.bar(view, x="Ticker", y=["CurrentWeight", "TargetWeight"], barmode="group", title=f"{selected}: current vs target weights")
        fig.update_yaxes(tickformat=".0%")
        st.plotly_chart(fig, use_container_width=True)
        st.dataframe(view, use_container_width=True, hide_index=True)
    else:
        st.info("Optimization outputs will appear after a successful research run with sufficient history.")
    if not expected.empty:
        st.markdown("**Expected-return / conviction inputs**")
        st.dataframe(expected, use_container_width=True, hide_index=True)
    if not active.empty:
        st.markdown("**Active Share detail**")
        st.dataframe(active, use_container_width=True, hide_index=True)

with TABS[7]:
    st.subheader("Factors")
    scores = read("factor_scores")
    exposure = read("factor_exposure")
    proxy = read("factor_proxy_sensitivity")
    if not exposure.empty:
        st.plotly_chart(px.bar(exposure, x="Factor", y="PortfolioExposure", title="Holding-relative factor scores"), use_container_width=True)
    if not proxy.empty:
        st.plotly_chart(px.bar(proxy, x="Proxy", y="BetaToProxy", title="Portfolio beta to public factor proxies"), use_container_width=True)
        st.dataframe(proxy, use_container_width=True, hide_index=True)
        st.caption("ETF proxy betas are diagnostics, not a commercial Barra/Axioma-style risk model.")
    if not scores.empty:
        factor_cols = ["Ticker", "ValueScore", "QualityScore", "GrowthScore", "MomentumScore", "LowVolatilityScore", "CompositeScore", "PortfolioWeight"]
        st.dataframe(scores[[c for c in factor_cols if c in scores]], use_container_width=True, hide_index=True)

with TABS[8]:
    st.subheader("Liquidity & capacity")
    liquidity = read("liquidity_analysis")
    if not liquidity.empty:
        st.plotly_chart(px.bar(liquidity, x="Ticker", y="EstimatedDaysToLiquidate", title="Estimated days to liquidate at configured ADV participation"), use_container_width=True)
        st.dataframe(liquidity, use_container_width=True, hide_index=True)
        st.caption("Public average-volume snapshots provide a first-pass capacity estimate, not an execution guarantee.")
    else:
        st.info("Liquidity output is not available in the latest research run.")

with TABS[9]:
    st.subheader("Monte Carlo")
    dist = read("monte_carlo_distribution")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Median 1Y return", pct(mc.get("median_return")))
    c2.metric("5th percentile", pct(mc.get("p05_return")))
    c3.metric("P(loss)", pct(mc.get("probability_loss")))
    c4.metric("P(loss >20%)", pct(mc.get("probability_loss_20pct")))
    if not dist.empty:
        fig = px.histogram(dist, x="TerminalReturn", nbins=80, title="Bootstrap Monte Carlo terminal return distribution")
        fig.update_xaxes(tickformat=".0%")
        st.plotly_chart(fig, use_container_width=True)
    st.caption("Historical bootstrap is a scenario engine, not a forecast.")

with TABS[10]:
    st.subheader("Forecast tracking")
    forecasts = read("forecast_accuracy")
    if forecasts.empty:
        st.info("No forecasts logged yet. Add rows to forecasts.csv before outcomes are known, then fill Actual and rerun.")
    else:
        completed = forecasts[forecasts["Actual"].notna()].copy() if "Actual" in forecasts else pd.DataFrame()
        if not completed.empty:
            c1, c2 = st.columns(2)
            c1.metric("Mean abs. forecast error", pct(completed["AbsoluteErrorPct"].mean()))
            c2.metric("Median abs. forecast error", pct(completed["AbsoluteErrorPct"].median()))
        st.dataframe(forecasts, use_container_width=True, hide_index=True)

with TABS[11]:
    st.subheader("Methodology & data integrity")
    st.markdown(
        """
        **Performance basis**  
        Current portfolio weights are applied across the available adjusted price history. The historical series is therefore a research/model diagnostic rather than a transaction-weighted realized track record.

        **Currency basis**  
        Current market-data downloads use listing-currency adjusted prices. Cross-currency holdings are not fully translated into a single reporting currency in the historical model.

        **Benchmark**  
        The configured market benchmark is used for relative return, beta, tracking error, information ratio and alpha diagnostics. A dated S&P 500 sector reference is used only for positioning comparisons.

        **Attribution**  
        Static return attribution uses current portfolio weights. It should not be interpreted as realized manager attribution unless point-in-time weights/transactions are supplied.

        **Research process**  
        The engine combines price history, public company information, factor diagnostics, reverse DCF, stress tests, Monte Carlo analysis, benchmark-relative metrics, alpha models, portfolio constraints and forecast tracking.
        """
    )
    config_path = BASE / "config.json"
    if config_path.exists():
        with st.expander("Current research configuration"):
            st.json(json.loads(config_path.read_text(encoding="utf-8")))
    st.markdown("**Private/local-only fields available here:** shares, average cost, market value, unrealized P&L, private thesis notes and detailed construction diagnostics.")

with TABS[12]:
    st.subheader("Power BI-ready output")
    st.write("Every analysis table is exported as a flat CSV under outputs/latest. Power BI can connect directly to this folder or individual CSV files.")
    st.code(str(OUT))
    st.write(
        "Professional outputs include holdings_analysis, portfolio_timeseries, risk_budget, benchmark_relative, concentration_summary, liquidity_analysis, factor_proxy_sensitivity, rolling_risk, historical_stress_windows, return_attribution, alpha_summary, portfolio_optimizations, constraint_report and active_share_detail."
    )
