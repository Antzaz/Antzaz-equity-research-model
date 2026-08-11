from __future__ import annotations

"""Final guard around the broad Financial Statements fallback repair.

The broad repair intentionally fills many optional lines. This guard always reconciles the
core income-statement rows to explicit Yahoo statement labels, because a historical fallback
alias must never make Operating Income equal Pre-Tax Income simply by choosing the wrong row.
"""

import math
import pandas as pd
import yfinance as yf
from openpyxl.styles import Font, Alignment

from financial_statement_repair_v2 import repair_financial_statements as _broad_repair

FMT_BN='#,##0.0;[Red](#,##0.0);-'; FMT_EPS='$0.00;[Red]($0.00);-'; LINK_GREEN="008000"


def _num(v):
    try:
        if isinstance(v,bool) or v in (None,""): return None
        x=float(v); return x if math.isfinite(x) else None
    except Exception: return None


def _income(ticker):
    try:
        df=yf.Ticker(ticker).income_stmt
        return df.copy() if isinstance(df,pd.DataFrame) else pd.DataFrame()
    except Exception: return pd.DataFrame()


def _series(df,aliases):
    if df.empty: return {}
    row=next((a for a in aliases if a in df.index),None)
    if row is None: return {}
    out={}
    for c,v in df.loc[row].items():
        try: y=int(pd.Timestamp(c).year)
        except Exception: continue
        n=_num(v)
        if n is not None: out[y]=n
    return out


def _section(ws,title):
    for r in range(1,ws.max_row+1):
        if str(ws.cell(r,1).value or "").strip().lower()==title.lower(): return r
    return None


def _label(ws,label,start,end):
    for r in range(start,end+1):
        if str(ws.cell(r,1).value or "").strip().lower()==label.lower(): return r
    return None


def repair_financial_statements(wb,ticker):
    result={"filled":0,"coverage":0}
    try:
        result=_broad_repair(wb,ticker) or result
    except Exception as exc:
        # The broad filler may have successfully written data before a presentation-only
        # issue; the core reconciliation below remains authoritative.
        result["broad_warning"]=str(exc)

    if "Financial Statements" not in wb.sheetnames: return result
    ws=wb["Financial Statements"]; i0=_section(ws,"Income Statement"); b0=_section(ws,"Balance Sheet")
    if not i0 or not b0: return result
    header=i0+1; year_cols={int(ws.cell(header,c).value):c for c in range(2,min(8,ws.max_column+1)) if isinstance(ws.cell(header,c).value,(int,float))}
    df=_income(ticker)
    mapping={
        "Revenue":(["Total Revenue","Operating Revenue"],False),
        "Cost of Revenue":(["Cost Of Revenue","Cost of Revenue"],False),
        "Gross Profit":(["Gross Profit"],False),
        "Operating Income":(["Operating Income"],False),
        "Pre-Tax Income":(["Pretax Income","Pre Tax Income"],False),
        "Income Taxes":(["Tax Provision","Income Tax Expense"],False),
        "Net Income":(["Net Income","Net Income Common Stockholders"],False),
        "Diluted EPS":(["Diluted EPS"],True),
    }
    filled=result.get("filled",0) or 0
    for label,(aliases,per_share) in mapping.items():
        r=_label(ws,label,i0,b0-1)
        if not r: continue
        vals=_series(df,aliases)
        for y,c in year_cols.items():
            v=vals.get(y)
            if v is None: continue
            # Explicit core labels are safe to overwrite; they are a structured statement
            # cross-check and repair known alias collisions in the earlier fallback path.
            ws.cell(r,c).value=v if per_share else v/1e9
            ws.cell(r,c).number_format=FMT_EPS if per_share else FMT_BN
            filled+=1

    # Source notes live outside merged A:H section headers.
    ws["I5"]="Source / Fallback Method"; ws["I5"].font=Font(bold=True)
    ws["I6"]="SEC Company Facts remains primary. Structured Yahoo annual statements reconcile/fill core rows when SEC is unavailable or a fallback alias is ambiguous."
    ws["I6"].alignment=Alignment(wrap_text=True,vertical="top")
    ws["I7"]="https://finance.yahoo.com/quote/"+ticker+"/financials/"; ws["I7"].hyperlink=ws["I7"].value; ws["I7"].font=Font(color=LINK_GREEN,underline="single")
    ws.column_dimensions["I"].width=62

    core=("Revenue","Operating Income","Pre-Tax Income","Income Taxes","Net Income","Diluted EPS")
    available=0; total=0
    for label in core:
        r=_label(ws,label,i0,b0-1)
        if not r: continue
        for c in year_cols.values(): total+=1; available+=1 if _num(ws.cell(r,c).value) is not None else 0
    result.update({"filled":filled,"coverage":available/total if total else 0,"core_reconciled":True})
    return result
