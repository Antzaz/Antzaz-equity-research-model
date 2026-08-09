"""Verified CVS filing fallback for the cross-company equity research workbook.

CVS discloses segment revenue and Adjusted Operating Income for four reportable segments.
The generic SEC HTML parser can fail on some local environments, so this narrow fallback
repairs only CVS and only with directly disclosed FY2023-FY2025 data. It also fills a few
high-confidence balance-sheet gaps that SEC Company Facts may miss.
"""

from openpyxl.styles import Font, Alignment

SEC_2025="https://www.sec.gov/Archives/edgar/data/64803/000006480326000010/cvs-20251231.htm"
INPUT_BLUE="0000FF"; LINK_GREEN="008000"; GREY="666666"; BLACK="000000"
FMT_BN='#,##0.0;[Red](#,##0.0);-'; FMT_PCT='0.0%;[Red](0.0%);-'

SEGMENTS=[
    ("Health Care Benefits",[105.646,130.665,143.354],[5.577,.307,2.939]),
    ("Health Services",[186.843,173.605,190.425],[7.312,7.243,7.151]),
    ("Pharmacy & Consumer Wellness",[116.763,124.500,139.367],[5.963,5.774,6.040]),
    ("Corporate/Other",[.451,.451,.484],[-1.318,-1.348,-1.687]),
]
BUSINESS=[
    ("Pharmacy",[223.452,210.272,229.025]),
    ("Front Store",[22.458,21.522,21.459]),
    ("Premiums",[99.192,122.896,134.751]),
    ("Net investment income",[1.153,2.153,2.233]),
    ("Other",[11.521,15.966,14.599]),
]
BS_GAPS={
    "Other Current Assets":[2.636,3.151,3.076,5.091],
    "Property & Equipment, Net":[12.873,13.183,12.993,13.083],
    "Accounts Payable":[14.838,14.897,15.892,17.641],
    "Long-Term Debt":[50.476,58.638,60.527,60.502],
}


def _find_row(ws,label):
    target=str(label).strip().lower()
    for r in range(1,ws.max_row+1):
        if str(ws.cell(r,1).value or "").strip().lower()==target: return r
    return None


def _repair_segment_analysis(wb):
    if "Segment Analysis" not in wb.sheetnames: return
    ws=wb["Segment Analysis"]
    ws["A3"]=("Standardized issuer-disclosed segment schema. Status: AUTO/FALLBACK — verified CVS FY2025 10-K. "
              "CVS's principal segment performance measure is Adjusted Operating Income; segment revenue includes intersegment activity before eliminations.")
    ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True)
    headers=["Segment","2023 Revenue","2024 Revenue","2025 Revenue","2025 Growth","2023–2025 CAGR","2023 Adj. Op. Income","2024 Adj. Op. Income","2025 Adj. Op. Income","2023 Adj. Margin","2024 Adj. Margin","2025 Adj. Margin","Margin Δ","2025 Revenue Mix"]
    for c,v in enumerate(headers,1): ws.cell(6,c,v)
    # Remove stale/manual rows before repopulating.
    for r in range(7,17):
        for c in range(1,15): ws.cell(r,c).value=None
    for r,(name,rev,op) in enumerate(SEGMENTS,7):
        ws.cell(r,1,name)
        for c,v in enumerate(rev,2): ws.cell(r,c,v); ws.cell(r,c).font=Font(color=INPUT_BLUE); ws.cell(r,c).number_format=FMT_BN
        ws.cell(r,5,f'=IFERROR(D{r}/C{r}-1,"")'); ws.cell(r,6,f'=IFERROR((D{r}/B{r})^(1/2)-1,"")')
        for c,v in enumerate(op,7): ws.cell(r,c,v); ws.cell(r,c).font=Font(color=INPUT_BLUE); ws.cell(r,c).number_format=FMT_BN
        ws.cell(r,10,f'=IFERROR(G{r}/B{r},"")'); ws.cell(r,11,f'=IFERROR(H{r}/C{r},"")'); ws.cell(r,12,f'=IFERROR(I{r}/D{r},"")'); ws.cell(r,13,f'=IFERROR(L{r}-K{r},"")'); ws.cell(r,14,f'=IFERROR(D{r}/SUM($D$7:$D$10),"")')
        for c in (5,6,10,11,12,13,14): ws.cell(r,c).number_format=FMT_PCT

    # Find the standardized business section; v2 currently places it on row 19.
    section=_find_row(ws,"Revenue by Business Line") or 19; header=section+1
    business_headers=["Business Line / Revenue Group","2023","2024","2025","2025 Growth","2023–2025 CAGR","2025 Mix","Source / Notes"]
    for c,v in enumerate(business_headers,1): ws.cell(header,c,v)
    for r in range(header+1,header+11):
        for c in range(1,9): ws.cell(r,c).value=None
    for r,(name,vals) in enumerate(BUSINESS,header+1):
        ws.cell(r,1,name)
        for c,v in enumerate(vals,2): ws.cell(r,c,v); ws.cell(r,c).font=Font(color=INPUT_BLUE); ws.cell(r,c).number_format=FMT_BN
        ws.cell(r,5,f'=IFERROR(D{r}/C{r}-1,"")'); ws.cell(r,6,f'=IFERROR((D{r}/B{r})^(1/2)-1,"")'); ws.cell(r,7,f'=IFERROR(D{r}/SUM($D${header+1}:$D${header+5}),"")'); ws.cell(r,8,SEC_2025); ws.cell(r,8).font=Font(color=LINK_GREEN)
        for c in (5,6,7): ws.cell(r,c).number_format=FMT_PCT

    source=_find_row(ws,"Source & Data Quality")
    if source:
        ws.cell(source+1,1,"SEC 10-K Source"); ws.cell(source+1,2,SEC_2025); ws.cell(source+1,2).font=Font(color=LINK_GREEN)
        ws.cell(source+2,1,"Extraction Status"); ws.cell(source+2,2,"AUTO/FALLBACK — verified CVS FY2025 10-K: 4 operating segments, 5 consolidated revenue groups")
        ws.cell(source+3,1,"Important"); ws.cell(source+3,2,"Segment profitability uses Adjusted Operating Income because CVS identifies it as the principal segment performance measure. Segment revenues are before eliminations; consolidated revenue groups reconcile to company revenue."); ws.cell(source+3,2).alignment=Alignment(wrap_text=True)


def _repair_financial_statements(wb):
    if "Financial Statements" not in wb.sheetnames: return
    ws=wb["Financial Statements"]
    for label,vals in BS_GAPS.items():
        r=_find_row(ws,label)
        if r:
            for c,v in enumerate(vals,2):
                # Fill only the 2022-2025 columns in the standardized balance-sheet block.
                ws.cell(r,c,v); ws.cell(r,c).font=Font(color=INPUT_BLUE); ws.cell(r,c).number_format=FMT_BN
    # Repair formula rows that should always be derived.
    r_op=_find_row(ws,"Operating Margin"); r_rev=_find_row(ws,"Revenue"); r_oi=_find_row(ws,"Operating Income")
    r_other=_find_row(ws,"Other Income / (Expense), Net"); r_pretax=_find_row(ws,"Pre-Tax Income")
    if r_op and r_rev and r_oi:
        for c in range(2,8):
            col=ws.cell(1,c).column_letter; ws.cell(r_op,c,f'=IFERROR({col}{r_oi}/{col}{r_rev},"")'); ws.cell(r_op,c).number_format=FMT_PCT; ws.cell(r_op,c).font=Font(color=BLACK)
    if r_other and r_pretax and r_oi:
        for c in range(2,8):
            col=ws.cell(1,c).column_letter; ws.cell(r_other,c,f"={col}{r_pretax}-{col}{r_oi}"); ws.cell(r_other,c).number_format=FMT_BN; ws.cell(r_other,c).font=Font(color=BLACK)
    r_ocf=_find_row(ws,"Operating Cash Flow"); r_cap=_find_row(ws,"Capital Expenditures"); r_fcf=_find_row(ws,"Free Cash Flow")
    if r_ocf and r_cap and r_fcf:
        for c in range(2,8):
            col=ws.cell(1,c).column_letter; ws.cell(r_fcf,c,f"={col}{r_ocf}+{col}{r_cap}"); ws.cell(r_fcf,c).number_format=FMT_BN; ws.cell(r_fcf,c).font=Font(color=BLACK)
    # Explain rather than estimate non-disclosed functional expense lines.
    note_row=(r_op+1) if r_op else 20
    ws.cell(note_row,1,"Intentional blanks: CVS does not separately disclose standardized R&D, Sales & Marketing, and G&A amounts in a form directly comparable with technology issuers."); ws.cell(note_row,1).font=Font(italic=True,color=GREY); ws.cell(note_row,1).alignment=Alignment(wrap_text=True)


def repair_cvs_model(wb,ticker):
    if str(ticker).upper()!="CVS": return
    _repair_segment_analysis(wb)
    _repair_financial_statements(wb)
