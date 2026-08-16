from __future__ import annotations

"""Commodity-producer valuation normalization.

A secular-growth DCF is a poor default for upstream/integrated oil & gas producers because
commodity prices, acquisition step-ups and refining cycles can make one or two years look like
permanent growth.  This module keeps the company-specific WACC, but normalizes revenue growth,
operating margins, capital intensity and terminal growth toward mid-cycle economics.

The framework is intentionally conservative about evidence:
- historical financials determine cycle-normalized margins and D&A;
- issuer guidance is used only when explicitly configured and sourced;
- unsupported segment/commodity forecasts are never invented;
- current/peak commodity conditions can affect near-term years but are faded out of perpetuity.
"""

import math
import statistics
from openpyxl.styles import Alignment, Font, PatternFill

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; GREY="666666"; GOLD="FFF2CC"; GREEN="E2F0D9"
FMT_PCT='0.0%;[Red](0.0%);-'; FMT_BN='#,##0.0;[Red](#,##0.0);-'; FMT_PRICE='$#,##0.00;[Red]($#,##0.00);-'

# Guidance is a valuation anchor/cross-check, not an instruction to force the model to the
# company's target.  Update these entries when the issuer publishes a new long-range framework.
ISSUER_COMMODITY_GUIDANCE={
    "CVX":{
        "framework":"Chevron 2025 Investor Day / 2026 capital budget",
        "reference_commodity":"Nominal $70 Brent planning framework",
        "production_cagr":"2–3% annual upstream production through 2030",
        "fcf_growth":">10% annual adjusted FCF growth over five years at $70 Brent",
        "capex_long_range":"$18–21B annual organic capex",
        "capex_2026":"$18–19B organic capex",
        "base_capex_2026":18.5,
        "base_capex_2027_2030":19.5,
        "bear_capex_2026":19.0,
        "bear_capex_2027_2030":21.0,
        "bull_capex_2026":18.0,
        "bull_capex_2027_2030":18.0,
        "source_date":"2025-11-12 / 2025-12-03",
        "source_url":"https://www.chevron.com/newsroom/2025/q4/chevron-outlines-plan-for-sustained-cash-flow-growth-at-investor-day",
        "capex_url":"https://www.chevron.com/newsroom/2025/q4/chevron-announces-2026-capex-budget-of-18-to-19-billion",
    }
}


def _num(v,default=None):
    try:
        if isinstance(v,bool) or v in (None,""): return default
        x=float(v)
        return x if math.isfinite(x) else default
    except Exception:
        return default


def _clip(x,lo,hi):
    x=_num(x,lo)
    return max(lo,min(hi,x))


def _fade(start,end,n):
    if n<=1: return [end]
    return [start+(end-start)*i/(n-1) for i in range(n)]


def _median(values,default):
    clean=[float(x) for x in values if isinstance(x,(int,float)) and not isinstance(x,bool) and math.isfinite(float(x))]
    return statistics.median(clean) if clean else default


def _identity(wb,ticker):
    sector=industry=""
    if "Company Data" in wb.sheetnames:
        sector=str(wb["Company Data"]["B6"].value or "").strip()
        industry=str(wb["Company Data"]["B7"].value or "").strip()
    return str(ticker).upper().strip(),sector,industry


def is_commodity_producer(wb,ticker):
    t,sector,industry=_identity(wb,ticker)
    if t in ISSUER_COMMODITY_GUIDANCE: return True
    if sector.lower()!="energy": return False
    text=industry.lower()
    producer_terms=(
        "oil & gas integrated","oil and gas integrated","exploration & production",
        "exploration and production","oil & gas e&p","oil and gas e&p",
    )
    return any(term in text for term in producer_terms)


def _historical_cycle(wb):
    if "Historical Financials" not in wb.sheetnames:
        return {"latest_revenue":None,"margin":.12,"da":.08,"capex":.09,"years":[]}
    ws=wb["Historical Financials"]; rows=[]
    for c in range(2,8):
        year=_num(ws.cell(3,c).value); rev=_num(ws.cell(4,c).value); op=_num(ws.cell(9,c).value)
        capex=_num(ws.cell(15,c).value); da=_num(ws.cell(18,c).value)
        if year and rev and rev>0:
            rows.append((int(year),rev,op,capex,da))
    rows=rows[-4:]
    margins=[op/rev for _,rev,op,_,_ in rows if op is not None]
    da_rates=[abs(da)/rev for _,rev,_,_,da in rows if da is not None]
    cap_rates=[abs(capex)/rev for _,rev,_,capex,_ in rows if capex is not None]
    return {
        "latest_revenue":rows[-1][1] if rows else None,
        "margin":_clip(_median(margins,.12),.04,.30),
        "da":_clip(_median(da_rates,.08),.02,.18),
        "capex":_clip(_median(cap_rates,.09),.02,.25),
        "years":[r[0] for r in rows],
    }


def _project_revenues(start,growths):
    out=[]; rev=float(start)
    for g in growths:
        rev*=1+g; out.append(rev)
    return out


def _capex_ratios_from_nominal(revenues,first,through_2030,start_year=2026):
    nominal=[]
    for i,_ in enumerate(revenues):
        year=start_year+i
        if year==start_year: value=first
        elif year<=2030: value=through_2030
        else:
            years_after=year-2030; value=through_2030*((1.02)**years_after)
        nominal.append(value)
    return [_clip(v/rev,.02,.25) if rev else .09 for v,rev in zip(nominal,revenues)],nominal


def _write_block(ws,cols,growths,margins,da_rate,capex_rates):
    for c,g,m,cp in zip(cols,growths,margins,capex_rates):
        ws.cell(12,c).value=g; ws.cell(14,c).value=m; ws.cell(18,c).value=da_rate; ws.cell(20,c).value=cp
        ws.cell(12,c).number_format=FMT_PCT; ws.cell(14,c).number_format=FMT_PCT
        ws.cell(18,c).number_format=FMT_PCT; ws.cell(20,c).number_format=FMT_PCT


def _style_sheet(ws):
    ws.sheet_view.showGridLines=False
    for c in range(1,9): ws.cell(1,c).fill=PatternFill("solid",fgColor=NAVY); ws.cell(2,c).fill=PatternFill("solid",fgColor=NAVY)
    ws["A1"].font=Font(bold=True,color=WHITE,size=18)
    for r in (5,12,22,31):
        for c in range(1,9): ws.cell(r,c).fill=PatternFill("solid",fgColor=NAVY); ws.cell(r,c).font=Font(bold=True,color=WHITE)
    for r in (6,13,23,32):
        for c in range(1,9): ws.cell(r,c).fill=PatternFill("solid",fgColor=BLUE); ws.cell(r,c).font=Font(bold=True,color=WHITE); ws.cell(r,c).alignment=Alignment(wrap_text=True)
    widths={"A":30,"B":24,"C":20,"D":20,"E":20,"F":28,"G":38,"H":58}
    for col,w in widths.items(): ws.column_dimensions[col].width=w
    for row in ws.iter_rows():
        for cell in row: cell.alignment=Alignment(wrap_text=True,vertical="top")


def _render_proof(wb,ticker,meta):
    if "Commodity Valuation" in wb.sheetnames: wb.remove(wb["Commodity Valuation"])
    ws=wb.create_sheet("Commodity Valuation")
    ws["A1"]=f"{ticker} — Commodity-Normalized Valuation Framework"
    ws["A3"]=("Commodity producers are valued on normalized mid-cycle economics rather than extrapolating a peak commodity price, "
               "acquisition step-up or refining cycle indefinitely. Company-specific WACC is retained; commodity risk is reflected primarily "
               "through normalized cash flows, capital intensity and a lower perpetual-growth assumption.")
    ws.merge_cells("A3:H3")
    ws["A5"]="Framework & Evidence"
    headers=["Item","Base / Value","Bear","Bull","Unit","Method","Primary Source","Audit Note"]
    for c,v in enumerate(headers,1): ws.cell(6,c,v)
    guidance=meta.get("guidance") or {}
    items=[
        ("Valuation framework","Mid-cycle DCF","Down-cycle","Strong cycle","","Normalize economics, then discount with company WACC","Project methodology","Not a price-target override"),
        ("Commodity reference",guidance.get("reference_commodity") or "Historical-cycle normalization","","","", "Issuer guidance when available; otherwise history","Issuer IR / historical financials","Refresh when long-range guidance changes"),
        ("Production / volume anchor",guidance.get("production_cagr") or "Not explicitly modeled","","","","Cross-check for long-run revenue growth","Issuer IR","Revenue growth should not structurally outrun volume + price/mix without evidence"),
        ("FCF guidance cross-check",guidance.get("fcf_growth") or "Not available","","","","Cross-check only; model does not force management target","Issuer IR","FCF growth can exceed revenue growth through mix, synergies and capital efficiency"),
        ("Long-range capex",guidance.get("capex_long_range") or f"Historical median {meta['hist_capex']:.1%} of revenue","","","","Nominal issuer guidance where available; otherwise historical normalized ratio","Issuer IR / historical financials","Capex is not allowed to collapse simply to boost DCF"),
        ("Calculated WACC",meta.get("wacc"),"+ stress shocks","- opportunity shocks","%","Existing dynamic CAPM/WACC engine retained","Cost of Capital","No arbitrary commodity WACC floor"),
        ("Terminal growth",meta.get("terminal_growth"),"Stress tests reduce further","Opportunity tests may raise modestly","%","Mature/depleting commodity terminal policy","Project methodology","Base capped at 2.0% for configured commodity framework"),
    ]
    for r,row in enumerate(items,7):
        for c,v in enumerate(row,1): ws.cell(r,c,v)
        if isinstance(row[1],float): ws.cell(r,2).number_format=FMT_PCT

    ws["A12"]="Historical Cycle Anchors"
    for c,v in enumerate(["Metric","Normalized Value","Years","Definition","Why It Matters","Source","Status","Caveat"],1): ws.cell(13,c,v)
    hist_rows=[
        ("Operating margin",meta["hist_margin"],", ".join(map(str,meta.get("hist_years",[]))),"Median operating income / revenue","Prevents peak margin from becoming perpetual","Historical Financials","Complete","Commodity cycles can remain above/below median for years"),
        ("D&A / Revenue",meta["hist_da"],", ".join(map(str,meta.get("hist_years",[]))),"Median D&A / revenue","Normalizes depletion/depreciation burden","Historical Financials","Complete","Accounting mix changes after acquisitions"),
        ("Capex / Revenue",meta["hist_capex"],", ".join(map(str,meta.get("hist_years",[]))),"Median capex / revenue","Capital intensity cross-check","Historical Financials","Complete","Issuer nominal guidance supersedes when available"),
    ]
    for r,row in enumerate(hist_rows,14):
        for c,v in enumerate(row,1): ws.cell(r,c,v)
        ws.cell(r,2).number_format=FMT_PCT

    ws["A22"]="Normalized Scenario Path"
    for c,v in enumerate(["Year","Bear Growth","Base Growth","Bull Growth","Bear EBIT Margin","Base EBIT Margin","Bull EBIT Margin","Base Capex ($bn / % revenue)"],1): ws.cell(23,c,v)
    for i,year in enumerate(meta["forecast_years"]):
        r=24+i
        vals=[year,meta["bear_growth"][i],meta["base_growth"][i],meta["bull_growth"][i],meta["bear_margin"][i],meta["base_margin"][i],meta["bull_margin"][i],f"{meta['base_capex_nominal'][i]:.1f} / {meta['base_capex_rates'][i]:.1%}"]
        for c,v in enumerate(vals,1): ws.cell(r,c,v)
        for c in range(2,8): ws.cell(r,c).number_format=FMT_PCT

    ws["A31"]="Source & Governance"
    for c,v in enumerate(["Source","Date","URL","Used For","Priority","Refresh Rule","Status","Notes"],1): ws.cell(32,c,v)
    sources=[]
    if guidance:
        sources.append((guidance.get("framework"),guidance.get("source_date"),guidance.get("source_url"),"$70 Brent / FCF / production / long-range capex anchors","Primary issuer","Refresh on new Investor Day / long-range plan","Current configured source","Management guidance is a cross-check, not a valuation target"))
        sources.append(("Chevron 2026 Capital Budget",guidance.get("source_date"),guidance.get("capex_url"),"2026 organic capex","Primary issuer","Refresh annually","Current configured source","Uses midpoint for base scenario"))
    sources.append(("Historical Financials","Latest completed fiscal years","Workbook","Mid-cycle operating margin, D&A and capital intensity","Audited/provider-reconciled history","Refresh each build","Automatic","Historical medians are not forecasts"))
    for r,row in enumerate(sources,33):
        for c,v in enumerate(row,1): ws.cell(r,c,v)
        if isinstance(row[2],str) and row[2].startswith("http"):
            ws.cell(r,3).hyperlink=row[2]; ws.cell(r,3).font=Font(color="008000",underline="single")
    _style_sheet(ws)
    ws.freeze_panes="A23"
    return ws


def apply_commodity_normalization(wb,ticker,info=None):
    """Normalize forecast assumptions for commodity producers; return audit metadata."""
    ticker=str(ticker).upper().strip()
    if not is_commodity_producer(wb,ticker):
        return {"applied":False,"ticker":ticker}
    if "Three-Case Scenarios" not in wb.sheetnames or "Historical Financials" not in wb.sheetnames:
        return {"applied":False,"ticker":ticker,"reason":"Required scenario/history sheets missing"}

    s=wb["Three-Case Scenarios"]; hist=_historical_cycle(wb); start=hist.get("latest_revenue")
    if not start or start<=0:
        return {"applied":False,"ticker":ticker,"reason":"Latest annual revenue unavailable"}

    cycle=hist["margin"]; da=hist["da"]; guidance=ISSUER_COMMODITY_GUIDANCE.get(ticker,{})
    base_cols=list(range(14,24)); bear_cols=list(range(2,12)); bull_cols=list(range(26,36))
    existing=[_num(s.cell(12,c).value,0) for c in base_cols]

    # Preserve a bounded near-term acquisition/cycle step-up, but force long-run growth back to a
    # production-plus-price/mix range instead of carrying a cyclical CAGR into perpetuity.
    base_growth=[_clip(existing[0],-.05,.25),_clip(existing[1],-.10,.10)]+_fade(.035,.020,8)
    bear_growth=[_clip(base_growth[0]-.05,-.15,.20),_clip(base_growth[1]-.025,-.15,.075)]+_fade(.010,0.0,8)
    bull_growth=[_clip(base_growth[0]+.05,-.02,.30),_clip(base_growth[1]+.04,-.05,.15)]+_fade(.050,.025,8)

    base_margin=[min(_num(s.cell(14,base_cols[0]).value,cycle),cycle+.05),min(_num(s.cell(14,base_cols[1]).value,cycle),cycle+.025)]+_fade(cycle+.015,cycle+.005,8)
    bear_margin=[max(.03,cycle-.01),max(.03,cycle-.015)]+_fade(max(.03,cycle-.015),max(.03,cycle-.025),8)
    bull_margin=[min(.30,cycle+.08),min(.28,cycle+.06)]+_fade(min(.25,cycle+.04),min(.23,cycle+.025),8)

    base_revs=_project_revenues(start,base_growth); bear_revs=_project_revenues(start,bear_growth); bull_revs=_project_revenues(start,bull_growth)
    years=list(range(2026,2036))
    if guidance:
        base_cap,base_nom=_capex_ratios_from_nominal(base_revs,guidance["base_capex_2026"],guidance["base_capex_2027_2030"])
        bear_cap,bear_nom=_capex_ratios_from_nominal(bear_revs,guidance["bear_capex_2026"],guidance["bear_capex_2027_2030"])
        bull_cap,bull_nom=_capex_ratios_from_nominal(bull_revs,guidance["bull_capex_2026"],guidance["bull_capex_2027_2030"])
    else:
        normalized=hist["capex"]
        base_cap=[normalized]*10; bear_cap=[min(.25,normalized+.02)]*10; bull_cap=[max(.02,normalized-.015)]*10
        base_nom=[r*c for r,c in zip(base_revs,base_cap)]; bear_nom=[r*c for r,c in zip(bear_revs,bear_cap)]; bull_nom=[r*c for r,c in zip(bull_revs,bull_cap)]

    _write_block(s,bear_cols,bear_growth,bear_margin,da,bear_cap)
    _write_block(s,base_cols,base_growth,base_margin,da,base_cap)
    _write_block(s,bull_cols,bull_growth,bull_margin,da,bull_cap)

    # A mature/depleting commodity producer should not receive a 3–3.5% perpetual nominal growth
    # assumption simply because the generic template uses one. WACC remains company-specific.
    s["C7"].value=.02; s["C7"].number_format=FMT_PCT

    wacc=_num(s["C6"].value)
    meta={
        "applied":True,"ticker":ticker,"framework":"Commodity-normalized mid-cycle DCF",
        "guidance":guidance,"hist_margin":cycle,"hist_da":da,"hist_capex":hist["capex"],"hist_years":hist["years"],
        "forecast_years":years,"bear_growth":bear_growth,"base_growth":base_growth,"bull_growth":bull_growth,
        "bear_margin":bear_margin,"base_margin":base_margin,"bull_margin":bull_margin,
        "bear_capex_rates":bear_cap,"base_capex_rates":base_cap,"bull_capex_rates":bull_cap,
        "base_capex_nominal":base_nom,"bear_capex_nominal":bear_nom,"bull_capex_nominal":bull_nom,
        "terminal_growth":.02,"wacc":wacc,
    }
    setattr(wb,"_commodity_valuation",meta)
    _render_proof(wb,ticker,meta)
    print(f"Commodity valuation normalization: {ticker}; cycle EBIT margin={cycle:.2%}; D&A/revenue={da:.2%}; terminal growth=2.00%; WACC retained={wacc:.2%}" if wacc is not None else f"Commodity valuation normalization: {ticker}; terminal growth=2.00%")
    return meta
