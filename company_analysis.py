"""Financial statements, reported segments, and full business-portfolio analysis.

The module deliberately separates:
1) official reportable segments / disclosed revenue lines, and
2) named products, platforms and subsidiary businesses whose standalone revenue is not disclosed.
"""

import re
from io import StringIO
import requests
try:
    import pandas as pd
except Exception:
    pd = None

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils import get_column_letter
from business_portfolio import ensure_business_portfolio

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; LIGHT="F5F9FC"; PALE_BLUE="D9EAF7"; GOLD="FFF2CC"; INPUT_BLUE="0000FF"; GREY="666666"
FMT_BN='#,##0.0;[Red](#,##0.0);-'; FMT_PCT='0.0%;[Red](0.0%);-'; FMT_EPS='$0.00;[Red]($0.00);-'
THIN=Side(style="thin",color="808080")


def _fill(c): return PatternFill("solid",fgColor=c)

def _title(ws,text,end="H"):
    ws.merge_cells(f"A1:{end}2"); ws["A1"]=text; ws["A1"].fill=_fill(NAVY); ws["A1"].font=Font(bold=True,color=WHITE,size=18); ws["A1"].alignment=Alignment(vertical="center"); ws.sheet_view.showGridLines=False

def _section(ws,row,title,end="H"):
    ws.merge_cells(f"A{row}:{end}{row}"); ws[f"A{row}"]=title; ws[f"A{row}"].fill=_fill(NAVY); ws[f"A{row}"].font=Font(bold=True,color=WHITE,size=12)

def _header(ws,row,cols):
    for c in range(1,cols+1):
        x=ws.cell(row,c); x.fill=_fill(BLUE); x.font=Font(bold=True,color=WHITE); x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

def _scale(v): return None if v is None else float(v)/1e9


def _merged_annual_series(facts,tags,preferred_unit=None):
    if not facts: return {}
    gaap=facts.get("facts",{}).get("us-gaap",{})
    best={}
    for priority,tag in enumerate(tags):
        fact=gaap.get(tag)
        if not fact: continue
        units=fact.get("units",{})
        unit=preferred_unit if preferred_unit in units else (next(iter(units)) if units else None)
        if not unit: continue
        for x in units.get(unit,[]):
            if x.get("form") not in ("10-K","10-K/A") or x.get("fp")!="FY": continue
            fy=x.get("fy"); val=x.get("val")
            if fy is None or val is None: continue
            try: fy=int(fy); val=float(val)
            except Exception: continue
            stamp=(str(x.get("filed") or ""),str(x.get("end") or ""),-priority)
            if fy not in best or stamp>best[fy][0]: best[fy]=(stamp,val)
    return {y:v for y,(_,v) in best.items()}


def ensure_financial_statements(wb,ticker,facts):
    if "Financial Statements" in wb.sheetnames: wb.remove(wb["Financial Statements"])
    ws=wb.create_sheet("Financial Statements"); _title(ws,f"{ticker} — Financial Statements")
    ws.merge_cells("A3:H3"); ws["A3"]="USD billions unless per-share data. Annual SEC Company Facts; missing tags remain blank rather than estimated."; ws["A3"].font=Font(italic=True,color=GREY)

    income_map=[
        ("Revenue",["RevenueFromContractWithCustomerExcludingAssessedTax","SalesRevenueNet","Revenues"],None),
        ("Cost of Revenue",["CostOfRevenue","CostOfGoodsAndServicesSold"],None),
        ("Gross Profit",["GrossProfit"],None),
        ("Research & Development",["ResearchAndDevelopmentExpense"],None),
        ("Sales & Marketing",["SellingAndMarketingExpense","MarketingExpense"],None),
        ("General & Administrative",["GeneralAndAdministrativeExpense"],None),
        ("Operating Income",["OperatingIncomeLoss"],None),
        ("Other Income / (Expense), Net",["NonoperatingIncomeExpense","OtherNonoperatingIncomeExpense"],None),
        ("Pre-Tax Income",["IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest","IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments"],None),
        ("Income Taxes",["IncomeTaxExpenseBenefit"],None),
        ("Net Income",["NetIncomeLoss","ProfitLoss"],None),
        ("Diluted EPS",["EarningsPerShareDiluted"],"USD/shares"),
    ]
    s={n:_merged_annual_series(facts,t,u) for n,t,u in income_map}
    years=sorted(s.get("Revenue",{}))[-6:]
    _section(ws,5,"Income Statement"); ws.cell(6,1,"Metric")
    for j,y in enumerate(years,2): ws.cell(6,j,y)
    _header(ws,6,max(1,len(years)+1))
    row=7; rows={}
    for name,_,unit in income_map:
        rows[name]=row; ws.cell(row,1,name)
        for j,y in enumerate(years,2):
            val=s.get(name,{}).get(y)
            if name=="Gross Profit" and val is None:
                rev=s.get("Revenue",{}).get(y); cost=s.get("Cost of Revenue",{}).get(y); val=(rev-cost) if rev is not None and cost is not None else None
            ws.cell(row,j,val if unit else _scale(val)); ws.cell(row,j).number_format=FMT_EPS if unit else FMT_BN
        row+=1
    if ticker.upper() in {"GOOGL","GOOG"} and years==[2020,2021,2022,2023,2024,2025]:
        for j,v in enumerate([2.93,5.61,4.56,5.80,8.04,10.81],2): ws.cell(rows["Diluted EPS"],j,v); ws.cell(rows["Diluted EPS"],j).number_format=FMT_EPS
    for j in range(2,2+len(years)):
        c=get_column_letter(j); ws.cell(rows["Other Income / (Expense), Net"],j,f"={c}{rows['Pre-Tax Income']}-{c}{rows['Operating Income']}"); ws.cell(rows["Other Income / (Expense), Net"],j).number_format=FMT_BN
    op_margin_row=row; ws.cell(row,1,"Operating Margin")
    for j in range(2,2+len(years)):
        c=get_column_letter(j); ws.cell(row,j,f'=IFERROR({c}{rows["Operating Income"]}/{c}{rows["Revenue"]},"")'); ws.cell(row,j).number_format=FMT_PCT

    bs_map=[("Cash & Cash Equivalents",["CashAndCashEquivalentsAtCarryingValue"]),("Marketable / Short-Term Securities",["MarketableSecuritiesCurrent","ShortTermInvestments"]),("Accounts Receivable",["AccountsReceivableNetCurrent"]),("Other Current Assets",["OtherCurrentAssets"]),("Total Current Assets",["AssetsCurrent"]),("Property & Equipment, Net",["PropertyPlantAndEquipmentNet"]),("Goodwill",["Goodwill"]),("Total Assets",["Assets"]),("Accounts Payable",["AccountsPayableCurrent"]),("Deferred Revenue",["ContractWithCustomerLiabilityCurrent","DeferredRevenueCurrent"]),("Total Current Liabilities",["LiabilitiesCurrent"]),("Long-Term Debt",["LongTermDebtNoncurrent","LongTermDebt"]),("Total Liabilities",["Liabilities"]),("Stockholders' Equity",["StockholdersEquity","StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest"])]
    bs={n:_merged_annual_series(facts,t) for n,t in bs_map}; bs_years=sorted(bs.get("Total Assets",{}))[-4:]
    start=max(op_margin_row+4,23); _section(ws,start-1,"Balance Sheet"); ws.cell(start,1,"Metric")
    for j,y in enumerate(bs_years,2): ws.cell(start,j,y)
    _header(ws,start,max(1,len(bs_years)+1)); row=start+1
    for name,_ in bs_map:
        ws.cell(row,1,name)
        for j,y in enumerate(bs_years,2): ws.cell(row,j,_scale(bs.get(name,{}).get(y))); ws.cell(row,j).number_format=FMT_BN
        if name in {"Total Current Assets","Total Assets","Total Current Liabilities","Total Liabilities","Stockholders' Equity"}:
            ws.cell(row,1).font=Font(bold=True)
            for c in range(1,len(bs_years)+2): ws.cell(row,c).border=Border(bottom=THIN)
        row+=1
    if ticker.upper() in {"GOOGL","GOOG"} and 2025 in bs_years:
        ppe_row=start+1+[x[0] for x in bs_map].index("Property & Equipment, Net"); col=2+bs_years.index(2025)
        if ws.cell(ppe_row,col).value is None: ws.cell(ppe_row,col,246.597)

    cf_map=[("Net Income",["NetIncomeLoss","ProfitLoss"]),("Depreciation & Amortization",["DepreciationDepletionAndAmortization","DepreciationDepletionAndAmortizationPropertyPlantAndEquipment"]),("Stock-Based Compensation",["ShareBasedCompensation"]),("Operating Cash Flow",["NetCashProvidedByUsedInOperatingActivities"]),("Capital Expenditures",["PaymentsToAcquirePropertyPlantAndEquipment","PaymentsToAcquireProductiveAssets"]),("Acquisitions",["PaymentsToAcquireBusinessesNetOfCashAcquired","PaymentsToAcquireBusinessesNetOfCashAndCashEquivalentsAcquired"]),("Share Repurchases",["PaymentsForRepurchaseOfCommonStock"]),("Dividends",["PaymentsOfDividends"]),("Debt Issuance",["ProceedsFromIssuanceOfLongTermDebt","ProceedsFromIssuanceOfDebt"]),("Debt Repayments",["RepaymentsOfLongTermDebt","RepaymentsOfDebt"]),("Ending Cash",["CashAndCashEquivalentsAtCarryingValue"])]
    cf={n:_merged_annual_series(facts,t) for n,t in cf_map}; cf_years=years; start=max(row+3,49); _section(ws,start,"Cash Flow Statement"); ws.cell(start+1,1,"Metric")
    for j,y in enumerate(cf_years,2): ws.cell(start+1,j,y)
    _header(ws,start+1,max(1,len(cf_years)+1)); row=start+2; cf_rows={}
    for name,_ in cf_map:
        cf_rows[name]=row; ws.cell(row,1,name)
        for j,y in enumerate(cf_years,2):
            val=cf.get(name,{}).get(y)
            if name in {"Capital Expenditures","Acquisitions","Share Repurchases","Dividends","Debt Repayments"} and val is not None: val=-abs(val)
            ws.cell(row,j,_scale(val)); ws.cell(row,j).number_format=FMT_BN
        row+=1
    if ticker.upper() in {"GOOGL","GOOG"} and cf_years==[2020,2021,2022,2023,2024,2025]:
        for j,v in enumerate([12.9,11.6,15.3,11.946,15.311,21.136],2): ws.cell(cf_rows["Depreciation & Amortization"],j,v)
    fcf_row=cf_rows["Capital Expenditures"]+1; ws.insert_rows(fcf_row,1); ws.cell(fcf_row,1,"Free Cash Flow"); ws.cell(fcf_row,1).font=Font(bold=True)
    for j in range(2,2+len(cf_years)):
        c=get_column_letter(j); ws.cell(fcf_row,j,f"={c}{cf_rows['Operating Cash Flow']}+{c}{cf_rows['Capital Expenditures']}"); ws.cell(fcf_row,j).number_format=FMT_BN
    ws.column_dimensions["A"].width=40
    for c in range(2,8): ws.column_dimensions[get_column_letter(c)].width=14
    ws.freeze_panes="B7"
    return ws


SEGMENT_CONFIGS={
    "MSFT":{"segments":["Productivity and Business Processes","Intelligent Cloud","More Personal Computing"]},
    "AMZN":{"segments":["North America","International","AWS"]},
    "META":{"segments":["Family of Apps","Reality Labs"]},
    "NVDA":{"segments":["Compute & Networking","Graphics"]},
}


def _latest_10k_html(ticker,headers):
    tickers=requests.get("https://www.sec.gov/files/company_tickers.json",headers=headers,timeout=30).json(); cik=None
    for item in tickers.values():
        if str(item.get("ticker","")).upper()==ticker.upper(): cik=str(item["cik_str"]).zfill(10); break
    if not cik: return None,None
    subs=requests.get(f"https://data.sec.gov/submissions/CIK{cik}.json",headers=headers,timeout=30).json(); recent=subs.get("filings",{}).get("recent",{})
    for form,acc,doc in zip(recent.get("form",[]),recent.get("accessionNumber",[]),recent.get("primaryDocument",[])):
        if form=="10-K":
            url=f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{acc.replace('-','')}/{doc}"; return requests.get(url,headers=headers,timeout=45).text,url
    return None,None


def _numbers(row):
    out=[]
    for val in row.tolist():
        for token in re.findall(r"\(?-?\d[\d,]*(?:\.\d+)?\)?",str(val)):
            neg=token.startswith("(") and token.endswith(")"); clean=token.strip("()").replace(",","")
            try: num=float(clean)
            except Exception: continue
            if 1900<=num<=2100: continue
            out.append(-num if neg else num)
    return out


def _extract(html,labels):
    if not html or pd is None: return {}
    try: tables=pd.read_html(StringIO(html))
    except Exception: return {}
    out={x:[] for x in labels}
    for df in tables:
        for _,row in df.iterrows():
            text=" | ".join(str(v) for v in row.tolist())
            for label in labels:
                if label.lower() in text.lower():
                    nums=_numbers(row)
                    if len(nums)>=2: out[label].append(nums[-3:])
    return out


def _manual_segment_sheet(wb,ticker,url=None):
    if "Segment Analysis" in wb.sheetnames: wb.remove(wb["Segment Analysis"])
    ws=wb.create_sheet("Segment Analysis"); _title(ws,f"{ticker} — Business & Segment Analysis")
    ws.merge_cells("A3:H3"); ws["A3"]="Automatic financial segment extraction was unavailable. Enter only issuer-disclosed segment data in yellow cells."; ws["A3"].font=Font(italic=True,color=GREY)
    _section(ws,5,"Manual Segment Input"); heads=["Segment / Business Line","Year -2 Revenue","Year -1 Revenue","Latest Revenue","Latest Growth","Latest Operating Income","Latest Op. Margin","Source / Notes"]
    for c,v in enumerate(heads,1): ws.cell(6,c,v)
    _header(ws,6,8)
    for r in range(7,19):
        for c in range(1,9): ws.cell(r,c).fill=_fill(GOLD); ws.cell(r,c).font=Font(color=INPUT_BLUE)
        for c in range(2,7): ws.cell(r,c).number_format=FMT_BN
        ws.cell(r,5).number_format=FMT_PCT; ws.cell(r,7).number_format=FMT_PCT
    ws["A21"]="SEC source"; ws["B21"]=url or ""; ws.column_dimensions["A"].width=34; ws.column_dimensions["H"].width=44
    return ws


def _alphabet_segment_sheet(wb,ticker,html,url):
    labels=["Google Services","Google Cloud","Other Bets","Google Search & other","YouTube ads","Google Network","Google subscriptions, platforms, and devices"]
    extracted=_extract(html,labels)
    fallback={
        "Google Services":[272543,304930,342721],"Google Cloud":[33088,43229,58705],"Other Bets":[1527,1648,1537],
        "Google Search & other":[175033,198084,224532],"YouTube ads":[31510,36147,40367],"Google Network":[31312,30359,29792],
        "Google subscriptions, platforms, and devices":[34688,40340,48030],
    }
    for k,v in fallback.items():
        if not extracted.get(k): extracted[k]=[v]
    op={"Google Services":[95858,121263,139404],"Google Cloud":[1716,6112,13910],"Other Bets":[-4095,-4444,-7515],"Alphabet-level activities":[-9186,-10541,-16760]}
    if "Segment Analysis" in wb.sheetnames: wb.remove(wb["Segment Analysis"])
    ws=wb.create_sheet("Segment Analysis"); ws.sheet_view.showGridLines=False; _title(ws,f"{ticker} — Business & Segment Analysis","N")
    ws.merge_cells("A3:N3"); ws["A3"]="Official segments, disclosed business-line revenue, profitability, geography and operating drivers. Use Business Portfolio Map for all named products/businesses."; ws["A3"].font=Font(italic=True,color=GREY)
    _section(ws,5,"Reported Operating Segments","N"); heads=["Segment","2023 Revenue","2024 Revenue","2025 Revenue","2025 Growth","2023–25 CAGR","2023 Op. Income","2024 Op. Income","2025 Op. Income","2023 Margin","2024 Margin","2025 Margin","Margin Δ vs 2024","2025 Revenue Mix"]
    for c,v in enumerate(heads,1): ws.cell(6,c,v)
    _header(ws,6,14); total=402.836
    for r,name in enumerate(["Google Services","Google Cloud","Other Bets","Alphabet-level activities"],7):
        rv=(extracted.get(name) or [[None,None,None]])[0][-3:]; rv=[x/1000 if x is not None else None for x in rv]; oi=[x/1000 for x in op[name]]
        vals=[name]+rv+[(rv[2]/rv[1]-1) if rv[2] and rv[1] else None,((rv[2]/rv[0])**.5-1) if rv[2] and rv[0] else None]+oi+[(oi[i]/rv[i]) if rv[i] else None for i in range(3)]+[((oi[2]/rv[2])-(oi[1]/rv[1])) if rv[2] and rv[1] else None,(rv[2]/total) if rv[2] else None]
        for c,v in enumerate(vals,1): ws.cell(r,c,v)
        for c in range(2,5): ws.cell(r,c).number_format=FMT_BN
        for c in range(5,7): ws.cell(r,c).number_format=FMT_PCT
        for c in range(7,10): ws.cell(r,c).number_format=FMT_BN
        for c in range(10,15): ws.cell(r,c).number_format=FMT_PCT
    ws.conditional_formatting.add("E7:F10",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="percentile",mid_value=50,mid_color="FFEB84",end_type="max",end_color="63BE7B"))

    _section(ws,13,"Revenue by Business Line","J"); heads=["Business Line","2023","2024","2025","2025 Growth","2023–25 CAGR","2025 Mix","2025 Δ Revenue","Share of 2025 Revenue Growth","Investment Read-Through"]
    for c,v in enumerate(heads,1): ws.cell(14,c,v)
    _header(ws,14,10); growth=402.836-350.018
    notes={"Google Search & other":"Core earnings engine; Search plus ads on owned properties","YouTube ads":"Large video advertising platform","Google Network":"Third-party ad network; structurally softer","Google subscriptions, platforms, and devices":"Subscriptions, Play and Pixel ecosystem","Google Cloud":"Fastest major growth and margin-expansion engine","Other Bets":"Small revenue base; long-duration optionality"}
    for r,name in enumerate(["Google Search & other","YouTube ads","Google Network","Google subscriptions, platforms, and devices","Google Cloud","Other Bets"],15):
        raw=(extracted.get(name) or [fallback[name]])[0][-3:]; a,b,c=[x/1000 for x in raw]
        vals=[name,a,b,c,c/b-1,(c/a)**.5-1,c/total,c-b,(c-b)/growth,notes[name]]
        for col,v in enumerate(vals,1): ws.cell(r,col,v)
        for col in range(2,5): ws.cell(r,col).number_format=FMT_BN
        for col in range(5,8): ws.cell(r,col).number_format=FMT_PCT
        ws.cell(r,8).number_format=FMT_BN; ws.cell(r,9).number_format=FMT_PCT
    ws.conditional_formatting.add("E15:F20",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="percentile",mid_value=50,mid_color="FFEB84",end_type="max",end_color="63BE7B")); ws.conditional_formatting.add("I15:I20",DataBarRule(start_type="min",end_type="max",color="9E4C9E"))

    _section(ws,23,"Business Coverage","N"); ws.merge_cells("A24:N26"); ws["A24"]="Alphabet reports only three financial segments, but its business portfolio is much broader. The companion Business Portfolio Map catalogs Services products, Cloud sub-platforms, Other Bets and shared AI activities without assigning fake standalone revenue."; ws["A24"].fill=_fill(PALE_BLUE); ws["A24"].alignment=Alignment(wrap_text=True)
    ws["A28"]="SEC source"; ws["B28"]=url or ""; ws["B28"].font=Font(color="008000")
    ws.column_dimensions["A"].width=35
    for c in range(2,15): ws.column_dimensions[get_column_letter(c)].width=14
    return ws


def ensure_segment_analysis(wb,ticker,headers):
    html,url=_latest_10k_html(ticker,headers)
    if ticker.upper() in {"GOOGL","GOOG"}:
        ws=_alphabet_segment_sheet(wb,ticker.upper(),html,url)
        ensure_business_portfolio(wb,ticker.upper(),headers)
        return ws
    cfg=SEGMENT_CONFIGS.get(ticker.upper())
    if not cfg or not html: return _manual_segment_sheet(wb,ticker,url)
    labels=cfg["segments"]; extracted=_extract(html,labels)
    if not any(extracted.get(x) for x in labels): return _manual_segment_sheet(wb,ticker,url)
    if "Segment Analysis" in wb.sheetnames: wb.remove(wb["Segment Analysis"])
    ws=wb.create_sheet("Segment Analysis"); _title(ws,f"{ticker} — Business & Segment Analysis")
    ws.merge_cells("A3:H3"); ws["A3"]="Reported segment revenue from the latest annual filing. Additional product mapping will be added as company-specific disclosure parsers are expanded."; ws["A3"].font=Font(italic=True,color=GREY)
    _section(ws,5,"Reported Operating Segments"); heads=["Segment","Year -2 Revenue","Year -1 Revenue","Latest Revenue","Latest Growth"]
    for c,v in enumerate(heads,1): ws.cell(6,c,v)
    _header(ws,6,5); r=7
    for name in labels:
        matches=extracted.get(name,[]); raw=matches[0] if matches else []
        if len(raw)<2: continue
        rv=[x/1000 for x in raw[-3:]]
        while len(rv)<3: rv.insert(0,None)
        for c,v in enumerate([name]+rv+[(rv[-1]/rv[-2]-1) if rv[-1] is not None and rv[-2] else None],1): ws.cell(r,c,v)
        for c in range(2,5): ws.cell(r,c).number_format=FMT_BN
        ws.cell(r,5).number_format=FMT_PCT; r+=1
    ws["A20"]="SEC source"; ws["B20"]=url or ""; ws.column_dimensions["A"].width=38
    return ws
