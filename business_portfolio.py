"""Hierarchical business-portfolio mapping for equity research workbooks.

Separates accounting segments from disclosed revenue groups and named businesses/products.
The goal is broad business coverage without inventing standalone financials that the issuer
has not disclosed.
"""

import html as html_lib
import re
import requests
from openpyxl.styles import PatternFill, Font, Alignment

NAVY = "17365D"
BLUE = "2F75B5"
WHITE = "FFFFFF"
LIGHT = "F5F9FC"
PALE_BLUE = "D9EAF7"
INPUT_BLUE = "0000FF"
GREY = "666666"
GREEN_FILL = "E2F0D9"
GOLD_FILL = "FFF2CC"

FMT_BN = '#,##0.0;[Red](#,##0.0);-'
FMT_PCT = '0.0%;[Red](0.0%);-'

ALPHABET_FAQ = "https://abc.xyz/investor/faqs-and-general-information/default.aspx"
ALPHABET_2025_10K = "https://www.sec.gov/Archives/edgar/data/1652044/000165204426000018/goog-20251231.htm"


def _fill(color):
    return PatternFill("solid", fgColor=color)


def _strip_html(text):
    text = re.sub(r"<script.*?</script>|<style.*?</style>", " ", text or "", flags=re.I | re.S)
    text = re.sub(r"<[^>]+>", " ", text)
    return re.sub(r"\s+", " ", html_lib.unescape(text)).strip()


def _discover_alphabet_names(headers):
    """Best-effort discovery of current product/business names from Alphabet investor FAQ.

    Static fallbacks remain authoritative if the page layout changes. Dynamic discovery is
    additive, allowing newly named products or Other Bets to appear without code changes.
    """
    discovered = {"services": [], "cloud": [], "other_bets": []}
    try:
        r = requests.get(ALPHABET_FAQ, headers=headers, timeout=30)
        r.raise_for_status()
        text = _strip_html(r.text)
    except Exception:
        return discovered

    patterns = {
        "services": r"Google Services Segment.*?Includes products and services such as:?\s*(.*?)\s*Google Services generates revenues",
        "other_bets": r"Other Bets Segment.*?Includes businesses such as\s*(.*?)\s*,?\s*among others",
    }
    for key, pat in patterns.items():
        m = re.search(pat, text, flags=re.I | re.S)
        if m:
            chunk = m.group(1).replace(" and ", ", ")
            names = [re.sub(r"\s+", " ", x).strip(" ,.;:") for x in chunk.split(",")]
            discovered[key] = [x for x in names if 1 < len(x) < 80]

    m = re.search(r"Google Cloud Segment.*?These services include:?\s*(.*?)\s*Google Cloud generates revenues", text, flags=re.I | re.S)
    if m:
        chunk = m.group(1)
        seeds = [
            "AI Infrastructure", "Cloud TPUs", "GPUs", "Vertex AI", "Cybersecurity",
            "Data and analytics", "Gemini Enterprise", "Gemini for Google Workspace",
            "Customer Engagement Suite", "Google Cloud Platform", "Google Workspace",
        ]
        discovered["cloud"] = [s for s in seeds if s.lower() in chunk.lower() or s in {"Google Cloud Platform", "Google Workspace"}]
    return discovered


def _alphabet_static_rows():
    faq = ALPHABET_FAQ
    sec = ALPHABET_2025_10K
    return [
        ["Google Services", "Google Search & other", "Google Search", "Search / advertising", "Included in disclosed revenue group", "Advertising + search distribution economics", "Core cash/earnings engine", faq],
        ["Google Services", "Google Search & other", "Gmail", "Consumer product / advertising", "Included in disclosed revenue group", "Advertising on owned-and-operated properties", "Engagement ecosystem / ad inventory", faq],
        ["Google Services", "Google Search & other", "Google Maps", "Consumer product / advertising", "Included in disclosed revenue group", "Advertising on owned-and-operated properties", "Local intent / commerce / ad inventory", faq],
        ["Google Services", "Google Search & other", "Google Play advertising", "Platform / advertising", "Included in disclosed revenue group", "Advertising on owned-and-operated properties", "Ecosystem monetization", faq],
        ["Google Services", "YouTube ads", "YouTube advertising", "Video advertising", "Standalone disclosed revenue line", "Advertising", "Large video ad platform", sec],
        ["Google Services", "Google Network", "AdSense", "Advertising network", "Included in disclosed revenue group", "Revenue share with publisher/network partners", "Third-party ad network", faq],
        ["Google Services", "Google Network", "AdMob", "Mobile advertising network", "Included in disclosed revenue group", "Revenue share / mobile advertising", "Mobile app monetization", faq],
        ["Google Services", "Google Network", "Google Ad Manager", "Ad-tech platform", "Included in disclosed revenue group", "Publisher ad serving / monetization", "Ad-tech infrastructure", faq],
        ["Google Services", "Google subscriptions, platforms, and devices", "YouTube TV", "Consumer subscription", "Included in disclosed revenue group", "Subscription fees", "Recurring subscription revenue", faq],
        ["Google Services", "Google subscriptions, platforms, and devices", "YouTube Music & Premium", "Consumer subscription", "Included in disclosed revenue group", "Subscription fees", "Recurring subscription / YouTube monetization", faq],
        ["Google Services", "Google subscriptions, platforms, and devices", "NFL Sunday Ticket", "Sports subscription / content", "Included in disclosed revenue group", "Subscription fees with content acquisition costs", "Premium sports content monetization", faq],
        ["Google Services", "Google subscriptions, platforms, and devices", "Google One Basic & Premium", "Cloud storage subscription", "Included in disclosed revenue group", "Consumer subscription", "Recurring consumer cloud revenue", faq],
        ["Google Services", "Google subscriptions, platforms, and devices", "Google AI Pro", "AI subscription", "Included in disclosed revenue group", "Consumer AI subscription", "Direct consumer AI monetization", faq],
        ["Google Services", "Google subscriptions, platforms, and devices", "Google AI Ultra", "AI subscription", "Included in disclosed revenue group", "Premium consumer AI subscription", "High-ARPU AI monetization", faq],
        ["Google Services", "Google subscriptions, platforms, and devices", "Google Play apps & in-app purchases", "Application platform", "Included in disclosed revenue group", "Net platform fees on app / in-app purchases", "App ecosystem economics", faq],
        ["Google Services", "Google subscriptions, platforms, and devices", "Pixel family", "Consumer hardware", "Included in disclosed revenue group", "Device sales", "Hardware / ecosystem control", faq],
        ["Google Services", "Other / not separately disclosed", "Android", "Operating system / platform", "No standalone revenue disclosed", "Ecosystem / licensing / indirect monetization", "Distribution moat / ecosystem", faq],
        ["Google Services", "Other / not separately disclosed", "Chrome", "Browser / platform", "No standalone revenue disclosed", "Indirect monetization / search distribution", "Search distribution / ecosystem", faq],
        ["Google Services", "Other / not separately disclosed", "Google Photos", "Consumer application", "No standalone revenue disclosed", "Freemium / subscription ecosystem", "Engagement / storage ecosystem", faq],
        ["Google Services", "Other / not separately disclosed", "YouTube core platform", "Video platform", "Revenue split across ads + subscriptions", "Ads + subscriptions + creator ecosystem", "Strategic media platform", sec],
        ["Google Cloud", "Google Cloud", "Google Cloud Platform (GCP)", "Cloud infrastructure/platform", "Included in segment revenue", "Consumption-based fees + subscriptions", "Primary enterprise cloud growth engine", faq],
        ["Google Cloud", "Google Cloud", "AI Infrastructure (Cloud TPUs & GPUs)", "AI infrastructure", "Included in segment revenue", "Compute consumption", "AI compute monetization", faq],
        ["Google Cloud", "Google Cloud", "Vertex AI", "AI developer platform", "Included in segment revenue", "Platform consumption / subscriptions", "Enterprise AI platform", faq],
        ["Google Cloud", "Google Cloud", "Cybersecurity", "Enterprise cloud service", "Included in segment revenue", "Consumption + subscription", "Cross-sell / workload expansion", faq],
        ["Google Cloud", "Google Cloud", "Data & analytics", "Enterprise cloud service", "Included in segment revenue", "Consumption + subscription", "Core enterprise workload", faq],
        ["Google Cloud", "Google Cloud", "Google Workspace", "Enterprise SaaS", "Included in segment revenue", "Per-seat subscriptions", "Recurring SaaS revenue", faq],
        ["Google Cloud", "Google Cloud", "Gemini for Google Workspace", "Enterprise AI application", "Included in segment revenue", "Workspace / AI subscription economics", "AI ARPU uplift / seat expansion", faq],
        ["Google Cloud", "Google Cloud", "Gemini Enterprise", "Enterprise AI application", "Included in segment revenue", "Enterprise subscriptions / consumption", "Enterprise AI monetization", faq],
        ["Google Cloud", "Google Cloud", "Customer Engagement Suite / agents", "Enterprise AI application", "Included in segment revenue", "Enterprise software / usage", "AI application layer", faq],
        ["Google Cloud", "Google Cloud", "Workspace apps: Gmail, Calendar, Docs, Drive, Meet", "Enterprise productivity suite", "Included in segment revenue", "Per-seat subscriptions", "Enterprise ecosystem / retention", sec],
        ["Other Bets", "Other Bets", "Waymo", "Autonomous transportation", "No standalone revenue disclosed", "Paid autonomous ride-hailing", "Largest commercial moonshot / optionality", faq],
        ["Other Bets", "Other Bets", "Verily", "Health technology", "No standalone revenue disclosed", "Healthcare-related services", "Healthcare optionality", faq],
        ["Other Bets", "Other Bets", "GFiber", "Internet services", "No standalone revenue disclosed", "Broadband / internet service fees", "Connectivity business", faq],
        ["Other Bets", "Other Bets", "Calico", "Biotechnology / longevity", "No standalone revenue disclosed", "Research partnerships / long-duration biotech economics", "Long-duration biotech optionality", faq],
        ["Other Bets", "Other Bets", "CapitalG", "Growth equity investing", "No standalone operating revenue disclosed", "Investment returns / portfolio appreciation", "Strategic investment exposure", faq],
        ["Other Bets", "Other Bets", "GV", "Venture capital investing", "No standalone operating revenue disclosed", "Investment returns / portfolio appreciation", "Early-stage innovation exposure", faq],
        ["Other Bets", "Other Bets", "Wing", "Autonomous delivery / drones", "No standalone revenue disclosed", "Delivery services / logistics platform", "Autonomous logistics optionality", faq],
        ["Other Bets", "Other Bets", "X", "Moonshot factory / R&D incubator", "No standalone revenue disclosed", "Research incubation / company creation", "Pipeline of future businesses", sec],
        ["Alphabet-level activities", "No revenue segment", "Shared general AI model R&D", "Shared AI R&D", "No standalone revenue", "Central R&D investment", "Supports multiple segments; margin drag today", faq],
        ["Alphabet-level activities", "No revenue segment", "Google DeepMind / frontier-model research", "AI research organization", "No standalone revenue", "Research / internal platform", "Strategic AI capability across Alphabet", sec],
    ]


def _financial_lookup(wb):
    out = {}
    if "Segment Analysis" not in wb.sheetnames:
        return out
    ws = wb["Segment Analysis"]
    for row in ws.iter_rows():
        label = row[0].value
        if not isinstance(label, str):
            continue
        vals = [c.value for c in row[1:6]]
        if len(vals) >= 3 and isinstance(vals[2], (int, float)):
            out[label.strip()] = float(vals[2])
    return out


def _build_alphabet_portfolio(wb, ticker, headers):
    discovered = _discover_alphabet_names(headers)
    base = _alphabet_static_rows()
    existing = {r[2].lower() for r in base}
    for name in discovered.get("services", []):
        if name.lower() not in existing:
            base.append(["Google Services", "Other / not separately disclosed", name, "Discovered Google Services product", "No standalone revenue disclosed", "Not separately disclosed", "Product ecosystem", ALPHABET_FAQ])
            existing.add(name.lower())
    for name in discovered.get("cloud", []):
        if name.lower() not in existing:
            base.append(["Google Cloud", "Google Cloud", name, "Discovered cloud product/service", "Included in segment revenue", "Cloud consumption / subscriptions", "Cloud product portfolio", ALPHABET_FAQ])
            existing.add(name.lower())
    for name in discovered.get("other_bets", []):
        if name.lower() not in existing:
            base.append(["Other Bets", "Other Bets", name, "Discovered Other Bet", "No standalone revenue disclosed", "Not separately disclosed", "Other Bets optionality", ALPHABET_FAQ])
            existing.add(name.lower())

    if "Business Portfolio Map" in wb.sheetnames:
        wb.remove(wb["Business Portfolio Map"])
    ws = wb.create_sheet("Business Portfolio Map")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:L2")
    ws["A1"] = f"{ticker} — Full Business Portfolio Map"
    ws["A1"].fill = _fill(NAVY)
    ws["A1"].font = Font(bold=True, color=WHITE, size=18)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells("A3:L4")
    ws["A3"] = (
        "Hierarchy of reported segments, disclosed revenue lines, products/platforms and Other Bets. "
        "Standalone revenue is shown only when Alphabet discloses it; otherwise the parent revenue line is shown."
    )
    ws["A3"].fill = _fill(LIGHT)
    ws["A3"].font = Font(italic=True, color=GREY)
    ws["A3"].alignment = Alignment(wrap_text=True)

    headers_row = [
        "Reported Segment", "Revenue Group", "Business / Product", "Category", "Disclosure Level",
        "2025 Standalone Revenue ($bn)", "2025 Parent Revenue ($bn)", "Parent Mix of Alphabet",
        "Monetization / Economic Model", "Key Evidence", "Investment Role", "Source",
    ]
    for c, v in enumerate(headers_row, 1):
        ws.cell(6, c, v)
        ws.cell(6, c).fill = _fill(BLUE)
        ws.cell(6, c).font = Font(bold=True, color=WHITE)
        ws.cell(6, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    financials = _financial_lookup(wb)
    total = None
    if "Historical Financials" in wb.sheetnames:
        total = wb["Historical Financials"]["G4"].value
        if not isinstance(total, (int, float)):
            total = None

    standalone = {"YouTube advertising": "YouTube ads"}
    for r_idx, item in enumerate(base, 7):
        segment, group, business, category, disclosure, monetization, role, source = item
        parent = financials.get(group)
        standalone_value = financials.get(standalone.get(business, "")) if business in standalone else None
        if business == "YouTube advertising" and standalone_value is not None:
            evidence = f"2025 YouTube ads revenue ${standalone_value:.1f}bn"
        elif business == "Google Cloud Platform (GCP)":
            evidence = "Google Cloud growth led by GCP and enterprise AI demand"
        elif business == "Waymo":
            evidence = "Other Bets revenue includes autonomous transportation; Waymo is scaling paid ride-hailing"
        elif business == "Google DeepMind / frontier-model research":
            evidence = "Shared frontier-model R&D is reported in Alphabet-level activities"
        elif parent is not None:
            evidence = f"Included in 2025 parent revenue of ${parent:.1f}bn"
        else:
            evidence = "No standalone revenue disclosed"
        vals = [segment, group, business, category, disclosure, standalone_value, parent, (parent / total if parent is not None and total else None), monetization, evidence, role, source]
        for c, v in enumerate(vals, 1):
            ws.cell(r_idx, c, v)
        for c in (6, 7):
            ws.cell(r_idx, c).number_format = FMT_BN
            ws.cell(r_idx, c).font = Font(color=INPUT_BLUE)
        ws.cell(r_idx, 8).number_format = FMT_PCT
        ws.cell(r_idx, 8).font = Font(color=INPUT_BLUE)
        for c in (1, 2, 3, 4, 5, 9, 10, 11, 12):
            ws.cell(r_idx, c).alignment = Alignment(wrap_text=True, vertical="top")
        if disclosure == "Standalone disclosed revenue line":
            ws.cell(r_idx, 5).fill = _fill(GREEN_FILL)
        elif "No standalone" in disclosure:
            ws.cell(r_idx, 5).fill = _fill(GOLD_FILL)

    end = 6 + len(base)
    summary = end + 3
    ws.merge_cells(start_row=summary, start_column=1, end_row=summary, end_column=12)
    ws.cell(summary, 1, "Portfolio Coverage Summary")
    ws.cell(summary, 1).fill = _fill(NAVY)
    ws.cell(summary, 1).font = Font(bold=True, color=WHITE)
    summary_rows = [
        ("Mapped businesses / products", len(base)),
        ("Official reportable segments", 3),
        ("Disclosed revenue groups", 6),
        ("Qualitative / parent-only items", sum(1 for r in base if "Standalone" not in r[4] or "No standalone" in r[4])),
    ]
    for i, (label, value) in enumerate(summary_rows, summary + 1):
        ws.cell(i, 1, label)
        ws.cell(i, 2, value)
    ws.cell(summary + 1, 1).fill = _fill(BLUE)
    ws.cell(summary + 1, 1).font = Font(bold=True, color=WHITE)
    ws.merge_cells(start_row=summary + 1, start_column=5, end_row=summary + 4, end_column=12)
    ws.cell(summary + 1, 5, (
        "Disclosure integrity: do not allocate estimated revenue to Android, Chrome, Maps, Cloud sub-products, "
        "Waymo, Verily, GFiber or other businesses when Alphabet does not publish standalone revenue. "
        "Use parent revenue groups for financial analysis and the portfolio map for strategic/business analysis."
    ))
    ws.cell(summary + 1, 5).fill = _fill(LIGHT)
    ws.cell(summary + 1, 5).alignment = Alignment(wrap_text=True, vertical="top")

    widths = {"A":22,"B":34,"C":34,"D":25,"E":30,"F":18,"G":18,"H":16,"I":34,"J":40,"K":34,"L":42}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A7"

    if "Segment Analysis" in wb.sheetnames:
        seg = wb["Segment Analysis"]
        row = max(seg.max_row + 2, 60)
        seg.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=min(14, seg.max_column or 14))
        seg.cell(row, 1, (
            f"Full Business Portfolio Map added: {len(base)} mapped Alphabet businesses/products. "
            "Use that sheet for product/business coverage; keep Segment Analysis as the official financial-segment view."
        ))
        seg.cell(row, 1).fill = _fill(PALE_BLUE)
        seg.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="center")
    return ws


def ensure_business_portfolio(wb, ticker, headers):
    if ticker.upper() in {"GOOGL", "GOOG"}:
        return _build_alphabet_portfolio(wb, ticker.upper(), headers)
    return None
