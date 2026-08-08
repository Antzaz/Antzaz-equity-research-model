"""Formula-cache-independent advanced analytics for the equity model.

Improvements: scorecard uses raw historical rows, loss years are excluded from P/E
statistics, DCF/Monte Carlo use the cash-conversion calibration, and Monte Carlo uses
correlated downside/base/strong-execution regimes rather than independent permanent
shocks around one base case.
"""

import random, statistics
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import LineChart, BarChart, RadarChart, Reference
from openpyxl.formatting.rule import DataBarRule
from advanced_analytics import _year_end_prices, _earnings_history, _solve_reverse_dcf

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; PALE_BLUE="D9EAF7"; GOLD="FFF2CC"; GREY="666666"
FMT_PCT='0.0%;[Red](0.0%);-'; FMT_PRICE='$#,##0.00;[Red]($#,##0.00);-'; FMT_BN='#,##0.0;[Red](#,##0.0);-'; FMT_MULT='0.0x;[Red](0.0x);-'

def _fill(c): return PatternFill("solid",fgColor=c)
def _num(v,default=None):
    try:
        if isinstance(v,bool): return default
        return float(v)
    except Exception: return default
def _latest_num(ws,row,start=2,end=7):
    for c in range(end,start-1,-1):
        v=ws.cell(row,c).value
        if isinstance(v,(int,float)) and not isinstance(v,bool): return float(v)
    return None
def _company_numbers(wb):
    ws=wb["Company Data"]; price=_num(ws["B8"].value,0); mc=_num(ws["B10"].value,0); cash=_num(ws["B12"].value,0); debt=_num(ws["B13"].value,0); shares=_num(ws["B9"].value) or ((mc/price) if price and mc else 1)
    return price,mc,shares,debt-cash
def _cash_adj(wb):
    if "Three-Case Scenarios" in wb.sheetnames:
        v=_num(wb["Three-Case Scenarios"]["AL4"].value)
        if v is not None: return max(-.05,min(.05,v))
    return 0.0

def _base_value(wb,growth_shock=0,margin_shock=0,capex_shock=0,wacc_shock=0,tgr_shock=0):
    s=wb["Three-Case Scenarios"]; h=wb["Historical Financials"]; _,_,shares,net_debt=_company_numbers(wb); rev=_latest_num(h,4) or 1; prev=rev; pv=0; fcf=0; wacc=max(.055,min(.16,_num(s["C6"].value,.09)+wacc_shock)); tgr=max(0,min(.06,_num(s["C7"].value,.03)+tgr_shock)); tax=_num(s["C8"].value,.21); adj=_cash_adj(wb); tgr=min(tgr,wacc-.005)
    for i,c in enumerate(range(14,24)):
        g=max(-.5,_num(s.cell(12,c).value,0)+growth_shock); margin=max(-.1,min(.6,_num(s.cell(14,c).value,.2)+margin_shock)); da=max(0,_num(s.cell(18,c).value,.04)); cap=max(.005,_num(s.cell(20,c).value,.1)+capex_shock); rev*=1+g; nwc=(rev-prev)*.01; fcf=rev*margin*(1-tax)+rev*da-rev*cap-nwc+rev*adj; pv+=fcf/((1+wacc)**(i+1)); prev=rev
    ev=pv+(fcf*(1+tgr)/(wacc-tgr))/((1+wacc)**10); return (ev-net_debt)/shares if shares else None

def _monte_carlo(wb,n=5000,seed=42):
    s=wb["Three-Case Scenarios"]; h=wb["Historical Financials"]; _,_,shares,net_debt=_company_numbers(wb); bg=[_num(s.cell(12,c).value,0) for c in range(14,24)]; bm=[_num(s.cell(14,c).value,.2) for c in range(14,24)]; bd=[_num(s.cell(18,c).value,.04) for c in range(14,24)]; bc=[_num(s.cell(20,c).value,.1) for c in range(14,24)]; tax=_num(s["C8"].value,.21); bw=_num(s["C6"].value,.09); bt=_num(s["C7"].value,.03); adj=_cash_adj(wb); start=_latest_num(h,4) or 1
    regimes=[(.25,-.025,-.020,.020,.0075,-.0025),(.50,0,0,0,0,0),(.25,.025,.020,-.015,-.003,.0025)]
    random.seed(seed); out=[]
    for _ in range(n):
        u=random.random(); cum=0; chosen=regimes[-1]
        for reg in regimes:
            cum+=reg[0]
            if u<=cum: chosen=reg; break
        _,rg,rm,rc,rw,rt=chosen; z=random.gauss(0,1); gs=rg+.012*z+random.gauss(0,.008); ms=rm+.010*z+random.gauss(0,.006); cs=rc-.008*z+random.gauss(0,.008); w=max(.055,min(.16,bw+rw-.002*z+random.gauss(0,.0025))); t=max(0,min(.06,bt+rt+.001*z+random.gauss(0,.0015))); t=min(t,w-.005); rev=start; prev=rev; pv=0; fcf=0
        for i in range(10):
            p=1-.35*(i/9); g=max(-.5,bg[i]+gs*p); margin=max(-.1,min(.6,bm[i]+ms*p)); cap=max(.005,bc[i]+cs*p); rev*=1+g; nwc=(rev-prev)*.01; fcf=rev*margin*(1-tax)+rev*bd[i]-rev*cap-nwc+rev*adj; pv+=fcf/((1+w)**(i+1)); prev=rev
        ev=pv+(fcf*(1+t)/(w-t))/((1+w)**10); out.append((ev-net_debt)/shares)
    return sorted(out)
def _scorecard(wb,current_price,forward_pe,base_value,severe_value):
    h=wb["Historical Financials"]; p=wb["Peer Comps"]; r0=_num(h["B4"].value); r1=_num(h["G4"].value); cagr=((r1/r0)**.2-1) if r0 and r1 and r0>0 and r1>0 else 0; rev=_num(h["G4"].value); op=_num(h["G9"].value); ocf=_num(h["G14"].value); cap=_num(h["G15"].value); opm=(op/rev) if rev and op is not None else 0; fcfm=((ocf-cap)/rev) if rev and ocf is not None and cap is not None else 0; pes=[_num(p.cell(r,3).value) for r in range(5,10)]; pes=[x for x in pes if x and x>0]; med=statistics.median(pes) if pes else forward_pe or 1; _,_,_,net_debt=_company_numbers(wb)
    return [("Growth",min(100,max(0,cagr/.20*100)),"5Y revenue CAGR"),("Profitability",min(100,max(0,opm/.40*100)),"Latest operating margin"),("FCF Quality",min(100,max(0,fcfm/.25*100)),"Latest reported FCF margin"),("Balance Sheet",85 if net_debt<0 else 55,"Net cash / debt profile"),("Absolute Valuation",min(100,max(0,base_value/current_price*100)) if current_price and base_value is not None else 0,"Base DCF vs market"),("Relative Valuation",min(100,max(0,med/forward_pe*70)) if forward_pe else 50,"Forward P/E vs peers"),("Stress Robustness",min(100,max(0,severe_value/current_price*100)) if current_price and severe_value is not None else 0,"Combined severe-bear value")]
def _section(ws,rng,title): ws.merge_cells(rng); c=ws[rng.split(":")[0]]; c.value=title; c.fill=_fill(NAVY); c.font=Font(bold=True,color=WHITE,size=12)
def _header(ws,row,start,end):
    for c in range(start,end+1): ws.cell(row,c).fill=_fill(BLUE); ws.cell(row,c).font=Font(bold=True,color=WHITE); ws.cell(row,c).alignment=Alignment(horizontal="center",wrap_text=True)

def ensure_advanced_analytics(wb,ticker,info=None):
    if not {"Company Data","Historical Financials","Three-Case Scenarios","Peer Comps"}.issubset(wb.sheetnames): return None
    if "Advanced Analytics" in wb.sheetnames: wb.remove(wb["Advanced Analytics"])
    ws=wb.create_sheet("Advanced Analytics"); ws.sheet_view.showGridLines=False; ws.merge_cells("A1:P2"); ws["A1"]="Advanced Valuation & Expectations Analytics"; ws["A1"].fill=_fill(NAVY); ws["A1"].font=Font(bold=True,color=WHITE,size=18); ws.merge_cells("A3:P3"); ws["A3"]="Historical valuation, earnings surprises, reverse DCF, regime-based Monte Carlo valuation and investment scorecard."; ws["A3"].font=Font(italic=True,color=GREY)
    h=wb["Historical Financials"]; s=wb["Three-Case Scenarios"]; price,mc,shares,net_debt=_company_numbers(wb); fpe=_num(wb["Company Data"]["B15"].value,_num((info or {}).get("forwardPE"),0))
    _section(ws,"A5:H5","Historical Price & Valuation"); years=[]; eps={}
    for c in range(2,8):
        y=h.cell(3,c).value
        if isinstance(y,(int,float)): years.append(int(y)); eps[int(y)]=_num(h.cell(12,c).value)
    prices=_year_end_prices(ticker,years)
    for c,v in enumerate(["Year","Year-End Price","Diluted EPS","Year-End P/E"],1): ws.cell(6,c,v)
    _header(ws,6,1,4); pe=[]; row=7
    for y in years:
        px=prices.get(y); e=eps.get(y); p_e=(px/e) if px and e and e>0 else None
        if p_e and p_e>0: pe.append(p_e)
        ws.cell(row,1,y); ws.cell(row,2,px); ws.cell(row,3,e); ws.cell(row,4,p_e); ws.cell(row,2).number_format=FMT_PRICE; ws.cell(row,3).number_format=FMT_PRICE; ws.cell(row,4).number_format=FMT_MULT; row+=1
    ws["F6"]="Valuation Statistic"; ws["G6"]="P/E"; _header(ws,6,6,7)
    for r,(label,val) in enumerate([("Historical Min",min(pe) if pe else None),("Historical Median",statistics.median(pe) if pe else None),("Historical Max",max(pe) if pe else None),("Current Forward P/E",fpe)],7): ws.cell(r,6,label); ws.cell(r,7,val); ws.cell(r,7).number_format=FMT_MULT
    if row>8:
        ch=LineChart(); ch.style=10; ch.title="Year-End Share Price"; ch.height=7; ch.width=13; ch.legend=None; ch.add_data(Reference(ws,min_col=2,min_row=6,max_row=row-1),titles_from_data=True); ch.set_categories(Reference(ws,min_col=1,min_row=7,max_row=row-1)); ch.visible_cells_only=False; ws.add_chart(ch,"A14")
    _section(ws,"I5:P5","Recent Earnings Surprises")
    for c,v in enumerate(["Reported","EPS Estimate","Actual EPS","EPS Surprise"],9): ws.cell(6,c,v)
    _header(ws,6,9,12); erow=7
    for e in _earnings_history(ticker,8): ws.cell(erow,9,e["date"]); ws.cell(erow,10,e["eps_est"]); ws.cell(erow,11,e["eps_actual"]); ws.cell(erow,12,e["eps_surprise"]); ws.cell(erow,10).number_format=FMT_PRICE; ws.cell(erow,11).number_format=FMT_PRICE; ws.cell(erow,12).number_format=FMT_PCT; erow+=1
    if erow==7: ws["I7"]="No earnings-surprise history returned by data provider."
    else:
        ch=BarChart(); ch.type="col"; ch.style=10; ch.title="EPS Surprise %"; ch.height=7; ch.width=13; ch.legend=None; ch.add_data(Reference(ws,min_col=12,min_row=6,max_row=erow-1),titles_from_data=True); ch.set_categories(Reference(ws,min_col=9,min_row=7,max_row=erow-1)); ch.y_axis.numFmt="0%"; ch.visible_cells_only=False; ws.add_chart(ch,"I14")
    ocf=_latest_num(h,14); capex=_latest_num(h,15); sbc=_latest_num(h,21) or 0; reported=(ocf-capex) if ocf is not None and capex is not None else None; normalized=(reported-sbc) if reported is not None else None; start_fcf=normalized if normalized is not None and normalized>0 else reported; wacc=_num(s["C6"].value,.09); tgr=_num(s["C7"].value,.03); target_ev=price*shares+net_debt; implied=_solve_reverse_dcf(start_fcf,target_ev,wacc,tgr)
    _section(ws,"A31:H31","Reverse DCF — Market-Implied Expectations"); ws["A32"]="Input / Output"; ws["B32"]="Value"; ws["C32"]="Interpretation"; _header(ws,32,1,3)
    rr=[("Current Price",price,FMT_PRICE,"Market price"),("Current Enterprise Value ($bn)",target_ev,FMT_BN,"Equity value plus net debt"),("Normalized Starting FCF ($bn)",start_fcf,FMT_BN,"Uses OCF-capex-SBC when positive; otherwise OCF-capex. See FCF & Capital Economics."),("WACC",wacc,FMT_PCT,"Base discount rate"),("Terminal Growth",tgr,FMT_PCT,"Base perpetual growth"),("Implied 10Y FCF CAGR",implied,FMT_PCT,"Constant FCF growth required by current price")]
    for r,(label,val,fmt,note) in enumerate(rr,33): ws.cell(r,1,label); ws.cell(r,2,val); ws.cell(r,2).number_format=fmt; ws.cell(r,3,note)
    ws["B38"].fill=_fill(GOLD); ws["B38"].font=Font(bold=True)
    base_val=_base_value(wb); severe=_base_value(wb,-.05,-.05,.05,.02,-.01); values=_monte_carlo(wb,5000); _section(ws,"I31:P31","Regime Monte Carlo Valuation — 5,000 Simulations"); ws["I32"]="Metric"; ws["J32"]="Result"; _header(ws,32,9,10)
    if values:
        pct=lambda p: values[int((len(values)-1)*p)]; summ=[("P10 Value / Share",pct(.10),FMT_PRICE),("P25 Value / Share",pct(.25),FMT_PRICE),("Median Value / Share",pct(.50),FMT_PRICE),("P75 Value / Share",pct(.75),FMT_PRICE),("P90 Value / Share",pct(.90),FMT_PRICE),("Probability > Current Price",sum(v>price for v in values)/len(values),FMT_PCT),("Probability > Base DCF",sum(v>base_val for v in values)/len(values),FMT_PCT)]
        for r,(label,val,fmt) in enumerate(summ,33): ws.cell(r,9,label); ws.cell(r,10,val); ws.cell(r,10).number_format=fmt
        lo,hi=min(values),max(values); width=(hi-lo)/20 if hi>lo else 1; ws["R32"]="Value / Share"; ws["S32"]="Frequency"
        for i in range(20):
            left=lo+i*width; right=left+width; ws.cell(33+i,18,(left+right)/2); ws.cell(33+i,19,sum(1 for v in values if v>=left and (v<right or i==19))); ws.cell(33+i,18).number_format=FMT_PRICE
        ch=BarChart(); ch.type="col"; ch.style=10; ch.title="Regime Monte Carlo Valuation Distribution"; ch.height=8; ch.width=13; ch.legend=None; ch.add_data(Reference(ws,min_col=19,min_row=32,max_row=52),titles_from_data=True); ch.set_categories(Reference(ws,min_col=18,min_row=33,max_row=52)); ch.visible_cells_only=False; ws.add_chart(ch,"I41")
    _section(ws,"A41:H41","Investment Scorecard"); ws["A42"]="Dimension"; ws["B42"]="Score (0–100)"; ws["C42"]="Comment"; _header(ws,42,1,3); scores=_scorecard(wb,price,fpe,base_val,severe)
    for r,(name,score,note) in enumerate(scores,43): ws.cell(r,1,name); ws.cell(r,2,score); ws.cell(r,3,note)
    composite=sum(x[1] for x in scores)/len(scores); ws["E42"]="Composite Score"; ws["F42"]=composite; ws["F42"].number_format="0.0"; ws["E42"].font=Font(bold=True); ws.conditional_formatting.add("B43:B49",DataBarRule(start_type="num",start_value=0,end_type="num",end_value=100,color="5B9BD5")); radar=RadarChart(); radar.type="filled"; radar.style=26; radar.title="Investment Scorecard Radar"; radar.height=8; radar.width=11; radar.add_data(Reference(ws,min_col=2,min_row=42,max_row=49),titles_from_data=True); radar.set_categories(Reference(ws,min_col=1,min_row=43,max_row=49)); radar.visible_cells_only=False; ws.add_chart(radar,"A52")
    ws.merge_cells("A70:P72"); ws["A70"]="Live sources: Yahoo Finance for price history and earnings dates; SEC-based workbook inputs for reported financials. Regime Monte Carlo uses correlated downside/base/strong-execution regimes. Results are analytical diagnostics, not recommendations."; ws["A70"].fill=_fill(PALE_BLUE); ws["A70"].alignment=Alignment(wrap_text=True); ws.column_dimensions["A"].width=29; ws.column_dimensions["C"].width=38
    for c in "BDEFGHIJKLMNOP": ws.column_dimensions[c].width=14
    ws.column_dimensions["R"].hidden=True; ws.column_dimensions["S"].hidden=True; return ws
