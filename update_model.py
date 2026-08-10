"""Cross-company equity research updater.

Usage: python update_model.py GOOGL

Reliability rules:
- SEC annual facts are selected from annual filing periods;
- observed history is right-aligned so column G is always the latest actual;
- target sector/industry is detected automatically from the target ticker;
- peer companies must match the target sector, with exact industry preferred;
- stale template peers are never used as a fallback;
- institutional sheets use Excel-safe worksheet names;
- final reliability/segment/peer controls run again before save.
"""

import importlib
import os
import re
import statistics
import subprocess
import sys
from datetime import date, datetime
from pathlib import Path

DEPS={"requests":"requests","yfinance":"yfinance","openpyxl":"openpyxl","pandas":"pandas","lxml":"lxml"}
for module,package in DEPS.items():
    try:
        importlib.import_module(module)
    except ImportError:
        subprocess.check_call([sys.executable,"-m","pip","install",package])

import requests
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.workbook.properties import CalcProperties

from stress_test import ensure_stress_test
import company_analysis as _company_analysis
from company_analysis import ensure_financial_statements, ensure_segment_analysis
from segment_analysis_v2 import ensure_segment_analysis_v2
from advanced_analytics_v2 import ensure_advanced_analytics
from visualization_v2 import ensure_visual_dashboard
from analysis_charts import ensure_analysis_charts
from model_quality_v3 import calibrate_scenario_cash_flow, ensure_model_quality
import institutional_layers as _institutional_layers
from institutional_layers import ensure_institutional_layers
from segment_chart_fix import repair_segment_charts
from cross_company_cleanup import refresh_cross_company_tabs
from dynamic_peer_engine import ensure_dynamic_peer_comps
from model_reliability import prepare_model_reliability
from investment_summary import ensure_investment_summary
from ownership_analysis import ensure_ownership_analysis
from currency_normalization import normalize_workbook_currency
from consensus_quality import normalize_expectations_consensus
from research_extensions import ensure_research_extensions

# institutional_layers historically created a 33-character temporary sheet name.
# Patch its sheet factory once so both institutional refreshes are warning-free.
_ORIGINAL_INSTITUTIONAL_CREATE=_institutional_layers._create
def _safe_institutional_create(wb,name):
    if name=="Base Rates & Thesis Probabilities":
        name="Base Rates & Probabilities"
    return _ORIGINAL_INSTITUTIONAL_CREATE(wb,name)
_institutional_layers._create=_safe_institutional_create

BASE=Path(__file__).resolve().parent
TEMPLATE=BASE/"GOOGL_Equity_Research_CLEAN_v7.xlsx"
OUTDIR=BASE/"updated_models"
OUTDIR.mkdir(exist_ok=True)
SEC_HEADERS={"User-Agent":os.getenv("SEC_USER_AGENT","Personal Equity Research Model contact@example.com")}


def sec_json(url):
    r=requests.get(url,headers=SEC_HEADERS,timeout=30)
    r.raise_for_status()
    return r.json()


def cik_for(ticker):
    for item in sec_json("https://www.sec.gov/files/company_tickers.json").values():
        if item["ticker"].upper()==ticker.upper():
            return str(item["cik_str"]).zfill(10)
    return None


def company_facts(ticker):
    cik=cik_for(ticker)
    return sec_json(f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json") if cik else None


def annual_series(facts,tags,preferred_unit=None):
    if not facts:
        return {}
    gaap=facts.get("facts",{}).get("us-gaap",{})
    best={}
    for rank,tag in enumerate(tags):
        fact=gaap.get(tag)
        if not fact:
            continue
        units=fact.get("units",{})
        unit=preferred_unit if preferred_unit in units else (next(iter(units)) if units else None)
        if not unit:
            continue
        for x in units.get(unit,[]):
            if x.get("form") not in ("10-K","10-K/A"):
                continue
            val=x.get("val"); end=x.get("end"); start=x.get("start")
            if val is None or not end:
                continue
            if start:
                try:
                    days=(date.fromisoformat(end)-date.fromisoformat(start)).days
                    if days<250 or days>450:
                        continue
                except Exception:
                    pass
            try:
                year=int(str(end)[:4]); val=float(val)
            except Exception:
                continue
            stamp=(str(x.get("filed") or ""),str(end),-rank)
            if year not in best or stamp>best[year][0]:
                best[year]=(stamp,val)
    return {y:v for y,(_,v) in best.items()}


def build_history(ticker,facts=None):
    facts=facts or company_facts(ticker)
    revenue=annual_series(facts,["RevenueFromContractWithCustomerExcludingAssessedTax","SalesRevenueNet","Revenues"])
    cost=annual_series(facts,["CostOfRevenue","CostOfGoodsAndServicesSold"])
    gross=annual_series(facts,["GrossProfit"])
    op=annual_series(facts,["OperatingIncomeLoss"])
    ni=annual_series(facts,["NetIncomeLoss","ProfitLoss"])
    eps=annual_series(facts,["EarningsPerShareDiluted"],"USD/shares")
    ocf=annual_series(facts,["NetCashProvidedByUsedInOperatingActivities"])
    capex=annual_series(facts,["PaymentsToAcquirePropertyPlantAndEquipment","PaymentsToAcquireProductiveAssets"])
    depr=annual_series(facts,["DepreciationDepletionAndAmortizationPropertyPlantAndEquipment","DepreciationDepletionAndAmortization"])
    rd=annual_series(facts,["ResearchAndDevelopmentExpense"])
    sbc=annual_series(facts,["ShareBasedCompensation"])
    years=sorted(revenue)[-6:] if revenue else []
    out={}
    for y in years:
        rev=revenue.get(y); c=cost.get(y); gp=gross.get(y)
        if gp is None and rev is not None and c is not None:
            gp=rev-c
        cx=abs(capex[y]) if y in capex else None
        fcf=ocf.get(y)-cx if ocf.get(y) is not None and cx is not None else None
        shares=ni[y]/eps[y] if y in ni and y in eps and eps[y] else None
        out[y]={"revenue":rev,"cost":c,"gross":gp,"op":op.get(y),"ni":ni.get(y),"eps":eps.get(y),"shares":shares,"ocf":ocf.get(y),"capex":cx,"fcf":fcf,"depr":depr.get(y),"rd":rd.get(y),"sbc":sbc.get(y)}
    return out


def yf_info(ticker):
    try:
        return yf.Ticker(ticker).info or {}
    except Exception:
        return {}


def put_company(wb,ticker,info):
    ws=wb["Company Data"]
    p=info.get("currentPrice") or info.get("regularMarketPrice")
    mc=info.get("marketCap"); cash=info.get("totalCash"); debt=info.get("totalDebt")
    vals={"B4":ticker,"B5":info.get("longName") or ticker,"B6":info.get("sector"),"B7":info.get("industry"),"B8":p,"B10":mc/1e9 if mc else None,"B11":info.get("enterpriseValue")/1e9 if info.get("enterpriseValue") else None,"B12":cash/1e9 if cash else None,"B13":debt/1e9 if debt else None,"B15":info.get("forwardPE")}
    if p and mc:
        vals["B9"]=(mc/1e9)/p
    if cash is not None and debt is not None:
        vals["B14"]=(debt-cash)/1e9
    for cell,value in vals.items():
        if value is not None:
            ws[cell]=value


def put_history(wb,hist):
    ws=wb["Historical Financials"]
    for r in (3,4,6,9,11,12,14,15,18,19,21):
        for c in range(2,8):
            ws.cell(r,c).value=None
    if not hist:
        print("Warning: SEC annual history unavailable; historical raw-input rows left blank for review.")
        return
    for i,y in enumerate(sorted(hist)[-6:],2):
        d=hist[y]
        scale=lambda x:x/1e9 if x is not None else None
        ws.cell(3,i,y); ws.cell(4,i,scale(d.get("revenue"))); ws.cell(6,i,scale(d.get("cost"))); ws.cell(9,i,scale(d.get("op"))); ws.cell(11,i,scale(d.get("ni"))); ws.cell(12,i,d.get("eps")); ws.cell(14,i,scale(d.get("ocf"))); ws.cell(15,i,scale(d.get("capex"))); ws.cell(18,i,scale(d.get("depr"))); ws.cell(19,i,scale(d.get("rd"))); ws.cell(21,i,scale(d.get("sbc")))


def fade(start,end):
    return [start+(end-start)*i/9 for i in range(10)]


def _template_history_fallback(wb):
    ws=wb["Historical Financials"]
    out={}
    for c in range(2,8):
        try:
            year=int(ws.cell(3,c).value)
        except Exception:
            continue
        raw=lambda v:float(v)*1e9 if isinstance(v,(int,float)) else None
        out[year]={"revenue":raw(ws.cell(4,c).value),"op":raw(ws.cell(9,c).value),"capex":raw(ws.cell(15,c).value),"depr":raw(ws.cell(18,c).value),"ocf":raw(ws.cell(14,c).value),"ni":raw(ws.cell(11,c).value),"eps":ws.cell(12,c).value if isinstance(ws.cell(12,c).value,(int,float)) else None,"sbc":raw(ws.cell(21,c).value)}
    return out


def update_scenarios(wb,hist,info):
    ws=wb["Three-Case Scenarios"]
    merged=dict(_template_history_fallback(wb)); merged.update(hist or {})
    years=sorted(y for y,d in merged.items() if d and d.get("revenue"))
    hc=info.get("revenueGrowth") or .10; opm=info.get("operatingMargins") or .20; cap=.10; dep=.04; cap_hist=[]
    if len(years)>=2:
        last=years[-1]; first=years[max(0,len(years)-6)]
        r0=merged[first].get("revenue"); r1=merged[last].get("revenue"); n=max(1,last-first)
        if r0 and r1 and r0>0 and r1>0:
            hc=(r1/r0)**(1/n)-1
        latest=merged[last]; rev=latest.get("revenue")
        if rev:
            if latest.get("op") is not None: opm=latest["op"]/rev
            if latest.get("capex") is not None: cap=abs(latest["capex"])/rev
            if latest.get("depr") is not None: dep=abs(latest["depr"])/rev
        for y in years[-3:]:
            d=merged[y]; rv=d.get("revenue"); cx=d.get("capex")
            if rv and cx is not None:
                cap_hist.append(abs(cx)/rv)
    hc=max(.01,min(.25,float(hc))); opm=max(.01,min(.50,float(opm))); cap=max(.01,min(.40,float(cap))); dep=max(.005,min(.15,float(dep)))
    base=max(.04,min(.20,hc)); growths=[fade(max(-.02,base-.05),.03),fade(base,.055),fade(min(.30,base+.04),.08)]; blocks=[range(2,12),range(14,24),range(26,36)]
    for cols,gs in zip(blocks,growths):
        for c,v in zip(cols,gs): ws.cell(12,c).value=v
    margins=[fade(max(-.05,opm-.03),max(-.05,opm-.04)),fade(opm,min(.55,opm+.01)),fade(min(.60,opm+.01),min(.60,opm+.04))]
    for cols,ms in zip(blocks,margins):
        for c,v in zip(cols,ms): ws.cell(14,c).value=v
    med_cap=statistics.median(cap_hist) if cap_hist else cap; normalized=max(.02,min(.22,max(dep,med_cap*.65))); starts=[min(.40,cap+.03),cap,max(.01,cap-.02)]; targets=[min(.30,normalized+.03),normalized,max(.015,normalized-.02)]
    for cols,start,target in zip(blocks,starts,targets):
        for c,v in zip(cols,fade(start,target)): ws.cell(20,c).value=v
    for cols in blocks:
        for c in cols: ws.cell(18,c).value=dep


def update_filings(wb,ticker):
    cik=cik_for(ticker)
    if not cik:
        return
    recent=sec_json(f"https://data.sec.gov/submissions/CIK{cik}.json").get("filings",{}).get("recent",{})
    ws=wb["Filings"]
    for r in range(4,20):
        for c in range(1,6): ws.cell(r,c).value=None
    row=4
    for form,period,filed,acc,doc in zip(recent.get("form",[]),recent.get("reportDate",[]),recent.get("filingDate",[]),recent.get("accessionNumber",[]),recent.get("primaryDocument",[])):
        if form not in {"10-K","10-Q","8-K","DEF 14A"}:
            continue
        url=f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{acc.replace('-','')}/{doc}"
        for c,v in enumerate([form,period,filed,url,"SEC filing"],1): ws.cell(row,c).value=v
        row+=1
        if row>15: break


def _fix_all_charts(wb):
    for ws in wb.worksheets:
        for ch in getattr(ws,"_charts",[]):
            try:
                ch.visible_cells_only=False; ch.display_blanks="gap"
            except Exception:
                pass


def _ensure_excel_sheet_names(wb):
    replacements={"Base Rates & Thesis Probabilities":"Base Rates & Probabilities"}
    for old,new in replacements.items():
        if old in wb.sheetnames:
            if new in wb.sheetnames and new!=old:
                wb.remove(wb[new])
            wb[old].title=new
    invalid=[]; bad_chars=re.compile(r'[:\\/?*\[\]]')
    for name in wb.sheetnames:
        if len(name)>31 or bad_chars.search(name): invalid.append(name)
    if invalid:
        raise ValueError("Invalid Excel worksheet name(s): "+", ".join(repr(x) for x in invalid))


def get_ticker():
    raw=sys.argv[1] if len(sys.argv)>1 else input("Ticker (e.g. GOOGL): ")
    raw=raw.strip()
    if " " in raw or ".py" in raw.lower():
        parts=raw.replace('"','').replace("'","").split(); candidates=[p.upper() for p in parts if re.fullmatch(r"[A-Za-z0-9.\-]{1,10}",p) and not p.lower().endswith(".py") and p.lower() not in {"python","py"}]
        if candidates: raw=candidates[-1]
    ticker=raw.upper().strip()
    if not re.fullmatch(r"[A-Z0-9.\-]{1,10}",ticker):
        raise ValueError(f"Invalid ticker: {ticker!r}. Enter only a ticker, e.g. GOOGL, MSFT, or NVDA.")
    return ticker


def main():
    ticker=get_ticker(); print(f"Ticker: {ticker}")
    if not TEMPLATE.exists(): raise FileNotFoundError(TEMPLATE)
    info=yf_info(ticker)
    try:
        facts=company_facts(ticker)
    except Exception as exc:
        print(f"Warning: SEC Company Facts unavailable: {exc}"); facts=None
    hist=build_history(ticker,facts)
    wb=load_workbook(TEMPLATE,data_only=False)
    put_company(wb,ticker,info); put_history(wb,hist)
    try:
        prepare_model_reliability(wb,ticker)
    except Exception as exc:
        print(f"Warning: initial history normalization failed: {exc}")
    try:
        print("Normalizing reporting currency and ADR units...")
        normalize_workbook_currency(wb,ticker,info)
    except Exception as exc:
        print(f"Warning: currency / ADR normalization failed: {exc}")
    update_scenarios(wb,hist,info); ensure_stress_test(wb); calibrate_scenario_cash_flow(wb)
    try:
        peers=ensure_dynamic_peer_comps(wb,ticker)
        sector=wb["Company Data"]["B6"].value; industry=wb["Company Data"]["B7"].value
        print(f"Peer classification: sector={sector!r}, industry={industry!r}, validated peers={len(peers)}")
    except Exception as exc:
        print(f"Warning: dynamic peer selection failed: {exc}")
    update_filings(wb,ticker)
    try: _company_analysis._merged_annual_series=annual_series
    except Exception: pass
    try: ensure_financial_statements(wb,ticker,facts)
    except Exception as exc: print(f"Warning: Financial Statements module failed: {exc}")
    try:
        if ticker in {"GOOGL","GOOG"}: ensure_segment_analysis(wb,ticker,SEC_HEADERS)
        else: ensure_segment_analysis_v2(wb,ticker,SEC_HEADERS)
    except Exception as exc: print(f"Warning: Segment Analysis module failed: {exc}")
    print("Building advanced valuation analytics...")
    try: ensure_advanced_analytics(wb,ticker,info)
    except Exception as exc: print(f"Warning: Advanced Analytics module failed: {exc}")
    try: ensure_visual_dashboard(wb,ticker)
    except Exception as exc: print(f"Warning: Visual Dashboard module failed: {exc}")
    try:
        if ticker not in {"GOOGL","GOOG"}: ensure_segment_analysis_v2(wb,ticker,SEC_HEADERS)
    except Exception as exc: print(f"Warning: Final Segment Analysis refresh failed: {exc}")
    try: ensure_analysis_charts(wb,ticker)
    except Exception as exc: print(f"Warning: Analysis Charts module failed: {exc}")
    print("Running model-quality and institutional controls...")
    try: ensure_model_quality(wb,ticker)
    except Exception as exc: print(f"Warning: Model Quality / Research Workbench failed: {exc}")
    try: ensure_institutional_layers(wb,ticker)
    except Exception as exc: print(f"Warning: Institutional Layers failed: {exc}")
    try: refresh_cross_company_tabs(wb,ticker)
    except Exception as exc: print(f"Warning: Cross-company cleanup failed: {exc}")
    print("Finalizing source checks, segment enrichment, valuation analytics and news...")
    try: repair_segment_charts(wb,ticker)
    except Exception as exc: print(f"Warning: Segment chart / final reliability repair failed: {exc}")
    try:
        normalize_workbook_currency(wb,ticker,info)
    except Exception as exc:
        print(f"Warning: final currency / ADR normalization failed: {exc}")
    print("Building final investment summary and ownership view...")
    try: ensure_investment_summary(wb,ticker)
    except Exception as exc: print(f"Warning: Investment Summary failed: {exc}")
    try: ensure_ownership_analysis(wb,ticker)
    except Exception as exc: print(f"Warning: Ownership & Holders failed: {exc}")

    # Run de-duplication and public-data quality controls only after every legacy generator
    # has finished, so removed presentation tabs are not recreated later in the pipeline.
    print("Consolidating duplicate tabs and extending market, workforce and leadership research...")
    try: normalize_expectations_consensus(wb,ticker,info)
    except Exception as exc: print(f"Warning: consensus currency alignment failed: {exc}")
    try: ensure_research_extensions(wb,ticker,info)
    except Exception as exc: print(f"Warning: research extension / workbook consolidation failed: {exc}")

    _ensure_excel_sheet_names(wb); _fix_all_charts(wb)
    if "Dashboard" in wb.sheetnames:
        wb["Dashboard"]["A1"]=f"{ticker} Long-Term Value Investing Dashboard"
    try:
        if getattr(wb,"calculation",None) is None:
            wb.calculation=CalcProperties(calcMode="auto",fullCalcOnLoad=True,forceFullCalc=True)
        else:
            wb.calculation.calcMode="auto"; wb.calculation.fullCalcOnLoad=True; wb.calculation.forceFullCalc=True
    except Exception as exc:
        print(f"Warning: could not set workbook recalculation flags: {exc}")
    out=OUTDIR/f"{ticker}_Equity_Research_{datetime.now():%Y%m%d_%H%M%S}.xlsx"; wb.save(out); print("Saved:",out)


if __name__=="__main__":
    main()