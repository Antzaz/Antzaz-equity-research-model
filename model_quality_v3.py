"""Cross-company model quality, cash-flow reconciliation and manual research workbench.

This layer is intentionally generic. It does not force every company into Alphabet-
specific logic. It reconciles cash-flow definitions, separates maintenance/growth
capex, adds balanced upside sensitivities, independent valuation cross-checks, data
quality checks, and a structured manual research workbench.
"""

import statistics
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; LIGHT="F5F9FC"; PALE_GREEN="E2F0D9"; PALE_RED="FCE4D6"; GOLD="FFF2CC"; GREY="666666"; INPUT_BLUE="0000FF"
FMT_PCT='0.0%;[Red](0.0%);-'; FMT_PRICE='$#,##0.00;[Red]($#,##0.00);-'; FMT_BN='#,##0.0;[Red](#,##0.0);-'
THIN=Side(style="thin", color="D9E1F2")


def _fill(c): return PatternFill("solid",fgColor=c)
def _num(v,default=None):
    try:
        if isinstance(v,bool): return default
        return float(v)
    except Exception: return default

def _median(xs,default=0):
    xs=[float(x) for x in xs if isinstance(x,(int,float)) and not isinstance(x,bool)]
    return statistics.median(xs) if xs else default

def _merge(ws,rng):
    if rng not in {str(x) for x in ws.merged_cells.ranges}: ws.merge_cells(rng)
def _section(ws,rng,title):
    _merge(ws,rng); c=ws[rng.split(":")[0]]; c.value=title; c.fill=_fill(NAVY); c.font=Font(bold=True,color=WHITE,size=11); c.alignment=Alignment(vertical="center")
def _header(ws,row,start,end):
    for c in range(start,end+1):
        x=ws.cell(row,c); x.fill=_fill(BLUE); x.font=Font(bold=True,color=WHITE); x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); x.border=Border(bottom=THIN)


def cash_conversion_adjustment(wb):
    """3Y median OCF bridge excluding SBC, capped to +/-5% of revenue."""
    if "Historical Financials" not in wb.sheetnames: return 0.0
    h=wb["Historical Financials"]; tax=.21; vals=[]
    if "Three-Case Scenarios" in wb.sheetnames: tax=_num(wb["Three-Case Scenarios"]["C8"].value,.21)
    for c in range(5,8):
        rev=_num(h.cell(4,c).value); op=_num(h.cell(9,c).value); ocf=_num(h.cell(14,c).value); da=_num(h.cell(18,c).value,0); sbc=_num(h.cell(21,c).value,0)
        if rev and op is not None and ocf is not None:
            vals.append((ocf-op*(1-tax)-da-sbc)/rev)
    return max(-.05,min(.05,_median(vals,0.0)))


def calibrate_scenario_cash_flow(wb):
    """Align forward FCFF and stress tests with historical cash conversion."""
    if "Three-Case Scenarios" not in wb.sheetnames: return 0.0
    ws=wb["Three-Case Scenarios"]; adj=cash_conversion_adjustment(wb)
    ws["AK3"]="Cash-Flow Calibration"; ws["AK4"]="Other cash conversion % revenue"; ws["AL4"]=adj; ws["AL4"].number_format=FMT_PCT
    ws["AK5"]="Definition"; ws["AL5"]="3Y median of OCF - NOPAT - D&A - SBC, divided by revenue; capped at +/-5%."
    ws["AK6"]="Purpose"; ws["AL6"]="Reconciles forward FCFF with historical cash conversion while treating SBC as an economic cost."
    ws.column_dimensions["AK"].hidden=True; ws.column_dimensions["AL"].hidden=True
    for start in (2,14,26):
        for c in range(start,start+10):
            col=get_column_letter(c)
            ws.cell(22,c).value=(f"={col}13*0.01-'Historical Financials'!G4*0.01" if c==start else f"={col}13*0.01-{get_column_letter(c-1)}13*0.01")
            ws.cell(23,c).value=f"={col}17+{col}19-{col}21-{col}22+{col}13*$AL$4"
            ws.cell(23,c).number_format=FMT_BN; ws.cell(24,c).value=f"=IFERROR({col}23/{col}13,\"\")"; ws.cell(24,c).number_format=FMT_PCT
    for c in (1,13,25): ws.cell(22,c).value="NWC Investment (1% of revenue change)"
    if ws["A48"].value=="Base Control":
        base_cols=[get_column_letter(14+i) for i in range(10)]; fcf_cols=[get_column_letter(12+i) for i in range(10)]; rev_cols=[get_column_letter(22+i) for i in range(10)]
        for r in range(48,59):
            for i,(fcf_c,rev_c,base_c) in enumerate(zip(fcf_cols,rev_cols,base_cols)):
                if i==0: prev_rev="'Historical Financials'!G4"; nwc=f"IFERROR({base_c}$22/({base_c}$13-'Historical Financials'!G4),0)"
                else: prev_rev=f"{rev_cols[i-1]}{r}"; prev_base=base_cols[i-1]; nwc=f"IFERROR({base_c}$22/({base_c}$13-{prev_base}$13),0)"
                ws[f"{fcf_c}{r}"]=(f"={rev_c}{r}*({base_c}$14+$C{r})*(1-$C$8)+{rev_c}{r}*{base_c}$18-{rev_c}{r}*MAX(0.005,({base_c}$20+$D{r}))-({rev_c}{r}-{prev_rev})*({nwc})+{rev_c}{r}*$AL$4")
                ws[f"{fcf_c}{r}"].number_format=FMT_BN
    return adj


def _company_numbers(wb):
    d=wb["Company Data"]; price=_num(d["B8"].value,0); mc=_num(d["B10"].value,0); cash=_num(d["B12"].value,0); debt=_num(d["B13"].value,0); shares=_num(d["B9"].value) or ((mc/price) if mc and price else 1)
    return price,shares,debt-cash

def _model_value(wb,g=0,m=0,cap=0,w=0,t=0):
    s=wb["Three-Case Scenarios"]; h=wb["Historical Financials"]; _,shares,net_debt=_company_numbers(wb); rev=_num(h["G4"].value,1); prev=rev; pv=0; fcf=0; tax=_num(s["C8"].value,.21); wacc=max(.055,min(.16,_num(s["C6"].value,.09)+w)); tgr=max(0,min(.06,_num(s["C7"].value,.03)+t)); tgr=min(tgr,wacc-.005); adj=_num(s["AL4"].value,0)
    for i,c in enumerate(range(14,24)):
        growth=max(-.4,_num(s.cell(12,c).value,0)+g); margin=max(-.1,min(.6,_num(s.cell(14,c).value,.2)+m)); da=max(0,_num(s.cell(18,c).value,.04)); cp=max(.005,_num(s.cell(20,c).value,.08)+cap)
        rev*=1+growth; nwc=(rev-prev)*.01; fcf=rev*margin*(1-tax)+rev*da-rev*cp-nwc+rev*adj; pv+=fcf/((1+wacc)**(i+1)); prev=rev
    ev=pv+(fcf*(1+tgr)/(wacc-tgr))/((1+wacc)**10); return (ev-net_debt)/shares if shares else None


def ensure_upside_sensitivity(wb):
    if "Three-Case Scenarios" not in wb.sheetnames: return
    ws=wb["Three-Case Scenarios"]; _section(ws,"A68:K68","Balanced Opportunity / Upside Sensitivity"); _merge(ws,"A69:K69"); ws["A69"]="Opportunity sensitivities mirror the downside stress tests so risk and upside are evaluated symmetrically; they are not forecasts."; ws["A69"].font=Font(italic=True,color=GREY)
    heads=["Opportunity Case","Revenue Growth Shock","EBIT Margin Shock","Capex / Revenue Shock","WACC Shock","Terminal Growth Shock","Value / Share","Upside / (Downside)","Value vs Base","Interpretation"]
    for c,v in enumerate(heads,1): ws.cell(70,c,v)
    _header(ws,70,1,10)
    cases=[("Base Control",0,0,0,0,0,"Unstressed Base case"),("Growth +3ppt",.03,0,0,0,0,"Faster demand / share capture"),("Growth +5ppt",.05,0,0,0,0,"Strong upside revenue execution"),("EBIT Margin +300bps",0,.03,0,0,0,"Operating leverage / mix improvement"),("EBIT Margin +500bps",0,.05,0,0,0,"Major structural margin expansion"),("Capex -300bps",0,0,-.03,0,0,"Faster capital-intensity normalization"),("WACC -100bps",0,0,0,-.01,0,"Lower required return / risk premium"),("Terminal Growth +50bps",0,0,0,0,.005,"Higher durable long-run growth"),("Strong Execution",.03,.03,-.02,-.005,.005,"Combined balanced upside")]
    base=_model_value(wb); price=_company_numbers(wb)[0]
    for r,(name,g,m,cap,w,t,note) in enumerate(cases,71):
        value=_model_value(wb,g,m,cap,w,t); vals=[name,g,m,cap,w,t,value,(value/price-1) if value and price else None,(value/base-1) if value and base else None,note]
        for c,v in enumerate(vals,1): ws.cell(r,c,v)
        for c in range(2,7): ws.cell(r,c).number_format=FMT_PCT
        ws.cell(r,7).number_format=FMT_PRICE; ws.cell(r,8).number_format=FMT_PCT; ws.cell(r,9).number_format=FMT_PCT
    ws.conditional_formatting.add("G71:G79",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="percentile",mid_value=50,mid_color="FFEB84",end_type="max",end_color="63BE7B"))


def ensure_fcf_reconciliation(wb,ticker):
    if "Historical Financials" not in wb.sheetnames: return
    if "FCF & Capital Economics" in wb.sheetnames: wb.remove(wb["FCF & Capital Economics"])
    ws=wb.create_sheet("FCF & Capital Economics"); ws.sheet_view.showGridLines=False; h=wb["Historical Financials"]; s=wb["Three-Case Scenarios"] if "Three-Case Scenarios" in wb.sheetnames else None
    _merge(ws,"A1:L2"); ws["A1"]=f"{ticker} — FCF Reconciliation & Capital Economics"; ws["A1"].fill=_fill(NAVY); ws["A1"].font=Font(bold=True,color=WHITE,size=18)
    _merge(ws,"A3:L3"); ws["A3"]="Reconciles reported OCF-capex with a DCF-consistent cash-flow bridge and separates maintenance from growth capex. Growth-capex returns are analytical proxies, not issuer disclosures."; ws["A3"].font=Font(italic=True,color=GREY)
    _section(ws,"A5:H5","Historical Cash-Flow Reconciliation"); headers=["Metric"]+[h.cell(3,c).value for c in range(2,8)]+["Interpretation"]
    for c,v in enumerate(headers,1): ws.cell(6,c,v)
    _header(ws,6,1,8); tax=_num(s["C8"].value,.21) if s else .21
    labels=["Revenue","EBIT","NOPAT @ normalized tax","D&A","Stock-based compensation","Operating cash flow","Capex","Reported FCF = OCF - Capex","SBC-adjusted FCF = OCF - Capex - SBC","Cash-conversion residual ex-SBC","Residual / Revenue"]
    for r,label in enumerate(labels,7):
        ws.cell(r,1,label)
        for j,c in enumerate(range(2,8),2):
            rev=_num(h.cell(4,c).value); op=_num(h.cell(9,c).value); ocf=_num(h.cell(14,c).value); cap=_num(h.cell(15,c).value); da=_num(h.cell(18,c).value,0); sbc=_num(h.cell(21,c).value,0)
            if label=="Revenue": v=rev
            elif label=="EBIT": v=op
            elif label.startswith("NOPAT"): v=op*(1-tax) if op is not None else None
            elif label=="D&A": v=da
            elif label.startswith("Stock"): v=sbc
            elif label.startswith("Operating cash"): v=ocf
            elif label=="Capex": v=cap
            elif label.startswith("Reported FCF"): v=ocf-cap if ocf is not None and cap is not None else None
            elif label.startswith("SBC-adjusted"): v=ocf-cap-sbc if ocf is not None and cap is not None else None
            else:
                residual=ocf-op*(1-tax)-da-sbc if ocf is not None and op is not None else None; v=(residual/rev if label.startswith("Residual /") and residual is not None and rev else residual)
            ws.cell(r,j,v); ws.cell(r,j).number_format=FMT_PCT if label.startswith("Residual /") else FMT_BN
    ws["H14"]="Common reported FCF definition"; ws["H15"]="Conservative metric when DCF share count is held static"; ws["H16"]="Working capital, deferred tax and other items after excluding SBC"
    _section(ws,"A20:H20","Maintenance vs Growth Capex"); ws["A21"]="Maintenance capex factor × D&A"; ws["B21"]=1.0; ws["B21"].fill=_fill(GOLD); ws["B21"].font=Font(color=INPUT_BLUE); ws["B21"].number_format="0.0x"; ws["C21"]="Editable proxy; not a reported split."
    heads=["Year","Total Capex","D&A","Maintenance Capex Proxy","Growth Capex Proxy","Growth % of Capex","Owner Earnings Proxy"]
    for c,v in enumerate(heads,1): ws.cell(23,c,v)
    _header(ws,23,1,7); growth_capex=[]; op_income=[]
    for r,c in enumerate(range(2,8),24):
        year=h.cell(3,c).value; cap=_num(h.cell(15,c).value,0); da=_num(h.cell(18,c).value,0); ocf=_num(h.cell(14,c).value,0); sbc=_num(h.cell(21,c).value,0); op=_num(h.cell(9,c).value); maintenance=min(cap,da); growth=max(0,cap-maintenance); owner=ocf-sbc-maintenance; growth_capex.append(growth); op_income.append(op)
        vals=[year,cap,da,maintenance,growth,(growth/cap if cap else None),owner]
        for cc,v in enumerate(vals,1): ws.cell(r,cc,v)
        for cc in range(2,6): ws.cell(r,cc).number_format=FMT_BN
        ws.cell(r,6).number_format=FMT_PCT; ws.cell(r,7).number_format=FMT_BN
    ws.conditional_formatting.add("F24:F29",DataBarRule(start_type="num",start_value=0,end_type="num",end_value=1,color="5B9BD5"))
    _section(ws,"A32:H32","Growth-Capex Return Proxy"); _merge(ws,"A33:H33"); ws["A33"]="Incremental EBIT two years later ÷ growth-capex proxy. Rough cohort diagnostic only: acquisitions, mix, pricing and accounting also affect EBIT."; ws["A33"].alignment=Alignment(wrap_text=True)
    heads=["Capex Year","Growth Capex Proxy","EBIT +2Y","Prior-year EBIT +2Y","Incremental EBIT","Pre-tax Return Proxy","Return vs Base WACC"]
    for c,v in enumerate(heads,1): ws.cell(35,c,v)
    _header(ws,35,1,7); base_w=_num(s["C6"].value,.09) if s else .09
    for idx in range(4):
        r=36+idx; gc=growth_capex[idx]; e2=op_income[idx+2]; e1=op_income[idx+1]; inc=(e2-e1) if e2 is not None and e1 is not None else None; ret=inc/gc if inc is not None and gc else None; vals=[h.cell(3,2+idx).value,gc,e2,e1,inc,ret,(ret-base_w) if ret is not None else None]
        for c,v in enumerate(vals,1): ws.cell(r,c,v)
        for c in range(2,6): ws.cell(r,c).number_format=FMT_BN
        ws.cell(r,6).number_format=FMT_PCT; ws.cell(r,7).number_format=FMT_PCT
    ws.conditional_formatting.add("G36:G39",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="num",mid_value=0,mid_color="FFEB84",end_type="max",end_color="63BE7B"))
    _section(ws,"J5:L5","Forward Capital-Economics Test"); rows=[("Base WACC",base_w,FMT_PCT,"From scenario model"),("Assumed mature return on growth capex",.15,FMT_PCT,"Editable judgment"),("Economic spread",None,FMT_PCT,"Return - WACC"),("Latest growth-capex proxy",growth_capex[-1] if growth_capex else None,FMT_BN,"Latest capex less maintenance proxy"),("Potential annual economic profit",None,FMT_BN,"Growth capex × economic spread")]
    for r,(lab,val,fmt,note) in enumerate(rows,6): ws.cell(r,10,lab); ws.cell(r,11,val); ws.cell(r,11).number_format=fmt; ws.cell(r,12,note)
    ws["K7"].fill=_fill(GOLD); ws["K7"].font=Font(color=INPUT_BLUE); ws["K8"]="=K7-K6"; ws["K10"]="=K9*K8"; ws["K8"].number_format=FMT_PCT; ws["K10"].number_format=FMT_BN
    ws["J12"]="Interpretation"; ws["K12"]="Positive spread indicates value-creating reinvestment if the mature return is achieved. Do not add this economic profit directly to DCF cash flow."; ws["K12"].alignment=Alignment(wrap_text=True)
    ws["J15"]="DCF cash-conversion calibration"; ws["K15"]=cash_conversion_adjustment(wb); ws["K15"].number_format=FMT_PCT; ws["L15"]="3Y median residual ex-SBC, capped +/-5%."
    for c,w in {"A":31,"B":15,"C":15,"D":18,"E":18,"F":16,"G":18,"H":34,"J":31,"K":18,"L":36}.items(): ws.column_dimensions[c].width=w


def ensure_valuation_cross_checks(wb,ticker):
    if not {"Company Data","Three-Case Scenarios","Advanced Analytics","Peer Comps"}.issubset(wb.sheetnames): return
    if "Valuation Cross-Checks" in wb.sheetnames: wb.remove(wb["Valuation Cross-Checks"])
    ws=wb.create_sheet("Valuation Cross-Checks"); ws.sheet_view.showGridLines=False; _merge(ws,"A1:H2"); ws["A1"]=f"{ticker} — Independent Valuation Cross-Checks"; ws["A1"].fill=_fill(NAVY); ws["A1"].font=Font(bold=True,color=WHITE,size=18); _merge(ws,"A3:H3"); ws["A3"]="Separates absolute, probabilistic, relative and market-implied views so shared DCF assumptions are not mistaken for independent confirmations."; ws["A3"].font=Font(italic=True,color=GREY)
    price,_,_=_company_numbers(wb); d=wb["Company Data"]; a=wb["Advanced Analytics"]; p=wb["Peer Comps"]; fpe=_num(d["B15"].value); pes=[_num(p.cell(r,3).value) for r in range(5,10)]; pes=[x for x in pes if x and x>0]; med=statistics.median(pes) if pes else None; fwd_eps=price/fpe if price and fpe else None; rel=fwd_eps*med if fwd_eps and med else None
    rows=[("Current Market Price",price,FMT_PRICE,"Market reference","Market"),("Bear DCF",_model_value(wb,-.05,-.03,.03,.015,-.005),FMT_PRICE,"Conservative absolute valuation","DCF"),("Base DCF",_model_value(wb),FMT_PRICE,"Core absolute valuation","DCF"),("Strong-Execution DCF",_model_value(wb,.03,.03,-.02,-.005,.005),FMT_PRICE,"Balanced upside sensitivity","DCF"),("Regime Monte Carlo Median",_num(a["J35"].value),FMT_PRICE,"Probabilistic valuation across regimes","Probabilistic"),("Peer P/E Implied Price",rel,FMT_PRICE,"Forward EPS × peer median forward P/E","Relative"),("Reverse DCF Implied FCF CAGR",_num(a["B38"].value),FMT_PCT,"Growth required by current price","Expectations")]
    for c,v in enumerate(["Lens","Result","Unit","What it answers","Method family","Vs Market"],1): ws.cell(5,c,v)
    _header(ws,5,1,6)
    for r,(lab,val,fmt,note,fam) in enumerate(rows,6): ws.cell(r,1,lab); ws.cell(r,2,val); ws.cell(r,2).number_format=fmt; ws.cell(r,3,"%" if fmt==FMT_PCT else "$ / share"); ws.cell(r,4,note); ws.cell(r,5,fam); ws.cell(r,6,(val/price-1) if val is not None and price and fmt!=FMT_PCT else None); ws.cell(r,6).number_format=FMT_PCT
    ws.conditional_formatting.add("F6:F11",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="num",mid_value=0,mid_color="FFEB84",end_type="max",end_color="63BE7B"))
    _section(ws,"A15:H15","Interpretation"); notes=["DCF, stress tests and Monte Carlo still share forecast economics; they are related evidence, not three votes.","Relative valuation is a genuinely different lens but can still be wrong if peers are mispriced.","Reverse DCF is an expectations test: compare market-required growth with your own forecast.","Use FCF & Capital Economics to judge whether high current capex is destructive or value-creating reinvestment."]
    for r,n in enumerate(notes,16): _merge(ws,f"A{r}:H{r}"); ws.cell(r,1,n); ws.cell(r,1).alignment=Alignment(wrap_text=True)
    for c,w in {"A":30,"B":18,"C":14,"D":46,"E":18,"F":16}.items(): ws.column_dimensions[c].width=w


def ensure_research_workbench(wb,ticker):
    company=wb["Company Data"]["B5"].value if "Company Data" in wb.sheetnames else ticker
    if "Research Notes" in wb.sheetnames: wb.remove(wb["Research Notes"])
    ws=wb.create_sheet("Research Notes"); ws.sheet_view.showGridLines=False; _merge(ws,"A1:H2"); ws["A1"]=f"{ticker} — Manual Investment Thesis & Research Notes"; ws["A1"].fill=_fill(NAVY); ws["A1"].font=Font(bold=True,color=WHITE,size=18); _merge(ws,"A3:H3"); ws["A3"]=f"Manual judgment layer for {company}: write what you believe, why the market may disagree, and what evidence would prove you wrong."; ws["A3"].font=Font(italic=True,color=GREY)
    sections=[(5,"Investment Thesis",["One-sentence thesis","Variant perception / what I believe differently","What the market appears to believe","Why mispricing may exist","Time horizon","Expected-return logic / required margin of safety"]),(14,"Business Quality & Moat",["Business model / how money is made","Durability of moat","Pricing power","Customer switching costs / network effects","Unit economics / incremental margins","Cyclicality / recurring revenue","Key KPI I would track every quarter"]),(24,"Market / Sector Analysis",["TAM / market growth","Industry structure / concentration","Competitive intensity","Market-share trend","Technology disruption","Regulation / policy","Customer power","Supplier / input risk","Where we are in the industry cycle"]),(36,"Management & Capital Allocation",["Management quality","Capital-allocation record","Buybacks / dilution / SBC","M&A discipline","Incentives / ownership","What management says vs what it does"]),(46,"Risks, Thesis Breakers & Catalysts",["Top 3 fundamental risks","Thesis breaker #1","Thesis breaker #2","What would make me sell","Bull catalysts","Bear catalysts","Leading indicators of deterioration"]),(57,"Behavioral / Feelings Check",["Current conviction (1–5)","Am I influenced by recent price action?","FOMO / fear level (1–5)","What evidence am I emotionally discounting?","What would make me wait rather than act?","What would change my mind?"]),(67,"Decision Summary",["Current stance: Buy / Watch / Avoid / Hold","Preferred entry valuation / price","Required margin of safety","Position-size thought","Next review trigger / date","Final decision rationale"])]
    for start,title,items in sections:
        _section(ws,f"A{start}:H{start}",title); r=start+1
        for item in items:
            ws.cell(r,1,item); _merge(ws,f"B{r}:H{r}"); ws.cell(r,2).fill=_fill(GOLD); ws.cell(r,2).font=Font(color=INPUT_BLUE); ws.cell(r,2).alignment=Alignment(wrap_text=True,vertical="top"); ws.row_dimensions[r].height=30; r+=1
    ws.column_dimensions["A"].width=34
    for c in range(2,9): ws.column_dimensions[get_column_letter(c)].width=15
    if "Research Checklist" in wb.sheetnames: wb.remove(wb["Research Checklist"])
    ck=wb.create_sheet("Research Checklist"); ck.sheet_view.showGridLines=False; _merge(ck,"A1:G2"); ck["A1"]=f"{ticker} — Investment Research Checklist"; ck["A1"].fill=_fill(NAVY); ck["A1"].font=Font(bold=True,color=WHITE,size=18); _merge(ck,"A3:G3"); ck["A3"]="Use status, evidence and confidence to separate completed diligence from assumptions or feelings."; ck["A3"].font=Font(italic=True,color=GREY)
    heads=["Category","Checklist Item","Status","Evidence / Conclusion","Source / Link","Confidence (1–5)","Last Reviewed"]
    for c,v in enumerate(heads,1): ck.cell(5,c,v)
    _header(ck,5,1,7)
    items=[("Data","Historical financials verified"),("Data","Current market data / share count verified"),("Data","Segment data complete or manual gaps identified"),("Business","Business model understood"),("Business","Moat and competitive advantage assessed"),("Business","Key unit economics / KPIs identified"),("Sector","TAM and market growth assessed"),("Sector","Competition / market share assessed"),("Sector","Industry-cycle position assessed"),("Sector","Regulatory / technology disruption reviewed"),("Management","Management quality assessed"),("Management","Capital allocation / M&A / buybacks reviewed"),("Management","SBC and dilution reviewed"),("Financial","Revenue / margin drivers modeled"),("Financial","Reported FCF reconciled to normalized FCFF"),("Financial","Maintenance vs growth capex reviewed"),("Financial","Reinvestment return / ROIC vs WACC reviewed"),("Financial","Balance sheet and liquidity reviewed"),("Valuation","Bear/Base/Bull assumptions reviewed"),("Valuation","Downside stress tests reviewed"),("Valuation","Upside opportunity sensitivities reviewed"),("Valuation","DCF sensitivity reviewed"),("Valuation","Relative valuation reviewed"),("Valuation","Reverse DCF / market expectations reviewed"),("Valuation","Independent valuation cross-checks reconciled"),("Thesis","Variant perception written"),("Thesis","Thesis breakers defined"),("Thesis","Catalysts and leading indicators defined"),("Behavior","FOMO / price-action bias checked"),("Behavior","Disconfirming evidence actively sought"),("Decision","Expected return compensates for risk"),("Decision","Entry price / margin of safety defined")]
    for r,(cat,item) in enumerate(items,6): ck.cell(r,1,cat); ck.cell(r,2,item); ck.cell(r,3,"Not Started")
    dv=DataValidation(type="list",formula1='"Not Started,In Progress,Done,N/A"'); ck.add_data_validation(dv); dv.add(f"C6:C{5+len(items)}"); dv2=DataValidation(type="list",formula1='"1,2,3,4,5"'); ck.add_data_validation(dv2); dv2.add(f"F6:F{5+len(items)}")
    for r in range(6,6+len(items)): ck.cell(r,3).fill=_fill(GOLD); ck.cell(r,3).font=Font(color=INPUT_BLUE); ck.cell(r,4).fill=_fill(LIGHT); ck.cell(r,5).fill=_fill(LIGHT); ck.cell(r,6).fill=_fill(GOLD); ck.cell(r,7).fill=_fill(LIGHT); ck.cell(r,4).alignment=Alignment(wrap_text=True,vertical="top")
    ck.column_dimensions["A"].width=15; ck.column_dimensions["B"].width=46; ck.column_dimensions["C"].width=16; ck.column_dimensions["D"].width=52; ck.column_dimensions["E"].width=34; ck.column_dimensions["F"].width=16; ck.column_dimensions["G"].width=15; ck.freeze_panes="A6"
    if "Research Journal" in wb.sheetnames: wb.remove(wb["Research Journal"])
    jr=wb.create_sheet("Research Journal"); jr.sheet_view.showGridLines=False; _merge(jr,"A1:J2"); jr["A1"]=f"{ticker} — Research Journal / Decision Log"; jr["A1"].fill=_fill(NAVY); jr["A1"].font=Font(bold=True,color=WHITE,size=18); _merge(jr,"A3:J3"); jr["A3"]="Record new evidence before checking the share-price reaction when possible. This reduces hindsight and narrative drift."; jr["A3"].font=Font(italic=True,color=GREY)
    heads=["Date","Source / Event","Observation / New Fact","Impact","Thesis Change?","Action / Follow-up","Confidence (1–5)","Share Price","What I Felt","Notes / What Would Falsify It"]
    for c,v in enumerate(heads,1): jr.cell(5,c,v)
    _header(jr,5,1,10)
    for r in range(6,56):
        for c in range(1,11): jr.cell(r,c).fill=_fill(LIGHT if c not in (4,5,7) else GOLD)
        jr.cell(r,3).alignment=Alignment(wrap_text=True,vertical="top"); jr.cell(r,10).alignment=Alignment(wrap_text=True,vertical="top")
    dv3=DataValidation(type="list",formula1='"Bullish,Neutral,Bearish,Mixed"'); jr.add_data_validation(dv3); dv3.add("D6:D55"); dv4=DataValidation(type="list",formula1='"No,Minor,Major"'); jr.add_data_validation(dv4); dv4.add("E6:E55"); dv5=DataValidation(type="list",formula1='"1,2,3,4,5"'); jr.add_data_validation(dv5); dv5.add("G6:G55")
    for c,w in {"A":13,"B":28,"C":48,"D":12,"E":13,"F":30,"G":15,"H":14,"I":24,"J":48}.items(): jr.column_dimensions[c].width=w
    jr.freeze_panes="A6"


def ensure_data_quality(wb,ticker):
    if "Data Quality" in wb.sheetnames: wb.remove(wb["Data Quality"])
    ws=wb.create_sheet("Data Quality"); ws.sheet_view.showGridLines=False; _merge(ws,"A1:F2"); ws["A1"]=f"{ticker} — Data Quality & Cross-Company Checks"; ws["A1"].fill=_fill(NAVY); ws["A1"].font=Font(bold=True,color=WHITE,size=18); _merge(ws,"A3:F3"); ws["A3"]="Automatic sanity checks help detect SEC period-selection errors, stale template data and incomplete company-specific modules."; ws["A3"].font=Font(italic=True,color=GREY)
    h=wb["Historical Financials"]; years=[h.cell(3,c).value for c in range(2,8)]; rev=[_num(h.cell(4,c).value) for c in range(2,8)]; cost=[_num(h.cell(6,c).value) for c in range(2,8)]; ascending=all(isinstance(years[i],(int,float)) and isinstance(years[i+1],(int,float)) and years[i]<years[i+1] for i in range(5)); cost_ok=all(c is None or r is None or c<=r*1.5 for c,r in zip(cost,rev)); seg=wb["Segment Analysis"] if "Segment Analysis" in wb.sheetnames else None; manual=bool(seg and any(seg.cell(r,1).value=="Manual Segment Input" for r in range(1,min(10,seg.max_row)+1))); google=ticker in {"GOOGL","GOOG"}; ai_tabs=all(x in wb.sheetnames for x in ("AI Analysis","AI Valuation"))
    checks=[("Ticker updated","PASS" if wb["Company Data"]["B4"].value==ticker else "FAIL",wb["Company Data"]["B4"].value,"Company Data should match requested ticker."),("Six annual periods","PASS" if len([x for x in years if x])==6 else "REVIEW",str(years),"Uses actual period-end year after SEC fix."),("Years unique and ascending","PASS" if ascending else "FAIL",str(years),"Detects mixed quarterly/annual period mapping."),("Revenue positive","PASS" if all(r is None or r>0 for r in rev) else "FAIL",str(rev),"Negative/zero revenue usually indicates extraction failure."),("Cost/revenue sanity","PASS" if cost_ok else "REVIEW","","Flags obviously inconsistent period selection."),("Segment analysis","MANUAL" if manual else "PASS","Manual input required" if manual else "Automated / enriched","Manual is preferable to invented segment revenue."),("Company-specific AI tabs","PASS" if (google==ai_tabs or (not google and not ai_tabs)) else "REVIEW",str(ai_tabs),"Alphabet-only sheets should not leak into unrelated tickers.")]
    for c,v in enumerate(["Check","Status","Observed","Why it matters"],1): ws.cell(5,c,v)
    _header(ws,5,1,4)
    for r,row in enumerate(checks,6):
        for c,v in enumerate(row,1): ws.cell(r,c,v)
        ws.cell(r,2).fill=_fill(PALE_GREEN if row[1]=="PASS" else GOLD if row[1] in {"REVIEW","MANUAL"} else PALE_RED)
    for c,w in {"A":30,"B":14,"C":45,"D":60}.items(): ws.column_dimensions[c].width=w


def improve_analysis_charts(wb,ticker):
    if "Analysis Charts" not in wb.sheetnames or "Historical Financials" not in wb.sheetnames: return
    ws=wb["Analysis Charts"]; h=wb["Historical Financials"]; years=[h.cell(3,c).value for c in range(2,8)]; rev=[_num(h.cell(4,c).value) for c in range(2,8)]; op=[_num(h.cell(9,c).value) for c in range(2,8)]; ocf=[_num(h.cell(14,c).value) for c in range(2,8)]; cap=[_num(h.cell(15,c).value) for c in range(2,8)]; fcf=[o-x if o is not None and x is not None else None for o,x in zip(ocf,cap)]
    def idx(vals):
        base=next((v for v in vals if isinstance(v,(int,float)) and v!=0),None); return [(v/base*100 if isinstance(v,(int,float)) and base else None) for v in vals]
    for c,v in enumerate(["Year","Revenue Index","Operating Income Index","FCF Index"],53): ws.cell(2,c,v)
    ri,oi,fi=idx(rev),idx(op),idx(fcf)
    for r,i in enumerate(range(6),3): ws.cell(r,53,years[i]); ws.cell(r,54,ri[i]); ws.cell(r,55,oi[i]); ws.cell(r,56,fi[i])
    for c in range(53,57): ws.column_dimensions[get_column_letter(c)].hidden=True
    if len(getattr(ws,"_charts",[]))>=4:
        try: ws._charts.pop(3)
        except Exception: pass
    ch=LineChart(); ch.style=10; ch.title="Indexed Financial Performance — First Year = 100"; ch.height=9; ch.width=13.5; ch.legend.position="b"; ch.y_axis.title="Index (first year = 100)"; ch.x_axis.title="Fiscal year"; ch.add_data(Reference(ws,min_col=54,max_col=56,min_row=2,max_row=8),titles_from_data=True); ch.set_categories(Reference(ws,min_col=53,min_row=3,max_row=8)); ch.visible_cells_only=False; ws.add_chart(ch,"I29")
    ws["I48"]="Indexed view separates operating growth from scale: 100 = first displayed year. Revenue, operating income and FCF are now comparable without revenue visually dwarfing FCF."; ws["I49"]="Reference: Historical Financials. See FCF & Capital Economics for normalization and capex decomposition."
    if ticker not in {"GOOGL","GOOG"}:
        ws["A70"]="Units: latest disclosed segment/business revenue ($bn)"; ws["F70"]="Source: Segment Analysis / issuer filing"; ws["I70"]="Margin = operating income ÷ segment revenue"; ws["M70"]="Source: Segment Analysis / issuer filing"; ws["I71"]="Use issuer-disclosed segments only; do not infer undisclosed sub-business revenue."; ws["A112"]="External segment source: see Segment Analysis source/notes. Company-specific segment data may require manual completion when SEC XBRL is not standardized."; ws["A113"]=""


def ensure_model_quality(wb,ticker):
    ticker=(ticker or "").upper()
    if ticker not in {"GOOGL","GOOG"}:
        for name in ("AI Analysis","AI Valuation"):
            if name in wb.sheetnames: wb.remove(wb[name])
    ensure_upside_sensitivity(wb); ensure_fcf_reconciliation(wb,ticker); ensure_valuation_cross_checks(wb,ticker); ensure_research_workbench(wb,ticker); ensure_data_quality(wb,ticker); improve_analysis_charts(wb,ticker)
    for sheet in wb.worksheets:
        for ch in getattr(sheet,"_charts",[]):
            try: ch.visible_cells_only=False; ch.display_blanks="gap"
            except Exception: pass
