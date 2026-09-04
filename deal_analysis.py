from __future__ import annotations

"""Source-backed recent material deals and transactions research.

The module is conservative by design: it scans ticker-specific public news / press releases and,
for SEC registrants, a small set of recent EDGAR filings. Transaction facts remain separate from
rule-based analyst lenses. Missing deal economics are left blank rather than inferred.
"""

from datetime import date, datetime, timezone, timedelta
from functools import lru_cache
import os
import re

import requests
try:
    import yfinance as yf
except Exception:  # pragma: no cover
    yf = None
try:
    from bs4 import BeautifulSoup
except Exception:  # pragma: no cover
    BeautifulSoup = None

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; GREY="666666"
PALE_GREEN="E2F0D9"; PALE_RED="FCE4D6"; PALE_YELLOW="FFF2CC"; PALE_BLUE="DDEBF7"; LINK_GREEN="008000"
THIN=Side(style="thin",color="D9E1F2")
DEFAULT_HORIZON_DAYS=730; DEFAULT_MAX_DEALS=12
SEC_FORMS={"8-K","10-Q","10-K","6-K","20-F","40-F"}
SEC_USER_AGENT=os.getenv("SEC_USER_AGENT","Antzaz Equity Research educational project https://antzaz.github.io/Antzaz-investment-research-showcase/")

DEAL_PATTERNS=[
    ("Acquisition / M&A",("acquire","acquisition","merger","takeover","buyout","purchase of")),
    ("Divestiture / Asset Sale",("divest","divestiture","sale of","sell its","sells its","spin-off","spinoff")),
    ("Strategic Investment",("investment in","invest in","invests in","stake in","equity stake","warrant","convertible")),
    ("Joint Venture",("joint venture","jv with","forms venture","venture with")),
    ("Major Commercial Contract",("contract","supply agreement","power purchase","offtake","licensing agreement","multi-year agreement","long-term agreement")),
    ("Strategic Partnership",("strategic partnership","partnership with","partners with","collaboration with","teams up with")),
    ("Financing / Funding",("funding round","financing","capital commitment","funding commitment","raises $","raises €","raises £")),
]
BASE_SCORE={"Acquisition / M&A":36,"Divestiture / Asset Sale":32,"Strategic Investment":27,"Joint Venture":24,"Major Commercial Contract":22,"Strategic Partnership":18,"Financing / Funding":17}
CURRENCIES={"$":"USD","US$":"USD","USD":"USD","€":"EUR","EUR":"EUR","£":"GBP","GBP":"GBP","¥":"JPY","JPY":"JPY","C$":"CAD","CAD":"CAD","A$":"AUD","AUD":"AUD"}
HIGH_PUBLISHERS=("reuters","bloomberg","associated press","financial times","wall street journal","wsj")
STOP={"the","and","for","with","from","deal","agreement","transaction","acquisition","acquire","acquires","merger","investment","partnership","strategic","announces","announced","completed","signs","signed","major","multi","year"}


def _fill(c): return PatternFill("solid",fgColor=c)
def _clean(v): return re.sub(r"\s+"," ",str(v or "")).strip()
def _clip(v,n=420):
    s=_clean(v)
    if len(s)<=n: return s
    cut=s[:n-1].rsplit(" ",1)[0]
    return (cut or s[:n-1])+"…"
def _nested(obj,*path):
    cur=obj
    for key in path:
        if not isinstance(cur,dict): return None
        cur=cur.get(key)
    return cur

def _date(v):
    if v in (None,""): return None
    if isinstance(v,datetime): return v.date().isoformat()
    if isinstance(v,date): return v.isoformat()
    if isinstance(v,(int,float)):
        try: return datetime.fromtimestamp(v,tz=timezone.utc).date().isoformat()
        except Exception: return None
    s=str(v).strip()
    try: return datetime.fromisoformat(s.replace("Z","+00:00")).date().isoformat()
    except Exception:
        m=re.search(r"20\d{2}-\d{2}-\d{2}",s); return m.group(0) if m else None


def _normalize_news(item,tab):
    if not isinstance(item,dict): return None
    c=item.get("content") if isinstance(item.get("content"),dict) else item
    title=_clean(c.get("title") or item.get("title"))
    if not title: return None
    summary=_clean(c.get("summary") or c.get("description") or item.get("summary") or item.get("description"))
    publisher=_clean(_nested(c,"provider","displayName") or item.get("publisher") or item.get("provider"))
    url=next((str(x) for x in (_nested(c,"canonicalUrl","url"),_nested(c,"clickThroughUrl","url"),c.get("link") if isinstance(c,dict) else None,item.get("link"),item.get("url")) if x),None)
    return {"title":title,"summary":summary,"date":_date(c.get("pubDate") or c.get("displayTime") or item.get("providerPublishTime") or item.get("pubDate")),"publisher":publisher or ("Company press release" if tab=="press releases" else "Yahoo Finance feed"),"url":url,"source_kind":"press_release" if tab=="press releases" else "news"}


def _yahoo_items(ticker,count=100):
    if yf is None: return []
    try: obj=yf.Ticker(ticker)
    except Exception: return []
    out=[]; seen=set()
    for tab in ("press releases","all"):
        try: raw=obj.get_news(count=count,tab=tab) or []
        except Exception:
            try: raw=(getattr(obj,"news",[]) or []) if tab=="all" else []
            except Exception: raw=[]
        for item in raw:
            x=_normalize_news(item,tab)
            if not x: continue
            key=(x.get("url") or "",x["title"].lower())
            if key in seen: continue
            seen.add(key); out.append(x)
    return sorted(out,key=lambda x:x.get("date") or "",reverse=True)


@lru_cache(maxsize=1)
def _sec_map():
    try:
        r=requests.get("https://www.sec.gov/files/company_tickers.json",headers={"User-Agent":SEC_USER_AGENT},timeout=15)
        r.raise_for_status(); data=r.json()
    except Exception: return {}
    out={}
    for row in (data or {}).values():
        try: out[str(row.get("ticker") or "").upper()]=int(row.get("cik_str"))
        except Exception: pass
    return out


def _sec_filings(ticker,horizon_days=DEFAULT_HORIZON_DAYS,max_filings=6):
    cik=_sec_map().get(str(ticker or "").upper().split(".")[0])
    if not cik: return []
    try:
        r=requests.get(f"https://data.sec.gov/submissions/CIK{cik:010d}.json",headers={"User-Agent":SEC_USER_AGENT},timeout=15)
        r.raise_for_status(); recent=(r.json().get("filings") or {}).get("recent") or {}
    except Exception: return []
    cutoff=date.today()-timedelta(days=horizon_days); events=[]; periodic=[]
    for form,fd,acc,doc in zip(recent.get("form") or [],recent.get("filingDate") or [],recent.get("accessionNumber") or [],recent.get("primaryDocument") or []):
        if form not in SEC_FORMS: continue
        try: dt=date.fromisoformat(str(fd))
        except Exception: continue
        if dt<cutoff: continue
        url=f"https://www.sec.gov/Archives/edgar/data/{cik}/{str(acc).replace('-','')}/{doc}"
        row={"form":form,"date":dt.isoformat(),"url":url}
        if form in {"8-K","6-K"} and len(events)<4: events.append(row)
        elif form not in {"8-K","6-K"} and len(periodic)<2: periodic.append(row)
        if len(events)>=4 and len(periodic)>=2: break
    rows=events+periodic; rows.sort(key=lambda x:x["date"],reverse=True); return rows[:max_filings]


def _html_text(html):
    if BeautifulSoup is not None:
        try:
            soup=BeautifulSoup(html,"lxml")
            for tag in soup(["script","style","noscript"]): tag.decompose()
            return _clean(soup.get_text(" ",strip=True))
        except Exception: pass
    return _clean(re.sub(r"<[^>]+>"," ",html))


def _transaction_windows(text,limit=6):
    pats=(r"completed (?:our )?acquisition",r"definitive agreement",r"pending acquisition",r"acquisitions? and divestitures?",r"entered into .*?agreement",r"joint venture",r"capital commitment",r"funding round",r"strategic investment",r"divestiture")
    hits=[]; low=text.lower()
    for pat in pats:
        for m in re.finditer(pat,low,re.I):
            snip=_clean(text[max(0,m.start()-220):min(len(text),m.end()+700)])
            if len(snip)<80: continue
            if not (re.search(r"(?:\$|€|£|¥|usd|eur|gbp)\s*\d",snip,re.I) or re.search(r"\d+(?:\.\d+)?\s*(?:billion|million|bn|mn)\b",snip,re.I) or re.search(r"(?:completed|closed|subject to|regulatory approval|definitive agreement)",snip,re.I)): continue
            key=re.sub(r"[^a-z0-9]+"," ",snip.lower())[:110]
            if any(key in re.sub(r"[^a-z0-9]+"," ",x.lower()) for x in hits): continue
            hits.append(_clip(snip,780))
            if len(hits)>=limit: return hits
    return hits


def _sec_items(ticker,horizon_days=DEFAULT_HORIZON_DAYS):
    out=[]
    for filing in _sec_filings(ticker,horizon_days):
        try:
            r=requests.get(filing["url"],headers={"User-Agent":SEC_USER_AGENT},timeout=18)
            r.raise_for_status(); text=_html_text(r.text)
        except Exception: continue
        for i,snip in enumerate(_transaction_windows(text),1):
            out.append({"title":f"SEC {filing['form']} transaction disclosure {i}","summary":snip,"date":filing["date"],"publisher":"SEC / EDGAR","url":filing["url"],"source_kind":"sec_filing"})
    return out


def _deal_type(text):
    low=_clean(text).lower()
    for kind,phrases in DEAL_PATTERNS:
        if any(p in low for p in phrases): return kind
    return None

def _status(text):
    low=_clean(text).lower()
    if any(x in low for x in ("terminated","scrapped","abandoned","called off")): return "Terminated"
    if any(x in low for x in ("completed","completes","closed","closes","completion of")): return "Completed / Closed"
    if any(x in low for x in ("approved","approval","cleared by","regulatory clearance")): return "Approved / Clearing"
    if any(x in low for x in ("definitive agreement","agreed to","will acquire","to acquire","pending","expected to close")): return "Announced / Pending"
    if any(x in low for x in ("renews","renewal","extends","expanded","expands")): return "Renewed / Expanded"
    if any(x in low for x in ("signs","signed","enters into","entered into","partners","partnership")): return "Signed / Active"
    return "Reported / Review"

_AMOUNT_RE=re.compile(r"(?:(?P<currency>US\$|C\$|A\$|USD|EUR|GBP|JPY|\$|€|£|¥)\s*)?(?P<number>\d{1,4}(?:\.\d+)?)\s*(?P<unit>billion|billions|bn|million|millions|mn)\b",re.I)
_VALUE_CONTEXT=re.compile(r"purchase price|purchase consideration|transaction value|deal value|acquisition|acquire|acquired|investment|invested|commitment|funding|consideration|proceeds|cash and|cash plus|debt assumed|assumption of debt|enterprise value|equity value|valued at|for approximately|for about",re.I)

def _amount(text):
    s=_clean(text); vals=[]
    for m in _AMOUNT_RE.finditer(s):
        try: n=float(m.group("number"))
        except Exception: continue
        unit=m.group("unit").lower(); rawcur=m.group("currency") or ""; cur=CURRENCIES.get(rawcur.upper() if rawcur.isalpha() else rawcur)
        vals.append({"raw":m.group(0).strip(),"value":n*(1e9 if unit.startswith("b") else 1e6),"currency":cur,"start":m.start(),"end":m.end()})
    if not vals: return None
    contexts=[m.span() for m in _VALUE_CONTEXT.finditer(s)]
    def rank(v):
        mid=(v["start"]+v["end"])/2
        dist=min((0 if a<=mid<=b else min(abs(mid-a),abs(mid-b)) for a,b in contexts),default=9999)
        return (1 if dist<=160 else 0,1 if v.get("currency") else 0,-dist,v["value"])
    best=max(vals,key=rank); return {k:v for k,v in best.items() if k not in {"start","end"}}


def _terms(text):
    low=_clean(text).lower(); out=[]
    for phrase,label in (("all-cash","All cash"),("cash consideration","Cash"),("all-stock","All stock"),("stock-for-stock","Stock-for-stock"),("assumption of debt","Debt assumed"),("convertible","Convertible security"),("warrant","Equity warrant / option"),("note receivable","Note receivable"),("minority stake","Minority stake"),("equity interest","Equity interest"),("multi-year","Multi-year"),("long-term","Long-term"),("power purchase","Power purchase"),("offtake","Offtake")):
        if phrase in low and label not in out: out.append(label)
    pct=re.findall(r"\b\d{1,3}(?:\.\d+)?%",text)
    if pct: out.append("Ownership / rate: "+", ".join(pct[:2]))
    cap=re.findall(r"\b\d+(?:\.\d+)?\s*(?:MW|GW|megawatts?|gigawatts?)\b",text,re.I)
    if cap: out.append("Capacity: "+", ".join(cap[:2]))
    return "; ".join(out) if out else "Not disclosed in retrieved source text"

def _source_quality(item):
    if item.get("source_kind")=="sec_filing": return "High — primary filing"
    if item.get("source_kind")=="press_release": return "High — company / press-release feed"
    pub=_clean(item.get("publisher")).lower()
    if any(x in pub for x in HIGH_PUBLISHERS): return "High — established news source"
    if pub and pub!="yahoo finance feed": return "Medium — publisher summary"
    return "Review — feed metadata only"

def _counterparty(title,ticker,info):
    aliases={str(ticker or "").upper().split(".")[0].lower()}
    for k in ("longName","shortName"):
        v=_clean((info or {}).get(k)).lower(); v=re.sub(r"\b(inc|corp|corporation|plc|ltd|limited|se|sa|ag|nv|company|co)\b","",v); v=re.sub(r"[^a-z0-9]+"," ",v).strip()
        if v: aliases|={v,v.split()[0]}
    pats=(r"(?:acquire|acquires|acquisition of|invest in|invests in|stake in|sale of)\s+([A-Z][A-Za-z0-9&'.\-]*(?:\s+[A-Z][A-Za-z0-9&'.\-]*){0,5})",r"(?:deal|agreement|partnership|collaboration|venture)\s+with\s+([A-Z][A-Za-z0-9&'.\-]*(?:\s+[A-Z][A-Za-z0-9&'.\-]*){0,5})")
    for pat in pats:
        m=re.search(pat,_clean(title))
        if not m: continue
        cand=re.split(r"\s+(?:for|to|in|on|as|after|amid)\s+",m.group(1),maxsplit=1)[0].strip(" ,.-")
        norm=re.sub(r"[^a-z0-9]+"," ",cand.lower()).strip()
        if cand and not any(norm==a or norm.startswith(a+" ") for a in aliases): return cand
    return "Review source / headline"

def _area(text):
    low=_clean(text).lower()
    specs=[
        (("ai","artificial intelligence","machine learning"),"AI / Machine Learning"),
        (("cloud","cybersecurity","security platform"),"Cloud / Cybersecurity"),
        (("data center","datacenter","compute","gpu","chip","semiconductor"),"Data Center / Compute"),
        (("power","energy","electricity","renewable","geothermal","solar","wind","battery"),"Power / Energy Infrastructure"),
        (("payment","payments","fintech"),"Payments / Fintech"),
        (("health","drug","pharma","biotech","clinical"),"Healthcare / Life Sciences"),
        (("advertising","ads","media","content","streaming"),"Advertising / Media"),
        (("logistics","warehouse","delivery","supply chain"),"Logistics / Supply Chain"),
        (("manufacturing","factory","industrial","automation"),"Industrial / Manufacturing"),
    ]
    labels=[label for phrases,label in specs if any(x in low for x in phrases)]
    return "; ".join(labels[:2]) if labels else "Corporate / Other"

def _materiality(amount,info,text):
    ratio=None; mc=(info or {}).get("marketCap"); quote=str((info or {}).get("currency") or "").upper()
    try:
        if amount and amount.get("currency")==quote and mc: ratio=amount["value"]/float(mc)
    except Exception: ratio=None
    if ratio is not None:
        if ratio>=.10: return "Transformational",ratio
        if ratio>=.03: return "Material",ratio
        if ratio>=.01: return "Notable",ratio
    if amount and amount.get("currency") in {"USD","EUR","GBP"}:
        if amount["value"]>=10e9: return "Major",ratio
        if amount["value"]>=1e9: return "Material",ratio
        if amount["value"]>=250e6: return "Notable",ratio
    low=_clean(text).lower()
    if any(x in low for x in ("largest-ever","largest ever","multi-year","long-term","definitive agreement")): return "Strategically material / value not normalized",ratio
    return "Review",ratio

def _score(kind,amount,materiality,quality,text,published):
    s=float(BASE_SCORE.get(kind,10)); value=(amount or {}).get("value") or 0
    s+=35 if value>=10e9 else 25 if value>=1e9 else 15 if value>=250e6 else 8 if value>=50e6 else 0
    s+={"Transformational":30,"Major":24,"Material":18,"Notable":10,"Strategically material / value not normalized":10}.get(materiality,0)
    if quality.startswith("High"): s+=8
    low=_clean(text).lower(); s+=8 if "largest" in low else 0; s+=5 if ("multi-year" in low or "long-term" in low) else 0
    if published:
        try:
            age=(date.today()-date.fromisoformat(published)).days; s+=8 if age<=90 else 4 if age<=365 else 0
        except Exception: pass
    return s

def _regulatory(text,status):
    low=_clean(text).lower(); flags=[]
    if any(x in low for x in ("antitrust","competition authority","regulator","regulatory approval","regulatory clearance")): flags.append("Competition / regulatory review")
    if any(x in low for x in ("shareholder approval","shareholders approved","vote")): flags.append("Shareholder approval")
    if "subject to" in low or "expected to close" in low: flags.append("Closing conditions remain")
    if status=="Announced / Pending" and not flags: flags.append("Verify customary closing / regulatory conditions")
    return "; ".join(flags) if flags else "No specific regulatory condition found in retrieved text"

def _rationale(kind,area):
    return {"Acquisition / M&A":f"Assess whether the acquisition strengthens {area.lower()} capabilities, expands TAM, accelerates capacity or removes a build-vs-buy bottleneck.","Divestiture / Asset Sale":"Assess strategic focus, proceeds, retained economics, stranded costs and capital redeployment.","Strategic Investment":f"Assess access to {area.lower()} capabilities, strategic optionality, supplier/customer alignment and governance exposure.","Joint Venture":f"Assess capital sharing, access to {area.lower()} capabilities, governance rights and economics split.","Major Commercial Contract":"Assess backlog/revenue visibility, duration, pricing, capacity obligations, concentration and required capex.","Strategic Partnership":f"Assess whether the partnership improves distribution, technology access, differentiation or adoption in {area.lower()}.","Financing / Funding":"Assess ownership dilution, capital commitment, funding runway, strategic control and future funding obligations."}.get(kind,"Assess strategic fit and expected return on capital.")
def _financial(kind,amount):
    note=f" Disclosed size: {amount['raw']}." if amount else " Deal value was not verified in retrieved text."
    return {"Acquisition / M&A":"Potential effects: cash/debt usage, goodwill/intangibles, acquired revenue/profit, integration costs, amortization, synergies and ROIC versus WACC.","Divestiture / Asset Sale":"Potential effects: proceeds, gain/loss, revenue/EBITDA deconsolidation, retained economics and capital redeployment.","Strategic Investment":"Potential effects: cash deployment, equity-method/fair-value accounting, strategic benefits and impairment risk.","Joint Venture":"Potential effects: contributed capital/assets, equity-method earnings, guarantees, shared capex and execution economics.","Major Commercial Contract":"Potential effects: backlog/revenue visibility, margin mix, capex/working-capital needs, guarantees and concentration.","Strategic Partnership":"Potential effects depend on commercialization: adoption, revenue share, distribution economics, development costs and exclusivity.","Financing / Funding":"Potential effects: liquidity, ownership/control, dilution or fair-value movements; verify parent versus subsidiary funding."}.get(kind,"Review cash flow, earnings and valuation implications.")+note
def _risks(kind,status,text):
    base={"Acquisition / M&A":["integration / execution","overpayment / ROIC","synergy delivery"],"Divestiture / Asset Sale":["lost earnings / optionality","separation execution"],"Strategic Investment":["valuation / impairment","counterparty dependency","governance rights"],"Joint Venture":["governance alignment","commitments","partner execution"],"Major Commercial Contract":["delivery / capacity","concentration","pricing / margin"],"Strategic Partnership":["commercialization","dependency / exclusivity","execution"],"Financing / Funding":["capital at risk","dilution / valuation","future funding obligations"]}.get(kind,[])
    low=_clean(text).lower()
    if any(x in low for x in ("antitrust","regulatory","approval","competition")): base.append("regulatory")
    if status=="Announced / Pending": base.append("closing risk")
    return ", ".join(dict.fromkeys(base)) or "Review transaction-specific execution and capital-allocation risk"
def _milestone(kind,status,text):
    m=re.search(r"expected to close(?:\s+in|\s+by)?\s+([^.;]{3,45})",_clean(text).lower())
    if m: return "Expected close: "+m.group(1).strip().rstrip(",")
    if status=="Announced / Pending": return "Regulatory / shareholder approvals and closing; then purchase-price allocation or integration"
    if kind=="Acquisition / M&A": return "Track purchase-price allocation, acquired growth/margins, integration costs, synergies and ROIC"
    if kind in {"Major Commercial Contract","Strategic Partnership"}: return "Track implementation, booked revenue/backlog, capacity, economics and renewal/expansion"
    if kind=="Strategic Investment": return "Track funding milestones, ownership changes, strategic deliverables and accounting effects"
    if kind=="Divestiture / Asset Sale": return "Track close, proceeds, stranded costs, retained stake and capital redeployment"
    return "Track execution milestones and the next filing / earnings disclosure"
def _tokens(text): return {w for w in re.findall(r"[a-z0-9]{3,}",_clean(text).lower()) if w not in STOP}
def _similar(a,b):
    if a.get("type")!=b.get("type"): return False
    ta,tb=_tokens(a.get("title")),_tokens(b.get("title")); overlap=len(ta&tb)/max(1,min(len(ta),len(tb))) if ta and tb else 0
    ca,cb=_clean(a.get("counterparty")).lower(),_clean(b.get("counterparty")).lower()
    return overlap>=.60 or bool(ca and cb and ca==cb and ca!="review source / headline" and overlap>=.30)
def _merge(a,b):
    primary,other=(b,a) if b.get("score",0)>=a.get("score",0) else (a,b); out=dict(primary)
    if not out.get("amount") and other.get("amount"): out["amount"]=other["amount"]; out["deal_value"]=other.get("deal_value")
    if out.get("counterparty")=="Review source / headline" and other.get("counterparty")!="Review source / headline": out["counterparty"]=other.get("counterparty")
    if other.get("source_quality","").startswith("High") and not out.get("source_quality","").startswith("High"):
        out["source_quality"]=other.get("source_quality"); out["publisher"]=other.get("publisher"); out["url"]=other.get("url")
    out["evidence"]=_clip(" | ".join(dict.fromkeys(x for x in (out.get("evidence"),other.get("evidence")) if x)),620); return out


def analyze_items(ticker,items,info=None,horizon_days=DEFAULT_HORIZON_DAYS,max_deals=DEFAULT_MAX_DEALS):
    info=info or {}; cutoff=date.today()-timedelta(days=horizon_days); deals=[]
    for item in items:
        published=_date(item.get("date"))
        try:
            if published and date.fromisoformat(published)<cutoff: continue
        except Exception: pass
        title=_clean(item.get("title")); summary=_clean(item.get("summary")); text=_clean(title+" "+summary); kind=_deal_type(text)
        if not kind: continue
        amount=_amount(text); quality=_source_quality(item); materiality,ratio=_materiality(amount,info,text); score=_score(kind,amount,materiality,quality,text,published)
        if score < (28 if kind in {"Acquisition / M&A","Divestiture / Asset Sale"} else 34): continue
        status=_status(text); area=_area(text)
        rec={"date":published,"status":status,"type":kind,"counterparty":_counterparty(title,ticker,info),"title":title,"amount":amount,"deal_value":amount.get("raw") if amount else None,"materiality":materiality,"market_cap_ratio":ratio,"strategic_area":area,"terms":_terms(text),"rationale":_rationale(kind,area),"financial_impact":_financial(kind,amount),"risks":_risks(kind,status,text),"regulatory":_regulatory(text,status),"next_milestone":_milestone(kind,status,text),"evidence":_clip(summary or title,560),"source_quality":quality,"publisher":_clean(item.get("publisher")),"url":item.get("url"),"score":score}
        for i,existing in enumerate(deals):
            if _similar(existing,rec): deals[i]=_merge(existing,rec); break
        else: deals.append(rec)
    deals.sort(key=lambda x:(x.get("score",0),x.get("date") or ""),reverse=True); return deals[:max_deals]


def collect_deals(ticker,info=None,horizon_days=DEFAULT_HORIZON_DAYS,max_deals=DEFAULT_MAX_DEALS):
    yahoo=_yahoo_items(ticker,100); sec=_sec_items(ticker,horizon_days); deals=analyze_items(ticker,yahoo+sec,info or {},horizon_days,max_deals)
    return deals,{"yahoo_items":len(yahoo),"sec_items":len(sec),"qualified_deals":len(deals),"horizon_days":horizon_days,"sources":"Yahoo Finance news/press releases + SEC/EDGAR when available"}


def _section(ws,row,title,end=18):
    for c in range(1,end+1): ws.cell(row,c).fill=_fill(NAVY); ws.cell(row,c).font=Font(bold=True,color=WHITE,size=11)
    ws.cell(row,1,title)
def _header(ws,row,heads):
    for c,v in enumerate(heads,1):
        x=ws.cell(row,c,v); x.fill=_fill(BLUE); x.font=Font(bold=True,color=WHITE); x.border=Border(bottom=THIN); x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)


def ensure_deal_analysis(wb,ticker,info=None,horizon_days=DEFAULT_HORIZON_DAYS,max_deals=DEFAULT_MAX_DEALS):
    """Create/refresh the professional Deals & Transactions worksheet."""
    info=info or {}; name="Deals & Transactions"
    if name in wb.sheetnames: wb.remove(wb[name])
    ws=wb.create_sheet(name); ws.sheet_view.showGridLines=False
    for c in range(1,19): ws.cell(1,c).fill=_fill(NAVY); ws.cell(2,c).fill=_fill(NAVY)
    ws["A1"]=f"{ticker} — Recent Material Deals & Transactions"; ws["A1"].font=Font(bold=True,color=WHITE,size=19)
    ws["A3"]="Source-backed transaction monitor. Facts from public press/news metadata and SEC filings are separated from rule-based analyst lenses. Missing deal economics remain blank. Open source links before changing forecasts, valuation or the investment thesis."
    ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True); ws.merge_cells("A3:R3"); ws.row_dimensions[3].height=38
    deals,meta=collect_deals(ticker,info,horizon_days,max_deals)
    _section(ws,5,"Deal Monitor Summary")
    disclosed=[x for x in deals if x.get("amount")]; largest=max(disclosed,key=lambda x:x["amount"].get("value") or 0) if disclosed else None
    pending=sum(x.get("status") in {"Announced / Pending","Approved / Clearing","Signed / Active"} for x in deals); types=sorted({x.get("type") for x in deals if x.get("type")})
    rows=[("Material transactions retained",len(deals),"Materiality and source-quality filtered"),("Pending / active items",pending,"Require closing, implementation or delivery monitoring"),("Largest disclosed transaction",largest.get("deal_value") if largest else "—",largest.get("title") if largest else "No disclosed amount verified"),("Deal types represented",", ".join(types) if types else "—","Classification from retrieved source text"),("Coverage window",f"{horizon_days} days",meta.get("sources")),("Source records reviewed",meta.get("yahoo_items",0)+meta.get("sec_items",0),f"News/press: {meta.get('yahoo_items',0)} | SEC excerpts: {meta.get('sec_items',0)}")]
    for r,(label,val,note) in enumerate(rows,6):
        ws.cell(r,1,label).font=Font(bold=True); ws.cell(r,2,val); ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=8); ws.cell(r,3,note); ws.cell(r,3).font=Font(color=GREY); ws.cell(r,3).alignment=Alignment(wrap_text=True)
    _section(ws,13,"Recent Material Deals — Facts + Analyst Lens")
    heads=["Date","Status","Type","Counterparty / Asset","Deal / Headline","Disclosed Value","Materiality","Strategic Area","Structure / Terms","Strategic Rationale — Analyst Lens","Financial / Valuation Impact — Analyst Lens","Key Risks — Analyst Lens","Regulatory / Closing","Next Milestone","Source Evidence","Source Quality","Publisher","Source Link"]
    _header(ws,14,heads); row=15
    for d in deals:
        vals=[d.get("date"),d.get("status"),d.get("type"),d.get("counterparty"),d.get("title"),d.get("deal_value") or "Value not verified",d.get("materiality"),d.get("strategic_area"),d.get("terms"),d.get("rationale"),d.get("financial_impact"),d.get("risks"),d.get("regulatory"),d.get("next_milestone"),d.get("evidence"),d.get("source_quality"),d.get("publisher"),d.get("url")]
        for c,v in enumerate(vals,1): ws.cell(row,c,v); ws.cell(row,c).alignment=Alignment(vertical="top",wrap_text=True)
        st=d.get("status"); ws.cell(row,2).fill=_fill(PALE_GREEN if st=="Completed / Closed" else PALE_YELLOW if st in {"Announced / Pending","Approved / Clearing","Signed / Active"} else PALE_RED if st=="Terminated" else PALE_BLUE); ws.cell(row,2).font=Font(bold=True)
        if d.get("url"): ws.cell(row,18).hyperlink=d["url"]; ws.cell(row,18).font=Font(color=LINK_GREEN,underline="single")
        ws.row_dimensions[row].height=82; row+=1
    if not deals:
        ws.cell(row,1,"No sufficiently material recent transaction was verified from the current public feeds / filings."); ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=18); ws.cell(row,1).font=Font(italic=True,color=GREY); row+=2
    pending_deals=[x for x in deals if x.get("status") in {"Announced / Pending","Approved / Clearing","Signed / Active"}]
    _section(ws,row+1,"Pending / Active Deal Watchlist"); _header(ws,row+2,["Deal","Status","Value","Next Milestone","Primary Risk","Source"]); rr=row+3
    for d in pending_deals[:8]:
        for c,v in enumerate([d.get("title"),d.get("status"),d.get("deal_value") or "—",d.get("next_milestone"),d.get("risks"),d.get("url")],1): ws.cell(rr,c,v); ws.cell(rr,c).alignment=Alignment(vertical="top",wrap_text=True)
        if d.get("url"): ws.cell(rr,6).hyperlink=d["url"]; ws.cell(rr,6).font=Font(color=LINK_GREEN,underline="single")
        ws.row_dimensions[rr].height=58; rr+=1
    if not pending_deals: ws.cell(rr,1,"No pending / active material deal requires a closing or implementation watch from the retrieved sources."); ws.cell(rr,1).font=Font(italic=True,color=GREY)
    widths={"A":12,"B":20,"C":25,"D":30,"E":52,"F":20,"G":24,"H":28,"I":34,"J":50,"K":55,"L":40,"M":38,"N":48,"O":62,"P":28,"Q":22,"R":52}
    for col,w in widths.items(): ws.column_dimensions[col].width=w
    ws.freeze_panes="A15"; ws.auto_filter.ref=f"A14:R{max(14,row-1)}"; return ws
