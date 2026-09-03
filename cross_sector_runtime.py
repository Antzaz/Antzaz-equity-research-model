from __future__ import annotations

"""Runtime installation of cross-sector model safeguards.

`safe_update_model` and `commodity_safe_runner` intentionally keep the legacy generator stable.
This module patches only the decision points that must be business-model aware: statement-profile
selection, score suitability, final segment fallback and Data Quality disclosure.
"""

from copy import deepcopy
from typing import Any

from openpyxl.styles import Alignment, Font, PatternFill

from business_model_registry import get_business_model_policy, workbook_policy
import issuer_statement_profiles as statement_profiles
import full_financial_statements_v3 as full_statements
import output_quality_v3 as output_quality
import score_integration_v2
import decision_view_v2
from sector_score_policy import compute_score_bundle, advanced_scorecard


_CONTEXT: dict[str, dict[str, Any]] = {}
_INSTALLED = False


def register_workbook_context(wb, ticker: str | None = None) -> dict[str, Any]:
    try:
        ws = wb["Company Data"]
        t = str(ticker or ws["B4"].value or "").upper().strip()
        ctx = {
            "ticker": t,
            "name": ws["B5"].value,
            "sector": ws["B6"].value,
            "industry": ws["B7"].value,
        }
        if t:
            _CONTEXT[t] = ctx
        return ctx
    except Exception:
        return {"ticker": str(ticker or "").upper().strip()}


def _context_for_ticker(ticker: str) -> dict[str, Any]:
    t = str(ticker or "").upper().strip()
    if t in _CONTEXT:
        return _CONTEXT[t]
    # Runtime market-data lookup is only a fallback.  Normal production builds register Company
    # Data context before the profile renderer is called.
    try:
        import yfinance as yf
        from runtime_data_guards import provider_symbol
        info = yf.Ticker(provider_symbol(t)).info or {}
        ctx = {
            "ticker": t,
            "name": info.get("longName") or info.get("shortName"),
            "sector": info.get("sector"),
            "industry": info.get("industry"),
        }
        _CONTEXT[t] = ctx
        return ctx
    except Exception:
        return {"ticker": t}


def contextual_statement_profile(ticker: str):
    """Use explicit issuer profiles first, then sector/business-model routing."""
    t = str(ticker or "").upper().strip()
    if t in statement_profiles.TICKER_PROFILE:
        return statement_profiles.get_statement_profile(t)

    ctx = _context_for_ticker(t)
    policy = get_business_model_policy(t, ctx.get("sector"), ctx.get("industry"), ctx.get("name"))

    if policy.key in {"bank", "capital_markets"}:
        profile = deepcopy(statement_profiles.PROFILES["bank"])
        profile["name"] = "Bank / capital-markets financial institution"
        return profile

    if policy.key == "insurance":
        # Berkshire's profile contains a conservative insurance-oriented taxonomy and deliberately
        # disables industrial FCF/net-debt treatment.  Generic insurers reuse those insurance rows;
        # non-applicable conglomerate lines remain blank rather than being fabricated.
        profile = deepcopy(statement_profiles.PROFILES["berkshire"])
        profile["name"] = "Insurance / reinsurance financial institution"
        profile["min_structure"] = (20, 24, 22)
        profile["min_mapped"] = (8, 11, 8)
        return profile

    if policy.key == "reit":
        profile = deepcopy(statement_profiles.PROFILES["default"])
        profile["name"] = "REIT / real-estate operating company"
        profile["derive_fcf"] = False
        profile["min_mapped"] = (10, 15, 10)
        return profile

    return statement_profiles.get_statement_profile(t)


def _meaningful_segment_names(wb) -> list[str]:
    if "Segment Analysis" not in wb.sheetnames:
        return []
    ws = wb["Segment Analysis"]
    ignore = {
        "segment", "business line / revenue group", "business line", "metric",
        "reported operating / reportable segments", "reported operating segments",
        "manual segment input", "source & data quality", "extraction status", "important",
        "sec 10-k source", "additional official/public sources", "revenue by business line / product group",
    }
    out = []
    for r in range(1, min(ws.max_row, 80) + 1):
        name = str(ws.cell(r, 1).value or "").strip()
        low = name.lower()
        if not name or low in ignore or low.startswith(("source", "status", "units", "research rule")):
            continue
        if any(token in low for token in ("analysis", "revenue by", "reported operating")):
            continue
        if name not in out:
            out.append(name)
    return out


def _numeric_segment_rows(wb) -> int:
    if "Segment Analysis" not in wb.sheetnames:
        return 0
    ws = wb["Segment Analysis"]
    count = 0
    for r in range(1, min(ws.max_row, 80) + 1):
        if str(ws.cell(r, 1).value or "").strip() and any(
            isinstance(ws.cell(r, c).value, (int, float)) for c in range(2, min(ws.max_column, 10) + 1)
        ):
            count += 1
    return count


def _company_data_business_names(wb) -> tuple[list[str], str]:
    if "Company Data" not in wb.sheetnames:
        return [], ""
    ws = wb["Company Data"]
    header = None
    for r in range(1, min(ws.max_row, 60) + 1):
        if str(ws.cell(r, 1).value or "").strip().lower() == "business / segment":
            header = r
            break
    if not header:
        return [], ""
    names, source = [], ""
    for r in range(header + 1, min(ws.max_row, header + 10) + 1):
        name = str(ws.cell(r, 1).value or "").strip()
        product = str(ws.cell(r, 2).value or "").strip()
        src = str(ws.cell(r, 4).value or "").strip()
        if src and not source:
            source = src
        if not product or not name or name.lower() in {"business profile", "reported business / segment"}:
            continue
        if name not in names:
            names.append(name)
    return names[:10], source


def ensure_descriptive_segment_contract(wb, ticker: str):
    """Guarantee a coherent Segment Analysis without inventing undisclosed economics.

    Numeric official segment data always wins.  If the parser could not even retain useful names,
    the final Company Data taxonomy is used as a descriptive-only fallback with blank financials.
    """
    if _numeric_segment_rows(wb) >= 3 or len(_meaningful_segment_names(wb)) >= 2:
        return {"changed": False, "reason": "existing official/discovered segment coverage retained"}

    names, source = _company_data_business_names(wb)
    if not names:
        return {"changed": False, "reason": "no reliable descriptive business taxonomy available"}

    try:
        from segment_analysis_v2 import _write_sheet, _years
        segments = [(name, [None, None, None], [None, None, None]) for name in names]
        status = (
            "DESCRIPTIVE FALLBACK — financial segment extraction was unavailable. Business/segment "
            "names come from the final Company Data issuer/profile taxonomy; financial cells remain "
            "blank because undisclosed segment economics are never estimated."
        )
        _write_sheet(wb, str(ticker).upper(), _years(wb), segments, [], source, status)
        return {"changed": True, "segments": names, "source": source}
    except Exception as exc:
        return {"changed": False, "reason": repr(exc)}


def append_business_model_quality(wb, ticker: str) -> None:
    if "Data Quality" not in wb.sheetnames:
        return
    policy = workbook_policy(wb, ticker)
    ws = wb["Data Quality"]
    label = "Business-model / valuation suitability"
    row = next(
        (r for r in range(1, ws.max_row + 1) if str(ws.cell(r, 1).value or "").strip() == label),
        ws.max_row + 1,
    )
    ws.cell(row, 1, label)
    ws.cell(row, 2, "PASS")
    ws.cell(row, 3, (
        f"{policy.label}. Primary framework: {policy.primary_valuation}. "
        + ("Industrial FCF/reverse DCF is permitted as a diagnostic." if policy.reverse_dcf_allowed else "Industrial FCF/reverse DCF is excluded from primary scoring/expectations analysis.")
        + (" Commodity mid-cycle normalization is required." if policy.commodity_normalization else "")
    ))
    ws.cell(row, 2).fill = PatternFill("solid", fgColor="E2F0D9")
    ws.cell(row, 2).font = Font(bold=True)
    ws.cell(row, 3).alignment = Alignment(wrap_text=True)

    # If an industrial DCF is not suitable, never leave the generic reliability row looking like
    # an unconditional PASS.  The workbook may still retain the DCF sheet for diagnostics/template
    # compatibility, but it is explicitly gated out of the score engine.
    if not policy.reverse_dcf_allowed:
        for r in range(1, ws.max_row + 1):
            if str(ws.cell(r, 1).value or "").strip() == "Valuation-model reliability gate":
                ws.cell(r, 2, "REVIEW")
                ws.cell(r, 2).fill = PatternFill("solid", fgColor="FFF2CC")
                ws.cell(r, 3, f"Business-model gate: conventional industrial DCF is diagnostic only. Prefer {policy.primary_valuation}.")
                break


def install_cross_sector_runtime(safe_module=None, update_model_module=None) -> None:
    global _INSTALLED
    if _INSTALLED:
        return

    # Functions imported with `from ... import` retain the function object but resolve their module
    # globals at execution time, so replacing these profile lookups is sufficient for statement and
    # quality logic.
    full_statements.get_statement_profile = contextual_statement_profile
    output_quality.get_statement_profile = contextual_statement_profile
    output_quality.compute_score_bundle = compute_score_bundle
    score_integration_v2.compute_score_bundle = compute_score_bundle
    decision_view_v2.compute_score_bundle = compute_score_bundle

    if safe_module is not None:
        safe_module.compute_score_bundle = compute_score_bundle
        safe_module.advanced_scorecard = advanced_scorecard

    if update_model_module is not None:
        current_statements = update_model_module.ensure_financial_statements
        if not getattr(current_statements, "_cross_sector_wrapped", False):
            def statements_wrapper(wb, ticker, facts):
                register_workbook_context(wb, ticker)
                return current_statements(wb, ticker, facts)
            statements_wrapper._cross_sector_wrapped = True
            update_model_module.ensure_financial_statements = statements_wrapper

        current_extensions = update_model_module.ensure_research_extensions
        if not getattr(current_extensions, "_cross_sector_wrapped", False):
            def extensions_wrapper(wb, ticker, info=None):
                register_workbook_context(wb, ticker)
                result = current_extensions(wb, ticker, info)
                seg = ensure_descriptive_segment_contract(wb, ticker)
                if seg.get("changed"):
                    print(f"Cross-sector Segment Analysis fallback: {ticker} -> {', '.join(seg.get('segments') or [])}")
                    # Segment coverage changed after the normal final quality pass; rerun the same
                    # public-data quality controls so the saved workbook reflects final state.
                    try:
                        if safe_module is not None:
                            safe_module.ensure_quality_checks(wb, ticker)
                        else:
                            output_quality.ensure_quality_checks(wb, ticker)
                    except Exception as exc:
                        print(f"Warning: post-segment cross-sector quality refresh failed: {exc}")
                append_business_model_quality(wb, ticker)
                return result
            extensions_wrapper._cross_sector_wrapped = True
            update_model_module.ensure_research_extensions = extensions_wrapper

    _INSTALLED = True
