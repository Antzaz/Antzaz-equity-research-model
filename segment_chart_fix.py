"""Final cross-company reliability, peer, segment-chart, AI and news refresh stage.

This stage deliberately runs after the first-pass workbook build. It repairs shorter
historical records, applies verified issuer fallbacks, anchors near-term revenue to current
consensus when available, recalibrates cash conversion, rebuilds analytics that depended
on stale history, enforces dynamic same-sector peers, recreates segment/business charts,
and generates the current news-impact research sheet.
"""

import os
import re
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ai_effect_analysis import ensure_ai_impact_analysis
from news_analysis import ensure_news_analysis
from segment_analysis_v2 import ensure_segment_analysis_v2
from amzn_model_repair import repair_amzn_model
from cvs_model_repair import repair_cvs_model
from gev_model_repair import repair_gev_model, apply_gev_guidance_assumptions, repair_gev_expectations
from consensus_rebase import rebase_near_term_revenue
from cross_company_cleanup import refresh_cross_company_tabs
from dynamic_peer_engine import ensure_dynamic_peer_comps
from model_reliability import prepare_model_reliability, finalize_model_reliability
from model_quality_v3 import calibrate_scenario_cash_flow, ensure_model_quality
from advanced_analytics_v2 import ensure_advanced_analytics
from institutional_layers import ensure_institutional_layers
from sector_peer_bayes import ensure_bayesian_base_rates, repair_cross_sheet_context

GREY="666666"; PALE_GREEN="E2F0D9"; GOLD="FFF2CC"
FMT_BN='#,##0.0;[Red](#,##0.0);-'; FMT_PCT='0.0%;[Red](0.0%);-'
SEGMENT_HEADERS={"User-Agent":os.getenv("SEC_USER_AGENT","Personal Equity Research Model contact@example.com")}

def _fill(c): return PatternFill("solid",fgColor=c)
def _find(ws,label):
    target=label.strip().lower()
    for r in range(1,ws.max_row+1):
        if str(ws.cell(r,1).value or "").strip().lower()==target: return r
    return None
def _find_any(ws,labels):
    for label in labels:
        r=_find(ws,label)
        if r is not None: return r
    return None

def _segment_key(name):
    text=str(name or "").strip(); m=re.search(r"\(([A-Z]{2,10})\)\s*$",text)
    if m: return m.group(1).lower()
    if re.fullmatch(r"[A-Z]{2,10}",text): return text.lower()
    return re.sub(r"[^a-z0-9]","",text.lower())
def _dedupe_segment_aliases(seg):
    section=_find_any(seg,("Reported Operating / Reportable Segments","Reported Operating Segments","Reported Segments"))
    if section is None: return
    header=section+1; stop=_find_any(seg,("Revenue by Business Line / Product Group","Revenue by Business Line","Revenue by Business Line / Disclosed Revenue Group")) or min(seg.max_row+1,header+20); seen={}
    for r in range(header+1,stop):
        name=seg.cell(r,1).value
        if not name: continue
        key=_segment_key(name)
        if not key: continue
        if key not in seen: seen[key]=r; continue
        keep=seen[key]
        for c in range(2,min(16,seg.max_column)+1):
            if seg.cell(keep,c).value in (None,"") and seg.cell(r,c).value not in (None,""): seg.cell(keep,c).value=seg.cell(r,c).value
        for c in range(1,min(16,seg.max_column)+1): seg.cell(r,c).value=None

def _chart_row(ch):
    try: return ch.anchor._from.row
    except Exception: return None
def _remove_target_charts(ws): ws._charts=[ch for ch in ws._charts if _chart_row(ch)!=52]
def _labels(ch):
    try: ch.dLbls=DataLabelList(); ch.dLbls.showVal=True
    except Exception: pass

def _shorten_final_titles(wb):
    if "Three-Case Scenarios" in wb.sheetnames: wb["Three-Case Scenarios"]["A1"]="Three-Case Scenarios — 10-Year DCF"
    if "AI Impact Analysis" in wb.sheetnames:
        ai=wb["AI Impact Analysis"]; ai["A3"]="Institutional AI lens: reported monetization, segment exposure, capital intensity, disruption risk, and AI upside/downside versus the existing Base DCF."; ai["A23"]="AI Segment Exposure & Scoring"; ai["A33"]="AI Surprise Scenarios vs Base Case"

def _peer_quality_check(wb,ticker,peers):
    if "Data Quality" not in wb.sheetnames: return
    target_sector=str(wb["Company Data"]["B6"].value or "") if "Company Data" in wb.sheetnames else ""; peer_ws=wb["Peer Comps"] if "Peer Comps" in wb.sheetnames else None; sectors=[]
    if peer_ws:
        for r in range(5,10):
            if peer_ws.cell(r,2).value: sectors.append(str(peer_ws.cell(r,9).value or ""))
    ok=bool(sectors) and all(s==target_sector for s in sectors); status="PASS" if ok and len(sectors)>=3 else ("REVIEW" if ok else "FAIL"); ws=wb["Data Quality"]; row=None
    for r in range(1,ws.max_row+1):
        if str(ws.cell(r,1).value or "").strip()=="Peer comps sector alignment": row=r; break
    row=row or ws.max_row+1; ws.cell(row,1,"Peer comps sector alignment"); ws.cell(row,2,status); ws.cell(row,3,f"{len(sectors)} peer(s); target sector={target_sector}; peer sectors={sectors}"); ws.cell(row,4,"Every comparison company must match the target company's detected sector; exact industry is preferred."); ws.cell(row,2).fill=_fill(PALE_GREEN if status=="PASS" else GOLD); ws.cell(row,2).font=Font(bold=True)
    for c in range(1,5): ws.cell(row,c).alignment=Alignment(wrap_text=True,vertical="top")

def _apply_final_controls(wb,ticker):
    prepare_model_reliability(wb,ticker); repair_amzn_model(wb,ticker); repair_cvs_model(wb,ticker); repair_gev_model(wb,ticker); prepare_model_reliability(wb,ticker)
    try: rebase_near_term_revenue(wb,ticker)
    except Exception as exc: print(f"Warning: consensus revenue rebase failed: {exc}")
    try: apply_gev_guidance_assumptions(wb,ticker)
    except Exception as exc: print(f"Warning: GEV guidance anchor failed: {exc}")
    try: calibrate_scenario_cash_flow(wb)
    except Exception as exc: print(f"Warning: final cash-flow calibration failed: {exc}")
    peers=[]
    try: peers=ensure_dynamic_peer_comps(wb,ticker)
    except Exception as exc: print(f"Warning: dynamic peer selection failed: {exc}")
    try: ensure_advanced_analytics(wb,ticker)
    except Exception as exc: print(f"Warning: final Advanced Analytics refresh failed: {exc}")
    try: ensure_model_quality(wb,ticker)
    except Exception as exc: print(f"Warning: final Model Quality refresh failed: {exc}")
    try: ensure_institutional_layers(wb,ticker)
    except Exception as exc: print(f"Warning: final Institutional Layers refresh failed: {exc}")
    try: repair_gev_expectations(wb,ticker)
    except Exception as exc: print(f"Warning: GEV expectations guidance control failed: {exc}")
    try: refresh_cross_company_tabs(wb,ticker)
    except Exception as exc: print(f"Warning: final cross-company refresh failed: {exc}")
    try: peers=ensure_dynamic_peer_comps(wb,ticker)
    except Exception as exc: print(f"Warning: final dynamic peer refresh failed: {exc}")
    try: ensure_bayesian_base_rates(wb,ticker)
    except Exception as exc: print(f"Warning: Bayesian base-rate refresh failed: {exc}")
    try: repair_cross_sheet_context(wb)
    except Exception as exc: print(f"Warning: cross-sheet context repair failed: {exc}")
    try: finalize_model_reliability(wb,ticker)
    except Exception as exc: print(f"Warning: final model reliability controls failed: {exc}")
    _peer_quality_check(wb,ticker,peers)

def repair_segment_charts(wb,ticker):
    # Alphabet used a legacy special-case parser earlier in the updater; replace it here with
    # the same resilient narrative/table parser used for all other companies.
    if ticker.upper() in {"GOOGL","GOOG"}:
        try: ensure_segment_analysis_v2(wb,ticker,SEGMENT_HEADERS)
        except Exception as exc: print(f"Warning: Alphabet resilient segment refresh failed: {exc}")
    _apply_final_controls(wb,ticker)
    if not {"Analysis Charts","Segment Analysis"}.issubset(wb.sheetnames):
        ensure_ai_impact_analysis(wb,ticker)
        try: ensure_news_analysis(wb,ticker)
        except Exception as exc: print(f"Warning: Recent News & Impact failed: {exc}")
        _shorten_final_titles(wb); return
    ws=wb["Analysis Charts"]; seg=wb["Segment Analysis"]; _dedupe_segment_aliases(seg); _remove_target_charts(ws)
    for row in ws.iter_rows(min_row=2,max_row=20,min_col=39,max_col=45):
        for cell in row: cell.value=None
    section=_find_any(seg,("Revenue by Business Line / Product Group","Revenue by Business Line","Revenue by Business Line / Disclosed Revenue Group")); business=[]; latest_year="Latest"
    if section:
        header=section+1; latest_col=4; h=str(seg.cell(header,latest_col).value or ""); latest_year=h if h else "Latest"
        for r in range(header+1,min(seg.max_row,header+20)+1):
            name=seg.cell(r,1).value; val=seg.cell(r,latest_col).value
            if name in (None,""):
                if business: break
                continue
            if val in (None,""): continue
            business.append((r,str(name)))
    ws["AM2"]="Business Line / Revenue Group"; ws["AN2"]=f"{latest_year} Revenue ($bn)"
    for out_r,(src_r,name) in enumerate(business[:10],3): ws.cell(out_r,39,name); ws.cell(out_r,40,f"='Segment Analysis'!D{src_r}"); ws.cell(out_r,40).number_format=FMT_BN
    if business:
        end=2+min(10,len(business)); ch=BarChart(); ch.type="bar"; ch.style=10; ch.title=f"{latest_year} Revenue Mix"; ch.height=8.5; ch.width=13.5; ch.legend=None; ch.add_data(Reference(ws,min_col=40,min_row=2,max_row=end),titles_from_data=True); ch.set_categories(Reference(ws,min_col=39,min_row=3,max_row=end)); ch.x_axis.numFmt="$0"; ch.x_axis.title=f"{latest_year} revenue ($bn)"; ch.visible_cells_only=False; ch.display_blanks="gap"; _labels(ch); ws.add_chart(ch,"A53"); ws["A70"]=f"Units: {latest_year} revenue ($bn)"; ws["D70"]="Issuer-disclosed segments / revenue groups"; ws["F70"]="Source: Segment Analysis / annual filing"; ws["A71"]="Business mix uses only disclosed categories; no standalone product revenue is invented."
    else: ws["A55"]="No reliable disclosed business-line revenue was extracted. Segment names may still be available on Segment Analysis; charts require financial values."; ws["A55"].font=Font(italic=True,color=GREY)
    seg_section=_find_any(seg,("Reported Operating / Reportable Segments","Reported Operating Segments","Reported Segments")); margins=[]; margin_col=None; profit_col=None; margin_header="Latest Margin"; profit_header="Segment profitability"
    if seg_section:
        header=seg_section+1
        for c in range(1,min(18,seg.max_column)+1):
            text=str(seg.cell(header,c).value or "")
            if "Margin" in text and "Δ" not in text: margin_col=c; margin_header=text
            if any(k in text for k in ("Op. Income","Operating Income","EBITDA","Segment Profit","Adjusted Earnings")): profit_col=c; profit_header=text
        if margin_col and profit_col:
            for r in range(header+1,min(seg.max_row,header+15)+1):
                name=seg.cell(r,1).value; profit=seg.cell(r,profit_col).value
                if name in (None,""):
                    if margins: break
                    continue
                if profit in (None,""): continue
                margins.append((r,str(name)))
    ws["AP2"]="Segment"; ws["AQ2"]=margin_header
    for out_r,(src_r,name) in enumerate(margins[:10],3): ws.cell(out_r,42,name); ws.cell(out_r,43,f"='Segment Analysis'!{get_column_letter(margin_col)}{src_r}"); ws.cell(out_r,43).number_format=FMT_PCT
    if margins:
        end=2+min(10,len(margins)); ch=BarChart(); ch.type="bar"; ch.style=10; ch.title=f"{margin_header} by Segment"; ch.height=8.5; ch.width=13.5; ch.legend=None; ch.add_data(Reference(ws,min_col=43,min_row=2,max_row=end),titles_from_data=True); ch.set_categories(Reference(ws,min_col=42,min_row=3,max_row=end)); ch.x_axis.numFmt="0%"; ch.x_axis.title=margin_header; ch.visible_cells_only=False; ch.display_blanks="gap"; _labels(ch); ws.add_chart(ch,"I53"); ws["I70"]=f"Margin = {profit_header} ÷ segment revenue"; ws["M70"]="Source: Segment Analysis / annual filing"; ws["I71"]="Profitability uses the issuer's disclosed segment performance measure; missing economics are not estimated."
    else: ws["I55"]="Segment profitability is not disclosed or not reliably extracted. Segment names remain visible, while the chart stays blank rather than estimating economics."; ws["I55"].font=Font(italic=True,color=GREY)
    ws["A112"]=f"External segment source: {ticker} annual filing / Segment Analysis sheet. Monte Carlo, scenario, stress and DCF values are model outputs based on workbook assumptions."; ws["A112"].font=Font(italic=True,color=GREY,size=9)
    ensure_ai_impact_analysis(wb,ticker)
    try: ensure_news_analysis(wb,ticker)
    except Exception as exc: print(f"Warning: Recent News & Impact failed: {exc}")
    finalize_model_reliability(wb,ticker); _shorten_final_titles(wb)
