"""Final cross-company reliability controls for the equity research workbook.

Key invariants:
- missing foreign-issuer history is recovered from issuer IR / 20-F IFRS / Yahoo annual statements;
- available historical periods are right-aligned in Historical Financials B:G so G is
  always the latest actual, without inventing missing history;
- derived historical formulas are blank-safe;
- dashboards and historical CAGR labels use the actual observed window;
- forward scenario D&A and normalized capex paths are rebuilt from repaired history;
- blank consensus/model fields do not become false -100% gaps;
- stress-test equity values are floored at zero;
- data-quality checks judge actual available history rather than requiring six years.
"""

import os
import statistics
import requests
import yfinance as yf
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from issuer_source_engine import (
    build_crossborder_history,
    patch_financial_statements_from_history,
    write_source_hierarchy,
)

PALE_GREEN="E2F0D9"; PALE_RED="FCE4D6"; GOLD="FFF2CC"; GREY="666666"
LINK_GREEN="008000"; BLACK="000000"
FMT_PCT='0.0%;[Red](0.0%);-'; FMT_BN='#,##0.0;[Red](#,##0.0);-'; FMT_PRICE='$#,##0.00;[Red]($#,##0.00);-'
RAW_ROWS=(3,4,6,9,11,12,14,15,18,19,21)
SEC_HEADERS={"User-Agent":os.getenv("SEC_USER_AGENT","Personal Equity Research Model contact@example.com")}
FOREIGN_FORMS={"20-F","20-F/A","40-F","40-F/A","6-K"}


def _fill(c): return PatternFill("solid",fgColor=c)

def _num(v,default=None):
    try:
        if isinstance(v,bool): return default
        return float(v)
    except Exception: return default


def _fade(start,end):
    return [start+(end-start)*i/9 for i in range(10)]


def history_points(wb,row=4):
    if "Historical Financials" not in wb.sheetnames: return []
    h=wb["Historical Financials"]; out=[]
    for c in range(2,8):
        y=h.cell(3,c).value; v=_num(h.cell(row,c).value)
        if isinstance(y,(int,float)) and v is not None: out.append((c,int(y),v))
    return out


def _fetch_sec_facts(ticker):
    try:
        items=requests.get("https://www.sec.gov/files/company_tickers.json",headers=SEC_HEADERS,timeout=25).json()
        cik=None
        for item in items.values():
            if str(item.get("ticker","")).upper()==str(ticker).upper():
                cik=str(item.get("cik_str")).zfill(10); break
        if not cik: return None
        r=requests.get(f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json",headers=SEC_HEADERS,timeout=30)
        r.raise_for_status(); return r.json()
    except Exception:
        return None


def _write_history_raw(wb,hist):
    if not hist or "Historical Financials" not in wb.sheetnames: return
    ws=wb["Historical Financials"]
    for r in RAW_ROWS:
        for c in range(2,8): ws.cell(r,c).value=None
    years=sorted(hist)[-6:]; start=8-len(years)
    for c,y in enumerate(years,start):
        d=hist[y]; scale=lambda x:x/1e9 if x is not None else None
        ws.cell(3,c).value=y
        ws.cell(4,c).value=scale(d.get("revenue"))
        ws.cell(6,c).value=scale(d.get("cost"))
        ws.cell(9,c).value=scale(d.get("op"))
        ws.cell(11,c).value=scale(d.get("ni"))
        ws.cell(12,c).value=d.get("eps")
        ws.cell(14,c).value=scale(d.get("ocf"))
        ws.cell(15,c).value=scale(d.get("capex"))
        ws.cell(18,c).value=scale(d.get("depr"))
        ws.cell(19,c).value=scale(d.get("rd"))
        ws.cell(21,c).value=scale(d.get("sbc"))


def _bank_mode(wb,ticker,info):
    sector=str((info or {}).get("sector") or "")
    industry=str((info or {}).get("industry") or "")
    is_bank=("bank" in industry.lower()) or (sector=="Financial Services" and "bank" in str(wb["Company Data"]["B7"].value or "").lower() if "Company Data" in wb.sheetnames else False)
    if not is_bank or "Historical Financials" not in wb.sheetnames: return
    h=wb["Historical Financials"]
    h["A9"]="Operating / Pre-Tax Profit"
    h["A10"]="Pre-Tax Margin"
    h["A14"]="Operating Cash Flow (not core bank metric)"
    h["A15"]="Capex (not core bank valuation metric)"
    h["A16"]="Free Cash Flow (not used for bank valuation)"
    h["A17"]="FCF Margin (not used for bank valuation)"
    h["A24"]="Bank methodology"
    h["B24"]="For banks, earnings, ROE/ROTE, capital ratios, book value, deposits/funding and credit quality are more decision-useful than industrial FCF/capex DCF. Historical income data is populated, but conventional FCF valuation should be treated as REVIEW."
    h["B24"].font=Font(italic=True,color=GREY); h["B24"].alignment=Alignment(wrap_text=True)


def _recover_crossborder_history(wb,ticker):
    """Recover missing history before downstream analytics; cache it on the workbook."""
    existing=history_points(wb,4)
    cached=getattr(wb,"_issuer_source_history",None)
    meta=getattr(wb,"_issuer_source_meta",None)
    info=getattr(wb,"_issuer_source_info",None)
    if cached:
        patch_financial_statements_from_history(wb,cached,meta)
        write_source_hierarchy(wb,ticker,meta or {})
        _bank_mode(wb,ticker,info or {})
        return cached
    # Only invoke network recovery when the main history is absent or extremely thin.
    if len(existing)>=2:
        return None
    try: info=yf.Ticker(ticker).info or {}
    except Exception: info={}
    facts=_fetch_sec_facts(ticker)
    try:
        hist,meta=build_crossborder_history(ticker,info,facts)
    except Exception as exc:
        print(f"Warning: issuer/source history recovery failed: {exc}"); return None
    if hist:
        _write_history_raw(wb,hist)
        setattr(wb,"_issuer_source_history",hist)
        setattr(wb,"_issuer_source_meta",meta)
        setattr(wb,"_issuer_source_info",info)
        write_source_hierarchy(wb,ticker,meta)
        patch_financial_statements_from_history(wb,hist,meta)
        _bank_mode(wb,ticker,info)
        print(f"Historical source recovery: ticker={ticker}, years={sorted(hist)}, priority=issuer IR -> SEC annual XBRL -> Yahoo fallback")
        return hist
    return None


def _repair_foreign_filings(wb,ticker):
    if "Filings" not in wb.sheetnames: return
    try:
        items=requests.get("https://www.sec.gov/files/company_tickers.json",headers=SEC_HEADERS,timeout=25).json(); cik=None
        for item in items.values():
            if str(item.get("ticker","")).upper()==str(ticker).upper(): cik=str(item.get("cik_str")).zfill(10); break
        if not cik: return
        recent=requests.get(f"https://data.sec.gov/submissions/CIK{cik}.json",headers=SEC_HEADERS,timeout=30).json().get("filings",{}).get("recent",{})
    except Exception: return
    forms=list(recent.get("form",[]))
    if not any(f in FOREIGN_FORMS for f in forms): return
    ws=wb["Filings"]
    for r in range(4,20):
        for c in range(1,6): ws.cell(r,c).value=None
    row=4
    allowed={"20-F","20-F/A","40-F","40-F/A","6-K","8-K","DEF 14A"}
    for form,period,filed,acc,doc in zip(forms,recent.get("reportDate",[]),recent.get("filingDate",[]),recent.get("accessionNumber",[]),recent.get("primaryDocument",[])):
        if form not in allowed: continue
        url=f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{str(acc).replace('-','')}/{doc}"
        note="Annual foreign private issuer filing" if form in {"20-F","20-F/A","40-F","40-F/A"} else "Foreign issuer current report"
        for c,v in enumerate([form,period,filed,url,note],1): ws.cell(row,c).value=v
        row+=1
        if row>15: break


def _install_blank_safe_formulas(h):
    cols=list(range(2,8))
    for c in cols:
        L=get_column_letter(c); prev=get_column_letter(c-1) if c>2 else None
        h.cell(5,c).value=(f'=IF(OR({L}4="",{prev}4=""),"",{L}4/{prev}4-1)' if prev else None)
        h.cell(7,c).value=f'=IF(OR({L}4="",{L}6=""),"",{L}4-{L}6)'
        h.cell(8,c).value=f'=IFERROR({L}7/{L}4,"")'
        h.cell(10,c).value=f'=IFERROR({L}9/{L}4,"")'
        h.cell(13,c).value=f'=IFERROR({L}11/{L}12,"")'
        h.cell(16,c).value=f'=IF(OR({L}14="",{L}15=""),"",{L}14-{L}15)'
        h.cell(17,c).value=f'=IFERROR({L}16/{L}4,"")'
        h.cell(20,c).value=f'=IF(OR({L}19="",{L}4=""),"",{L}19/{L}4)'
        h.cell(22,c).value=f'=IF(OR({L}21="",{L}4=""),"",{L}21/{L}4)'
        for r in (5,8,10,17,20,22): h.cell(r,c).number_format=FMT_PCT
        for r in (7,16): h.cell(r,c).number_format=FMT_BN


def right_align_history(wb):
    if "Historical Financials" not in wb.sheetnames: return []
    h=wb["Historical Financials"]; source_cols=[]
    for c in range(2,8):
        y=h.cell(3,c).value
        if isinstance(y,(int,float)): source_cols.append(c)
    if not source_cols:
        _install_blank_safe_formulas(h); return []
    years=[int(h.cell(3,c).value) for c in source_cols]
    order=sorted(range(len(years)),key=lambda i:years[i])[-6:]
    source_cols=[source_cols[i] for i in order]; years=[years[i] for i in order]
    payload={r:[h.cell(r,c).value for c in source_cols] for r in RAW_ROWS}
    for r in RAW_ROWS:
        for c in range(2,8): h.cell(r,c).value=None
    start=8-len(years)
    for j,y in enumerate(years,start): h.cell(3,j).value=y
    for r in RAW_ROWS:
        if r==3: continue
        vals=payload[r]
        for j,v in enumerate(vals,start): h.cell(r,j).value=v
    _install_blank_safe_formulas(h)
    if len(years)<6:
        h["A24"]="Coverage note"
        h["B24"]=f"{len(years)} comparable annual periods are available ({years[0]}–{years[-1]}). They are right-aligned so the latest actual is in column G; missing history is not invented."
        h["B24"].font=Font(italic=True,color=GREY); h["B24"].alignment=Alignment(wrap_text=True)
    return years


def sync_scenario_d_and_a(wb):
    if not {"Historical Financials","Three-Case Scenarios"}.issubset(wb.sheetnames): return
    h=wb["Historical Financials"]; s=wb["Three-Case Scenarios"]
    rev=_num(h["G4"].value); da=_num(h["G18"].value)
    if not rev or da is None or da<0: return
    ratio=max(.002,min(.25,da/rev))
    for start in (2,14,26):
        for c in range(start,start+10): s.cell(18,c).value=ratio; s.cell(18,c).number_format=FMT_PCT


def sync_scenario_capex(wb):
    if not {"Historical Financials","Three-Case Scenarios"}.issubset(wb.sheetnames): return
    h=wb["Historical Financials"]; s=wb["Three-Case Scenarios"]
    rev=_num(h["G4"].value); latest_cap=_num(h["G15"].value); latest_da=_num(h["G18"].value)
    if not rev or latest_cap is None: return
    cap=max(.01,min(.40,abs(latest_cap)/rev)); dep=max(.002,min(.25,abs(latest_da)/rev)) if latest_da is not None else .04
    cap_hist=[]
    for c in range(2,8):
        rv=_num(h.cell(4,c).value); cx=_num(h.cell(15,c).value)
        if rv and cx is not None: cap_hist.append(abs(cx)/rv)
    med_cap=statistics.median(cap_hist[-3:]) if cap_hist else cap
    normalized=max(.02,min(.22,max(dep,med_cap*.65))); starts=[min(.40,cap+.03),cap,max(.01,cap-.02)]; targets=[min(.30,normalized+.03),normalized,max(.015,normalized-.02)]
    for cols,start,target in zip((range(2,12),range(14,24),range(26,36)),starts,targets):
        for c,v in zip(cols,_fade(start,target)): s.cell(20,c).value=v; s.cell(20,c).number_format=FMT_PCT


def _available_cagr(wb,row=4):
    pts=history_points(wb,row)
    if len(pts)<2: return None,None,None,None
    _,y0,v0=pts[0]; _,y1,v1=pts[-1]; n=max(1,y1-y0)
    if v0<=0 or v1<=0: return None,y0,y1,n
    return (v1/v0)**(1/n)-1,y0,y1,n


def _positive_cagr(wb,row=16):
    pts=[p for p in history_points(wb,row) if p[2]>0]
    if len(pts)<2: return None,None,None,None
    _,y0,v0=pts[0]; _,y1,v1=pts[-1]; n=max(1,y1-y0)
    return (v1/v0)**(1/n)-1,y0,y1,n


def repair_history_dashboards(wb):
    rev_cagr,y0,y1,_=_available_cagr(wb,4); fcf_cagr,fy0,fy1,_=_positive_cagr(wb,16)
    if "Dashboard" in wb.sheetnames:
        ws=wb["Dashboard"]
        if y0 is not None: ws["A8"]=f"Revenue CAGR ({y0}–{str(y1)[-2:]})"; ws["B8"]=rev_cagr; ws["B8"].number_format=FMT_PCT
        else: ws["A8"]="Available-history Revenue CAGR"; ws["B8"]=None
        if fy0 is not None: ws["A10"]=f"FCF CAGR ({fy0}–{str(fy1)[-2:]})"; ws["B10"]=fcf_cagr; ws["B10"].number_format=FMT_PCT
        else: ws["A10"]="Positive-history FCF CAGR"; ws["B10"]=None
    if "Visual Dashboard" in wb.sheetnames:
        ws=wb["Visual Dashboard"]
        if y0 is not None: ws["A10"]=f"Revenue CAGR ({y0}–{str(y1)[-2:]})"; ws["A11"]="=Dashboard!B8"; ws["A11"].number_format=FMT_PCT


def repair_expectation_gaps(wb):
    if "Expectations & Consensus" not in wb.sheetnames: return
    ws=wb["Expectations & Consensus"]
    for r in range(7,min(ws.max_row,40)+1):
        metric=str(ws.cell(r,1).value or "")
        if not metric: continue
        ws.cell(r,5).value=f'=IF(OR(C{r}="",D{r}=""),"",D{r}-C{r})'
        if metric in {"EBIT Margin","Capex / Revenue"}: ws.cell(r,6).value=f'=IF(OR(C{r}="",D{r}=""),"",D{r}-C{r})'
        else: ws.cell(r,6).value=f'=IF(OR(C{r}="",D{r}=""),"",IFERROR(D{r}/C{r}-1,""))'
        ws.cell(r,6).number_format=FMT_PCT


def repair_available_history_labels(wb):
    cagr,y0,y1,_=_available_cagr(wb,4)
    if "Moat & Competitive Advantage" in wb.sheetnames:
        ws=wb["Moat & Competitive Advantage"]
        for r in range(18,min(ws.max_row,28)+1):
            if "Revenue CAGR" in str(ws.cell(r,1).value or ""):
                ws.cell(r,1).value="Available-history Revenue CAGR"; ws.cell(r,2).value=cagr; ws.cell(r,2).number_format=FMT_PCT
                if y0 is not None: ws.cell(r,4).value=f"Historical growth context using observed {y0}–{y1} annual periods"
    if "Advanced Analytics" in wb.sheetnames:
        ws=wb["Advanced Analytics"]
        for r in range(40,min(ws.max_row,55)+1):
            if str(ws.cell(r,1).value or "")=="Growth":
                score=min(100,max(0,(cagr or 0)/.20*100)); ws.cell(r,2).value=score; ws.cell(r,3).value=(f"Available-history revenue CAGR ({y0}–{y1})" if y0 is not None else "Available-history revenue CAGR")
                vals=[_num(ws.cell(rr,2).value) for rr in range(r,r+7)]; vals=[v for v in vals if v is not None]
                if vals: ws["F42"]=statistics.mean(vals)
                break
    if "Base Rates & Probabilities" in wb.sheetnames:
        ws=wb["Base Rates & Probabilities"]
        if ws.max_row>=33:
            ws["A33"]="Available company history"; ws["B33"]=(f"Revenue CAGR over observed {y0}–{y1} period" if y0 is not None else "Revenue CAGR over available history"); ws["C33"]=(f"{y1-y0} years" if y0 is not None else ""); ws["E33"]=cagr; ws["E33"].number_format=FMT_PCT; ws["G33"]="Company-specific historical anchor; uses only observed comparable annual periods."


def floor_equity_values(wb):
    if "Three-Case Scenarios" not in wb.sheetnames: return
    ws=wb["Three-Case Scenarios"]
    for cell in ("B39","C39","D39"):
        v=ws[cell].value
        if isinstance(v,str) and v.startswith("=") and not v.upper().startswith("=MAX(0,"): ws[cell]=f"=MAX(0,{v[1:]})"
    for r in range(48,59):
        v=ws.cell(r,7).value
        if isinstance(v,str) and v.startswith("=") and not v.upper().startswith("=MAX(0,"): ws.cell(r,7).value=f"=MAX(0,{v[1:]})"
    if ws["A63"].value=="Severe Bear Value / Share": ws["B63"]="=MAX(0,G58)"; ws["B64"]='=IFERROR(B63/\'Company Data\'!B8-1,"")'


def clear_nonperiod_placeholders(wb):
    if "FCF & Capital Economics" in wb.sheetnames and "Historical Financials" in wb.sheetnames:
        h=wb["Historical Financials"]; ws=wb["FCF & Capital Economics"]
        for c in range(2,8):
            if not isinstance(h.cell(3,c).value,(int,float)):
                for r in range(7,18): ws.cell(r,c).value=None
                rr=24+(c-2)
                if rr<=ws.max_row:
                    for cc in range(1,8): ws.cell(rr,cc).value=None


def repair_data_quality(wb,ticker):
    if "Data Quality" not in wb.sheetnames: return
    ws=wb["Data Quality"]; pts=history_points(wb,4); years=[y for _,y,_ in pts]; rows={str(ws.cell(r,1).value or "").strip():r for r in range(1,ws.max_row+1)}
    def put(label,status,obs,why):
        r=rows.get(label)
        if r is None: r=ws.max_row+1; rows[label]=r
        ws.cell(r,1,label); ws.cell(r,2,status); ws.cell(r,3,obs); ws.cell(r,4,why); ws.cell(r,2).fill=_fill(PALE_GREEN if status=="PASS" else (GOLD if status=="REVIEW" else PALE_RED)); ws.cell(r,2).font=Font(bold=True)
        for c in range(1,5): ws.cell(r,c).alignment=Alignment(wrap_text=True,vertical="top")
    fixed=rows.get("Six annual periods") or rows.get("Annual history coverage")
    if fixed:
        ws.cell(fixed,1,"Annual history coverage"); status="PASS" if len(years)>=3 else ("REVIEW" if len(years)==2 else "FAIL"); ws.cell(fixed,2,status); ws.cell(fixed,3,f"{len(years)} comparable annual periods: {years}" if years else "No annual periods"); ws.cell(fixed,4,"Foreign issuers and newer/spun companies are evaluated on available comparable annual history, not a hard-coded six-year requirement."); ws.cell(fixed,2).fill=_fill(PALE_GREEN if status=="PASS" else GOLD)
    put("Years unique and ascending","PASS" if years==sorted(set(years)) and bool(years) else "FAIL",str(years),"Tests only observed nonblank annual periods.")
    put("Latest actual alignment","PASS" if years and wb["Historical Financials"]["G3"].value==max(years) else "FAIL",f"Historical Financials!G = {wb['Historical Financials']['G3'].value}","Downstream forecasts require G to hold the latest actual.")
    if getattr(wb,"_issuer_source_meta",None): put("Historical source hierarchy","PASS", "Issuer IR / SEC annual XBRL / Yahoo fallback", "Foreign-issuer history is no longer dependent on US 10-K us-gaap tags alone.")


def prepare_model_reliability(wb,ticker):
    _recover_crossborder_history(wb,ticker)
    _repair_foreign_filings(wb,ticker)
    years=right_align_history(wb)
    # On later passes Financial Statements already exists; reuse the cached source history to patch it.
    cached=getattr(wb,"_issuer_source_history",None)
    if cached: patch_financial_statements_from_history(wb,cached,getattr(wb,"_issuer_source_meta",None))
    sync_scenario_d_and_a(wb); sync_scenario_capex(wb)
    return years


def finalize_model_reliability(wb,ticker):
    repair_history_dashboards(wb); repair_expectation_gaps(wb); repair_available_history_labels(wb); floor_equity_values(wb); clear_nonperiod_placeholders(wb); repair_data_quality(wb,ticker)
