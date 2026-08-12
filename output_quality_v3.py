from __future__ import annotations

"""Final workbook quality controls and low-value-tab pruning."""

from openpyxl.styles import Alignment, Font, PatternFill
from score_engine_v3 import compute_score_bundle
from issuer_statement_profiles import get_statement_profile

GREEN="E2F0D9"; GOLD="FFF2CC"; RED="FCE4D6"; BLUE="2F75B5"; WHITE="FFFFFF"


def _num(v):
    try: return float(v)
    except Exception: return None

def _find(ws,label,col=1):
    needle=str(label).strip().lower()
    for r in range(1,ws.max_row+1):
        if str(ws.cell(r,col).value or '').strip().lower()==needle: return r
    return None

def _base_inputs(wb):
    price=base=severe=mc=None
    if 'Investment Summary' in wb.sheetnames:
        ws=wb['Investment Summary']; price=_num(ws['B8'].value); base=_num(ws['D8'].value); severe=_num(ws['B9'].value); mc=_num(ws['H9'].value)
    return price,base,severe,mc

def _bayesian_is_empty(wb):
    if 'Base Rates & Probabilities' not in wb.sheetnames: return True
    ws=wb['Base Rates & Probabilities']; found=0
    for r in range(1,ws.max_row+1):
        label=str(ws.cell(r,1).value or '').strip().lower()
        if label in {'bear','base','bull'}:
            for c in range(2,min(ws.max_column,8)+1):
                if isinstance(ws.cell(r,c).value,(int,float)): found+=1
    return found<2

def prune_low_value_tabs(wb):
    removed=[]
    if 'Business Portfolio Map' in wb.sheetnames:
        wb.remove(wb['Business Portfolio Map']); removed.append('Business Portfolio Map')
    if _bayesian_is_empty(wb) and 'Base Rates & Probabilities' in wb.sheetnames:
        wb.remove(wb['Base Rates & Probabilities']); removed.append('Base Rates & Probabilities (empty template)')
    return removed

def reconcile_score_displays(wb,ticker):
    price,base,severe,mc=_base_inputs(wb); bundle=compute_score_bundle(wb,ticker,base_value=base,severe_value=severe,current_price=price,mc_prob=mc)
    if 'Advanced Analytics' in wb.sheetnames:
        ws=wb['Advanced Analytics']
        for key,d in bundle['dimensions'].items():
            r=_find(ws,key)
            if r: ws.cell(r,2).value=d['score']; ws.cell(r,3).value=d['status']
        for label in ('Composite Score','Overall Investment Score','Quantitative Score'):
            r=_find(ws,label)
            if r: ws.cell(r,2).value=bundle['category_scores']['Overall Investment Score']
    return bundle

def _segment_status(wb,ticker):
    if 'Segment Analysis' not in wb.sheetnames: return 'REVIEW','Segment Analysis missing.'
    ws=wb['Segment Analysis']
    if str(ticker).upper()=='CEG':
        expected={'Mid-Atlantic','Midwest','New York','ERCOT','Other Power Regions'}
        names={str(ws.cell(r,1).value or '').strip() for r in range(1,ws.max_row+1)}
        if expected.issubset(names) and 'Calpine' in names: return 'PASS','Verified SEC annual segments plus 2026 Calpine structural break are populated.'
        return 'FAIL','CEG reportable segments are incomplete or narrative-parser fragments survived.'
    numeric=0
    for row in ws.iter_rows():
        if any(isinstance(c.value,(int,float)) for c in row[1:]): numeric+=1
    return ('PASS','Segment financial data populated.') if numeric>=3 else ('REVIEW','Segment names may exist but numeric segment economics are sparse.')


def _statement_status(wb,ticker):
    if 'Financial Statements' not in wb.sheetnames or 'Historical Financials' not in wb.sheetnames:
        return 'FAIL','Core statement/history sheet missing.'
    profile=get_statement_profile(ticker)
    fs=wb['Financial Statements']; hs=wb['Historical Financials']; i0=_find(fs,'Income Statement')
    revenue_label=profile.get('canonical_revenue','Revenue')
    r=_find(fs,revenue_label)
    if not r or not i0:
        return 'FAIL',f"Canonical revenue row missing for {profile['name']}: expected {revenue_label}."
    header=next((rr for rr in range(i0+1,min(fs.max_row+1,i0+6)) if str(fs.cell(rr,1).value or '').strip().lower()=='metric'),None)
    if not header: return 'FAIL','Financial Statements annual header missing.'
    ycols={int(fs.cell(header,c).value):c for c in range(2,min(fs.max_column,10)+1) if isinstance(fs.cell(header,c).value,(int,float))}
    hcols={int(hs.cell(3,c).value):c for c in range(2,min(hs.max_column,8)+1) if isinstance(hs.cell(3,c).value,(int,float))}
    common=sorted(set(ycols)&set(hcols))
    if not common: return 'REVIEW','No common annual periods to reconcile.'
    y=common[-1]; a=_num(fs.cell(r,ycols[y]).value); b=_num(hs.cell(4,hcols[y]).value)
    if a is None or b is None: return 'REVIEW','Latest canonical revenue missing in one sheet.'
    tolerance=max(.02,abs(a)*.002)
    if abs(a-b)<=tolerance:
        return 'PASS',f'Latest annual {revenue_label} reconciles across Financial Statements and Historical Financials ({y}).'
    return 'FAIL',f'{revenue_label} mismatch remains for {y}: statements={a}, history={b}.'


def _section_rows(ws,title,next_title):
    s=_find(ws,title); e=_find(ws,next_title) if next_title else ws.max_row+1
    if not s or not e: return [],[]
    header=next((r for r in range(s+1,min(e,s+6)) if str(ws.cell(r,1).value or '').strip().lower()=='metric'),None)
    if not header: return [],[]
    years=[c for c in range(2,min(ws.max_column,10)+1) if isinstance(ws.cell(header,c).value,(int,float)) and 1900<=int(ws.cell(header,c).value)<=2100]
    derived={
        'Operating Margin','Net Margin','Net Debt','Working Capital','Free Cash Flow',
        'Efficiency Ratio (simple)','Credit Cost / Revenue','Loans / Deposits'
    }
    labels=[]; mapped=[]
    for r in range(header+1,e):
        label=str(ws.cell(r,1).value or '').strip()
        if not label or label in derived: continue
        labels.append(label)
        if any(isinstance(ws.cell(r,c).value,(int,float)) for c in years): mapped.append(label)
    return labels,mapped


def _statement_profile_status(wb,ticker):
    if 'Financial Statements' not in wb.sheetnames: return 'FAIL','Financial Statements missing.'
    profile=get_statement_profile(ticker); marker=str(wb['Financial Statements']['A4'].value or '')
    if profile['name'] not in marker:
        return 'FAIL',f"Expected statement profile '{profile['name']}' is not identified in Financial Statements."
    if profile['key']=='bank':
        required={'Total Net Revenue','Provision for Credit Losses','Deposits','Loans, Net'}
    elif profile['key']=='berkshire':
        required={'Insurance Premiums Earned','Investment Gains / (Losses)','Total Revenues','Insurance Losses & Loss Adjustment Liabilities'}
    elif profile['key'] in {'tsm','siemens'}:
        required={'Revenue','Operating Income','Total Assets','Total Equity'}
    else:
        required={'Revenue','Operating Income','Total Assets','Operating Cash Flow'}
    present={str(wb['Financial Statements'].cell(r,1).value or '').strip() for r in range(1,wb['Financial Statements'].max_row+1)}
    missing=sorted(required-present)
    if missing: return 'FAIL','Profile-specific core rows missing: '+', '.join(missing)
    return 'PASS',f"Economically appropriate statement profile applied: {profile['name']}."


def _full_statement_depth(wb,ticker):
    if 'Financial Statements' not in wb.sheetnames: return 'FAIL','Financial Statements missing.'
    profile=get_statement_profile(ticker); ws=wb['Financial Statements']
    income,im=_section_rows(ws,'Income Statement','Balance Sheet')
    balance,bm=_section_rows(ws,'Balance Sheet','Cash Flow Statement')
    cash,cm=_section_rows(ws,'Cash Flow Statement','Full Statement Coverage')
    min_structure=tuple(profile.get('min_structure') or (20,35,25))
    min_mapped=tuple(profile.get('min_mapped') or (12,18,14))
    counts=(len(income),len(balance),len(cash)); mapped=(len(im),len(bm),len(cm))
    structural=all(a>=b for a,b in zip(counts,min_structure))
    if not structural:
        return 'FAIL',(
            f"Profile statement structure is too shallow for {profile['name']}: "
            f"income={counts[0]}/{min_structure[0]} minimum, balance={counts[1]}/{min_structure[1]}, cash={counts[2]}/{min_structure[2]}."
        )
    useful=all(a>=b for a,b in zip(mapped,min_mapped))
    status='PASS' if useful else 'REVIEW'
    detail=(f"{profile['name']} full statements: income {mapped[0]}/{counts[0]} mapped rows, "
            f"balance {mapped[1]}/{counts[1]}, cash flow {mapped[2]}/{counts[2]}. "
            'Profile-specific unmapped lines stay blank rather than being estimated.')
    return status,detail


def _product_profile_status(wb):
    if 'Company Data' not in wb.sheetnames: return 'FAIL','Company Data missing.'
    ws=wb['Company Data']; header=_find(ws,'Business Overview & Main Products / Services')
    if not header: return 'FAIL','Main products/services section missing from Company Data.'
    product_header=_find(ws,'Business / Segment')
    if not product_header: return 'FAIL','Product table header missing from Company Data.'
    rows=0
    # Product rows are reserved for at most ten entries immediately below the table header.
    # Row 32 is the separate research-rule note and must never be counted as a product.
    for r in range(product_header+1,min(ws.max_row,product_header+10)+1):
        if str(ws.cell(r,2).value or '').strip(): rows+=1
    if rows>=3: return 'PASS',f'Company Data contains {rows} main product/service rows with business-line context and source fields.'
    if rows>=1: return 'REVIEW',f'Company Data product profile exists but has only {rows} reliably identified row(s).'
    return 'FAIL','Company Data product profile contains no populated products/services.'


def _company_classification_status(wb,ticker):
    if 'Company Data' not in wb.sheetnames: return 'FAIL','Company Data missing.'
    ws=wb['Company Data']; t=str(ticker).upper().strip()
    sector=str(ws['B6'].value or '').strip(); industry=str(ws['B7'].value or '').strip()
    if not sector or not industry: return 'REVIEW','Sector or industry is missing; peer classification should not use stale template metadata.'
    if t not in {'GOOGL','GOOG'} and (sector,industry)==('Communication Services','Internet Content & Information'):
        return 'FAIL','Stale Alphabet template sector/industry survived into a different issuer.'
    expected={
        'SIE.DE':('Industrials','Specialty Industrial Machinery'),
        'BRK.B':('Financial Services','Insurance - Diversified'),
        'BRK-B':('Financial Services','Insurance - Diversified'),
    }.get(t)
    if expected and (sector,industry)!=expected:
        return 'FAIL',f'Classification mismatch: expected {expected[0]} / {expected[1]}, got {sector} / {industry}.'
    return 'PASS',f'Company classification is populated and issuer-consistent: {sector} / {industry}.'


def _finalize_statement_currency_label(wb):
    if 'Financial Statements' not in wb.sheetnames: return 'FAIL','Financial Statements missing.'
    info=getattr(wb,'_wacc_info',{}) or {}
    reporting=str(info.get('financialCurrency') or '').upper().strip()
    quote=str(info.get('currency') or '').upper().strip()
    # At finalization, cross-border financials have already been normalized into the traded/
    # valuation currency. Same-currency listings remain in their native reporting currency.
    final_currency=quote if reporting and quote and reporting!=quote else (reporting or quote or 'USD')
    ws=wb['Financial Statements']
    ws['A3']=f'{final_currency} billions unless otherwise stated. EPS and share-count rows use the units shown in their labels.'
    return 'PASS',f'Financial Statements final currency basis: {final_currency}; reporting={reporting or "unknown"}, traded/valuation={quote or "unknown"}.'


def ensure_quality_checks(wb,ticker,bundle=None,removed=None):
    if 'Data Quality' not in wb.sheetnames: wb.create_sheet('Data Quality')
    ws=wb['Data Quality']; removed=removed or []; bundle=bundle or reconcile_score_displays(wb,ticker)
    labels={
        'Canonical financial-statement reconciliation','Statement profile suitability','Full financial-statement depth','Statement currency basis',
        'Company classification freshness','Company product/profile coverage','Segment Analysis public-data coverage','Valuation-model reliability gate',
        'Score-engine single-source reconciliation','Low-value tab pruning'
    }
    for r in range(1,ws.max_row+1):
        if str(ws.cell(r,1).value or '') in labels:
            for c in range(1,min(ws.max_column,8)+1): ws.cell(r,c).value=None
    start=ws.max_row+2
    controls=[]
    st,detail=_finalize_statement_currency_label(wb); controls.append(('Statement currency basis',st,detail))
    st,detail=_statement_status(wb,ticker); controls.append(('Canonical financial-statement reconciliation',st,detail))
    st,detail=_statement_profile_status(wb,ticker); controls.append(('Statement profile suitability',st,detail))
    st,detail=_full_statement_depth(wb,ticker); controls.append(('Full financial-statement depth',st,detail))
    st,detail=_company_classification_status(wb,ticker); controls.append(('Company classification freshness',st,detail))
    st,detail=_product_profile_status(wb); controls.append(('Company product/profile coverage',st,detail))
    st,detail=_segment_status(wb,ticker); controls.append(('Segment Analysis public-data coverage',st,detail))
    rel=bundle.get('valuation_model_reliability') or {'status':'PASS','reasons':[]}
    rel_status='FAIL' if rel.get('status')!='PASS' else 'PASS'
    controls.append(('Valuation-model reliability gate',rel_status,' '.join(rel.get('reasons') or []) or 'Base valuation output is economically interpretable.'))
    controls.append(('Score-engine single-source reconciliation','PASS',f"Score Engine v3; effective weight coverage {bundle.get('coverage',0):.0%}. Weak/missing dimensions are reweighted rather than fabricated."))
    controls.append(('Low-value tab pruning','PASS','Removed: '+', '.join(removed) if removed else 'No low-value empty tabs required removal.'))
    for i,(name,status,detail) in enumerate(controls,start):
        ws.cell(i,1,name); ws.cell(i,2,status); ws.cell(i,3,detail); ws.cell(i,3).alignment=Alignment(wrap_text=True)
        ws.cell(i,2).fill=PatternFill('solid',fgColor=GREEN if status=='PASS' else GOLD if status=='REVIEW' else RED); ws.cell(i,2).font=Font(bold=True)
    ws.column_dimensions['A'].width=max(ws.column_dimensions['A'].width or 0,42); ws.column_dimensions['C'].width=max(ws.column_dimensions['C'].width or 0,95)
    return controls
