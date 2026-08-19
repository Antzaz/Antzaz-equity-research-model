from __future__ import annotations

"""Sync recruiter-safe investment-thesis inputs from the local Excel workbook to GitHub Actions.

The workbook stays local/private. This script reads only the explicitly public fields,
serializes them to compact JSON, and stores that JSON in the PORTFOLIO_THESIS_JSON Actions
secret. Private notes are never included.
"""

import argparse
import json
import subprocess
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_PATH = ROOT / "institutional_research" / "portfolio_thesis.xlsx"
DEFAULT_REPO = "Antzaz/Antzaz-equity-research-model"

SCORE_FIELDS = {
    "Business Quality (1-5)": "business_quality",
    "Moat (1-5)": "moat_score",
    "Management & Capital Allocation (1-5)": "management_capital_allocation",
    "Balance Sheet (1-5)": "balance_sheet",
    "Growth (1-5)": "growth",
    "Valuation (1-5)": "valuation",
    "Risk/Resilience (1-5)": "risk_resilience",
}

PUBLIC_TEXT_FIELDS = {
    "Status": "status",
    "Time Horizon": "time_horizon",
    "Investment Thesis": "investment_thesis",
    "Why I Own It": "why_owned",
    "Competitive Advantage / Moat": "competitive_advantage",
    "Growth Drivers": "growth_drivers",
    "Valuation Rationale": "valuation_rationale",
    "Catalysts": "catalysts",
    "Key Risks": "key_risks",
    "Falsification / Sell Condition": "sell_condition",
    "Monitoring KPI": "monitoring_kpi",
    "Public Notes": "public_notes",
}


def _yes(value) -> bool:
    return str(value or "").strip().lower() in {"yes", "y", "true", "1", "publish"}


def _clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def _number(value):
    try:
        if value in (None, ""):
            return None
        if isinstance(value, str):
            text = value.strip()
            if text.endswith("%"):
                return float(text[:-1].strip()) / 100.0
        return float(value)
    except Exception:
        return None


def _read_company_theses(wb) -> list[dict]:
    if "Company Theses" not in wb.sheetnames:
        raise ValueError("Workbook is missing the 'Company Theses' sheet.")
    ws = wb["Company Theses"]
    headers = {str(cell.value).strip(): idx + 1 for idx, cell in enumerate(ws[3]) if cell.value}
    required = {"Publish", "Ticker", "Company"}
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError("Company Theses is missing required columns: " + ", ".join(missing))

    rows: list[dict] = []
    for row_idx in range(4, ws.max_row + 1):
        ticker = _clean(ws.cell(row_idx, headers["Ticker"]).value)
        if not ticker:
            continue
        if not _yes(ws.cell(row_idx, headers["Publish"]).value):
            continue

        record: dict = {
            "ticker": str(ticker).upper(),
            "company": _clean(ws.cell(row_idx, headers["Company"]).value),
        }
        for excel_name, json_name in PUBLIC_TEXT_FIELDS.items():
            if excel_name in headers:
                record[json_name] = _clean(ws.cell(row_idx, headers[excel_name]).value)

        if "Conviction (1-5)" in headers:
            record["conviction"] = _number(ws.cell(row_idx, headers["Conviction (1-5)"]).value)
        if "Expected Annual Return %" in headers:
            record["expected_annual_return"] = _number(
                ws.cell(row_idx, headers["Expected Annual Return %"]).value
            )
        if "Review Date" in headers:
            value = ws.cell(row_idx, headers["Review Date"]).value
            if value is not None:
                record["review_date"] = value.isoformat() if hasattr(value, "isoformat") else str(value)

        scores: dict[str, float] = {}
        for excel_name, json_name in SCORE_FIELDS.items():
            if excel_name in headers:
                val = _number(ws.cell(row_idx, headers[excel_name]).value)
                if val is not None:
                    if not 1 <= val <= 5:
                        raise ValueError(f"{ticker}: {excel_name} must be between 1 and 5.")
                    scores[json_name] = val
        record["scores"] = scores
        record["composite_score"] = (
            sum(scores.values()) / len(scores) if scores else None
        )

        # Do not publish empty thesis rows just because Publish defaults to Yes.
        meaningful = any(
            record.get(k)
            for k in [
                "investment_thesis",
                "why_owned",
                "competitive_advantage",
                "growth_drivers",
                "valuation_rationale",
                "key_risks",
                "sell_condition",
            ]
        ) or bool(scores)
        if meaningful:
            rows.append(record)
    return rows


def _read_philosophy(wb) -> dict:
    if "Portfolio Philosophy" not in wb.sheetnames:
        return {}
    ws = wb["Portfolio Philosophy"]
    fields: dict[str, object] = {}
    publish = False
    for row_idx in range(4, ws.max_row + 1):
        field = _clean(ws.cell(row_idx, 1).value)
        value = _clean(ws.cell(row_idx, 2).value)
        if not field:
            continue
        if str(field).strip().lower() == "publish":
            publish = _yes(value)
            continue
        if value is not None:
            key = str(field).strip().lower().replace("&", "and").replace("/", "_")
            key = "_".join(key.split())
            fields[key] = value
    return fields if publish else {}


def build_payload(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    return {
        "version": 1,
        "portfolio_philosophy": _read_philosophy(wb),
        "company_theses": _read_company_theses(wb),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", default=str(DEFAULT_PATH))
    parser.add_argument("--repo", default=DEFAULT_REPO)
    args = parser.parse_args()

    path = Path(args.file).expanduser().resolve()
    if not path.exists():
        raise SystemExit(f"Thesis workbook not found: {path}")

    payload = build_payload(path)
    compact = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    size_bytes = len(compact.encode("utf-8"))
    if size_bytes > 44_000:
        raise SystemExit(
            "Public thesis payload is too large for the current GitHub-secret workflow. "
            "Shorten recruiter-facing thesis text or publish fewer rows."
        )

    subprocess.run(
        ["gh", "secret", "set", "PORTFOLIO_THESIS_JSON", "--repo", args.repo],
        input=compact,
        text=True,
        check=True,
    )
    theses = payload["company_theses"]
    expected_count = sum(1 for item in theses if item.get("expected_annual_return") is not None)
    print("Recruiter-safe investment thesis inputs synced successfully.")
    print(f"Published company theses: {len(theses)}")
    print(f"Expected annual returns populated: {expected_count} / {len(theses)}")
    if expected_count == 0 and theses:
        print("WARNING: No Expected Annual Return values were read from the workbook. Enter percentages such as 12% in the 'Expected Annual Return %' column and save the workbook before syncing.")
    print(f"Portfolio philosophy fields: {len(payload['portfolio_philosophy'])}")
    print("Private Notes were not included in the GitHub secret.")


if __name__ == "__main__":
    main()
