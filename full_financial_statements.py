from __future__ import annotations

"""Expand the model's Financial Statements sheet from a condensed summary into a standardized
three-statement view while preserving canonical issuer/SEC core values.

Priority per cell:
1) already-reconciled values from the existing Financial Statements sheet;
2) SEC Company Facts for mapped US-GAAP tags;
3) Yahoo Finance annual statement rows as a structured fallback.

The expanded sheet intentionally leaves unsupported company-specific lines blank instead of
inventing values. A Source / Definition column makes the lineage visible to the analyst.
"""

import math
from collections import defaultdict

import yfinance as yf
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; GREY="666666"; LIGHT="F5F9FC"
FMT_BN='#,##0.0;[Red](#,##0.0);-'; FMT_PCT='0.0%;[Red](0.0%);-'; FMT_EPS='$0.00;[Red]($0.00);-'; FMT_SHARES='#,##0.0;[Red](#,##0.0);-'
THIN=Side(style="thin",color="808080")


def _fill(c): return PatternFill("solid",fgColor=c)
def _num(v):
    try:
        x=float(v); return x if math.isfinite(x) else None
    except Exception: return None

def _find(ws,label,start=1,end=None):
    needle=str(label).strip().lower(); end=end or ws.max_row
    for r in range(start,min(end,ws.max_row)+1):
        if str(ws.cell(r,1).value or "").strip().lower()==needle: return r
    return None

def _section_bounds(ws,title,next_title=None):
    s=_find(ws,title)
    if not s: return None,None
    e=_find(ws,next_title,s+1) if next_title else ws.max_row+1
    return s,(e or ws.max_row+1)

def _year_cols(ws,header_row):
    out={}
    for c in range(2,min(ws.max_column,12)+1):
        v=ws.cell(header_row,c).value
        if isinstance(v,(int,float)) and 1900<=int(v)<=2100: out[int(v)]=c
    return out


def _capture_existing(wb):
    out={"income":defaultdict(dict),"balance":defaultdict(dict),"cash":defaultdict(dict)}
    if "Financial Statements" not in wb.sheetnames: return out
    ws=wb["Financial Statements"]
    specs=[("income","Income Statement","Balance Sheet"),("balance","Balance Sheet","Cash Flow Statement"),("cash","Cash Flow Statement","Financial Statement Integrity & Source Reconciliation")]
    for key,title,nxt in specs:
        s,e=_section_bounds(ws,title,nxt)
        if not s: continue
        header=next((r for r in range(s+1,min(e,s+5)) if str(ws.cell(r,1).value or "").strip().lower()=="metric"),None)
        if not header: continue
        yc=_year_cols(ws,header)
        for r in range(header+1,e):
            label=str(ws.cell(r,1).value or "").strip()
            if not label: continue
            for y,c in yc.items():
                v=_num(ws.cell(r,c).value)
                if v is not None: out[key][label][y]=v
    return out


def _sec_series(facts,tags,unit="USD"):
    if not facts: return {}
    gaap=(facts.get("facts") or {}).get("us-gaap") or {}; best={}
    for priority,tag in enumerate(tags):
        fact=gaap.get(tag)
        if not fact: continue
        units=fact.get("units") or {}; values=units.get(unit)
        if not values and units: values=next(iter(units.values()))
        for x in values or []:
            if x.get("form") not in {"10-K","10-K/A"}: continue
            fy=x.get("fy"); val=x.get("val")
            try: fy=int(fy); val=float(val)
            except Exception: continue
            stamp=(str(x.get("filed") or ""),str(x.get("end") or ""),-priority)
            if fy not in best or stamp>best[fy][0]: best[fy]=(stamp,val)
    return {y:v for y,(_,v) in best.items()}


def _yf_frames(ticker):
    try:
        t=yf.Ticker(ticker)
        return {"income":t.income_stmt,"balance":t.balance_sheet,"cash":t.cashflow}
    except Exception:
        return {"income":None,"balance":None,"cash":None}


def _yf_series(df,candidates):
    if df is None or getattr(df,"empty",True): return {}
    row=None
    for name in candidates:
        if name in df.index:
            row=df.loc[name]; break
    if row is None: return {}
    out={}
    for col,val in row.items():
        try: y=int(getattr(col,"year",str(col)[:4])); v=float(val)
        except Exception: continue
        if math.isfinite(v): out[y]=v
    return out


def _get(existing,key,label,years,sec,yf,unit="money",aliases=()):
    values={}; sources={}
    old={}
    for nm in (label,)+tuple(aliases): old.update(existing[key].get(nm,{}) or {})
    scale=1.0 if unit=="eps" else 1e9
    for y in years:
        if y in old:
            values[y]=old[y]; sources[y]="Canonical issuer/SEC-reconciled"
        elif y in sec:
            values[y]=sec[y] if unit=="eps" else sec[y]/scale; sources[y]="SEC Company Facts"
        elif y in yf:
            values[y]=yf[y] if unit=="eps" else yf[y]/scale; sources[y]="Yahoo annual statement fallback"
    return values,sources


INCOME=[
    ("Revenue",["Total Revenue","Operating Revenue"],["RevenueFromContractWithCustomerExcludingAssessedTax","SalesRevenueNet","Revenues"],"money",()),
    ("Cost of Revenue",["Cost Of Revenue"],["CostOfRevenue","CostOfGoodsAndServicesSold"],"money",()),
    ("Gross Profit",["Gross Profit"],["GrossProfit"],"money",()),
    ("Research & Development",["Research And Development"],["ResearchAndDevelopmentExpense"],"money",()),
    ("Selling, General & Administrative",["Selling General And Administration"],["SellingGeneralAndAdministrativeExpense"],"money",("General & Administrative","Sales & Marketing")),
    ("Other Operating Expenses",["Other Operating Expenses"],["OtherOperatingExpenses"],"money",()),
    ("Total Operating Expenses",["Operating Expense"],["OperatingExpenses"],"money",()),
    ("Operating Income",["Operating Income"],["OperatingIncomeLoss"],"money",()),
    ("Interest Income",["Interest Income Non Operating"],["InterestIncomeExpenseNonoperatingNet","InvestmentIncomeInterest"],"money",()),
    ("Interest Expense",["Interest Expense Non Operating"],["InterestExpenseNonOperating","InterestExpense"],"money",()),
    ("Net Interest Income / (Expense)",["Net Non Operating Interest Income Expense"],["InterestIncomeExpenseNonoperatingNet"],"money",()),
    ("Other Non-Operating Income / (Expense)",["Other Non Operating Income Expenses","Total Other Finance Cost"],["NonoperatingIncomeExpense","OtherNonoperatingIncomeExpense"],"money",("Other Income / (Expense), Net",)),
    ("Pre-Tax Income",["Pretax Income"],["IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest","IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments"],"money",()),
    ("Income Taxes",["Tax Provision"],["IncomeTaxExpenseBenefit"],"money",()),
    ("Income After Tax",["Net Income Continuous Operations"],["IncomeLossFromContinuingOperations"],"money",()),
    ("Minority / Noncontrolling Interests",["Net Income Including Noncontrolling Interests"],["NetIncomeLossAttributableToNoncontrollingInterest"],"money",()),
    ("Equity in Affiliates",["Otherunder Preferred Stock Dividend"],["IncomeLossFromEquityMethodInvestments"],"money",()),
    ("Net Income",["Net Income","Net Income Common Stockholders"],["NetIncomeLoss","ProfitLoss"],"money",()),
    ("Net Income Available to Common",["Net Income Common Stockholders"],["NetIncomeLossAvailableToCommonStockholdersBasic"],"money",()),
    ("Basic EPS",["Basic EPS"],["EarningsPerShareBasic"],"eps",()),
    ("Diluted EPS",["Diluted EPS"],["EarningsPerShareDiluted"],"eps",()),
    ("Basic Weighted Average Shares",["Basic Average Shares"],["WeightedAverageNumberOfSharesOutstandingBasic"],"shares",()),
    ("Diluted Weighted Average Shares",["Diluted Average Shares"],["WeightedAverageNumberOfDilutedSharesOutstanding"],"shares",()),
]

BALANCE=[
    ("Cash & Cash Equivalents",["Cash And Cash Equivalents"],["CashAndCashEquivalentsAtCarryingValue"],"money",("Cash, Cash Equivalents & Restricted Cash",)),
    ("Short-Term Investments / Marketable Securities",["Other Short Term Investments"],["MarketableSecuritiesCurrent","ShortTermInvestments"],"money",("Marketable / Short-Term Securities",)),
    ("Cash + Short-Term Investments",["Cash Cash Equivalents And Short Term Investments"],["CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents"],"money",()),
    ("Accounts Receivable",["Accounts Receivable"],["AccountsReceivableNetCurrent"],"money",()),
    ("Inventory",["Inventory"],["InventoryNet"],"money",()),
    ("Other Current Assets",["Other Current Assets"],["OtherCurrentAssets"],"money",()),
    ("Total Current Assets",["Current Assets"],["AssetsCurrent"],"money",()),
    ("Gross Property, Plant & Equipment",["Gross PPE"],["PropertyPlantAndEquipmentGross"],"money",()),
    ("Accumulated Depreciation",["Accumulated Depreciation"],["AccumulatedDepreciationDepletionAndAmortizationPropertyPlantAndEquipment"],"money",()),
    ("Property & Equipment, Net",["Net PPE"],["PropertyPlantAndEquipmentNet"],"money",()),
    ("Operating Lease Assets",["Operating Lease Right Of Use Asset"],["OperatingLeaseRightOfUseAsset"],"money",()),
    ("Goodwill",["Goodwill"],["Goodwill"],"money",()),
    ("Other Intangible Assets",["Other Intangible Assets"],["FiniteLivedIntangibleAssetsNet","IndefiniteLivedIntangibleAssetsExcludingGoodwill"],"money",()),
    ("Investments & Other Financial Assets",["Investments And Other Financial Assets"],["LongTermInvestments"],"money",()),
    ("Other Non-Current Assets",["Other Non Current Assets"],["OtherAssetsNoncurrent"],"money",()),
    ("Total Non-Current Assets",["Total Non Current Assets"],["AssetsNoncurrent"],"money",()),
    ("Total Assets",["Total Assets"],["Assets"],"money",()),
    ("Accounts Payable",["Accounts Payable"],["AccountsPayableCurrent"],"money",()),
    ("Accrued Expenses",["Payables And Accrued Expenses"],["AccountsPayableAndOtherAccruedLiabilitiesCurrent"],"money",()),
    ("Deferred Revenue / Contract Liabilities",["Current Deferred Revenue"],["ContractWithCustomerLiabilityCurrent","DeferredRevenueCurrent"],"money",("Deferred Revenue",)),
    ("Short-Term Debt",["Current Debt"],["ShortTermBorrowings"],"money",()),
    ("Current Portion Long-Term Debt",["Current Debt And Capital Lease Obligation"],["LongTermDebtCurrent"],"money",()),
    ("Other Current Liabilities",["Other Current Liabilities"],["OtherLiabilitiesCurrent"],"money",()),
    ("Total Current Liabilities",["Current Liabilities"],["LiabilitiesCurrent"],"money",()),
    ("Long-Term Debt",["Long Term Debt"],["LongTermDebtNoncurrent","LongTermDebt"],"money",()),
    ("Operating Lease Liabilities",["Long Term Capital Lease Obligation"],["OperatingLeaseLiabilityNoncurrent"],"money",()),
    ("Deferred Tax Liabilities",["Trade And Other Payables Non Current"],["DeferredTaxLiabilitiesNoncurrent"],"money",()),
    ("Pension / Post-Retirement Liabilities",["Other Non Current Liabilities"],["PensionAndOtherPostretirementBenefitsLiabilitiesNoncurrent"],"money",()),
    ("Other Non-Current Liabilities",["Other Non Current Liabilities"],["OtherLiabilitiesNoncurrent"],"money",()),
    ("Total Non-Current Liabilities",["Total Non Current Liabilities Net Minority Interest"],["LiabilitiesNoncurrent"],"money",()),
    ("Total Liabilities",["Total Liabilities Net Minority Interest"],["Liabilities"],"money",()),
    ("Preferred Stock",["Preferred Stock"],["PreferredStocksIncludingAdditionalPaidInCapital"],"money",()),
    ("Common Stock",["Common Stock"],["CommonStocksIncludingAdditionalPaidInCapital"],"money",()),
    ("Additional Paid-In Capital",["Additional Paid In Capital"],["AdditionalPaidInCapital"],"money",()),
    ("Retained Earnings",["Retained Earnings"],["RetainedEarningsAccumulatedDeficit"],"money",()),
    ("Accumulated Other Comprehensive Income",["Gains Losses Not Affecting Retained Earnings"],["AccumulatedOtherComprehensiveIncomeLossNetOfTax"],"money",()),
    ("Treasury Stock",["Treasury Stock"],["TreasuryStockValue"],"money",()),
    ("Noncontrolling Interest",["Minority Interest"],["MinorityInterest"],"money",()),
    ("Stockholders' Equity",["Stockholders Equity"],["StockholdersEquity","StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest"],"money",()),
    ("Total Equity",["Total Equity Gross Minority Interest"],["StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest"],"money",()),
    ("Total Liabilities & Equity",["Total Liabilities Net Minority Interest"],["LiabilitiesAndStockholdersEquity"],"money",()),
]

CASH=[
    ("Net Income",["Net Income From Continuing Operations","Net Income"],["NetIncomeLoss","ProfitLoss"],"money",()),
    ("Depreciation, Amortization & Accretion",["Depreciation Amortization Depletion","Depreciation And Amortization"],["DepreciationDepletionAndAmortization","DepreciationDepletionAndAmortizationPropertyPlantAndEquipment"],"money",()),
    ("Stock-Based Compensation",["Stock Based Compensation"],["ShareBasedCompensation"],"money",()),
    ("Deferred Income Taxes",["Deferred Tax"],["DeferredIncomeTaxExpenseBenefit"],"money",()),
    ("Asset Impairments",["Impairment Of Capital Assets"],["AssetImpairmentCharges"],"money",()),
    ("Unrealized / Fair-Value Gains & Losses",["Operating Gains Losses"],["GainLossOnInvestments"],"money",()),
    ("Other Non-Cash Items",["Other Non Cash Items"],["OtherNoncashIncomeExpense"],"money",()),
    ("Change in Accounts Receivable",["Change In Accounts Receivable"],["IncreaseDecreaseInAccountsReceivable"],"money",()),
    ("Change in Inventory",["Change In Inventory"],["IncreaseDecreaseInInventories"],"money",()),
    ("Change in Accounts Payable / Accrued Liabilities",["Change In Payables And Accrued Expense","Change In Account Payable"],["IncreaseDecreaseInAccountsPayableAndAccruedLiabilities","IncreaseDecreaseInAccountsPayable"],"money",()),
    ("Change in Deferred Revenue / Contract Liabilities",["Change In Other Current Liabilities"],["IncreaseDecreaseInContractWithCustomerLiability"],"money",()),
    ("Change in Other Working Capital",["Change In Other Working Capital"],["IncreaseDecreaseInOtherOperatingAssets","IncreaseDecreaseInOtherOperatingLiabilities"],"money",()),
    ("Change in Working Capital",["Change In Working Capital"],["IncreaseDecreaseInOperatingAssets"],"money",()),
    ("Operating Cash Flow",["Operating Cash Flow"],["NetCashProvidedByUsedInOperatingActivities"],"money",()),
    ("Capital Expenditures",["Capital Expenditure","Purchase Of PPE"],["PaymentsToAcquirePropertyPlantAndEquipment","PaymentsToAcquireProductiveAssets"],"money",()),
    ("Acquisitions",["Net Business Purchases","Purchase Of Business"],["PaymentsToAcquireBusinessesNetOfCashAcquired","PaymentsToAcquireBusinessesNetOfCashAndCashEquivalentsAcquired"],"money",()),
    ("Purchases of Investments",["Investment Purchases"],["PaymentsToAcquireInvestments"],"money",()),
    ("Proceeds from Investment Sales / Maturities",["Investment Sales"],["ProceedsFromSaleMaturityAndPrepaymentOfInvestments"],"money",()),
    ("Other Investing Activities",["Net Other Investing Changes"],["PaymentsForProceedsFromOtherInvestingActivities"],"money",()),
    ("Investing Cash Flow",["Investing Cash Flow"],["NetCashProvidedByUsedInInvestingActivities"],"money",()),
    ("Debt Issuance",["Issuance Of Debt","Long Term Debt Issuance"],["ProceedsFromIssuanceOfLongTermDebt","ProceedsFromIssuanceOfDebt"],"money",()),
    ("Debt Repayments",["Repayment Of Debt","Long Term Debt Payments"],["RepaymentsOfLongTermDebt","RepaymentsOfDebt"],"money",()),
    ("Common Stock Issuance",["Issuance Of Capital Stock","Common Stock Issuance"],["ProceedsFromStockOptionsExercised","ProceedsFromIssuanceOfCommonStock"],"money",()),
    ("Share Repurchases",["Repurchase Of Capital Stock","Common Stock Payments"],["PaymentsForRepurchaseOfCommonStock"],"money",()),
    ("Dividends",["Cash Dividends Paid"],["PaymentsOfDividends"],"money",()),
    ("Other Financing Activities",["Net Other Financing Charges"],["ProceedsFromPaymentsForOtherFinancingActivities"],"money",()),
    ("Financing Cash Flow",["Financing Cash Flow"],["NetCashProvidedByUsedInFinancingActivities"],"money",()),
    ("Effect of FX on Cash",["Effect Of Exchange Rate Changes"],["EffectOfExchangeRateOnCashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents"],"money",()),
    ("Net Change in Cash",["Changes In Cash"],["CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect"],"money",()),
    ("Beginning Cash",["Beginning Cash Position"],["CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents"],"money",()),
    ("Ending Cash",["End Cash Position","Cash And Cash Equivalents"],["CashAndCashEquivalentsAtCarryingValue"],"money",()),
]

OUTFLOW_LABELS={"Capital Expenditures","Acquisitions","Purchases of Investments","Debt Repayments","Share Repurchases","Dividends"}
TOTAL_LABELS={"Revenue","Gross Profit","Total Operating Expenses","Operating Income","Pre-Tax Income","Net Income","Total Current Assets","Total Non-Current Assets","Total Assets","Total Current Liabilities","Total Non-Current Liabilities","Total Liabilities","Stockholders' Equity","Total Equity","Total Liabilities & Equity","Operating Cash Flow","Investing Cash Flow","Financing Cash Flow","Net Change in Cash","Ending Cash","Free Cash Flow"}


def _series(existing,frames,facts,key,label,yf_names,sec_tags,unit,aliases,years):
    sec=_sec_series(facts,sec_tags,"USD/shares" if unit=="eps" else "shares" if unit=="shares" else "USD")
    yfs=_yf_series(frames.get(key),yf_names)
    vals,src=_get(existing,key,label,years,sec,yfs,unit,aliases)
    if label in OUTFLOW_LABELS:
        vals={y:-abs(v) for y,v in vals.items()}
    return vals,src


def _available_years(existing,frames,facts):
    years=set()
    for label,vals in existing["income"].items(): years.update(vals)
    df=frames.get("income")
    if df is not None and not getattr(df,"empty",True):
        for c in df.columns:
            try: years.add(int(getattr(c,"year",str(c)[:4])))
            except Exception: pass
    years.update(_sec_series(facts,["RevenueFromContractWithCustomerExcludingAssessedTax","SalesRevenueNet","Revenues"]).keys())
    completed=max(existing["income"].get("Revenue",{}).keys() or years or [0])
    valid=sorted(y for y in years if 2000<=y<=completed)
    return valid[-6:]


def _title(ws,ticker,last_col):
    ws.merge_cells(start_row=1,start_column=1,end_row=2,end_column=last_col)
    ws.cell(1,1,f"{ticker} — Full Financial Statements"); ws.cell(1,1).fill=_fill(NAVY); ws.cell(1,1).font=Font(bold=True,color=WHITE,size=18); ws.cell(1,1).alignment=Alignment(vertical="center")
    ws.merge_cells(start_row=3,start_column=1,end_row=3,end_column=last_col)
    ws.cell(3,1,"USD billions unless per-share/share-count data. Canonical issuer/SEC values are preserved; SEC Company Facts is preferred for mapped lines and Yahoo annual statements fill unsupported standardized lines. Blank means not reliably mapped, not zero.")
    ws.cell(3,1).font=Font(italic=True,color=GREY); ws.cell(3,1).alignment=Alignment(wrap_text=True)


def _write_section(ws,start,title,rows,years,existing,frames,facts,key):
    source_col=2+len(years)
    ws.merge_cells(start_row=start,start_column=1,end_row=start,end_column=source_col)
    ws.cell(start,1,title); ws.cell(start,1).fill=_fill(NAVY); ws.cell(start,1).font=Font(bold=True,color=WHITE,size=12)
    hr=start+1; ws.cell(hr,1,"Metric")
    for j,y in enumerate(years,2): ws.cell(hr,j,y)
    ws.cell(hr,source_col,"Source / Definition")
    for c in range(1,source_col+1): ws.cell(hr,c).fill=_fill(BLUE); ws.cell(hr,c).font=Font(bold=True,color=WHITE); ws.cell(hr,c).alignment=Alignment(horizontal="center",wrap_text=True)
    row=hr+1; row_map={}
    for label,yf_names,sec_tags,unit,aliases in rows:
        row_map[label]=row; ws.cell(row,1,label)
        vals,sources=_series(existing,frames,facts,key,label,yf_names,sec_tags,unit,aliases,years)
        for j,y in enumerate(years,2):
            v=vals.get(y); ws.cell(row,j,v)
            ws.cell(row,j).number_format=FMT_EPS if unit=="eps" else FMT_SHARES if unit=="shares" else FMT_BN
        uniq=[]
        for y in years:
            s=sources.get(y)
            if s and s not in uniq: uniq.append(s)
        ws.cell(row,source_col," → ".join(uniq) if uniq else "Not reliably mapped")
        ws.cell(row,source_col).alignment=Alignment(wrap_text=True,vertical="top")
        if label in TOTAL_LABELS:
            ws.cell(row,1).font=Font(bold=True)
            for c in range(1,source_col+1): ws.cell(row,c).border=Border(top=THIN)
        row+=1
    return row,row_map,source_col


def expand_financial_statements(wb,ticker,facts=None):
    existing=_capture_existing(wb); frames=_yf_frames(ticker); years=_available_years(existing,frames,facts)
    if not years:
        return {"expanded":False,"income_rows":0,"balance_rows":0,"cash_rows":0,"years":[]}
    idx=wb.sheetnames.index("Financial Statements") if "Financial Statements" in wb.sheetnames else len(wb.sheetnames)
    if "Financial Statements" in wb.sheetnames: wb.remove(wb["Financial Statements"])
    ws=wb.create_sheet("Financial Statements",idx); ws.sheet_view.showGridLines=False
    last_col=2+len(years); _title(ws,ticker,last_col)

    row,imap,source_col=_write_section(ws,5,"Income Statement",INCOME,years,existing,frames,facts,"income")
    # Derived margins remain formulas and therefore never replace reported statement lines.
    for label,numer in (("Operating Margin","Operating Income"),("Net Margin","Net Income")):
        ws.cell(row,1,label); ws.cell(row,source_col,"Derived: selected income ÷ revenue")
        for j,y in enumerate(years,2):
            c=get_column_letter(j); ws.cell(row,j,f'=IFERROR({c}{imap[numer]}/{c}{imap["Revenue"]},"")'); ws.cell(row,j).number_format=FMT_PCT
        row+=1

    row+=2; row,bmap,source_col=_write_section(ws,row,"Balance Sheet",BALANCE,years,existing,frames,facts,"balance")
    for label,formula in (("Net Debt",("Long-Term Debt","Cash & Cash Equivalents")),("Working Capital",("Total Current Assets","Total Current Liabilities"))):
        ws.cell(row,1,label); ws.cell(row,source_col,"Derived balance-sheet metric")
        a,b=formula
        for j,y in enumerate(years,2):
            c=get_column_letter(j); ws.cell(row,j,f'=IFERROR({c}{bmap[a]}-{c}{bmap[b]},"")'); ws.cell(row,j).number_format=FMT_BN
        row+=1

    row+=2; row,cmap,source_col=_write_section(ws,row,"Cash Flow Statement",CASH,years,existing,frames,facts,"cash")
    ws.cell(row,1,"Free Cash Flow"); ws.cell(row,1).font=Font(bold=True); ws.cell(row,source_col,"Derived: Operating Cash Flow + Capital Expenditures")
    for j,y in enumerate(years,2):
        c=get_column_letter(j); ws.cell(row,j,f'=IFERROR({c}{cmap["Operating Cash Flow"]}+{c}{cmap["Capital Expenditures"]},"")'); ws.cell(row,j).number_format=FMT_BN
    for c in range(1,source_col+1): ws.cell(row,c).border=Border(top=THIN)
    row+=3

    # Visible completeness control. This is not a substitute for accounting reconciliation.
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=source_col); ws.cell(row,1,"Full Statement Coverage"); ws.cell(row,1).fill=_fill(NAVY); ws.cell(row,1).font=Font(bold=True,color=WHITE)
    row+=1; ws.cell(row,1,"Section"); ws.cell(row,2,"Mapped rows"); ws.cell(row,3,"Total standardized rows"); ws.cell(row,4,"Coverage")
    for c in range(1,5): ws.cell(row,c).fill=_fill(BLUE); ws.cell(row,c).font=Font(bold=True,color=WHITE)
    stats={}
    for key,name,row_map in (("income","Income Statement",imap),("balance","Balance Sheet",bmap),("cash","Cash Flow Statement",cmap)):
        mapped=0
        for label,r in row_map.items():
            if any(_num(ws.cell(r,j).value) is not None for j in range(2,2+len(years))): mapped+=1
        stats[key]=(mapped,len(row_map)); row+=1; ws.cell(row,1,name); ws.cell(row,2,mapped); ws.cell(row,3,len(row_map)); ws.cell(row,4,mapped/len(row_map) if row_map else 0); ws.cell(row,4).number_format=FMT_PCT

    ws.column_dimensions["A"].width=44
    for c in range(2,2+len(years)): ws.column_dimensions[get_column_letter(c)].width=14
    ws.column_dimensions[get_column_letter(source_col)].width=46
    for r in range(1,ws.max_row+1): ws.cell(r,source_col).alignment=Alignment(wrap_text=True,vertical="top")
    ws.freeze_panes="B7"
    wb.calculation.fullCalcOnLoad=True
    return {"expanded":True,"income_rows":stats["income"][0],"income_total":stats["income"][1],"balance_rows":stats["balance"][0],"balance_total":stats["balance"][1],"cash_rows":stats["cash"][0],"cash_total":stats["cash"][1],"years":years}
