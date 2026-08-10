from __future__ import annotations

"""People, leadership, market-position and workbook de-duplication extensions.

The module keeps facts, calculated proxies and analyst inference separate. It does not
invent employee-happiness or market-share data when a comparable source is unavailable.
"""

import math
import yfinance as yf
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from market_context import business_market_share_records, market_share_record
from source_registry import SPECIALIST_MARKET_SOURCES, issuer_sources

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; GOLD="FFF2CC"
PALE_GREEN="E2F0D9"; GREY="666666"; LINK_GREEN="008000"
THIN=Side(style="thin",color="D9E1F2")
FMT_PCT='0.0%;[Red](0.0%);-'; FMT_SCORE='0.0'


def _fill(c): return PatternFill("solid",fgColor=c)

def _num(v,default=None):
    try:
        if isinstance(v,bool) or v in (None,""): return default
        x=float(v)
        return x if math.isfinite(x) else default
    except Exception: return default

def _section(ws,row,title,end=10):
    for c in range(1,end+1):
        ws.cell(row,c).fill=_fill(NAVY); ws.cell(row,c).font=Font(bold=True,color=WHITE,size=11)
    ws.cell(row,1,title)

def _header(ws,row,start,end):
    for c in range(start,end+1):
        x=ws.cell(row,c); x.fill=_fill(BLUE); x.font=Font(bold=True,color=WHITE)
        x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); x.border=Border(bottom=THIN)

def _safe_info(ticker,info=None):
    if info: return info
    try: return yf.Ticker(ticker).info or {}
    except Exception: return {}

def _sources(ticker,info):
    return issuer_sources(ticker,(info or {}).get("website"))

def _employee_signal(ticker,info):
    src=_sources(ticker,info)
    if ticker.upper()=="TSM":
        return {
            "score":96.0,
            "scope":"Global Inclusive Workplace learning-program satisfaction; not a company-wide engagement score",
            "period":"2024-2025 program",
            "source":src.get("workplace") or src.get("sustainability"),
            "status":"SCOPE-LIMITED",
            "evidence":"TSMC reports >100,500 participants and an average satisfaction score of 96 points for the inclusion-learning program.",
        }
    source=src.get("workplace_governance") or src.get("additional_information") or src.get("sustainability") or src.get("annual_reports") or src.get("investor")
    return {
        "score":None,
        "scope":"No comparable company-wide employee happiness/engagement score was automatically verified",
        "period":None,
        "source":source,
        "status":"REVIEW",
        "evidence":"Use issuer human-capital, sustainability and annual-report disclosures first; employee-review platforms are optional manual corroboration.",
    }

def _history_metrics(wb):
    if "Historical Financials" not in wb.sheetnames: return {}
    h=wb["Historical Financials"]; pts=[]
    for c in range(2,8):
        y=h.cell(3,c).value; rev=_num(h.cell(4,c).value)
        if isinstance(y,(int,float)) and rev and rev>0: pts.append((int(y),rev))
    cagr=None
    if len(pts)>=2:
        y0,v0=pts[0]; y1,v1=pts[-1]; n=max(1,y1-y0)
        if v0>0 and v1>0: cagr=(v1/v0)**(1/n)-1
    rev=_num(h["G4"].value); op=_num(h["G9"].value); ocf=_num(h["G14"].value); cap=_num(h["G15"].value)
    return {
        "revenue_cagr":cagr,
        "operating_margin":op/rev if rev and op is not None else None,
        "fcf_margin":(ocf-cap)/rev if rev and ocf is not None and cap is not None else None,
    }

def _leadership_proxy(wb,ticker,info,employee):
    hist=_history_metrics(wb); cagr=hist.get("revenue_cagr"); opm=hist.get("operating_margin"); fcfm=hist.get("fcf_margin")
    d=wb["Company Data"] if "Company Data" in wb.sheetnames else None
    net_debt=_num(d["B14"].value) if d else None
    execution=50.0
    if cagr is not None and opm is not None:
        execution=max(0,min(100,50*min(1,max(0,cagr)/.20)+50*min(1,max(0,opm)/.40)))
    capital=50.0
    if fcfm is not None:
        capital=35+45*min(1,max(0,fcfm)/.25)
        if net_debt is not None and net_debt<0: capital+=20
        capital=min(100,capital)
    officers=info.get("companyOfficers") or []
    depth=min(100,35+8*len([x for x in officers if isinstance(x,dict) and x.get("name")]))
    src=_sources(ticker,info)
    if src.get("leadership") or src.get("governance"): depth=max(depth,65)
    culture=_num(employee.get("score"),50.0)
    governance=75.0 if src.get("governance") else (60.0 if src.get("investor") else 50.0)
    score=.30*execution+.25*capital+.15*depth+.15*culture+.15*governance
    rows=[
        ("Execution track record",execution,"Revenue growth and operating-margin history; performance proxy, not direct causality"),
        ("Capital allocation / cash generation",capital,"FCF margin and net cash/debt profile"),
        ("Leadership depth / disclosure",depth,"Public officer depth plus issuer leadership/governance disclosure"),
        ("Employee / culture signal",culture,employee.get("scope")),
        ("Governance disclosure",governance,"Availability of official governance / investor disclosure"),
    ]
    return score,rows

def _pct_rank(values,value,higher=True):
    vals=sorted(v for v in values if v is not None and math.isfinite(v))
    if value is None or len(vals)<2: return None
    below=sum(v<value for v in vals); equal=sum(v==value for v in vals)
    pct=(below+.5*max(1,equal-1))/(len(vals)-1)*100
    pct=max(0,min(100,pct))
    return pct if higher else 100-pct

def _peer_rows(wb):
    if "Peer Comps" not in wb.sheetnames: return []
    ws=wb["Peer Comps"]; rows=[]
    for r in range(4,min(ws.max_row,20)+1):
        ticker=str(ws.cell(r,2).value or "").strip().upper()
        if not ticker or ticker.startswith("REVIEW"): continue
        rows.append({
            "ticker":ticker,"company":ws.cell(r,1).value,
            "pe":_num(ws.cell(r,3).value),"ev_ebitda":_num(ws.cell(r,5).value),
            "growth":_num(ws.cell(r,6).value),"margin":_num(ws.cell(r,7).value),
            "roe":_num(ws.cell(r,8).value),
            "market_share":_num(ws.cell(r,13).value) if ws.max_column>=13 else None,
            "peer_set_weight":_num(ws.cell(r,14).value) if ws.max_column>=14 else None,
        })
    return rows

def _alternative_screen(wb,target_ticker):
    rows=_peer_rows(wb)
    if len(rows)<2: return {"result":"No validated peer set available","candidate":None}
    share_ok=sum(x.get("market_share") is not None for x in rows)>=2
    specs=[
        ("growth",True,.22 if share_ok else .24),("margin",True,.22 if share_ok else .24),
        ("roe",True,.13 if share_ok else .14),("pe",False,.18 if share_ok else .20),
        ("ev_ebitda",False,.15 if share_ok else .18),
    ]
    if share_ok: specs.append(("market_share",True,.10))
    for row in rows:
        total=0.0; weight=0.0; used=0
        for key,higher,w in specs:
            value=row.get(key); vals=[x.get(key) for x in rows if x.get(key) is not None]
            rank=_pct_rank(vals,value,higher) if value is not None else None
            if rank is None: continue
            total+=rank*w; weight+=w; used+=1
        row["screen_score"]=total/weight if weight else None; row["screen_metrics"]=used
    target=next((x for x in rows if x["ticker"]==target_ticker.upper()),None)
    candidates=[x for x in rows if x["ticker"]!=target_ticker.upper() and x.get("screen_score") is not None and x.get("screen_metrics",0)>=4]
    if not target or target.get("screen_score") is None or target.get("screen_metrics",0)<4 or not candidates:
        return {"result":"Peer metrics are incomplete; no robust alternative candidate","candidate":None,"target":target}
    best=max(candidates,key=lambda x:x["screen_score"]); gap=best["screen_score"]-target["screen_score"]
    reasons=[]
    for key,label,higher in [
        ("growth","revenue growth",True),("margin","operating margin",True),("roe","ROE",True),
        ("pe","forward P/E",False),("ev_ebitda","EV/EBITDA",False),
        ("market_share","comparable industry market share",True),
    ]:
        a=best.get(key); b=target.get(key)
        if a is None or b is None: continue
        better=(a>b) if higher else (a<b)
        if better: reasons.append(label)
    clear=gap>=10 and len(reasons)>=3
    return {
        "result":f"{best['ticker']} merits deeper research on the current peer screen" if clear else "No clearly superior same-sector company on the current peer screen",
        "candidate":best if clear else None,"best":best,"target":target,"score_gap":gap,
        "reasons":reasons,"market_share_used":share_ok,
    }

def _write_leadership(wb,ticker,info,employee,leadership_score,leadership_rows,alt):
    if "Leadership & Culture" in wb.sheetnames: wb.remove(wb["Leadership & Culture"])
    ws=wb.create_sheet("Leadership & Culture"); ws.sheet_view.showGridLines=False
    for c in range(1,9): ws.cell(1,c).fill=_fill(NAVY); ws.cell(2,c).fill=_fill(NAVY)
    ws["A1"]=f"{ticker} — Leadership, Workforce & Alternative Screen"; ws["A1"].font=Font(bold=True,color=WHITE,size=18)
    ws["A3"]="Evidence is source-scoped. Leadership scores and the alternative-company screen are transparent research aids, not factual ratings or personalized recommendations."
    ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True)

    _section(ws,5,"Worker Happiness / Employee Experience Evidence",8)
    for c,v in enumerate(["Metric","Value","Scope","Period","Status","Source"],1): ws.cell(6,c,v)
    _header(ws,6,1,6)
    vals=["Worker happiness / satisfaction signal",employee.get("score"),employee.get("scope"),employee.get("period"),employee.get("status"),employee.get("source")]
    for c,v in enumerate(vals,1): ws.cell(7,c,v)
    ws["B7"].number_format=FMT_SCORE; ws["A8"]="Evidence"; ws["B8"]=employee.get("evidence"); ws.merge_cells("B8:F8"); ws["B8"].alignment=Alignment(wrap_text=True)
    if employee.get("source"): ws["F7"].hyperlink=employee["source"]; ws["F7"].font=Font(color=LINK_GREEN,underline="single")

    _section(ws,10,"Leadership Evidence Score — Transparent Proxy",8)
    ws["A11"]="Composite proxy / 100"; ws["B11"]=leadership_score; ws["B11"].number_format=FMT_SCORE
    ws["C11"]="Execution, capital allocation, leadership depth, culture and governance disclosure; not a factual management rating."
    ws.merge_cells("C11:F11"); ws["C11"].alignment=Alignment(wrap_text=True)
    for c,v in enumerate(["Dimension","Score / 100","Evidence / Caveat"],1): ws.cell(13,c,v)
    _header(ws,13,1,3)
    for r,(name,score,note) in enumerate(leadership_rows,14):
        ws.cell(r,1,name); ws.cell(r,2,score); ws.cell(r,2).number_format=FMT_SCORE; ws.cell(r,3,note); ws.cell(r,3).alignment=Alignment(wrap_text=True)

    _section(ws,21,"Executive Team — Public Snapshot",8)
    for c,v in enumerate(["Name","Title","Age / Birth Year","Reported Pay","Issuer Source"],1): ws.cell(22,c,v)
    _header(ws,22,1,5)
    src=_sources(ticker,info); leadership_src=src.get("leadership") or src.get("governance") or src.get("annual_reports") or src.get("investor")
    row=23
    for officer in (info.get("companyOfficers") or [])[:8]:
        if not isinstance(officer,dict): continue
        for c,v in enumerate([officer.get("name"),officer.get("title"),officer.get("age") or officer.get("yearBorn"),officer.get("totalPay"),leadership_src],1): ws.cell(row,c,v)
        if leadership_src: ws.cell(row,5).hyperlink=leadership_src; ws.cell(row,5).font=Font(color=LINK_GREEN,underline="single")
        row+=1
    if row==23: ws["A23"]="Public officer data unavailable; use the issuer source registry."

    _section(ws,33,"Business Market Position — Source-Scoped Snapshots",8)
    for c,v in enumerate(["Metric","Share","Period","Market Definition","Method","Provider","Source"],1): ws.cell(34,c,v)
    _header(ws,34,1,7)
    positions=business_market_share_records(ticker)
    for rr,rec in enumerate(positions[:6],35):
        values=[rec.get("metric"),rec.get("share"),rec.get("period"),rec.get("basis"),rec.get("method"),rec.get("provider"),rec.get("source")]
        for c,v in enumerate(values,1): ws.cell(rr,c,v)
        ws.cell(rr,2).number_format=FMT_PCT
        if rec.get("source"): ws.cell(rr,7).hyperlink=rec["source"]; ws.cell(rr,7).font=Font(color=LINK_GREEN,underline="single")
    if not positions: ws["A35"]="No source-scoped business market-share snapshot is mapped; blank is preferable to false precision."

    _section(ws,43,"Same-Sector Alternative Screen",8)
    for c,v in enumerate(["Item","Result","Score","Why / Evidence"],1): ws.cell(44,c,v)
    _header(ws,44,1,4)
    target=alt.get("target") or {}; best=alt.get("best") or {}; candidate=alt.get("candidate") or {}
    rows=[
        ("Current company peer-screen score",target.get("ticker") or ticker,target.get("screen_score"),f"{target.get('screen_metrics',0)} comparable metrics"),
        ("Best peer on current metrics",best.get("ticker"),best.get("screen_score"),", ".join(alt.get("reasons") or [])),
        ("Research conclusion",alt.get("result"),None,"Screen only; validate business quality, balance sheet, market structure and risks."),
    ]
    for rr,rowvals in enumerate(rows,45):
        for c,v in enumerate(rowvals,1): ws.cell(rr,c,v)
        ws.cell(rr,4).alignment=Alignment(wrap_text=True)
    if candidate:
        ws["A48"]="Candidate for deeper research"; ws["B48"]=candidate.get("ticker"); ws["C48"]=candidate.get("screen_score")
        ws["D48"]="Requires full fundamental and valuation work before preferring it."; ws["D48"].alignment=Alignment(wrap_text=True)

    _section(ws,51,"Research Source Registry",8)
    for c,v in enumerate(["Topic","Source / Provider","URL","Use"],1): ws.cell(52,c,v)
    _header(ws,52,1,4); rr=53
    for topic,url in src.items():
        ws.cell(rr,1,topic.replace("_"," ").title()); ws.cell(rr,2,"Issuer-owned"); ws.cell(rr,3,url)
        ws.cell(rr,4,"Preferred company, filing, governance or workforce evidence")
        if str(url).startswith("http"): ws.cell(rr,3).hyperlink=url; ws.cell(rr,3).font=Font(color=LINK_GREEN,underline="single")
        rr+=1
    for topic,s in SPECIALIST_MARKET_SOURCES.items():
        ws.cell(rr,1,topic.replace("_"," ").title()); ws.cell(rr,2,s.get("provider")); ws.cell(rr,3,s.get("url")); ws.cell(rr,4,s.get("purpose"))
        ws.cell(rr,3).hyperlink=s.get("url"); ws.cell(rr,3).font=Font(color=LINK_GREEN,underline="single"); rr+=1
    ws.cell(rr,1,"Employee reviews — optional"); ws.cell(rr,2,"Glassdoor / Indeed / Comparably")
    ws.cell(rr,4,"Manual corroboration only; not auto-scraped or treated as audited company-wide evidence")
    for col,w in {"A":38,"B":28,"C":58,"D":48,"E":24,"F":26,"G":60}.items(): ws.column_dimensions[col].width=w
    ws.freeze_panes="A6"

def _dashboard_peer_direct(wb):
    if "Dashboard" not in wb.sheetnames or "Peer Comps" not in wb.sheetnames: return
    ws=wb["Dashboard"]
    for r,col,higher in [(19,"C",False),(20,"D",False),(21,"E",False),(22,"F",True),(23,"G",True),(24,"H",True)]:
        ws.cell(r,2,f"='Peer Comps'!{col}4")
        ws.cell(r,3,f"=IFERROR(MEDIAN('Peer Comps'!{col}5:{col}9),\"\")")
        comp=">" if higher else "<"; good="Better" if higher else "Attractive"; bad="Worse" if higher else "Premium"
        ws.cell(r,4,f'=IF(OR(B{r}="",C{r}=""),"",IF(B{r}{comp}C{r},"{good}","{bad}"))')

def _block_row(ws,title,min_row,height=12,width=8):
    for r in range(1,ws.max_row+1):
        if str(ws.cell(r,1).value or "").strip()==title:
            for rr in range(r,min(ws.max_row,r+height)+1):
                for c in range(1,width+1): ws.cell(rr,c).value=None
            return r
    return max(min_row,ws.max_row+2)

def _dashboard(wb,ticker,employee,leadership_score,alt):
    if "Dashboard" not in wb.sheetnames: return
    ws=wb["Dashboard"]; title="People, Leadership & Market Position"; start=_block_row(ws,title,33,12,4)
    ws.cell(start,1,title)
    for c in range(1,5): ws.cell(start,c).fill=_fill(NAVY); ws.cell(start,c).font=Font(bold=True,color=WHITE)
    peers=_peer_rows(wb); target=next((x for x in peers if x["ticker"]==ticker.upper()),{})
    comparable=market_share_record(ticker)
    rows=[
        ("Worker happiness / satisfaction signal",employee.get("score") if employee.get("score") is not None else employee.get("status"),employee.get("scope")),
        ("Workforce evidence status",employee.get("status"),employee.get("source") or "No verified issuer source mapped"),
        ("Leadership evidence proxy / 100",leadership_score,"See Leadership & Culture for component scores"),
        ("Comparable industry market share",comparable.get("share"),(comparable.get("basis") or "No like-for-like peer market-share source mapped")+(f"; {comparable.get('period')}" if comparable else "")),
        ("Selected peer-set market value %",target.get("peer_set_weight"),"Calculated from selected target + peers; not an industry market share"),
    ]
    for rec in business_market_share_records(ticker)[:2]:
        rows.append((rec.get("metric"),rec.get("share"),f"{rec.get('period')} — {rec.get('provider')}; {rec.get('basis')}"))
    rows.append(("Same-sector alternative screen",(alt.get("candidate") or {}).get("ticker") or "None clearly superior",alt.get("result")))
    for rr,(label,value,note) in enumerate(rows,start+1):
        ws.cell(rr,1,label); ws.cell(rr,2,value); ws.cell(rr,3,note); ws.cell(rr,3).alignment=Alignment(wrap_text=True)
        if isinstance(value,(int,float)) and ("share" in label.lower() or "%" in label): ws.cell(rr,2).number_format=FMT_PCT
        elif isinstance(value,(int,float)): ws.cell(rr,2).number_format=FMT_SCORE
    ws.column_dimensions["C"].width=max(ws.column_dimensions["C"].width or 0,56)

def _summary(wb,ticker,employee,leadership_score,alt):
    if "Investment Summary" not in wb.sheetnames: return
    ws=wb["Investment Summary"]; title="People, Leadership & Competitive Position"; row=_block_row(ws,title,ws.max_row+2,8,8)
    for c in range(1,9): ws.cell(row,c).fill=_fill(NAVY); ws.cell(row,c).font=Font(bold=True,color=WHITE)
    ws.cell(row,1,title)
    pos=business_market_share_records(ticker)
    pos_text="; ".join(f"{x.get('metric')}: {x.get('share'):.1%} ({x.get('period')})" for x in pos[:2] if isinstance(x.get("share"),(int,float))) or "No source-scoped market-position snapshot mapped"
    values=[
        ("Worker happiness / employee signal",employee.get("score") if employee.get("score") is not None else employee.get("status"),employee.get("scope")),
        ("Leadership evidence proxy / 100",leadership_score,"See Leadership & Culture"),
        ("Business market position",pos_text,"See Leadership & Culture for market definition and source"),
        ("Same-sector alternative",(alt.get("candidate") or {}).get("ticker") or "None clearly superior",alt.get("result")),
    ]
    for rr,(label,value,note) in enumerate(values,row+1):
        ws.cell(rr,1,label); ws.cell(rr,2,value); ws.cell(rr,3,note)
        ws.merge_cells(start_row=rr,start_column=3,end_row=rr,end_column=8); ws.cell(rr,3).alignment=Alignment(wrap_text=True)

def _quality(wb,ticker,info,employee,alt):
    if "Data Quality" not in wb.sheetnames: return
    ws=wb["Data Quality"]; existing={str(ws.cell(r,1).value or "").strip():r for r in range(1,ws.max_row+1)}
    src=_sources(ticker,info); share=market_share_record(ticker); target=alt.get("target") or {}
    checks=[
        ("Issuer source registry","PASS" if src else "REVIEW",f"{len(src)} issuer-owned source(s) mapped" if src else "No explicit issuer source mapped","Issuer pages improve auditability; regulator and transparent fallback sources remain separate."),
        ("Market-share comparability","PASS" if share else "REVIEW",share.get("basis") if share else "No comparable public industry-share source mapped","Peer market share is used only on a like-for-like market definition."),
        ("Employee sentiment scope","PASS" if employee.get("score") is not None else "REVIEW",employee.get("scope"),"Program satisfaction, engagement and company-wide happiness must not be conflated."),
        ("Alternative-company evidence coverage","PASS" if target.get("screen_metrics",0)>=4 else "REVIEW",f"{target.get('screen_metrics',0)} comparable metric(s) used for target","A candidate requires adequate peer evidence plus a material score gap."),
    ]
    for name,status,observed,why in checks:
        r=existing.get(name) or ws.max_row+1; ws.cell(r,1,name); ws.cell(r,2,status); ws.cell(r,3,observed); ws.cell(r,4,why)
        ws.cell(r,2).fill=_fill(PALE_GREEN if status=="PASS" else GOLD); ws.cell(r,2).font=Font(bold=True)
        for c in range(1,5): ws.cell(r,c).alignment=Alignment(wrap_text=True,vertical="top")

def _dedupe(wb):
    _dashboard_peer_direct(wb)
    for name in ("Visual Dashboard","Comparative Analysis","Valuation Cross-Checks","AI Analysis","AI Valuation"):
        if name in wb.sheetnames: wb.remove(wb[name])

def ensure_research_extensions(wb,ticker,info=None):
    info=_safe_info(ticker,info); employee=_employee_signal(ticker,info)
    leadership_score,leadership_rows=_leadership_proxy(wb,ticker,info,employee); alt=_alternative_screen(wb,ticker)
    _write_leadership(wb,ticker,info,employee,leadership_score,leadership_rows,alt)
    _dashboard(wb,ticker,employee,leadership_score,alt); _summary(wb,ticker,employee,leadership_score,alt)
    _quality(wb,ticker,info,employee,alt); _dedupe(wb)
    return {"employee":employee,"leadership_score":leadership_score,"alternative":alt,"business_market_position":business_market_share_records(ticker),"issuer_sources":_sources(ticker,info)}
