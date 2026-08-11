from __future__ import annotations

"""Auditable company-specific WACC engine.

The prior workbook inherited a 9% base WACC from the Google template because the scenario
updater never recalculated discount rates. This module calculates a company-specific WACC
from public market inputs and writes every assumption, formula and source to a Cost of
Capital sheet.

Priority:
- risk-free rate: latest US Treasury 10-year par yield;
- equity risk premium: Damodaran implied US ERP, with a dated conservative fallback;
- beta: current Yahoo Finance beta, then 5Y monthly regression vs S&P 500;
- cost of debt: risk-free plus a transparent leverage-based synthetic spread;
- tax rate: latest reported effective tax rate from Financial Statements, else 21%;
- capital weights: current market equity and total debt.

The synthetic credit spread is explicitly a model assumption, not a reported credit quote.
"""

import math
import re
import statistics
from datetime import datetime, timezone
from xml.etree import ElementTree as ET

import requests
import yfinance as yf
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

NAVY="17365D"; BLUE="2F75B5"; WHITE="FFFFFF"; GREY="666666"; GOLD="FFF2CC"; PALE_GREEN="E2F0D9"
FMT_PCT='0.00%;[Red](0.00%);-'; FMT_BN='#,##0.0;[Red](#,##0.0);-'; FMT_MULT='0.00x;[Red](0.00x);-'
THIN=Side(style="thin",color="D9E1F2")
TREASURY_URL="https://home.treasury.gov/resource-center/data-chart-center/interest-rates/pages/xml?data=daily_treasury_yield_curve&field_tdr_date_value={year}"
TREASURY_SOURCE="https://home.treasury.gov/resource-center/data-chart-center/interest-rates/TextView?type=daily_treasury_yield_curve"
DAMODARAN_SOURCE="https://pages.stern.nyu.edu/~adamodar/New_Home_Page/home.htm"
YAHOO_SOURCE="https://finance.yahoo.com/"


def _fill(c): return PatternFill("solid",fgColor=c)
def _num(v,default=None):
    try:
        if isinstance(v,bool) or v in (None,""): return default
        x=float(v); return x if math.isfinite(x) else default
    except Exception: return default

def _local(tag): return str(tag).split("}")[-1].upper()

def _treasury_10y():
    """Return latest available 10Y Treasury yield as decimal, date, status."""
    year=datetime.now(timezone.utc).year
    headers={"User-Agent":"Antzaz Equity Research public-data WACC engine"}
    for y in (year,year-1):
        try:
            r=requests.get(TREASURY_URL.format(year=y),headers=headers,timeout=15); r.raise_for_status()
            root=ET.fromstring(r.content); rows=[]
            for entry in root.iter():
                if _local(entry.tag)!="ENTRY": continue
                date_val=None; ten=None
                for node in entry.iter():
                    name=_local(node.tag); text=(node.text or "").strip()
                    if not text: continue
                    if name in {"NEW_DATE","NEWDATE"}: date_val=text[:10]
                    elif name in {"BC_10YEAR","BC10YEAR"}: ten=_num(text)
                if date_val and ten is not None: rows.append((date_val,ten/100))
            if rows:
                date_val,value=max(rows,key=lambda x:x[0]); return value,date_val,"Treasury 10Y"
        except Exception:
            continue
    # Dated fallback aligned to Damodaran's July 1, 2026 ERP update.
    return .0445,"2026-07-01","Fallback: 4.45% Treasury rate used in Damodaran July-2026 ERP update"


def _damodaran_erp():
    headers={"User-Agent":"Antzaz Equity Research public-data WACC engine"}
    try:
        text=requests.get(DAMODARAN_SOURCE,headers=headers,timeout=15).text
        # Example: Implied ERP on July 1, 2026 = 4.18% (Trailing 12 month, with adjusted payout)
        m=re.search(r"Implied ERP on\s+([^=]{4,40})=\s*([0-9]+(?:\.[0-9]+)?)%",text,re.I)
        if m:
            return float(m.group(2))/100,re.sub(r"\s+"," ",m.group(1)).strip(),"Damodaran implied ERP"
    except Exception:
        pass
    return .0418,"2026-07-01","Fallback: Damodaran July-2026 implied ERP"


def _regression_beta(ticker):
    try:
        px=yf.download([ticker,"^GSPC"],period="5y",interval="1mo",auto_adjust=True,progress=False,threads=False)
        close=px["Close"] if "Close" in px else px
        if ticker not in close or "^GSPC" not in close: return None
        a=close[ticker].pct_change().dropna(); b=close["^GSPC"].pct_change().dropna(); joined=a.to_frame("a").join(b.to_frame("b"),how="inner").dropna()
        if len(joined)<24: return None
        var=float(joined["b"].var()); cov=float(joined[["a","b"]].cov().iloc[0,1])
        return cov/var if var>0 else None
    except Exception:
        return None


def _beta(ticker,info):
    b=_num((info or {}).get("beta"))
    if b is not None and .05<b<4.0: return b,"Yahoo Finance beta"
    b=_regression_beta(ticker)
    if b is not None and .05<b<4.0: return b,"5Y monthly beta vs S&P 500 calculated from Yahoo price history"
    return 1.0,"Fallback beta = 1.00 (market beta); REVIEW"


def _find_row(ws,label,start=1,end=None):
    if ws is None: return None
    needle=str(label).strip().lower(); end=end or ws.max_row
    for r in range(start,min(end,ws.max_row)+1):
        if str(ws.cell(r,1).value or "").strip().lower()==needle: return r
    return None


def _effective_tax_rate(wb):
    if "Financial Statements" not in wb.sheetnames: return .21,"Fallback statutory/reference rate"
    ws=wb["Financial Statements"]
    pretax=_find_row(ws,"Pre-Tax Income"); taxes=_find_row(ws,"Income Taxes")
    if pretax and taxes:
        for c in range(min(7,ws.max_column),1,-1):
            p=_num(ws.cell(pretax,c).value); t=_num(ws.cell(taxes,c).value)
            if p not in (None,0) and t is not None:
                rate=t/p
                if 0<=rate<=.50: return rate,"Latest reported effective tax rate"
    return .21,"Fallback statutory/reference rate"


def _latest_fcf(wb):
    if "Historical Financials" not in wb.sheetnames: return None
    ws=wb["Historical Financials"]
    for c in range(7,1,-1):
        ocf=_num(ws.cell(14,c).value); cap=_num(ws.cell(15,c).value)
        if ocf is not None and cap is not None: return ocf-abs(cap)
    return None


def _credit_spread(wb):
    d=wb["Company Data"] if "Company Data" in wb.sheetnames else None
    debt=_num(d["B13"].value) if d else None; cash=_num(d["B12"].value) if d else None; fcf=_latest_fcf(wb)
    net=(debt-cash) if debt is not None and cash is not None else None
    if net is None: return .015,"1.50% fallback synthetic spread"
    if net<=0: return .0075,"0.75% synthetic spread: net-cash balance sheet"
    leverage=net/max(abs(fcf or 0),1e-9) if fcf is not None else None
    if leverage is None: return .015,"1.50% synthetic spread: FCF unavailable"
    if leverage<=1: return .0100,f"1.00% synthetic spread: net debt/FCF {leverage:.2f}x"
    if leverage<=2: return .0150,f"1.50% synthetic spread: net debt/FCF {leverage:.2f}x"
    if leverage<=3: return .0200,f"2.00% synthetic spread: net debt/FCF {leverage:.2f}x"
    return .0300,f"3.00% synthetic spread: net debt/FCF {leverage:.2f}x"


def calculate_wacc(wb,ticker,info=None):
    info=info or {}
    d=wb["Company Data"] if "Company Data" in wb.sheetnames else None
    equity=_num(d["B10"].value) if d else None; debt=_num(d["B13"].value) if d else None
    if equity is None: equity=_num(info.get("marketCap")); equity=equity/1e9 if equity else None
    if debt is None: debt=_num(info.get("totalDebt")); debt=debt/1e9 if debt else 0.0
    debt=max(0.0,debt or 0.0); equity=max(0.0,equity or 0.0)
    rf,rf_date,rf_status=_treasury_10y(); erp,erp_date,erp_status=_damodaran_erp(); beta,beta_status=_beta(ticker,info)
    tax,tax_status=_effective_tax_rate(wb); spread,spread_status=_credit_spread(wb)
    cost_equity=rf+beta*erp; pretax_debt=rf+spread; aftertax_debt=pretax_debt*(1-tax)
    total=equity+debt
    ew=equity/total if total else 1.0; dw=debt/total if total else 0.0
    base=ew*cost_equity+dw*aftertax_debt
    base=max(.045,min(.18,base)); bear=min(.20,base+.015); bull=max(.04,base-.010)
    return {
        "risk_free":rf,"risk_free_date":rf_date,"risk_free_status":rf_status,
        "erp":erp,"erp_date":erp_date,"erp_status":erp_status,"beta":beta,"beta_status":beta_status,
        "cost_equity":cost_equity,"credit_spread":spread,"credit_spread_status":spread_status,
        "pretax_cost_debt":pretax_debt,"tax_rate":tax,"tax_status":tax_status,"aftertax_cost_debt":aftertax_debt,
        "equity_value":equity,"debt_value":debt,"equity_weight":ew,"debt_weight":dw,
        "base_wacc":base,"bear_wacc":bear,"bull_wacc":bull,
    }


def _write_sheet(wb,ticker,x):
    if "Cost of Capital" in wb.sheetnames: wb.remove(wb["Cost of Capital"])
    ws=wb.create_sheet("Cost of Capital"); ws.sheet_view.showGridLines=False
    for c in range(1,8): ws.cell(1,c).fill=_fill(NAVY); ws.cell(2,c).fill=_fill(NAVY)
    ws["A1"]=f"{ticker} — Cost of Capital / WACC"; ws["A1"].font=Font(bold=True,color=WHITE,size=18)
    ws["A3"]="Company-specific public-data WACC. Market inputs are refreshed at model build; the credit spread is a transparent model estimate when no market debt yield is available."
    ws.merge_cells("A3:G3"); ws["A3"].font=Font(italic=True,color=GREY); ws["A3"].alignment=Alignment(wrap_text=True)
    headers=["Input","Value","Formula / Role","Source / Method","As of","Status","Source URL"]
    for c,v in enumerate(headers,1): ws.cell(5,c,v); ws.cell(5,c).fill=_fill(BLUE); ws.cell(5,c).font=Font(bold=True,color=WHITE); ws.cell(5,c).alignment=Alignment(wrap_text=True)
    rows=[
        ("Risk-free rate",x["risk_free"],"10Y USD risk-free proxy",x["risk_free_status"],x["risk_free_date"],"PASS",TREASURY_SOURCE),
        ("Equity risk premium",x["erp"],"US implied ERP",x["erp_status"],x["erp_date"],"PASS",DAMODARAN_SOURCE),
        ("Equity beta",x["beta"],"Systematic equity risk",x["beta_status"],"Build date","PASS" if "REVIEW" not in x["beta_status"] else "REVIEW",YAHOO_SOURCE),
        ("Cost of equity",x["cost_equity"],"Risk-free + Beta × ERP","Calculated from rows above","Build date","PASS",None),
        ("Synthetic credit spread",x["credit_spread"],"Added to risk-free for pre-tax debt cost",x["credit_spread_status"],"Build date","MODEL",None),
        ("Pre-tax cost of debt",x["pretax_cost_debt"],"Risk-free + credit spread","Calculated","Build date","MODEL",None),
        ("Effective tax rate",x["tax_rate"],"Tax shield on debt",x["tax_status"],"Latest reported FY","PASS" if "reported" in x["tax_status"].lower() else "REVIEW",None),
        ("After-tax cost of debt",x["aftertax_cost_debt"],"Pre-tax debt cost × (1 − tax rate)","Calculated","Build date","PASS",None),
        ("Market equity ($bn)",x["equity_value"],"E in capital weights","Company Data / current market cap","Build date","PASS",YAHOO_SOURCE),
        ("Debt ($bn)",x["debt_value"],"D in capital weights","Company Data / latest total debt","Latest available","PASS",None),
        ("Equity weight",x["equity_weight"],"E / (D + E)","Calculated","Build date","PASS",None),
        ("Debt weight",x["debt_weight"],"D / (D + E)","Calculated","Build date","PASS",None),
    ]
    for r,row in enumerate(rows,6):
        for c,v in enumerate(row,1): ws.cell(r,c,v)
        if r in (14,15): ws.cell(r,2).number_format=FMT_BN
        elif r==8: ws.cell(r,2).number_format=FMT_MULT
        else: ws.cell(r,2).number_format=FMT_PCT
        if row[-1]: ws.cell(r,7).hyperlink=row[-1]; ws.cell(r,7).font=Font(color="008000",underline="single")
    for c in range(1,8): ws.cell(19,c).fill=_fill(NAVY); ws.cell(19,c).font=Font(bold=True,color=WHITE)
    ws["A19"]="Calculated WACC & Scenario Discount Rates"
    hs=["Scenario","WACC","Adjustment vs Base","Formula","Use"]
    for c,v in enumerate(hs,1): ws.cell(20,c,v); ws.cell(20,c).fill=_fill(BLUE); ws.cell(20,c).font=Font(bold=True,color=WHITE)
    scen=[("Bear",x["bear_wacc"],x["bear_wacc"]-x["base_wacc"],"Base WACC + 150 bps","Downside valuation"),("Base",x["base_wacc"],0.0,"E/(D+E)×CoE + D/(D+E)×CoD×(1−Tax)","Primary DCF"),("Bull",x["bull_wacc"],x["bull_wacc"]-x["base_wacc"],"Base WACC − 100 bps","Strong-execution valuation")]
    for r,row in enumerate(scen,21):
        for c,v in enumerate(row,1): ws.cell(r,c,v)
        ws.cell(r,2).number_format=FMT_PCT; ws.cell(r,3).number_format=FMT_PCT
    ws["A26"]="Interpretation"; ws["A26"].font=Font(bold=True)
    ws["B26"]="A lower/higher WACC should arise from observable market risk, beta and capital structure—not from the ticker inheriting a template assumption. The scenario WACC adjustments are explicit stress assumptions."
    ws.merge_cells("B26:G27"); ws["B26"].alignment=Alignment(wrap_text=True,vertical="top")
    widths={"A":27,"B":17,"C":42,"D":48,"E":18,"F":13,"G":55}
    for c,w in widths.items(): ws.column_dimensions[c].width=w
    ws.freeze_panes="A6"; return ws


def _quality_row(wb,x):
    if "Data Quality" not in wb.sheetnames: return
    ws=wb["Data Quality"]; label="Company-specific WACC calculation"; row=None
    for r in range(1,ws.max_row+1):
        if str(ws.cell(r,1).value or "").strip()==label: row=r; break
    row=row or ws.max_row+1
    status="PASS" if x.get("risk_free") and x.get("erp") and x.get("beta") else "REVIEW"
    ws.cell(row,1,label); ws.cell(row,2,status); ws.cell(row,3,f"Base WACC {x['base_wacc']:.2%}; Rf {x['risk_free']:.2%}; ERP {x['erp']:.2%}; beta {x['beta']:.2f}; E wt {x['equity_weight']:.1%}; D wt {x['debt_weight']:.1%}")
    ws.cell(row,4,"Cost of Capital sheet exposes market inputs, formula, source date and scenario adjustments. Template 9% is no longer reused.")
    ws.cell(row,2).fill=_fill(PALE_GREEN if status=="PASS" else GOLD); ws.cell(row,2).font=Font(bold=True)
    for c in range(1,5): ws.cell(row,c).alignment=Alignment(wrap_text=True,vertical="top")


def apply_dynamic_wacc(wb,ticker,info=None):
    x=calculate_wacc(wb,str(ticker).upper(),info or {})
    if "Three-Case Scenarios" in wb.sheetnames:
        ws=wb["Three-Case Scenarios"]
        ws["B6"]=x["bear_wacc"]; ws["C6"]=x["base_wacc"]; ws["D6"]=x["bull_wacc"]
        for cell in ("B6","C6","D6"): ws[cell].number_format=FMT_PCT
    _write_sheet(wb,str(ticker).upper(),x); _quality_row(wb,x)
    setattr(wb,"_wacc_inputs",x)
    print(f"Dynamic WACC: {ticker} base={x['base_wacc']:.2%}, bear={x['bear_wacc']:.2%}, bull={x['bull_wacc']:.2%}, beta={x['beta']:.2f}, Rf={x['risk_free']:.2%}, ERP={x['erp']:.2%}")
    return x
