from __future__ import annotations

"""Enrich Segment Analysis from issuer-owned and regulatory public sources.

This module is intentionally a second pass. It never replaces a working issuer-specific
segment repair. It adds missing official segment names and fills blank economics only when
an official/regulatory table can be matched with reasonable confidence.
"""

import re
import statistics

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from segment_source_engine import collect_segment_documents, verified_fallback
from segment_analysis_v2 import _tables, _discover_segments, _extract_known, _pick

GOLD = "FFF2CC"
INPUT_BLUE = "0000FF"
LINK_GREEN = "008000"
GREY = "666666"
FMT_BN = '#,##0.0;[Red](#,##0.0);-'
FMT_PCT = '0.0%;[Red](0.0%);-'


def _fill(color):
    return PatternFill("solid", fgColor=color)


def _num(value):
    try:
        if isinstance(value, bool):
            return None
        return float(value)
    except Exception:
        return None


def _clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def _key(name):
    text = _clean(name)
    low = re.sub(r"[^a-z0-9]", "", text.lower())
    if "groupfunctions" in low or "groupitems" in low:
        return "groupfunctions"
    m = re.search(r"\(([A-Z]{2,10})\)\s*$", text)
    if m:
        return m.group(1).lower()
    if re.fullmatch(r"[A-Z]{2,10}", text):
        return text.lower()
    return low


def _parse_business_division_list(fragment):
    text = _clean(fragment).strip(" .:;-")
    if not text:
        return []
    # Stop before the next explanatory sentence.
    text = re.split(r"\.\s+(?:All |Together |These |The |Group |Financial |Our |We )", text, maxsplit=1)[0]
    parts = [x.strip(" .:;-") for x in re.split(r"\s*;\s*|\s*,\s*", text) if x.strip()]
    if len(parts) >= 2:
        last = parts[-1]
        # In a comma-separated list, split the first title-cased final conjunction only.
        m = re.search(r"\s+and\s+(?=(?:the\s+)?[A-Z])", last)
        if m:
            left = last[:m.start()].strip(); right = last[m.end():].strip()
            if left and right:
                parts[-1:] = [left, right]
    out = []
    bad = {"group", "company", "business", "businesses", "segment", "segments", "division", "divisions"}
    for item in parts:
        item = re.sub(r"^(?:and|the)\s+", "", item, flags=re.I).strip(" .:;-")
        if not item or len(item) > 100 or not re.search(r"[A-Za-z]", item):
            continue
        if item.lower() in bad:
            continue
        if item not in out:
            out.append(item)
    return out if 2 <= len(out) <= 15 else []


def _narrative_divisions(text):
    text = _clean(text)
    if not text:
        return []
    patterns = [
        r"(?:businesses|operations)\s+(?:are|is)\s+organized(?:\s+globally)?\s+into\s+(?:the\s+following\s+)?(?:\w+\s+)?business\s+divisions?\s*:\s*(.{8,650}?)\.",
        r"operational\s+structure\s+(?:of\s+the\s+group\s+)?(?:comprises|consists\s+of|includes)\s+(?:the\s+)?(?:group\s+functions?\s+and\s+)?(?:\w+\s+)?business\s+divisions?\s*:\s*(.{8,650}?)\.",
        r"(?:reportable|operating)\s+segments?\s+(?:are|include|consist\s+of)\s*:?\s*(.{8,650}?)\.",
        r"(?:business\s+divisions?|operating\s+divisions?)\s+(?:are|include|consist\s+of)\s*:?\s*(.{8,650}?)\.",
    ]
    best = []
    for pattern in patterns:
        for match in re.finditer(pattern, text, flags=re.I):
            names = _parse_business_division_list(match.group(1))
            if len(names) > len(best):
                best = names
    return best


def _find_row(ws, label):
    target = str(label).strip().lower()
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip().lower() == target:
            return r
    return None


def _segment_block(ws):
    section = _find_row(ws, "Reported Operating / Reportable Segments")
    if section is None:
        section = _find_row(ws, "Reported Operating Segments")
    if section is None:
        section = _find_row(ws, "Reported Segments")
    business = _find_row(ws, "Revenue by Business Line / Product Group")
    if business is None:
        business = _find_row(ws, "Revenue by Business Line")
    if section is None:
        return None, None, None
    header = section + 1
    end = (business - 1) if business else min(ws.max_row, header + 15)
    return header, header + 1, end


def _company_revenue(wb):
    if "Historical Financials" not in wb.sheetnames:
        return None
    return _num(wb["Historical Financials"]["G4"].value)


def _scale_for(raw_values, company_revenue):
    latest = []
    for vals in raw_values:
        if vals and _num(vals[-1]) is not None:
            latest.append(abs(float(vals[-1])))
    if not latest:
        return 1.0
    med = statistics.median(latest)
    if company_revenue and med > company_revenue * 10:
        return 0.001
    if med > 10000:
        return 0.001
    return 1.0


def _format_new_row(ws, row, block_end):
    for c in range(2, 10):
        if ws.cell(row, c).value is None:
            ws.cell(row, c).fill = _fill(GOLD)
            ws.cell(row, c).font = Font(color=INPUT_BLUE)
            ws.cell(row, c).number_format = FMT_BN
    ws.cell(row, 5, f'=IFERROR(D{row}/C{row}-1,"")')
    ws.cell(row, 6, f'=IFERROR((D{row}/B{row})^(1/2)-1,"")')
    for rev_col, profit_col, margin_col in ((2, 7, 10), (3, 8, 11), (4, 9, 12)):
        ws.cell(row, margin_col, f'=IFERROR({get_column_letter(profit_col)}{row}/{get_column_letter(rev_col)}{row},"")')
    ws.cell(row, 13, f'=IFERROR(L{row}-K{row},"")')
    ws.cell(row, 14, f'=IFERROR(D{row}/SUM($D$7:$D${block_end}),"")')
    for c in (5, 6, 10, 11, 12, 13, 14):
        ws.cell(row, c).number_format = FMT_PCT


def _set_source(cell, text, url=None):
    cell.value = text
    cell.font = Font(color=LINK_GREEN)
    if url:
        try:
            cell.hyperlink = url
        except Exception:
            pass


def enrich_segment_analysis(wb, ticker, headers):
    ticker = str(ticker).upper().strip()
    if "Segment Analysis" not in wb.sheetnames:
        return False
    ws = wb["Segment Analysis"]

    try:
        docs = collect_segment_documents(ticker, headers)
    except Exception as exc:
        print(f"Warning: official segment source discovery failed: {exc}")
        docs = []

    discovered = []
    source_for = {}
    numeric_tables = []
    parser_notes = []
    for doc in docs:
        html = doc.get("html")
        text = doc.get("text") or ""
        names = []
        if html:
            tables, parser = _tables(html)
            parser_notes.append(f"{doc.get('kind')}: {parser}")
            if tables:
                numeric_tables.append((doc, tables))
                try:
                    names.extend(_discover_segments(tables, html))
                except Exception:
                    pass
        try:
            names.extend(_narrative_divisions(text))
        except Exception:
            pass
        for name in names:
            k = _key(name)
            if not k:
                continue
            if all(_key(x) != k for x in discovered):
                discovered.append(name)
                source_for[k] = (doc.get("kind", "Official source"), doc.get("url"))

    fallback = verified_fallback(ticker)
    fallback_names = list(fallback.get("segments", []))
    # Verified fallback protects against PDFs / JavaScript IR pages that cannot be parsed.
    # Live discovery still takes precedence for naming and source attribution.
    if len(discovered) < 2 and fallback_names:
        for name in fallback_names:
            k = _key(name)
            if all(_key(x) != k for x in discovered):
                discovered.append(name)
                source_for[k] = ("Verified issuer fallback", fallback.get("source"))

    header_row, first_row, block_end = _segment_block(ws)
    if header_row is None:
        return False

    existing = {}
    for r in range(first_row, block_end + 1):
        name = ws.cell(r, 1).value
        if name:
            existing[_key(name)] = r

    missing = [name for name in discovered if _key(name) not in existing]
    blank_rows = [r for r in range(first_row, block_end + 1) if not ws.cell(r, 1).value]
    if len(missing) > len(blank_rows):
        business_row = block_end + 1
        add = len(missing) - len(blank_rows)
        ws.insert_rows(business_row, amount=add)
        blank_rows.extend(range(business_row, business_row + add))
        block_end += add

    for name, row in zip(missing, blank_rows):
        ws.cell(row, 1, name)
        _format_new_row(ws, row, block_end)
        kind, url = source_for.get(_key(name), ("Official/public source", None))
        ws.cell(row, 15, "Name disclosed; numbers unavailable")
        _set_source(ws.cell(row, 16), kind, url)
        existing[_key(name)] = row

    # Use all names currently on the sheet as extraction targets. This allows official
    # 20-F/6-K/annual-report tables to fill rows originally created by the older parser.
    labels = [str(ws.cell(r, 1).value) for r in range(first_row, block_end + 1) if ws.cell(r, 1).value]
    raw_by_label = {lab: {"revenue": None, "op": None, "source": None} for lab in labels}

    # Alias Group Functions / Group Items because UBS and other banks may use both labels
    # across narrative and tabular disclosures.
    extract_labels = list(labels)
    alias_owner = {}
    for lab in labels:
        if _key(lab) == "groupfunctions":
            for alias in ("Group Functions", "Group functions", "Group Items"):
                if alias not in extract_labels:
                    extract_labels.append(alias)
                alias_owner[alias] = lab

    for doc, tables in numeric_tables:
        try:
            extracted = _extract_known(tables, extract_labels)
        except Exception:
            continue
        for label in extract_labels:
            owner = alias_owner.get(label, label)
            if owner not in raw_by_label:
                continue
            rev = _pick(extracted.get(label, {}).get("revenue", []))
            op = _pick(extracted.get(label, {}).get("op", []))
            if rev and raw_by_label[owner]["revenue"] is None:
                raw_by_label[owner]["revenue"] = rev
                raw_by_label[owner]["source"] = doc
            if op and raw_by_label[owner]["op"] is None:
                raw_by_label[owner]["op"] = op
                raw_by_label[owner]["source"] = raw_by_label[owner]["source"] or doc

    scale = _scale_for(
        [v["revenue"] for v in raw_by_label.values() if v.get("revenue")],
        _company_revenue(wb),
    )

    filled_numeric = 0
    for label, payload in raw_by_label.items():
        row = existing.get(_key(label))
        if row is None:
            continue
        rev = payload.get("revenue")
        op = payload.get("op")
        any_fill = False
        if rev:
            vals = list(rev[-3:])
            while len(vals) < 3:
                vals.insert(0, None)
            for c, value in enumerate(vals, 2):
                num = _num(value)
                if num is not None and ws.cell(row, c).value in (None, ""):
                    ws.cell(row, c).value = num * scale
                    ws.cell(row, c).number_format = FMT_BN
                    ws.cell(row, c).fill = PatternFill(fill_type=None)
                    ws.cell(row, c).font = Font(color=INPUT_BLUE)
                    any_fill = True
        if op:
            vals = list(op[-3:])
            while len(vals) < 3:
                vals.insert(0, None)
            for c, value in enumerate(vals, 7):
                num = _num(value)
                if num is not None and ws.cell(row, c).value in (None, ""):
                    ws.cell(row, c).value = num * scale
                    ws.cell(row, c).number_format = FMT_BN
                    ws.cell(row, c).fill = PatternFill(fill_type=None)
                    ws.cell(row, c).font = Font(color=INPUT_BLUE)
                    any_fill = True
        if any_fill:
            filled_numeric += 1
            have_rev = any(_num(ws.cell(row, c).value) is not None for c in (2, 3, 4))
            have_op = any(_num(ws.cell(row, c).value) is not None for c in (7, 8, 9))
            ws.cell(row, 15, "Revenue + profitability" if have_rev and have_op else ("Revenue only" if have_rev else "Profitability only"))
            doc = payload.get("source") or {}
            _set_source(ws.cell(row, 16), doc.get("kind", "Official/regulatory report"), doc.get("url"))

    # Add transparent source block without removing the original parser's source note.
    source_section = _find_row(ws, "Source & Data Quality")
    if source_section:
        row = source_section + 4
        ws.cell(row, 1, "Additional official/public sources")
        ws.cell(row, 1).font = Font(bold=True)
        unique = []
        seen_urls = set()
        for doc in docs:
            url = doc.get("url")
            if url and url not in seen_urls:
                unique.append(doc); seen_urls.add(url)
        for doc in unique[:8]:
            row += 1
            ws.cell(row, 1, doc.get("kind", "Official source"))
            _set_source(ws.cell(row, 2), doc.get("url", ""), doc.get("url"))
        if fallback.get("note"):
            row += 1
            ws.cell(row, 1, "Verified structure note")
            ws.cell(row, 2, fallback.get("note"))
            ws.cell(row, 2).alignment = Alignment(wrap_text=True)

    if discovered or filled_numeric:
        old = str(ws["A3"].value or "")
        ws["A3"] = (
            f"{old} Official-source enrichment: {len(discovered)} live/fallback segment name(s) identified; "
            f"{filled_numeric} row(s) received additional official/regulatory financial data. "
            "Source priority: issuer website/IR and annual/results reports, then regulatory filings."
        )
        ws["A3"].font = Font(italic=True, color=GREY)
        ws["A3"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[3].height = max(ws.row_dimensions[3].height or 42, 60)

    print(
        f"Segment source enrichment: ticker={ticker}, official_docs={len(docs)}, "
        f"discovered_names={len(discovered)}, added_names={len(missing)}, numeric_rows_filled={filled_numeric}"
    )
    return bool(discovered or filled_numeric)
