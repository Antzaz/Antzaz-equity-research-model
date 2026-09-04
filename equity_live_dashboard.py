from __future__ import annotations

"""Private hosted dashboard for the equity-research models in the portfolio universe."""

from pathlib import Path
import sys

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
INST = ROOT / "institutional_research"
if str(INST) not in sys.path:
    sys.path.insert(0, str(INST))

from src.live_data import download_latest_live_bundle  # noqa: E402

st.set_page_config(page_title="Portfolio Company Research", layout="wide")
st.title("Portfolio Company Equity Research")

CACHE = ROOT / ".equity_live_cache"


def _secret(name: str, default=None):
    try:
        group = st.secrets["live_data"]
        return group.get(name, default)
    except Exception:
        return default


@st.cache_resource(ttl=3600)
def _hydrate():
    return download_latest_live_bundle(
        repository=_secret("repository", "Antzaz/Antzaz-equity-research-model"),
        token=_secret("github_token"),
        password=_secret("bundle_password"),
        target=CACHE,
        workflow_file=_secret("workflow_file", "daily-portfolio-refresh.yml"),
    )


try:
    manifest = _hydrate()
except Exception as exc:
    st.error("Could not load the latest encrypted daily research bundle.")
    st.caption(str(exc))
    st.stop()

companies = [x for x in manifest.get("companies", []) if x.get("status") == "ok" and x.get("file")]
if not companies:
    st.warning("The latest daily bundle contains no successful company models.")
    st.stop()

company_by_ticker = {x["ticker"]: x for x in companies}
ticker = st.selectbox("Company", sorted(company_by_ticker))
entry = company_by_ticker[ticker]
model_path = CACHE / entry["file"]

if not model_path.exists():
    st.error("The selected company workbook is missing from the encrypted bundle.")
    st.stop()

wb = load_workbook(model_path, data_only=False, read_only=False)


def _num(value):
    try:
        if isinstance(value, bool) or value in (None, ""):
            return None
        return float(value)
    except Exception:
        return None


def _pct(value):
    value = _num(value)
    return "—" if value is None else f"{value:.1%}"


def _price(value):
    value = _num(value)
    return "—" if value is None else f"${value:,.2f}"


def _right_of_label(ws, label):
    if ws is None:
        return None
    target = str(label).strip().lower()
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value or "").strip().lower() == target:
                return ws.cell(cell.row, cell.column + 1).value
    return None


def _row_value(ws, label, value_col=2):
    if ws is None:
        return None
    target = str(label).strip().lower()
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip().lower() == target:
            return ws.cell(r, value_col).value
    return None


company = wb["Company Data"] if "Company Data" in wb.sheetnames else None
name = company["B5"].value if company else ticker
sector = company["B6"].value if company else None
industry = company["B7"].value if company else None
current_price = company["B8"].value if company else None
market_cap = company["B10"].value if company else None
forward_pe = company["B15"].value if company else None

summary = wb["Investment Summary"] if "Investment Summary" in wb.sheetnames else None
verdict = _right_of_label(summary, "Model View") if summary else None
quant_score = _right_of_label(summary, "Quant Score / 100") if summary else None
why = _right_of_label(summary, "Why") if summary else None

people = wb["Leadership & Culture"] if "Leadership & Culture" in wb.sheetnames else None
worker_score = _row_value(people, "Worker happiness / satisfaction signal")
worker_scope = _row_value(people, "Worker happiness / satisfaction signal", 3)
leadership_score = _row_value(people, "Composite proxy / 100")
peer_ws = wb["Peer Comps"] if "Peer Comps" in wb.sheetnames else None
market_share = _num(peer_ws["M4"].value) if peer_ws and peer_ws.max_column >= 13 else None
market_share_basis = peer_ws["O4"].value if peer_ws and peer_ws.max_column >= 15 else None

st.subheader(str(name or ticker))
st.caption(" • ".join(str(x) for x in (sector, industry) if x))

c1, c2, c3, c4, c5 = st.columns(5)
c1.metric("Current price", _price(current_price))
c2.metric("Market cap ($bn)", "—" if _num(market_cap) is None else f"{_num(market_cap):,.1f}")
c3.metric("Forward P/E", "—" if _num(forward_pe) is None else f"{_num(forward_pe):.1f}x")
c4.metric("Quant score", "—" if _num(quant_score) is None else f"{_num(quant_score):.1f}/100")
c5.metric("Model view", str(verdict or "—"))
if why:
    st.info(str(why))

p1, p2, p3 = st.columns(3)
p1.metric("Industry market share", _pct(market_share))
p2.metric("Employee signal", "—" if _num(worker_score) is None else f"{_num(worker_score):.0f}/100*")
p3.metric("Leadership proxy", "—" if _num(leadership_score) is None else f"{_num(leadership_score):.1f}/100")
if market_share_basis:
    st.caption(f"Market share: {market_share_basis}")
if worker_scope:
    st.caption(f"* Employee signal scope: {worker_scope}")

with open(model_path, "rb") as f:
    st.download_button(
        "Download latest Excel model",
        data=f.read(),
        file_name=f"{ticker}_latest_equity_research.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


tab_hist, tab_segments, tab_people, tab_deals, tab_news, tab_sources = st.tabs(
    ["Historical Financials", "Segments", "Leadership & Culture", "Deals & Transactions", "Recent News", "Workbook Sources"]
)

with tab_hist:
    if "Historical Financials" not in wb.sheetnames:
        st.info("No Historical Financials sheet is available.")
    else:
        ws = wb["Historical Financials"]
        rows = {
            "Revenue ($bn)": 4,
            "Operating Income ($bn)": 9,
            "Net Income ($bn)": 11,
            "Diluted EPS": 12,
            "Operating Cash Flow ($bn)": 14,
            "Capital Expenditures ($bn)": 15,
        }
        records = []
        for c in range(2, 8):
            year = ws.cell(3, c).value
            if not isinstance(year, (int, float)):
                continue
            rec = {"Year": int(year)}
            for label, r in rows.items():
                rec[label] = _num(ws.cell(r, c).value)
            ocf = rec.get("Operating Cash Flow ($bn)")
            capex = rec.get("Capital Expenditures ($bn)")
            rec["FCF ($bn)"] = (ocf - abs(capex)) if ocf is not None and capex is not None else None
            records.append(rec)
        if records:
            st.dataframe(pd.DataFrame(records).set_index("Year"), use_container_width=True)
        else:
            st.info("No comparable annual history is currently available in the model.")

with tab_segments:
    if "Segment Analysis" not in wb.sheetnames:
        st.info("No Segment Analysis sheet is available.")
    else:
        ws = wb["Segment Analysis"]
        business_row = None
        for r in range(1, ws.max_row + 1):
            if str(ws.cell(r, 1).value or "").strip() in {"Revenue by Business Line / Product Group", "Revenue by Business Line"}:
                business_row = r
                break
        end = (business_row - 1) if business_row else min(ws.max_row, 22)
        records = []
        for r in range(7, end + 1):
            seg = ws.cell(r, 1).value
            if not seg:
                continue
            records.append({
                "Segment": seg,
                "Year -2 Revenue": _num(ws.cell(r, 2).value),
                "Year -1 Revenue": _num(ws.cell(r, 3).value),
                "Latest Revenue": _num(ws.cell(r, 4).value),
                "Year -2 Segment Profit": _num(ws.cell(r, 7).value),
                "Year -1 Segment Profit": _num(ws.cell(r, 8).value),
                "Latest Segment Profit": _num(ws.cell(r, 9).value),
                "Data Status": ws.cell(r, 15).value,
                "Source": ws.cell(r, 16).value,
            })
        if records:
            st.dataframe(pd.DataFrame(records), use_container_width=True)
        else:
            st.info("No segment names have been recovered yet.")

with tab_people:
    if people is None:
        st.info("No Leadership & Culture research sheet is available in this workbook yet.")
    else:
        l1, l2 = st.columns(2)
        with l1:
            st.markdown("#### Workforce signal")
            st.metric("Satisfaction / employee signal", "—" if _num(worker_score) is None else f"{_num(worker_score):.0f}/100")
            if worker_scope:
                st.write(worker_scope)
            evidence = _row_value(people, "Evidence")
            if evidence:
                st.caption(str(evidence))
        with l2:
            st.markdown("#### Leadership research")
            st.metric("Leadership evidence proxy", "—" if _num(leadership_score) is None else f"{_num(leadership_score):.1f}/100")
            st.caption("Transparent research proxy based on execution, capital allocation, leadership depth, culture and governance disclosure—not a factual management rating.")

        dimensions=[]
        for r in range(14,21):
            name_v=people.cell(r,1).value
            score_v=_num(people.cell(r,2).value)
            note_v=people.cell(r,3).value
            if name_v and score_v is not None:
                dimensions.append({"Dimension":name_v,"Score / 100":score_v,"Evidence / Caveat":note_v})
        if dimensions:
            st.dataframe(pd.DataFrame(dimensions),use_container_width=True,hide_index=True)

        conclusion=_row_value(people,"Research conclusion")
        best_peer=_row_value(people,"Best peer on current metrics")
        candidate=_row_value(people,"Candidate for deeper research")
        st.markdown("#### Same-sector alternative screen")
        if conclusion:
            st.write(conclusion)
        if candidate:
            st.success(f"Candidate for deeper research: {candidate}")
        elif best_peer:
            st.caption(f"Best current peer-screen score: {best_peer}; threshold for a clear alternative was not met.")

with tab_deals:
    if "Deals & Transactions" not in wb.sheetnames:
        st.info("No Deals & Transactions sheet is available in this workbook yet.")
    else:
        ws = wb["Deals & Transactions"]
        deals = []
        known_types = {
            "Acquisition / M&A", "Divestiture / Asset Sale", "Strategic Investment", "Joint Venture",
            "Major Commercial Contract", "Strategic Partnership", "Financing / Funding",
        }
        for r in range(15, ws.max_row + 1):
            deal_type = str(ws.cell(r, 3).value or "").strip()
            headline = ws.cell(r, 5).value
            if deal_type not in known_types or not headline:
                continue
            deals.append({
                "Date": ws.cell(r, 1).value,
                "Status": ws.cell(r, 2).value,
                "Type": deal_type,
                "Counterparty / Asset": ws.cell(r, 4).value,
                "Deal / Headline": headline,
                "Disclosed Value": ws.cell(r, 6).value,
                "Materiality": ws.cell(r, 7).value,
                "Strategic Area": ws.cell(r, 8).value,
                "Structure / Terms": ws.cell(r, 9).value,
                "Strategic Rationale — Analyst Lens": ws.cell(r, 10).value,
                "Financial / Valuation Impact — Analyst Lens": ws.cell(r, 11).value,
                "Key Risks — Analyst Lens": ws.cell(r, 12).value,
                "Regulatory / Closing": ws.cell(r, 13).value,
                "Next Milestone": ws.cell(r, 14).value,
                "Source Quality": ws.cell(r, 16).value,
                "Publisher": ws.cell(r, 17).value,
                "Source Link": ws.cell(r, 18).value,
            })
        if deals:
            df = pd.DataFrame(deals)
            st.caption("Material transaction monitor. Analyst-lens columns are inference; source-backed facts and links remain separate.")
            st.dataframe(df, use_container_width=True, hide_index=True)
            active = df[df["Status"].isin(["Announced / Pending", "Approved / Clearing", "Signed / Active"])]
            if not active.empty:
                st.markdown("#### Pending / active deal watch")
                st.dataframe(active[["Deal / Headline", "Status", "Disclosed Value", "Next Milestone", "Key Risks — Analyst Lens", "Source Link"]], use_container_width=True, hide_index=True)
        else:
            st.info("No sufficiently material recent deal was verified in the latest refresh.")

with tab_news:
    if "Recent News & Impact" not in wb.sheetnames:
        st.info("No Recent News & Impact sheet is available.")
    else:
        ws = wb["Recent News & Impact"]
        stories = []
        for r in range(1, ws.max_row + 1):
            tone = str(ws.cell(r, 2).value or "").strip()
            if tone not in {"Positive", "Negative", "Mixed / Neutral"}:
                continue
            stories.append({
                "Date": ws.cell(r, 1).value,
                "Tone": tone,
                "Headline": ws.cell(r, 3).value,
                "Publisher": ws.cell(r, 4).value,
                "Summary": ws.cell(r, 5).value,
                "Potential Impact": ws.cell(r, 6).value,
                "Link": ws.cell(r, 7).value,
            })
        if stories:
            st.dataframe(pd.DataFrame(stories), use_container_width=True, hide_index=True)
        else:
            st.info("No recent qualifying stories were returned in the latest refresh.")

with tab_sources:
    st.write(f"Daily bundle generated: {manifest.get('generated_utc', '—')}")
    st.write(f"GitHub refresh run: {manifest.get('github_run_updated_at', '—')}")
    st.write("The downloadable Excel workbook contains the detailed source notes and links used by each research module.")
